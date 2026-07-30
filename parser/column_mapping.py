"""
T32 — Column Mapping Wizard helpers.

Chức năng:
  1. `read_headers_and_preview(filepath, preview_rows_count)`
     → Đọc header row 1 + N dòng đầu để FE hiển thị preview trước khi map.
  2. `IHRP_STANDARD_COLUMNS` — danh sách cột chuẩn iHRP mà user cần map.
  3. `suggest_mapping(actual_headers, ihrp_columns)`
     → Fuzzy match từng cột iHRP với header thực tế; return top-3 candidate
     kèm score dùng `difflib.SequenceMatcher` (stdlib, không dep mới).

Preset lưu trữ trong `.project_store/<slug>/excel_mapping_presets.json`
qua `analyzer.project_store.load_mapping_presets` etc.

Wizard flow:
  User pick file → POST /api/upload-preview → nhận headers + preview + suggestion
  → sửa mapping trong UI → POST /api/upload-confirm với `column_mapping`
  → backend gọi FunctionListParser().parse(filepath, column_mapping=...).
"""
from __future__ import annotations

import re
from difflib import SequenceMatcher
from typing import Any, Optional

import openpyxl


# ---------------------------------------------------------------------------
# Standard iHRP column set — user maps đến 1 trong các cột này
# ---------------------------------------------------------------------------
#
# Meta columns: tên chuẩn có trong META_KEYWORDS của parser (auto-detect
# hoạt động nếu header match keyword; user chỉ cần map nếu file dùng tên lạ).
#
# Phase columns: sinh động theo pattern `<Phase> - <Attr>`. Danh sách này
# là gợi ý phổ biến nhất; user vẫn có thể map thêm phase custom bằng cách
# nhập tên iHRP tự do (VD "MyPhase - Start").
#
# FE dùng list này để render bảng mapping. Không phải mọi user đều cần map
# hết — chỉ cột thực sự tồn tại trong file.

_META_STANDARD = [
    "Mã CN",
    "Tên chức năng",
    "Module",
    "Priority",
    "Complexity",
    "FIT/GAP",
    "Giai đoạn",
    "Quy trình",
    "System",
    "Mô tả",
    "Function liên quan",
    "Risk/Blocker",
    "Last Updated Date",
]

_PHASE_NAMES = [
    "Analysis",
    "Dev",
    "Config Local",
    "Config UAT",
    "Test",
    "UAT",
    "Doc",
    "Golive",
]

_PHASE_ATTRS = ["Start", "End", "Status", "PIC", "Estimate MH"]

# Danh sách flat cho FE render bảng — thứ tự có ý nghĩa (meta trước, phase sau).
IHRP_STANDARD_COLUMNS: list[str] = list(_META_STANDARD) + [
    f"{phase} - {attr}" for phase in _PHASE_NAMES for attr in _PHASE_ATTRS
]


# ---------------------------------------------------------------------------
# Read headers + preview
# ---------------------------------------------------------------------------

def read_headers_and_preview(
    filepath: str,
    preview_rows_count: int = 5,
) -> tuple[list[str], list[list[Any]], Optional[str]]:
    """
    Mở workbook, đọc header row 1 + N dòng preview.

    Return (headers, preview_rows, sheet_name).
      - headers: list[str] giữ nguyên thứ tự cột trong file (cột trống → "").
      - preview_rows: list[list[Any]] mỗi row = list cell value (chuyển
        datetime → ISO string để JSON serializable).
      - sheet_name: tên sheet được đọc (để log).

    KHÔNG raise: file lỗi trả (empty_headers, empty_rows, None) để FE xử lý.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return [], [], None

    try:
        # Ưu tiên sheet "Function List" giống parser chính
        ws = None
        for name in wb.sheetnames:
            lname = name.lower().replace(" ", "")
            if "function" in lname or "functionlist" in lname:
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if ws is None:
            return [], [], None
        sheet_name = ws.title

        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration:
            return [], [], sheet_name
        headers = [str(v).strip() if v is not None else "" for v in first]

        preview: list[list[Any]] = []
        for _ in range(preview_rows_count):
            try:
                row = next(rows_iter)
            except StopIteration:
                break
            preview.append([_stringify_cell(v) for v in row])

        return headers, preview, sheet_name
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _stringify_cell(v: Any) -> Any:
    """Convert cell value → JSON-safe type. Datetime → ISO string."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):  # date / datetime
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    return v


# ---------------------------------------------------------------------------
# Bilingual alias hint — English ↔ Vietnamese
# ---------------------------------------------------------------------------
#
# Pure SequenceMatcher không giải quyết được case bilingual (VD "Module" vs
# "Phân hệ" — 0 ký tự trùng). Bảng alias dưới đây là hint tối thiểu cho
# các cột iHRP phổ biến; nếu 1 iHRP col hoặc actual header chứa keyword
# alias thì fuzzy score được cộng bonus 0.5. Không cần dep translation.
#
# Format: {ihrp_std_lowercased: [aliases_lowercased_có_trong_actual_header]}
_ALIAS_HINTS: dict[str, list[str]] = {
    "mã cn": ["function code", "code", "func code", "fcode", "ma cn", "ma_cn", "mã chức năng"],
    "tên chức năng": ["function name", "name", "func name", "fname", "ten cn", "ten chuc nang"],
    "module": ["phân hệ", "phan he", "sub-system", "subsystem", "phanhe"],
    "priority": ["ưu tiên", "uu tien", "do uu tien", "độ ưu tiên"],
    "complexity": ["độ phức tạp", "do phuc tap", "phuc tap"],
    "fit/gap": ["fit gap", "fit/gap", "fitgap"],
    "giai đoạn": ["stage", "phase (metadata)", "giai doan", "wave"],
    "quy trình": ["process", "business process", "quy trinh"],
    "system": ["hệ thống", "he thong"],
    "mô tả": ["description", "desc", "mo ta"],
    "function liên quan": ["related function", "related", "function lien quan"],
    "risk/blocker": ["risk", "blocker", "issue"],
    "last updated date": ["last updated", "updated at", "ngày cập nhật", "ngay cap nhat"],
}


def _alias_bonus(ihrp: str, actual: str) -> float:
    """
    Nếu ihrp có alias trong bảng và actual (normalized) chứa 1 alias
    → bonus 0.5. Cover case bilingual & viết tắt phổ biến.
    """
    ihrp_key = str(ihrp).strip().lower()
    aliases = _ALIAS_HINTS.get(ihrp_key)
    if not aliases:
        # Cũng match theo bảng ngược: nếu actual là 1 key trong bảng
        # và ihrp là alias của actual → bonus.
        for base_key, alias_list in _ALIAS_HINTS.items():
            if str(actual).strip().lower() == base_key:
                if any(a in ihrp_key or ihrp_key in a for a in alias_list):
                    return 0.5
        return 0.0
    actual_low = str(actual).strip().lower()
    for a in aliases:
        if a in actual_low or actual_low in a:
            return 0.5
    return 0.0


# ---------------------------------------------------------------------------
# Fuzzy matching (SequenceMatcher — stdlib, không thêm dep)
# ---------------------------------------------------------------------------

def _normalize_for_match(s: str) -> str:
    """
    Chuẩn hoá text trước khi compare: lowercase + strip diacritic đơn giản +
    bỏ ký tự phi alphanumeric để `Analysis-Start`, `analysis_start`,
    `AnalysisStart`, `Analysis Start` khớp với nhau.

    KHÔNG dùng unicodedata.normalize để tránh strip dấu tiếng Việt (mà user
    có thể muốn giữ để phân biệt "Ma CN" vs "Mã CN"). Chỉ normalize
    case + separator.
    """
    if s is None:
        return ""
    return re.sub(r"[^0-9a-zA-Zà-ỹÀ-Ỹ]+", "", str(s).strip().lower())


def _fuzzy_score(a: str, b: str) -> float:
    """
    Trả 0.0-1.0. Kết hợp 4 signal:
      1. Ratio SequenceMatcher trên bản normalized (order-insensitive về sep).
      2. Substring bonus: nếu 1 chuỗi chứa toàn bộ chuỗi kia → +0.15.
      3. Token overlap: chia bởi whitespace/[-_/.] rồi tính tỷ lệ token trùng.
      4. Alias hint bilingual: match qua bảng _ALIAS_HINTS → +0.5 (cover case
         "Module" vs "Phân hệ", "Mã CN" vs "Function Code" mà pure string
         không giải quyết được).
    """
    if not a or not b:
        return 0.0
    na, nb = _normalize_for_match(a), _normalize_for_match(b)
    if not na or not nb:
        return 0.0
    ratio = SequenceMatcher(None, na, nb).ratio()

    bonus = 0.0
    if na in nb or nb in na:
        bonus += 0.15

    tokens_a = {t for t in re.split(r"[\s\-_/.]+", a.lower()) if t}
    tokens_b = {t for t in re.split(r"[\s\-_/.]+", b.lower()) if t}
    if tokens_a and tokens_b:
        overlap = len(tokens_a & tokens_b) / max(len(tokens_a | tokens_b), 1)
        bonus += 0.1 * overlap

    # Alias hint (bilingual English ↔ Vietnamese) — check cả 2 hướng.
    bonus += max(_alias_bonus(a, b), _alias_bonus(b, a))

    return min(1.0, ratio + bonus)


def suggest_mapping(
    actual_headers: list[str],
    ihrp_columns: Optional[list[str]] = None,
    top_k: int = 3,
    min_score: float = 0.35,
) -> dict[str, list[dict]]:
    """
    Với mỗi cột iHRP chuẩn, gợi ý top-K header thực tế match nhất.

    Args:
        actual_headers: list header từ file user upload.
        ihrp_columns: danh sách cột iHRP cần suggest (default: IHRP_STANDARD_COLUMNS).
        top_k: số candidate max per iHRP column.
        min_score: score < min_score → drop khỏi kết quả (giảm nhiễu).

    Return:
        {ihrp_col: [{"header": actual_str, "score": 0.0-1.0}, ...], ...}
        List sorted desc theo score. iHRP col không có candidate nào >= min_score
        → giá trị list rỗng (FE render "(không có)" xám nhạt).
    """
    ihrp_cols = ihrp_columns or IHRP_STANDARD_COLUMNS
    out: dict[str, list[dict]] = {}
    # Chuẩn hoá list actual 1 lần để dùng lại
    actual_clean = [h for h in actual_headers if h]
    for ihrp in ihrp_cols:
        scores = []
        for actual in actual_clean:
            s = _fuzzy_score(ihrp, actual)
            if s >= min_score:
                scores.append({"header": actual, "score": round(s, 3)})
        scores.sort(key=lambda x: x["score"], reverse=True)
        out[ihrp] = scores[:top_k]
    return out


# ---------------------------------------------------------------------------
# Sanitize column_mapping input
# ---------------------------------------------------------------------------

def sanitize_column_mapping(mapping: Any) -> dict[str, str]:
    """
    Validate + normalize column_mapping từ FE. Chấp nhận dict {ihrp: actual}.

    - Reject non-dict → return {}.
    - Trim key/value; drop entry rỗng.
    - Giới hạn 200 entries + mỗi key/value tối đa 200 ký tự (chống abuse).
    """
    if not isinstance(mapping, dict):
        return {}
    out: dict[str, str] = {}
    for k, v in list(mapping.items())[:200]:
        key = str(k or "").strip()[:200]
        val = str(v or "").strip()[:200]
        if key and val:
            out[key] = val
    return out
