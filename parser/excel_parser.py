"""
Excel Function List Parser — Tự động phát hiện cấu trúc cột.
KHÔNG hardcode index cột. Mọi thứ dựa trên header text row 1.
"""
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional
import openpyxl


# === Trạng thái hợp lệ ===
VALID_STATUSES = {"Open", "Assigned", "In-progress", "Resolved", "Closed", "Pending", "Cancelled"}

# "Not Started" / "Chưa bắt đầu" — KHÔNG alias tĩnh.
# Map phụ thuộc PIC phase: không PIC → Open; có PIC → Assigned.
_NOT_STARTED_TOKENS = {
    "not started",
    "not_started",
    "notstarted",
    "chưa bắt đầu",
    "chua bat dau",
}

# Alias tĩnh phổ biến từ API/Excel → canonical VALID_STATUSES (key = lower-case).
STATUS_ALIASES = {
    # --- Open ---
    "open": "Open",
    "mới": "Open",
    "moi": "Open",

    # --- Assigned ---
    "assign": "Assigned",
    "assigned": "Assigned",
    "đã giao": "Assigned",
    "da giao": "Assigned",

    # --- In-progress ---
    "in progress": "In-progress",
    "inprogress": "In-progress",
    "in_progress": "In-progress",
    "in-progress": "In-progress",
    "đang làm": "In-progress",
    "dang lam": "In-progress",
    "đang thực hiện": "In-progress",
    "dang thuc hien": "In-progress",

    # --- Resolved ---
    "resolve": "Resolved",
    "resolved": "Resolved",
    "đã xử lý": "Resolved",
    "da xu ly": "Resolved",

    # --- Closed ---
    "closed": "Closed",
    "đóng": "Closed",
    "dong": "Closed",
    "finished": "Closed",
    "done": "Closed",
    "complete": "Closed",
    "completed": "Closed",
    "hoàn thành": "Closed",
    "hoan thanh": "Closed",
    "xong": "Closed",

    # --- Pending ---
    "pending": "Pending",
    "waiting": "Pending",
    "wait": "Pending",
    "chờ": "Pending",
    "cho": "Pending",
    "on hold": "Pending",
    "onhold": "Pending",
    "on_hold": "Pending",

    # --- Cancelled ---
    "cancel": "Cancelled",
    "canceled": "Cancelled",  # US spelling
    "cancelled": "Cancelled",
    "hủy": "Cancelled",
    "huy": "Cancelled",
    "đã hủy": "Cancelled",
    "da huy": "Cancelled",
}

# Ngưỡng outlier Estimate MH (1 ô phase). > 500 thường là Excel date-serial
# bị lệch cột (VD ~45000 ngày) hoặc timestamp kiểu ~1.7e12 — không phải MH thật.
ESTIMATE_MH_MAX = 500.0

# === Keyword mapping để detect cột meta ===
META_KEYWORDS = {
    "stt": ["STT", "No", "#"],
    "ma_cn": ["Mã CN", "Mã chức năng", "Function Code", "Code"],
    "ten_cn": ["Tên chức năng", "Function Name", "Tên CN"],
    "module": ["Module", "Phân hệ"],
    "system": ["System", "Hệ thống"],
    "fid": ["FID", "Function ID"],
    "quy_trinh": ["Quy trình", "Process", "Business Process"],
    "requirement_id": ["Requirement ID", "Req ID"],
    "fit_gap": ["FIT/GAP", "FIT GAP", "Fit/Gap"],
    "giai_doan": ["Giai đoạn", "Stage"],
    "priority": ["Priority", "Ưu tiên", "Độ ưu tiên"],
    "complexity": ["Complexity", "Độ phức tạp"],
    "mo_ta": ["Mô tả", "Description"],
    "function_lq": ["Function liên quan", "Related Function"],
    "risk_blocker": ["Risk/Blocker", "Risk", "Blocker"],
    "last_updated": ["Last Updated Date", "Last Updated", "Ngày cập nhật"],
    "remark": ["Remark", "Ghi chú chung"],
    # Mã dự án (iHRP Task Daily API: JSON key `project`, VD "MPHG_IHRP_2025_PM")
    "ma_du_an": ["Mã dự án", "Project Code", "Project"],
    # Phase C — Change Request / scope creep (KHÔNG dùng keyword ngắn "CR"
    # trong list này vì partial match sẽ dính "Description"; exact "CR"
    # được xử lý riêng trong _detect_meta_columns).
    "is_cr": [
        "Change Request", "Phát sinh", "Phat sinh", "Scope Creep",
        "Is CR", "CR Flag", "CR?", "CR No", "CR Number", "CR ID",
    ],
    "cr_date": [
        "Ngày phát sinh", "Ngay phat sinh", "CR Date", "CR Raised",
        "Change Request Date", "Ngày CR", "Ngay CR",
    ],
    # Phase E — UAT / customer feedback quality (KHÔNG dùng keyword ngắn "Bug"
    # trong partial match vì dính "Debug"; exact Bug/Defect xử lý riêng).
    "defect_count": [
        "Defect Count", "Bug Count", "Số lỗi", "So loi", "Số bug", "So bug",
        "Số Defect", "So Defect", "Defect Qty", "Bug Qty",
        "Defects", "Defect/Bug", "Bug/Defect", "Số Defects",
    ],
    "feedback_count": [
        "Feedback Count", "Số phản hồi", "So phan hoi",
        "Customer Feedback", "Phản hồi", "Phan hoi", "Feedback",
    ],
    "reopen_count": [
        "Reopen Count", "Re-open Count", "Số lần reopen", "So lan reopen",
        "Số reopen", "So reopen", "Reopen Times", "Reopens",
        "Re-open", "Reopen",
    ],
    "uat_cycle": [
        "UAT Cycle Count", "UAT Cycles", "Số vòng UAT", "So vong UAT",
        "Vòng UAT", "Vong UAT", "UAT Cycle", "Cycle UAT", "UAT vòng",
    ],
}

# === Mapping phase name → task type (tiếng Việt) ===
TASK_TYPE_RULES = [
    (r"(?i)analy",                          "Phân tích"),
    (r"(?i)\bdev\b",                        "Lập trình"),
    (r"(?i)(?:local|test)",                 "Kiểm thử"),
    (r"(?i)config.*uat",                    "Cấu hình UAT"),
    (r"(?i)^uat$",                          "UAT"),
    (r"(?i)doc",                            "Tài liệu"),
    (r"(?i)(?:prod|golive|go.?live)",       "Cấu hình Golive"),
]


@dataclass
class PhaseGroup:
    """Nhóm cột cho 1 phase (VD: Analysis, Dev, Config UAT...)"""
    name: str
    attributes: dict[str, int] = field(default_factory=dict)  # attr_name → col_index (1-based)

    @property
    def start_col(self) -> Optional[int]:
        return self.attributes.get("Start") or self.attributes.get("From") or self.attributes.get("Planned")

    @property
    def end_col(self) -> Optional[int]:
        return self.attributes.get("End") or self.attributes.get("To") or self.attributes.get("Actual")

    @property
    def status_col(self) -> Optional[int]:
        return self.attributes.get("Status")

    @property
    def pic_cols(self) -> list[int]:
        return [v for k, v in self.attributes.items() if "PIC" in k.upper()]

    @property
    def estimate_col(self) -> Optional[int]:
        return self.attributes.get("Estimate MH")

    @property
    def note_col(self) -> Optional[int]:
        return self.attributes.get("Note")

    @property
    def task_type(self) -> str:
        """Map tên phase → loại công việc tiếng Việt."""
        for pattern, label in TASK_TYPE_RULES:
            if re.search(pattern, self.name):
                return label
        return self.name


@dataclass
class PhaseData:
    """Dữ liệu của 1 function tại 1 phase."""
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    status: Optional[str] = None
    pics: list[str] = field(default_factory=list)
    estimate_mh: Optional[float] = None
    note: Optional[str] = None
    extra: dict[str, Any] = field(default_factory=dict)
    # True khi raw status là "Not Started"/"Chưa bắt đầu" rồi map → Open/Assigned.
    # Unassigned start-gate không được coi đây là "đang làm" nếu chưa tới Start/End.
    from_not_started: bool = False


@dataclass
class FunctionRow:
    """1 dòng = 1 chức năng trong Function List."""
    row_num: int
    meta: dict[str, Any] = field(default_factory=dict)
    phases: dict[str, PhaseData] = field(default_factory=dict)


@dataclass
class ParsedData:
    """Kết quả parse toàn bộ file."""
    headers: dict[str, int]                  # header_text → col_index
    meta_columns: dict[str, Optional[int]]   # meta_key → col_index
    phase_groups: list[PhaseGroup]
    rows: list[FunctionRow]

    # Danh sách unique values (sorted)
    all_modules: list[str] = field(default_factory=list)
    all_phases: list[str] = field(default_factory=list)
    all_pics: list[str] = field(default_factory=list)
    all_statuses: list[str] = field(default_factory=list)
    all_priorities: list[str] = field(default_factory=list)
    all_complexities: list[str] = field(default_factory=list)
    all_giai_doan: list[str] = field(default_factory=list)
    all_processes: list[str] = field(default_factory=list)  # danh sách Quy trình
    all_project_codes: list[str] = field(default_factory=list)  # Mã dự án (meta.ma_du_an)

    # Data-quality log: các token PIC bị blacklist khi parse.
    # Mỗi entry: {row_index, phase_name, header_text, raw_value, matched_keyword,
    #             ma_cn, module}. Chỉ log token thuộc VALID_STATUSES bị lộn cột
    # (không log "-" / "n/a" — đó là placeholder trống, không phải bug lệch cột).
    pic_blacklisted: list[dict] = field(default_factory=list)

    # Data-quality log: ô Estimate MH bị reject (datetime / outlier > ESTIMATE_MH_MAX).
    # Mỗi entry: {row_index, phase_name, header, raw_value, reason, ma_cn, module}.
    estimate_mh_rejected: list[dict] = field(default_factory=list)


class FunctionListParser:
    """
    Parser chính — auto-detect cấu trúc file Function List.
    Không hardcode cột, hoạt động với bất kỳ file nào có cùng pattern header.
    """

    def parse(self, filepath: str,
              column_mapping: Optional[dict[str, str]] = None) -> ParsedData:
        """
        Đọc file Excel, trả về ParsedData.

        Args:
            filepath: đường dẫn file .xlsx / .xls.
            column_mapping (T32 — Column Mapping Wizard): dict {ihrp_std: actual}
                map header chuẩn iHRP → header thực tế trong file. Nếu có →
                RENAME header trong `headers` dict trước khi detect meta/phase
                → cho phép user upload file có tên cột lạ mà không cần rewrite
                Excel. VD: {"Tên chức năng": "Function Name",
                             "Analysis - Start": "Ngày bắt đầu phân tích"}.
                Bỏ qua entry có value rỗng (user chưa map cột đó).

        Tối ưu tốc độ: mở workbook ở chế độ read_only + data_only,
        đọc dữ liệu 1 lần qua iter_rows thay vì random-access ws.cell().
        """
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            # Tìm sheet chính: ưu tiên "Function List", fallback sheet đầu
            ws = None
            for name in wb.sheetnames:
                if "function" in name.lower() or "functionlist" in name.lower().replace(" ", ""):
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb[wb.sheetnames[0]]

            # Streaming: đọc tất cả row vào memory 1 lần
            # (nhanh hơn nhiều so với ws.cell(r, c).value trong read_only mode)
            all_rows_matrix: list[tuple] = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not all_rows_matrix:
            return ParsedData(
                headers={}, meta_columns={}, phase_groups=[], rows=[],
                all_modules=[], all_phases=[], all_pics=[], all_statuses=[],
                all_priorities=[], all_complexities=[], all_giai_doan=[],
                all_processes=[], all_project_codes=[],
            )

        # 1. Row 1 = headers → dict {header_text: col_index 1-based}
        header_tuple = all_rows_matrix[0]
        headers: dict[str, int] = {}
        for col_idx_0, val in enumerate(header_tuple):
            if val is not None:
                headers[str(val).strip()] = col_idx_0 + 1

        # 1.5. Apply column_mapping (T32) — user map header lạ → header iHRP
        #      chuẩn. Chỉ ADD alias, KHÔNG remove header gốc → detect
        #      keyword-based vẫn hoạt động cho các cột auto-match được.
        if column_mapping:
            headers = self._apply_column_mapping(headers, column_mapping)

        # 2. Detect meta columns
        meta_columns = self._detect_meta_columns(headers)

        # 3. Detect phase groups
        phase_groups = self._detect_phase_groups(headers)

        # 4. Parse data rows từ row 2 trở đi
        # Trả kèm PIC-blacklist + Estimate MH reject log (data-quality).
        rows, pic_blacklisted, estimate_mh_rejected = self._parse_rows_from_matrix(
            all_rows_matrix[1:], meta_columns, phase_groups, headers
        )

        # 4.5. Normalize PIC names case-insensitively
        # Nếu file có cả "SonHN6" và "SONHN6" thì coi là 1 người,
        # dùng version xuất hiện đầu tiên có mix case (không phải all-caps).
        self._normalize_pic_names(rows)

        # 5. Collect unique values
        all_modules = sorted({r.meta.get("module", "") for r in rows if r.meta.get("module")})
        all_phases = [pg.name for pg in phase_groups]
        all_priorities = sorted({r.meta.get("priority", "") for r in rows if r.meta.get("priority")})
        all_complexities = sorted({r.meta.get("complexity", "") for r in rows if r.meta.get("complexity")})
        all_giai_doan = sorted({str(r.meta.get("giai_doan", "")) for r in rows if r.meta.get("giai_doan")})
        all_processes = sorted({r.meta.get("quy_trinh", "") for r in rows if r.meta.get("quy_trinh")})
        all_project_codes = sorted({
            str(r.meta.get("ma_du_an", "")).strip()
            for r in rows
            if r.meta.get("ma_du_an") and str(r.meta.get("ma_du_an")).strip()
        })

        pics_set = set()
        statuses_set = set()
        for r in rows:
            for pd in r.phases.values():
                pics_set.update(pd.pics)
                if pd.status:
                    statuses_set.add(pd.status)

        # (workbook đã được close trong `finally` block ở trên)
        return ParsedData(
            headers=headers,
            meta_columns=meta_columns,
            phase_groups=phase_groups,
            rows=rows,
            all_modules=all_modules,
            all_phases=all_phases,
            all_pics=sorted(pics_set),
            all_statuses=sorted(statuses_set),
            all_priorities=all_priorities,
            all_complexities=all_complexities,
            all_giai_doan=all_giai_doan,
            pic_blacklisted=pic_blacklisted,
            estimate_mh_rejected=estimate_mh_rejected,
            all_processes=all_processes,
            all_project_codes=all_project_codes,
        )

    # ------------------------------------------------------------------
    # T32: Column Mapping Wizard — apply mapping thủ công từ user
    # ------------------------------------------------------------------

    def _apply_column_mapping(
        self,
        headers: dict[str, int],
        column_mapping: dict[str, str],
    ) -> dict[str, int]:
        """
        Thêm alias iHRP-standard vào `headers` dict dựa vào mapping user cấu hình.

        Args:
            headers: dict {header_text_trong_file: col_index_1_based}.
            column_mapping: dict {ihrp_std_name: actual_header_in_file}.
                VD: {"Mã CN": "Function Code", "Analysis - Start": "AnalysisStart"}.
                Entry có value rỗng → skip (user chưa map cột đó).

        Returns:
            Bản copy của headers với các iHRP standard name được add vào
            cùng col_index của header thực tế. Header gốc VẪN GIỮ để không
            phá auto-detect (nếu 1 cột được cả auto + manual match, kết
            quả giống nhau vì cùng col_index).

        Behavior:
            - Nếu `actual` không tồn tại trong headers → skip (log warning).
            - Nếu `ihrp_std` đã tồn tại → OVERRIDE bằng col của actual
              (user muốn thay vì auto-detect thì tôn trọng).
        """
        new_headers = dict(headers)
        for ihrp_std, actual in column_mapping.items():
            if not ihrp_std or not actual:
                continue
            ihrp_std = str(ihrp_std).strip()
            actual = str(actual).strip()
            if not ihrp_std or not actual:
                continue
            # Match theo original text đầu tiên; fallback lowercase match
            col_idx = headers.get(actual)
            if col_idx is None:
                for h, idx in headers.items():
                    if h.lower() == actual.lower():
                        col_idx = idx
                        break
            if col_idx is None:
                # Header user map không tồn tại trong file — bỏ qua thầm lặng
                # (FE đã pre-validate ở upload-preview endpoint).
                continue
            new_headers[ihrp_std] = col_idx
        return new_headers

    # ------------------------------------------------------------------
    # Detect meta columns (bằng keyword matching)
    # ------------------------------------------------------------------

    def _detect_meta_columns(self, headers: dict[str, int]) -> dict[str, Optional[int]]:
        """Tìm cột meta bằng keyword, trả về {meta_key: col_index}."""
        result: dict[str, Optional[int]] = {}
        header_lower_map = {h.lower(): (h, idx) for h, idx in headers.items()}

        for meta_key, keywords in META_KEYWORDS.items():
            found = None
            for kw in keywords:
                kw_lower = kw.lower()
                # Exact match trước
                for h_lower, (h_orig, idx) in header_lower_map.items():
                    if h_lower == kw_lower:
                        found = idx
                        break
                if found:
                    break
                # Partial match
                for h_lower, (h_orig, idx) in header_lower_map.items():
                    if kw_lower in h_lower and " - " not in h_orig:
                        found = idx
                        break
                if found:
                    break
            result[meta_key] = found

        # Exact header "CR" — không dùng partial (tránh "Description" chứa "cr")
        if result.get("is_cr") is None:
            for h, idx in headers.items():
                if " - " in h:
                    continue
                if h.strip().upper() == "CR":
                    result["is_cr"] = idx
                    break

        # Exact Bug/Defect — không partial (tránh "Debug" chứa "bug")
        if result.get("defect_count") is None:
            for h, idx in headers.items():
                if " - " in h:
                    continue
                if h.strip().lower() in ("bug", "bugs", "defect", "defects"):
                    result["defect_count"] = idx
                    break

        return result

    # ------------------------------------------------------------------
    # Detect phase groups (bằng pattern "PhaseName - Attribute")
    # ------------------------------------------------------------------

    def _detect_phase_groups(self, headers: dict[str, int]) -> list[PhaseGroup]:
        """Tìm nhóm phase bằng pattern 'Phase - Attr' trong header."""
        groups: dict[str, PhaseGroup] = {}
        phase_order: list[str] = []  # Giữ thứ tự xuất hiện trong file

        for header_text, col_idx in headers.items():
            if " - " in header_text:
                # Tách: "Config UAT - Status" → phase="Config UAT", attr="Status"
                parts = header_text.rsplit(" - ", 1)
                if len(parts) == 2:
                    phase_name = parts[0].strip()
                    attr_name = parts[1].strip()

                    if phase_name not in groups:
                        groups[phase_name] = PhaseGroup(name=phase_name)
                        phase_order.append(phase_name)

                    groups[phase_name].attributes[attr_name] = col_idx

        # Return theo thứ tự xuất hiện
        return [groups[name] for name in phase_order if name in groups]

    # ------------------------------------------------------------------
    # Parse data rows
    # ------------------------------------------------------------------

    def _parse_rows_from_matrix(
        self,
        data_rows: list[tuple],
        meta_columns: dict[str, Optional[int]],
        phase_groups: list[PhaseGroup],
        headers: dict[str, int] | None = None,
    ) -> tuple[list[FunctionRow], list[dict], list[dict]]:
        """
        Parse từ matrix data đã load sẵn (list of tuples).
        Truy cập theo tuple index (0-based) nên nhanh hơn ws.cell nhiều lần.

        Return:
            (rows, pic_blacklisted, estimate_mh_rejected)
              - rows: list FunctionRow
              - pic_blacklisted: list dict data-quality log
                {row_index, phase_name, header_text, raw_value, matched_keyword,
                 ma_cn, module}
              - estimate_mh_rejected: list dict
                {row_index, phase_name, header, raw_value, reason, ma_cn, module}
        """
        rows: list[FunctionRow] = []
        pic_blacklisted: list[dict] = []
        estimate_mh_rejected: list[dict] = []

        # Reverse map col_index (1-based) → header text để log rõ ràng
        # (VD "UAT - PIC FPT" giúp user tra cell nào bị lệch cột)
        col_to_header: dict[int, str] = {}
        if headers:
            for h, idx in headers.items():
                col_to_header[idx] = h

        # Cache column index (chuyển 1-based → 0-based 1 lần thay vì mỗi row)
        stt_col = meta_columns.get("stt")
        ma_col = meta_columns.get("ma_cn")
        ten_col = meta_columns.get("ten_cn")
        stt_idx = (stt_col - 1) if stt_col else None
        ma_idx = (ma_col - 1) if ma_col else None
        ten_idx = (ten_col - 1) if ten_col else None

        # Cache meta column indices (0-based)
        meta_indices = {k: (v - 1) for k, v in meta_columns.items() if v is not None}

        # Cache phase column indices
        known_attrs = {"Start", "End", "From", "To", "Planned", "Actual",
                       "Status", "Estimate MH", "Note"}

        phase_cache = []
        for pg in phase_groups:
            # Với mỗi PIC col: giữ cả 0-based index + header text gốc để log
            pic_cols_info = [
                {"idx": c - 1, "header": col_to_header.get(c, f"col_{c}")}
                for c in pg.pic_cols
            ]
            phase_cache.append({
                "name": pg.name,
                "start_idx": (pg.start_col - 1) if pg.start_col else None,
                "end_idx": (pg.end_col - 1) if pg.end_col else None,
                "status_idx": (pg.status_col - 1) if pg.status_col else None,
                "pic_cols_info": pic_cols_info,
                "estimate_idx": (pg.estimate_col - 1) if pg.estimate_col else None,
                "note_idx": (pg.note_col - 1) if pg.note_col else None,
                "extra_attrs": [
                    (name, col - 1)
                    for name, col in pg.attributes.items()
                    if name not in known_attrs and "PIC" not in name.upper()
                ],
            })

        for row_offset, row_tuple in enumerate(data_rows):
            # row_offset = 0 tương ứng row Excel số 2
            row_num = row_offset + 2

            # Row trống → skip
            has_data = False
            for check_idx in (stt_idx, ma_idx, ten_idx):
                if check_idx is not None and check_idx < len(row_tuple) and row_tuple[check_idx] is not None:
                    has_data = True
                    break
            if not has_data:
                continue

            def _get(idx: Optional[int]):
                """Helper: đọc giá trị tại 0-based index, trả None nếu out-of-range."""
                if idx is None or idx >= len(row_tuple):
                    return None
                return row_tuple[idx]

            # Parse meta
            meta = {}
            for key, idx in meta_indices.items():
                val = _get(idx)
                if val is not None:
                    if isinstance(val, (int, float, datetime)):
                        meta[key] = val
                    else:
                        meta[key] = str(val).strip()
                else:
                    meta[key] = None

            # Chuẩn hóa giai_doan → string
            gd = meta.get("giai_doan")
            if gd is not None:
                if isinstance(gd, (int, float)):
                    try:
                        if float(gd).is_integer():
                            meta["giai_doan"] = str(int(gd))
                        else:
                            meta["giai_doan"] = str(gd)
                    except (TypeError, ValueError):
                        meta["giai_doan"] = str(gd)
                else:
                    meta["giai_doan"] = str(gd).strip()

            # Parse phases
            phases: dict[str, PhaseData] = {}
            for pc in phase_cache:
                pd = PhaseData()

                if pc["start_idx"] is not None:
                    pd.start_date = self._normalize_date(_get(pc["start_idx"]))
                if pc["end_idx"] is not None:
                    pd.end_date = self._normalize_date(_get(pc["end_idx"]))

                # PIC trước Status — "Not Started" map Open/Assigned phụ thuộc PIC
                for pic_info in pc["pic_cols_info"]:
                    pv = _get(pic_info["idx"])
                    if pv:
                        # Dùng version có log để thu thập token blacklist
                        valid, dropped = self._parse_pics_with_log(pv)
                        pd.pics.extend(valid)
                        # Ghi log với đầy đủ context (row + phase + header)
                        for d in dropped:
                            pic_blacklisted.append({
                                "row_index": row_num,
                                "phase_name": pc["name"],
                                "header_text": pic_info["header"],
                                "raw_value": d["raw_value"],
                                "matched_keyword": d["matched_keyword"],
                                # ma_cn + module để user tra cứu nhanh trong Excel
                                "ma_cn": meta.get("ma_cn") or "",
                                "module": meta.get("module") or "",
                            })

                if pc["status_idx"] is not None:
                    raw_status = _get(pc["status_idx"])
                    pd.status = self._normalize_status(
                        raw_status,
                        has_pic=bool(pd.pics),
                    )
                    pd.from_not_started = self._is_not_started_token(raw_status)

                if pc["estimate_idx"] is not None:
                    est = _get(pc["estimate_idx"])
                    est_header = col_to_header.get(
                        (pc["estimate_idx"] + 1) if pc["estimate_idx"] is not None else -1,
                        f"{pc['name']} - Estimate MH",
                    )
                    mh_val, reject_reason = self._normalize_estimate_mh(est)
                    if reject_reason:
                        # raw_value serialize được JSON (datetime → iso string)
                        if isinstance(est, datetime):
                            raw_repr: Any = est.isoformat(sep=" ", timespec="seconds")
                        elif isinstance(est, date):
                            raw_repr = est.isoformat()
                        else:
                            raw_repr = est
                        estimate_mh_rejected.append({
                            "row_index": row_num,
                            "phase_name": pc["name"],
                            "header": est_header,
                            "raw_value": raw_repr,
                            "reason": reject_reason,
                            "ma_cn": meta.get("ma_cn") or "",
                            "module": meta.get("module") or "",
                        })
                        pd.estimate_mh = None
                    else:
                        pd.estimate_mh = mh_val

                if pc["note_idx"] is not None:
                    nv = _get(pc["note_idx"])
                    if nv:
                        pd.note = str(nv).strip()

                for attr_name, attr_idx in pc["extra_attrs"]:
                    v = _get(attr_idx)
                    if v is not None:
                        pd.extra[attr_name] = v

                phases[pc["name"]] = pd

            rows.append(FunctionRow(row_num=row_num, meta=meta, phases=phases))

        return rows, pic_blacklisted, estimate_mh_rejected

    # ------------------------------------------------------------------
    # Normalize helpers
    # ------------------------------------------------------------------

    def _normalize_date(self, value) -> Optional[date]:
        """Chuẩn hóa date từ nhiều format."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            return None  # Không phải date

        s = str(value).strip()
        if not s:
            return None

        # Thử các format
        for fmt in ["%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _normalize_estimate_mh(self, value) -> tuple[Optional[float], Optional[str]]:
        """
        Chuẩn hóa Estimate MH. Reject datetime/date và outlier > ESTIMATE_MH_MAX.

        Root cause số ~1.7e12 / ~45000: ô Estimate bị lệch cột Date → openpyxl
        trả datetime hoặc Excel date-serial (float lớn). Không phải man-hour.

        Return:
            (mh_float | None, reason | None)
            - Giá trị hợp lệ → (float, None)
            - Ô trống / không phải số → (None, None) — không log
            - Reject → (None, reason_str) — caller ghi vào estimate_mh_rejected
        """
        if value is None:
            return None, None
        # datetime/date — lệch cột Start/End sang Estimate
        if isinstance(value, datetime):
            return None, "datetime_not_allowed"
        if isinstance(value, date):
            return None, "date_not_allowed"
        # bool là subclass của int trong Python — không coi là MH
        if isinstance(value, bool):
            return None, None
        if isinstance(value, (int, float)):
            mh = float(value)
            if mh > ESTIMATE_MH_MAX:
                return None, f"outlier_gt_{int(ESTIMATE_MH_MAX)}"
            return mh, None
        # Chuỗi số hợp lệ (hiếm, khi cell format Text)
        s = str(value).strip()
        if not s:
            return None, None
        try:
            mh = float(s.replace(",", "."))
        except (TypeError, ValueError):
            return None, None
        if mh > ESTIMATE_MH_MAX:
            return None, f"outlier_gt_{int(ESTIMATE_MH_MAX)}"
        return mh, None

    @staticmethod
    def _is_not_started_token(value) -> bool:
        """Raw status có phải Not Started / Chưa bắt đầu không (trước khi map)."""
        if value is None or isinstance(value, (int, float)):
            return False
        return str(value).strip().lower() in _NOT_STARTED_TOKENS

    def _normalize_status(self, value, has_pic: bool = False) -> Optional[str]:
        """Chuẩn hóa status về VALID_STATUSES.

        - Số (int/float): bỏ qua → None (Estimate MH lệch cột).
        - "Not Started" / "Chưa bắt đầu": không PIC → Open; có PIC → Assigned.
        - Alias / match VALID_STATUSES: trả canonical.
        - Status lạ / unknown → None (coi blank).
        """
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return None  # Lỗi: giá trị Estimate MH bị lệch sang cột Status
        s = str(value).strip()
        if not s:
            return None
        low = s.lower()
        # Not Started phụ thuộc PIC — không dùng STATUS_ALIASES tĩnh
        if low in _NOT_STARTED_TOKENS:
            return "Assigned" if has_pic else "Open"
        # Alias tĩnh (VD "In Progress" → "In-progress", "Finished" → "Closed")
        alias = STATUS_ALIASES.get(low)
        if alias:
            return alias
        # Match case-insensitive với VALID_STATUSES
        for valid in VALID_STATUSES:
            if low == valid.lower():
                return valid
        return None  # Status lạ → blank

    def _parse_pics(self, value) -> list[str]:
        """
        Tách nhiều PIC từ 1 cell.

        Filter dữ liệu bẩn:
        - Rỗng, "-", "n/a"
        - Trùng tên với 1 trong VALID_STATUSES (case-insensitive) — case này xảy ra
          khi cột Status bị lệch qua cột PIC do user paste sai. Nếu người thật có
          tên trùng status ("Open", "Closed"…) sẽ bị bỏ, nhưng xác suất cực thấp.
          Xem screenshot 2 bug 3: "Closed" bị lẫn vào dropdown PIC.

        Wrapper backward-compat: chỉ trả list PIC hợp lệ.
        Muốn biết token nào bị bỏ → dùng `_parse_pics_with_log`.
        """
        pics, _ = self._parse_pics_with_log(value)
        return pics

    def _parse_pics_with_log(self, value) -> tuple[list[str], list[dict]]:
        """
        Giống `_parse_pics` nhưng còn trả về log các token BỊ BLACKLIST
        (chỉ những token match VALID_STATUSES — dấu hiệu lệch cột thật sự).

        Return:
            (pics_giữ_lại, blacklist_tokens)
            Mỗi blacklist_token: {"raw_value": <str>, "matched_keyword": <str canonical>}

        Không log "-" / "n/a" — đó là placeholder trống của user, không phải
        dữ liệu bẩn cần cảnh báo (log sẽ nhiễu report).
        """
        if value is None:
            return [], []
        s = str(value).strip()
        if not s:
            return [], []

        # Map lower → canonical version của VALID_STATUSES để log giữ chuẩn "Closed"
        # thay vì "closed" (dù user gõ "CLOSED" hay "closed" đều map về "Closed").
        blacklist_canonical = {st.lower(): st for st in VALID_STATUSES}
        parts = re.split(r'[,;+\n]+', s)
        result: list[str] = []
        dropped: list[dict] = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            low = p.lower()
            if low in ("-", "n/a"):
                # Placeholder trống, không log — xem docstring
                continue
            if low in blacklist_canonical:
                dropped.append({
                    "raw_value": p,
                    "matched_keyword": blacklist_canonical[low],
                })
                continue
            result.append(p)
        return result, dropped

    def _normalize_pic_names(self, rows: list[FunctionRow]) -> None:
        """
        Normalize tên PIC case-insensitively.
        VD: "SonHN6" + "SONHN6" → chuẩn hóa thành "SonHN6".
        Ưu tiên version có mix case (không phải toàn upper hoặc toàn lower).
        """
        # Collect all PIC names
        all_pics: set[str] = set()
        for r in rows:
            for pd in r.phases.values():
                all_pics.update(pd.pics)

        if not all_pics:
            return

        # Group theo lowercase
        groups: dict[str, list[str]] = {}
        for p in all_pics:
            key = p.lower()
            groups.setdefault(key, []).append(p)

        # Chọn canonical: ưu tiên (1) không phải all-upper, (2) alphabet đầu tiên
        canonical_map: dict[str, str] = {}
        for key, variants in groups.items():
            if len(variants) == 1:
                canonical_map[key] = variants[0]
            else:
                # Sort: prefer version có ít nhất 1 chữ thường (không phải ALL CAPS)
                variants.sort(key=lambda x: (x.isupper(), x))
                canonical_map[key] = variants[0]

        # Apply mapping
        for r in rows:
            for pd in r.phases.values():
                pd.pics = [canonical_map.get(p.lower(), p) for p in pd.pics]
