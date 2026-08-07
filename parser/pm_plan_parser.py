# -*- coding: utf-8 -*-
"""
Parser kế hoạch dự án PM (KeHoachDuAn Excel).

Sheet roles (auto-propose):
  gantt          — WBS / milestone (Gantt-chart)
  gantt_old      — bản cũ (bỏ qua khi parse chính, vẫn liệt kê)
  schedule       — lịch trình chi tiết (UAT/Golive): ngày + PIC
  deliverables   — sản phẩm bàn giao
  team_vendor    — đội FPT
  team_client    — đội khách hàng
  ignore         — bỏ qua
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any, Optional

import openpyxl

# Vai trò sheet chuẩn cho mapping UI
SHEET_ROLES = [
    "gantt",
    "gantt_old",
    "schedule",
    "deliverables",
    "team_vendor",
    "team_client",
    "ignore",
]

ROLE_LABELS_VI = {
    "gantt": "Gantt / WBS (chính)",
    "gantt_old": "Gantt cũ (bỏ qua parse)",
    "schedule": "Lịch trình chi tiết (ngày + PIC)",
    "deliverables": "Sản phẩm bàn giao",
    "team_vendor": "Đội dự án FPT (vendor)",
    "team_client": "Đội dự án khách hàng",
    "ignore": "Bỏ qua",
}


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _to_iso_date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s or s.upper() in ("#REF!", "#N/A", "#VALUE!", "N/A"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _split_pics(raw: Any) -> list[str]:
    """Tách PIC giống Function List: dấu phẩy / ; / xuống dòng."""
    if raw is None or raw == "":
        return []
    text = str(raw).replace("\r\n", "\n").replace("\r", "\n")
    parts = re.split(r"[,;\n]+", text)
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        s = p.strip().lstrip("+").strip()
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return out


def propose_sheet_mapping(sheet_names: list[str]) -> dict[str, str]:
    """Gợi ý role cho từng sheet theo keyword tên."""
    mapping: dict[str, str] = {}
    for name in sheet_names:
        n = (name or "").strip().lower()
        role = "ignore"
        if "gantt" in n and "old" in n:
            role = "gantt_old"
        elif "gantt" in n:
            role = "gantt"
        elif any(k in n for k in ("chi tiết", "chi tiet", "thực hiện", "thuc hien", "detail")):
            role = "schedule"
        elif any(k in n for k in ("khung", "master plan", "masterplan")):
            role = "gantt"
        elif any(k in n for k in ("lịch trình", "lich trinh", "uat", "golive")):
            role = "schedule"
        elif any(k in n for k in ("sản phẩm", "san pham", "bàn giao", "ban giao", "deliverable")):
            role = "deliverables"
        elif any(k in n for k in ("đội", "doi ", "team")) and "fpt" in n:
            role = "team_vendor"
        elif any(k in n for k in ("đội", "doi ", "team")) and "fpt" not in n:
            role = "team_client"
        elif "fpt" in n and any(k in n for k in ("đội", "doi", "nhân sự", "nhan su")):
            role = "team_vendor"
        mapping[name] = role

    # Đảm bảo có đúng 1 gantt chính nếu có nhiều
    gantts = [s for s, r in mapping.items() if r == "gantt"]
    if len(gantts) > 1:
        # Giữ sheet không có "old", còn lại → gantt_old
        for s in gantts:
            if "old" in s.lower():
                mapping[s] = "gantt_old"
    return mapping


def preview_plan_workbook(filepath: str) -> dict[str, Any]:
    """Đọc tên sheet + vài dòng mẫu + đề xuất mapping (chưa parse đầy đủ)."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            sample: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(max_row=6, max_col=10, values_only=True)):
                sample.append([_cell_str(c) if c is not None else "" for c in row])
                if i >= 5:
                    break
            sheets.append({"name": name, "sample": sample})
    finally:
        wb.close()
    mapping = propose_sheet_mapping([s["name"] for s in sheets])
    return {
        "sheets": sheets,
        "proposed_mapping": mapping,
        "role_labels": ROLE_LABELS_VI,
        "roles": SHEET_ROLES,
    }


def parse_plan(
    filepath: str,
    sheet_mapping: Optional[dict[str, str]] = None,
) -> dict[str, Any]:
    """
    Parse KeHoachDuAn → dict chuẩn hoá.

    sheet_mapping: {sheet_name: role}; nếu None → auto-propose.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        names = list(wb.sheetnames)
        mapping = sheet_mapping or propose_sheet_mapping(names)
        # Chỉ map sheet tồn tại
        mapping = {k: v for k, v in mapping.items() if k in names}

        result: dict[str, Any] = {
            "sheet_mapping": mapping,
            "sheet_names": names,
            "weeks": [],
            "milestones": [],
            "schedule": [],
            "deliverables": [],
            "team_vendor": [],
            "team_client": [],
        }

        for sheet_name, role in mapping.items():
            if role in ("ignore", "gantt_old"):
                continue
            ws = wb[sheet_name]
            if role == "gantt":
                weeks, milestones = _parse_gantt(ws)
                result["weeks"] = weeks
                result["milestones"] = milestones
            elif role == "schedule":
                meta = _scan_project_meta(ws)
                if meta.get("project_start"):
                    result["project_start"] = meta["project_start"]
                sched, day_cols = _parse_schedule(ws)
                result["schedule"] = sched
                if day_cols:
                    result["day_columns"] = day_cols
            elif role == "deliverables":
                result["deliverables"] = _parse_deliverables(ws)
            elif role == "team_vendor":
                result["team_vendor"] = _parse_team(ws, side="vendor")
            elif role == "team_client":
                result["team_client"] = _parse_team(ws, side="client")

        result["summary"] = {
            "milestone_count": len(result["milestones"]),
            "schedule_count": len(result["schedule"]),
            "deliverable_count": len(
                [d for d in result["deliverables"] if not d.get("is_group")]
            ),
            "team_vendor_count": len(result["team_vendor"]),
            "team_client_count": len(result["team_client"]),
        }
        return result
    finally:
        wb.close()


def _parse_gantt(ws) -> tuple[list[dict], list[dict]]:
    """Đọc WBS: STT + Công việc + danh sách tuần (W1..). Thanh Gantt thường là shape → không lấy được week_start/end từ fill."""
    weeks: list[dict] = []
    # Row 2 = tháng, row 3 = W1..; hoặc tìm row có W1
    week_row = None
    month_row = None
    for r in range(1, 8):
        vals = [_cell_str(ws.cell(r, c).value) for c in range(1, min(60, (ws.max_column or 1) + 1))]
        if any(re.fullmatch(r"W\d+", v) for v in vals):
            week_row = r
            month_row = r - 1 if r > 1 else None
            break
        # Khung tuần kiểu Vietinak: hàng số 1,2,3… (không có W prefix)
        num_cells = [v for v in vals if re.fullmatch(r"\d+", v)]
        if len(num_cells) >= 4:
            week_row = r
            month_row = r - 2 if r > 2 else (r - 1 if r > 1 else None)
            break
    if week_row:
        current_month = ""
        for c in range(1, min(60, (ws.max_column or 1) + 1)):
            wlabel = _cell_str(ws.cell(week_row, c).value)
            if month_row:
                m = _cell_str(ws.cell(month_row, c).value)
                if m:
                    current_month = m
            label = None
            if re.fullmatch(r"W\d+", wlabel):
                label = wlabel
            elif re.fullmatch(r"\d+", wlabel):
                label = f"W{wlabel}"
            if label:
                weeks.append({"col": c, "label": label, "month": current_month})

    milestones: list[dict] = []
    start_r = (week_row or 3) + 1
    for r in range(start_r, (ws.max_row or 0) + 1):
        stt = ws.cell(r, 1).value
        name = _cell_str(ws.cell(r, 2).value)
        if not name:
            continue
        # Bỏ header lặp
        if name.lower() in ("công việc", "cong viec"):
            continue
        milestones.append({
            "stt": _cell_str(stt) if stt is not None else "",
            "name": name,
            "week_start": None,
            "week_end": None,
            "note": "Thanh Gantt (shape) — tuần chi tiết xem sheet Lịch trình",
        })
    return weeks, milestones


def _scan_project_meta(ws) -> dict[str, Any]:
    """Đọc Project Start Date / meta từ vùng header sheet lịch trình."""
    meta: dict[str, Any] = {}
    for r in range(1, 15):
        for c in range(1, 20):
            label = _cell_str(ws.cell(r, c).value).lower()
            if not label:
                continue
            if "project start" in label or (
                ("ngày bắt đầu" in label or "ngay bat dau" in label)
                and ("dự án" in label or "du an" in label)
            ):
                for nc in (c + 1, c):
                    d = _to_iso_date(ws.cell(r, nc).value)
                    if d:
                        meta["project_start"] = d
                        return meta
    return meta


from parser.excel_parser import STATUS_ALIASES


def _normalize_pm_status(raw: Any) -> Optional[str]:
    """Chuẩn hóa trạng thái PM plan → canonical (Open, Closed, …)."""
    s = _cell_str(raw)
    if not s:
        return None
    low = s.lower().strip()
    if low == "overdue":
        return "In-progress"
    alias = STATUS_ALIASES.get(low)
    if alias:
        return alias
    # Title-case fallback cho giá trị đã canonical
    for canon in ("Open", "Assigned", "In-progress", "Resolved", "Closed", "Pending", "Cancelled"):
        if low == canon.lower():
            return canon
    return s


def _parse_percent(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        v = float(raw)
        if 0 <= v <= 1:
            return round(v * 100, 1)
        if 0 <= v <= 100:
            return round(v, 1)
        return None
    s = _cell_str(raw).replace("%", "").replace(",", ".").strip()
    if not s:
        return None
    try:
        v = float(s)
        if 0 <= v <= 1:
            return round(v * 100, 1)
        if 0 <= v <= 100:
            return round(v, 1)
    except ValueError:
        return None
    return None


def _detect_day_columns(ws, header_row: int) -> list[dict]:
    """Cột ngày M/T/W… trên sheet chi tiết — date ở hàng ngay trên header."""
    day_columns: list[dict] = []
    date_row = header_row - 1
    if date_row < 1:
        return day_columns
    max_c = min((ws.max_column or 1) + 1, 220)
    for c in range(1, max_c):
        lbl = _cell_str(ws.cell(header_row, c).value).upper()
        if lbl not in ("M", "T", "W", "F", "S"):
            continue
        dt = _to_iso_date(ws.cell(date_row, c).value)
        if not dt:
            continue
        day_columns.append({"col": c, "label": lbl, "date": dt})
    return day_columns


def _parse_schedule(ws) -> tuple[list[dict], list[dict]]:
    """Lịch trình: Công việc + ngày + PIC; hỗ trợ thêm Status / Actual End / % (MasterPlan Detail)."""
    header_row = None
    col_map: dict[str, int] = {}
    max_scan_row = min(20, (ws.max_row or 1) + 1)
    max_scan_col = min(25, (ws.max_column or 1) + 1)

    for r in range(1, max_scan_row):
        row_vals = {
            c: _cell_str(ws.cell(r, c).value).replace("\n", " ")
            for c in range(1, max_scan_col)
        }
        joined = " | ".join(row_vals.values()).lower()
        has_task = "công việc" in joined or "cong viec" in joined or "task" in joined
        has_date = any(
            k in joined
            for k in ("ngày", "ngay", "bắt đầu", "bat dau", "kết thúc", "ket thuc", "start", "end")
        )
        if has_task and has_date:
            header_row = r
            for c, v in row_vals.items():
                vl = v.lower().strip()
                if not vl:
                    continue
                if vl in ("stt", "#", "no", "no."):
                    col_map["stt"] = c
                elif "công việc" in vl or "cong viec" in vl or vl == "task":
                    col_map["name"] = c
                elif "từ ngày" in vl or "tu ngay" in vl or vl == "start" or vl == "from":
                    col_map["start"] = c
                elif "đến ngày" in vl or "den ngay" in vl or vl in ("end", "to"):
                    col_map["end"] = c
                elif vl == "bắt đầu" or vl == "bat dau" or "bắt đầu" in vl:
                    if "start" not in col_map:
                        col_map["start"] = c
                elif vl == "kết thúc" or vl == "ket thuc" or "kết thúc" in vl:
                    if "end" not in col_map:
                        col_map["end"] = c
                elif "trạng thái" in vl or "trang thai" in vl or vl == "status" or vl == "tt":
                    col_map["status"] = c
                elif "ngày hoàn thành" in vl or "ngay hoan thanh" in vl or "actual end" in vl:
                    col_map["actual_end"] = c
                elif "%" in vl and ("hoàn" in vl or "hoan" in vl or "complete" in vl):
                    col_map["percent_complete"] = c
                elif "phụ trách" in vl and "fpt" in vl:
                    col_map["pic_fpt"] = c
                elif "hỗ trợ" in vl and "fpt" in vl:
                    col_map["support_fpt"] = c
                elif "phụ trách" in vl and ("mphg" in vl or "khách" in vl or "client" in vl):
                    col_map["pic_client"] = c
                elif "hỗ trợ" in vl and ("mphg" in vl or "khách" in vl or "client" in vl):
                    col_map["support_client"] = c
                elif vl == "fpt" and "pic_fpt" not in col_map:
                    col_map["pic_fpt"] = c
                elif vl in ("vietinak", "client", "khách hàng", "khach hang", "mphg") and "pic_client" not in col_map:
                    col_map["pic_client"] = c
                elif "ghi chú" in vl or "ghi chu" in vl or vl == "note":
                    col_map["note"] = c
            break

    if not header_row:
        header_row = 2
        col_map = {
            "name": 1, "start": 2, "end": 3,
            "pic_fpt": 4, "support_fpt": 5,
            "pic_client": 6, "support_client": 7, "note": 8,
        }

    def _cell_val(row: int, key: str) -> Any:
        if key not in col_map:
            return None
        return ws.cell(row, col_map[key]).value

    items: list[dict] = []
    current_phase = ""
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        name = _cell_str(_cell_val(r, "name"))
        if not name:
            continue

        stt_raw = _cell_val(r, "stt")
        stt_s = _cell_str(stt_raw)
        start = _to_iso_date(_cell_val(r, "start"))
        end = _to_iso_date(_cell_val(r, "end"))
        status = _normalize_pm_status(_cell_val(r, "status"))
        actual_end = _to_iso_date(_cell_val(r, "actual_end"))
        percent_complete = _parse_percent(_cell_val(r, "percent_complete"))

        pic_fpt = _split_pics(_cell_val(r, "pic_fpt")) if "pic_fpt" in col_map else []
        support_fpt = _split_pics(_cell_val(r, "support_fpt")) if "support_fpt" in col_map else []
        pic_client = _split_pics(_cell_val(r, "pic_client")) if "pic_client" in col_map else []
        support_client = _split_pics(_cell_val(r, "support_client")) if "support_client" in col_map else []
        note = _cell_str(_cell_val(r, "note"))

        # PIC cột org tag (FPT / Vietinak)
        if "pic_fpt" in col_map and not pic_fpt:
            org = _cell_str(_cell_val(r, "pic_fpt"))
            if org:
                pic_fpt = [org]
        if "pic_client" in col_map and not pic_client:
            org = _cell_str(_cell_val(r, "pic_client"))
            if org:
                pic_client = [org]

        is_section_letter = bool(re.fullmatch(r"[A-Z]", stt_s))
        is_phase = bool(re.match(r"(?i)^giai\s*đoạn|^giai\s*doan|^phase\b", name))
        has_any_pic = bool(pic_fpt or support_fpt or pic_client or support_client)
        is_header = is_section_letter or is_phase or (
            start and end and not has_any_pic and name.lower().startswith("giai")
        )

        if is_header:
            current_phase = name
            items.append({
                "name": name,
                "stt": stt_s,
                "start": start,
                "end": end,
                "status": status,
                "actual_end": actual_end,
                "percent_complete": percent_complete,
                "pic_fpt": [],
                "support_fpt": [],
                "pic_client": [],
                "support_client": [],
                "note": note,
                "is_phase_header": True,
                "phase": name,
            })
            continue

        items.append({
            "name": name,
            "stt": stt_s,
            "start": start,
            "end": end,
            "status": status,
            "actual_end": actual_end,
            "percent_complete": percent_complete,
            "pic_fpt": pic_fpt,
            "support_fpt": support_fpt,
            "pic_client": pic_client,
            "support_client": support_client,
            "note": note,
            "is_phase_header": False,
            "phase": current_phase,
        })
    day_columns = _detect_day_columns(ws, header_row) if header_row else []
    return items, day_columns


def _parse_deliverables(ws) -> list[dict]:
    """Sản phẩm bàn giao — group header (STT text) + item rows."""
    header_row = None
    for r in range(1, 6):
        vals = [_cell_str(ws.cell(r, c).value).lower() for c in range(1, 12)]
        if any("sản phẩm" in v or "san pham" in v or "deliverable" in v for v in vals):
            header_row = r
            break
        if "stt" in vals and any("bàn giao" in v or "ban giao" in v for v in vals):
            header_row = r
            break
    if not header_row:
        header_row = 2

    items: list[dict] = []
    current_group = ""
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        stt_raw = ws.cell(r, 1).value
        name = _cell_str(ws.cell(r, 2).value)
        stt_s = _cell_str(stt_raw)
        # Group: có text ở col1, không có name hoặc name trống
        if stt_s and not str(stt_raw).replace(".", "").isdigit() and not name:
            current_group = stt_s
            items.append({
                "stt": "",
                "name": stt_s,
                "is_group": True,
                "group": stt_s,
                "due_date": None,
                "type": "",
                "note": "",
                "author": "",
                "reviewer_fpt": "",
                "approver_fpt": "",
            })
            continue
        if not name:
            continue
        items.append({
            "stt": stt_s,
            "name": name,
            "is_group": False,
            "group": current_group,
            "due_date": _to_iso_date(ws.cell(r, 3).value),
            "type": _cell_str(ws.cell(r, 4).value),
            "note": _cell_str(ws.cell(r, 5).value),
            "author": _cell_str(ws.cell(r, 6).value),
            "reviewer_fpt": _cell_str(ws.cell(r, 7).value),
            "approver_fpt": _cell_str(ws.cell(r, 8).value),
        })
    return items


def _parse_team(ws, side: str = "vendor") -> list[dict]:
    """Đội dự án — group + members (Họ tên, Chức vụ, Trách nhiệm, Email)."""
    header_row = None
    for r in range(1, 5):
        vals = [_cell_str(ws.cell(r, c).value).lower() for c in range(1, 8)]
        if any("họ" in v or "tên" in v or "name" in v for v in vals):
            header_row = r
            break
    if not header_row:
        header_row = 1

    # Map columns loosely
    headers = {
        c: _cell_str(ws.cell(header_row, c).value).lower()
        for c in range(1, min(8, (ws.max_column or 1) + 1))
    }
    col_name = col_title = col_role = col_resp = col_email = None
    for c, h in headers.items():
        if "họ" in h or "tên" in h or h == "name":
            col_name = c
        elif "chức vụ" in h or "chuc vu" in h or "title" in h:
            col_title = c
        elif "vai trò" in h or "vai tro" in h or h == "role":
            col_role = c
        elif "trách nhiệm" in h or "trach nhiem" in h:
            col_resp = c
        elif "email" in h:
            col_email = c
    col_name = col_name or 2
    col_title = col_title or 3
    col_resp = col_resp or (5 if side == "client" else 4)
    col_email = col_email or (6 if side == "client" else 5)

    members: list[dict] = []
    current_group = ""
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        stt_raw = ws.cell(r, 1).value
        name = _cell_str(ws.cell(r, col_name).value)
        stt_s = _cell_str(stt_raw)
        if stt_s and not str(stt_raw).replace(".", "").isdigit() and not name:
            current_group = stt_s
            continue
        if not name:
            continue
        members.append({
            "group": current_group,
            "stt": stt_s,
            "name": name,
            "title": _cell_str(ws.cell(r, col_title).value) if col_title else "",
            "role": _cell_str(ws.cell(r, col_role).value) if col_role else "",
            "responsibility": _cell_str(ws.cell(r, col_resp).value) if col_resp else "",
            "email": _cell_str(ws.cell(r, col_email).value) if col_email else "",
            "side": side,
        })
    return members
