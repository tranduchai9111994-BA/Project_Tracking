"""
iHRP Function List Tracker — Flask Server (V2 + Projects)

V2 features:
- Advanced analytics: Unassigned, Duration, Stalled, Risk Score, Effort, Process, Timeline
- Snapshot & Compare mode per project
- Multi-project support: mỗi khách hàng/dự án 1 workspace riêng
- Auto-migrate snapshot cũ vào project "Default"
- Cross-project compare (Phase 2)
- Export/Import project package (zip)
- Multi-sheet export, export-by-PIC, export-compare
- Trim response payload cho tốc độ (top N items thay vì all)
"""
import io
import os
import re
import sys
import shutil
import tempfile
import zipfile
from datetime import date, datetime
from typing import Any, Optional

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for

from parser.excel_parser import FunctionListParser
from analyzer.dashboard_engine import DashboardEngine
from analyzer.snapshot_manager import SnapshotManager
from analyzer.compare_engine import CompareEngine
from analyzer.project_manager import ProjectManager, Project
from analyzer.drill_down import drill_down as drill_down_fn, build_title as build_drill_title, SUPPORTED_CHARTS
from analyzer.portfolio import (
    search_across_projects,
    compare_projects,
    aggregate_rollup,
    rollup_summary_override,
    SEARCH_SCOPES,
)
from exporter.excel_exporter import (
    export_overdue_report,
    export_stalled_report,
    export_full_report,
    export_by_pic,
    export_compare_report,
    export_drill_down,
    export_pic_blacklist_report,
    export_portfolio_compare,
    export_chart,
    export_audit_report,
    export_sla_report,
    export_capacity_report,
    export_slow_heatmap_report,
    export_baseline_variance_report,
    export_fitgap_report,
    export_function_diff_report,
    SUPPORTED_EXPORT_CHARTS,
)
from exporter.weekly_mom import export_weekly_mom
from exporter.pm_exporter import export_pm_report

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["PROJECTS_FOLDER"] = os.path.join(app.config["UPLOAD_FOLDER"], "projects")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["PROJECTS_FOLDER"], exist_ok=True)

# --------------------------------------------------------------------------
# Session auth — bắt buộc đăng nhập trước khi vào dashboard / API nội bộ.
# Public API (/public/*, /embed/*) vẫn dùng token riêng (không đụng session).
# Tắt tạm: IHRP_DISABLE_AUTH=1 (chỉ debug).
# --------------------------------------------------------------------------
from analyzer import auth_store as _auth_store

app.secret_key = _auth_store.ensure_secret_key()
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# Tạo admin/admin nếu chưa có user nào
if _auth_store.ensure_default_admin():
    print(
        "[auth] Đã tạo tài khoản mặc định admin / admin — hãy đổi mật khẩu ngay.",
        file=sys.stderr,
    )

# --------------------------------------------------------------------------
# Cache-busting cho static assets — mtime của file thành ?v= query param.
# Khi dev sửa dashboard.js / style.css, mtime đổi → browser tự fetch bản mới,
# tránh bug UI "hiện số 0" do trình duyệt cache JS/CSS cũ.
# --------------------------------------------------------------------------
_STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")


def _static_ver(rel_path: str) -> str:
    """Trả về mtime của file static để làm cache-busting query param."""
    full = os.path.join(_STATIC_DIR, rel_path.replace("/", os.sep))
    try:
        return str(int(os.path.getmtime(full)))
    except OSError:
        return "0"


@app.context_processor
def _inject_static_ver():
    return {"static_ver": _static_ver}

# ==========================================================================
# Global state — mỗi project 1 slot, cache metrics trong memory
# ==========================================================================
# App chạy local single-user nên dict trong memory là đủ.
# Cache TTL implicit: cho đến khi user upload file mới hoặc chuyển project.
_state: dict[str, dict[str, Any]] = {}
# _state[slug] = {"data": ParsedData, "metrics": dict, "filename": str, "upload_time": datetime}

_project_mgr = ProjectManager(app.config["PROJECTS_FOLDER"])
# Đảm bảo tồn tại project "default" ngay khi khởi động
_project_mgr.get_or_create_default()

# --------------------------------------------------------------------------
# T34 Task 2 — LAN secure self-host: cài admin guard + access log ngay từ
# đầu để mọi request đều đi qua.
# --------------------------------------------------------------------------
from analyzer import lan_security as _lansec

# Path log — nằm trong .project_store để user backup cùng data
_ACCESS_LOG_PATH = os.path.join(
    os.path.dirname(__file__), ".project_store", "access.log",
)
# Cho phép override qua env (VD tester chỉ định temp dir)
_ACCESS_LOG_PATH = os.environ.get("IHRP_ACCESS_LOG", _ACCESS_LOG_PATH)

# Admin guard — chặn POST/PUT/DELETE từ non-localhost. Tắt bằng env
# IHRP_DISABLE_ADMIN_GUARD=1 (không nên dùng, chỉ để debug).
if os.environ.get("IHRP_DISABLE_ADMIN_GUARD", "").strip() != "1":
    _lansec.install_admin_guard(app)

# Access log — luôn bật nhưng có thể tắt qua env IHRP_DISABLE_ACCESS_LOG=1.
if os.environ.get("IHRP_DISABLE_ACCESS_LOG", "").strip() != "1":
    _lansec.install_access_log(app, _ACCESS_LOG_PATH)

# Session login gate — bảo vệ UI + /api/* (trừ auth endpoints).
# Public API / embed vẫn token-based. Tắt: IHRP_DISABLE_AUTH=1.
_AUTH_DISABLED = os.environ.get("IHRP_DISABLE_AUTH", "").strip() == "1"
_AUTH_PUBLIC_PREFIXES = (
    "/static/",
    "/public/",
    "/embed/",
)
_AUTH_PUBLIC_EXACT = {
    "/login",
    "/logout",
    "/api/auth/login",
    "/api/health",
}


def _auth_current_user() -> Optional[dict]:
    """User từ Flask session (đã verify id còn tồn tại)."""
    uid = session.get("user_id")
    if not uid:
        return None
    user = _auth_store.get_user_by_id(uid)
    if not user:
        session.clear()
        return None
    return user


def _auth_is_public_path(path: str) -> bool:
    if path in _AUTH_PUBLIC_EXACT:
        return True
    for prefix in _AUTH_PUBLIC_PREFIXES:
        if path.startswith(prefix):
            return True
    return False


if not _AUTH_DISABLED:
    @app.before_request
    def _session_login_gate():
        # OPTIONS preflight (CORS public API) — không yêu cầu session
        if request.method == "OPTIONS":
            return None
        path = request.path or ""
        if _auth_is_public_path(path):
            return None
        if _auth_current_user():
            return None
        # HTML page → redirect login; API / XHR → 401 JSON
        wants_html = (
            path == "/"
            or (not path.startswith("/api/") and "text/html" in (request.accept_mimetypes.best or ""))
        )
        if wants_html and request.method == "GET":
            return redirect(url_for("login_page", next=path))
        return jsonify({
            "error": "Chưa đăng nhập.",
            "code": "AUTH_REQUIRED",
            "login_url": "/login",
        }), 401

# Phase 7 Slim: dọn export / snapshot / synced tạm / PM PPTX trùng khi start
try:
    from analyzer.disk_janitor import (
        purge_old_exports,
        purge_excess_snapshots,
        purge_excess_synced_all,
        purge_duplicate_pm_weekly_all,
        MAX_SYNCED_XLSX,
    )
    _n_exp = purge_old_exports(app.config["PROJECTS_FOLDER"], max_age_days=7)
    _n_snap = 0
    for _slug_name in os.listdir(app.config["PROJECTS_FOLDER"]):
        _snap_dir = os.path.join(app.config["PROJECTS_FOLDER"], _slug_name, "snapshots")
        _n_snap += purge_excess_snapshots(_snap_dir, keep=15)
    _n_synced = purge_excess_synced_all(
        app.config["PROJECTS_FOLDER"], keep=MAX_SYNCED_XLSX,
    )
    _n_pptx = purge_duplicate_pm_weekly_all(app.config["PROJECTS_FOLDER"])
    if _n_exp or _n_snap or _n_synced or _n_pptx:
        print(
            f"[janitor] Đã xóa {_n_exp} export, {_n_snap} snapshot, "
            f"{_n_synced} synced_*.xlsx, {_n_pptx} PM PPTX trùng",
            file=sys.stderr,
        )
except Exception as _janitor_err:
    print(f"[janitor] Bỏ qua: {_janitor_err}", file=sys.stderr)


def _run_startup_digest_scheduler() -> None:
    """T26 — Chạy cron-lite digest ngay khi start Flask (safe: không block start)."""
    try:
        from analyzer import digest as _digest_mod
        # State loader dùng chung với dashboard endpoints, gọi _get_state đã cache
        results = _digest_mod.run_scheduler(_project_mgr, _get_state)
        for r in results:
            if r.get("status") == "ok":
                print(f"[digest] Đã sinh {r['filename']} cho project {r['slug']}", file=sys.stderr)
    except Exception as _digest_err:
        print(f"[digest] Bỏ qua: {_digest_err}", file=sys.stderr)


def _run_startup_auto_archive() -> None:
    """
    T-AA — Background thread: auto-archive snapshot cũ cho project có
    auto_run_on_startup=True. Log vào access.log + stderr.
    """
    import threading

    def _worker():
        try:
            from analyzer import project_store as ps
            from analyzer import archive_manager as am
            from analyzer import lan_security as lansec
            projects = _project_mgr.list_projects(include_archived=False)
            total_archived = 0
            total_purged = 0
            for proj in projects:
                try:
                    folder = _project_mgr.get_project_folder(proj.slug)
                    settings = ps.load_archive_settings(folder)
                    if not settings.get("enabled") or not settings.get("auto_run_on_startup"):
                        continue
                    smgr = _project_mgr.get_snapshot_manager(proj.slug)
                    days = int(settings.get("archive_after_days") or 0)
                    archived = am.auto_archive_project(smgr.dir, days=days) if days > 0 else []
                    purge_days = int(settings.get("purge_after_days") or 0)
                    purged = am.purge_archive(smgr.dir, days=purge_days) if purge_days > 0 else []
                    total_archived += len(archived)
                    total_purged += len(purged)
                    if archived or purged:
                        msg = (
                            f"[archive] project={proj.slug} "
                            f"archived={len(archived)} purged={len(purged)}"
                        )
                        print(msg, file=sys.stderr)
                        try:
                            # Ghi thêm dòng vào access.log (best-effort)
                            with open(_ACCESS_LOG_PATH, "a", encoding="utf-8") as lf:
                                lf.write(
                                    f"{datetime.now().isoformat(timespec='seconds')} "
                                    f"STARTUP ARCHIVE {proj.slug} "
                                    f"archived={len(archived)} purged={len(purged)}\n"
                                )
                        except OSError:
                            pass
                except Exception as e:
                    print(f"[archive] Bỏ qua {getattr(proj, 'slug', '?')}: {e}", file=sys.stderr)
            if total_archived or total_purged:
                print(
                    f"[archive] Tổng: archived={total_archived} purged={total_purged}",
                    file=sys.stderr,
                )
        except Exception as e:
            print(f"[archive] Startup worker lỗi: {e}", file=sys.stderr)

    t = threading.Thread(target=_worker, name="auto-archive", daemon=True)
    t.start()


# ==========================================================================
# Helpers
# ==========================================================================

# Trim payload để response nhẹ hơn ~40%. Frontend chỉ cần top N cho card overview.
PAYLOAD_LIMITS = {
    "risk_scores": 50,           # frontend hiển thị top 20
    "duration_items": 200,       # bảng chi tiết
    # Stalled: local Module filter rebuild transitions từ items trên FE —
    # trim thấp (200) làm lệch số APP (vd 15/21) so với funnel/matrix.
    "stalled_items": 5000,
    "unassigned_tasks": 300,     # bảng dài nhất, cần pagination FE
}


def _trim_payload(metrics: dict) -> dict:
    """
    Trả về bản copy metrics đã trim để giảm bandwidth.
    Tổng items và list gốc vẫn còn ở backend cho endpoint chi tiết.
    """
    trimmed = dict(metrics)  # shallow copy

    # Trim risk_scores → top 50
    rs = metrics.get("risk_scores")
    if isinstance(rs, list) and len(rs) > PAYLOAD_LIMITS["risk_scores"]:
        trimmed["risk_scores"] = rs[:PAYLOAD_LIMITS["risk_scores"]]
        trimmed["risk_scores_total"] = len(rs)

    # Trim duration_analysis.items → top 200
    da = metrics.get("duration_analysis")
    if isinstance(da, dict):
        items = da.get("items", [])
        if len(items) > PAYLOAD_LIMITS["duration_items"]:
            da_copy = dict(da)
            da_copy["items"] = items[:PAYLOAD_LIMITS["duration_items"]]
            da_copy["items_total"] = len(items)
            trimmed["duration_analysis"] = da_copy

    # Trim stalled_tasks.items → top 200
    st = metrics.get("stalled_tasks")
    if isinstance(st, dict):
        items = st.get("items", [])
        if len(items) > PAYLOAD_LIMITS["stalled_items"]:
            st_copy = dict(st)
            st_copy["items"] = items[:PAYLOAD_LIMITS["stalled_items"]]
            st_copy["items_total"] = len(items)
            trimmed["stalled_tasks"] = st_copy

    # Trim unassigned_tasks → top 300
    ua = metrics.get("unassigned_tasks", [])
    if isinstance(ua, list) and len(ua) > PAYLOAD_LIMITS["unassigned_tasks"]:
        trimmed["unassigned_tasks"] = ua[:PAYLOAD_LIMITS["unassigned_tasks"]]
        trimmed["unassigned_tasks_total"] = len(ua)

    return trimmed


def _resolve_slug(explicit_slug: Optional[str] = None) -> str:
    """
    Trả về slug hiện hành:
    - Nếu có explicit → dùng nó nếu tồn tại
    - Nếu không → query param `project` → default
    """
    if explicit_slug:
        if _project_mgr.project_exists(explicit_slug):
            return explicit_slug
        raise ValueError(f"Không tìm thấy project: {explicit_slug}")

    q = request.args.get("project")
    if q and _project_mgr.project_exists(q):
        return q
    return "default"


def _load_state_from_disk(slug: str) -> Optional[dict]:
    """
    Load state từ snapshot mới nhất của project nếu có, hoặc current.xlsx.
    Dùng khi server restart / sau sync invalidate cache.

    QUAN TRỌNG: Ưu tiên snapshot mới nhất (không phải current.xlsx).
    Sync với target_action=snapshot ghi đè pickle/index nhưng trước đây
    không luôn update current.xlsx → nếu ưu tiên current.xlsx, dashboard
    sau sync vẫn hiện data upload cũ (timestamp Upload cũ, overdue cũ).
    """
    proj = _project_mgr.get_project(slug)
    if not proj:
        return None

    smgr = _project_mgr.get_snapshot_manager(slug)
    snaps = smgr.list_snapshots()
    if snaps:
        latest = snaps[0]
        try:
            loaded = smgr.load_snapshot(latest["date"])
            if loaded and loaded.get("parsed") is not None:
                data = loaded["parsed"]
                _apply_module_order_to_data(slug, data)
                metrics = DashboardEngine().compute_all(data)
                meta = loaded.get("meta") or latest
                upload_time = None
                ut = meta.get("upload_time")
                if ut:
                    try:
                        upload_time = datetime.fromisoformat(str(ut))
                    except (TypeError, ValueError):
                        upload_time = None
                if upload_time is None:
                    pkl_name = meta.get("pickle") or latest.get("pickle") or ""
                    pkl_path = os.path.join(smgr.dir, pkl_name) if pkl_name else ""
                    if pkl_path and os.path.isfile(pkl_path):
                        upload_time = datetime.fromtimestamp(os.path.getmtime(pkl_path))
                    else:
                        upload_time = datetime.now()
                return {
                    "data": data,
                    "metrics": metrics,
                    "filename": meta.get("filename") or latest.get("filename") or "snapshot",
                    "upload_time": upload_time,
                }
        except Exception as e:
            import sys
            print(
                f"[project={slug}] Không load được snapshot '{latest.get('date')}': "
                f"{type(e).__name__}: {e} — fallback current.xlsx",
                file=sys.stderr,
            )

    # Fallback: current.xlsx (project chưa có snapshot / snapshot corrupt)
    current_path = _project_mgr.get_current_file_path(slug)
    if not os.path.isfile(current_path):
        return None

    try:
        data = FunctionListParser().parse(current_path)
        _apply_module_order_to_data(slug, data)
        metrics = DashboardEngine().compute_all(data)
        return {
            "data": data,
            "metrics": metrics,
            "filename": "current.xlsx",
            "upload_time": datetime.fromtimestamp(os.path.getmtime(current_path)),
        }
    except Exception as e:
        import sys
        print(
            f"[project={slug}] Không load được state từ '{current_path}': "
            f"{type(e).__name__}: {e}",
            file=sys.stderr,
        )
        return None


def _get_state(slug: str) -> Optional[dict]:
    """Lấy state (memory hoặc load from disk). Return None nếu chưa upload."""
    if slug in _state:
        return _state[slug]
    loaded = _load_state_from_disk(slug)
    if loaded:
        _state[slug] = loaded
    return loaded


def _need_state(slug: str):
    """Helper: nếu chưa có state → trả về error response."""
    st = _get_state(slug)
    if st is None:
        return None, (
            jsonify({
                "error": f"Project '{slug}' chưa có file. Upload Function List trước.",
                "code": "NO_FILE",
            }),
            404,
        )
    return st, None


def _filter_parsed_data(
    data,
    modules: Optional[list[str]] = None,
    processes: Optional[list[str]] = None,
    pics: Optional[list[str]] = None,
    project_codes: Optional[list[str]] = None,
    # Backward compat: 2 kwargs cũ (single string) — Wave 1 API + tests cũ
    module: str = "",
    process: str = "",
):
    """
    Tạo bản sao ParsedData chỉ chứa rows match module/process/pic/project_code filter.

    Semantics:
    - OR trong 1 chiều filter (VD: modules=[A,B] → module ∈ {A,B})
    - AND giữa các chiều (modules AND processes AND pics AND project_codes)
    - PIC match: row match nếu BẤT KỲ phase nào của row có PIC ∈ pics
    - project_codes: match meta.ma_du_an

    Args:
        data: ParsedData gốc
        modules: list module cần lọc (None/empty = không lọc chiều này)
        processes: list quy trình cần lọc
        pics: list PIC cần lọc
        project_codes: list Mã dự án cần lọc
        module, process: legacy single-value (Wave 1) — auto convert sang list

    Returns:
        ParsedData mới với subset rows (giữ nguyên phase structure).
    """
    from parser.excel_parser import ParsedData

    # Merge legacy single-value vào list (backward compat với call site cũ + tests cũ)
    mod_list = list(modules) if modules else []
    if module:
        mod_list.append(module)
    proc_list = list(processes) if processes else []
    if process:
        proc_list.append(process)
    pic_list = list(pics) if pics else []
    pc_list = list(project_codes) if project_codes else []

    # Dedupe + loại rỗng
    mod_set = {m for m in mod_list if m}
    proc_set = {p for p in proc_list if p}
    pic_set = {p for p in pic_list if p}
    pc_set = {p for p in pc_list if p}

    if not mod_set and not proc_set and not pic_set and not pc_set:
        return data  # short-circuit — không lọc

    def _match(row) -> bool:
        # AND giữa các chiều, OR trong mỗi chiều
        if mod_set and row.meta.get("module", "") not in mod_set:
            return False
        if proc_set and row.meta.get("quy_trinh", "") not in proc_set:
            return False
        if pc_set:
            code = str(row.meta.get("ma_du_an") or "").strip()
            if code not in pc_set:
                return False
        if pic_set:
            # Row match nếu có ÍT NHẤT 1 phase chứa ÍT NHẤT 1 PIC thuộc pic_set
            row_pics = set()
            for pd in row.phases.values():
                row_pics.update(pd.pics)
            if not (row_pics & pic_set):
                return False
        return True

    filtered_rows = [r for r in data.rows if _match(r)]

    # Recompute all_* fields (nhưng giữ nguyên phases từ data gốc để chart cấu trúc không đổi)
    # Module: giữ thứ tự từ data.all_modules (đã apply module_order) — KHÔNG re-alpha.
    from analyzer.module_order import sort_modules
    present_modules = {r.meta.get("module", "") for r in filtered_rows if r.meta.get("module")}
    all_modules = sort_modules(present_modules, data.all_modules)
    all_priorities = sorted({r.meta.get("priority", "") for r in filtered_rows if r.meta.get("priority")})
    all_complexities = sorted({r.meta.get("complexity", "") for r in filtered_rows if r.meta.get("complexity")})
    all_giai_doan = sorted({str(r.meta.get("giai_doan", "")) for r in filtered_rows if r.meta.get("giai_doan")})
    all_processes = sorted({r.meta.get("quy_trinh", "") for r in filtered_rows if r.meta.get("quy_trinh")})
    all_project_codes = sorted({
        str(r.meta.get("ma_du_an", "")).strip()
        for r in filtered_rows
        if r.meta.get("ma_du_an") and str(r.meta.get("ma_du_an")).strip()
    })

    pics_out = set()
    statuses_set = set()
    for r in filtered_rows:
        for pd in r.phases.values():
            pics_out.update(pd.pics)
            if pd.status:
                statuses_set.add(pd.status)

    return ParsedData(
        headers=data.headers,
        meta_columns=data.meta_columns,
        phase_groups=data.phase_groups,
        rows=filtered_rows,
        all_modules=all_modules,
        all_phases=data.all_phases,  # phase structure không đổi
        all_pics=sorted(pics_out),
        all_statuses=sorted(statuses_set),
        all_priorities=all_priorities,
        all_complexities=all_complexities,
        all_giai_doan=all_giai_doan,
        all_processes=all_processes,
        all_project_codes=all_project_codes,
        # Giữ data-quality log nhưng chỉ entry thuộc rows còn lại sau filter
        pic_blacklisted=_filter_dq_log(
            getattr(data, "pic_blacklisted", []) or [], filtered_rows
        ),
        estimate_mh_rejected=_filter_dq_log(
            getattr(data, "estimate_mh_rejected", []) or [], filtered_rows
        ),
    )


def _filter_dq_log(items: list[dict], filtered_rows) -> list[dict]:
    """Lọc data-quality log theo row_index còn lại sau filter rows."""
    keep_rows = {r.row_num for r in filtered_rows}
    return [it for it in items if it.get("row_index") in keep_rows]


def _project_to_dict(p: Project) -> dict:
    """Serialize Project + thêm số snapshot cho FE hiển thị nhanh."""
    smgr = _project_mgr.get_snapshot_manager(p.slug)
    snap_count = len(smgr.list_snapshots())
    return {
        "slug": p.slug,
        "name": p.name,
        "description": p.description,
        "created_at": p.created_at,
        "last_upload_at": p.last_upload_at,
        "is_archived": p.is_archived,
        "tags": p.tags,
        "snapshot_count": snap_count,
    }


# ==========================================================================
# Auth — login / logout / account management
# ==========================================================================

@app.route("/api/health", methods=["GET"])
def api_health():
    """Health check công khai (không cần login) — cho LAN / monitor."""
    return jsonify({"ok": True, "auth": not _AUTH_DISABLED})


@app.route("/login", methods=["GET", "POST"])
def login_page():
    """Trang đăng nhập (HTML). Đã login → về dashboard."""
    if not _AUTH_DISABLED and _auth_current_user():
        return redirect("/")
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        user = _auth_store.authenticate(username, password)
        if user:
            session.clear()
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            nxt = (request.args.get("next") or request.form.get("next") or "/").strip()
            if not nxt.startswith("/") or nxt.startswith("//"):
                nxt = "/"
            return redirect(nxt)
        error = "Sai tên đăng nhập hoặc mật khẩu."
    return render_template("login.html", error=error)


@app.route("/logout", methods=["GET", "POST"])
def logout_page():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    """JSON login — dùng cho test / client không form."""
    body = request.get_json(silent=True) or {}
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    user = _auth_store.authenticate(username, password)
    if not user:
        return jsonify({"error": "Sai tên đăng nhập hoặc mật khẩu.", "code": "AUTH_FAILED"}), 401
    session.clear()
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = user["role"]
    return jsonify({"success": True, "user": user})


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/auth/me", methods=["GET"])
def api_auth_me():
    if _AUTH_DISABLED:
        return jsonify({
            "user": {"id": None, "username": "dev", "role": "admin"},
            "auth_disabled": True,
        })
    user = _auth_current_user()
    if not user:
        return jsonify({"error": "Chưa đăng nhập.", "code": "AUTH_REQUIRED"}), 401
    return jsonify({"user": user, "auth_disabled": False})


@app.route("/api/auth/users", methods=["GET"])
def api_auth_list_users():
    """Admin: danh sách tài khoản (không kèm hash)."""
    user = _auth_current_user()
    if _AUTH_DISABLED:
        user = {"role": "admin"}
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Chỉ admin mới xem danh sách tài khoản.", "code": "FORBIDDEN"}), 403
    return jsonify({"users": _auth_store.list_users()})


@app.route("/api/auth/users", methods=["POST"])
def api_auth_create_user():
    """Admin: tạo user mới. Body: {username, password, role?}."""
    user = _auth_current_user()
    if _AUTH_DISABLED:
        user = {"role": "admin"}
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Chỉ admin mới tạo tài khoản.", "code": "FORBIDDEN"}), 403
    body = request.get_json(silent=True) or {}
    try:
        created = _auth_store.create_user(
            body.get("username") or "",
            body.get("password") or "",
            role=body.get("role") or "viewer",
        )
        return jsonify({"success": True, "user": created}), 201
    except _auth_store.AuthError as e:
        return jsonify({"error": str(e), "code": "AUTH_ERROR"}), e.status_code


@app.route("/api/auth/change-password", methods=["POST"])
def api_auth_change_password():
    """
    Đổi mật khẩu.

    - User thường: body {current_password, new_password} — đổi của chính mình.
    - Admin: body {user_id, new_password} — đổi cho user khác (không cần current).
      Hoặc bỏ user_id → đổi của chính mình (vẫn cần current_password).
    """
    me = _auth_current_user()
    if _AUTH_DISABLED:
        return jsonify({"error": "Auth đang tắt (IHRP_DISABLE_AUTH=1)."}), 400
    if not me:
        return jsonify({"error": "Chưa đăng nhập.", "code": "AUTH_REQUIRED"}), 401

    body = request.get_json(silent=True) or {}
    new_password = body.get("new_password") or ""
    target_id = (body.get("user_id") or "").strip()

    try:
        if target_id and target_id != me["id"]:
            if me.get("role") != "admin":
                return jsonify({
                    "error": "Chỉ admin mới đổi mật khẩu người khác.",
                    "code": "FORBIDDEN",
                }), 403
            _auth_store.change_password(target_id, new_password)
        else:
            _auth_store.change_password(
                me["id"],
                new_password,
                current_password=body.get("current_password") or "",
                require_current=True,
            )
        return jsonify({"success": True})
    except _auth_store.AuthError as e:
        return jsonify({"error": str(e), "code": "AUTH_ERROR"}), e.status_code


# ==========================================================================
# Frontend routing
# ==========================================================================

@app.route("/")
def index():
    return render_template("index.html")


# ==========================================================================
# Project CRUD
# ==========================================================================

@app.route("/api/projects", methods=["GET"])
def list_projects():
    """Danh sách project. Query ?include_archived=1 để hiện cả archived."""
    include_archived = request.args.get("include_archived") in ("1", "true", "yes")
    projects = _project_mgr.list_projects(include_archived=include_archived)
    return jsonify({
        "success": True,
        "projects": [_project_to_dict(p) for p in projects],
    })


@app.route("/api/projects", methods=["POST"])
def create_project():
    """Tạo project mới. Body JSON: {name, description}"""
    body = request.get_json(silent=True) or {}
    name = body.get("name", "").strip()
    desc = body.get("description", "").strip()
    if not name:
        return jsonify({"error": "Tên project không được rỗng"}), 400
    try:
        proj = _project_mgr.create_project(name, desc)
        return jsonify({"success": True, "project": _project_to_dict(proj)}), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/projects/<slug>", methods=["GET"])
def get_project(slug):
    proj = _project_mgr.get_project(slug)
    if not proj:
        return jsonify({"error": "Không tìm thấy project"}), 404
    return jsonify({"success": True, "project": _project_to_dict(proj)})


@app.route("/api/projects/<slug>", methods=["PUT"])
def rename_project(slug):
    """Rename hoặc update description. Body: {name, description}."""
    body = request.get_json(silent=True) or {}
    name = body.get("name", "").strip()
    desc = body.get("description")
    if not name:
        return jsonify({"error": "Tên project không được rỗng"}), 400
    ok = _project_mgr.rename_project(slug, name, desc)
    if not ok:
        return jsonify({"error": "Không tìm thấy project"}), 404
    return jsonify({"success": True, "project": _project_to_dict(_project_mgr.get_project(slug))})


@app.route("/api/projects/<slug>", methods=["DELETE"])
def delete_project(slug):
    """
    Xóa project. Query ?soft=1 → chỉ archive (khôi phục được).
    Không xóa được 'default' để tránh mất state.
    """
    if slug == "default":
        return jsonify({"error": "Không được xóa project 'default'"}), 400
    soft = request.args.get("soft") in ("1", "true", "yes")
    if soft:
        ok = _project_mgr.archive_project(slug, True)
    else:
        ok = _project_mgr.delete_project(slug)
        _state.pop(slug, None)  # Xóa cache
    if not ok:
        return jsonify({"error": "Không tìm thấy project"}), 404
    return jsonify({"success": True})


@app.route("/api/projects/<slug>/restore", methods=["POST"])
def restore_project(slug):
    ok = _project_mgr.archive_project(slug, False)
    if not ok:
        return jsonify({"error": "Không tìm thấy project"}), 404
    return jsonify({"success": True})


# ==========================================================================
# Upload & core metrics (per-project)
# ==========================================================================

def _upload_and_process(slug: str) -> tuple:
    """
    Nội bộ: nhận multipart/form-data với file, xử lý và lưu state.
    Return: (response_json, status_code)
    """
    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy file trong request"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Chưa chọn file"}), 400

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Chỉ hỗ trợ file .xlsx"}), 400

    threshold = request.args.get("threshold", default=3, type=int)

    # Lưu vào current.xlsx của project
    filepath = _project_mgr.get_current_file_path(slug)
    file.save(filepath)

    try:
        parser = FunctionListParser()
        data = parser.parse(filepath)
        _apply_module_order_to_data(slug, data)

        engine = DashboardEngine(long_duration_threshold=threshold)
        metrics = engine.compute_all(data)

        upload_time = datetime.now()
        _state[slug] = {
            "data": data,
            "metrics": metrics,
            "filename": file.filename,
            "upload_time": upload_time,
        }

        # Auto lưu snapshot vào folder snapshots của project
        smgr = _project_mgr.get_snapshot_manager(slug)
        smgr.save_snapshot(filepath, data, metrics, source="upload")

        # Cập nhật last_upload_at cho project
        _project_mgr.touch_last_upload(slug)

        # Upload history (meta only) + soft validation warnings
        from analyzer import project_store as ps
        from analyzer.audit_report import build_audit_issues
        import hashlib
        checksum = ""
        try:
            with open(filepath, "rb") as _bf:
                checksum = hashlib.md5(_bf.read()).hexdigest()
        except OSError:
            pass
        folder = _project_mgr.get_project_folder(slug)
        ps.append_upload_history(
            folder,
            filename=file.filename,
            row_count=len(data.rows),
            checksum=checksum,
            source="upload",
            extra={
                "modules": len(data.all_modules),
                "phases": len(data.all_phases),
            },
        )

        warnings: list[dict] = []
        if len(data.rows) == 0:
            warnings.append({"level": "critical", "code": "empty_rows", "message": "File không có dòng chức năng nào"})
        if not data.all_phases:
            warnings.append({"level": "critical", "code": "no_phases", "message": "Không phát hiện phase group nào (pattern 'Phase - Attr')"})
        rej = getattr(data, "estimate_mh_rejected", []) or []
        if data.rows and len(rej) / max(len(data.rows), 1) > 0.1:
            warnings.append({
                "level": "warning",
                "code": "estimate_reject_rate",
                "message": f"{len(rej)} giá trị Estimate MH bị loại (>10% số function)",
            })
        try:
            audit = build_audit_issues(data, metrics)
            if audit["summary"].get("missing_meta_count", 0) > 0:
                warnings.append({
                    "level": "info",
                    "code": "missing_meta",
                    "message": f"{audit['summary']['missing_meta_count']} function thiếu meta (module/priority/...)",
                })
        except Exception:
            pass

        settings = ps.load_project_settings(folder)

        current_meta = {
            "date": upload_time.date().isoformat(),
            "filename": file.filename,
            "total_functions": len(data.rows),
        }
        auto_diff = _build_auto_diff_summary(slug, data, current_meta)

        return jsonify({
            "success": True,
            "project": _project_to_dict(_project_mgr.get_project(slug)),
            "filename": file.filename,
            "rows_count": len(data.rows),
            "upload_time": upload_time.isoformat(),
            "metrics": _trim_payload(metrics),
            "snapshots": smgr.list_snapshots(),
            # Số token PIC bị parser blacklist — FE hiển thị badge count.
            # Không nằm trong metrics vì blacklist là data-quality info,
            # không cascade theo global filter.
            "pic_blacklist_count": len(getattr(data, "pic_blacklisted", []) or []),
            "warnings": warnings,
            "settings": settings,
            "auto_diff": auto_diff,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi parse file: {str(e)}"}), 500


@app.route("/api/projects/<slug>/upload", methods=["POST"])
def upload_to_project(slug):
    """Upload xlsx vào project cụ thể."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Không tìm thấy project"}), 404
    result = _upload_and_process(slug)
    return result if isinstance(result, tuple) else result


@app.route("/api/upload", methods=["POST"])
def upload_legacy():
    """Legacy: upload không chỉ định project → project 'default'."""
    slug = request.args.get("project", "default")
    if not _project_mgr.project_exists(slug):
        _project_mgr.get_or_create_default()
        slug = "default"
    return _upload_and_process(slug)


# ==========================================================================
# T32: Column Mapping Wizard — 2 endpoint mới (preview + confirm)
# ==========================================================================
# Flow:
#   1) POST /api/upload-preview: nhận file, LƯU vào uploads/tmp/<uuid>.xlsx
#      trả về headers + preview rows + suggestion + tmp_id.
#   2) User điều chỉnh mapping trong modal wizard.
#   3) POST /api/upload-confirm: nhận tmp_id + column_mapping + project_slug
#      → parse với mapping → save như _upload_and_process → trả dashboard payload.
#
# Preset (per-project): GET/POST/DELETE /api/projects/<slug>/mapping-presets
# lưu ở excel_mapping_presets.json để lần sau apply nhanh cho file cùng schema.

_TMP_UPLOAD_DIR = os.path.join(app.config["UPLOAD_FOLDER"], "tmp")
os.makedirs(_TMP_UPLOAD_DIR, exist_ok=True)


def _prune_old_tmp_uploads(max_age_hours: int = 24) -> None:
    """Xoá file .xlsx trong uploads/tmp/ cũ hơn N giờ để không tràn ổ."""
    import time as _time
    now = _time.time()
    cutoff = now - max_age_hours * 3600
    try:
        for name in os.listdir(_TMP_UPLOAD_DIR):
            fp = os.path.join(_TMP_UPLOAD_DIR, name)
            try:
                if os.path.isfile(fp) and os.path.getmtime(fp) < cutoff:
                    os.remove(fp)
            except OSError:
                pass
    except OSError:
        pass


@app.route("/api/upload-preview", methods=["POST"])
def upload_preview():
    """
    T32 — Upload file, đọc header + preview + fuzzy match suggestion.
    KHÔNG parse full data + KHÔNG save vào project — chỉ chuẩn bị cho
    modal Column Mapping Wizard.

    Response:
        {
          "tmp_id": "<uuid>",
          "filename": "original.xlsx",
          "sheet_name": "Function List",
          "headers": ["Function Code", "Function Name", ...],
          "preview_rows": [[...], [...], ...],
          "ihrp_columns": ["Mã CN", "Tên chức năng", ...],
          "auto_suggest": {"Mã CN": [{"header": "Function Code", "score": 0.87}, ...], ...},
          "presets": [...]  // nếu có project_slug trong query
        }
    """
    import uuid as _uuid
    from parser import column_mapping as cm_mod

    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy file trong request"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Chưa chọn file"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Chỉ hỗ trợ file .xlsx"}), 400

    _prune_old_tmp_uploads()
    tmp_id = _uuid.uuid4().hex[:16]
    tmp_path = os.path.join(_TMP_UPLOAD_DIR, f"{tmp_id}.xlsx")
    try:
        file.save(tmp_path)
    except OSError as e:
        return jsonify({"error": f"Không lưu được file tạm: {e}"}), 500

    headers, preview, sheet_name = cm_mod.read_headers_and_preview(tmp_path)
    if not headers:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        return jsonify({
            "error": "Không đọc được header từ file (sheet trống hoặc format lỗi)"
        }), 400

    suggestion = cm_mod.suggest_mapping(headers, cm_mod.IHRP_STANDARD_COLUMNS)

    # T34 Task 3 (A+B) — Sample preview + type inference cho mỗi header
    from analyzer.type_infer import infer_all_headers
    column_types = infer_all_headers(headers, preview)

    # Nếu FE có gửi project_slug trong query → trả kèm list preset đã lưu
    presets: list[dict] = []
    matched_preset = None
    fingerprint = cm_mod.header_fingerprint(headers)
    slug = (request.args.get("project_slug") or "").strip()
    if slug and _project_mgr.project_exists(slug):
        from analyzer import project_store as ps
        presets = ps.list_mapping_presets(_project_dir_for(slug))
        matched_preset = cm_mod.match_preset_by_fingerprint(presets, fingerprint)

    return jsonify({
        "success": True,
        "tmp_id": tmp_id,
        "filename": file.filename,
        "sheet_name": sheet_name,
        "headers": headers,
        "preview_rows": preview,
        "ihrp_columns": cm_mod.IHRP_STANDARD_COLUMNS,
        "auto_suggest": suggestion,
        "column_types": column_types,  # T34 Task 3 (A+B): {header: {type, badge, samples}}
        "presets": presets,
        "header_fingerprint": fingerprint,
        "matched_preset": matched_preset,  # U06 — auto-apply nếu khớp
    })


@app.route("/api/upload-confirm", methods=["POST"])
def upload_confirm():
    """
    T32 — Xác nhận + parse file đã upload-preview với column_mapping thủ công.

    Body JSON:
        {
          "tmp_id": "<uuid>",
          "project_slug": "<slug>",   // bắt buộc; project phải tồn tại
          "column_mapping": {"Mã CN": "Function Code", ...},   // optional
          "threshold": 3    // long duration threshold, optional
        }

    Nếu `column_mapping` rỗng → parser hoạt động ở chế độ auto-detect thuần
    (fallback tương thích cho file chuẩn iHRP).
    """
    from parser import column_mapping as cm_mod

    body = request.get_json(silent=True) or {}
    tmp_id = str(body.get("tmp_id") or "").strip()
    slug = str(body.get("project_slug") or "").strip()
    mapping = cm_mod.sanitize_column_mapping(body.get("column_mapping") or {})
    threshold = int(body.get("threshold") or 3)

    if not tmp_id:
        return jsonify({"error": "Thiếu 'tmp_id'"}), 400
    # Path traversal guard — tmp_id là uuid hex nên chỉ chấp nhận [a-f0-9]
    if not all(c in "abcdef0123456789" for c in tmp_id.lower()):
        return jsonify({"error": "tmp_id không hợp lệ"}), 400
    tmp_path = os.path.join(_TMP_UPLOAD_DIR, f"{tmp_id}.xlsx")
    if not os.path.isfile(tmp_path):
        return jsonify({
            "error": "File tạm không tồn tại (có thể đã hết hạn). Upload lại."
        }), 404
    if not slug or not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404

    # Copy tmp → project current.xlsx
    filepath = _project_mgr.get_current_file_path(slug)
    try:
        shutil.copy2(tmp_path, filepath)
    except OSError as e:
        return jsonify({"error": f"Không copy file: {e}"}), 500

    try:
        parser_obj = FunctionListParser()
        data = parser_obj.parse(filepath, column_mapping=mapping or None)
        _apply_module_order_to_data(slug, data)
        engine = DashboardEngine(long_duration_threshold=threshold)
        metrics = engine.compute_all(data)

        upload_time = datetime.now()
        original_name = str(body.get("filename") or f"synced_{tmp_id}.xlsx")
        _state[slug] = {
            "data": data,
            "metrics": metrics,
            "filename": original_name,
            "upload_time": upload_time,
        }

        smgr = _project_mgr.get_snapshot_manager(slug)
        smgr.save_snapshot(filepath, data, metrics, source="upload")
        _project_mgr.touch_last_upload(slug)

        # Cleanup tmp file (đã copy vào project)
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        return jsonify({
            "success": True,
            "project": _project_to_dict(_project_mgr.get_project(slug)),
            "filename": original_name,
            "rows_count": len(data.rows),
            "upload_time": upload_time.isoformat(),
            "metrics": _trim_payload(metrics),
            "snapshots": smgr.list_snapshots(),
            "pic_blacklist_count": len(getattr(data, "pic_blacklisted", []) or []),
            "column_mapping_applied": bool(mapping),
            "column_mapping_count": len(mapping),
            "auto_diff": _build_auto_diff_summary(slug, data, {
                "date": upload_time.date().isoformat(),
                "filename": original_name,
                "total_functions": len(data.rows),
            }),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi parse file: {str(e)}"}), 500


@app.route("/api/projects/<slug>/mapping-presets",
           methods=["GET", "POST"])
def project_mapping_presets(slug: str):
    """GET → list preset. POST → save preset {name, mapping}."""
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"presets": ps.list_mapping_presets(folder)})
    body = request.get_json(silent=True) or {}
    name = str(body.get("name") or "").strip()
    mapping = body.get("mapping") or {}
    if not name:
        return jsonify({"error": "Thiếu 'name'"}), 400
    if not isinstance(mapping, dict):
        return jsonify({"error": "'mapping' phải là object"}), 400
    try:
        presets = ps.save_mapping_preset(
            folder, name, mapping,
            fingerprint=str(body.get("fingerprint") or body.get("header_fingerprint") or ""),
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"presets": presets}), 201


@app.route("/api/projects/<slug>/mapping-presets/<name>",
           methods=["DELETE"])
def project_mapping_preset_delete(slug: str, name: str):
    """Xoá 1 preset theo name."""
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    deleted, presets = ps.delete_mapping_preset(folder, name)
    if not deleted:
        return jsonify({"error": "Không tìm thấy preset"}), 404
    return jsonify({"success": True, "presets": presets})


# ==========================================================================
# T34 Task 3E — Validate mapping dry-run (test parse 5 row đầu)
# ==========================================================================

@app.route("/api/validate-mapping", methods=["POST"])
def validate_mapping():
    """
    Chạy parser dry-run trên tmp file với mapping user chọn → trả preview
    5 record + errors/warnings để user check trước khi confirm.

    Body JSON:
      {
        "tmp_id": "<uuid từ upload-preview>",
        "column_mapping": {"Mã CN": "Function Code", ...},
        "n_rows": 5    // optional, default 5, max 20
      }

    Response:
      {
        "success": true,
        "rows": [{ma_cn, ten_cn, phases: {...}}, ...],
        "errors": [{row_idx, col, msg}, ...],
        "warnings": ["...", ...],
        "row_count_scanned": N
      }
    """
    from analyzer.type_infer import validate_mapping_dry_run
    from parser import column_mapping as cm_mod

    body = request.get_json(silent=True) or {}
    tmp_id = str(body.get("tmp_id") or "").strip()
    mapping = cm_mod.sanitize_column_mapping(body.get("column_mapping") or {})
    n_rows = int(body.get("n_rows") or 5)
    n_rows = max(1, min(n_rows, 20))

    if not tmp_id:
        return jsonify({"error": "Thiếu 'tmp_id'"}), 400
    if not all(c in "abcdef0123456789" for c in tmp_id.lower()):
        return jsonify({"error": "tmp_id không hợp lệ"}), 400

    tmp_path = os.path.join(_TMP_UPLOAD_DIR, f"{tmp_id}.xlsx")
    if not os.path.isfile(tmp_path):
        return jsonify({
            "error": "tmp_id không tồn tại hoặc đã hết hạn (24h)"
        }), 404

    result = validate_mapping_dry_run(tmp_path, mapping, n_rows=n_rows)
    return jsonify(result)


# ==========================================================================
# T34 Task 3C — Integration mapping preset CRUD (JSON API preset per integ)
# ==========================================================================

@app.route(
    "/api/projects/<slug>/integrations/<integration_id>/mapping-presets",
    methods=["GET", "POST"],
)
def integration_mapping_presets(slug: str, integration_id: str):
    """
    GET → list preset của integration_id.
    POST → save preset {name, mapping}.
    """
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)

    if request.method == "GET":
        return jsonify({
            "presets": ps.list_integration_mapping_presets(folder, integration_id),
        })

    body = request.get_json(silent=True) or {}
    name = str(body.get("name") or "").strip()
    mapping = body.get("mapping") or {}
    if not name:
        return jsonify({"error": "Thiếu 'name'"}), 400
    if not isinstance(mapping, dict):
        return jsonify({"error": "'mapping' phải là object"}), 400
    try:
        presets = ps.save_integration_mapping_preset(folder, integration_id, name, mapping)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"presets": presets}), 201


@app.route(
    "/api/projects/<slug>/integrations/<integration_id>/mapping-presets/<name>",
    methods=["DELETE"],
)
def integration_mapping_preset_delete(slug: str, integration_id: str, name: str):
    """Xoá 1 preset của integration."""
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    deleted, presets = ps.delete_integration_mapping_preset(folder, integration_id, name)
    if not deleted:
        return jsonify({"error": "Không tìm thấy preset"}), 404
    return jsonify({"success": True, "presets": presets})


def _flatten_records_for_preview(records: list) -> list[dict]:
    """
    Flatten list of dict (chỉ 1 level nested — nếu value là dict → prefix
    dot). Value list/scalar → giữ nguyên.

    Ví dụ:
      [{"code": "A", "meta": {"module": "HR"}}]
      → [{"code": "A", "meta.module": "HR"}]

    Không dùng recursive để tránh explode với json phức tạp.
    """
    if not isinstance(records, list):
        return []
    out: list[dict] = []
    for r in records:
        if not isinstance(r, dict):
            continue
        flat: dict = {}
        for k, v in r.items():
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat[f"{k}.{k2}"] = v2
            else:
                flat[k] = v
        out.append(flat)
    return out


def _parse_multi_arg(name: str) -> list[str]:
    """
    Parse query param dạng multi-value. Hỗ trợ CẢ 2 pattern (tương thích ngược):
    - Repeated: ?module=A&module=B
    - Comma-separated: ?module=A,B,C
    - Kết hợp: ?module=A,B&module=C  → [A, B, C]
    Trim + dedupe (giữ thứ tự xuất hiện lần đầu để URL của user không bị đảo).
    """
    raw_values = request.args.getlist(name)
    result: list[str] = []
    seen = set()
    for raw in raw_values:
        # 1 item có thể chứa comma → split thêm
        for part in raw.split(","):
            v = part.strip()
            if v and v not in seen:
                seen.add(v)
                result.append(v)
    return result


def _parse_project_code_args() -> list[str]:
    """
    Đọc Mã dự án từ query — alias theo thứ tự ưu tiên:
    g_project / g_ma_du_an / g_project_code / project_code / ma_du_an / project.
    """
    for key in (
        "g_project",
        "g_ma_du_an",
        "g_project_code",
        "project_code",
        "ma_du_an",
        "project",
    ):
        vals = _parse_multi_arg(key)
        if vals:
            return vals
    return []


def _project_codes_from_body(body: dict) -> list[str]:
    """Đọc Mã dự án từ POST JSON body (nhiều alias)."""
    if not body:
        return []

    def _as_list(val) -> list[str]:
        if not val:
            return []
        if isinstance(val, list):
            out: list[str] = []
            for it in val:
                out.extend(_as_list(it))
            return out
        return [x.strip() for x in str(val).split(",") if x.strip()]

    for key in (
        "g_project",
        "g_ma_du_an",
        "g_project_code",
        "g_project_codes",
        "project_codes",
        "project_code",
        "ma_du_an",
        "project",
    ):
        vals = _as_list(body.get(key))
        if vals:
            return vals
    return []


@app.route("/api/projects/<slug>/dashboard")
def dashboard_of_project(slug):
    st, err = _need_state(slug)
    if err:
        return err

    # Global filter (module / quy trình / pic / mã dự án) — multi-value
    fmodules = _parse_multi_arg("module")
    fprocesses = _parse_multi_arg("process")
    fpics = _parse_multi_arg("pic")
    fproject_codes = _parse_project_code_args()

    if fmodules or fprocesses or fpics or fproject_codes:
        filtered_data = _filter_parsed_data(
            st["data"],
            modules=fmodules,
            processes=fprocesses,
            pics=fpics,
            project_codes=fproject_codes,
        )
        engine = DashboardEngine()
        metrics = engine.compute_all(filtered_data)
        applied_filter = {
            "modules": fmodules,
            "processes": fprocesses,
            "pics": fpics,
            "project_codes": fproject_codes,
            "row_count": len(filtered_data.rows),
        }
    else:
        metrics = st["metrics"]
        applied_filter = None

    return jsonify({
        "success": True,
        "project": _project_to_dict(_project_mgr.get_project(slug)),
        "filename": st["filename"],
        "upload_time": st["upload_time"].isoformat() if st["upload_time"] else None,
        "metrics": _trim_payload(metrics),
        "snapshots": _project_mgr.get_snapshot_manager(slug).list_snapshots(),
        "applied_filter": applied_filter,
        # Count blacklist luôn tính từ st["data"] (parsed data gốc, không filter),
        # để badge count nhất quán khi user đổi filter.
        "pic_blacklist_count": len(getattr(st["data"], "pic_blacklisted", []) or []),
    })


@app.route("/api/dashboard")
def dashboard_legacy():
    """Legacy: dashboard mặc định = project 'default'."""
    try:
        slug = _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    return dashboard_of_project(slug)


# ==========================================================================
# Advanced analytics (per-project + legacy fallback)
# ==========================================================================

def _apply_list_filters(items: list, filter_map: dict) -> list:
    """Filter list of dicts theo dict {key: value}, bỏ qua value falsy."""
    result = items
    for k, v in filter_map.items():
        if not v:
            continue
        if k == "pic":
            result = [i for i in result if v in i.get("pic", [])]
        elif k == "min_days":
            try:
                min_v = int(v)
                result = [i for i in result if i.get("days_overdue", 0) >= min_v]
            except (TypeError, ValueError):
                pass
        else:
            result = [i for i in result if i.get(k) == v]
    return result


@app.route("/api/projects/<slug>/overdue")
@app.route("/api/overdue")
def get_overdue(slug=None):
    """Overdue full list (chưa trim). Support paginate qua ?limit=&offset=."""
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    overdue = st["metrics"].get("overdue_list", [])
    filters = {
        "module": request.args.get("module"),
        "pic": request.args.get("pic"),
        "phase": request.args.get("phase"),
        "min_days": request.args.get("min_days"),
    }
    overdue = _apply_list_filters(overdue, filters)

    total = len(overdue)
    limit = request.args.get("limit", type=int)
    offset = request.args.get("offset", default=0, type=int)
    if limit:
        overdue = overdue[offset:offset + limit]

    return jsonify({"success": True, "overdue": overdue, "total": total, "offset": offset})


@app.route("/api/projects/<slug>/unassigned")
@app.route("/api/unassigned")
def get_unassigned(slug=None):
    """Unassigned với pagination."""
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    items = st["metrics"].get("unassigned_tasks", [])
    filters = {"module": request.args.get("module"), "phase": request.args.get("phase")}
    items = _apply_list_filters(items, filters)

    total = len(items)
    limit = request.args.get("limit", type=int)
    offset = request.args.get("offset", default=0, type=int)
    if limit:
        items = items[offset:offset + limit]

    return jsonify({"success": True, "items": items, "total": total, "offset": offset})


@app.route("/api/projects/<slug>/long-duration")
@app.route("/api/long-duration")
def get_long_duration(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err
    return jsonify({"success": True, "data": st["metrics"].get("duration_analysis", {})})


@app.route("/api/projects/<slug>/stalled")
@app.route("/api/stalled")
def get_stalled(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err
    return jsonify({"success": True, "data": st["metrics"].get("stalled_tasks", {})})


@app.route("/api/projects/<slug>/risk-scores")
@app.route("/api/risk-scores")
def get_risk_scores(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err
    all_scores = st["metrics"].get("risk_scores", [])
    top = request.args.get("top", default=20, type=int)
    items = all_scores[:top]
    return jsonify({"success": True, "items": items, "total": len(all_scores)})


# ==========================================================================
# Drill-Down — click biểu đồ → list function chi tiết + export Excel
# ==========================================================================

def _parse_drill_filters(source: dict) -> dict:
    """Extract các filter key hợp lệ từ query string hoặc JSON body."""
    # BUG FIX: thiếu 'task_type', 'ma_cn', 'level' → drill task_type / risk / timeline
    # nhận filters rỗng, trả 0 items. Bổ sung mọi filter key mà analyzer/drill_down.py cần.
    valid_keys = (
        "module", "phase", "status", "pic", "priority",
        "complexity", "fit_gap", "giai_doan", "process",
        "task_type", "ma_cn", "level",
    )
    return {k: source.get(k, "") for k in valid_keys if source.get(k, "") != ""}


@app.route("/api/projects/<slug>/drill-down")
@app.route("/api/drill-down")
def drill_down_endpoint(slug=None):
    """
    GET params:
      - chart: 1 trong SUPPORTED_CHARTS
      - filters: module, phase, status, pic, priority, complexity, fit_gap, giai_doan, process
    Trả: {items: [...], total: N, title: "..."}
    """
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    chart = request.args.get("chart", "")
    if chart not in SUPPORTED_CHARTS:
        return jsonify({
            "error": f"Chart không hỗ trợ: '{chart}'. Supported: {list(SUPPORTED_CHARTS)}"
        }), 400

    filters = _parse_drill_filters(request.args)
    # Global filter (từ filter bar chính) — pre-filter data trước khi run drill.
    # Prefix "_g_" để không conflict với chart-specific filter (VD `module` là chart filter).
    data_scoped = _apply_global_filter_from_request(st["data"], request.args)
    try:
        items = drill_down_fn(data_scoped, chart, filters)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    title = build_drill_title(chart, filters)
    return jsonify({
        "success": True,
        "chart": chart,
        "filters": filters,
        "title": title,
        "items": items,
        "total": len(items),
    })


def _apply_global_filter_from_request(data, source):
    """Đọc _g_module / _g_process / _g_pic / _g_project* → apply _filter_parsed_data."""
    def _mv(key: str) -> list[str]:
        # Hỗ trợ cả query args (getlist) lẫn plain dict (comma-split)
        if hasattr(source, "getlist"):
            vals = source.getlist(key)
        else:
            raw = source.get(key, "")
            vals = [raw] if isinstance(raw, str) else list(raw or [])
        out = []
        for v in vals:
            for part in str(v).split(","):
                part = part.strip()
                if part:
                    out.append(part)
        return out

    g_modules = _mv("_g_module")
    g_processes = _mv("_g_process")
    g_pics = _mv("_g_pic")
    g_project_codes = (
        _mv("_g_project")
        or _mv("_g_ma_du_an")
        or _mv("_g_project_code")
    )
    if not (g_modules or g_processes or g_pics or g_project_codes):
        return data
    return _filter_parsed_data(
        data,
        modules=g_modules,
        processes=g_processes,
        pics=g_pics,
        project_codes=g_project_codes,
    )


@app.route("/api/projects/<slug>/drill-down/export", methods=["POST"])
@app.route("/api/drill-down/export", methods=["POST"])
def drill_down_export(slug=None):
    """
    POST JSON: {chart, filters: {...}}
    Trả file Excel .xlsx.
    """
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    body = request.get_json(silent=True) or {}
    chart = body.get("chart", "")
    if chart not in SUPPORTED_CHARTS:
        return jsonify({
            "error": f"Chart không hỗ trợ: '{chart}'. Supported: {list(SUPPORTED_CHARTS)}"
        }), 400

    filters = _parse_drill_filters(body.get("filters", {}))
    # Global filter (từ body.global_filter — FE gửi khi export)
    gf = body.get("global_filter", {}) or {}
    data_scoped = _apply_global_filter_from_request(st["data"], {
        "_g_module": gf.get("modules", []),
        "_g_process": gf.get("processes", []),
        "_g_pic": gf.get("pics", []),
        "_g_project": gf.get("project_codes", []) or gf.get("projectCodes", []) or gf.get("g_project", []),
    })
    try:
        items = drill_down_fn(data_scoped, chart, filters)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    title = build_drill_title(chart, filters)
    filepath = export_drill_down(
        items,
        title=title,
        subtitle=f"Project: {slug} | Tổng: {len(items)} function",
        output_dir=_project_mgr.get_export_dir(slug),
    )
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


# ==========================================================================
# Snapshot & Compare (per-project + cross-project)
# ==========================================================================

@app.route("/api/projects/<slug>/snapshots", methods=["GET"])
@app.route("/api/snapshots", methods=["GET"])
def list_snapshots(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    smgr = _project_mgr.get_snapshot_manager(slug)
    return jsonify({"success": True, "snapshots": smgr.list_snapshots()})


@app.route("/api/projects/<slug>/snapshots/<snapshot_date>", methods=["DELETE"])
@app.route("/api/snapshots/<snapshot_date>", methods=["DELETE"])
def delete_snapshot(snapshot_date: str, slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    smgr = _project_mgr.get_snapshot_manager(slug)
    ok = smgr.delete_snapshot(snapshot_date)
    if not ok:
        return jsonify({"error": "Không tìm thấy snapshot"}), 404
    return jsonify({"success": True, "snapshots": smgr.list_snapshots()})


@app.route("/api/projects/<slug>/compare")
@app.route("/api/compare")
def compare_snapshots(slug=None):
    """Compare 2 snapshot trong CÙNG 1 project."""
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404

    old_date = request.args.get("old")
    new_date = request.args.get("new")
    if not old_date or not new_date:
        return jsonify({"error": "Thiếu tham số old/new"}), 400

    smgr = _project_mgr.get_snapshot_manager(slug)
    old_data = smgr.load_snapshot(old_date)
    new_data = smgr.load_snapshot(new_date)
    if not old_data or not new_data:
        return jsonify({"error": "Không tìm thấy snapshot"}), 404

    result = CompareEngine().compare(
        old_data["parsed"], new_data["parsed"],
        old_date=old_date, new_date=new_date,
    )
    return jsonify({"success": True, "result": result})


@app.route("/api/compare-cross")
def compare_cross_project():
    """
    Compare snapshot giữa 2 project khác nhau (Phase 2).
    Query: ?project_a=<slug>&snap_a=<date>&project_b=<slug>&snap_b=<date>
    """
    pa = request.args.get("project_a")
    sa = request.args.get("snap_a")
    pb = request.args.get("project_b")
    sb = request.args.get("snap_b")
    if not (pa and sa and pb and sb):
        return jsonify({"error": "Thiếu tham số project_a/snap_a/project_b/snap_b"}), 400

    smgr_a = _project_mgr.get_snapshot_manager(pa)
    smgr_b = _project_mgr.get_snapshot_manager(pb)
    a = smgr_a.load_snapshot(sa)
    b = smgr_b.load_snapshot(sb)
    if not a or not b:
        return jsonify({"error": "Không tìm thấy 1 trong 2 snapshot"}), 404

    result = CompareEngine().compare(
        a["parsed"], b["parsed"],
        old_date=f"{pa}:{sa}", new_date=f"{pb}:{sb}",
    )
    proj_a = _project_mgr.get_project(pa)
    proj_b = _project_mgr.get_project(pb)
    return jsonify({
        "success": True,
        "project_a": _project_to_dict(proj_a) if proj_a else None,
        "project_b": _project_to_dict(proj_b) if proj_b else None,
        "snap_a": sa,
        "snap_b": sb,
        "result": result,
    })


@app.route("/api/projects/<slug>/upload-compare", methods=["POST"])
@app.route("/api/upload-compare", methods=["POST"])
def upload_compare(slug=None):
    """Upload file cũ để so sánh với file hiện tại của project (không lưu snapshot)."""
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy file"}), 400
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "File không hợp lệ"}), 400

    tmp_path = os.path.join(app.config["UPLOAD_FOLDER"], f"tmp_compare_{slug}.xlsx")
    file.save(tmp_path)

    try:
        old_data = FunctionListParser().parse(tmp_path)
        result = CompareEngine().compare(
            old_data, st["data"],
            old_date="uploaded_file",
            new_date=date.today().isoformat(),
        )
        return jsonify({"success": True, "result": result})
    except Exception as e:
        return jsonify({"error": f"Lỗi so sánh: {str(e)}"}), 500
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass


# ==========================================================================
# Exports (per-project + legacy)
# ==========================================================================

@app.route("/api/projects/<slug>/export-overdue", methods=["GET", "POST"])
@app.route("/api/export-overdue", methods=["GET", "POST"])
def export_overdue(slug=None):
    """
    Xuất Excel danh sách task trễ.

    Chấp nhận 2 loại filter (áp tuần tự):
      1. Global filter (header dashboard) — g_module / g_process / g_pic:
         → filter parsed data → recompute overdue_list. Đảm bảo file xuất
         ra khớp với danh sách user đang thấy trên grid sau khi apply filter
         header (không lộn với `module`/`pic` cấp local widget).
      2. Local widget filter (widget riêng của section Overdue) —
         module / pic / phase: áp lên overdue_list đã filter global để
         thu hẹp thêm.

    Backward compat: nếu client cũ chỉ gửi `module`/`pic`/`phase` (không có
    g_*), giữ semantics cũ = local filter, không apply global.
    """
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    def _as_list(val) -> list[str]:
        if not val:
            return []
        if isinstance(val, list):
            out: list[str] = []
            for it in val:
                out.extend(_as_list(it))
            return out
        return [x.strip() for x in str(val).split(",") if x.strip()]

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        # Local widget filter (giữ shape cũ để backward compat)
        filters = {
            "module": body.get("module"),
            "pic": body.get("pic"),
            "phase": body.get("phase"),
        }
        # Global filter — key mới g_* để không đè local
        g_modules = _as_list(body.get("g_module") or body.get("g_modules"))
        g_processes = _as_list(body.get("g_process") or body.get("g_processes"))
        g_pics = _as_list(body.get("g_pic") or body.get("g_pics"))
        g_project_codes = _project_codes_from_body(body)
        mode = (body.get("mode") or "both").strip().lower()
    else:
        filters = {
            "module": request.args.get("module"),
            "pic": request.args.get("pic"),
            "phase": request.args.get("phase"),
        }
        g_modules = _parse_multi_arg("g_module")
        g_processes = _parse_multi_arg("g_process")
        g_pics = _parse_multi_arg("g_pic")
        g_project_codes = _parse_project_code_args()
        mode = (request.args.get("mode") or "both").strip().lower()

    filters = {k: v for k, v in filters.items() if v}

    try:
        # Bước 1: apply global filter → recompute overdue_list nếu có
        if g_modules or g_processes or g_pics or g_project_codes:
            filtered_data = _filter_parsed_data(
                st["data"],
                modules=g_modules,
                processes=g_processes,
                pics=g_pics,
                project_codes=g_project_codes,
            )
            overdue_list = DashboardEngine().compute_all(filtered_data).get(
                "overdue_list", []
            )
        else:
            overdue_list = st["metrics"].get("overdue_list", [])

        filepath = export_overdue_report(
            overdue_list=overdue_list,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters if filters else None,
            mode=mode,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất file: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-stalled", methods=["GET", "POST"])
@app.route("/api/export-stalled", methods=["GET", "POST"])
def export_stalled(slug=None):
    """
    Xuất Excel danh sách task bị đình trệ.

    Chấp nhận 2 loại filter (áp tuần tự, giống export-overdue):
      1. Global filter — g_module / g_process / g_pic:
         → filter parsed data → recompute stalled_tasks.
      2. Local widget filter — module (comma-sep multi):
         → thu hẹp thêm trên stalled items.
    """
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    def _as_list(val) -> list[str]:
        if not val:
            return []
        if isinstance(val, list):
            out: list[str] = []
            for it in val:
                out.extend(_as_list(it))
            return out
        return [x.strip() for x in str(val).split(",") if x.strip()]

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        filters = {"module": body.get("module")}
        g_modules = _as_list(body.get("g_module") or body.get("g_modules"))
        g_processes = _as_list(body.get("g_process") or body.get("g_processes"))
        g_pics = _as_list(body.get("g_pic") or body.get("g_pics"))
        g_project_codes = _project_codes_from_body(body)
        mode = (body.get("mode") or "both").strip().lower()
    else:
        filters = {"module": request.args.get("module")}
        g_modules = _parse_multi_arg("g_module")
        g_processes = _parse_multi_arg("g_process")
        g_pics = _parse_multi_arg("g_pic")
        g_project_codes = _parse_project_code_args()
        mode = (request.args.get("mode") or "both").strip().lower()

    filters = {k: v for k, v in filters.items() if v}

    try:
        if g_modules or g_processes or g_pics or g_project_codes:
            filtered_data = _filter_parsed_data(
                st["data"],
                modules=g_modules,
                processes=g_processes,
                pics=g_pics,
                project_codes=g_project_codes,
            )
            stalled_items = (
                DashboardEngine().compute_all(filtered_data)
                .get("stalled_tasks", {})
                .get("items", [])
            )
        else:
            stalled_items = (
                st["metrics"].get("stalled_tasks", {}) or {}
            ).get("items", [])

        filepath = export_stalled_report(
            stalled_items=stalled_items,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters if filters else None,
            mode=mode,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất file: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-full-report")
@app.route("/api/export-full-report")
def export_full(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err
    try:
        filepath = export_full_report(st["metrics"], _project_mgr.get_export_dir(slug))
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất file: {str(e)}"}), 500

@app.route("/api/projects/<slug>/export-weekly-mom")
@app.route("/api/export-weekly-mom")
def export_weekly_mom_api(slug=None):
    """Xuất báo cáo tuần MoM (mẫu W30) + sheet PM Dashboard."""
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err
    try:
        from analyzer import pm_store as _pm_store
        proj = _project_mgr.get_project(slug)
        project_code = (proj.name if proj else None) or slug
        pdir = _project_mgr.get_project_folder(slug)
        pm_plan = _pm_store.load_plan(pdir)
        filepath = export_weekly_mom(
            st["metrics"],
            _project_mgr.get_export_dir(slug),
            project_code=project_code,
            parsed_data=st.get("data"),
            pm_plan=pm_plan,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất MoM tuần: {str(e)}"}), 500


# ==========================================================================
# Chiều PM — Kế hoạch dự án (Excel) + Weekly Report (PPT)
# ==========================================================================

def _pm_save_tmp(file_storage, allowed_ext: tuple[str, ...]) -> tuple:
    """Lưu file upload vào uploads/tmp/. Return (tmp_id, filename, path) | (None, err, status)."""
    import uuid as _uuid
    if not file_storage or not file_storage.filename:
        return None, "Chưa chọn file", 400
    fname = file_storage.filename
    ext = os.path.splitext(fname)[1].lower()
    if ext not in allowed_ext:
        return None, f"Chỉ hỗ trợ file {', '.join(allowed_ext)}", 400
    tmp_dir = os.path.join(app.config["UPLOAD_FOLDER"], "tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    _prune_old_tmp_uploads()
    tmp_id = _uuid.uuid4().hex[:16]
    tmp_path = os.path.join(tmp_dir, f"{tmp_id}{ext}")
    file_storage.save(tmp_path)
    return tmp_id, fname, tmp_path


def _pm_tmp_path(tmp_id: str) -> Optional[str]:
    """Tìm file tmp theo id (xlsx/pptx/xls)."""
    if not tmp_id or not re.fullmatch(r"[0-9a-f]{8,32}", tmp_id):
        return None
    tmp_dir = os.path.join(app.config["UPLOAD_FOLDER"], "tmp")
    for ext in (".xlsx", ".pptx", ".xls"):
        fp = os.path.join(tmp_dir, f"{tmp_id}{ext}")
        if os.path.isfile(fp):
            return fp
    return None


@app.route("/api/projects/<slug>/pm", methods=["GET"])
def pm_get(slug: str):
    """Lấy bundle chiều PM đã lưu (+ optional join Function List)."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    from analyzer import pm_store as _pm_store
    pdir = _project_mgr.get_project_folder(slug)
    bundle = _pm_store.load_pm_bundle(pdir)
    st = _get_state(slug)
    fl_links = _pm_store.link_with_function_list(
        bundle.get("plan"),
        bundle.get("weekly"),
        st.get("data") if st else None,
    )
    bundle["fl_links"] = fl_links
    return jsonify(bundle)


@app.route("/api/projects/<slug>/pm/plan/preview", methods=["POST"])
def pm_plan_preview(slug: str):
    """Upload KeHoachDuAn → đề xuất sheet mapping."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    from parser.pm_plan_parser import preview_plan_workbook
    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy file"}), 400
    tmp_id, fname, tmp_path = _pm_save_tmp(request.files["file"], (".xlsx", ".xls"))
    if tmp_id is None:
        return jsonify({"error": fname}), tmp_path  # fname=err, tmp_path=status
    try:
        preview = preview_plan_workbook(tmp_path)
    except Exception as e:
        return jsonify({"error": f"Không đọc được Excel: {e}"}), 400
    return jsonify({"tmp_id": tmp_id, "filename": fname, **preview})


@app.route("/api/projects/<slug>/pm/plan/confirm", methods=["POST"])
def pm_plan_confirm(slug: str):
    """Xác nhận mapping + parse + lưu plan.json."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    from parser.pm_plan_parser import parse_plan, propose_sheet_mapping
    from analyzer import pm_store as _pm_store
    body = request.get_json(silent=True) or {}
    tmp_id = (body.get("tmp_id") or "").strip()
    tmp_path = _pm_tmp_path(tmp_id)
    if not tmp_path:
        return jsonify({"error": "tmp_id không hợp lệ hoặc file tạm đã hết hạn"}), 400
    mapping = body.get("sheet_mapping") or {}
    if not isinstance(mapping, dict):
        mapping = {}
    try:
        parsed = parse_plan(tmp_path, sheet_mapping=mapping or None)
        if not mapping:
            parsed["sheet_mapping"] = propose_sheet_mapping(parsed.get("sheet_names") or [])
        pdir = _project_mgr.get_project_folder(slug)
        saved = _pm_store.save_plan(
            pdir, parsed,
            source_filename=body.get("filename") or os.path.basename(tmp_path),
            source_path=tmp_path,
        )
        return jsonify({
            "ok": True,
            "summary": saved.get("summary"),
            "sheet_mapping": saved.get("sheet_mapping"),
            "imported_at": saved.get("imported_at"),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi parse kế hoạch: {e}"}), 500


@app.route("/api/projects/<slug>/pm/weekly/preview", methods=["POST"])
def pm_weekly_preview(slug: str):
    """Upload Weekly PPT → tóm tắt slides."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    from parser.pm_weekly_parser import preview_weekly
    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy file"}), 400
    tmp_id, fname, tmp_path = _pm_save_tmp(request.files["file"], (".pptx",))
    if tmp_id is None:
        return jsonify({"error": fname}), tmp_path
    try:
        preview = preview_weekly(tmp_path)
    except Exception as e:
        return jsonify({"error": f"Không đọc được PPT: {e}"}), 400
    return jsonify({"tmp_id": tmp_id, "filename": fname, **preview})


@app.route("/api/projects/<slug>/pm/weekly/confirm", methods=["POST"])
def pm_weekly_confirm(slug: str):
    """Parse + lưu weekly.json."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    from parser.pm_weekly_parser import parse_weekly
    from analyzer import pm_store as _pm_store
    body = request.get_json(silent=True) or {}
    tmp_id = (body.get("tmp_id") or "").strip()
    tmp_path = _pm_tmp_path(tmp_id)
    if not tmp_path:
        return jsonify({"error": "tmp_id không hợp lệ hoặc file tạm đã hết hạn"}), 400
    try:
        parsed = parse_weekly(tmp_path)
        pdir = _project_mgr.get_project_folder(slug)
        saved = _pm_store.save_weekly(
            pdir, parsed,
            source_filename=body.get("filename") or os.path.basename(tmp_path),
            source_path=tmp_path,
        )
        return jsonify({
            "ok": True,
            "summary": saved.get("summary"),
            "period_start": saved.get("period_start"),
            "period_end": saved.get("period_end"),
            "imported_at": saved.get("imported_at"),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi parse weekly: {e}"}), 500


@app.route("/api/projects/<slug>/pm/export", methods=["GET"])
def pm_export(slug: str):
    """Xuất Excel tổng hợp chiều PM."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    from analyzer import pm_store as _pm_store
    pdir = _project_mgr.get_project_folder(slug)
    plan = _pm_store.load_plan(pdir)
    weekly = _pm_store.load_weekly(pdir)
    if not plan and not weekly:
        return jsonify({"error": "Chưa có dữ liệu chiều PM — hãy import kế hoạch hoặc weekly"}), 400
    st = _get_state(slug)
    fl_links = _pm_store.link_with_function_list(
        plan, weekly, st.get("data") if st else None,
    )
    proj = _project_mgr.get_project(slug)
    project_code = (proj.name if proj else None) or slug
    try:
        filepath = export_pm_report(
            plan, weekly,
            _project_mgr.get_export_dir(slug),
            project_code=project_code,
            fl_links=fl_links,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi xuất chiều PM: {e}"}), 500


# ==========================================================================
# T34 Task 1 — Xuất "Toàn bộ vấn đề" ra 1 Excel workbook (8 sheet)
# ==========================================================================

@app.route("/api/projects/<slug>/export-all-issues", methods=["GET", "POST"])
def project_export_all_issues(slug: str):
    """
    Xuất 1 workbook Excel duy nhất chứa mọi loại vấn đề, mỗi loại 1 sheet.

    Cover + Overdue + Chưa PIC + Đình trệ + High Risk + Aging WIP +
    Data Quality + Bookmark.

    Query params (đồng nhất với các endpoint export khác):
      g_module / g_process / g_pic  → global filter (comma-separated
        hoặc lặp nhiều lần).
      threshold                     → ngưỡng aging WIP (default 14).

    POST body: JSON { g_module: [...], g_process: [...], g_pic: [...],
      threshold: int }.

    Áp filter global 1 lần (tại _filter_parsed_data), sau đó compute mọi
    loại vấn đề trên filtered data. Tránh recompute nhiều lần.
    """
    from analyzer.data_quality import compute_data_quality
    from analyzer.advanced_metrics import compute_aging_wip
    from analyzer.risk_scorer import compute_all_risk_scores
    from analyzer import project_store as ps
    from exporter.export_all_issues import export_all_issues

    state, err = _require_state(slug)
    if err:
        return err

    # Parse global filter — hỗ trợ cả POST body và GET query string
    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        fmodules = body.get("g_module") or body.get("modules") or []
        fprocesses = body.get("g_process") or body.get("processes") or []
        fpics = body.get("g_pic") or body.get("pics") or []
        fproject_codes = _project_codes_from_body(body)
        threshold = body.get("threshold") or 14
        if isinstance(fmodules, str):
            fmodules = [x.strip() for x in fmodules.split(",") if x.strip()]
        if isinstance(fprocesses, str):
            fprocesses = [x.strip() for x in fprocesses.split(",") if x.strip()]
        if isinstance(fpics, str):
            fpics = [x.strip() for x in fpics.split(",") if x.strip()]
    else:
        fmodules = _parse_multi_arg("g_module") or _parse_multi_arg("module")
        fprocesses = _parse_multi_arg("g_process") or _parse_multi_arg("process")
        fpics = _parse_multi_arg("g_pic") or _parse_multi_arg("pic")
        fproject_codes = _parse_project_code_args()
        threshold = request.args.get("threshold") or 14

    try:
        threshold = int(threshold)
        threshold = max(1, min(365, threshold))
    except (ValueError, TypeError):
        threshold = 14

    # Apply filter global 1 lần cho toàn bộ compute
    if fmodules or fprocesses or fpics or fproject_codes:
        filtered = _filter_parsed_data(
            state["data"],
            modules=fmodules, processes=fprocesses, pics=fpics,
            project_codes=fproject_codes,
        )
    else:
        filtered = state["data"]

    # Compute 7 loại vấn đề — reuse các function đã có
    engine = DashboardEngine()
    metrics = engine.compute_all(filtered)

    overdue_list = metrics.get("overdue_list", []) or []
    unassigned_list = metrics.get("unassigned_tasks", []) or []
    stalled_list = (metrics.get("stalled_tasks", {}) or {}).get("items", []) or []

    risk_scores = compute_all_risk_scores(filtered, date.today(), 3)
    risk_list = [r for r in risk_scores if r.get("risk_score", 0) >= 30]

    aging_payload = compute_aging_wip(filtered, threshold_days=threshold)
    aging_items = aging_payload.get("items", []) or []

    dq_payload = compute_data_quality(filtered)
    dq_issues = dq_payload.get("issues", []) or []

    # Bookmark — dùng list ma_cn từ store rồi map sang FunctionRow của filtered.
    # Nếu function bị filter loại → không đưa vào bookmark sheet (nhất quán
    # với các sheet khác).
    project_dir = _project_dir_for(slug)
    bookmarked_codes = set(ps.load_bookmarks(project_dir) or [])
    bookmark_functions = []
    if bookmarked_codes:
        for r in filtered.rows:
            mc = str(r.meta.get("ma_cn") or "").strip()
            if mc and mc in bookmarked_codes:
                bookmark_functions.append({
                    "ma_cn": mc,
                    "ten_cn": r.meta.get("ten_cn", ""),
                    "module": r.meta.get("module", ""),
                    "quy_trinh": r.meta.get("quy_trinh") or r.meta.get("process", ""),
                    "priority": r.meta.get("priority", ""),
                    "complexity": r.meta.get("complexity", ""),
                    "giai_doan": r.meta.get("giai_doan", ""),
                    "fit_gap": r.meta.get("fit_gap", ""),
                })

    project = _project_mgr.get_project(slug)
    project_name = project.name if project else slug

    # Ngôn ngữ export (sheet/header) — theo UI đang chọn
    from analyzer.i18n import normalize_lang
    lang_raw = request.args.get("lang")
    if request.method == "POST":
        _body = request.get_json(silent=True) or {}
        lang_raw = _body.get("lang") or lang_raw
    lang = normalize_lang(lang_raw)

    try:
        filepath = export_all_issues(
            project_name=project_name,
            slug=slug,
            overdue_list=overdue_list,
            unassigned_list=unassigned_list,
            stalled_list=stalled_list,
            risk_list=risk_list,
            aging_wip_items=aging_items,
            data_quality_issues=dq_issues,
            bookmark_functions=bookmark_functions,
            filter_info={
                "modules": fmodules,
                "processes": fprocesses,
                "pics": fpics,
                "project_codes": fproject_codes,
            },
            output_dir=_project_mgr.get_export_dir(slug),
            lang=lang,
        )
        return send_file(filepath, as_attachment=True,
                         download_name=os.path.basename(filepath))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất file: {str(e)}"}), 500


# ==========================================================================
# FL re-import — xuất issues đúng format Function List + mẫu schema
# ==========================================================================

@app.route("/api/projects/<slug>/export-fl-reimport", methods=["GET", "POST"])
def project_export_fl_reimport(slug: str):
    """
    Xuất Excel Function List (header row 1) chỉ các CN dính issue:
    overdue / unassigned / stalled / anomalies (+ missing_deadline).
    Tôn trọng global filter. Tô vàng PIC/Status; xanh nhạt date-chain.
    """
    from analyzer.data_quality import ANOMALY_CODES, compute_data_quality
    from exporter.fl_reimport_export import collect_issue_hits, export_fl_reimport

    state, err = _require_state(slug)
    if err:
        return err

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        fmodules = body.get("g_module") or body.get("modules") or []
        fprocesses = body.get("g_process") or body.get("processes") or []
        fpics = body.get("g_pic") or body.get("pics") or []
        fproject_codes = _project_codes_from_body(body)
        if isinstance(fmodules, str):
            fmodules = [x.strip() for x in fmodules.split(",") if x.strip()]
        if isinstance(fprocesses, str):
            fprocesses = [x.strip() for x in fprocesses.split(",") if x.strip()]
        if isinstance(fpics, str):
            fpics = [x.strip() for x in fpics.split(",") if x.strip()]
    else:
        fmodules = _parse_multi_arg("g_module") or _parse_multi_arg("module")
        fprocesses = _parse_multi_arg("g_process") or _parse_multi_arg("process")
        fpics = _parse_multi_arg("g_pic") or _parse_multi_arg("pic")
        fproject_codes = _parse_project_code_args()

    if fmodules or fprocesses or fpics or fproject_codes:
        filtered = _filter_parsed_data(
            state["data"],
            modules=fmodules, processes=fprocesses, pics=fpics,
            project_codes=fproject_codes,
        )
    else:
        filtered = state["data"]

    engine = DashboardEngine()
    metrics = engine.compute_all(filtered)
    overdue_list = metrics.get("overdue_list", []) or []
    unassigned_list = metrics.get("unassigned_tasks", []) or []
    stalled_list = (metrics.get("stalled_tasks", {}) or {}).get("items", []) or []

    dq = compute_data_quality(filtered)
    # Anomaly + DQ liên quan re-import (deadline / PIC / status)
    reimport_codes = set(ANOMALY_CODES) | {
        "missing_deadline", "blank_pic", "invalid_status", "closed_no_end",
    }
    anomaly_issues = [
        it for it in (dq.get("issues") or [])
        if it.get("code") in reimport_codes
    ]

    hits = collect_issue_hits(
        overdue_list=overdue_list,
        unassigned_list=unassigned_list,
        stalled_list=stalled_list,
        anomaly_issues=anomaly_issues,
    )
    if not hits:
        return jsonify({"error": "Không có function nào dính issue để xuất."}), 400

    project_dir = _project_dir_for(slug)
    # Nguồn row: template đã lưu → current.xlsx → snapshot path trong state
    source_xlsx = None
    from exporter.fl_export_schema import template_xlsx_path
    tpl = template_xlsx_path(project_dir)
    if os.path.isfile(tpl):
        source_xlsx = tpl
    else:
        cur = _project_mgr.get_current_file_path(slug)
        if os.path.isfile(cur):
            source_xlsx = cur
        elif state.get("filepath") and os.path.isfile(state["filepath"]):
            source_xlsx = state["filepath"]

    try:
        filepath = export_fl_reimport(
            filtered,
            hits=hits,
            output_dir=_project_mgr.get_export_dir(slug),
            project_dir=project_dir,
            source_xlsx=source_xlsx,
            project_slug=slug,
        )
        return send_file(
            filepath, as_attachment=True,
            download_name=os.path.basename(filepath),
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất FL re-import: {str(e)}"}), 500


@app.route("/api/projects/<slug>/fl-export-template", methods=["GET", "POST", "DELETE"])
def project_fl_export_template(slug: str):
    """
    GET: schema đã lưu + tip.
    POST multipart file=xlsx → parse → trả review payload (chưa lưu nếu
         query save=0; mặc định lưu luôn khi không gửi slot_assignments).
         Body JSON alternate: {tmp_id} hoặc {slot_assignments} + save.
    DELETE: xoá mẫu + schema.
    """
    from exporter.fl_export_schema import (
        delete_fl_export_template,
        load_fl_export_schema,
        review_from_xlsx,
        save_fl_export_template,
        apply_slot_overrides_to_schema,
        template_xlsx_path,
    )

    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    project_dir = _project_dir_for(slug)

    if request.method == "GET":
        schema = load_fl_export_schema(project_dir)
        has_tpl = os.path.isfile(template_xlsx_path(project_dir))
        return jsonify({
            "has_template": has_tpl,
            "schema": schema,
            "template_file": "fl_export_template.xlsx" if has_tpl else None,
            "tip": (
                "Upload file Function List mẫu → hệ thống auto-map cột. "
                "Review kéo-thả nếu cần → Lưu. Export FL re-import dùng schema này."
            ),
        })

    if request.method == "DELETE":
        delete_fl_export_template(project_dir)
        return jsonify({"success": True, "has_template": False})

    # POST — upload hoặc confirm mapping
    body = request.get_json(silent=True) if request.is_json else None
    save_flag = request.args.get("save", "1")
    want_save = str(save_flag).lower() not in ("0", "false", "no")

    tmp_path = None
    source_filename = ""
    slot_assignments = None

    if body and body.get("slot_assignments") is not None and not request.files:
        # Confirm mapping trên file đã lưu / tmp
        slot_assignments = body.get("slot_assignments") or {}
        schema_base = load_fl_export_schema(project_dir)
        if not schema_base:
            return jsonify({"error": "Chưa có schema — upload mẫu trước"}), 400
        schema = apply_slot_overrides_to_schema(schema_base, slot_assignments)
        if want_save:
            # Giữ file template hiện có
            tpl = template_xlsx_path(project_dir)
            if not os.path.isfile(tpl):
                return jsonify({"error": "Thiếu file mẫu trên disk"}), 400
            schema = save_fl_export_template(project_dir, tpl, schema)
        return jsonify({
            "success": True,
            "saved": want_save,
            "schema": schema,
            "review": {
                "slots": schema.get("slots"),
                "note_column": schema.get("note_column"),
                "headers": [
                    {"header": h, "group": h.rsplit(" - ", 1)[0] if " - " in h else "Meta"}
                    for h in (schema.get("headers") or [])
                ],
            },
        })

    if "file" in request.files and request.files["file"].filename:
        f = request.files["file"]
        source_filename = f.filename or "template.xlsx"
        if not source_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return jsonify({"error": "Chỉ nhận file Excel (.xlsx)"}), 400
        os.makedirs(_TMP_UPLOAD_DIR, exist_ok=True)
        import uuid
        tmp_id = uuid.uuid4().hex[:12]
        tmp_path = os.path.join(_TMP_UPLOAD_DIR, f"fl_tpl_{tmp_id}.xlsx")
        f.save(tmp_path)
    elif body and body.get("tmp_id"):
        tmp_path = os.path.join(_TMP_UPLOAD_DIR, f"fl_tpl_{body['tmp_id']}.xlsx")
        if not os.path.isfile(tmp_path):
            return jsonify({"error": "tmp_id hết hạn / không tồn tại"}), 400
        source_filename = body.get("filename") or "template.xlsx"
        slot_assignments = body.get("slot_assignments")
    else:
        return jsonify({"error": "Thiếu file mẫu Function List"}), 400

    try:
        if slot_assignments:
            from exporter.fl_export_schema import schema_from_xlsx, build_review_payload
            from parser.excel_parser import FunctionListParser
            parsed = FunctionListParser().parse(tmp_path)
            review = build_review_payload(
                parsed,
                source_filename=source_filename,
                slot_overrides=slot_assignments,
            )
        else:
            review = review_from_xlsx(tmp_path, source_filename)
        schema = review["schema"]
        saved = False
        if want_save and not slot_assignments:
            # Upload lần đầu: lưu luôn với auto-detect; user có thể edit sau
            schema = save_fl_export_template(project_dir, tmp_path, schema)
            saved = True
        elif want_save and slot_assignments:
            schema = save_fl_export_template(project_dir, tmp_path, schema)
            saved = True
        return jsonify({
            "success": True,
            "saved": saved,
            "tmp_id": os.path.basename(tmp_path).replace("fl_tpl_", "").replace(".xlsx", "") if tmp_path else None,
            "review": review,
            "schema": schema,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Không parse được mẫu: {e}"}), 400
    finally:
        # Nếu đã save thì có thể xoá tmp; nếu chưa save giữ tmp cho confirm
        if want_save and tmp_path and os.path.isfile(tmp_path):
            # Đã copy vào project_dir — xoá tmp
            try:
                if os.path.isfile(template_xlsx_path(project_dir)):
                    # chỉ xoá nếu không phải cùng path
                    if os.path.abspath(tmp_path) != os.path.abspath(template_xlsx_path(project_dir)):
                        os.remove(tmp_path)
            except OSError:
                pass


@app.route("/api/projects/<slug>/fl-export-template/review", methods=["POST"])
def project_fl_export_template_review(slug: str):
    """Upload mẫu → chỉ trả review mapping (không lưu). save=0."""
    # Reuse handler với save=0
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    # Force query
    from werkzeug.datastructures import ImmutableMultiDict
    args = request.args.to_dict(flat=True)
    args["save"] = "0"
    request.args = ImmutableMultiDict(args)
    return project_fl_export_template(slug)


@app.route("/api/projects/<slug>/export-chart", methods=["GET", "POST"])
def export_chart_endpoint(slug):
    """
    Xuất Excel 1 chart từ metrics.
    Query/body: chart (bắt buộc) + mode=summary|detail|both (default both)
      + optional module/process/pic filters.
    Khi có filter → recompute metrics trên subset rồi export.
    Sheets: Tong_hop / Chi_tiet (+ Theo_nhom nếu chart có group, kèm summary).
    """
    st, err = _need_state(slug)
    if err:
        return err

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        chart = (body.get("chart") or "").strip()
        fmodules = body.get("modules") or body.get("module") or []
        fprocesses = body.get("processes") or body.get("process") or []
        fpics = body.get("pics") or body.get("pic") or []
        fproject_codes = _project_codes_from_body(body)
        group_by = (body.get("group_by") or "module").strip().lower()
        mode = (body.get("mode") or "both").strip().lower()
        if isinstance(fmodules, str):
            fmodules = [x.strip() for x in fmodules.split(",") if x.strip()]
        if isinstance(fprocesses, str):
            fprocesses = [x.strip() for x in fprocesses.split(",") if x.strip()]
        if isinstance(fpics, str):
            fpics = [x.strip() for x in fpics.split(",") if x.strip()]
    else:
        chart = (request.args.get("chart") or "").strip()
        fmodules = _parse_multi_arg("module")
        fprocesses = _parse_multi_arg("process")
        fpics = _parse_multi_arg("pic")
        fproject_codes = _parse_project_code_args()
        group_by = (request.args.get("group_by") or "module").strip().lower()
        mode = (request.args.get("mode") or "both").strip().lower()

    if not chart:
        return jsonify({"error": "Thiếu tham số chart"}), 400
    if chart not in SUPPORTED_EXPORT_CHARTS:
        return jsonify({
            "error": f"Chart không hỗ trợ: {chart}",
            "supported": sorted(SUPPORTED_EXPORT_CHARTS),
        }), 400

    try:
        if fmodules or fprocesses or fpics or fproject_codes:
            filtered = _filter_parsed_data(
                st["data"], modules=fmodules, processes=fprocesses, pics=fpics,
                project_codes=fproject_codes,
            )
            metrics = DashboardEngine().compute_all(filtered)
            data_for_export = filtered
            subtitle = (
                f"Filter: module={fmodules or '-'} · process={fprocesses or '-'} · "
                f"pic={fpics or '-'} · project_code={fproject_codes or '-'} · "
                f"{len(filtered.rows)} function"
            )
        else:
            metrics = st["metrics"]
            data_for_export = st["data"]
            subtitle = ""

        filepath = export_chart(
            chart=chart,
            metrics=metrics,
            output_dir=_project_mgr.get_export_dir(slug),
            subtitle=subtitle,
            parsed_data=data_for_export,
            group_by=group_by,
            mode=mode,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất chart: {str(e)}"}), 500


@app.route("/api/projects/<slug>/audit-report", methods=["GET", "POST"])
def audit_report_endpoint(slug):
    """
    Xuất Report Đánh giá (11 sheet).
    Query/body: scope=all|filtered (+ module/process/pic nếu filtered).
    """
    st, err = _need_state(slug)
    if err:
        return err

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        scope = (body.get("scope") or "all").strip().lower()
        fmodules = body.get("modules") or body.get("module") or []
        fprocesses = body.get("processes") or body.get("process") or []
        fpics = body.get("pics") or body.get("pic") or []
        fproject_codes = _project_codes_from_body(body)
        if isinstance(fmodules, str):
            fmodules = [x.strip() for x in fmodules.split(",") if x.strip()]
        if isinstance(fprocesses, str):
            fprocesses = [x.strip() for x in fprocesses.split(",") if x.strip()]
        if isinstance(fpics, str):
            fpics = [x.strip() for x in fpics.split(",") if x.strip()]
    else:
        scope = (request.args.get("scope") or "all").strip().lower()
        fmodules = _parse_multi_arg("module")
        fprocesses = _parse_multi_arg("process")
        fpics = _parse_multi_arg("pic")
        fproject_codes = _parse_project_code_args()

    try:
        if scope == "filtered" and (fmodules or fprocesses or fpics or fproject_codes):
            data = _filter_parsed_data(
                st["data"], modules=fmodules, processes=fprocesses, pics=fpics,
                project_codes=fproject_codes,
            )
            metrics = DashboardEngine().compute_all(data)
            subtitle = (
                f"Scope: filtered · module={fmodules or '-'} · process={fprocesses or '-'} · "
                f"pic={fpics or '-'} · project_code={fproject_codes or '-'} · "
                f"{len(data.rows)} function"
            )
        else:
            data = st["data"]
            metrics = st["metrics"]
            subtitle = "Scope: all"

        filepath = export_audit_report(
            data, metrics,
            output_dir=_project_mgr.get_export_dir(slug),
            subtitle=subtitle,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi khi xuất audit report: {str(e)}"}), 500


@app.route("/api/projects/<slug>/pic-blacklist", methods=["GET"])
@app.route("/api/pic-blacklist", methods=["GET"])
def get_pic_blacklist(slug=None):
    """
    Trả về danh sách PIC bị blacklist trong file hiện tại của project.
    Response: {success, items, total, keywords}
      - items: list dict metadata cell bị bỏ
      - keywords: các keyword unique đã match (VD ["Closed", "In-progress"])

    Đây là báo cáo data-quality — giúp user phát hiện cell Excel bị lệch cột
    Status sang PIC. Không filter theo global filter — luôn là toàn dataset.
    """
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    data = st["data"]
    items = list(getattr(data, "pic_blacklisted", []) or [])
    # Unique keyword list — cho FE hiển thị chip filter/summary
    keywords = sorted({it.get("matched_keyword", "") for it in items if it.get("matched_keyword")})

    return jsonify({
        "success": True,
        "items": items,
        "total": len(items),
        "keywords": keywords,
    })


@app.route("/api/projects/<slug>/pic-blacklist/export", methods=["GET", "POST"])
@app.route("/api/pic-blacklist/export", methods=["GET", "POST"])
def export_pic_blacklist(slug=None):
    """Xuất Excel .xlsx báo cáo PIC bị blacklist. Không filter — toàn dataset."""
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    items = list(getattr(st["data"], "pic_blacklisted", []) or [])
    try:
        filepath = export_pic_blacklist_report(
            items=items,
            output_dir=_project_mgr.get_export_dir(slug),
            subtitle=(
                f"Project: {slug} | Tổng: {len(items)} token | "
                f"File: {st.get('filename', '')}"
            ),
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất file: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-by-pic")
@app.route("/api/export-by-pic")
def export_pic(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    st, err = _need_state(slug)
    if err:
        return err

    pic = request.args.get("pic")
    if not pic:
        return jsonify({"error": "Thiếu tham số pic"}), 400
    try:
        filepath = export_by_pic(st["metrics"], pic, _project_mgr.get_export_dir(slug))
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất file: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-compare")
@app.route("/api/export-compare")
def export_compare(slug=None):
    try:
        slug = slug or _resolve_slug()
    except ValueError as e:
        return jsonify({"error": str(e)}), 404

    old_date = request.args.get("old")
    new_date = request.args.get("new")
    if not old_date or not new_date:
        return jsonify({"error": "Thiếu tham số old/new"}), 400

    smgr = _project_mgr.get_snapshot_manager(slug)
    old = smgr.load_snapshot(old_date)
    new = smgr.load_snapshot(new_date)
    if not old or not new:
        return jsonify({"error": "Không tìm thấy snapshot"}), 404

    result = CompareEngine().compare(
        old["parsed"], new["parsed"],
        old_date=old_date, new_date=new_date,
    )
    try:
        filepath = export_compare_report(
            result, old_date, new_date, _project_mgr.get_export_dir(slug)
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất: {str(e)}"}), 500


# ==========================================================================
# Project package: Export/Import (Phase 2)
# ==========================================================================

@app.route("/api/projects/<slug>/export-package")
def export_package(slug):
    """
    Đóng gói toàn bộ project (current.xlsx + snapshots + meta) thành zip.
    Dùng để backup hoặc chuyển sang máy khác.
    """
    proj = _project_mgr.get_project(slug)
    if not proj:
        return jsonify({"error": "Không tìm thấy project"}), 404

    pdir = _project_mgr._project_dir(slug)  # OK use internal
    if not os.path.isdir(pdir):
        return jsonify({"error": "Thư mục project trống"}), 400

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(pdir):
            for f in files:
                abs_path = os.path.join(root, f)
                rel_path = os.path.relpath(abs_path, pdir)
                zf.write(abs_path, arcname=os.path.join(slug, rel_path))
    buf.seek(0)

    download_name = f"project_{slug}_{date.today().isoformat()}.zip"
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=download_name,
    )


@app.route("/api/projects/import-package", methods=["POST"])
def import_package():
    """
    Import project từ file zip (đã xuất bằng export-package).
    Zip cấu trúc: {slug}/meta.json, {slug}/current.xlsx, {slug}/snapshots/...
    Nếu slug đã tồn tại → generate slug mới (VD: dev-2).
    """
    if "file" not in request.files:
        return jsonify({"error": "Thiếu file"}), 400
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".zip"):
        return jsonify({"error": "Chỉ hỗ trợ .zip"}), 400

    tmp_zip = os.path.join(app.config["UPLOAD_FOLDER"], "_import_tmp.zip")
    file.save(tmp_zip)

    try:
        with zipfile.ZipFile(tmp_zip, "r") as zf:
            # Xác định slug gốc từ prefix folder trong zip
            names = zf.namelist()
            if not names:
                return jsonify({"error": "Zip rỗng"}), 400
            first_seg = names[0].split("/")[0]
            if not first_seg:
                return jsonify({"error": "Cấu trúc zip không hợp lệ"}), 400

            # Đọc meta.json để lấy tên
            meta_content = None
            for name in names:
                if name.endswith("meta.json"):
                    with zf.open(name) as f:
                        import json as _json
                        meta_content = _json.load(f)
                        break

            new_name = meta_content.get("name") if meta_content else first_seg
            new_desc = meta_content.get("description", "") if meta_content else ""

            # Tạo project mới với tên (slug tự sinh unique)
            new_proj = _project_mgr.create_project(new_name, new_desc)

            # Extract toàn bộ zip vào folder mới
            target_dir = _project_mgr._project_dir(new_proj.slug)
            with tempfile.TemporaryDirectory() as td:
                zf.extractall(td)
                src_root = os.path.join(td, first_seg)
                if not os.path.isdir(src_root):
                    return jsonify({"error": "Zip không có folder project ở gốc"}), 400
                # Move files sang target
                for item in os.listdir(src_root):
                    s = os.path.join(src_root, item)
                    d = os.path.join(target_dir, item)
                    if os.path.isfile(s):
                        shutil.copy2(s, d)
                    elif os.path.isdir(s):
                        if os.path.exists(d):
                            shutil.rmtree(d)
                        shutil.copytree(s, d)

            return jsonify({
                "success": True,
                "project": _project_to_dict(_project_mgr.get_project(new_proj.slug)),
            })
    except zipfile.BadZipFile:
        return jsonify({"error": "File zip không hợp lệ"}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi import: {str(e)}"}), 500
    finally:
        try:
            os.remove(tmp_zip)
        except OSError:
            pass


# ==========================================================================
# Portfolio (Cross-project) — Level 1/2/3
# ==========================================================================
# 4 endpoint mới, KHÔNG sửa endpoint cũ. Reuse `_get_state` để tận dụng cache
# in-memory + auto-load từ disk (đã có sẵn).

def _portfolio_state_loader(slug: str):
    """State loader cho portfolio.py — bridge `_get_state` vào contract."""
    return _get_state(slug)


def _baseline_parsed_loader(slug: str):
    """
    Phase A — load ParsedData của snapshot baseline (nếu project đã đánh dấu).
    Trả None nếu chưa set / snapshot không còn / lỗi load.
    """
    from analyzer import project_store as ps
    try:
        folder = _project_dir_for(slug)
        settings = ps.load_project_settings(folder)
        snap_id = (settings.get("baseline_snapshot_id") or "").strip()
        if not snap_id:
            return None
        smgr = _project_mgr.get_snapshot_manager(slug)
        loaded = smgr.load_snapshot(snap_id)
        if not loaded:
            return None
        return loaded.get("parsed")
    except Exception:
        return None


@app.route("/api/portfolio/search", methods=["GET"])
def portfolio_search():
    """
    Global search text qua tất cả project.
    Query params:
      - q: từ khoá (bắt buộc, min 1 ký tự)
      - scope: "name" | "code" | "pic" | "process" | "all" (default: all)
      - limit: giới hạn kết quả (default 50, max 200)
      - include_archived: 1/0 — search cả project archived (default 0)
    """
    q = (request.args.get("q") or "").strip()
    scope = (request.args.get("scope") or "all").strip().lower()
    if scope not in SEARCH_SCOPES:
        scope = "all"
    try:
        limit = int(request.args.get("limit") or 50)
    except (TypeError, ValueError):
        limit = 50
    limit = max(1, min(limit, 200))
    include_archived = request.args.get("include_archived") in ("1", "true", "yes")

    result = search_across_projects(
        _project_mgr,
        _portfolio_state_loader,
        query=q,
        scope=scope,
        limit=limit,
        include_archived=include_archived,
    )
    return jsonify({
        "success": True,
        "query": q,
        "scope": scope,
        **result,
    })


@app.route("/api/portfolio/compare", methods=["POST"])
def portfolio_compare():
    """
    Compare 2-4 project. Body JSON: {slugs: ["a", "b", ...]}
    """
    body = request.get_json(silent=True) or {}
    slugs = body.get("slugs") or []
    if not isinstance(slugs, list):
        return jsonify({"error": "slugs phải là list"}), 400
    slugs = [str(s).strip() for s in slugs if s]
    if len(slugs) < 2:
        return jsonify({"error": "Cần chọn ít nhất 2 project để so sánh"}), 400

    result = compare_projects(_project_mgr, _portfolio_state_loader, slugs)
    return jsonify({"success": True, **result})


@app.route("/api/portfolio/compare/export", methods=["POST"])
def portfolio_compare_export():
    """
    Export bảng compare summary ra Excel. Body JSON: {slugs: [...]}
    """
    body = request.get_json(silent=True) or {}
    slugs = body.get("slugs") or []
    if not isinstance(slugs, list) or len(slugs) < 2:
        return jsonify({"error": "Cần ít nhất 2 project"}), 400
    slugs = [str(s).strip() for s in slugs if s]

    result = compare_projects(_project_mgr, _portfolio_state_loader, slugs)
    try:
        # Lưu vào folder uploads chung (không thuộc project cụ thể nào)
        filepath = export_portfolio_compare(result, app.config["UPLOAD_FOLDER"])
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất Excel: {str(e)}"}), 500


@app.route("/api/portfolio/rollup", methods=["GET"])
def portfolio_rollup():
    """
    Portfolio Dashboard Rollup — gộp N project → 1 virtual dashboard.
    Query params:
      - slugs: CSV list slug (default: all non-archived)
      - include_archived: 1/0 (áp dụng khi slugs rỗng)
    """
    slugs_raw = request.args.get("slugs")
    if slugs_raw:
        slugs = [s.strip() for s in slugs_raw.split(",") if s.strip()]
    else:
        slugs = None
    include_archived = request.args.get("include_archived") in ("1", "true", "yes")

    rollup = aggregate_rollup(
        _project_mgr,
        _portfolio_state_loader,
        slugs=slugs,
        include_archived=include_archived,
    )

    if rollup["projects_count"] == 0:
        return jsonify({
            "error": "Không có project nào có file để rollup",
            "skipped": rollup["skipped"],
        }), 404

    engine = DashboardEngine()
    metrics = engine.compute_all(rollup["aggregated"])
    overridden = rollup_summary_override(rollup, metrics)

    return jsonify({
        "success": True,
        "metrics": _trim_payload(overridden),
        "per_project": rollup["per_project"],
        "skipped": rollup["skipped"],
        "projects_count": rollup["projects_count"],
    })


# ==========================================================================
# PIC Overload — đa dự án (ngày / tuần / tháng)
# ==========================================================================

def _pic_overload_thresholds_from_request() -> dict:
    """Merge settings đã lưu + query/body overrides."""
    from analyzer.pic_overload import load_overload_settings, merge_thresholds
    base = load_overload_settings(app.config["PROJECTS_FOLDER"])
    src = request.args if request.method == "GET" else (request.get_json(silent=True) or {})
    overrides = {}
    for k in (
        "day_max_tasks",
        "week_min_overload_days",
        "month_min_overload_days",
        "week_max_task_days",
        "month_max_task_days",
    ):
        if k in src and src.get(k) not in (None, ""):
            overrides[k] = src.get(k)
    if "phase_keywords" in src:
        overrides["phase_keywords"] = src.get("phase_keywords")
    return merge_thresholds({**base, **overrides})


@app.route("/api/pic-overload/settings", methods=["GET", "PUT"])
def pic_overload_settings():
    """Settings ngưỡng PIC Overload (global, đa dự án)."""
    from analyzer.pic_overload import load_overload_settings, save_overload_settings
    folder = app.config["PROJECTS_FOLDER"]
    if request.method == "GET":
        return jsonify({"success": True, "settings": load_overload_settings(folder)})
    body = request.get_json(silent=True) or {}
    saved = save_overload_settings(folder, body)
    return jsonify({"success": True, "settings": saved})


@app.route("/api/pic-overload", methods=["GET"])
def pic_overload():
    """
    PIC Overload đa dự án.
    Query:
      grain=day|week|month
      from=, to= (YYYY-MM-DD)
      pic= (optional filter)
      include_archived=0|1
      day_max_tasks= (override)
      phase_keywords=Dev,Config (optional)
    """
    from analyzer.pic_overload import compute_pic_overload, VALID_GRAINS
    grain = (request.args.get("grain") or "day").strip().lower()
    if grain not in VALID_GRAINS:
        grain = "day"
    thr = _pic_overload_thresholds_from_request()
    include_archived = request.args.get("include_archived") in ("1", "true", "yes")
    pic = (request.args.get("pic") or "").strip() or None
    try:
        detail_limit = int(request.args.get("detail_limit") or 5000)
    except (TypeError, ValueError):
        detail_limit = 5000
    detail_limit = max(100, min(detail_limit, 20000))

    result = compute_pic_overload(
        _project_mgr,
        _portfolio_state_loader,
        grain=grain,
        date_from=request.args.get("from"),
        date_to=request.args.get("to"),
        thresholds=thr,
        pic_filter=pic,
        include_archived=include_archived,
        detail_limit=detail_limit,
    )
    return jsonify({"success": True, **result})


@app.route("/api/pic-overload/export", methods=["GET", "POST"])
def pic_overload_export():
    """
    Xuất Excel PIC Overload.
    Body/query: grain, from, to, mode=summary|detail|both, include_fl=0|1, pic=
    """
    from analyzer.pic_overload import compute_pic_overload, VALID_GRAINS
    from exporter.pic_overload_exporter import export_pic_overload_report

    body = request.get_json(silent=True) or {}
    src = {**request.args.to_dict(), **body}
    grain = (src.get("grain") or "day").strip().lower()
    if grain not in VALID_GRAINS:
        grain = "day"
    mode = (src.get("mode") or "both").strip().lower()
    include_fl = str(src.get("include_fl") or "0").lower() in ("1", "true", "yes")
    thr = _pic_overload_thresholds_from_request()
    # Re-merge body overrides for thresholds
    for k in (
        "day_max_tasks",
        "week_min_overload_days",
        "month_min_overload_days",
        "week_max_task_days",
        "month_max_task_days",
        "phase_keywords",
    ):
        if k in src and src.get(k) not in (None, ""):
            thr[k] = src[k]
    from analyzer.pic_overload import merge_thresholds
    thr = merge_thresholds(thr)

    pic = (src.get("pic") or "").strip() or None
    result = compute_pic_overload(
        _project_mgr,
        _portfolio_state_loader,
        grain=grain,
        date_from=src.get("from"),
        date_to=src.get("to"),
        thresholds=thr,
        pic_filter=pic,
        include_archived=str(src.get("include_archived") or "").lower() in ("1", "true", "yes"),
    )

    project_data = None
    project_dirs = None
    if include_fl:
        project_data = {}
        project_dirs = {}
        slugs = {d.get("project_slug") for d in (result.get("detail") or []) if d.get("project_slug")}
        for slug in slugs:
            st = _get_state(slug)
            if st and st.get("data"):
                project_data[slug] = st["data"]
                try:
                    project_dirs[slug] = _project_dir_for(slug)
                except Exception:
                    pass

    try:
        filepath = export_pic_overload_report(
            result,
            output_dir=app.config["UPLOAD_FOLDER"],
            mode=mode,
            include_fl=include_fl,
            project_data=project_data,
            project_dirs=project_dirs,
        )
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath),
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất Excel: {str(e)}"}), 500


@app.route("/api/forecast-gantt", methods=["GET", "POST"])
def forecast_gantt():
    """
    Forecast Gantt — tháng UAT/Golive (+ Phân tích/Dev/Cấu hình) đa dự án.

    Query/body:
      slugs=a,b,c  (optional; trống = mọi project active có file)
      include_archived=0|1
    """
    from analyzer.forecast_gantt import compute_forecast_gantt

    body = request.get_json(silent=True) or {}
    raw = request.args.get("slugs") or body.get("slugs") or ""
    if isinstance(raw, list):
        slugs = [str(s).strip() for s in raw if str(s).strip()]
    else:
        slugs = [s.strip() for s in str(raw).split(",") if s.strip()]
    include_archived = (
        request.args.get("include_archived") in ("1", "true", "yes")
        or body.get("include_archived") in (True, 1, "1", "true", "yes")
    )
    result = compute_forecast_gantt(
        _project_mgr,
        _portfolio_state_loader,
        slugs=slugs or None,
        include_archived=include_archived,
        baseline_loader=_baseline_parsed_loader,
    )
    return jsonify({"success": True, **result})


@app.route("/api/forecast-gantt/export", methods=["GET", "POST"])
def forecast_gantt_export():
    """Xuất Excel nhẹ Forecast Gantt (cùng filter slugs)."""
    from analyzer.forecast_gantt import compute_forecast_gantt
    from exporter.forecast_gantt_exporter import export_forecast_gantt

    body = request.get_json(silent=True) or {}
    raw = request.args.get("slugs") or body.get("slugs") or ""
    if isinstance(raw, list):
        slugs = [str(s).strip() for s in raw if str(s).strip()]
    else:
        slugs = [s.strip() for s in str(raw).split(",") if s.strip()]
    include_archived = (
        request.args.get("include_archived") in ("1", "true", "yes")
        or body.get("include_archived") in (True, 1, "1", "true", "yes")
    )
    result = compute_forecast_gantt(
        _project_mgr,
        _portfolio_state_loader,
        slugs=slugs or None,
        include_archived=include_archived,
        baseline_loader=_baseline_parsed_loader,
    )
    try:
        filepath = export_forecast_gantt(result, app.config["UPLOAD_FOLDER"])
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath),
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất Excel: {str(e)}"}), 500


# ==========================================================================
# Forecast Manpower — MH / MD / MM + nhu cầu tuyển theo công đoạn
# ==========================================================================

def _forecast_manpower_params():
    """Parse query/body chung cho forecast-manpower."""
    body = request.get_json(silent=True) or {}
    src = {**request.args.to_dict(), **body}
    basis = (src.get("basis") or "unit").strip().lower()
    if basis not in ("unit", "duration"):
        basis = "unit"
    display_unit = (src.get("unit") or src.get("display_unit") or "manhour").strip().lower()
    if display_unit in ("mh", "hour", "hours"):
        display_unit = "manhour"
    if display_unit in ("md", "day", "days"):
        display_unit = "manday"
    if display_unit in ("mm", "month", "months"):
        display_unit = "manmonth"
    if display_unit not in ("manhour", "manday", "manmonth"):
        display_unit = "manhour"
    try:
        default_mh = float(src.get("default_mh") or 8)
    except (TypeError, ValueError):
        default_mh = 8.0
    try:
        target_months = float(src.get("target_months") or 1)
    except (TypeError, ValueError):
        target_months = 1.0
    headcount = src.get("headcount") or {}
    if isinstance(headcount, str):
        import json as _json
        raw_hc = headcount.strip()
        headcount = {}
        if raw_hc.startswith("{"):
            try:
                headcount = _json.loads(raw_hc)
            except Exception:
                headcount = {}
        else:
            for part in raw_hc.split(","):
                if ":" in part:
                    k, v = part.split(":", 1)
                    try:
                        headcount[k.strip()] = float(v.strip())
                    except ValueError:
                        pass
    if not isinstance(headcount, dict):
        headcount = {}
    # Also accept flat keys hc_dev, hc_impl, hc_<stage>
    for k, v in list(src.items()):
        if k.startswith("hc_") and v not in (None, ""):
            try:
                headcount[k[3:]] = float(v)
            except (TypeError, ValueError):
                pass
    modules = src.get("module") or src.get("modules") or ""
    processes = src.get("process") or src.get("processes") or ""
    pics = src.get("pic") or src.get("pics") or ""
    if isinstance(modules, str):
        modules = [m.strip() for m in modules.split(",") if m.strip()]
    if isinstance(processes, str):
        processes = [p.strip() for p in processes.split(",") if p.strip()]
    if isinstance(pics, str):
        pics = [p.strip() for p in pics.split(",") if p.strip()]
    hc_clean = {}
    for k, v in headcount.items():
        try:
            if v is not None and str(v).strip() != "":
                hc_clean[str(k)] = float(v)
        except (TypeError, ValueError):
            pass
    return {
        "basis": basis,
        "display_unit": display_unit,
        "default_mh": default_mh,
        "target_months": target_months,
        "headcount": hc_clean,
        "modules": modules or None,
        "processes": processes or None,
        "pics": pics or None,
        "mode": (src.get("mode") or "both").strip().lower(),
    }


@app.route("/api/projects/<slug>/forecast-manpower", methods=["GET", "POST"])
def forecast_manpower(slug: str):
    """
    Ước lượng manhours / mandays / manmonths + nhu cầu tuyển theo công đoạn.

    Params: basis=unit|duration, unit=manhour|manday|manmonth,
            default_mh=8, target_months=1, headcount={...} hoặc hc_dev / hc_impl,
            module/process/pic filters.
    """
    from analyzer.forecast_manpower import compute_forecast_manpower

    state = _get_state(slug)
    if not state or not state.get("data"):
        return jsonify({"error": "Chưa có dữ liệu. Hãy upload Function List."}), 400
    p = _forecast_manpower_params()
    data = _filter_parsed_data(
        state["data"],
        modules=p["modules"],
        processes=p["processes"],
        pics=p["pics"],
    )
    result = compute_forecast_manpower(
        data,
        basis=p["basis"],
        display_unit=p["display_unit"],
        default_mh=p["default_mh"],
        target_months=p["target_months"],
        headcount=p["headcount"],
    )
    return jsonify({"success": True, **result})


@app.route("/api/projects/<slug>/export-forecast-manpower", methods=["GET", "POST"])
def export_forecast_manpower_api(slug: str):
    """Xuất Excel Forecast Manpower (Tong_hop / Chi_tiet / both)."""
    from analyzer.forecast_manpower import compute_forecast_manpower
    from exporter.forecast_manpower_exporter import export_forecast_manpower

    state = _get_state(slug)
    if not state or not state.get("data"):
        return jsonify({"error": "Chưa có dữ liệu. Hãy upload Function List."}), 400
    p = _forecast_manpower_params()
    data = _filter_parsed_data(
        state["data"],
        modules=p["modules"],
        processes=p["processes"],
        pics=p["pics"],
    )
    result = compute_forecast_manpower(
        data,
        basis=p["basis"],
        display_unit=p["display_unit"],
        default_mh=p["default_mh"],
        target_months=p["target_months"],
        headcount=p["headcount"],
    )
    try:
        filepath = export_forecast_manpower(
            result, app.config["UPLOAD_FOLDER"], mode=p["mode"]
        )
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath),
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất Excel: {str(e)}"}), 500


# ==========================================================================
# Estimate Ratio — ước lượng theo hệ số (parametric, không thay Forecast MH)
# ==========================================================================

def _estimate_ratio_filters():
    """Parse module/process/pic filter từ query hoặc JSON body."""
    body = request.get_json(silent=True) or {}
    src = {**request.args.to_dict(), **body}
    modules = src.get("module") or src.get("modules") or ""
    processes = src.get("process") or src.get("processes") or ""
    pics = src.get("pic") or src.get("pics") or ""
    if isinstance(modules, str):
        modules = [m.strip() for m in modules.split(",") if m.strip()]
    if isinstance(processes, str):
        processes = [p.strip() for p in processes.split(",") if p.strip()]
    if isinstance(pics, str):
        pics = [p.strip() for p in pics.split(",") if p.strip()]
    return {
        "modules": modules or None,
        "processes": processes or None,
        "pics": pics or None,
    }


@app.route("/api/projects/<slug>/estimate-ratio", methods=["GET", "POST"])
def estimate_ratio_api(slug: str):
    """
    GET: tính ước lượng theo hệ số (+ trả params hiện tại).
    POST body ``{"action":"save","scope":"project|global","params":{...}}``
         lưu estimation_params.json rồi tính lại.
    POST không action / action=compute: tính với params override trong body (không lưu).
    """
    from analyzer.estimate_ratio import (
        compute_estimate_ratio,
        load_estimation_params,
        normalize_params,
        save_estimation_params,
    )

    state = _get_state(slug)
    if not state or not state.get("data"):
        return jsonify({"error": "Chưa có dữ liệu. Hãy upload Function List."}), 400

    project_dir = _project_dir_for(slug)
    projects_folder = _project_mgr.base_dir
    body = request.get_json(silent=True) or {}
    action = (body.get("action") or request.args.get("action") or "").strip().lower()

    if request.method == "POST" and action == "save":
        scope = (body.get("scope") or "project").strip().lower()
        if scope not in ("project", "global"):
            scope = "project"
        try:
            params = save_estimation_params(
                project_dir,
                body.get("params") or {},
                scope=scope,
                projects_folder=projects_folder,
            )
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
    else:
        params = load_estimation_params(project_dir, projects_folder)
        # Override tạm (không lưu) nếu client gửi params
        if isinstance(body.get("params"), dict) and body["params"]:
            from analyzer.estimate_ratio import _deep_merge
            raw = {k: v for k, v in params.items() if not k.startswith("_")}
            params = normalize_params(_deep_merge(raw, body["params"]))

    filt = _estimate_ratio_filters()
    data = _filter_parsed_data(
        state["data"],
        modules=filt["modules"],
        processes=filt["processes"],
        pics=filt["pics"],
    )
    # Giữ metadata nguồn/paths trên response params
    stored = load_estimation_params(project_dir, projects_folder)
    result = compute_estimate_ratio(data, params)
    result["params_meta"] = {
        "source": stored.get("_source"),
        "paths": stored.get("_paths"),
    }
    # Trả params đã dùng (đã normalize); kèm source nếu không override
    if not (isinstance(body.get("params"), dict) and body["params"] and action != "save"):
        result["params"]["_source"] = stored.get("_source")
        result["params"]["_paths"] = stored.get("_paths")
    return jsonify({"success": True, **result})


@app.route("/api/projects/<slug>/estimation-params", methods=["GET", "PUT", "POST"])
def estimation_params_api(slug: str):
    """GET/PUT estimation_params.json (project hoặc global qua ?scope=)."""
    from analyzer.estimate_ratio import load_estimation_params, save_estimation_params

    project_dir = _project_dir_for(slug)
    projects_folder = _project_mgr.base_dir
    if request.method == "GET":
        params = load_estimation_params(project_dir, projects_folder)
        return jsonify({"success": True, "params": params})
    body = request.get_json(silent=True) or {}
    scope = (
        body.get("scope")
        or request.args.get("scope")
        or "project"
    ).strip().lower()
    if scope not in ("project", "global"):
        scope = "project"
    payload = body.get("params") if isinstance(body.get("params"), dict) else body
    try:
        saved = save_estimation_params(
            project_dir,
            payload,
            scope=scope,
            projects_folder=projects_folder,
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"success": True, "params": saved})


# ==========================================================================
# Project stores — capacity / saved views / history / settings / aliases
# ==========================================================================

def _project_dir_for(slug: str) -> str:
    return _project_mgr.get_project_folder(slug)


def _apply_module_order_to_data(slug: str, data) -> None:
    """Áp module_order.json lên ParsedData.all_modules (in-place)."""
    from analyzer import project_store as ps
    from analyzer.module_order import apply_module_order
    order = ps.load_module_order(_project_dir_for(slug))
    data.all_modules = apply_module_order(list(data.all_modules or []), order)


def _recompute_metrics_with_module_order(slug: str) -> Optional[dict]:
    """
    Re-apply module order lên state đang cache + recompute metrics.
    Gọi sau khi user lưu/reset thứ tự Module.
    """
    st = _get_state(slug)
    if not st or st.get("data") is None:
        return None
    _apply_module_order_to_data(slug, st["data"])
    st["metrics"] = DashboardEngine().compute_all(st["data"])
    return st["metrics"]


@app.route("/api/projects/<slug>/capacity", methods=["GET", "PUT"])
def project_capacity(slug: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify(ps.load_capacity(folder))
    body = request.get_json(silent=True) or {}
    return jsonify(ps.save_capacity(folder, body))


@app.route("/api/projects/<slug>/saved-views", methods=["GET", "POST"])
def project_saved_views(slug: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"views": ps.load_saved_views(folder)})
    body = request.get_json(silent=True) or {}
    views = ps.upsert_saved_view(folder, body)
    return jsonify({"views": views})


@app.route("/api/projects/<slug>/saved-views/<view_id>", methods=["DELETE"])
def project_delete_saved_view(slug: str, view_id: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    views = ps.delete_saved_view(_project_dir_for(slug), view_id)
    return jsonify({"views": views})


# ==========================================================================
# Section order (Task 4b — drag-drop customize)
# ==========================================================================

@app.route("/api/projects/<slug>/section-order", methods=["GET", "POST"])
def project_section_order(slug: str):
    """
    GET → trả section_order hiện tại (list id string, hoặc [] nếu chưa set).
    POST → body {order: [id1, id2, ...]} — lưu và trả về order đã lưu.
    """
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"order": ps.load_section_order(folder)})
    body = request.get_json(silent=True) or {}
    order = body.get("order") or body.get("section_order") or []
    if not isinstance(order, list):
        return jsonify({"error": "order phải là list"}), 400
    saved = ps.save_section_order(folder, order)
    return jsonify({"order": saved})


@app.route("/api/projects/<slug>/section-order/reset", methods=["POST"])
def project_section_order_reset(slug: str):
    """Xoá custom section_order → FE fallback về HTML default."""
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    ps.reset_section_order(_project_dir_for(slug))
    return jsonify({"order": []})


# ==========================================================================
# Module order — thứ tự Module dùng chung toàn dashboard
# ==========================================================================

@app.route("/api/projects/<slug>/module-order", methods=["GET", "POST", "PUT"])
def project_module_order(slug: str):
    """
    GET  → {order: [...], detected: [...]}  (detected = modules trong data hiện tại)
    POST/PUT → body {order: [TMS, HR, ...]} — lưu + recompute metrics in-memory.
    """
    from analyzer import project_store as ps
    from analyzer.module_order import sort_modules
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    st = _get_state(slug)
    detected: list[str] = []
    if st and st.get("data") is not None:
        detected = list(getattr(st["data"], "all_modules", []) or [])

    if request.method == "GET":
        saved = ps.load_module_order(folder)
        # Trả order hiệu lực (saved + module mới alpha ở cuối) để UI hiển thị đủ
        effective = sort_modules(detected, saved) if detected else list(saved)
        return jsonify({
            "order": saved,
            "effective": effective,
            "detected": detected,
        })

    body = request.get_json(silent=True) or {}
    order = body.get("order") or body.get("module_order") or []
    if not isinstance(order, list):
        return jsonify({"error": "order phải là list"}), 400
    saved = ps.save_module_order(folder, order)
    metrics = _recompute_metrics_with_module_order(slug)
    st2 = _get_state(slug)
    detected_after: list[str] = []
    if st2 and st2.get("data") is not None:
        detected_after = list(getattr(st2["data"], "all_modules", []) or [])
    return jsonify({
        "order": saved,
        "effective": detected_after or sort_modules(detected, saved),
        "detected": detected_after or detected,
        "metrics_updated": metrics is not None,
    })


@app.route("/api/projects/<slug>/module-order/reset", methods=["POST"])
def project_module_order_reset(slug: str):
    """Xoá module_order.json → alphabetical + recompute."""
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    ps.reset_module_order(_project_dir_for(slug))
    metrics = _recompute_metrics_with_module_order(slug)
    st = _get_state(slug)
    detected: list[str] = []
    if st and st.get("data") is not None:
        detected = list(getattr(st["data"], "all_modules", []) or [])
    return jsonify({
        "order": [],
        "effective": detected,
        "detected": detected,
        "metrics_updated": metrics is not None,
    })


# ==========================================================================
# Chart notes (T28 — comment per-chart + tóm tắt chung cho PDF export)
# ==========================================================================

@app.route("/api/projects/<slug>/chart-notes", methods=["GET", "PUT"])
def project_chart_notes(slug: str):
    """
    GET → trả về {summary: str, notes: {section_id: text}}.
    PUT body {summary?: str, notes?: {section_id: text}} → merge và lưu.
    - summary: tóm tắt chung của báo cáo (max 500 ký tự) — hiển thị ở cover PDF.
    - notes: comment per-chart (max 200 ký tự / chart) — value rỗng = xoá key.
    """
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify(ps.load_chart_notes(folder))
    body = request.get_json(silent=True) or {}
    saved = ps.save_chart_notes(folder, body)
    return jsonify(saved)


# ==========================================================================
# Chart configs (Task 6 — Phase A: inline edit title/caption/visibility)
# ==========================================================================

@app.route("/api/projects/<slug>/chart-config", methods=["GET", "POST", "DELETE"])
def project_chart_config(slug: str):
    """
    GET → trả full map {target_id: {title?, caption?, hidden?}}.
    POST → body {target_id, title?, caption?, hidden?} → upsert 1 chart.
           Nếu tất cả field trong body rỗng → xoá config cho target đó.
    DELETE:
      · ?target=<id> → xoá 1 target.
      · Không query → reset toàn bộ chart configs.
    """
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)

    if request.method == "GET":
        return jsonify({"configs": ps.load_chart_configs(folder)})

    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        target_id = body.get("target_id") or body.get("chart_id") or ""
        if not target_id:
            return jsonify({"error": "target_id bắt buộc"}), 400
        # Pass through TOÀN BỘ field (Phase A + Phase B). Sanitize logic ở
        # project_store._sanitize_chart_config sẽ chỉ giữ key hợp lệ.
        entry: dict = {
            "title": body.get("title"),
            "caption": body.get("caption"),
            "hidden": bool(body.get("hidden")),
            "type": body.get("type") or body.get("chart_type"),
            "x_field": body.get("x_field"),
            "y_measure": body.get("y_measure"),
            "series_field": body.get("series_field"),
            "palette": body.get("palette") or body.get("colors"),
            "filter_override": body.get("filter_override"),
        }
        all_cfg = ps.upsert_chart_config(folder, target_id, entry)
        return jsonify({"configs": all_cfg})

    # DELETE
    target_id = request.args.get("target", "").strip()
    if target_id:
        all_cfg = ps.delete_chart_config(folder, target_id)
        return jsonify({"configs": all_cfg})
    ps.reset_chart_configs(folder)
    return jsonify({"configs": {}})


@app.route("/api/projects/<slug>/chart-config/visibility", methods=["PUT"])
def project_chart_visibility(slug: str):
    """
    Bulk toggle hiển thị nhiều section cùng lúc — dùng cho tab "Hiển thị"
    trong Settings modal.

    Body: {"visibility": {"section-xxx": true, "section-yyy": false, ...}}
      · true  → hiển thị (xoá cờ hidden nếu có)
      · false → ẩn (đặt cờ hidden = true, preserve các field khác)

    Trả về: {"configs": <full chart_configs map>} — FE có thể apply ngay.
    """
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    body = request.get_json(silent=True) or {}
    mapping = body.get("visibility")
    if not isinstance(mapping, dict):
        return jsonify({"error": "visibility phải là dict {section_id: bool}"}), 400
    # Coerce sang bool để không lệ thuộc payload gửi 0/1/"true"/"false"
    coerced = {str(k): bool(v) for k, v in mapping.items() if k}
    folder = _project_dir_for(slug)
    all_cfg = ps.set_chart_config_visibility(folder, coerced)
    return jsonify({"configs": all_cfg})


@app.route("/api/projects/<slug>/chart-fields")
def project_chart_fields(slug: str):
    """
    Task 8: trả về danh sách field / measure / chart_type / palette phù hợp
    với FE dropdown. Không phụ thuộc file uploaded (static enum).
    """
    from analyzer.generic_chart import get_available_fields
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    return jsonify(get_available_fields())


# ==========================================================================
# Custom dashboards (Task 9 — Dynamic Dashboard Builder)
# ==========================================================================

@app.route("/api/projects/<slug>/custom-dashboard", methods=["GET", "POST"])
def project_custom_dashboards(slug: str):
    """GET → list toàn bộ. POST → thêm/update 1 custom dashboard."""
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"items": ps.load_custom_dashboards(folder)})
    body = request.get_json(silent=True) or {}
    saved = ps.upsert_custom_dashboard(folder, body)
    if not saved:
        return jsonify({"error": "Dữ liệu không hợp lệ (thiếu title hoặc x_field)"}), 400
    return jsonify({"item": saved, "items": ps.load_custom_dashboards(folder)})


@app.route("/api/projects/<slug>/custom-dashboard/<item_id>", methods=["PUT", "DELETE"])
def project_custom_dashboard_item(slug: str, item_id: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "PUT":
        body = request.get_json(silent=True) or {}
        body["id"] = item_id
        saved = ps.upsert_custom_dashboard(folder, body)
        if not saved:
            return jsonify({"error": "Dữ liệu không hợp lệ"}), 400
        return jsonify({"item": saved})
    items = ps.delete_custom_dashboard(folder, item_id)
    return jsonify({"items": items})


@app.route("/api/projects/<slug>/custom-dashboard/<item_id>/data")
def project_custom_dashboard_data(slug: str, item_id: str):
    """Aggregate data cho 1 custom dashboard — reuse aggregate_chart từ Task 8."""
    from analyzer import project_store as ps
    from analyzer.generic_chart import aggregate_chart
    state, err = _require_state(slug)
    if err:
        return err
    items = ps.load_custom_dashboards(_project_dir_for(slug))
    item = next((i for i in items if i.get("id") == item_id), None)
    if not item:
        return jsonify({"error": "Custom dashboard không tồn tại"}), 404
    filters = dict(item.get("filters") or {})
    # Merge global filter nếu query có
    for k, gk in [("modules", "module"), ("processes", "process"), ("pics", "pic")]:
        gv = _parse_multi_arg(gk)
        if gv:
            filters[k] = list(set(filters.get(k, []) + gv)) or gv
    try:
        result = aggregate_chart(
            state["data"],
            x_field=item["x_field"],
            y_measure=item.get("y_measure", "count"),
            series_field=item.get("series_field") or None,
            filters=filters,
        )
    except Exception as e:
        return jsonify({"error": f"Aggregate failed: {e}"}), 400
    result["config"] = item
    return jsonify(result)


@app.route("/api/projects/<slug>/custom-dashboard/<item_id>/drill")
def project_custom_dashboard_drill(slug: str, item_id: str):
    """T27 — Drill-down cho 1 bucket của custom dashboard.

    Query params:
        x_value        (bắt buộc) — label bucket trên trục X
        series_value   (optional) — nếu chart có series/group field
        limit          (optional, default 500) — cap số row trả
    """
    from analyzer import project_store as ps
    from analyzer.generic_chart import drill_chart
    state, err = _require_state(slug)
    if err:
        return err
    items = ps.load_custom_dashboards(_project_dir_for(slug))
    item = next((i for i in items if i.get("id") == item_id), None)
    if not item:
        return jsonify({"error": "Custom dashboard không tồn tại"}), 404
    x_value = request.args.get("x_value")
    if not x_value:
        return jsonify({"error": "Thiếu x_value"}), 400
    series_value = request.args.get("series_value") or ""
    try:
        limit = max(1, min(2000, int(request.args.get("limit", 500))))
    except (TypeError, ValueError):
        limit = 500
    filters = dict(item.get("filters") or {})
    # Merge global filter
    for k, gk in [("modules", "module"), ("processes", "process"), ("pics", "pic")]:
        gv = _parse_multi_arg(gk)
        if gv:
            filters[k] = list(set(filters.get(k, []) + gv)) or gv
    try:
        result = drill_chart(
            state["data"],
            x_field=item["x_field"],
            x_value=x_value,
            series_field=item.get("series_field") or None,
            series_value=series_value or None,
            filters=filters,
            limit=limit,
        )
    except Exception as e:
        return jsonify({"error": f"Drill failed: {e}"}), 400
    # Kèm meta dashboard để FE hiển thị title/caption trong modal
    result["dashboard"] = {
        "id": item.get("id"),
        "title": item.get("title"),
        "caption": item.get("caption"),
        "x_field": item.get("x_field"),
        "y_measure": item.get("y_measure"),
        "series_field": item.get("series_field"),
    }
    return jsonify(result)


@app.route("/api/projects/<slug>/custom-dashboard/<item_id>/export")
def project_custom_dashboard_export(slug: str, item_id: str):
    """Xuất Excel — 1 sheet chứa aggregated data + 1 sheet metadata."""
    from analyzer import project_store as ps
    from analyzer.generic_chart import aggregate_chart, FIELDS, MEASURES
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from tempfile import mkdtemp
    from datetime import datetime
    import os
    state, err = _require_state(slug)
    if err:
        return err
    items = ps.load_custom_dashboards(_project_dir_for(slug))
    item = next((i for i in items if i.get("id") == item_id), None)
    if not item:
        return jsonify({"error": "Custom dashboard không tồn tại"}), 404
    agg = aggregate_chart(
        state["data"],
        x_field=item["x_field"],
        y_measure=item.get("y_measure", "count"),
        series_field=item.get("series_field") or None,
        filters=item.get("filters") or {},
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    xlabel = FIELDS.get(item["x_field"], item["x_field"])
    headers = [xlabel] + [ds["label"] for ds in agg["datasets"]]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")
    for i, x in enumerate(agg["labels"]):
        row = [x] + [ds["data"][i] for ds in agg["datasets"]]
        ws.append(row)
    for col_idx, _ in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)].width = 18
    meta_ws = wb.create_sheet("Info")
    meta_rows = [
        ["Title", item.get("title")],
        ["Caption", item.get("caption")],
        ["X field", xlabel],
        ["Y measure", MEASURES.get(item.get("y_measure"), item.get("y_measure"))],
        ["Series field", FIELDS.get(item.get("series_field"), item.get("series_field") or "-")],
        ["Chart type", item.get("chart_type")],
        ["Palette", item.get("palette")],
        ["Filters", str(item.get("filters") or {})],
        ["Created at", item.get("created_at")],
        ["Total rows after filter", agg["meta"].get("total_rows_after_filter")],
        ["Exported at", datetime.now().isoformat(timespec="seconds")],
    ]
    for r in meta_rows:
        meta_ws.append(r)
    meta_ws.column_dimensions["A"].width = 22
    meta_ws.column_dimensions["B"].width = 60
    for c in meta_ws["A"]:
        c.font = Font(bold=True)
    out_dir = mkdtemp(prefix="custom_dash_")
    safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in item.get("title", "custom"))[:40]
    fpath = os.path.join(out_dir, f"custom_{safe_name}_{item_id[:12]}.xlsx")
    wb.save(fpath)
    return send_file(fpath, as_attachment=True,
                     download_name=os.path.basename(fpath))


# ==========================================================================
# T30: Registry API + Đồng bộ dữ liệu
# ==========================================================================
# Cho phép user cấu hình danh sách API/endpoint từ nhiều ứng dụng nguồn (iHRP
# prod, workload report, GAP list…). Bấm "Sync" → tự login → tải Excel → parse
# → tạo snapshot mới. Credential lưu ở `.env`, KHÔNG trong JSON.
#
# Storage: uploads/projects/<slug>/integrations.json (list các integration).
# Module logic: analyzer/integrations.py.
# ==========================================================================


@app.route("/api/projects/<slug>/integrations", methods=["GET", "POST"])
def project_integrations(slug: str):
    """GET → danh sách integrations. POST → tạo mới, body JSON schema xem docs."""
    from analyzer import integrations as integ_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({
            "integrations": integ_mod.list_integrations(
                folder,
                source_app=request.args.get("source_app") or "",
                env=request.args.get("env") or "",
                visibility=request.args.get("visibility") or "",
                q=request.args.get("q") or "",
            ),
            "capabilities": integ_mod.integration_capabilities(),
        })
    body = request.get_json(silent=True) or {}
    try:
        created = integ_mod.create_integration(folder, body)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"integration": created}), 201


@app.route("/api/projects/<slug>/integrations/<integration_id>",
           methods=["GET", "PUT", "DELETE"])
def project_integration_detail(slug: str, integration_id: str):
    from analyzer import integrations as integ_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        integ = integ_mod.get_integration(folder, integration_id)
        if not integ:
            return jsonify({"error": "Không tìm thấy integration"}), 404
        return jsonify({"integration": integ})
    if request.method == "PUT":
        body = request.get_json(silent=True) or {}
        try:
            updated = integ_mod.update_integration(folder, integration_id, body)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        if not updated:
            return jsonify({"error": "Không tìm thấy integration"}), 404
        return jsonify({"integration": updated})
    # DELETE
    ok = integ_mod.delete_integration(folder, integration_id)
    if not ok:
        return jsonify({"error": "Không tìm thấy integration"}), 404
    return jsonify({"success": True})


@app.route("/api/projects/<slug>/integrations/<integration_id>/test",
           methods=["POST"])
def project_integration_test(slug: str, integration_id: str):
    """Test login only — không tải Excel, chỉ verify creds + URL."""
    from analyzer import integrations as integ_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    result = integ_mod.test_integration(folder, integration_id)
    # Trả 200 kể cả khi status=error để FE hiển thị message rõ; HTTP 500 để
    # trường hợp integration không tồn tại thì tiện log.
    status_code = 200 if result.get("status") == "ok" else 200
    if "Không tìm thấy" in (result.get("message") or ""):
        status_code = 404
    return jsonify(result), status_code


@app.route("/api/projects/<slug>/integrations/<integration_id>/test-db",
           methods=["POST"])
def project_integration_test_db(slug: str, integration_id: str):
    """
    T31 — Test kết nối DB cho integration có auth.method='database'.
    Chỉ mở connection + ping `SELECT 1`, KHÔNG chạy query của endpoint.
    Alias tiện lợi cho FE — logic backend giống hệt `/test` với method=database.
    """
    from analyzer import integrations as integ_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    integ = integ_mod.get_integration(folder, integration_id)
    if not integ:
        return jsonify({"status": "error", "message": "Không tìm thấy integration"}), 404
    method = (integ.get("auth") or {}).get("method") or ""
    if method != "database":
        return jsonify({
            "status": "error",
            "message": f"Endpoint này dùng auth.method='{method}', không phải database. Dùng /test.",
        }), 400
    result = integ_mod.test_database_connection(integ.get("auth") or {})
    integ_mod._update_last_status(  # type: ignore[attr-defined]
        folder, integration_id, result.get("status", "error"), result.get("message", ""))
    return jsonify(result)


@app.route("/api/projects/<slug>/integrations/<integration_id>/sync",
           methods=["POST"])
def project_integration_sync(slug: str, integration_id: str):
    """Đồng bộ 1 endpoint.

    Body:
      - endpoint_id (required)
      - selected_map (optional): {mã nguồn: slug local} — chỉ sync mã đã chọn
      - project_code_filter (optional): str | list mã
      - persist_map (optional, default true): lưu selected_map vào config
    """
    from analyzer import integrations as integ_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    body = request.get_json(silent=True) or {}
    endpoint_id = (body.get("endpoint_id") or "").strip()
    if not endpoint_id:
        return jsonify({"status": "error", "message": "Thiếu 'endpoint_id'"}), 400

    selected_map = body.get("selected_map")
    if selected_map is None and "project_code_map" in body:
        # Alias: FE có thể gửi project_code_map thay selected_map
        selected_map = body.get("project_code_map")
    if selected_map is not None and not isinstance(selected_map, dict):
        return jsonify({
            "status": "error",
            "message": "'selected_map' phải là object {mã: slug}",
        }), 400

    project_code_filter = body.get("project_code_filter")
    persist_map = body.get("persist_map")
    if persist_map is None:
        persist_map = True

    folder = _project_dir_for(slug)
    result = integ_mod.sync_integration(
        project_dir=folder,
        integration_id=integration_id,
        endpoint_id=endpoint_id,
        project_manager=_project_mgr,
        project_slug=slug,
        selected_map=selected_map,
        project_code_filter=project_code_filter,
        persist_map=bool(persist_map),
    )

    # Nếu sync ok → invalidate + eager-reload state từ snapshot mới nhất.
    # Chỉ pop() rồi để /dashboard tự load dễ race: request song song (DQ /
    # aging / gantt…) có thể nạp lại pickle cũ vào _state SAU khi pop nhưng
    # TRƯỚC khi FE gọi /dashboard → UI giữ Sync timestamp / metrics cũ.
    # Multi-project routing: refresh tất cả slug đã nhận data.
    if result.get("status") == "ok":
        synced = list(result.get("synced_slugs") or [slug])
        for s in synced:
            _state.pop(s, None)
            loaded = _load_state_from_disk(s)
            if loaded:
                _state[s] = loaded
        # Metadata để FE cập nhật header Sync ngay (không đợi parse response dashboard).
        primary = slug if slug in synced else (synced[0] if synced else slug)
        st_primary = _state.get(primary)
        if st_primary:
            ut = st_primary.get("upload_time")
            if ut is not None and hasattr(ut, "isoformat"):
                result["upload_time"] = ut.isoformat(timespec="seconds")
            elif ut:
                result["upload_time"] = str(ut)
            result["filename"] = st_primary.get("filename")
            result["dashboard_ready"] = True

    return jsonify(result)


@app.route("/api/projects/<slug>/integrations/<integration_id>/project-codes",
           methods=["POST"])
def project_integration_project_codes(slug: str, integration_id: str):
    """
    Preview unique mã dự án từ endpoint (không tạo snapshot).
    Body: {"endpoint_id": "..."}.
    Trả: {status, project_codes:[{code,count}], project_code_map, routing_available}.
    """
    from analyzer import integrations as integ_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    body = request.get_json(silent=True) or {}
    endpoint_id = (body.get("endpoint_id") or "").strip()
    if not endpoint_id:
        return jsonify({"status": "error", "message": "Thiếu 'endpoint_id'"}), 400
    folder = _project_dir_for(slug)
    result = integ_mod.list_endpoint_project_codes(
        project_dir=folder,
        integration_id=integration_id,
        endpoint_id=endpoint_id,
    )
    return jsonify(result)


@app.route("/api/projects/<slug>/integrations/<integration_id>/preview-json",
           methods=["POST"])
def project_integration_preview_json(slug: str, integration_id: str):
    """
    Preview 1 endpoint JSON để FE auto-suggest field_mapping.
    Body: {"endpoint_id": "..."}.
    Trả: {status, sample_records, flat_keys, record_count, field_types}.
    Không tạo snapshot, không thay đổi state.

    T34 Task 3A — Thêm `field_types` với sample values + type inference cho
    mỗi JSON path để UI Field Mapping panel hiển thị badge.
    """
    from analyzer import integrations as integ_mod
    from analyzer.type_infer import infer_all_headers
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    body = request.get_json(silent=True) or {}
    endpoint_id = (body.get("endpoint_id") or "").strip()
    if not endpoint_id:
        return jsonify({"status": "error", "message": "Thiếu 'endpoint_id'"}), 400
    folder = _project_dir_for(slug)
    result = integ_mod.preview_json_endpoint(
        project_dir=folder,
        integration_id=integration_id,
        endpoint_id=endpoint_id,
    )

    # T34 Task 3A+B — Sinh field_types dùng type inference nếu có sample_records
    samples = result.get("sample_records") or []
    if samples:
        all_fields = sorted({k for r in _flatten_records_for_preview(samples)
                              for k in r.keys()})
        flat_samples = _flatten_records_for_preview(samples)
        fake_headers = all_fields
        fake_preview = [
            [r.get(field, None) for field in all_fields]
            for r in flat_samples
        ]
        result["field_types"] = infer_all_headers(fake_headers, fake_preview)
    else:
        result["field_types"] = {}

    return jsonify(result)


# ==========================================================================
# Task 17: overview theo `module | process | both` (rebuild bảng khi user
# đổi segmented control mà không cần full re-fetch dashboard).
# ==========================================================================

@app.route("/api/projects/<slug>/phase-matrix")
def project_phase_matrix(slug: str):
    """b9: Matrix Phase × (Module|Quy trình) theo global filter.

    Tránh double-compute khi user chỉ toggle mode: FE gọi endpoint này với
    group_by=process khi user chuyển; group_by=module dùng data cached từ
    /dashboard (không cần fetch lại). Endpoint áp dụng
    ``_filtered_data_from_request`` để tôn trọng module/process/pic global filter.
    """
    from analyzer.dashboard_engine import DashboardEngine
    state, err = _require_state(slug)
    if err:
        return err
    group_by = (request.args.get("group_by") or "module").lower()
    if group_by not in ("module", "process"):
        group_by = "module"
    data = _filtered_data_from_request(state)
    engine = DashboardEngine()
    return jsonify({
        "group_by": group_by,
        **engine._phase_status_matrix(data, group_by=group_by),
        "applied_filter": {
            "modules": _parse_multi_arg("module"),
            "processes": _parse_multi_arg("process"),
            "pics": _parse_multi_arg("pic"),
            "project_codes": _parse_project_code_args(),
        },
    })


@app.route("/api/projects/<slug>/module-overview")
def project_module_overview(slug: str):
    """
    BUG P0-B fix: endpoint TRƯỚC ĐÂY dùng `state["data"]` (raw, không filter).
    Khi user chọn group_by=process kèm global filter (VD chỉ 1 module), backend
    vẫn tính trên toàn bộ data → bảng hiển thị ALL processes, không apply filter.
    Fix: dùng `_filtered_data_from_request()` để tôn trọng module/process/pic
    truyền qua query params.
    """
    from analyzer.dashboard_engine import DashboardEngine
    state, err = _require_state(slug)
    if err:
        return err
    group_by = (request.args.get("group_by") or "module").lower()
    if group_by not in ("module", "process", "both"):
        group_by = "module"
    data = _filtered_data_from_request(state)
    # DashboardEngine.__init__(today=None, ...): pass no arg → dùng date.today().
    engine = DashboardEngine()
    return jsonify({
        "group_by": group_by,
        "rows": engine._overview_by(data, group_by=group_by),
        "applied_filter": {
            "modules": _parse_multi_arg("module"),
            "processes": _parse_multi_arg("process"),
            "pics": _parse_multi_arg("pic"),
            "project_codes": _parse_project_code_args(),
        },
    })


# ==========================================================================
# T24 — Bookmarks + Notes
# ==========================================================================

@app.route("/api/projects/<slug>/bookmarks", methods=["GET"])
def project_bookmarks_get(slug: str):
    from analyzer import project_store as ps
    state, err = _require_state(slug)
    if err:
        return err
    pdir = _project_dir_for(slug)
    bookmarks = ps.load_bookmarks(pdir)
    notes = ps.load_function_notes(pdir)
    # Enrich với thông tin function (ten_cn, module) để hiển thị nhanh
    by_ma = {}
    for row in state["data"].rows:
        mc = str(row.meta.get("ma_cn") or "").strip()
        if mc:
            by_ma[mc] = {
                "row_num": row.row_num,
                "ten_cn": str(row.meta.get("ten_cn") or ""),
                "module": str(row.meta.get("module") or ""),
                "quy_trinh": str(row.meta.get("quy_trinh") or ""),
            }
    items = []
    for mc in bookmarks:
        base = by_ma.get(mc, {"ten_cn": "(đã xóa/không tìm thấy)", "module": ""})
        items.append({
            "ma_cn": mc,
            "note": (notes.get(mc) or {}).get("note", ""),
            "note_updated_at": (notes.get(mc) or {}).get("updated_at", ""),
            **base,
        })
    return jsonify({"items": items, "count": len(items)})


@app.route("/api/projects/<slug>/bookmarks/toggle", methods=["POST"])
def project_bookmark_toggle(slug: str):
    from analyzer import project_store as ps
    state, err = _require_state(slug)
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    ma_cn = str(payload.get("ma_cn") or "").strip()
    if not ma_cn:
        return jsonify({"error": "ma_cn required"}), 400
    is_now, all_bm = ps.toggle_bookmark(_project_dir_for(slug), ma_cn)
    return jsonify({"bookmarked": is_now, "bookmarks": all_bm})


@app.route("/api/projects/<slug>/tags", methods=["GET"])
def project_tags_get(slug: str):
    """Danh sách tag theo Mã CN (đã review / escalate / chờ khách / CR / UAT issue)."""
    from analyzer import project_store as ps
    _, err = _require_state(slug)
    if err:
        return err
    tags = ps.load_function_tags(_project_dir_for(slug))
    return jsonify({
        "tags": tags,
        "valid_tags": list(ps.VALID_FUNCTION_TAGS),
        "count": len(tags),
    })


@app.route("/api/projects/<slug>/tags/bulk", methods=["POST"])
def project_tags_bulk(slug: str):
    """
    Bulk tag nhiều function trong drill-down.
    Body: { ma_cns: [...], tag: "đã review"|"escalate"|"chờ khách"|"CR"|"UAT issue", action: "add"|"remove"|"toggle" }
    """
    from analyzer import project_store as ps
    _, err = _require_state(slug)
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    ma_cns = payload.get("ma_cns") or []
    if not isinstance(ma_cns, list) or not ma_cns:
        return jsonify({"error": "ma_cns required (list)"}), 400
    tag = str(payload.get("tag") or "").strip()
    if not tag:
        return jsonify({"error": "tag required"}), 400
    action = str(payload.get("action") or "add").strip().lower()
    if action not in ("add", "remove", "toggle"):
        action = "add"
    tags = ps.bulk_tag_functions(
        _project_dir_for(slug), ma_cns, tag, action=action,
    )
    return jsonify({
        "ok": True,
        "tag": tag,
        "action": action,
        "affected": len([m for m in ma_cns if str(m).strip()]),
        "tags": tags,
    })


@app.route("/api/projects/<slug>/pic-upcoming")
def project_pic_upcoming(slug: str):
    """PIC × upcoming weeks — task đến hạn theo tuần tới."""
    from analyzer.pic_upcoming import compute_pic_upcoming_weeks
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    weeks = request.args.get("weeks", default=4, type=int)
    payload = compute_pic_upcoming_weeks(data, weeks=weeks)
    return jsonify(payload)


@app.route("/api/projects/<slug>/fl-reimport-verify")
def project_fl_reimport_verify(slug: str):
    """
    Sau re-upload yellow-cell export: so ô vàng snapshot trước vs hiện tại.
    Query: vs=previous (default) hoặc YYYY-MM-DD.
    """
    from analyzer.fl_reimport_verify import verify_fl_reimport
    from analyzer.data_quality import compute_data_quality
    from exporter.fl_reimport_export import collect_issue_hits

    state, err = _require_state(slug)
    if err:
        return err
    vs = request.args.get("vs", "previous").strip() or "previous"
    prev_data, prev_meta, err2 = _resolve_diff_previous_snapshot(slug, vs)
    if err2:
        return err2

    # Issue hits từ snapshot TRƯỚC (baseline yellow cells)
    try:
        prev_engine = DashboardEngine()
        prev_metrics = prev_engine.compute_all(prev_data)
        prev_dq = compute_data_quality(prev_data, today=prev_engine.today)
        hits_lists = {
            "overdue_list": (prev_metrics.get("overdue_list") or [])[:],
            "unassigned_list": (prev_metrics.get("unassigned_tasks") or [])[:],
            "stalled_list": list((prev_metrics.get("stalled_tasks") or {}).get("items") or []),
            "anomaly_issues": list(prev_dq.get("issues") or []),
        }
    except Exception as e:
        return jsonify({"error": f"Không tính issue hits snapshot trước: {e}"}), 500

    report = verify_fl_reimport(
        prev_data,
        state["data"],
        overdue_list=hits_lists["overdue_list"],
        unassigned_list=hits_lists["unassigned_list"],
        stalled_list=hits_lists["stalled_list"],
        anomaly_issues=hits_lists["anomaly_issues"],
    )
    report["previous_snapshot"] = prev_meta or {}
    report["current_filename"] = state.get("filename")
    return jsonify(report)


@app.route("/api/projects/<slug>/notes/<path:ma_cn>", methods=["GET"])
def project_note_get(slug: str, ma_cn: str):
    from analyzer import project_store as ps
    _, err = _require_state(slug)
    if err:
        return err
    notes = ps.load_function_notes(_project_dir_for(slug))
    n = notes.get(ma_cn) or {}
    return jsonify({"ma_cn": ma_cn, "note": n.get("note", ""), "updated_at": n.get("updated_at", "")})


@app.route("/api/projects/<slug>/notes/<path:ma_cn>", methods=["PUT", "DELETE"])
def project_note_save(slug: str, ma_cn: str):
    from analyzer import project_store as ps
    _, err = _require_state(slug)
    if err:
        return err
    if request.method == "DELETE":
        ps.save_function_note(_project_dir_for(slug), ma_cn, "")
        return jsonify({"ma_cn": ma_cn, "deleted": True})
    payload = request.get_json(silent=True) or {}
    note = str(payload.get("note") or "").strip()
    ps.save_function_note(_project_dir_for(slug), ma_cn, note)
    return jsonify({"ma_cn": ma_cn, "note": note, "saved": True})


# ==========================================================================
# T22 — Aging WIP tracking
# ==========================================================================

@app.route("/api/projects/<slug>/aging-wip")
def project_aging_wip(slug: str):
    """
    Trả danh sách phase In-progress vượt threshold ngày.
    Query params:
      threshold: int (mặc định 14)
      + module/process/pic để filter (global filter).
    """
    from analyzer.advanced_metrics import compute_aging_wip
    state, err = _require_state(slug)
    if err:
        return err
    try:
        threshold = int(request.args.get("threshold") or 14)
        threshold = max(1, min(365, threshold))  # clamp 1-365
    except ValueError:
        threshold = 14
    data = _filtered_data_from_request(state)
    return jsonify(compute_aging_wip(data, threshold_days=threshold))


@app.route("/api/projects/<slug>/export-aging-wip")
def project_export_aging_wip(slug: str):
    from analyzer.advanced_metrics import compute_aging_wip
    from exporter.excel_exporter import export_aging_wip_report
    state, err = _require_state(slug)
    if err:
        return err
    try:
        threshold = int(request.args.get("threshold") or 14)
        threshold = max(1, min(365, threshold))
    except ValueError:
        threshold = 14
    data = _filtered_data_from_request(state)
    payload = compute_aging_wip(data, threshold_days=threshold)
    project = _project_mgr.get_project(slug)
    subtitle = (
        f"Project: {project.name if project else slug} | "
        f"Ngưỡng: {threshold} ngày | Ngày: {date.today().strftime('%d/%m/%Y')}"
    )
    filepath = export_aging_wip_report(payload, output_dir="uploads", subtitle=subtitle)
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


# ==========================================================================
# Gantt Calendar — Excel-style timeline (Month/Week/Day header 3 tầng)
# ==========================================================================

def _parse_gantt_params():
    """Parse group_by + granularity từ query string (default: module/week)."""
    group_by = (request.args.get("group_by") or "module").strip().lower()
    granularity = (request.args.get("granularity") or "auto").strip().lower()
    return group_by, granularity


@app.route("/api/projects/<slug>/gantt-calendar")
def project_gantt_calendar(slug: str):
    """
    Gantt Calendar — dashboard timeline Excel-style.

    Query params:
      group_by: "module" | "phan_he" | "process" | "quy_trinh" | "function"
      granularity: "day" | "week" | "month" | "auto"
      + module/process/pic để filter (global filter — reuse contract chung).
    """
    from analyzer.gantt_calendar import compute_gantt_calendar
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    group_by, granularity = _parse_gantt_params()
    payload = compute_gantt_calendar(data, group_by=group_by, granularity=granularity)
    return jsonify(payload)


@app.route("/api/projects/<slug>/export-gantt-calendar")
def project_export_gantt_calendar(slug: str):
    """Xuất Gantt Calendar sang Excel (merge cell Month/Week + fill màu category)."""
    from analyzer.gantt_calendar import compute_gantt_calendar
    from exporter.excel_exporter import export_gantt_calendar_report
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    group_by, granularity = _parse_gantt_params()
    payload = compute_gantt_calendar(data, group_by=group_by, granularity=granularity)
    project = _project_mgr.get_project(slug)
    subtitle = (
        f"Project: {project.name if project else slug} | "
        f"Group={group_by} | Granularity={payload.get('granularity','')} | "
        f"Ngày: {date.today().strftime('%d/%m/%Y')}"
    )
    filepath = export_gantt_calendar_report(
        payload, output_dir=_project_mgr.get_export_dir(slug), subtitle=subtitle,
    )
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


# ==========================================================================
# T21 — Data Quality Panel
# ==========================================================================

@app.route("/api/projects/<slug>/data-quality")
def project_data_quality(slug: str):
    """
    Trả về data quality issues (list + summary).
    Hỗ trợ global filter module/process/pic để user zoom vào 1 subset.
    Kèm ownership (PIC + target date) + SLA / resolution rate WoW.
    """
    from analyzer import project_store as ps
    from analyzer.data_quality import compute_data_quality
    from analyzer.dq_ownership import attach_ownership, compute_dq_sla_stats

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    result = compute_data_quality(data)
    folder = _project_dir_for(slug)
    ownership = ps.load_dq_ownership(folder)

    # prior open count từ snapshot trước (nếu có)
    prior_open = None
    try:
        snaps = _project_mgr.get_snapshot_manager(slug).list_snapshots() or []
        if len(snaps) >= 2:
            prev = snaps[1]
            # Dùng tổng issue approximation từ metrics nếu có; else load parse
            loaded = _project_mgr.get_snapshot_manager(slug).load_snapshot(prev["date"])
            if loaded and loaded.get("parsed"):
                prior_dq = compute_data_quality(loaded["parsed"])
                prior_open = int((prior_dq.get("summary") or {}).get("total_issues") or 0)
    except Exception:
        prior_open = None

    result["issues"] = attach_ownership(result.get("issues") or [], ownership)
    result["ownership_stats"] = compute_dq_sla_stats(
        result["issues"], ownership, prior_open_count=prior_open,
    )
    result["ownership"] = ownership
    return jsonify(result)


@app.route("/api/projects/<slug>/dq-ownership", methods=["GET", "POST"])
def project_dq_ownership(slug: str):
    """GET list / POST upsert ownership cho 1 DQ issue key."""
    from analyzer import project_store as ps

    _, err = _require_state(slug)
    if err:
        return err
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"items": ps.load_dq_ownership(folder)})
    payload = request.get_json(silent=True) or {}
    key = str(payload.get("key") or "").strip()
    if not key:
        return jsonify({"error": "key required (ma_cn|phase|code)"}), 400
    user = _auth_current_user() or {}
    items = ps.save_dq_ownership(
        folder,
        key,
        owner_pic=str(payload.get("owner_pic") or ""),
        target_date=str(payload.get("target_date") or ""),
        assigned_by=str(user.get("username") or ""),
        note=str(payload.get("note") or ""),
        resolved=payload.get("resolved"),
        delete=bool(payload.get("delete")),
    )
    return jsonify({"ok": True, "items": items})


@app.route("/api/projects/<slug>/export-data-quality")
def project_export_data_quality(slug: str):
    """Xuất báo cáo Data Quality ra Excel."""
    from analyzer.data_quality import compute_data_quality
    from exporter.excel_exporter import export_data_quality_report
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    payload = compute_data_quality(data)
    project = _project_mgr.get_project(slug)
    subtitle = f"Project: {project.name if project else slug} | Ngày: {date.today().strftime('%d/%m/%Y')}"
    filepath = export_data_quality_report(payload, output_dir="uploads", subtitle=subtitle)
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


@app.route("/api/projects/<slug>/export-rlog-weekly", methods=["GET", "POST"])
def project_export_rlog_weekly(slug: str):
    """
    Xuất Excel Rlog coded tuần này + kế hoạch tuần tới.

    Tôn trọng global filter (module / process / pic / g_project) giống section
    đang xem. Payload reuse analyzer.rlog_weekly.compute_rlog_weekly.
    """
    from analyzer.rlog_weekly import compute_rlog_weekly
    from exporter.rlog_exporter import export_rlog_weekly_report

    state, err = _require_state(slug)
    if err:
        return err

    # POST body có thể gửi filter; GET dùng query params.
    if request.method == "POST":
        body = request.get_json(silent=True) or {}

        def _as_list(val) -> list[str]:
            if not val:
                return []
            if isinstance(val, list):
                out: list[str] = []
                for it in val:
                    out.extend(_as_list(it))
                return out
            return [x.strip() for x in str(val).split(",") if x.strip()]

        modules = _as_list(
            body.get("module") or body.get("g_module") or body.get("g_modules")
        )
        processes = _as_list(
            body.get("process") or body.get("g_process") or body.get("g_processes")
        )
        pics = _as_list(body.get("pic") or body.get("g_pic") or body.get("g_pics"))
        project_codes = _project_codes_from_body(body)
        if modules or processes or pics or project_codes:
            data = _filter_parsed_data(
                state["data"],
                modules=modules,
                processes=processes,
                pics=pics,
                project_codes=project_codes,
            )
        else:
            data = state["data"]
    else:
        data = _filtered_data_from_request(state)

    try:
        # Đồng bộ today với DashboardEngine (metrics.rlog_weekly trên UI)
        payload = compute_rlog_weekly(data, today=DashboardEngine().today)
        project = _project_mgr.get_project(slug)
        subtitle = (
            f"Project: {project.name if project else slug} | "
            f"Ngày: {date.today().strftime('%d/%m/%Y')}"
        )
        filepath = export_rlog_weekly_report(
            payload,
            output_dir=_project_mgr.get_export_dir(slug),
            subtitle=subtitle,
        )
        return send_file(
            filepath, as_attachment=True, download_name=os.path.basename(filepath)
        )
    except Exception as e:
        return jsonify({"error": f"Lỗi khi xuất Rlog weekly: {str(e)}"}), 500


# ==========================================================================
# Kanban (Task 10 — Kanban theo tuần)
# ==========================================================================

@app.route("/api/projects/<slug>/kanban")
def project_kanban(slug: str):
    """
    Query params:
      week_offset: int (default 0) — 0=tuần này, 1=tuần sau, -1=tuần trước
      module, process, pic, role: multi (comma hoặc repeat)
      search: text
    """
    from analyzer import project_store as ps
    from analyzer.kanban import compute_kanban
    state, err = _require_state(slug)
    if err:
        return err
    try:
        offset = int(request.args.get("week_offset") or 0)
    except ValueError:
        offset = 0
    filters = {
        "modules": _parse_multi_arg("module"),
        "processes": _parse_multi_arg("process"),
        "pics": _parse_multi_arg("pic"),
        "roles": _parse_multi_arg("role"),
        "search": request.args.get("search") or "",
    }
    # Task 14: role_map auto-detect từ phase (không đọc pic_role_map cũ nữa).
    from analyzer.kanban import infer_pic_roles
    role_map = infer_pic_roles(state["data"])
    return jsonify(compute_kanban(
        state["data"],
        week_offset=offset,
        pic_role_map=role_map,
        filters=filters,
    ))


@app.route("/api/projects/<slug>/pic-roles", methods=["GET", "POST"])
def project_pic_roles(slug: str):
    """Task 14: role auto-detect từ phase — GET vẫn trả map derived, POST no-op.

    - GET → `{map, all_pics, all_roles: ["BA","Dev"]}` — map derived,
      không đọc từ store.
    - POST → return 200 với warning "Role auto-detected, manual map disabled".
    """
    from analyzer.kanban import unique_pics, infer_pic_roles
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    state = _get_state(slug)
    all_pics = unique_pics(state["data"]) if state and state.get("data") else []
    derived_map = infer_pic_roles(state["data"]) if state and state.get("data") else {}
    if request.method == "GET":
        return jsonify({
            "map": derived_map,
            "all_pics": all_pics,
            "all_roles": ["BA", "Dev"],
            "auto_detected": True,
        })
    # POST — no-op (backward-compat). Client cũ có thể vẫn call save.
    app.logger.warning("[pic-roles] POST ignored: role auto-detected from phase (Task 14).")
    return jsonify({
        "map": derived_map,
        "all_pics": all_pics,
        "all_roles": ["BA", "Dev"],
        "auto_detected": True,
        "warning": "Role auto-detected từ phase, manual map đã bị vô hiệu.",
    })


@app.route("/api/projects/<slug>/chart-aggregate", methods=["POST"])
def project_chart_aggregate(slug: str):
    """
    Task 8/9: aggregate data để render chart tuỳ chỉnh.
    Body:
      {
        "x_field": "module",
        "y_measure": "count",
        "series_field": "status" | null,
        "filters": {"modules":[...], "processes":[...], ...},
        "apply_global_filter": true  # nếu true → merge với globalFilters từ query
      }
    """
    from analyzer.generic_chart import aggregate_chart
    state, err = _require_state(slug)
    if err:
        return err
    body = request.get_json(silent=True) or {}
    data = state["data"]
    filters = body.get("filters") or {}
    # Merge global filter từ query (nếu FE muốn)
    if body.get("apply_global_filter"):
        gmods = _parse_multi_arg("module")
        gprocs = _parse_multi_arg("process")
        gpics = _parse_multi_arg("pic")
        if gmods:
            filters["modules"] = list(set(filters.get("modules", []) + gmods)) or gmods
        if gprocs:
            filters["processes"] = list(set(filters.get("processes", []) + gprocs)) or gprocs
        if gpics:
            filters["pics"] = list(set(filters.get("pics", []) + gpics)) or gpics
    try:
        result = aggregate_chart(
            data,
            x_field=body.get("x_field") or "module",
            y_measure=body.get("y_measure") or "count",
            series_field=body.get("series_field") or None,
            filters=filters,
            limit_x=int(body.get("limit_x") or 50),
        )
    except Exception as e:
        return jsonify({"error": f"Aggregate failed: {e}"}), 400
    return jsonify(result)


@app.route("/api/projects/<slug>/upload-history")
def project_upload_history(slug: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    items = ps.load_upload_history(_project_dir_for(slug))
    return jsonify({
        "items": items,
        "max_entries": ps.MAX_UPLOAD_HISTORY,
    })


@app.route("/api/projects/<slug>/settings", methods=["GET", "PUT", "POST"])
def project_settings(slug: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify(ps.load_project_settings(folder))
    body = request.get_json(silent=True) or {}
    return jsonify(ps.save_project_settings(folder, body))


# --------------------------------------------------------------------------
# T-AA — Snapshot archive settings / run / restore
# --------------------------------------------------------------------------

@app.route("/api/projects/<slug>/archive-settings", methods=["GET", "PUT"])
def project_archive_settings(slug: str):
    """GET/PUT cấu hình auto-archive per-project."""
    from analyzer import project_store as ps
    from analyzer import archive_manager as am
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        settings = ps.load_archive_settings(folder)
        smgr = _project_mgr.get_snapshot_manager(slug)
        snaps = smgr.list_snapshots()
        usage = am.snapshot_disk_usage(smgr.dir)
        return jsonify({
            "settings": settings,
            "snapshots": snaps,
            "archived_count": sum(1 for s in snaps if s.get("archived")),
            "hot_count": sum(1 for s in snaps if not s.get("archived")),
            "disk": usage,
        })
    body = request.get_json(silent=True) or {}
    return jsonify({"settings": ps.save_archive_settings(folder, body)})


@app.route("/api/projects/<slug>/archive-run", methods=["POST"])
def project_archive_run(slug: str):
    """
    Manual trigger archive + optional purge.
    Body: { days?: int, purge_days?: int, purge?: bool }
    """
    from analyzer import project_store as ps
    from analyzer import archive_manager as am
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    settings = ps.load_archive_settings(folder)
    body = request.get_json(silent=True) or {}
    days = body.get("days")
    if days is None:
        days = settings.get("archive_after_days", 90)
    try:
        days = int(days)
    except (TypeError, ValueError):
        days = 90

    smgr = _project_mgr.get_snapshot_manager(slug)
    archived = am.auto_archive_project(smgr.dir, days=days)
    purged: list[str] = []
    do_purge = bool(body.get("purge"))
    purge_days = body.get("purge_days")
    if purge_days is None:
        purge_days = settings.get("purge_after_days", 365)
    try:
        purge_days = int(purge_days)
    except (TypeError, ValueError):
        purge_days = 365
    if do_purge and purge_days > 0:
        purged = am.purge_archive(smgr.dir, days=purge_days)

    return jsonify({
        "success": True,
        "archived": [e.get("date") for e in archived],
        "archived_count": len(archived),
        "purged": purged,
        "purged_count": len(purged),
        "disk": am.snapshot_disk_usage(smgr.dir),
    })


@app.route("/api/projects/<slug>/snapshots/<snap_id>/restore", methods=["POST"])
def project_snapshot_restore(slug: str, snap_id: str):
    """Rã đông 1 snapshot đã archive."""
    from analyzer import archive_manager as am
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    smgr = _project_mgr.get_snapshot_manager(slug)
    try:
        entry = am.restore_snapshot(smgr.dir, snap_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Restore lỗi: {e}"}), 500
    return jsonify({"success": True, "entry": entry})


@app.route("/api/projects/<slug>/snapshots/<snap_id>/archive", methods=["POST"])
def project_snapshot_archive_one(slug: str, snap_id: str):
    """Archive thủ công 1 snapshot."""
    from analyzer import archive_manager as am
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    smgr = _project_mgr.get_snapshot_manager(slug)
    try:
        entry = am.archive_snapshot(smgr.dir, snap_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Archive lỗi: {e}"}), 500
    return jsonify({"success": True, "entry": entry})


# --------------------------------------------------------------------------
# T26 — Weekly Digest endpoints
# --------------------------------------------------------------------------

@app.route("/api/projects/<slug>/digests", methods=["GET", "POST"])
def project_digests(slug: str):
    """GET: list history; POST: generate ngay (manual trigger)."""
    from analyzer import digest as digest_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"items": digest_mod.list_digests(folder)})
    # POST — generate ngay
    state = _get_state(slug)
    if not state or not state.get("metrics"):
        return jsonify({"error": "Project chưa có dữ liệu — upload trước"}), 400
    path = digest_mod.generate_digest_now(folder, state["metrics"])
    if not path:
        return jsonify({"error": "Không sinh được digest"}), 500
    return jsonify({
        "filename": os.path.basename(path),
        "items": digest_mod.list_digests(folder),
    })


@app.route("/api/projects/<slug>/digests/<path:filename>", methods=["GET", "DELETE"])
def project_digest_file(slug: str, filename: str):
    """GET: download 1 file; DELETE: xoá."""
    from analyzer import digest as digest_mod
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    path = digest_mod.get_digest_path(folder, filename)
    if not path:
        return jsonify({"error": "File digest không tồn tại"}), 404
    if request.method == "DELETE":
        try:
            os.remove(path)
        except OSError as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"items": digest_mod.list_digests(folder)})
    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/projects/<slug>/phase-aliases", methods=["GET", "PUT"])
def project_phase_aliases(slug: str):
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"aliases": ps.load_phase_aliases(folder)})
    body = request.get_json(silent=True) or {}
    aliases = body.get("aliases", body)
    return jsonify({"aliases": ps.save_phase_aliases(folder, aliases)})


# --------------------------------------------------------------------------
# P4/P5 analytics endpoints — mỗi endpoint mỏng, chỉ pipe engine → JSON.
# Tất cả nhận query params: module, process, pic (comma-separated) để filter cascade.
# --------------------------------------------------------------------------

def _require_state(slug: str):
    """Helper — trả (state, err_response). None state nghĩa là chưa upload."""
    if not _project_mgr.project_exists(slug):
        return None, (jsonify({"error": "Project không tồn tại"}), 404)
    state = _get_state(slug)
    if not state or not state.get("data"):
        return None, (jsonify({"error": "Chưa upload file"}), 404)
    return state, None


def _filtered_data_from_request(state):
    """Đọc query params (module/process/pic/g_project) → ParsedData đã filter."""
    fmodules = _parse_multi_arg("module")
    fprocesses = _parse_multi_arg("process")
    fpics = _parse_multi_arg("pic")
    fproject_codes = _parse_project_code_args()
    if not (fmodules or fprocesses or fpics or fproject_codes):
        return state["data"]
    return _filter_parsed_data(
        state["data"],
        modules=fmodules,
        processes=fprocesses,
        pics=fpics,
        project_codes=fproject_codes,
    )


@app.route("/api/projects/<slug>/capacity-load")
def project_capacity_load(slug: str):
    """Remaining MH vs capacity — hỗ trợ filter module/process/pic."""
    from analyzer import project_store as ps
    from analyzer.advanced_metrics import compute_capacity_load
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    cap = ps.load_capacity(_project_dir_for(slug))
    return jsonify(compute_capacity_load(data, cap))


@app.route("/api/projects/<slug>/burndown")
def project_burndown(slug: str):
    """b12: hỗ trợ query param `phase` để giới hạn scope theo 1 phase cụ thể.

    Không có param → đếm mọi phase Closed (behavior cũ). Có param → chỉ
    đếm phase name matched.
    """
    from analyzer.advanced_metrics import compute_burndown_velocity
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    phase = (request.args.get("phase") or "").strip()
    return jsonify(compute_burndown_velocity(data, phase=phase or None))


@app.route("/api/projects/<slug>/sla")
def project_sla(slug: str):
    from analyzer import project_store as ps
    from analyzer.advanced_metrics import compute_sla_violations
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    settings = ps.load_project_settings(_project_dir_for(slug))
    sla = settings.get("sla") or {}
    return jsonify(compute_sla_violations(
        data,
        must_have_days=int(sla.get("must_have_days", 3)),
        should_have_days=int(sla.get("should_have_days", 7)),
    ))


@app.route("/api/projects/<slug>/slow-heatmap")
def project_slow_heatmap(slug: str):
    from analyzer.advanced_metrics import compute_slow_heatmap
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    return jsonify(compute_slow_heatmap(data))


@app.route("/api/projects/<slug>/dependency-blockers")
def project_dependency_blockers(slug: str):
    from analyzer.advanced_metrics import compute_dependency_blockers
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    return jsonify(compute_dependency_blockers(data))


@app.route("/api/projects/<slug>/baseline-variance")
def project_baseline_variance(slug: str):
    from analyzer.advanced_metrics import compute_baseline_variance
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    return jsonify(compute_baseline_variance(data))


# --------------------------------------------------------------------------
# Phase A — Baseline snapshot (kế hoạch gốc) + SV + Predictive completion
# --------------------------------------------------------------------------

@app.route("/api/projects/<slug>/baseline", methods=["GET", "PUT", "POST", "DELETE"])
def project_baseline(slug: str):
    """
    Đánh dấu / đọc / xóa snapshot baseline của project.

    GET  → {baseline_snapshot_id, meta?, snapshots[]}
    PUT/POST body: {baseline_snapshot_id: "YYYY-MM-DD"|null|""}
         null/"" → clear baseline
    DELETE → clear baseline
    """
    from analyzer import project_store as ps
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    folder = _project_dir_for(slug)
    smgr = _project_mgr.get_snapshot_manager(slug)
    snaps = smgr.list_snapshots()

    if request.method == "GET":
        settings = ps.load_project_settings(folder)
        snap_id = (settings.get("baseline_snapshot_id") or "").strip()
        meta = next((s for s in snaps if s.get("date") == snap_id), None)
        return jsonify({
            "baseline_snapshot_id": snap_id or None,
            "meta": meta,
            "snapshots": snaps,
        })

    if request.method == "DELETE":
        settings = ps.save_project_settings(folder, {"baseline_snapshot_id": ""})
        return jsonify({
            "baseline_snapshot_id": None,
            "settings": settings,
            "message": "Đã xóa baseline.",
        })

    body = request.get_json(silent=True) or {}
    raw = body.get("baseline_snapshot_id", "")
    if raw is None or str(raw).strip() == "":
        settings = ps.save_project_settings(folder, {"baseline_snapshot_id": ""})
        return jsonify({
            "baseline_snapshot_id": None,
            "settings": settings,
            "message": "Đã xóa baseline.",
        })

    snap_id = str(raw).strip()
    # Validate snapshot tồn tại
    if not any(s.get("date") == snap_id for s in snaps):
        return jsonify({
            "error": f"Snapshot '{snap_id}' không tồn tại trong project này.",
        }), 400
    settings = ps.save_project_settings(folder, {"baseline_snapshot_id": snap_id})
    meta = next((s for s in snaps if s.get("date") == snap_id), None)
    return jsonify({
        "baseline_snapshot_id": snap_id,
        "meta": meta,
        "settings": settings,
        "message": f"Đã đặt baseline = {snap_id}.",
    })


@app.route("/api/projects/<slug>/baseline-sv")
def project_baseline_sv(slug: str):
    """
    Schedule Variance vs snapshot baseline (Phase A).

    SV = end_hiện_tại − end_baseline (ngày). Cần đã đánh dấu baseline.
    """
    from analyzer import project_store as ps
    from analyzer.baseline_sv import compute_baseline_sv

    state, err = _require_state(slug)
    if err:
        return err
    folder = _project_dir_for(slug)
    settings = ps.load_project_settings(folder)
    snap_id = (settings.get("baseline_snapshot_id") or "").strip()
    if not snap_id:
        return jsonify({
            "error": "Chưa đánh dấu baseline. Chọn 1 snapshot làm kế hoạch gốc.",
            "baseline_snapshot_id": None,
            "summary": None,
        }), 400

    smgr = _project_mgr.get_snapshot_manager(slug)
    loaded = smgr.load_snapshot(snap_id)
    if not loaded or not loaded.get("parsed"):
        return jsonify({
            "error": f"Không load được snapshot baseline '{snap_id}'.",
            "baseline_snapshot_id": snap_id,
        }), 404

    data = _filtered_data_from_request(state)
    top_raw = request.args.get("top")
    top: int | None = 200
    if top_raw is not None:
        try:
            top = int(top_raw)
            if top <= 0:
                top = None
        except (TypeError, ValueError):
            top = 200

    result = compute_baseline_sv(
        data,
        loaded["parsed"],
        baseline_snapshot_id=snap_id,
        top_functions=top,
    )
    return jsonify(result)


@app.route("/api/projects/<slug>/completion-forecast")
def project_completion_forecast(slug: str):
    """
    Dự báo ngày xong từ velocity 4 tuần (remaining ÷ Closed/tuần).
    Query: phase= (optional — scope burndown giống /burndown).
    """
    from analyzer.completion_forecast import compute_completion_forecast

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    phase = (request.args.get("phase") or "").strip()
    return jsonify(compute_completion_forecast(data, phase=phase or None))


def _load_baseline_for_project(slug: str):
    """Trả (baseline ParsedData|None, snap_id str)."""
    from analyzer import project_store as ps
    folder = _project_dir_for(slug)
    settings = ps.load_project_settings(folder)
    snap_id = (settings.get("baseline_snapshot_id") or "").strip()
    baseline = None
    if snap_id:
        loaded = _project_mgr.get_snapshot_manager(slug).load_snapshot(snap_id)
        if loaded and loaded.get("parsed"):
            baseline = loaded["parsed"]
    return baseline, snap_id


def _snapshot_series_for_project(slug: str, current_data=None):
    """
    List (date, ParsedData) tăng dần từ snapshot history + current.
    """
    from datetime import date as _date
    smgr = _project_mgr.get_snapshot_manager(slug)
    series = []
    for entry in reversed(smgr.list_snapshots() or []):  # list desc → reverse asc
        d_str = entry.get("date") or ""
        try:
            as_of = _date.fromisoformat(d_str[:10])
        except ValueError:
            continue
        loaded = smgr.load_snapshot(d_str)
        if loaded and loaded.get("parsed"):
            series.append((as_of, loaded["parsed"]))
    # Đảm bảo điểm hiện tại (hôm nay) có mặt
    if current_data is not None:
        today = _date.today()
        if not series or series[-1][0] != today:
            series.append((today, current_data))
        else:
            series[-1] = (today, current_data)
    return series


@app.route("/api/projects/<slug>/earned-value")
def project_earned_value(slug: str):
    """
    Earned Value (Phase B) — EV / PV / AC → SPI / CPI.

    PV cần baseline snapshot; không có baseline vẫn trả EV/AC/CPI với SPI=null.
    """
    from analyzer.earned_value import compute_earned_value

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    baseline, snap_id = _load_baseline_for_project(slug)

    result = compute_earned_value(
        data,
        baseline=baseline,
        baseline_snapshot_id=snap_id or None,
        today=None,
    )
    if snap_id and baseline is None:
        result["messages"] = list(result.get("messages") or []) + [
            f"Không load được snapshot baseline '{snap_id}' — SPI tạm N/A."
        ]
        result["has_baseline"] = False
        result["summary"]["pv"] = None
        result["summary"]["spi"] = None
        result["summary"]["spi_label"] = "N/A"
        result["summary"]["pv_pct_bac"] = None
    return jsonify(result)


@app.route("/api/projects/<slug>/earned-value-scurve")
def project_earned_value_scurve(slug: str):
    """S-curve EV/PV/AC theo snapshot (weekly)."""
    from analyzer.earned_value import compute_evm_scurve

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    baseline, snap_id = _load_baseline_for_project(slug)
    series = _snapshot_series_for_project(slug, current_data=data)
    weekly = (request.args.get("weekly", "1") or "1").strip() != "0"
    return jsonify(compute_evm_scurve(
        series,
        baseline=baseline,
        baseline_snapshot_id=snap_id or None,
        weekly=weekly,
    ))


@app.route("/api/projects/<slug>/scope-creep")
def project_scope_creep(slug: str):
    """
    Change Request / Scope Creep (Phase C).

    Primary: cột Excel CR / Phát sinh (auto-detect).
    Fallback: tag «CR» hoặc settings.cr_function_codes.
    """
    from analyzer import project_store as ps
    from analyzer.scope_creep import compute_scope_creep

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    settings = ps.load_project_settings(folder)
    tags = ps.load_function_tags(folder)
    result = compute_scope_creep(
        data,
        function_tags=tags,
        cr_function_codes=settings.get("cr_function_codes") or [],
    )
    return jsonify(result)


@app.route("/api/projects/<slug>/uat-quality")
def project_uat_quality(slug: str):
    """
    Phase E — UAT / Customer feedback quality.

    Auto-detect cột Defect/Bug/Feedback/Reopen/UAT cycle.
    Không có cột → empty metrics + optional tag «UAT issue» (không bịa số lỗi).
    """
    from analyzer import project_store as ps
    from analyzer.uat_quality import compute_uat_quality

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    tags = ps.load_function_tags(folder)
    return jsonify(compute_uat_quality(data, function_tags=tags))


# ==========================================================================
# PMO/BA analytics — Excel export (Tong_hop + Chi_tiet)
# ==========================================================================

def _export_mode_from_request() -> str:
    body = request.get_json(silent=True) or {}
    mode = (request.args.get("mode") or body.get("mode") or "both").strip().lower()
    return mode if mode in ("summary", "detail", "both") else "both"


@app.route("/api/projects/<slug>/export-baseline-sv", methods=["GET", "POST"])
def export_baseline_sv_api(slug: str):
    """Xuất Excel Schedule Variance vs snapshot baseline (ALL function×phase)."""
    from analyzer import project_store as ps
    from analyzer.baseline_sv import compute_baseline_sv
    from exporter.pmo_analytics_exporter import export_baseline_sv_report

    state, err = _require_state(slug)
    if err:
        return err
    folder = _project_dir_for(slug)
    settings = ps.load_project_settings(folder)
    snap_id = (settings.get("baseline_snapshot_id") or "").strip()
    if not snap_id:
        return jsonify({"error": "Chưa đánh dấu baseline. Chọn snapshot làm kế hoạch gốc."}), 400
    smgr = _project_mgr.get_snapshot_manager(slug)
    loaded = smgr.load_snapshot(snap_id)
    if not loaded or not loaded.get("parsed"):
        return jsonify({"error": f"Không load được snapshot baseline '{snap_id}'."}), 404
    data = _filtered_data_from_request(state)
    result = compute_baseline_sv(
        data, loaded["parsed"],
        baseline_snapshot_id=snap_id,
        top_functions=None,
    )
    try:
        filepath = export_baseline_sv_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Baseline SV: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-earned-value", methods=["GET", "POST"])
def export_earned_value_api(slug: str):
    """Xuất Excel Earned Value (SPI/CPI)."""
    from analyzer.earned_value import compute_earned_value
    from exporter.pmo_analytics_exporter import export_earned_value_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    baseline, snap_id = _load_baseline_for_project(slug)
    result = compute_earned_value(
        data, baseline=baseline, baseline_snapshot_id=snap_id or None, today=None
    )
    try:
        filepath = export_earned_value_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất EVM: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-earned-value-scurve", methods=["GET", "POST"])
def export_earned_value_scurve_api(slug: str):
    """Xuất Excel EVM S-curve."""
    from analyzer.earned_value import compute_evm_scurve
    from exporter.pmo_analytics_exporter import export_evm_scurve_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    baseline, snap_id = _load_baseline_for_project(slug)
    series = _snapshot_series_for_project(slug, current_data=data)
    result = compute_evm_scurve(
        series, baseline=baseline, baseline_snapshot_id=snap_id or None, weekly=True,
    )
    try:
        filepath = export_evm_scurve_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất S-curve: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-scope-creep", methods=["GET", "POST"])
def export_scope_creep_api(slug: str):
    """Xuất Excel Scope Creep (ALL CR)."""
    from analyzer import project_store as ps
    from analyzer.scope_creep import compute_scope_creep
    from exporter.pmo_analytics_exporter import export_scope_creep_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    settings = ps.load_project_settings(folder)
    tags = ps.load_function_tags(folder)
    result = compute_scope_creep(
        data,
        function_tags=tags,
        cr_function_codes=settings.get("cr_function_codes") or [],
        detail_limit=None,
    )
    try:
        filepath = export_scope_creep_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Scope Creep: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-uat-quality", methods=["GET", "POST"])
def export_uat_quality_api(slug: str):
    """Xuất Excel UAT Quality (ALL function có dữ liệu)."""
    from analyzer import project_store as ps
    from analyzer.uat_quality import compute_uat_quality
    from exporter.pmo_analytics_exporter import export_uat_quality_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    tags = ps.load_function_tags(folder)
    result = compute_uat_quality(data, function_tags=tags, detail_limit=None)
    try:
        filepath = export_uat_quality_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất UAT Quality: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-completion-forecast", methods=["GET", "POST"])
def export_completion_forecast_api(slug: str):
    """Xuất Excel dự báo ngày xong (velocity)."""
    from analyzer.completion_forecast import compute_completion_forecast
    from exporter.pmo_analytics_exporter import export_completion_forecast_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    phase = (request.args.get("phase") or "").strip()
    result = compute_completion_forecast(data, phase=phase or None)
    try:
        filepath = export_completion_forecast_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Completion Forecast: {str(e)}"}), 500


@app.route("/api/projects/<slug>/executive-dashboard")
def project_executive_dashboard(slug: str):
    """PM Executive Dashboard — 1 trang tổng hợp."""
    from analyzer import project_store as ps
    from analyzer.completion_forecast import compute_completion_forecast
    from analyzer.earned_value import compute_earned_value
    from analyzer.executive_dashboard import build_executive_dashboard
    from analyzer.risk_tracking import attach_mitigations
    from analyzer.scope_creep import compute_scope_creep

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    baseline, snap_id = _load_baseline_for_project(slug)
    tags = ps.load_function_tags(folder)
    settings = ps.load_project_settings(folder)
    mitigations = ps.load_risk_mitigations(folder)

    evm = compute_earned_value(
        data, baseline=baseline, baseline_snapshot_id=snap_id or None,
    )
    fc = compute_completion_forecast(data)
    sc = compute_scope_creep(
        data,
        function_tags=tags,
        cr_function_codes=settings.get("cr_function_codes") or [],
    )
    risk_scores = attach_mitigations(
        list((state.get("metrics") or {}).get("risk_scores") or []),
        mitigations,
    )
    project = _project_mgr.get_project(slug)
    payload = build_executive_dashboard(
        data=data,
        metrics=state.get("metrics") or {},
        earned_value=evm,
        completion_forecast=fc,
        scope_creep=sc,
        risk_scores=risk_scores,
        mitigations=mitigations,
        project_name=(project.name if project else slug),
    )
    return jsonify(payload)


@app.route("/api/projects/<slug>/export-executive-dashboard", methods=["GET", "POST"])
def export_executive_dashboard_api(slug: str):
    """Xuất Excel PM Executive Dashboard."""
    from analyzer import project_store as ps
    from analyzer.completion_forecast import compute_completion_forecast
    from analyzer.earned_value import compute_earned_value
    from analyzer.executive_dashboard import build_executive_dashboard
    from analyzer.risk_tracking import attach_mitigations
    from analyzer.scope_creep import compute_scope_creep
    from exporter.pmo_analytics_exporter import export_executive_dashboard_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    baseline, snap_id = _load_baseline_for_project(slug)
    tags = ps.load_function_tags(folder)
    settings = ps.load_project_settings(folder)
    mitigations = ps.load_risk_mitigations(folder)
    evm = compute_earned_value(
        data, baseline=baseline, baseline_snapshot_id=snap_id or None,
    )
    fc = compute_completion_forecast(data)
    sc = compute_scope_creep(
        data,
        function_tags=tags,
        cr_function_codes=settings.get("cr_function_codes") or [],
    )
    risk_scores = attach_mitigations(
        list((state.get("metrics") or {}).get("risk_scores") or []),
        mitigations,
    )
    project = _project_mgr.get_project(slug)
    payload = build_executive_dashboard(
        data=data,
        metrics=state.get("metrics") or {},
        earned_value=evm,
        completion_forecast=fc,
        scope_creep=sc,
        risk_scores=risk_scores,
        mitigations=mitigations,
        project_name=(project.name if project else slug),
    )
    try:
        filepath = export_executive_dashboard_report(
            payload, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Executive: {str(e)}"}), 500


@app.route("/api/projects/<slug>/risk-trend")
def project_risk_trend(slug: str):
    """Xu hướng risk score theo snapshot + danh sách mitigation."""
    from analyzer import project_store as ps
    from analyzer.risk_tracking import attach_mitigations, compute_risk_trend

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    series = _snapshot_series_for_project(slug, current_data=data)
    trend = compute_risk_trend(series, weekly=True)
    mitigations = ps.load_risk_mitigations(folder)
    scores = attach_mitigations(
        list((state.get("metrics") or {}).get("risk_scores") or [])[:50],
        mitigations,
    )
    trend["mitigations"] = mitigations
    trend["risk_scores"] = scores
    return jsonify(trend)


@app.route("/api/projects/<slug>/risk-mitigation", methods=["GET", "POST"])
def project_risk_mitigation(slug: str):
    """GET all / POST upsert mitigation (key = ma_cn hoặc module:X)."""
    from analyzer import project_store as ps

    _, err = _require_state(slug)
    if err:
        return err
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"items": ps.load_risk_mitigations(folder)})
    payload = request.get_json(silent=True) or {}
    key = str(payload.get("key") or payload.get("ma_cn") or "").strip()
    if not key:
        return jsonify({"error": "key / ma_cn required"}), 400
    user = _auth_current_user() or {}
    items = ps.save_risk_mitigation(
        folder,
        key,
        note=str(payload.get("note") or ""),
        owner=str(payload.get("owner") or ""),
        target_date=str(payload.get("target_date") or ""),
        updated_by=str(user.get("username") or ""),
        delete=bool(payload.get("delete")),
    )
    return jsonify({"ok": True, "items": items})


@app.route("/api/projects/<slug>/diff-review", methods=["GET", "POST"])
def project_diff_review(slug: str):
    """
    Diff approval nhẹ: gắn tag «đã review» + audit trail.
    POST body: { ma_cns: [...], reviewed: true|false, vs: "previous"|date }
    """
    from analyzer import project_store as ps

    _, err = _require_state(slug)
    if err:
        return err
    folder = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({
            "tags": ps.load_function_tags(folder),
            "reviews": ps.load_diff_reviews(folder),
            "review_tag": "đã review",
        })
    payload = request.get_json(silent=True) or {}
    ma_cns = payload.get("ma_cns") or []
    if isinstance(ma_cns, str):
        ma_cns = [ma_cns]
    if not isinstance(ma_cns, list) or not ma_cns:
        return jsonify({"error": "ma_cns required"}), 400
    reviewed = payload.get("reviewed", True)
    action = "review" if reviewed else "unreview"
    vs = str(payload.get("vs") or "")
    user = _auth_current_user() or {}
    username = str(user.get("username") or "")
    tag_action = "add" if reviewed else "remove"
    tags = ps.bulk_tag_functions(folder, ma_cns, "đã review", action=tag_action)
    reviews = ps.load_diff_reviews(folder)
    for ma in ma_cns:
        reviews = ps.append_diff_review(
            folder, str(ma), reviewed_by=username, vs=vs, action=action,
        )
    return jsonify({
        "ok": True,
        "tag": "đã review",
        "action": action,
        "affected": len([m for m in ma_cns if str(m).strip()]),
        "tags": tags,
        "reviews": reviews,
    })


@app.route("/api/projects/<slug>/insight-module-deltas")
def project_insight_module_deltas(slug: str):
    """Delta OD/UA/ST theo module vs snapshot trước."""
    from analyzer.insight_module_deltas import compute_module_issue_deltas

    state, err = _require_state(slug)
    if err:
        return err
    vs = (request.args.get("vs") or "previous").strip() or "previous"
    prev_data, prev_meta, err2 = _resolve_diff_previous_snapshot(slug, vs)
    if err2:
        # Không có snapshot trước → empty
        return jsonify({
            "modules": [],
            "totals": {},
            "message": "Chưa có snapshot trước để so sánh module delta.",
            "available": False,
        })
    result = compute_module_issue_deltas(prev_data, state["data"])
    result["previous_snapshot"] = prev_meta or {}
    result["available"] = True
    return jsonify(result)


@app.route("/api/projects/<slug>/export-pic-upcoming", methods=["GET", "POST"])
def export_pic_upcoming_api(slug: str):
    """Xuất Excel PIC × tuần tới."""
    from analyzer.pic_upcoming import compute_pic_upcoming_weeks
    from exporter.pmo_analytics_exporter import export_pic_upcoming_report

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    try:
        weeks = int(request.args.get("weeks") or 4)
    except (TypeError, ValueError):
        weeks = 4
    result = compute_pic_upcoming_weeks(data, weeks=weeks)
    try:
        filepath = export_pic_upcoming_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất PIC tuần tới: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-estimate-ratio", methods=["GET", "POST"])
def export_estimate_ratio_api(slug: str):
    """Xuất Excel ước lượng theo hệ số (parametric)."""
    from analyzer.estimate_ratio import compute_estimate_ratio, load_estimation_params
    from exporter.pmo_analytics_exporter import export_estimate_ratio_report

    state, err = _require_state(slug)
    if err:
        return err
    filt = _estimate_ratio_filters()
    data = _filter_parsed_data(
        state["data"],
        modules=filt["modules"],
        processes=filt["processes"],
        pics=filt["pics"],
    )
    params = load_estimation_params(_project_dir_for(slug), _project_mgr.base_dir)
    # Cho phép override params từ body (không lưu) — giống API compute
    body = request.get_json(silent=True) or {}
    if isinstance(body.get("params"), dict) and body["params"]:
        from analyzer.estimate_ratio import _deep_merge, normalize_params
        raw = {k: v for k, v in params.items() if not k.startswith("_")}
        params = normalize_params(_deep_merge(raw, body["params"]))
    result = compute_estimate_ratio(data, params)
    try:
        filepath = export_estimate_ratio_report(
            result, _project_mgr.get_export_dir(slug), mode=_export_mode_from_request()
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Estimate Ratio: {str(e)}"}), 500


@app.route("/api/projects/<slug>/pmo-risk")
def project_pmo_risk(slug: str):
    """
    Phase D — Risk + chiều Resource (PIC overload) + Dependency (cascade module).

    Query:
      cross_project=1 → lấy overload PIC từ quét đa dự án (pic_overload).
    """
    from analyzer import project_store as ps
    from analyzer.risk_scorer import compute_pmo_risk
    from analyzer.pic_overload import (
        compute_pic_overload,
        load_overload_settings,
        overloaded_pics_for_data,
    )

    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    folder = _project_dir_for(slug)
    module_order = ps.load_module_order(folder)
    thr = load_overload_settings(app.config["PROJECTS_FOLDER"])

    overloaded: set[str] | None = None
    overload_source = "single_project"
    cross = (request.args.get("cross_project") or "").strip().lower() in (
        "1", "true", "yes", "y",
    )
    if cross:
        try:
            result_ol = compute_pic_overload(
                _project_mgr,
                lambda s: _get_state(s),
                grain="day",
                thresholds=thr,
                today=date.today(),
                detail_limit=1,
            )
            overloaded = {
                p["pic"] for p in (result_ol.get("by_pic") or [])
                if p.get("is_overload")
            }
            overload_source = "cross_project"
        except Exception:
            overloaded = overloaded_pics_for_data(
                data, today=date.today(), thresholds=thr,
            )
            overload_source = "single_project_fallback"
    else:
        overloaded = overloaded_pics_for_data(
            data, today=date.today(), thresholds=thr,
        )

    result = compute_pmo_risk(
        data,
        today=date.today(),
        module_order=module_order or None,
        overloaded_pics=overloaded,
        overload_thresholds=thr,
    )
    result["overload_source"] = overload_source
    return jsonify(result)


# ==========================================================================
# Excel exports cho 4 section Phase 4/5 (Vấn đề 3 + Rule V4 "xuất ALL").
# Accept GET (dùng query string) và POST (body JSON) — cả 2 đều đọc global
# filter từ query string (?module=, ?process=, ?pic=) để tương thích form-submit
# cũ. FE mới nên POST JSON body {module: [], process: [], pic: []} → convert
# sang args-like dict để reuse _filtered_data_from_request contract.
# ==========================================================================


def _filter_state_from_body_or_args(state):
    """
    Đọc filter từ (a) POST body JSON hoặc (b) query args, apply _filter_parsed_data.

    Body JSON format: {"module": ["A","B"], "process": ["X"], "pic": ["Y"]}
    Query args: module=A,B&process=X&pic=Y (repeated hoặc comma-separated).
    Return (filtered_data, filters_dict_for_subtitle).
    """
    body = request.get_json(silent=True) or {} if request.method == "POST" else {}

    def _as_list(v) -> list[str]:
        if v is None:
            return []
        if isinstance(v, str):
            return [x.strip() for x in v.split(",") if x.strip()]
        if isinstance(v, list):
            out = []
            for it in v:
                out.extend(_as_list(it))
            return out
        return []

    # Ưu tiên body, fallback về query args
    fmods = _as_list(body.get("module") or body.get("modules")) or _parse_multi_arg("module")
    fprocs = _as_list(body.get("process") or body.get("processes")) or _parse_multi_arg("process")
    fpics = _as_list(body.get("pic") or body.get("pics")) or _parse_multi_arg("pic")
    fpcodes = _project_codes_from_body(body) or _parse_project_code_args()

    if not (fmods or fprocs or fpics or fpcodes):
        return state["data"], {
            "modules": [], "processes": [], "pics": [], "project_codes": [],
        }
    filtered = _filter_parsed_data(
        state["data"],
        modules=fmods,
        processes=fprocs,
        pics=fpics,
        project_codes=fpcodes,
    )
    return filtered, {
        "modules": fmods, "processes": fprocs, "pics": fpics, "project_codes": fpcodes,
    }


@app.route("/api/projects/<slug>/export-sla", methods=["GET", "POST"])
def export_sla(slug: str):
    """Xuất Excel SLA — vi phạm deadline theo Priority. Áp global filter."""
    from analyzer import project_store as ps
    from analyzer.advanced_metrics import compute_sla_violations
    state, err = _require_state(slug)
    if err:
        return err
    try:
        data, filters = _filter_state_from_body_or_args(state)
        settings = ps.load_project_settings(_project_dir_for(slug))
        sla_cfg = settings.get("sla") or {}
        payload = compute_sla_violations(
            data,
            must_have_days=int(sla_cfg.get("must_have_days", 3)),
            should_have_days=int(sla_cfg.get("should_have_days", 7)),
        )
        filepath = export_sla_report(
            payload=payload,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất SLA: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-capacity", methods=["GET", "POST"])
def export_capacity(slug: str):
    """Xuất Excel Capacity PIC — remaining MH vs công suất."""
    from analyzer import project_store as ps
    from analyzer.advanced_metrics import compute_capacity_load
    state, err = _require_state(slug)
    if err:
        return err
    try:
        data, filters = _filter_state_from_body_or_args(state)
        cap = ps.load_capacity(_project_dir_for(slug))
        payload = compute_capacity_load(data, cap)
        filepath = export_capacity_report(
            payload=payload,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Capacity: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-slow", methods=["GET", "POST"])
def export_slow(slug: str):
    """Xuất Excel Slow Heatmap — Ai đang chậm (PIC × Phase)."""
    from analyzer.advanced_metrics import compute_slow_heatmap
    state, err = _require_state(slug)
    if err:
        return err
    try:
        data, filters = _filter_state_from_body_or_args(state)
        payload = compute_slow_heatmap(data)
        filepath = export_slow_heatmap_report(
            payload=payload,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Slow Heatmap: {str(e)}"}), 500


@app.route("/api/projects/<slug>/export-baseline", methods=["GET", "POST"])
def export_baseline(slug: str):
    """
    Xuất Excel Baseline vs Actual — variance ngày.
    Rule V4: dùng top=None để compute trả ALL items, KHÔNG cắt 200 như API JSON.
    """
    from analyzer.advanced_metrics import compute_baseline_variance
    state, err = _require_state(slug)
    if err:
        return err
    try:
        data, filters = _filter_state_from_body_or_args(state)
        # top=None → không cắt top 200 (rule "xuất ALL record" cho Excel)
        full_payload = compute_baseline_variance(data, top=None)
        filepath = export_baseline_variance_report(
            payload=full_payload,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters,
            all_items=full_payload.get("items"),
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Baseline: {str(e)}"}), 500


# ==========================================================================
# Function Traceability (BA — Task 1)
# ==========================================================================
# BA nhập mã CN / tên chức năng → autocomplete + modal "Hồ sơ chức năng"
# hiển thị full lifecycle: meta + từng phase (Start/End/Status/PIC) + summary
# (đang ở phase nào, có trễ không, next deadline). Auto-detect cột theo .cursorrules.
# ==========================================================================


@app.route("/api/projects/<slug>/function-search")
def function_search(slug: str):
    """
    Autocomplete tra cứu function. Query params:
    - q: chuỗi search (mã CN / tên / module / quy trình), tối thiểu 1 ký tự
    - limit: số kết quả max (default 10)

    Trả `{items: [...]}` — mỗi item có row_num để mở detail modal.
    """
    from analyzer.function_traceability import search_functions
    state, err = _require_state(slug)
    if err:
        return err
    q = (request.args.get("q") or "").strip()
    try:
        limit = int(request.args.get("limit", 10))
    except (TypeError, ValueError):
        limit = 10
    limit = max(1, min(limit, 50))  # kẹp 1–50

    items = search_functions(state["data"], q, limit=limit)
    return jsonify({"query": q, "count": len(items), "items": items})


@app.route("/api/projects/<slug>/function-detail/<int:row_num>")
def function_detail(slug: str, row_num: int):
    """Full lifecycle của 1 function (theo row_num Excel gốc)."""
    from analyzer.function_traceability import get_function_detail
    state, err = _require_state(slug)
    if err:
        return err
    result = get_function_detail(state["data"], row_num)
    if result is None:
        return jsonify({"error": f"Không tìm thấy function row {row_num}"}), 404
    return jsonify(result)


# ==========================================================================
# FIT/GAP Dashboard (BA — Task 2)
# ==========================================================================
# BA cần section riêng để quản lý lifecycle GAP: cards summary + 3 chart
# (module / quy trình / priority) + bảng aging GAP > 14 ngày. Global filter
# (module/process/pic) apply. Xuất Excel tuân rule V4 (xuất ALL).
# ==========================================================================


@app.route("/api/projects/<slug>/fitgap-analytics")
def fitgap_analytics(slug: str):
    """Cards + 3 chart data + aging list. Apply global filter module/process/pic."""
    from analyzer.fitgap_analytics import compute_fitgap_analytics
    state, err = _require_state(slug)
    if err:
        return err
    data = _filtered_data_from_request(state)
    # Cho phép override aging threshold qua query param (BA có thể thử 7/21/30 ngày)
    try:
        thr = int(request.args.get("aging_threshold_days", 14))
    except (TypeError, ValueError):
        thr = 14
    thr = max(1, min(thr, 365))
    return jsonify(compute_fitgap_analytics(data, aging_threshold_days=thr))


@app.route("/api/projects/<slug>/export-fitgap", methods=["GET", "POST"])
def export_fitgap(slug: str):
    """Xuất Excel FIT/GAP Dashboard (multi-sheet, xuất ALL). Áp global filter."""
    from analyzer.fitgap_analytics import compute_fitgap_analytics
    state, err = _require_state(slug)
    if err:
        return err
    try:
        data, filters = _filter_state_from_body_or_args(state)
        try:
            thr = int(request.args.get("aging_threshold_days", 14))
        except (TypeError, ValueError):
            thr = 14
        payload = compute_fitgap_analytics(data, aging_threshold_days=thr)
        filepath = export_fitgap_report(
            payload=payload,
            output_dir=_project_mgr.get_export_dir(slug),
            filters=filters,
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất FIT/GAP: {str(e)}"}), 500


# ==========================================================================
# Function Diff (BA — Task 3)
# ==========================================================================
# So sánh state hiện tại với snapshot upload trước. Snapshot đã có sẵn dưới
# dạng pickled ParsedData (SnapshotManager). Không cần tạo JSON snapshot
# riêng — pickle rẻ hơn và đã có sẵn full data.
#
# Match function bằng Mã CN, fallback theo (Tên + Module). Trả 6 tab data:
# Added / Deleted / PIC / Priority-Complexity / FIT-GAP / Phase Status.
# ==========================================================================


def _resolve_diff_previous_snapshot(slug: str, vs: str):
    """
    Chọn snapshot đóng vai "trước" khi diff.

    Args:
        vs: 'previous' (mặc định) → snapshot ngay trước current;
            hoặc YYYY-MM-DD → snapshot cụ thể.

    Return: (parsed_data_prev, meta_dict, error_response_or_None)
    """
    smgr = _project_mgr.get_snapshot_manager(slug)
    snaps = smgr.list_snapshots()
    if not snaps:
        return None, None, (
            jsonify({"error": "Chưa có snapshot nào để so sánh", "code": "NO_SNAPSHOT"}),
            404,
        )

    if vs and vs != "previous":
        # snapshot cụ thể theo date
        loaded = smgr.load_snapshot(vs)
        if not loaded:
            return None, None, (jsonify({"error": f"Snapshot {vs} không tồn tại"}), 404)
        return loaded["parsed"], loaded["meta"], None

    # vs=previous → snapshot ngay trước current.
    # snapshot mới nhất (index 0) là snapshot chính hôm nay (thường trùng current).
    # → 'previous' đúng phải là index 1. Nếu chỉ có 1 snapshot → không diff được.
    if len(snaps) < 2:
        return None, None, (
            jsonify({
                "error": "Chỉ có 1 snapshot — chưa có snapshot trước để so sánh. "
                         "Upload file lần thứ 2 để bắt đầu track diff.",
                "code": "SINGLE_SNAPSHOT",
                "current_snapshot": snaps[0] if snaps else None,
            }),
            404,
        )

    prev_entry = snaps[1]
    loaded = smgr.load_snapshot(prev_entry["date"])
    if not loaded:
        return None, None, (jsonify({"error": "Không load được snapshot trước"}), 500)
    return loaded["parsed"], loaded["meta"], None


def _build_auto_diff_summary(slug: str, current_data, current_meta: Optional[dict] = None) -> Optional[dict]:
    """
    Sau upload/sync: so với snapshot trước → badges ngắn cho summary cards.
    Return None nếu chưa đủ 2 snapshot.
    """
    from analyzer.function_diff import compute_function_diff

    prev_data, prev_meta, err = _resolve_diff_previous_snapshot(slug, "previous")
    if err:
        return None
    try:
        payload = compute_function_diff(
            current=current_data,
            previous=prev_data,
            current_meta=current_meta or {},
            previous_meta=prev_meta or {},
        )
        badges = payload.get("badges") or {}
        return {
            "available": True,
            "previous_date": (prev_meta or {}).get("date"),
            "current_date": (current_meta or {}).get("date"),
            "badges": badges,
            "counts": payload.get("counts") or {},
            # Lists rút gọn cho modal click-through (cap 50)
            "lists": {
                "added": (payload.get("added") or [])[:50],
                "status_rollback": (payload.get("status_rollback") or [])[:50],
                "pic_changed": (payload.get("pic_changed") or [])[:50],
                "deleted": (payload.get("deleted") or [])[:50],
            },
        }
    except Exception:
        return None


@app.route("/api/projects/<slug>/function-diff")
def function_diff(slug: str):
    """
    So sánh state hiện tại với snapshot trước.
    Query params:
    - vs: 'previous' (default) hoặc YYYY-MM-DD của snapshot cụ thể.
    """
    from analyzer.function_diff import compute_function_diff
    state, err = _require_state(slug)
    if err:
        return err

    vs = request.args.get("vs", "previous").strip() or "previous"
    prev_data, prev_meta, err2 = _resolve_diff_previous_snapshot(slug, vs)
    if err2:
        return err2

    # Meta cho snapshot "current" — dùng thông tin _state
    current_meta = {
        "date": state["upload_time"].date().isoformat() if state.get("upload_time") else None,
        "filename": state.get("filename", "current.xlsx"),
        "total_functions": len(state["data"].rows),
    }

    payload = compute_function_diff(
        current=state["data"],
        previous=prev_data,
        current_meta=current_meta,
        previous_meta=prev_meta,
    )
    # List snapshots để FE hiển thị dropdown "so với snapshot nào"
    payload["available_snapshots"] = _project_mgr.get_snapshot_manager(slug).list_snapshots()
    return jsonify(payload)


@app.route("/api/projects/<slug>/export-function-diff")
def export_function_diff(slug: str):
    """Xuất Excel Function Diff (multi-sheet, xuất ALL). Query: ?vs="""
    from analyzer.function_diff import compute_function_diff
    state, err = _require_state(slug)
    if err:
        return err

    vs = request.args.get("vs", "previous").strip() or "previous"
    prev_data, prev_meta, err2 = _resolve_diff_previous_snapshot(slug, vs)
    if err2:
        return err2

    current_meta = {
        "date": state["upload_time"].date().isoformat() if state.get("upload_time") else None,
        "filename": state.get("filename", "current.xlsx"),
        "total_functions": len(state["data"].rows),
    }

    try:
        payload = compute_function_diff(
            current=state["data"],
            previous=prev_data,
            current_meta=current_meta,
            previous_meta=prev_meta,
        )
        filepath = export_function_diff_report(
            payload=payload,
            output_dir=_project_mgr.get_export_dir(slug),
        )
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        return jsonify({"error": f"Lỗi xuất Function Diff: {str(e)}"}), 500


# ==========================================================================
# T33 — Public API (REST + iframe + PNG snapshot)
# ==========================================================================
#
# Admin CRUD (yêu cầu là "chủ project" — hiện app chỉ có 1 user local nên
# không có auth layer riêng; production sẽ cần thêm session/JWT):
#     GET    /api/projects/<slug>/public-tokens             — list
#     POST   /api/projects/<slug>/public-tokens             — create
#     DELETE /api/projects/<slug>/public-tokens/<token_id>  — revoke
#     GET    /api/projects/<slug>/public-scopes             — metadata cho FE
#
# Public read (header X-API-Key hoặc ?token=pub_xxx):
#     GET    /public/api/v1/projects/<slug>/summary
#     GET    /public/api/v1/projects/<slug>/charts/<chart_id>
#     GET    /public/api/v1/projects/<slug>/functions?page=&size=
#
# Rate limit: 60 req / 60s / token — vượt trả 429 + Retry-After.
# CORS: allow-all (mission: embed vào Confluence/Word/email public), method GET
#       only, header X-API-Key allowed. Không credentials/cookie.
# ==========================================================================

from functools import wraps
from analyzer import public_api as _pubapi


def _add_cors_headers(resp):
    """Add CORS headers cho public API — allow all origin, method GET/OPTIONS."""
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "X-API-Key, Content-Type"
    resp.headers["Access-Control-Max-Age"] = "3600"
    return resp


def _extract_public_token() -> str:
    """
    Thứ tự ưu tiên: header X-API-Key → query ?token=.
    Trả string (rỗng nếu không có).
    """
    tok = (request.headers.get("X-API-Key") or "").strip()
    if not tok:
        tok = (request.args.get("token") or "").strip()
    return tok


def require_public_token(scope: Optional[str] = None):
    """
    Decorator xác thực token cho public API.

    Args:
        scope: scope key cần thiết (VD "summary", "module-overview"). Nếu
               token có scope "*" → cho phép. Nếu None → chỉ verify token
               hợp lệ, không check scope.
    """
    def deco(fn):
        @wraps(fn)
        def wrapper(slug, *args, **kwargs):
            # Preflight OPTIONS → return 204 với CORS headers (không auth)
            if request.method == "OPTIONS":
                return _add_cors_headers(app.response_class(status=204))
            if not _project_mgr.project_exists(slug):
                return _add_cors_headers(jsonify({"error": "Project không tồn tại"})), 404
            project_dir = _project_dir_for(slug)
            token = _extract_public_token()
            try:
                entry = _pubapi.verify_token(project_dir, token, required_scope=scope)
                _pubapi.check_rate_limit(entry["id"])
            except _pubapi.RateLimitError as e:
                resp = jsonify({"error": str(e), "retry_after": e.retry_after})
                resp = _add_cors_headers(resp)
                resp.headers["Retry-After"] = str(e.retry_after)
                return resp, 429
            except _pubapi.PublicApiError as e:
                return _add_cors_headers(jsonify({"error": str(e)})), e.status_code

            # Update last_used_at best-effort (silent-fail)
            _pubapi.touch_last_used(project_dir, entry["id"])
            resp = fn(slug, *args, **kwargs)
            # Nếu view trả tuple (resp, code) → apply CORS lên resp
            if isinstance(resp, tuple):
                r0 = resp[0]
                _add_cors_headers(r0)
                return resp
            return _add_cors_headers(resp)

        return wrapper
    return deco


# --- Admin CRUD (không cần token, chỉ cần project tồn tại) ---

@app.route("/api/projects/<slug>/public-tokens", methods=["GET", "POST"])
def public_tokens_collection(slug: str):
    """List tokens hoặc create token mới."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    project_dir = _project_dir_for(slug)
    if request.method == "GET":
        return jsonify({"tokens": _pubapi.list_tokens(project_dir)})
    body = request.get_json(silent=True) or {}
    try:
        plaintext, entry = _pubapi.create_token(
            project_dir,
            name=body.get("name"),
            scope=body.get("scope"),
            expires_in_days=body.get("expires_in_days"),
        )
    except _pubapi.PublicApiError as e:
        return jsonify({"error": str(e)}), e.status_code
    # Trả plaintext token CHÍNH XÁC 1 LẦN — FE phải cảnh báo user copy ngay.
    return jsonify({
        "token": plaintext,
        "entry": entry,
        "warning": "Token này chỉ hiển thị 1 lần — copy ngay và lưu chỗ an toàn.",
    })


@app.route("/api/projects/<slug>/public-tokens/<token_id>", methods=["DELETE"])
def public_tokens_delete(slug: str, token_id: str):
    """Revoke token — không xoá entry (giữ audit trail)."""
    if not _project_mgr.project_exists(slug):
        return jsonify({"error": "Project không tồn tại"}), 404
    ok = _pubapi.revoke_token(_project_dir_for(slug), token_id)
    if not ok:
        return jsonify({"error": "Không tìm thấy token"}), 404
    return jsonify({"ok": True, "revoked": token_id})


# ==========================================================================
# T34 Task 2 — LAN admin API: URL LAN + access log tail (localhost only)
# ==========================================================================

@app.route("/api/lan/info")
def lan_info():
    """
    Trả thông tin LAN cho UI: URL truy cập từ mọi interface + info bảo mật.

    Response:
      {
        "urls": [{"ip": "...", "url": "...", "label": "..."}],
        "port": 5000,
        "admin_guard": bool,  // guard có bật không
        "access_log": bool,   // log có bật không
        "extra_admin_allowlist": ["192.168.1.10", ...],
        "is_localhost_request": bool,  // request hiện tại có phải localhost không
      }

    Không cần auth — chỉ trả metadata public. Localhost check hiển thị để
    FE có thể hiện badge "🔒 Bạn đang truy cập từ LAN, admin bị khóa".
    """
    from analyzer import lan_security as ls
    return jsonify({
        "urls": ls.detect_lan_ips(port=5000),
        "port": 5000,
        "admin_guard": os.environ.get("IHRP_DISABLE_ADMIN_GUARD", "") != "1",
        "access_log": os.environ.get("IHRP_DISABLE_ACCESS_LOG", "") != "1",
        "extra_admin_allowlist": sorted(ls._extra_allow_list()),
        "is_localhost_request": ls.is_localhost_request(request),
    })


@app.route("/api/lan/access-log")
def lan_access_log():
    """
    Trả tail của access log để UI Settings hiển thị lịch sử truy cập.

    Query: limit (default 100, max 500).
    Chỉ cho phép localhost xem — thông tin IP nội bộ là nhạy cảm.
    """
    from analyzer import lan_security as ls
    if not ls.is_localhost_request(request):
        return jsonify({
            "error": "Access log chỉ xem được từ máy chủ (localhost).",
            "code": "LOCALHOST_ONLY",
        }), 403
    try:
        limit = int(request.args.get("limit") or 100)
    except (TypeError, ValueError):
        limit = 100
    limit = max(1, min(limit, 500))
    entries = ls.read_access_log_tail(_ACCESS_LOG_PATH, limit=limit)
    return jsonify({
        "entries": entries,
        "count": len(entries),
        "log_path": _ACCESS_LOG_PATH,
    })


@app.route("/api/projects/<slug>/public-scopes")
def public_scopes(slug: str):
    """
    Metadata cho FE build multi-select scope. Trả danh sách scope + label
    tiếng Việt. Endpoint này không cần project exist (metadata static) nhưng
    giữ URL pattern để nhất quán.
    """
    return jsonify({"scopes": _pubapi.PUBLIC_SCOPES})


# --- Public read API (yêu cầu token) ---

def _public_metrics_or_error(slug: str):
    """Load metrics hiện tại; nếu chưa upload → trả JSON error + status."""
    st = _get_state(slug)
    if not st or not st.get("metrics"):
        return None, (jsonify({"error": "Project chưa có data (chưa upload)"}), 404)
    return st, None


@app.route("/public/api/v1/projects/<slug>/summary", methods=["GET", "OPTIONS"])
@require_public_token(scope="summary")
def public_summary(slug: str):
    st, err = _public_metrics_or_error(slug)
    if err:
        return err
    m = st["metrics"]
    summary = m.get("summary") or {}
    project = _project_mgr.get_project(slug)
    return jsonify({
        "project": {
            "slug": slug,
            "name": project.name if project else slug,
        },
        "summary": summary,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    })


# Map chart_id (public API) → key trong metrics dict.
# Định nghĩa tường minh để user copy snippet đúng, tránh nhầm với chart_id nội bộ.
_PUBLIC_CHART_MAP: dict[str, str] = {
    "module-overview": "module_overview",
    "phase-matrix": "phase_status_matrix",
    "phase-stacked": "phase_progress_stacked",
    "progress-task-type": "progress_by_task_type",
    "pic-workload": "pic_workload",
    "priority": "priority_breakdown",
    "complexity": "complexity_breakdown",
    "fit-gap": "fit_gap_analysis",
    "giai-doan": "giai_doan_progress",
    "overdue": "overdue_list",
    "unassigned": "unassigned_tasks",
    "stalled": "stalled_tasks",
    "risk": "risk_scores",
    "effort-heatmap": "effort_analysis",
    "process": "process_analysis",
}


@app.route("/public/api/v1/projects/<slug>/charts/<chart_id>", methods=["GET", "OPTIONS"])
def public_chart(slug: str, chart_id: str):
    """
    Public chart endpoint — scope check dynamic theo `chart_id`.

    Không dùng decorator require_public_token cố định vì scope phụ thuộc
    chart_id. Verify inline.
    """
    if request.method == "OPTIONS":
        return _add_cors_headers(app.response_class(status=204))
    if not _project_mgr.project_exists(slug):
        return _add_cors_headers(jsonify({"error": "Project không tồn tại"})), 404
    if chart_id not in _PUBLIC_CHART_MAP:
        return _add_cors_headers(jsonify({
            "error": f"Chart không hỗ trợ: {chart_id}",
            "supported": sorted(_PUBLIC_CHART_MAP.keys()),
        })), 400

    project_dir = _project_dir_for(slug)
    token = _extract_public_token()
    try:
        entry = _pubapi.verify_token(project_dir, token, required_scope=chart_id)
        _pubapi.check_rate_limit(entry["id"])
    except _pubapi.RateLimitError as e:
        resp = jsonify({"error": str(e), "retry_after": e.retry_after})
        resp = _add_cors_headers(resp)
        resp.headers["Retry-After"] = str(e.retry_after)
        return resp, 429
    except _pubapi.PublicApiError as e:
        return _add_cors_headers(jsonify({"error": str(e)})), e.status_code

    _pubapi.touch_last_used(project_dir, entry["id"])

    st, err = _public_metrics_or_error(slug)
    if err:
        # Tuple → phải wrap CORS thủ công
        r0, code = err
        _add_cors_headers(r0)
        return r0, code

    metrics_key = _PUBLIC_CHART_MAP[chart_id]
    data = st["metrics"].get(metrics_key)
    resp = jsonify({
        "chart_id": chart_id,
        "data": data if data is not None else {},
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    })
    return _add_cors_headers(resp)


@app.route("/public/api/v1/projects/<slug>/functions", methods=["GET", "OPTIONS"])
@require_public_token(scope="functions")
def public_functions(slug: str):
    """
    List danh sách function (paginated). Trả subset field an toàn — không
    expose PIC email/phone chi tiết. Query: ?page=1&size=50 (max size=200).
    """
    st, err = _public_metrics_or_error(slug)
    if err:
        return err
    try:
        page = max(1, int(request.args.get("page") or 1))
    except ValueError:
        page = 1
    try:
        size = min(200, max(1, int(request.args.get("size") or 50)))
    except ValueError:
        size = 50

    rows = st["data"].rows if st.get("data") else []
    total = len(rows)
    start = (page - 1) * size
    end = start + size
    page_rows = rows[start:end]
    items: list[dict] = []
    for r in page_rows:
        # Progress % của function — reuse public metric nếu có
        # Compute nhanh: tổng phase có status Closed / total phase
        phase_stats = {"total": 0, "closed": 0, "open": 0}
        for pd in (r.phases or {}).values():
            phase_stats["total"] += 1
            if pd.status == "Closed":
                phase_stats["closed"] += 1
            elif pd.status in ("Open", None, ""):
                phase_stats["open"] += 1
        items.append({
            "ma_cn": r.meta.get("ma_cn") or "",
            "ten_cn": r.meta.get("ten_cn") or "",
            "module": r.meta.get("module") or "",
            "process": r.meta.get("quy_trinh") or "",
            "priority": r.meta.get("priority") or "",
            "complexity": r.meta.get("complexity") or "",
            "fit_gap": r.meta.get("fit_gap") or "",
            "giai_doan": r.meta.get("giai_doan") or "",
            "phase_stats": phase_stats,
        })
    return jsonify({
        "items": items,
        "page": page,
        "size": size,
        "total": total,
        "total_pages": (total + size - 1) // size if size else 1,
    })


# ==========================================================================
# T33 Task 2B — iframe embed + PNG snapshot (Playwright)
# ==========================================================================

# Label tiếng Việt cho từng chart_id — hiển thị làm title trong iframe/PNG.
_PUBLIC_CHART_LABELS: dict[str, str] = {
    "module-overview":    "Tổng quan theo Module",
    "phase-matrix":       "Phase × Status Matrix",
    "phase-stacked":      "Tiến độ theo Phase",
    "progress-task-type": "Tiến độ theo loại công việc",
    "pic-workload":       "Khối lượng PIC",
    "priority":           "Phân bố Priority",
    "complexity":         "Phân bố Complexity",
    "fit-gap":            "Phân tích FIT/GAP",
    "giai-doan":          "Tiến độ theo Giai đoạn",
    "overdue":            "Danh sách trễ deadline",
    "unassigned":         "Task chưa có PIC",
    "stalled":            "Task đình trệ",
    "risk":               "Top 20 Risk Score",
    "effort-heatmap":     "Effort Heatmap",
    "process":            "Phân tích Quy trình",
}


@app.route("/embed/<slug>/<chart_id>")
def embed_chart(slug: str, chart_id: str):
    """
    Trang embed chart cho iframe / PNG snapshot.

    Query:
      token=pub_...     (bắt buộc — verify inside embed JS gọi public API)
      bg=transparent    (optional — nền trong suốt cho blend UI)

    KHÔNG verify token ở đây — token verify khi FE gọi
    /public/api/v1/.../charts/<chart_id>. Điều này giúp OPTIONS preflight
    và HTML render nhanh; nếu token sai FE sẽ hiển error box.

    X-Frame-Options: bỏ hoàn toàn (mặc định Flask không set). Nếu deploy
    sau reverse proxy có set → override thêm response header
    'X-Frame-Options: ALLOWALL'.
    """
    if chart_id not in _PUBLIC_CHART_MAP:
        return jsonify({
            "error": f"Chart không hỗ trợ embed: {chart_id}",
            "supported": sorted(_PUBLIC_CHART_MAP.keys()),
        }), 400
    project = _project_mgr.get_project(slug)
    project_name = project.name if project else slug
    bg = (request.args.get("bg") or "").strip().lower()
    resp = app.make_response(render_template(
        "embed.html",
        slug=slug,
        chart_id=chart_id,
        chart_label=_PUBLIC_CHART_LABELS.get(chart_id, chart_id),
        project_name=project_name,
        bg=bg,
    ))
    # Cho phép nhúng vào bất kỳ site nào — override reverse-proxy default
    resp.headers["X-Frame-Options"] = "ALLOWALL"
    # CSP tối thiểu: script từ jsdelivr (Chart.js), style inline (ok cho embed)
    resp.headers["Content-Security-Policy"] = (
        "default-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors *"
    )
    return resp


# --- PNG snapshot (Playwright) ---

# Cache config: TTL 5 phút — refresh khi dashboard update tương đối chậm rãi.
_PNG_CACHE_TTL_SEC = 300
_PNG_CACHE_DIR_NAME = "public_cache"


def _png_cache_dir(slug: str) -> str:
    """Thư mục cache PNG per-project."""
    d = os.path.join(_project_dir_for(slug), _PNG_CACHE_DIR_NAME)
    os.makedirs(d, exist_ok=True)
    return d


def _png_cache_key(chart_id: str, w: int, h: int, bg: str) -> str:
    """
    Cache key = chart_id_WxH_bg.png.
    Không hash token — cache là public per-project (mọi token cùng scope
    xem cùng ảnh). Đảm bảo revoke token vẫn hoạt động vì verify vẫn chạy
    trước khi serve cache.
    """
    bg_tag = bg or "white"
    return f"{chart_id}_{w}x{h}_{bg_tag}.png"


def _try_playwright_screenshot(url: str, w: int, h: int, out_path: str,
                               wait_selector: str = "body[data-chart-ready]",
                               timeout_ms: int = 15000) -> Optional[str]:
    """
    Chạy Playwright headless chromium để chụp ảnh. Trả None nếu thành công,
    hoặc chuỗi error nếu fail (VD Playwright chưa install).

    Playwright là **optional dep**. Nếu chưa install → return error message
    hướng dẫn user cài. Endpoint gọi hàm này trả 503.
    """
    try:
        from playwright.sync_api import sync_playwright  # type: ignore
    except ImportError:
        return (
            "Playwright chưa cài. Chạy: pip install playwright "
            "&& python -m playwright install chromium (~200MB)"
        )
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            try:
                context = browser.new_context(viewport={"width": w, "height": h})
                page = context.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                # Chờ chart render xong — JS embed set data-chart-ready
                page.wait_for_selector(wait_selector, timeout=timeout_ms)
                # Extra 200ms cho animation cuối
                page.wait_for_timeout(200)
                page.screenshot(path=out_path, full_page=False,
                                clip={"x": 0, "y": 0, "width": w, "height": h})
                context.close()
            finally:
                browser.close()
        return None
    except Exception as e:
        return f"Playwright chụp ảnh lỗi: {e}"


@app.route("/public/api/v1/projects/<slug>/charts/<chart_id>/image",
           methods=["GET", "OPTIONS"])
def public_chart_image(slug: str, chart_id: str):
    """
    PNG snapshot của 1 chart. Query:
      w=800, h=400     (viewport size — default 800x400, max 1920x1200)
      bg=transparent   (optional)
      token=pub_...    (bắt buộc — verify scope <chart_id>)

    Flow:
      1. Verify token + scope + rate limit (giống public_chart).
      2. Check cache PNG → serve nếu chưa TTL.
      3. Playwright: mở http://localhost:<port>/embed/<slug>/<chart_id>?token=
         → screenshot → lưu cache → serve.

    Playwright chưa install → HTTP 503 với hướng dẫn cài.
    """
    if request.method == "OPTIONS":
        return _add_cors_headers(app.response_class(status=204))
    if not _project_mgr.project_exists(slug):
        return _add_cors_headers(jsonify({"error": "Project không tồn tại"})), 404
    if chart_id not in _PUBLIC_CHART_MAP:
        return _add_cors_headers(jsonify({
            "error": f"Chart không hỗ trợ: {chart_id}",
            "supported": sorted(_PUBLIC_CHART_MAP.keys()),
        })), 400

    project_dir = _project_dir_for(slug)
    token = _extract_public_token()
    try:
        entry = _pubapi.verify_token(project_dir, token, required_scope=chart_id)
        _pubapi.check_rate_limit(entry["id"])
    except _pubapi.RateLimitError as e:
        resp = jsonify({"error": str(e), "retry_after": e.retry_after})
        resp = _add_cors_headers(resp)
        resp.headers["Retry-After"] = str(e.retry_after)
        return resp, 429
    except _pubapi.PublicApiError as e:
        return _add_cors_headers(jsonify({"error": str(e)})), e.status_code
    _pubapi.touch_last_used(project_dir, entry["id"])

    # Parse w/h với cap để tránh abuse (chụp 10000x10000 nướng CPU)
    try:
        w = max(200, min(1920, int(request.args.get("w") or 800)))
    except ValueError:
        w = 800
    try:
        h = max(150, min(1200, int(request.args.get("h") or 400)))
    except ValueError:
        h = 400
    bg = (request.args.get("bg") or "").strip().lower()

    cache_dir = _png_cache_dir(slug)
    cache_name = _png_cache_key(chart_id, w, h, bg)
    cache_path = os.path.join(cache_dir, cache_name)

    # Cache hit → serve trực tiếp
    if os.path.isfile(cache_path):
        age = time.time() - os.path.getmtime(cache_path)
        if age < _PNG_CACHE_TTL_SEC:
            resp = send_file(cache_path, mimetype="image/png", as_attachment=False,
                             download_name=cache_name, max_age=_PNG_CACHE_TTL_SEC)
            _add_cors_headers(resp)
            resp.headers["X-Cache"] = "HIT"
            resp.headers["X-Cache-Age"] = str(int(age))
            return resp

    # Cache miss → screenshot
    # Build embed URL nội bộ — dùng request.host_url để hỗ trợ mọi host
    embed_qs = f"?token={token}"
    if bg == "transparent":
        embed_qs += "&bg=transparent"
    embed_url = f"{request.host_url.rstrip('/')}/embed/{slug}/{chart_id}{embed_qs}"

    err = _try_playwright_screenshot(embed_url, w, h, cache_path)
    if err:
        # ImportError message → 503; runtime error → 500
        status = 503 if "chưa cài" in err else 500
        return _add_cors_headers(jsonify({
            "error": err,
            "hint": "Xem docs/PUBLIC_API_GUIDE.md section Playwright install.",
        })), status

    resp = send_file(cache_path, mimetype="image/png", as_attachment=False,
                     download_name=cache_name, max_age=_PNG_CACHE_TTL_SEC)
    _add_cors_headers(resp)
    resp.headers["X-Cache"] = "MISS"
    return resp


# Import time module ở top-level cho _png_cache
import time  # noqa: E402


# ==========================================================================
# Main
# ==========================================================================

if __name__ == "__main__":
    import sys

    debug_mode = "--debug" in sys.argv
    reloader = debug_mode  # chỉ auto-reload khi user bật --debug
    # Solo-safe: mặc định 127.0.0.1. Mở LAN: IHRP_LAN=1 hoặc IHRP_BIND_LOCAL_ONLY=0.
    bind_host = _lansec.resolve_bind_host()
    port = 5000

    print("\n" + "=" * 60)
    print("  iHRP Function List Tracker (V3)")
    print(f"  Listen: {bind_host}:{port}")
    print(f"  Local URL: http://localhost:{port}")
    if bind_host == "0.0.0.0":
        print("  [CANH BAO] Bind 0.0.0.0 — GET dashboard mo tren LAN.")
        print("             Khong dung tren WiFi cong cong; chi mang noi bo cong ty.")
    else:
        print("  Bind: LOCAL-ONLY (mac dinh). Mo LAN: set IHRP_LAN=1")
    if debug_mode:
        print("  Mode: DEBUG (auto-reload BAT khi sua file)")
    else:
        print("  Mode: PRODUCTION (on dinh, khong auto-reload)")
    print("=" * 60 + "\n")
    # T26: check digest scheduler ngay khi start (không block server)
    # Chỉ chạy 1 lần với reloader parent — tránh spawn duplicate digest
    # khi Flask debug=True fork worker.
    if not reloader or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        _run_startup_digest_scheduler()
        _run_startup_auto_archive()
    try:
        app.run(debug=debug_mode, use_reloader=reloader, port=port, host=bind_host)
    except Exception as e:
        print(f"\n[LOI] Server crash: {e}")
        import traceback
        traceback.print_exc()
        input("\nNhan Enter de thoat...")
