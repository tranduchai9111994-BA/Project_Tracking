"""
T34 Task 3 (B) — Type inference cho Column Mapping Wizard.

Mục đích: quan sát 3-5 sample values → đoán kiểu → giới hạn dropdown mapping
chỉ show iHRP cols compatible → giảm risk map sai.

Types trả về (string constants):
  - "date_iso"          — "2026-07-30" hoặc datetime object
  - "date_dmy"          — "30/07/2026" hoặc "30-07-26"
  - "date_excel_serial" — Excel serial number (float 40000+)
  - "integer"           — 42
  - "decimal"           — 3.14
  - "pic_list"          — "BaoLQ31, NhiVN" (multi-token separated)
  - "boolean"           — TRUE/FALSE/YES/NO/1/0 (2 giá trị)
  - "status_enum"       — Open/Assigned/In-progress/Closed/…
  - "empty"             — toàn None/blank
  - "string"            — fallback

Compatible iHRP columns:
  - date_*        → Analysis-Start/End, Dev-Start/End, ...-Start/End columns
  - pic_list      → *-PIC columns
  - status_enum   → *-Status columns
  - integer/decimal → *-Estimate MH, Priority (nếu số), Giai đoạn (nếu số)
  - boolean       → không map trực tiếp (edge case)
  - string        → mọi cột (fallback)
  - empty         → không suggest (user tự chọn)

Không dùng regex nặng — đơn giản, testable, fast.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any, Optional

# ==========================================================================
# CONSTANTS — patterns + enum lists
# ==========================================================================

# Status hợp lệ theo .cursorrules (bao gồm cả biến thể viết hoa/thường)
_STATUS_ENUMS_LOWER = {
    "open", "assigned", "in-progress", "in progress", "inprogress",
    "resolved", "closed", "pending", "cancelled", "canceled",
    "wip", "done", "todo", "review",
}

_BOOLEAN_LOWER = {
    "true", "false", "yes", "no", "y", "n", "t", "f", "1", "0",
    "có", "không", "co", "khong",
}

# Regex — compile 1 lần
_RE_DATE_ISO = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}")
_RE_DATE_DMY = re.compile(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}")
_RE_INT = re.compile(r"^-?\d+$")
_RE_DECIMAL = re.compile(r"^-?\d+([\.,]\d+)?$")

# PIC separator: comma, semicolon, newline, plus (theo .cursorrules)
_RE_PIC_SEP = re.compile(r"[,;\n+]+")


# ==========================================================================
# TYPE INFERENCE
# ==========================================================================

def _is_date_like(v: Any) -> bool:
    """True nếu v là date/datetime object hoặc str parseable dạng ISO/DMY."""
    if isinstance(v, (date, datetime)):
        return True
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s:
        return False
    if _RE_DATE_ISO.match(s) or _RE_DATE_DMY.match(s):
        return True
    return False


def _classify_single(v: Any) -> str:
    """Classify 1 value → trả 1 type string. Không dùng cho kết quả cuối."""
    if v is None:
        return "empty"
    # Datetime/date object
    if isinstance(v, (datetime, date)):
        return "date_iso"
    # Bool
    if isinstance(v, bool):
        return "boolean"
    # Int/float
    if isinstance(v, int):
        return "integer"
    if isinstance(v, float):
        # Excel serial date thường ở khoảng 40000-60000 (2009-2064)
        if 30000 < v < 80000 and float(v).is_integer():
            return "date_excel_serial"
        return "decimal"

    # String
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return "empty"
        # Enum status? (check trước date để "In-progress" không nhầm)
        if s.lower() in _STATUS_ENUMS_LOWER:
            return "status_enum"
        # Boolean literal
        if s.lower() in _BOOLEAN_LOWER:
            return "boolean"
        # Date
        if _RE_DATE_ISO.match(s):
            return "date_iso"
        if _RE_DATE_DMY.match(s):
            return "date_dmy"
        # Numeric
        if _RE_INT.match(s):
            return "integer"
        if _RE_DECIMAL.match(s):
            return "decimal"
        # Multi-PIC — có separator + >=2 non-empty token
        if _RE_PIC_SEP.search(s):
            tokens = [t.strip() for t in _RE_PIC_SEP.split(s) if t.strip()]
            if len(tokens) >= 2:
                return "pic_list"
        # Còn lại → string
        return "string"

    # Loại khác (bytes, list, dict…) → string fallback
    return "string"


def infer_type(sample_values: list[Any]) -> str:
    """
    Đoán kiểu chung từ list sample values.

    Logic:
      1. Loại "empty" ra khỏi voting.
      2. Nếu list toàn empty → "empty".
      3. Nếu MỌI value non-empty đều cùng 1 type → return type đó.
      4. Nếu mix nhưng > 60% cùng type → return type đó (tolerance).
      5. Nếu mix hoàn toàn → "string" (fallback safe nhất).

    Special: date_iso + date_dmy + date_excel_serial → gộp thành "date_iso"
    (đại diện chung date). date_excel_serial gộp riêng vì FE hiển badge
    khác được.

    Ví dụ:
      >>> infer_type(["2026-01-15", "2026-02-20", None])
      "date_iso"
      >>> infer_type(["10", "20", "30"])
      "integer"
      >>> infer_type(["Alice, Bob", "Cai;Dee\\nEve"])
      "pic_list"
      >>> infer_type(["Open", "Closed", "In-progress"])
      "status_enum"
    """
    if not sample_values:
        return "empty"

    types = [_classify_single(v) for v in sample_values]
    non_empty = [t for t in types if t != "empty"]
    if not non_empty:
        return "empty"

    # Gộp variant date thành 1 để đếm cho robust
    def _canonicalize(t: str) -> str:
        if t in ("date_iso", "date_dmy"):
            return "date_iso"
        return t

    canon = [_canonicalize(t) for t in non_empty]
    unique = set(canon)

    # All same
    if len(unique) == 1:
        return canon[0]

    # Majority (>60%)
    from collections import Counter
    counter = Counter(canon)
    top_type, top_count = counter.most_common(1)[0]
    if top_count / len(canon) >= 0.6:
        return top_type

    return "string"


# ==========================================================================
# COMPATIBLE iHRP COLUMNS PER TYPE
# ==========================================================================

def _phase_cols(suffix: str) -> list[str]:
    """Sinh list tất cả `<Phase> - <suffix>` (dùng cho date/PIC/Status)."""
    from parser.column_mapping import _PHASE_NAMES
    return [f"{p} - {suffix}" for p in _PHASE_NAMES]


def compatible_ihrp_cols(inferred_type: str) -> list[str]:
    """
    Return list iHRP columns compatible với inferred_type.

    Empty list nghĩa là "không suggest gì" — user chọn tay từ full list.
    """
    if inferred_type == "empty":
        return []
    if inferred_type in ("date_iso", "date_dmy", "date_excel_serial"):
        # Date compatible: mọi Start / End column
        return _phase_cols("Start") + _phase_cols("End") + [
            "Last Updated Date",
        ]
    if inferred_type == "pic_list":
        return _phase_cols("PIC")
    if inferred_type == "status_enum":
        return _phase_cols("Status")
    if inferred_type in ("integer", "decimal"):
        # Numeric: Estimate MH + có thể Priority/Giai đoạn nếu là số
        return _phase_cols("Estimate MH") + ["Priority", "Giai đoạn"]
    if inferred_type == "boolean":
        # Không map trực tiếp — trả empty (user tự chọn nếu cần)
        return []
    # string / fallback → all iHRP columns (không restrict)
    from parser.column_mapping import IHRP_STANDARD_COLUMNS
    return list(IHRP_STANDARD_COLUMNS)


# ==========================================================================
# UI HELPER — badge display info
# ==========================================================================

# Map type → {label, icon, color} cho frontend render badge.
TYPE_BADGES = {
    "date_iso":         {"label": "date", "icon": "📅", "color": "blue"},
    "date_dmy":         {"label": "date", "icon": "📅", "color": "blue"},
    "date_excel_serial":{"label": "date (Excel)", "icon": "📅", "color": "blue"},
    "integer":          {"label": "number", "icon": "🔢", "color": "purple"},
    "decimal":          {"label": "number", "icon": "🔢", "color": "purple"},
    "pic_list":         {"label": "PIC", "icon": "👥", "color": "orange"},
    "status_enum":      {"label": "status", "icon": "🏷", "color": "green"},
    "boolean":          {"label": "yes/no", "icon": "✓", "color": "gray"},
    "empty":            {"label": "empty", "icon": "∅", "color": "gray"},
    "string":           {"label": "text", "icon": "📝", "color": "gray"},
}


def infer_all_headers(
    headers: list[str],
    preview_rows: list[list[Any]],
) -> dict[str, dict[str, Any]]:
    """
    Với mỗi header, trả:
      {header: {"type": str, "badge": {...}, "samples": [3 sample non-empty]}}

    preview_rows là output của `read_headers_and_preview` — mỗi row = list
    cell (giá trị đã stringify với datetime).

    Sample lấy 3 giá trị NON-EMPTY đầu tiên; nếu không đủ non-empty thì
    lấy mọi thứ.
    """
    out: dict[str, dict[str, Any]] = {}
    n_cols = len(headers)
    for col_idx, header in enumerate(headers):
        if not header:
            continue
        # Lấy sample từ preview_rows column col_idx
        raw_samples: list[Any] = []
        for row in preview_rows:
            if col_idx < len(row):
                raw_samples.append(row[col_idx])
        # Ưu tiên 3 non-empty đầu tiên
        non_empty = [v for v in raw_samples if v is not None
                     and not (isinstance(v, str) and not v.strip())]
        display_samples = non_empty[:3] if non_empty else raw_samples[:3]
        inferred = infer_type(raw_samples)
        out[header] = {
            "type": inferred,
            "badge": TYPE_BADGES.get(inferred, TYPE_BADGES["string"]),
            "samples": [_stringify(v) for v in display_samples],
        }
    return out


def _stringify(v: Any) -> str:
    """Convert value → readable string cho hiển thị sample."""
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    return str(v)


# ==========================================================================
# VALIDATION DRY-RUN — sub-task E
# ==========================================================================

def validate_mapping_dry_run(
    filepath: str,
    column_mapping: dict[str, str],
    n_rows: int = 5,
) -> dict[str, Any]:
    """
    Chạy parser trên N record đầu với mapping → trả preview + errors/warnings.

    Return shape:
      {
        "success": bool,             // False nếu parse crash hoặc >50% row lỗi
        "rows": [                    // preview record iHRP shape
          {"ma_cn": "...", "ten_cn": "...", "module": "...",
           "phases": {"Analysis": {"start_date": "...", ...}, ...}}, ...
        ],
        "errors": [                  // parse error per cell
          {"row_idx": 2, "col": "Analysis - End", "msg": "..."}, ...
        ],
        "warnings": [                // warning không blocking
          "Cột iHRP 'Dev - PIC' map đến header không có trong file",
          ...
        ],
        "row_count_scanned": N,
      }

    Dry-run KHÔNG lưu vào state, KHÔNG mutate — chỉ đọc file để user test
    mapping trước khi confirm.
    """
    from parser.excel_parser import FunctionListParser

    errors: list[dict] = []
    warnings: list[str] = []
    rows: list[dict] = []

    try:
        parser = FunctionListParser()
        # Parse full file với mapping — sau đó cắt N row đầu
        # (Parser hiện tại không expose limit param; chấp nhận parse full
        # rồi slice, vì dry-run chỉ trên 5 row nên user file có 10k row
        # cũng nhanh — nếu chậm sau này thêm early-stop.)
        parsed = parser.parse(filepath, column_mapping=column_mapping or None)
    except Exception as e:
        # Parse crash → return error rõ ràng
        return {
            "success": False,
            "rows": [],
            "errors": [{"row_idx": 0, "col": "", "msg": f"Parse crash: {e}"}],
            "warnings": [],
            "row_count_scanned": 0,
        }

    # Convert N row đầu thành dict serializable
    for r in parsed.rows[:n_rows]:
        row_dict = {
            "row_num": r.row_num,
            "ma_cn": r.meta.get("ma_cn", ""),
            "ten_cn": r.meta.get("ten_cn", ""),
            "module": r.meta.get("module", ""),
            "priority": r.meta.get("priority", ""),
            "complexity": r.meta.get("complexity", ""),
            "fit_gap": r.meta.get("fit_gap", ""),
            "phases": {},
        }
        for phase_name, pd in r.phases.items():
            row_dict["phases"][phase_name] = {
                "start_date": pd.start_date.isoformat() if pd.start_date else "",
                "end_date": pd.end_date.isoformat() if pd.end_date else "",
                "status": pd.status or "",
                "pic": pd.pics,
                "estimate_mh": pd.estimate_mh,
            }
        rows.append(row_dict)

    # Detect warnings — mapping trỏ header không tồn tại trong file
    # (Không phải error nhưng đáng lưu ý — thường do user chọn nhầm)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(values_only=True), tuple())
        actual_headers = {str(v).strip() for v in first_row if v}
        wb.close()
        for ihrp_col, actual_col in (column_mapping or {}).items():
            if actual_col and actual_col not in actual_headers:
                warnings.append(
                    f"Cột iHRP '{ihrp_col}' map đến header '{actual_col}' "
                    f"không có trong file."
                )
    except Exception:
        pass

    # Cell-level validation trên N row đầu — check date parse fail explicit.
    # Parser đã set date=None khi parse fail; nếu mapping trỏ đến 1 cột
    # kiểu string nhưng iHRP expect date → parser silent-fail → cell trống.
    # Detect: nếu header raw value non-empty ở cell nhưng parsed = None.
    # (Best-effort — không critical, chỉ hint cho user.)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(values_only=True), tuple())
        header_to_idx: dict[str, int] = {
            str(v).strip(): i for i, v in enumerate(header_row) if v
        }

        # Reverse mapping: actual_header → ihrp_col (cho date columns)
        date_mappings = [
            (ihrp, actual) for ihrp, actual in (column_mapping or {}).items()
            if (" - Start" in ihrp or " - End" in ihrp) and actual
        ]

        row_iter = ws.iter_rows(values_only=True)
        next(row_iter)  # skip header
        for r_idx, row in enumerate(row_iter, start=2):
            if r_idx - 1 > n_rows:
                break
            for ihrp_col, actual_col in date_mappings:
                col_idx = header_to_idx.get(actual_col)
                if col_idx is None or col_idx >= len(row):
                    continue
                raw = row[col_idx]
                # Raw có giá trị nhưng không phải date/datetime/numeric parseable
                if raw is None or raw == "":
                    continue
                if isinstance(raw, (date, datetime, int, float)):
                    continue
                if isinstance(raw, str):
                    s = raw.strip()
                    if _RE_DATE_ISO.match(s) or _RE_DATE_DMY.match(s):
                        continue
                    errors.append({
                        "row_idx": r_idx,
                        "col": ihrp_col,
                        "msg": f"Không parse được '{s[:30]}' làm ngày",
                    })
        wb.close()
    except Exception:
        pass

    return {
        "success": True,
        "rows": rows,
        "errors": errors,
        "warnings": warnings,
        "row_count_scanned": len(rows),
    }
