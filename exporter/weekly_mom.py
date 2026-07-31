"""
Xuất báo cáo tuần MoM (Meeting Minutes) theo mẫu W30.

Sheets:
  1. Cover Page     — mục lục cuộc họp + mã dự án
  2. Master plan    — WBS Module × Phase từ Function List (Start–End, % Closed, PIC)
  3. Gantt          — timeline tuần (ô tô màu theo Start–End)
  4. MoM_W{n}       — biên bản họp tuần (kế hoạch tuần này / tuần tới + overdue)
  5. Risk Analysis  — rủi ro đa chiều (Overdue / Unassigned / Stalled / DQ / Rlog)
  6. PM Dashboard   — số liệu + vài biểu đồ Excel tiêu biểu

Ô thiếu data → để trống / "N/A" + ghi chú, không bịa số.
"""
from __future__ import annotations

import os
from collections import Counter
from datetime import date, datetime, timedelta
from typing import Any, Optional

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from exporter.excel_exporter import (
    BODY_ALIGN,
    GREEN_FILL,
    ORANGE_FILL,
    RED_FILL,
    THIN_BORDER,
    YELLOW_FILL,
    _fill_by_days,
)
from parser.excel_parser import VALID_STATUSES

# Styles MoM — hierarchy mẫu W30:
#   H1 Title   = navy #1F4E79, trắng, 14pt bold, merge full-width
#   H2 Section = vàng #FFC000 (hoặc xanh #0070C0 cho "Nội dung"), trắng, 12pt bold
#   H3 Header  = xanh #0070C0, trắng, 11pt bold, center
MOM_HEADER_FILL = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
MOM_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
MOM_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
TITLE_BAR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_BAR_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
CONTENT_BAR_FILL = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
CONTENT_BAR_FONT = Font(name="Times New Roman", bold=True, size=12, color="FFFFFF")
NOTE_FONT = Font(name="Times New Roman", italic=True, size=10, color="666666")
GUIDE_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
META_LABEL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
META_VALUE_FONT = Font(name="Times New Roman", size=12, color="0000FF")
BODY_MOM_FONT = Font(name="Calibri", size=10)
BODY_MOM_BOLD = Font(name="Calibri", size=10, bold=True)
MOM_BODY_FONT = Font(name="Times New Roman", size=11)
MOM_BODY_BOLD = Font(name="Times New Roman", size=11, bold=True)
MOM_LABEL_FONT = Font(name="Times New Roman", bold=True, size=12)
LINK_FONT = Font(name="Times New Roman", size=11, color="0070C0", underline="single")
MODULE_FILL = PatternFill(start_color="A8D08D", end_color="A8D08D", fill_type="solid")
LU_DATE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BANNER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BANNER_FONT = Font(name="Calibri", bold=True, size=11, color="9C5700")
GANTT_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
GANTT_MOD_BAR = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GANTT_PHASE_BAR = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
GANTT_NOW_BAR = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
GANTT_TODAY_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
GANTT_EMPTY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
STATUS_DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
STATUS_OPEN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
STATUS_WIP_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
KPI_FILL = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
PHASE_HEADER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
MEDIUM_BOTTOM = Border(bottom=Side(style="medium", color="1F4E79"))
NA = "N/A"

# Giới hạn dòng gợi ý trên MoM (tránh flood khi nhiều phase overlap dài)
_WEEK_PLAN_LIMIT = 40


def _iso_week_label(d: date) -> str:
    """VD: W30 (ISO week)."""
    return f"W{d.isocalendar()[1]:02d}"


def _parse_date(v: Any) -> Optional[date]:
    """Chuẩn hóa mọi kiểu date thường gặp về date | None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # ISO có thể dài hơn 10 (datetime)
    for fmt, n in (("%Y-%m-%d", 10), ("%d/%m/%Y", 10), ("%Y/%m/%d", 10), ("%d-%m-%Y", 10)):
        try:
            return datetime.strptime(s[:n], fmt).date()
        except ValueError:
            continue
    return None


def _fmt_date(v: Any) -> str:
    """Format hiển thị dd/MM/yyyy; rỗng nếu không parse được."""
    d = _parse_date(v)
    return d.strftime("%d/%m/%Y") if d else ""


def _normalize_date_pair(
    start: Any, end: Any,
) -> tuple[Optional[date], Optional[date], bool]:
    """
    Parse Start/End; nếu cả hai có và start > end → swap.

    Returns: (start, end, swapped)
    """
    s = _parse_date(start)
    e = _parse_date(end)
    swapped = False
    if s is not None and e is not None and s > e:
        s, e = e, s
        swapped = True
    return s, e, swapped


def _week_range(d: date) -> tuple[date, date]:
    """Thứ 2 → Chủ nhật của tuần ISO chứa d."""
    start = d - timedelta(days=d.weekday())
    return start, start + timedelta(days=6)


def _ranges_overlap(
    a_start: Optional[date],
    a_end: Optional[date],
    b_start: date,
    b_end: date,
) -> bool:
    """True nếu [a_start, a_end] giao [b_start, b_end] (điểm đơn = start=end)."""
    if a_start is None and a_end is None:
        return False
    s = a_start if a_start is not None else a_end
    e = a_end if a_end is not None else a_start
    assert s is not None and e is not None
    return s <= b_end and e >= b_start


def _normalize_status(status: Any) -> str:
    """Trả status chuẩn nếu khớp VALID_STATUSES (case-insensitive); else giữ nguyên trimmed."""
    if status is None:
        return ""
    s = str(status).strip()
    if not s:
        return ""
    # Số thuần → lỗi lệch cột Estimate MH
    if s.isdigit():
        return ""
    for valid in VALID_STATUSES:
        if s.lower() == valid.lower():
            return valid
    return s


def _style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    """H3 — table header xanh #0070C0, chữ trắng, center."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = MOM_HEADER_FONT
        cell.fill = MOM_HEADER_FILL
        cell.alignment = MOM_HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def _apply_border_row(ws, row: int, start_col: int, end_col: int) -> None:
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if cell.font is None or cell.font.name is None:
            cell.font = BODY_MOM_FONT
        if cell.alignment is None or cell.alignment.vertical is None:
            cell.alignment = BODY_ALIGN


def _style_title_bar(ws, row: int, title: str, end_col: int, *, start_col: int = 1) -> None:
    """H1 — thanh tiêu đề navy #1F4E79 + chữ trắng (mọi sheet)."""
    ws.merge_cells(
        start_row=row, start_column=start_col, end_row=row, end_column=end_col,
    )
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = TITLE_BAR_FONT
    cell.fill = TITLE_BAR_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = TITLE_BAR_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def _style_section_bar(
    ws, row: int, title: str, end_col: int, *, start_col: int = 1, letter: str = "",
) -> None:
    """H2 — section vàng #FFC000 (MoM A/B, Risk/Dashboard blocks)."""
    if letter:
        ws.cell(row=row, column=start_col, value=letter)
        ws.cell(row=row, column=start_col + 1, value=title)
    else:
        ws.merge_cells(
            start_row=row, start_column=start_col, end_row=row, end_column=end_col,
        )
        ws.cell(row=row, column=start_col, value=title)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_CENTER
    ws.row_dimensions[row].height = 22


def _set_print_landscape(ws, *, fit_width: bool = True) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1 if fit_width else 0
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _set_print_portrait(ws) -> None:
    """MoM biên bản — portrait như mẫu W30."""
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6


def _sheet_link(ws, cell_ref: str, target_sheet: str, label: str) -> None:
    """Hyperlink nội bộ tới sheet (giống Cover mẫu W30)."""
    cell = ws[cell_ref]
    cell.value = label
    cell.hyperlink = f"#'{target_sheet}'!A1"
    cell.font = LINK_FONT


def _status_fill(status: str) -> Optional[PatternFill]:
    s = (status or "").strip().lower()
    if s in ("closed", "done", "resolved"):
        return STATUS_DONE_FILL
    if s in ("open", "pending", "assigned"):
        return STATUS_OPEN_FILL
    if s in ("in-progress", "in progress"):
        return STATUS_WIP_FILL
    return None


def _pct_fill(pct: float) -> Optional[PatternFill]:
    if pct >= 80:
        return GREEN_FILL
    if pct >= 50:
        return STATUS_WIP_FILL
    if pct > 0:
        return YELLOW_FILL
    return None


def _collect_week_plan(
    parsed_data,
    week_start: date,
    week_end: date,
    *,
    limit: int = _WEEK_PLAN_LIMIT,
) -> list[dict]:
    """
    Công việc gợi ý trong tuần ISO (đủ dùng PM, không flood overlap dài):
      - Phase chưa Closed/Cancelled
      - Sau normalize/swap Start–End: End nằm trong tuần HOẶC Start nằm trong tuần
      - Không lấy phase chỉ "giao tuần" dài (overlap-only) — gây trùng A/B và đầy limit
    Ưu tiên: End trong tuần → Start trong tuần.
    """
    if parsed_data is None:
        return []
    items: list[dict] = []
    for r in parsed_data.rows:
        ma = r.meta.get("ma_cn") or ""
        ten = r.meta.get("ten_cn") or ""
        module = r.meta.get("module") or ""
        for phase_name, pd in (r.phases or {}).items():
            st = _normalize_status(pd.status)
            if st in ("Closed", "Cancelled"):
                continue
            s, e, swapped = _normalize_date_pair(pd.start_date, pd.end_date)
            end_in = e is not None and week_start <= e <= week_end
            start_in = s is not None and week_start <= s <= week_end
            if not end_in and not start_in:
                continue
            priority = 0 if end_in else 1
            label = f"[{phase_name}] {ma} — {ten}".strip(" —")
            if not ma and not ten:
                label = f"[{phase_name}]"
            pics = ", ".join(pd.pics or [])
            note = (("swap Start↔End; " if swapped else "") + (pd.note or "")).strip()
            items.append({
                "ten": label,
                "pic": pics,
                "from": _fmt_date(s),
                "to": _fmt_date(e),
                "status": st,
                "note": note[:80].rstrip("; "),
                "module": module,
                "phase": phase_name,
                "priority": priority,
                "_end": e or date.max,
                "_start": s or date.max,
            })
    items.sort(key=lambda x: (x["priority"], x["_end"], x["module"], x["phase"]))
    out = items[:limit]
    for it in out:
        it.pop("_end", None)
        it.pop("_start", None)
        it.pop("priority", None)
    return out


def _phase_dates_from_rows(
    parsed_data,
    module: str,
    phase: str,
    today: Optional[date] = None,
) -> tuple[Optional[date], Optional[date], int]:
    """
    Min Start / max End sau swap từng function; bỏ outlier năm.
    Returns: (start, end, n_dated).
    """
    from analyzer.gantt_calendar import _is_outlier_date

    starts: list[date] = []
    ends: list[date] = []
    if parsed_data is None:
        return None, None, 0
    for r in parsed_data.rows:
        if (r.meta.get("module") or "") != module:
            continue
        pd = (r.phases or {}).get(phase)
        if not pd:
            continue
        ps, pe, _ = _normalize_date_pair(pd.start_date, pd.end_date)
        if ps and not _is_outlier_date(ps, today):
            starts.append(ps)
        if pe and not _is_outlier_date(pe, today):
            ends.append(pe)
    return (
        min(starts) if starts else None,
        max(ends) if ends else None,
        len(starts) + len(ends),
    )


def _format_phase_status(pct: Any, top_status: str) -> str:
    """Tránh chữ dư kiểu 'Closed · 100.0% Closed'."""
    if pct is not None and pct != "":
        try:
            pct_f = float(pct)
            pct_txt = f"{pct_f:g}% Closed"
        except (TypeError, ValueError):
            pct_txt = f"{pct}% Closed"
        if top_status and top_status not in ("Closed", "—"):
            return f"{top_status} · {pct_txt}"
        return pct_txt
    return top_status or ""


def _build_master_rows(
    metrics: dict,
    parsed_data,
    today: Optional[date] = None,
) -> list[dict]:
    """
    WBS-like: mỗi Module 1 dòng cha + mỗi Phase có date 1 dòng con.
    % / n từ timeline_data; Start–End ưu tiên aggregate từ ParsedData (đã swap).
    """
    timeline = (metrics or {}).get("timeline_data") or {}
    modules = timeline.get("modules") or []
    phases = timeline.get("phases") or []
    data = timeline.get("data") or {}
    if not modules and parsed_data is not None:
        modules = list(parsed_data.all_modules or [])
        phases = list(parsed_data.all_phases or [])

    pic_map: dict[tuple[str, str], list[str]] = {}
    status_map: dict[tuple[str, str], dict[str, int]] = {}
    if parsed_data is not None:
        for r in parsed_data.rows:
            mod = r.meta.get("module") or ""
            for ph, pd in (r.phases or {}).items():
                key = (mod, ph)
                for p in (pd.pics or []):
                    pic_map.setdefault(key, [])
                    if p not in pic_map[key]:
                        pic_map[key].append(p)
                st = _normalize_status(pd.status)
                if st:
                    status_map.setdefault(key, {})
                    status_map[key][st] = status_map[key].get(st, 0) + 1

    rows: list[dict] = []
    stt_mod = 0
    for module in modules:
        mod_data = data.get(module) or {}
        mod_starts: list[date] = []
        mod_ends: list[date] = []
        phase_children: list[dict] = []
        for ph in phases:
            cell = mod_data.get(ph) or {}
            s, e, n_dated = _phase_dates_from_rows(parsed_data, module, ph, today)
            if s is None and e is None:
                s, e, _ = _normalize_date_pair(cell.get("start"), cell.get("end"))
            if s is None and e is None and not cell:
                continue
            if s is None and e is None and not (cell.get("total") or cell.get("pct_closed") is not None):
                continue
            if s:
                mod_starts.append(s)
            if e:
                mod_ends.append(e)
            pct = cell.get("pct_closed")
            closed = cell.get("closed")
            total = cell.get("total")
            if pct is None and total:
                pct = round(100.0 * int(closed or 0) / int(total), 1)
            counts = status_map.get((module, ph)) or {}
            top_status = max(counts.items(), key=lambda kv: kv[1])[0] if counts else ""
            status_txt = _format_phase_status(pct, top_status)
            pics = pic_map.get((module, ph)) or []
            note_bits = []
            if total:
                note_bits.append(f"n={total}")
            elif n_dated:
                note_bits.append(f"dated={n_dated}")
            phase_children.append({
                "level": 1,
                "name": f"    {ph}",
                "status": status_txt,
                "pic_kdg": ", ".join(pics[:4]) if pics else "",
                "start": s,
                "end": e,
                "note": " · ".join(note_bits),
                "pct": pct if pct is not None else "",
            })

        if not phase_children and not mod_starts and not mod_ends:
            continue
        stt_mod += 1
        mod_ov = None
        for m in (metrics or {}).get("module_overview") or []:
            if (m.get("module") or m.get("label")) == module:
                mod_ov = m
                break
        mod_pct = mod_ov.get("progress_pct") if mod_ov else None
        mod_od = mod_ov.get("overdue_count") if mod_ov else None
        active = (mod_ov or {}).get("active_phase") or ""
        note_parts = []
        if mod_ov:
            note_parts.append(f"{mod_ov.get('total', 0)} CN")
        if mod_od:
            note_parts.append(f"overdue={mod_od}")
        if active:
            note_parts.append(f"phase: {active}")
        rows.append({
            "level": 0,
            "stt": str(stt_mod),
            "name": module,
            "status": f"{mod_pct:g}% Closed" if isinstance(mod_pct, (int, float)) else (
                f"{mod_pct}% Closed" if mod_pct is not None else ""
            ),
            "pic_kdg": "",
            "start": min(mod_starts) if mod_starts else None,
            "end": max(mod_ends) if mod_ends else None,
            "note": " · ".join(note_parts),
            "pct": mod_pct if mod_pct is not None else "",
        })
        for i, ch in enumerate(phase_children, 1):
            ch["stt"] = f"{stt_mod}.{i}"
            rows.append(ch)
    return rows


def _write_cover(
    wb,
    project_code: str,
    week_label: str,
    meeting_sheet_name: str,
    today: date,
) -> None:
    """Cover sạch kiểu W30: H1 title + meta + TOC hyperlink."""
    ws = wb.create_sheet("Cover Page", 0)
    widths = {
        "A": 2, "B": 20, "C": 16, "D": 42, "E": 10, "F": 12,
        "G": 30, "H": 8,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.sheet_view.showGridLines = False

    # H1
    _style_title_bar(ws, 2, "MEETING MINUTES — iHRP Tracker", 7, start_col=2)

    ws["B4"] = "Project Code"
    ws["B4"].font = Font(name="Calibri", size=10, color="808080")
    ws["B4"].fill = META_LABEL_FILL
    ws["B4"].alignment = LEFT_CENTER
    ws["B4"].border = THIN_BORDER
    ws["C4"] = project_code or NA
    ws["C4"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws["C4"].alignment = LEFT_CENTER
    ws["C4"].border = THIN_BORDER

    ws["F4"] = "Meeting Minutes Records"
    ws["F4"].font = TITLE_FONT
    ws["F4"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("F4:G4")

    ws["B5"] = "Version"
    ws["B5"].font = Font(name="Calibri", size=10, color="808080")
    ws["B5"].fill = META_LABEL_FILL
    ws["B5"].alignment = LEFT_CENTER
    ws["B5"].border = THIN_BORDER
    ws["C5"] = "v1.4 (auto từ iHRP Tracker)"
    ws["C5"].font = Font(name="Calibri", size=10)
    ws["C5"].alignment = LEFT_CENTER
    ws["C5"].border = THIN_BORDER

    ws["E5"] = "Tuần"
    ws["E5"].font = LABEL_FONT
    ws["E5"].fill = META_LABEL_FILL
    ws["E5"].alignment = CENTER
    ws["E5"].border = THIN_BORDER
    ws["F5"] = week_label
    ws["F5"].font = Font(name="Calibri", bold=True, size=14, color="C00000")
    ws["F5"].alignment = CENTER
    ws["F5"].fill = BANNER_FILL
    ws["F5"].border = THIN_BORDER

    # H2 mục lục
    _style_section_bar(ws, 7, "MỤC LỤC BÁO CÁO", 7, start_col=2)

    headers = ["Sheet Name", "Date", "Review subject"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=8, column=i, value=h)
    _style_header_row(ws, 8, 2, 7)
    ws.merge_cells("D8:G8")

    toc = [
        ("Master plan", "WBS Module × Phase (Start–End + % Closed từ Function List)"),
        ("Gantt", "Timeline tuần quanh ngày xuất (ô tô màu theo kế hoạch)"),
        (meeting_sheet_name, f"Họp định kỳ dự án — {week_label}"),
        ("Risk Analysis", "Rủi ro đa chiều: Overdue / Unassigned / Stalled / High risk / Rlog"),
        ("PM Dashboard", "Tóm tắt metrics + biểu đồ Excel"),
    ]
    for i, (sheet, subject) in enumerate(toc):
        r = 9 + i
        _sheet_link(ws, f"B{r}", sheet, sheet)
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", bold=True, size=10, color="0070C0", underline="single",
        )
        ws.cell(row=r, column=3, value=today.strftime("%d/%m/%Y")).font = BODY_MOM_FONT
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=4, value=subject).font = BODY_MOM_FONT
        ws.cell(row=r, column=4).alignment = LEFT_CENTER
        ws.merge_cells(f"D{r}:G{r}")
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if i % 2 == 1:
                if not cell.fill or cell.fill.fgColor is None or (
                    getattr(cell.fill.fgColor, "rgb", None) in (None, "00000000")
                ):
                    if c != 2:
                        cell.fill = ALT_ROW_FILL
        ws.row_dimensions[r].height = 18

    # Ô trống cho MoM tuần sau (giống mẫu W30)
    for r in range(14, 20):
        ws.merge_cells(f"D{r}:G{r}")
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 16

    ws.merge_cells("B21:G21")
    ws["B21"] = (
        f"Xuất {today.strftime('%d/%m/%Y')} · Master/Gantt/Risk từ Function List · "
        f"MoM gợi ý deadline — điền giờ họp & người tham gia trước khi gửi."
    )
    ws["B21"].font = GUIDE_FONT

    _set_print_landscape(ws)


def _write_master_plan(wb, metrics: dict, parsed_data, today: Optional[date] = None) -> None:
    """Master plan: WBS Module × Phase từ timeline Function List."""
    ws = wb.create_sheet("Master plan")
    widths = {
        "A": 3, "B": 7, "C": 56, "D": 16, "E": 10, "F": 18,
        "G": 14, "H": 14, "I": 14, "J": 14, "K": 26,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False

    # H1 + badge cập nhật
    _style_title_bar(ws, 1, "MASTER PLAN — WBS Module × Phase (Function List)", 8, start_col=2)
    ws["I1"] = "Cập nhật liên tục"
    ws["I1"].font = BANNER_FONT
    ws["I1"].fill = BANNER_FILL
    ws["I1"].alignment = CENTER
    ws.merge_cells("I1:J1")
    ws["I1"].border = THIN_BORDER
    ws["J1"].border = THIN_BORDER
    ws["J1"].fill = BANNER_FILL
    ws["K1"].fill = TITLE_BAR_FILL
    ws["K1"].border = THIN_BORDER

    ws["B2"] = "STT"
    ws["C2"] = "Công việc"
    ws["D2"] = "Tình trạng"
    ws["E2"] = "PIC"
    ws["G2"] = "Kế hoạch v1.0"
    ws["I2"] = "Kế hoạch LU"
    ws["K2"] = "Ghi chú"
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    ws.merge_cells("E2:F2")
    ws.merge_cells("G2:H2")
    ws.merge_cells("I2:J2")
    ws.merge_cells("K2:K3")
    ws["E3"] = "FIS"
    ws["F3"] = "KDG"
    ws["G3"] = "Thời gian bắt đầu"
    ws["H3"] = "Thời gian kết thúc"
    ws["I3"] = "Thời gian bắt đầu"
    ws["J3"] = "Thời gian kết thúc"

    for r in (2, 3):
        for c in range(2, 12):
            cell = ws.cell(row=r, column=c)
            cell.font = MOM_HEADER_FONT
            cell.fill = MOM_HEADER_FILL
            cell.alignment = MOM_HEADER_ALIGN
            cell.border = THIN_BORDER
        ws.row_dimensions[r].height = 22

    master_rows = _build_master_rows(metrics, parsed_data, today=today)
    if not master_rows:
        ws["B4"] = ""
        ws["C4"] = (
            "N/A — chưa có Start/End phase trong Function List để dựng Master plan."
        )
        ws["C4"].font = NOTE_FONT
        ws.merge_cells("C4:K4")
        for r in range(5, 10):
            ws.cell(row=r, column=2, value=r - 4)
            for c in range(2, 12):
                ws.cell(row=r, column=c).border = THIN_BORDER
                ws.cell(row=r, column=c).font = BODY_MOM_FONT
        return

    row = 4
    for item in master_rows:
        ws.cell(row=row, column=2, value=item.get("stt", ""))
        name = (item.get("name") or "").strip()
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=item.get("status", ""))
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value=item.get("pic_kdg", ""))
        ws.cell(row=row, column=7, value="")
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value=_fmt_date(item.get("start")))
        ws.cell(row=row, column=10, value=_fmt_date(item.get("end")))
        ws.cell(row=row, column=11, value=item.get("note", ""))
        is_mod = item.get("level") == 0
        for c in range(2, 12):
            cell = ws.cell(row=row, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_MOM_BOLD if is_mod else BODY_MOM_FONT
            if c == 3:
                cell.alignment = LEFT_CENTER
                if not is_mod:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True, indent=1,
                    )
            elif c in (4, 6, 11):
                cell.alignment = LEFT_CENTER
            else:
                cell.alignment = CENTER
            if is_mod:
                cell.fill = MODULE_FILL
            else:
                cell.fill = WHITE_FILL
                if c in (9, 10) and cell.value:
                    cell.fill = LU_DATE_FILL
        row += 1

    ws.cell(
        row=row + 1, column=2,
        value="Nguồn FL (đã swap From>To). PIC FIS trống · LU = mốc hiện tại · v1.0 = baseline PM điền.",
    )
    ws.cell(row=row + 1, column=2).font = GUIDE_FONT
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=11)
    ws.freeze_panes = "C4"
    ws.print_title_rows = "1:3"
    _set_print_landscape(ws)


def _iter_weeks(start: date, end: date) -> list[tuple[date, date, str]]:
    """Danh sách (week_start, week_end, label Wxx) bao phủ [start, end]."""
    if start > end:
        start, end = end, start
    cur = start - timedelta(days=start.weekday())
    weeks = []
    guard = 0
    while cur <= end and guard < 80:
        we = cur + timedelta(days=6)
        weeks.append((cur, we, _iso_week_label(cur)))
        cur = cur + timedelta(days=7)
        guard += 1
    return weeks


def _write_gantt_sheet(wb, metrics: dict, parsed_data, today: date) -> None:
    """Sheet Gantt: hàng = Module × Phase giao cửa sổ; cột = tuần cố định quanh today."""
    ws = wb.create_sheet("Gantt")
    min_d = today - timedelta(weeks=6)
    max_d = today + timedelta(weeks=14)
    weeks = _iter_weeks(min_d, max_d)
    win_start, win_end = weeks[0][0], weeks[-1][1]

    all_rows = _build_master_rows(metrics, parsed_data, today=today)
    visible: list[dict] = []
    i = 0
    while i < len(all_rows):
        item = all_rows[i]
        if item.get("level") == 0:
            children = []
            j = i + 1
            while j < len(all_rows) and all_rows[j].get("level") == 1:
                ch = all_rows[j]
                s, e, _ = _normalize_date_pair(ch.get("start"), ch.get("end"))
                if _ranges_overlap(s, e, win_start, win_end):
                    children.append(ch)
                j += 1
            s0, e0, _ = _normalize_date_pair(item.get("start"), item.get("end"))
            if children or _ranges_overlap(s0, e0, win_start, win_end):
                visible.append(item)
                visible.extend(children)
            i = j
        else:
            s, e, _ = _normalize_date_pair(item.get("start"), item.get("end"))
            if _ranges_overlap(s, e, win_start, win_end):
                visible.append(item)
            i += 1

    end_col = 4 + max(len(weeks), 1)
    if not visible:
        _style_title_bar(ws, 1, "GANTT", 6)
        ws["A3"] = (
            f"N/A — không có Module/Phase giao cửa sổ "
            f"{win_start.strftime('%d/%m/%Y')}–{win_end.strftime('%d/%m/%Y')}."
        )
        ws["A3"].font = NOTE_FONT
        return

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    for i in range(len(weeks)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 7
    ws.sheet_view.showGridLines = False

    title = (
        f"GANTT — Module × Phase theo tuần "
        f"({win_start.strftime('%d/%m')}–{win_end.strftime('%d/%m/%Y')}, xuất {today.strftime('%d/%m/%Y')})"
    )
    _style_title_bar(ws, 1, title, end_col)

    # Legend rõ ràng
    ws["A2"] = "Chú thích:"
    ws["A2"].font = LABEL_FONT
    legends = [
        (2, "Module", GANTT_MOD_BAR, "FFFFFF"),
        (3, "Phase", GANTT_PHASE_BAR, "FFFFFF"),
        (4, "Tuần này ∩ bar", GANTT_NOW_BAR, "FFFFFF"),
        (5, "Tuần hiện tại", GANTT_TODAY_FILL, "000000"),
    ]
    for col, text, fill, color in legends:
        cell = ws.cell(row=2, column=col, value=text)
        cell.fill = fill
        cell.font = Font(name="Calibri", size=8, bold=True, color=color)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=min(end_col, 12))
    ws.cell(
        row=2, column=6,
        value="Bar tô theo Start–End giao tuần · Cam đậm = bar giao tuần xuất",
    ).font = GUIDE_FONT
    ws.row_dimensions[2].height = 18

    headers = ["STT", "Công việc", "Từ ngày", "Đến ngày"] + [w[2] for w in weeks]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, 1, end_col)

    today_week = _week_range(today)[0]
    for wi, (ws_d, _we_d, _) in enumerate(weeks):
        if ws_d == today_week:
            cell = ws.cell(row=3, column=5 + wi)
            cell.fill = GANTT_TODAY_FILL
            cell.font = Font(name="Calibri", bold=True, size=9, color="000000")

    for ri, item in enumerate(visible):
        row = 4 + ri
        s, e, _ = _normalize_date_pair(item.get("start"), item.get("end"))
        name = (item.get("name") or "").strip()
        vals = [item.get("stt", ""), name, _fmt_date(s), _fmt_date(e)]
        is_mod = item.get("level") == 0
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            cell.font = BODY_MOM_BOLD if is_mod else Font(name="Calibri", size=9)
            if c == 2:
                cell.alignment = LEFT_CENTER if is_mod else Alignment(
                    horizontal="left", vertical="center", wrap_text=True, indent=1,
                )
            else:
                cell.alignment = CENTER
            if is_mod:
                cell.fill = MODULE_FILL
        for wi, (ws_d, we_d, _) in enumerate(weeks):
            cell = ws.cell(row=row, column=5 + wi, value="")
            cell.border = THIN_BORDER
            has_bar = False
            if s is not None or e is not None:
                if _ranges_overlap(s, e, ws_d, we_d):
                    has_bar = True
                    if ws_d == today_week:
                        cell.fill = GANTT_NOW_BAR
                    elif is_mod:
                        cell.fill = GANTT_MOD_BAR
                    else:
                        cell.fill = GANTT_PHASE_BAR
            if not has_bar:
                if ws_d == today_week:
                    cell.fill = GANTT_TODAY_FILL
                elif wi % 2 == 1:
                    cell.fill = GANTT_EMPTY_FILL
        ws.row_dimensions[row].height = 15

    note_row = 4 + len(visible) + 1
    ws.cell(
        row=note_row, column=1,
        value="Chỉ hiện hàng giao cửa sổ ±6…+14 tuần quanh ngày xuất.",
    )
    ws.cell(row=note_row, column=1).font = GUIDE_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    ws.freeze_panes = "E4"
    ws.print_title_rows = "1:3"
    _set_print_landscape(ws, fit_width=False)


def _write_mom_sheet(
    wb,
    sheet_name: str,
    week_label: str,
    today: date,
    metrics: dict,
    parsed_data=None,
) -> None:
    """Biên bản họp tuần — cấu trúc bám sheet mẫu W30 (portrait, TNR)."""
    ws = wb.create_sheet(sheet_name)
    col_widths = {1: 3, 2: 5, 3: 52, 4: 14, 5: 36, 6: 12, 7: 12, 8: 12}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = False

    week_start, week_end = _week_range(today)
    next_start = week_start + timedelta(days=7)
    next_end = week_end + timedelta(days=7)

    # H1
    _style_title_bar(
        ws, 1,
        f"BIÊN BẢN HỌP — {sheet_name}  ({week_start.strftime('%d/%m')}–{week_end.strftime('%d/%m/%Y')})",
        8, start_col=1,
    )

    meta_rows = [
        (2, "Ngày/Date:", today.strftime("%d/%m/%Y"), True),
        (3, "Giờ/Time:", "", False),
        (4, "Tên/Subject:", f"Họp định kỳ dự án — {week_label}", True),
        (5, "Người tham gia/Committee:", "", False),
    ]
    for r, label, value, filled in meta_rows:
        ws.cell(row=r, column=2, value=label).font = MOM_LABEL_FONT
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=4, value=value)
        if r == 5:
            ws.merge_cells(f"D{r}:H{r}")
            ws.row_dimensions[r].height = 40
        elif r == 4:
            ws.merge_cells(f"D{r}:H{r}")
        else:
            ws.merge_cells(f"D{r}:E{r}")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c <= 3:
                cell.fill = META_LABEL_FILL
                cell.alignment = LEFT_CENTER
            else:
                cell.font = META_VALUE_FONT if filled or value else MOM_BODY_FONT
                cell.alignment = LEFT_CENTER if r in (4, 5) else CENTER

    _sheet_link(ws, "F2", "Cover Page", "Return to Cover Page")
    ws["F2"].alignment = CENTER
    ws.merge_cells("F2:H2")

    # H2 nội dung
    ws.merge_cells("B7:H7")
    ws["B7"] = "Nội dung / Content"
    for c in range(2, 9):
        ws.cell(row=7, column=c).fill = CONTENT_BAR_FILL
        ws.cell(row=7, column=c).border = THIN_BORDER
        ws.cell(row=7, column=c).font = CONTENT_BAR_FONT
        ws.cell(row=7, column=c).alignment = LEFT_CENTER
    ws.row_dimensions[7].height = 22

    headers = ["STT", "Công việc", "PIC", "Ghi chú", "Từ ngày", "Đến ngày", "Tình trạng"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=8, column=i, value=h)
    _style_header_row(ws, 8, 1, 8)

    def _write_plan_section(row: int, letter: str, title: str, items: list[dict], empty_hint: str) -> int:
        _style_section_bar(ws, row, title, 8, start_col=2, letter=letter)
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1
        if items:
            for idx, it in enumerate(items, 1):
                vals = [idx, it["ten"], it["pic"], it.get("note", ""), it["from"], it["to"], it["status"]]
                for c, v in enumerate(vals, 2):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.font = MOM_BODY_FONT
                    if c == 2:
                        cell.alignment = CENTER
                    elif c in (6, 7):
                        cell.alignment = CENTER
                    elif c == 8:
                        cell.alignment = CENTER
                        sf = _status_fill(str(v or ""))
                        if sf:
                            cell.fill = sf
                    elif c == 3:
                        cell.alignment = LEFT_TOP
                    else:
                        cell.alignment = LEFT_CENTER
                _apply_border_row(ws, row, 1, 8)
                ten_len = len(str(it.get("ten") or ""))
                ws.row_dimensions[row].height = 28 if ten_len > 55 else 18
                row += 1
        else:
            ws.cell(row=row, column=3, value=empty_hint).font = NOTE_FONT
            _apply_border_row(ws, row, 1, 8)
            row += 1
            for _ in range(2):
                _apply_border_row(ws, row, 1, 8)
                row += 1
        return row

    this_week = _collect_week_plan(parsed_data, week_start, week_end)
    next_week = _collect_week_plan(parsed_data, next_start, next_end)

    row = 9
    row = _write_plan_section(
        row, "A",
        f"KẾ HOẠCH TUẦN NÀY ({week_start.strftime('%d/%m')}–{week_end.strftime('%d/%m/%Y')})",
        this_week,
        "Không có phase Start/End trong tuần này. PM điền tay.",
    )

    row += 1
    row = _write_plan_section(
        row, "B",
        f"KẾ HOẠCH TUẦN TỚI ({next_start.strftime('%d/%m')}–{next_end.strftime('%d/%m/%Y')})",
        next_week,
        "Không có phase Start/End trong tuần tới. PM điền tay.",
    )

    # Hành động — header giống mẫu W30
    row += 1
    ws.cell(row=row, column=2, value="Hành động")
    ws.cell(row=row, column=6, value="PIC")
    ws.cell(row=row, column=7, value="Ngày")
    ws.cell(row=row, column=8, value="Trạng thái")
    ws.merge_cells(f"B{row}:E{row}")
    for c in range(1, 9):
        cell = ws.cell(row=row, column=c)
        cell.font = MOM_HEADER_FONT
        cell.fill = MOM_HEADER_FILL
        cell.alignment = MOM_HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 20

    overdue = sorted(
        list(metrics.get("overdue_list") or []),
        key=lambda x: (-int(x.get("days_overdue") or 0), x.get("end_date") or ""),
    )[:15]
    row += 1
    if overdue:
        for idx, it in enumerate(overdue, 1):
            ten = (
                f"Xử lý trễ [{it.get('phase', '')}] "
                f"{it.get('ma_cn') or ''} — {it.get('ten_cn') or ''}"
            ).strip(" —")
            pic = ", ".join(it.get("pic") or []) or NA
            status = _normalize_status(it.get("status")) or "Open"
            ws.cell(row=row, column=2, value=idx).alignment = CENTER
            ws.cell(row=row, column=3, value=ten)
            ws.merge_cells(f"C{row}:E{row}")
            ws.cell(row=row, column=6, value=pic)
            ws.cell(row=row, column=7, value=_fmt_date(it.get("end_date")))
            ws.cell(row=row, column=8, value=status)
            fill = _fill_by_days(int(it.get("days_overdue") or 0))
            for c in range(1, 9):
                cell = ws.cell(row=row, column=c)
                cell.border = THIN_BORDER
                cell.font = MOM_BODY_FONT
                cell.alignment = CENTER if c >= 6 else LEFT_TOP
                if fill and c >= 2:
                    cell.fill = fill
            sf = _status_fill(status)
            if sf:
                ws.cell(row=row, column=8).fill = sf
            ws.row_dimensions[row].height = 28 if len(ten) > 50 else 18
            row += 1
    else:
        ws.cell(row=row, column=3, value="Không có task overdue.").font = NOTE_FONT
        ws.merge_cells(f"C{row}:E{row}")
        _apply_border_row(ws, row, 1, 8)
        row += 1
        for _ in range(3):
            ws.merge_cells(f"C{row}:E{row}")
            _apply_border_row(ws, row, 1, 8)
            row += 1

    row += 1
    ws.merge_cells(f"B{row}:H{row}")
    ws.cell(
        row=row, column=2,
        value=f"Gợi ý auto {today.strftime('%d/%m/%Y')} · Điền giờ họp / người tham gia trước khi gửi.",
    ).font = GUIDE_FONT
    ws.freeze_panes = "A9"
    ws.print_title_rows = "7:8"
    _set_print_portrait(ws)


def _block_title(ws, row: int, title: str, n_cols: int = 8) -> int:
    """H2 section cho Risk / PM Dashboard."""
    _style_section_bar(ws, row, title, n_cols, start_col=1)
    return row + 1


def _write_pm_dashboard(wb, metrics: dict, today: date, week_label: str) -> None:
    """
    PM Dashboard — bảng trái (A–G), chart phải (cột J+) không tạo khoảng trống dọc lớn.
    """
    ws = wb.create_sheet("PM Dashboard")
    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 11
    ws.sheet_view.showGridLines = False

    _style_title_bar(
        ws, 1,
        f"BẢNG ĐIỀU KHIỂN PM — Function List ({week_label})",
        12,
    )

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Xuất ngày {today.strftime('%d/%m/%Y')} — số liệu từ metrics engine"
    ws["A2"].font = GUIDE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    summary = metrics.get("summary") or {}
    row = 4

    # --- 1. Summary KPI ngang ---
    row = _block_title(ws, row, "1. TÓM TẮT DỰ ÁN", n_cols=8)
    cards = [
        ("Tổng CN", summary.get("total_functions", NA)),
        ("Tiến độ %", summary.get("overall_progress_pct", NA)),
        ("CN trễ", summary.get("total_overdue", NA)),
        ("Phase trễ", summary.get("total_overdue_records", NA)),
        ("Thiếu PIC", summary.get("unassigned_count", NA)),
        ("Risk ≥50", summary.get("high_risk_count", NA)),
        ("Module", summary.get("modules_count", NA)),
        ("Phase", summary.get("phases_count", NA)),
    ]
    for i, (label, val) in enumerate(cards):
        c = i + 1
        h = ws.cell(row=row, column=c, value=label)
        h.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        h.fill = MOM_HEADER_FILL
        h.alignment = CENTER
        h.border = THIN_BORDER
        v = ws.cell(row=row + 1, column=c, value=val if val is not None else NA)
        v.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
        v.fill = KPI_FILL
        v.alignment = CENTER
        v.border = THIN_BORDER
        if isinstance(val, float):
            v.number_format = "0.0"
        if label in ("CN trễ", "Phase trễ", "Risk ≥50") and isinstance(val, (int, float)) and val > 0:
            v.fill = RED_FILL
    # Giữ cột A/B label-value cũ cho test đọc "Chỉ số"? Tests check A column text TÓM TẮT etc.
    row += 3

    # --- 2. Overdue ---
    row = _block_title(ws, row, "2. CÔNG VIỆC TRỄ (top 20 — sắp theo số ngày trễ ↓)", n_cols=9)
    overdue = sorted(
        list(metrics.get("overdue_list") or []),
        key=lambda x: (-int(x.get("days_overdue") or 0), x.get("module") or "", x.get("ma_cn") or ""),
    )[:20]
    ov_headers = ["STT", "Mã CN", "Tên chức năng", "Module", "Phase", "Deadline", "Ngày trễ", "Status", "PIC"]
    for i, h in enumerate(ov_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 1, 9)
    row += 1
    if not overdue:
        ws.cell(row=row, column=1, value="Không có task overdue.").font = NOTE_FONT
        row += 1
    else:
        for idx, it in enumerate(overdue, 1):
            vals = [
                idx,
                it.get("ma_cn", ""),
                it.get("ten_cn", ""),
                it.get("module", ""),
                it.get("phase", ""),
                _fmt_date(it.get("end_date")),
                it.get("days_overdue", 0),
                _normalize_status(it.get("status")),
                ", ".join(it.get("pic") or []),
            ]
            fill = _fill_by_days(int(it.get("days_overdue") or 0))
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY_MOM_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT_TOP if c == 3 else CENTER if c in (1, 6, 7) else LEFT_CENTER
                if fill:
                    cell.fill = fill
            ten = str(it.get("ten_cn") or "")
            ws.row_dimensions[row].height = 26 if len(ten) > 40 else 16
            row += 1

    row += 1
    # --- 3. Phase progress + chart bên phải ---
    row = _block_title(ws, row, "3. TIẾN ĐỘ THEO PHASE (% Closed trên status đã fill)", n_cols=4)
    stacked = metrics.get("phase_progress_stacked") or {}
    phases = list(stacked.get("phases") or [])
    statuses = list(stacked.get("statuses") or [])
    pdata = stacked.get("data") or {}
    phase_chart_start = None
    phase_chart_end = None
    if not phases:
        ws.cell(row=row, column=1, value="Chưa có dữ liệu phase.").font = NOTE_FONT
        row += 1
    else:
        hdr = ["Phase", "% Closed", "Closed", "Tổng có status"]
        for i, h in enumerate(hdr, 1):
            ws.cell(row=row, column=i, value=h)
        _style_header_row(ws, row, 1, 4)
        row += 1
        phase_chart_start = row
        status_keys = [s for s in statuses if s != "(Blank)"]
        phase_rows_tmp = []
        for ph in phases:
            counts = pdata.get(ph) or {}
            closed = int(counts.get("Closed", 0) or 0)
            total = sum(int(counts.get(s, 0) or 0) for s in status_keys)
            pct = round(closed / total * 100, 1) if total > 0 else 0.0
            phase_rows_tmp.append((ph, pct, closed, total))
        for ph, pct, closed, total in phase_rows_tmp:
            vals = [ph, pct, closed, total]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY_MOM_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER if c > 1 else LEFT_CENTER
                if c == 2:
                    cell.number_format = "0.0"
            pf = _pct_fill(pct) if total > 0 else None
            if pf:
                ws.cell(row=row, column=1).fill = pf
            row += 1
        phase_chart_end = row - 1

        if phase_chart_start and phase_chart_end >= phase_chart_start:
            chart = BarChart()
            chart.type = "col"
            chart.title = "% Closed theo Phase"
            chart.y_axis.title = "%"
            data_ref = Reference(ws, min_col=2, min_row=phase_chart_start - 1, max_row=phase_chart_end)
            cats = Reference(ws, min_col=1, min_row=phase_chart_start, max_row=phase_chart_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "J" + str(phase_chart_start - 1))

    row += 1
    # --- 4. Module ---
    row = _block_title(ws, row, "4. TỔNG QUAN MODULE (sắp theo số trễ ↓, rồi tiến độ %)", n_cols=6)
    modules = list(metrics.get("module_overview") or [])
    modules.sort(
        key=lambda m: (-int(m.get("overdue_count") or 0), float(m.get("progress_pct") or 0), m.get("module") or ""),
    )
    mo_headers = ["STT", "Module", "Tổng CN", "Tiến độ %", "Phase active", "Overdue"]
    for i, h in enumerate(mo_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 1, 6)
    row += 1
    mod_chart_start = None
    mod_chart_end = None
    if not modules:
        ws.cell(row=row, column=1, value="Chưa có module_overview.").font = NOTE_FONT
        row += 1
    else:
        mod_chart_start = row
        for idx, m in enumerate(modules, 1):
            vals = [
                idx,
                m.get("module") or m.get("label") or "",
                m.get("total", 0),
                m.get("progress_pct", 0),
                m.get("active_phase", ""),
                m.get("overdue_count", 0),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY_MOM_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER if c != 2 and c != 5 else LEFT_CENTER
                if c == 4 and isinstance(v, (int, float)):
                    cell.number_format = "0.0"
            od = int(m.get("overdue_count") or 0)
            if od > 0:
                ws.cell(row=row, column=6).fill = ORANGE_FILL if od < 5 else RED_FILL
            row += 1
        mod_chart_end = row - 1
        chart_n = mod_chart_end - mod_chart_start + 1
        if chart_n > 0:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Tiến độ % theo Module"
            data_ref = Reference(
                ws, min_col=4, min_row=mod_chart_start - 1,
                max_row=mod_chart_end,
            )
            cats = Reference(
                ws, min_col=2, min_row=mod_chart_start,
                max_row=mod_chart_end,
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 12
            chart.height = max(6, min(10, chart_n * 0.6 + 3))
            ws.add_chart(chart, "J" + str(mod_chart_start - 1))

    row += 1
    # --- 5. PIC ---
    row = _block_title(ws, row, "5. TẢI VIỆC PIC (top 25 — sắp theo tổng task ↓)", n_cols=7)
    pics = sorted(
        list(metrics.get("pic_workload") or []),
        key=lambda p: (-int(p.get("total_tasks") or 0), -int(p.get("overdue") or 0), p.get("pic") or ""),
    )[:25]
    pic_headers = ["STT", "PIC", "Tổng", "Closed", "In-progress", "Assigned", "Overdue"]
    for i, h in enumerate(pic_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 1, 7)
    row += 1
    pic_chart_start = None
    if not pics:
        ws.cell(row=row, column=1, value="Chưa có pic_workload.").font = NOTE_FONT
        row += 1
    else:
        pic_chart_start = row
        for idx, p in enumerate(pics, 1):
            vals = [
                idx,
                p.get("pic", ""),
                p.get("total_tasks", 0),
                p.get("closed", 0),
                p.get("in_progress", 0),
                p.get("assigned", 0),
                p.get("overdue", 0),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY_MOM_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER if c != 2 else LEFT_CENTER
            if int(p.get("overdue") or 0) > 0:
                ws.cell(row=row, column=7).fill = RED_FILL
            row += 1
        pic_chart_end = row - 1
        top_n = min(10, pic_chart_end - pic_chart_start + 1)
        if top_n > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Top PIC theo tổng task"
            data_ref = Reference(
                ws, min_col=3, min_row=pic_chart_start - 1,
                max_row=pic_chart_start + top_n - 1,
            )
            cats = Reference(
                ws, min_col=2, min_row=pic_chart_start,
                max_row=pic_chart_start + top_n - 1,
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "J" + str(pic_chart_start - 1))

    # Pie — data A/B, chart cột J (không đè bảng)
    pct = summary.get("overall_progress_pct")
    if isinstance(pct, (int, float)):
        pie_row = row + 1
        ws.cell(row=pie_row, column=1, value="Phân bổ tiến độ chung").font = Font(
            name="Calibri", bold=True, size=11, color="1F4E79",
        )
        closed_pct = round(float(pct), 1)
        remain_pct = round(max(0.0, 100.0 - closed_pct), 1)
        ws.cell(row=pie_row + 1, column=1, value="Đã Closed (ước lượng %)")
        ws.cell(row=pie_row + 1, column=2, value=closed_pct)
        ws.cell(row=pie_row + 2, column=1, value="Còn lại")
        ws.cell(row=pie_row + 2, column=2, value=remain_pct)
        for r in (pie_row + 1, pie_row + 2):
            for c in (1, 2):
                ws.cell(row=r, column=c).border = THIN_BORDER
                ws.cell(row=r, column=c).font = BODY_MOM_FONT
            ws.cell(row=r, column=2).number_format = "0.0"
        pie = PieChart()
        pie.title = "Tiến độ chung (%)"
        labels = Reference(ws, min_col=1, min_row=pie_row + 1, max_row=pie_row + 2)
        data = Reference(ws, min_col=2, min_row=pie_row + 1, max_row=pie_row + 2)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.width = 10
        pie.height = 8
        ws.add_chart(pie, "J" + str(pie_row))
        row = pie_row + 4

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(
        row=row, column=1,
        value="Biểu đồ đặt cột J trở đi — không đè bảng số liệu.",
    ).font = GUIDE_FONT

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:2"
    _set_print_landscape(ws)


# ---------------------------------------------------------------------------
# Risk Analysis — đa chiều từ metrics sẵn có
# ---------------------------------------------------------------------------

_RISK_SEV_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
_RISK_SEV_FILL = {
    "Critical": RED_FILL,
    "High": ORANGE_FILL,
    "Medium": YELLOW_FILL,
    "Low": GREEN_FILL,
}


def _risk_severity_overdue(days: int) -> str:
    if days >= 14:
        return "Critical"
    if days >= 7:
        return "High"
    return "Medium"


def _pics_join(pic) -> str:
    if isinstance(pic, (list, tuple)):
        return ", ".join(str(p) for p in pic if p)
    return str(pic or "")


def _collect_risk_items(metrics: dict) -> list[dict]:
    """
    Gom risk đa chiều từ metrics (không bịa):
      Overdue, Unassigned, Stalled, High risk score, Missing deadline (summary),
      Rlog plan thiếu PIC, Rlog coded thấp vs plan.
    """
    items: list[dict] = []
    seen: set[tuple] = set()

    def _add(it: dict) -> None:
        key = (
            it.get("risk_type"),
            it.get("ma_cn"),
            it.get("phase"),
            it.get("detail", "")[:40],
        )
        if key in seen:
            return
        seen.add(key)
        items.append(it)

    for it in metrics.get("overdue_list") or []:
        days = int(it.get("days_overdue") or 0)
        sev = _risk_severity_overdue(days)
        _add({
            "severity": sev,
            "risk_type": "Overdue",
            "ma_cn": it.get("ma_cn") or "",
            "rlog_id": it.get("rlog_id") or "",
            "ten_cn": it.get("ten_cn") or "",
            "module": it.get("module") or "",
            "phase": it.get("phase") or "",
            "pic": _pics_join(it.get("pic")),
            "detail": f"Trễ {days} ngày · deadline {_fmt_date(it.get('end_date'))} · {it.get('status') or ''}",
            "suggestion": "Ưu tiên Closed hoặc cập nhật End/Status; escalate PIC nếu >7 ngày.",
            "_sort": (_RISK_SEV_ORDER[sev], -days),
        })

    for it in metrics.get("unassigned_tasks") or []:
        days = int(it.get("days_overdue") or 0)
        sev = "High" if it.get("is_overdue") or days > 0 else "Medium"
        _add({
            "severity": sev,
            "risk_type": "Unassigned",
            "ma_cn": it.get("ma_cn") or "",
            "rlog_id": it.get("rlog_id") or "",
            "ten_cn": it.get("ten_cn") or "",
            "module": it.get("module") or "",
            "phase": it.get("phase") or "",
            "pic": "",
            "detail": f"Thiếu PIC · status {it.get('status') or ''} · phase trước đã Closed",
            "suggestion": "Gán PIC ngay cho phase đang tới lượt.",
            "_sort": (_RISK_SEV_ORDER[sev], -days),
        })

    stalled = metrics.get("stalled_tasks") or {}
    for it in (stalled.get("items") if isinstance(stalled, dict) else None) or []:
        wait = int(it.get("wait_days") or 0)
        if wait < 3:
            continue  # nhiễu ngắn
        sev = "High" if wait >= 7 else "Medium"
        _add({
            "severity": sev,
            "risk_type": "Stalled",
            "ma_cn": it.get("ma_cn") or "",
            "rlog_id": "",
            "ten_cn": it.get("ten_cn") or "",
            "module": it.get("module") or "",
            "phase": f"{it.get('completed_phase') or ''}→{it.get('waiting_phase') or ''}",
            "pic": "",
            "detail": f"Đình trệ {wait} ngày sau Closed {it.get('completed_phase') or ''}",
            "suggestion": f"Khởi động phase {it.get('waiting_phase') or 'tiếp theo'} hoặc cập nhật status.",
            "_sort": (_RISK_SEV_ORDER[sev], -wait),
        })

    for it in metrics.get("risk_scores") or []:
        score = int(it.get("risk_score") or 0)
        if score < 50:
            continue
        sev = "Critical" if score >= 70 else "High"
        factors = it.get("risk_factors") or []
        factor_txt = ", ".join(str(f) for f in factors[:4]) if factors else "composite score"
        _add({
            "severity": sev,
            "risk_type": "High risk score",
            "ma_cn": it.get("ma_cn") or "",
            "rlog_id": "",
            "ten_cn": it.get("ten_cn") or "",
            "module": it.get("module") or "",
            "phase": "",
            "pic": "",
            "detail": f"Score {score} · {factor_txt}",
            "suggestion": "Review factors (overdue/duration/unassigned) và lập action.",
            "_sort": (_RISK_SEV_ORDER[sev], -score),
        })

    # DQ: ưu tiên list issues (high + missing_deadline); fallback summary
    dq = metrics.get("data_quality") or {}
    dq_issues = list(dq.get("issues") or []) if isinstance(dq, dict) else []
    md_from_list = 0
    high_from_list = 0
    for it in dq_issues:
        code = (it.get("code") or "").strip()
        sev_raw = (it.get("severity") or "").strip().lower()
        label = (it.get("label") or code or "DQ").strip()
        if sev_raw == "high" and high_from_list < 25:
            high_from_list += 1
            _add({
                "severity": "High",
                "risk_type": f"DQ high: {label}",
                "ma_cn": it.get("ma_cn") or "",
                "rlog_id": it.get("rlog_id") or "",
                "ten_cn": it.get("ten_cn") or "",
                "module": it.get("module") or "",
                "phase": it.get("phase") or "",
                "pic": "",
                "detail": it.get("detail") or label,
                "suggestion": it.get("suggestion") or "Sửa dữ liệu Function List.",
                "_sort": (_RISK_SEV_ORDER["High"], -high_from_list),
            })
        elif code == "missing_deadline" and md_from_list < 20:
            md_from_list += 1
            _add({
                "severity": "Medium",
                "risk_type": "Missing deadline",
                "ma_cn": it.get("ma_cn") or "",
                "rlog_id": it.get("rlog_id") or "",
                "ten_cn": it.get("ten_cn") or "",
                "module": it.get("module") or "",
                "phase": it.get("phase") or "",
                "pic": "",
                "detail": it.get("detail") or "Phase đang làm thiếu End date",
                "suggestion": it.get("suggestion") or "Bổ sung End date cho phase đang mở.",
                "_sort": (_RISK_SEV_ORDER["Medium"], 0),
            })

    summary = metrics.get("summary") or {}
    md_count = int(summary.get("missing_deadline_count") or 0)
    md_rec = int(summary.get("missing_deadline_records") or 0)
    if md_from_list == 0 and (md_count > 0 or md_rec > 0):
        sev = "High" if md_count >= 10 else "Medium"
        _add({
            "severity": sev,
            "risk_type": "Missing deadline",
            "ma_cn": "",
            "rlog_id": "",
            "ten_cn": f"{md_count} function · {md_rec} phase thiếu End",
            "module": "(tóm tắt)",
            "phase": "",
            "pic": "",
            "detail": "DQ: phase active thiếu deadline (End date)",
            "suggestion": "Bổ sung End date cho phase đang mở để theo dõi đúng.",
            "_sort": (_RISK_SEV_ORDER[sev], -md_count),
        })

    anomaly_count = int(summary.get("anomaly_count") or 0)
    anomaly_rec = int(summary.get("anomaly_records") or 0)
    if high_from_list == 0 and (anomaly_count > 0 or anomaly_rec > 0):
        sev = "High" if anomaly_count >= 5 else "Medium"
        _add({
            "severity": sev,
            "risk_type": "DQ high / anomaly",
            "ma_cn": "",
            "rlog_id": "",
            "ten_cn": f"{anomaly_count} function · {anomaly_rec} bản ghi bất thường",
            "module": "(tóm tắt)",
            "phase": "",
            "pic": "",
            "detail": "DQ: end<start / trùng Mã CN / overlap / estimate lệch…",
            "suggestion": "Mở Data Quality trên dashboard, sửa các issue severity high.",
            "_sort": (_RISK_SEV_ORDER[sev], -anomaly_count),
        })

    rlog = metrics.get("rlog_weekly") or {}
    plan = (rlog.get("rlog_plan_next_week") or {}) if isinstance(rlog, dict) else {}
    coded = (rlog.get("rlog_coded_this_week") or {}) if isinstance(rlog, dict) else {}
    plan_items = list(plan.get("items") or [])
    plan_count = int(plan.get("count") or len(plan_items))
    coded_count = int(coded.get("count") or 0)

    no_pic_plan = 0
    for it in plan_items:
        pics = it.get("pic") or []
        if not pics:
            no_pic_plan += 1
            _add({
                "severity": "Medium",
                "risk_type": "Rlog plan thiếu PIC",
                "ma_cn": it.get("ma_cn") or "",
                "rlog_id": it.get("rlog_id") or "",
                "ten_cn": it.get("ten_cn") or "",
                "module": it.get("module") or "",
                "phase": it.get("phase") or "Dev",
                "pic": "",
                "detail": "Kế hoạch Dev tuần tới chưa có PIC",
                "suggestion": "Gán PIC Dev trước khi bắt đầu tuần tới.",
                "_sort": (_RISK_SEV_ORDER["Medium"], 0),
            })

    if plan_count > 0 and coded_count * 2 < plan_count:
        sev = "High" if coded_count == 0 and plan_count >= 3 else "Medium"
        _add({
            "severity": sev,
            "risk_type": "Rlog coded thấp",
            "ma_cn": "",
            "rlog_id": "",
            "ten_cn": f"Coded tuần này {coded_count} · Plan tuần tới {plan_count}",
            "module": "(tóm tắt)",
            "phase": "Dev",
            "pic": "",
            "detail": f"Coded/Plan = {coded_count}/{plan_count}"
                      + (f" · plan thiếu PIC: {no_pic_plan}" if no_pic_plan else ""),
            "suggestion": "Rà soát capacity Dev tuần này vs kế hoạch tuần tới.",
            "_sort": (_RISK_SEV_ORDER[sev], -(plan_count - coded_count)),
        })

    items.sort(key=lambda x: x.get("_sort", (9, 0)))
    for it in items:
        it.pop("_sort", None)
    return items


def _pivot_count(items: list[dict], key_a: str, key_b: str) -> list[tuple]:
    """Đếm (a, b) → count, sort giảm dần."""
    c: Counter = Counter()
    for it in items:
        a = (it.get(key_a) or "(blank)").strip() or "(blank)"
        b = (it.get(key_b) or "(blank)").strip() or "(blank)"
        if a == "(tóm tắt)":
            continue
        c[(a, b)] += 1
    return sorted(c.items(), key=lambda x: (-x[1], x[0][0], x[0][1]))


def _enrich_metrics_for_risk(metrics: dict, parsed_data, today: date) -> dict:
    """Gắn data_quality vào metrics nếu chưa có (để list DQ high / missing deadline)."""
    m = dict(metrics or {})
    if m.get("data_quality") or parsed_data is None:
        return m
    try:
        from analyzer.data_quality import compute_data_quality
        m["data_quality"] = compute_data_quality(parsed_data, today=today)
    except Exception:
        pass
    return m


def _write_risk_sheet(
    wb, metrics: dict, today: date, week_label: str, parsed_data=None,
) -> None:
    """Sheet Risk Analysis — summary đa chiều + top risks."""
    ws = wb.create_sheet("Risk Analysis")
    widths = {
        "A": 12, "B": 20, "C": 16, "D": 14, "E": 12, "F": 32,
        "G": 12, "H": 16, "I": 38, "J": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False

    _style_title_bar(
        ws, 1,
        f"PHÂN TÍCH RỦI RO ĐA CHIỀU — {week_label} (xuất {today.strftime('%d/%m/%Y')})",
        10,
    )
    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "Nguồn: overdue · unassigned (pred Closed + Start đã đến) · stalled · "
        "risk_scores ≥50 · DQ high / missing deadline · rlog plan thiếu PIC"
    )
    ws["A2"].font = GUIDE_FONT

    enriched = _enrich_metrics_for_risk(metrics or {}, parsed_data, today)
    items = _collect_risk_items(enriched)
    row = 4

    # KPI theo loại
    row = _block_title(ws, row, "1. TÓM TẮT THEO LOẠI RỦI RO", n_cols=4)
    by_type = Counter(it["risk_type"] for it in items)
    by_sev = Counter(it["severity"] for it in items)
    ws.cell(row=row, column=1, value="Loại risk")
    ws.cell(row=row, column=2, value="Số mục")
    ws.cell(row=row, column=3, value="Mức độ")
    ws.cell(row=row, column=4, value="Số mục")
    _style_header_row(ws, row, 1, 4)
    row += 1
    type_rows = sorted(by_type.items(), key=lambda x: -x[1])
    sev_rows = sorted(by_sev.items(), key=lambda x: _RISK_SEV_ORDER.get(x[0], 9))
    max_r = max(len(type_rows), len(sev_rows), 1)
    for i in range(max_r):
        if i < len(type_rows):
            ws.cell(row=row, column=1, value=type_rows[i][0])
            ws.cell(row=row, column=2, value=type_rows[i][1])
        if i < len(sev_rows):
            ws.cell(row=row, column=3, value=sev_rows[i][0])
            ws.cell(row=row, column=4, value=sev_rows[i][1])
            sf = _RISK_SEV_FILL.get(sev_rows[i][0])
            if sf:
                ws.cell(row=row, column=3).fill = sf
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).font = BODY_MOM_FONT
            ws.cell(row=row, column=c).alignment = CENTER if c in (2, 4) else LEFT_CENTER
        row += 1
    if not items:
        ws.cell(row=row, column=1, value="Không phát hiện risk từ metrics hiện tại.").font = NOTE_FONT
        row += 1

    row += 1
    # Module × loại
    row = _block_title(ws, row, "2. MODULE × LOẠI RISK (top 20)", n_cols=3)
    for i, h in enumerate(["Module", "Loại risk", "Số mục"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 1, 3)
    row += 1
    piv = _pivot_count(items, "module", "risk_type")[:20]
    if not piv:
        ws.cell(row=row, column=1, value="—").font = NOTE_FONT
        row += 1
    else:
        for (mod, rtype), cnt in piv:
            ws.cell(row=row, column=1, value=mod)
            ws.cell(row=row, column=2, value=rtype)
            ws.cell(row=row, column=3, value=cnt)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).font = BODY_MOM_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            row += 1

    row += 1
    row = _block_title(ws, row, "3. PHASE × LOẠI RISK (top 20)", n_cols=3)
    for i, h in enumerate(["Phase", "Loại risk", "Số mục"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 1, 3)
    row += 1
    piv = _pivot_count(items, "phase", "risk_type")[:20]
    if not piv:
        ws.cell(row=row, column=1, value="—").font = NOTE_FONT
        row += 1
    else:
        for (ph, rtype), cnt in piv:
            ws.cell(row=row, column=1, value=ph)
            ws.cell(row=row, column=2, value=rtype)
            ws.cell(row=row, column=3, value=cnt)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).font = BODY_MOM_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            row += 1

    row += 1
    row = _block_title(ws, row, "4. PIC × OVERDUE (top 15)", n_cols=2)
    ws.cell(row=row, column=1, value="PIC")
    ws.cell(row=row, column=2, value="Overdue count")
    _style_header_row(ws, row, 1, 2)
    row += 1
    pic_od = Counter()
    for it in items:
        if it.get("risk_type") != "Overdue":
            continue
        pic = (it.get("pic") or "").strip() or "(chưa gán)"
        for p in [x.strip() for x in pic.replace(";", ",").split(",") if x.strip()]:
            pic_od[p] += 1
    if not pic_od:
        ws.cell(row=row, column=1, value="Không có overdue gắn PIC.").font = NOTE_FONT
        row += 1
    else:
        for pic, cnt in sorted(pic_od.items(), key=lambda x: -x[1])[:15]:
            ws.cell(row=row, column=1, value=pic)
            ws.cell(row=row, column=2, value=cnt)
            for c in range(1, 3):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).font = BODY_MOM_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            if cnt >= 3:
                ws.cell(row=row, column=2).fill = ORANGE_FILL
            row += 1

    row += 1
    row = _block_title(ws, row, "5. TOP RISKS (severityity ↓, tối đa 60)", n_cols=10)
    headers = [
        "STT", "Mức", "Loại risk", "Mã CN", "RlogID", "Tên chức năng",
        "Module", "Phase", "Chi tiết", "Gợi ý",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 1, 10)
    row += 1

    top = items[:60]
    if not top:
        ws.cell(row=row, column=1, value="Không có mục risk.").font = NOTE_FONT
        row += 1
    else:
        for idx, it in enumerate(top, 1):
            vals = [
                idx,
                it.get("severity", ""),
                it.get("risk_type", ""),
                it.get("ma_cn", ""),
                it.get("rlog_id", ""),
                it.get("ten_cn", ""),
                it.get("module", ""),
                it.get("phase", ""),
                it.get("detail", ""),
                it.get("suggestion", ""),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY_MOM_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER if c in (1, 2, 4, 5, 7) else LEFT_TOP
            sev = it.get("severity") or ""
            sf = _RISK_SEV_FILL.get(sev)
            if sf:
                ws.cell(row=row, column=2).fill = sf
            if idx % 2 == 0:
                for c in (3, 6, 8, 9, 10):
                    if not ws.cell(row=row, column=c).fill or ws.cell(row=row, column=c).fill.fgColor is None:
                        ws.cell(row=row, column=c).fill = ALT_ROW_FILL
            ws.row_dimensions[row].height = 28 if len(str(it.get("ten_cn") or "")) > 40 else 18
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(
        row=row, column=1,
        value="Gợi ý mang tính hỗ trợ PM — xác nhận lại trên Function List trước khi gửi khách.",
    ).font = GUIDE_FONT

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:2"
    _set_print_landscape(ws)


def export_weekly_mom(
    metrics: dict,
    output_dir: str = "uploads",
    *,
    project_code: str = "",
    parsed_data=None,
    today: Optional[date] = None,
    pm_plan: Optional[dict] = None,
) -> str:
    """
    Sinh workbook báo cáo tuần MoM (mẫu W30) + Master/Gantt + Risk + PM Dashboard.

    Args:
        metrics: dict từ DashboardEngine.compute()
        output_dir: thư mục lưu file
        project_code: mã dự án (slug hoặc tên) — hiện trên Cover
        parsed_data: ParsedData optional — kế hoạch tuần / PIC / Master plan
        today: ngày báo cáo (test có thể cố định)
        pm_plan: dict kế hoạch chiều PM (KeHoachDuAn) — nếu có, thêm sheet PM Lịch trình

    Returns:
        Đường dẫn file .xlsx đã lưu.
    """
    today = today or date.today()
    week_label = _iso_week_label(today)
    meeting_sheet = f"MoM_{week_label}"
    code = (project_code or "").strip() or NA

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_cover(wb, code, week_label, meeting_sheet, today)
    _write_master_plan(wb, metrics or {}, parsed_data, today=today)
    _write_gantt_sheet(wb, metrics or {}, parsed_data, today)
    _write_mom_sheet(wb, meeting_sheet, week_label, today, metrics or {}, parsed_data)
    _write_risk_sheet(wb, metrics or {}, today, week_label, parsed_data=parsed_data)
    _write_pm_dashboard(wb, metrics or {}, today, week_label)
    if pm_plan and (pm_plan.get("schedule") or pm_plan.get("milestones")):
        _write_pm_kehoach_sheet(wb, pm_plan)

    os.makedirs(output_dir, exist_ok=True)
    safe_code = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in code)[:40]
    filename = f"{safe_code}_MoM_{today.year}.{week_label}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def _write_pm_kehoach_sheet(wb, pm_plan: dict) -> None:
    """Sheet bổ sung từ KeHoachDuAn (chiều PM) — đồng bộ H1/H3 với MoM."""
    ws = wb.create_sheet("PM Lịch trình")
    ws.sheet_view.showGridLines = False
    for col, width in enumerate([22, 45, 12, 12, 18, 18, 18, 25], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _style_title_bar(ws, 1, "PM LỊCH TRÌNH — Kế hoạch dự án (KeHoachDuAn)", 8)
    ws.merge_cells("A2:H2")
    src = pm_plan.get("source_filename") or ""
    imported = pm_plan.get("imported_at") or ""
    ws["A2"] = f"Nguồn: {src or '—'}  ·  Import: {imported or '—'}"
    ws["A2"].font = GUIDE_FONT
    ws["A2"].alignment = LEFT_CENTER
    ws.row_dimensions[2].height = 16

    headers = [
        "Giai đoạn", "Công việc", "Từ ngày", "Đến ngày",
        "PIC FPT", "Hỗ trợ FPT", "PIC KH", "Ghi chú",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    _style_header_row(ws, 3, 1, 8)

    row = 4
    schedule = pm_plan.get("schedule") or []
    if schedule:
        for item in schedule:
            vals = [
                item.get("phase") or "",
                item.get("name") or "",
                item.get("start") or "",
                item.get("end") or "",
                ", ".join(item.get("pic_fpt") or []),
                ", ".join(item.get("support_fpt") or []),
                ", ".join(item.get("pic_client") or []),
                item.get("note") or "",
            ]
            is_phase = bool(item.get("is_phase_header"))
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row, c, v)
                cell.border = THIN_BORDER
                cell.alignment = LEFT_TOP if c in (2, 8) else CENTER if c in (3, 4) else LEFT_CENTER
                cell.font = BODY_MOM_BOLD if is_phase else BODY_MOM_FONT
                if is_phase:
                    cell.fill = PHASE_HEADER_FILL
                elif row % 2 == 0:
                    cell.fill = ALT_ROW_FILL
            ws.row_dimensions[row].height = 18
            row += 1
    elif pm_plan.get("milestones"):
        ws.cell(3, 1, "STT")
        ws.cell(3, 2, "Milestone (WBS)")
        for c in (1, 2):
            ws.cell(3, c).fill = MOM_HEADER_FILL
            ws.cell(3, c).font = MOM_HEADER_FONT
            ws.cell(3, c).alignment = MOM_HEADER_ALIGN
            ws.cell(3, c).border = THIN_BORDER
        for m in pm_plan["milestones"]:
            ws.cell(row, 1, m.get("stt") or "").border = THIN_BORDER
            ws.cell(row, 1).font = BODY_MOM_FONT
            ws.cell(row, 1).alignment = CENTER
            ws.cell(row, 2, m.get("name") or "").border = THIN_BORDER
            ws.cell(row, 2).font = BODY_MOM_FONT
            ws.cell(row, 2).alignment = LEFT_CENTER
            row += 1
    else:
        ws.cell(row, 1, "N/A — chưa có lịch trình / milestone trong kế hoạch PM.").font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(
        row, 1,
        "Không ghi đè Master plan từ Function List — sheet này chỉ chiếu kế hoạch PM.",
    ).font = GUIDE_FONT

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"
    _set_print_landscape(ws)
