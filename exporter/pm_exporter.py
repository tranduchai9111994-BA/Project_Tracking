# -*- coding: utf-8 -*-
"""
Xuất Excel tổng hợp chiều PM (kế hoạch + weekly snapshot).

Style hierarchy đồng bộ MoM W30:
  H1 Title   = navy #1F4E79, trắng, 14pt bold, merge full-width
  H2 Section = vàng #FFC000, trắng, 12pt bold (khi cần)
  H3 Header  = xanh #0070C0, trắng, 11pt bold, center
"""
from __future__ import annotations

import os
from datetime import date, datetime
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Palette đồng bộ weekly_mom / mẫu W30
HEADER_FILL = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_BAR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_BAR_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
PHASE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GROUP_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
GUIDE_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
BODY_FONT = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_title_bar(ws, row: int, title: str, end_col: int) -> None:
    """H1 — thanh tiêu đề navy."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, title)
    cell.font = TITLE_BAR_FONT
    cell.fill = TITLE_BAR_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, end_col + 1):
        ws.cell(row, c).fill = TITLE_BAR_FILL
        ws.cell(row, c).border = THIN
    ws.row_dimensions[row].height = 28


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    """H3 — table header xanh #0070C0."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[row].height = 24


def _style_body_row(
    ws, row: int, n_cols: int, *, alt: bool = False, fill: Optional[PatternFill] = None,
) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        cell.border = THIN
        if cell.alignment is None or cell.alignment.vertical is None:
            cell.alignment = WRAP
        if cell.font is None or cell.font.name is None:
            cell.font = BODY_FONT
        if fill is not None:
            cell.fill = fill
        elif alt:
            cell.fill = ALT_ROW_FILL


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_pm_report(
    plan: Optional[dict[str, Any]],
    weekly: Optional[dict[str, Any]],
    output_dir: str,
    *,
    project_code: str = "",
    fl_links: Optional[dict[str, Any]] = None,
) -> str:
    """Sinh workbook chiều PM. Returns filepath."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_cover(wb, plan, weekly, project_code)
    if plan:
        _write_milestones(wb, plan)
        _write_schedule(wb, plan)
        _write_deliverables(wb, plan)
        _write_teams(wb, plan)
    if weekly:
        _write_weekly(wb, weekly)
    if fl_links and fl_links.get("link_count"):
        _write_fl_links(wb, fl_links)

    os.makedirs(output_dir, exist_ok=True)
    today = date.today().isoformat()
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (project_code or "project"))[:40]
    filename = f"{safe}_ChieuPM_{today}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def _write_cover(wb, plan, weekly, project_code: str) -> None:
    ws = wb.create_sheet("Tổng quan", 0)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [28, 14, 72])

    _style_title_bar(ws, 1, "BÁO CÁO CHIỀU PM — Kế hoạch & Weekly", 3)
    ws.merge_cells("A2:C2")
    ws["A2"] = (
        f"Dự án: {project_code or '—'}  ·  "
        f"Xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = GUIDE_FONT
    ws["A2"].alignment = LEFT_CENTER
    ws.row_dimensions[2].height = 16

    _header_row(ws, ["Nguồn", "Trạng thái", "Chi tiết"], row=4)

    row = 5
    if plan:
        s = plan.get("summary") or {}
        ws.cell(row, 1, "Kế hoạch dự án (Excel)")
        ws.cell(row, 2, "Đã import")
        ws.cell(
            row, 3,
            f"{plan.get('source_filename', '')} | "
            f"WBS {s.get('milestone_count', 0)} | "
            f"Lịch trình {s.get('schedule_count', 0)} | "
            f"Deliverable {s.get('deliverable_count', 0)} | "
            f"Import {plan.get('imported_at', '')}",
        )
    else:
        ws.cell(row, 1, "Kế hoạch dự án (Excel)")
        ws.cell(row, 2, "Chưa có")
        ws.cell(row, 3, "Upload file KeHoachDuAn (.xlsx)")
    _style_body_row(ws, row, 3)
    row += 1

    if weekly:
        s = weekly.get("summary") or {}
        period = f"{weekly.get('period_start') or '?'} → {weekly.get('period_end') or '?'}"
        ws.cell(row, 1, "Weekly Report (PPT)")
        ws.cell(row, 2, "Đã import")
        ws.cell(
            row, 3,
            f"{weekly.get('source_filename', '')} | {period} | "
            f"Done {s.get('done_count', 0)} | Next {s.get('next_count', 0)} | "
            f"Import {weekly.get('imported_at', '')}",
        )
    else:
        ws.cell(row, 1, "Weekly Report (PPT)")
        ws.cell(row, 2, "Chưa có")
        ws.cell(row, 3, "Upload file Weekly Report (.pptx)")
    _style_body_row(ws, row, 3, alt=True)

    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:4"


def _write_milestones(wb, plan: dict) -> None:
    ws = wb.create_sheet("WBS Gantt")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [8, 48, 36])

    _style_title_bar(ws, 1, "WBS / MILESTONE — từ Gantt kế hoạch dự án", 3)
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Nguồn: {plan.get('source_filename') or '—'}  ·  Import: {plan.get('imported_at') or '—'}"
    ws["A2"].font = GUIDE_FONT
    ws["A2"].alignment = LEFT_CENTER

    _header_row(ws, ["STT", "Công việc (milestone)", "Ghi chú"], row=3)
    milestones = plan.get("milestones") or []
    for i, m in enumerate(milestones, 1):
        r = i + 3
        ws.cell(r, 1, m.get("stt") or i)
        ws.cell(r, 2, m.get("name"))
        ws.cell(r, 3, m.get("note") or "")
        ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 1).font = BODY_FONT
        ws.cell(r, 2).font = BODY_FONT
        ws.cell(r, 3).font = BODY_FONT
        _style_body_row(ws, r, 3, alt=(i % 2 == 0))

    weeks = plan.get("weeks") or []
    if weeks:
        r = len(milestones) + 5
        ws.cell(r, 1, "Trục tuần (từ Gantt):").font = LABEL_FONT
        ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3)
        ws.cell(r + 1, 1, ", ".join(
            f"{w['label']}({w.get('month') or ''})" for w in weeks[:40]
        )).font = GUIDE_FONT

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"


def _write_schedule(wb, plan: dict) -> None:
    ws = wb.create_sheet("Lịch trình")
    ws.sheet_view.showGridLines = False
    headers = [
        "Giai đoạn", "Công việc", "Từ ngày", "Đến ngày",
        "PIC FPT", "Hỗ trợ FPT", "PIC KH", "Hỗ trợ KH", "Ghi chú", "Phase header?",
    ]
    _set_col_widths(ws, [18, 36, 12, 12, 16, 14, 14, 14, 28, 12])

    _style_title_bar(ws, 1, "LỊCH TRÌNH — KeHoachDuAn", 10)
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Nguồn: {plan.get('source_filename') or '—'}  ·  Import: {plan.get('imported_at') or '—'}"
    ws["A2"].font = GUIDE_FONT
    ws["A2"].alignment = LEFT_CENTER

    _header_row(ws, headers, row=3)
    for i, item in enumerate(plan.get("schedule") or [], 1):
        r = i + 3
        vals = [
            item.get("phase") or "",
            item.get("name") or "",
            item.get("start") or "",
            item.get("end") or "",
            ", ".join(item.get("pic_fpt") or []),
            ", ".join(item.get("support_fpt") or []),
            ", ".join(item.get("pic_client") or []),
            ", ".join(item.get("support_client") or []),
            item.get("note") or "",
            "Yes" if item.get("is_phase_header") else "",
        ]
        is_phase = bool(item.get("is_phase_header"))
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = BODY_BOLD if is_phase else BODY_FONT
            cell.alignment = CENTER if c in (3, 4, 10) else LEFT_CENTER
        fill = PHASE_FILL if is_phase else (ALT_ROW_FILL if i % 2 == 0 else None)
        _style_body_row(ws, r, 10, fill=fill)
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"


def _write_deliverables(wb, plan: dict) -> None:
    ws = wb.create_sheet("Sản phẩm bàn giao")
    ws.sheet_view.showGridLines = False
    headers = ["Nhóm", "STT", "Sản phẩm", "Ngày", "Loại", "Ghi chú", "Người lập"]
    _set_col_widths(ws, [18, 8, 40, 12, 14, 28, 16])

    _style_title_bar(ws, 1, "SẢN PHẨM BÀN GIAO", 7)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Nguồn: {plan.get('source_filename') or '—'}"
    ws["A2"].font = GUIDE_FONT

    _header_row(ws, headers, row=3)
    for i, d in enumerate(plan.get("deliverables") or [], 1):
        r = i + 3
        vals = [
            d.get("group") or "",
            d.get("stt") or "",
            d.get("name") or "",
            d.get("due_date") or "",
            d.get("type") or "",
            d.get("note") or "",
            d.get("author") or "",
        ]
        is_group = bool(d.get("is_group"))
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = BODY_BOLD if is_group else BODY_FONT
            cell.alignment = CENTER if c in (2, 4) else LEFT_CENTER
        fill = GROUP_FILL if is_group else (ALT_ROW_FILL if i % 2 == 0 else None)
        _style_body_row(ws, r, 7, fill=fill)

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"


def _write_teams(wb, plan: dict) -> None:
    ws = wb.create_sheet("Đội dự án")
    ws.sheet_view.showGridLines = False
    headers = ["Side", "Nhóm", "STT", "Họ tên", "Chức vụ", "Vai trò", "Email"]
    _set_col_widths(ws, [12, 16, 8, 28, 18, 22, 28])

    _style_title_bar(ws, 1, "ĐỘI DỰ ÁN — FPT & Khách hàng", 7)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Nguồn: {plan.get('source_filename') or '—'}"
    ws["A2"].font = GUIDE_FONT

    _header_row(ws, headers, row=3)
    rows = []
    for m in plan.get("team_vendor") or []:
        rows.append(("FPT", m))
    for m in plan.get("team_client") or []:
        rows.append(("Khách hàng", m))
    for i, (side, m) in enumerate(rows, 1):
        r = i + 3
        vals = [
            side,
            m.get("group") or "",
            m.get("stt") or "",
            m.get("name") or "",
            m.get("title") or "",
            m.get("role") or "",
            m.get("email") or "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = BODY_FONT
            cell.alignment = CENTER if c == 3 else LEFT_CENTER
        _style_body_row(ws, r, 7, alt=(i % 2 == 0))

    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"


def _write_weekly(wb, weekly: dict) -> None:
    period = f"{weekly.get('period_start') or '?'} → {weekly.get('period_end') or '?'}"
    title = weekly.get("title") or "Weekly"
    project_title = weekly.get("project_title") or ""

    # --- Done ---
    ws = wb.create_sheet("Weekly Done")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [8, 40, 16, 12, 14, 28])
    _style_title_bar(ws, 1, f"WEEKLY DONE — {title}", 6)
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Kỳ: {period}  ·  {project_title}".strip(" ·")
    ws["A2"].font = GUIDE_FONT
    ws["A2"].alignment = LEFT_CENTER
    _header_row(ws, ["STT", "Công việc", "Đơn vị", "Ngày", "Tình trạng", "Ghi chú"], row=3)
    for i, item in enumerate(weekly.get("done") or [], 1):
        r = i + 3
        for c, v in enumerate([
            item.get("stt"), item.get("task"), item.get("unit"),
            item.get("date"), item.get("status"), item.get("note"),
        ], 1):
            cell = ws.cell(r, c, v or "")
            cell.font = BODY_FONT
            cell.alignment = CENTER if c in (1, 4) else LEFT_CENTER
        _style_body_row(ws, r, 6, alt=(i % 2 == 0))
    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"

    # --- Next ---
    ws2 = wb.create_sheet("Weekly Next")
    ws2.sheet_view.showGridLines = False
    _set_col_widths(ws2, [8, 40, 16, 12, 12, 28])
    _style_title_bar(ws2, 1, f"WEEKLY NEXT — {title}", 6)
    ws2.merge_cells("A2:F2")
    ws2["A2"] = f"Kỳ: {period}"
    ws2["A2"].font = GUIDE_FONT
    _header_row(ws2, ["STT", "Công việc", "Đơn vị", "Bắt đầu", "Kết thúc", "Ghi chú"], row=3)
    for i, item in enumerate(weekly.get("next") or [], 1):
        r = i + 3
        for c, v in enumerate([
            item.get("stt"), item.get("task"), item.get("unit"),
            item.get("start"), item.get("end"), item.get("note"),
        ], 1):
            cell = ws2.cell(r, c, v or "")
            cell.font = BODY_FONT
            cell.alignment = CENTER if c in (1, 4, 5) else LEFT_CENTER
        _style_body_row(ws2, r, 6, alt=(i % 2 == 0))
    ws2.freeze_panes = "A4"
    ws2.print_title_rows = "1:3"

    # --- Risk ---
    ws3 = wb.create_sheet("Weekly Risk")
    ws3.sheet_view.showGridLines = False
    _set_col_widths(ws3, [12, 72])
    _style_title_bar(ws3, 1, f"WEEKLY ISSUE / RISK — {title}", 2)
    ws3.merge_cells("A2:B2")
    ws3["A2"] = f"Kỳ: {period}"
    ws3["A2"].font = GUIDE_FONT
    _header_row(ws3, ["Loại", "Nội dung"], row=3)
    r = 4
    idx = 0
    for t in weekly.get("issues") or []:
        ws3.cell(r, 1, "Issue")
        ws3.cell(r, 2, t)
        ws3.cell(r, 1).font = BODY_BOLD
        ws3.cell(r, 2).font = BODY_FONT
        _style_body_row(ws3, r, 2, alt=(idx % 2 == 1))
        r += 1
        idx += 1
    for t in weekly.get("risks") or []:
        ws3.cell(r, 1, "Risk")
        ws3.cell(r, 2, t)
        ws3.cell(r, 1).font = BODY_BOLD
        ws3.cell(r, 2).font = BODY_FONT
        _style_body_row(ws3, r, 2, alt=(idx % 2 == 1))
        r += 1
        idx += 1
    ws3.freeze_panes = "A4"
    ws3.print_title_rows = "1:3"


def _write_fl_links(wb, fl_links: dict) -> None:
    ws = wb.create_sheet("FL Links")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [14, 40, 24, 20, 24])
    _style_title_bar(ws, 1, "LIÊN KẾT FUNCTION LIST ↔ KẾ HOẠCH / WEEKLY", 5)
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Số link: {fl_links.get('link_count') or 0}"
    ws["A2"].font = GUIDE_FONT
    _header_row(ws, ["Nguồn", "Tên công việc", "Module trùng", "Phase trùng", "PIC trùng"], row=3)
    r = 4
    idx = 0
    for link in fl_links.get("schedule_links") or []:
        ws.cell(r, 1, "Lịch trình")
        ws.cell(r, 2, link.get("name"))
        ws.cell(r, 3, ", ".join(link.get("modules") or []))
        ws.cell(r, 4, ", ".join(link.get("phases") or []))
        ws.cell(r, 5, ", ".join(link.get("pics") or []))
        for c in range(1, 6):
            ws.cell(r, c).font = BODY_FONT
        _style_body_row(ws, r, 5, alt=(idx % 2 == 1))
        r += 1
        idx += 1
    for link in fl_links.get("weekly_links") or []:
        ws.cell(r, 1, f"Weekly/{link.get('kind')}")
        ws.cell(r, 2, link.get("task"))
        ws.cell(r, 3, ", ".join(link.get("modules") or []))
        for c in range(1, 6):
            ws.cell(r, c).font = BODY_FONT
        _style_body_row(ws, r, 5, alt=(idx % 2 == 1))
        r += 1
        idx += 1
    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"
