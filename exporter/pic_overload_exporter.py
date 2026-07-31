"""
Export PIC Overload — Tong_hop + Chi_tiet + optional FL re-import per project.

FL re-import best-effort:
  - Dùng header hiện tại của ParsedData (hoặc schema đã lưu nếu có)
  - Chỉ ghi các row liên quan overload
  - Tô vàng cột PIC / Status của phase dính overload + ghi note vào Remark nếu có
"""
from __future__ import annotations

import os
from collections import defaultdict
from datetime import date
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from exporter.excel_exporter import (
    BODY_ALIGN,
    BODY_FONT,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FONT,
    RED_FILL,
    THIN_BORDER,
    YELLOW_FILL,
    _normalize_export_mode,
    _want_detail,
    _want_summary,
    _write_sheet,
)
from parser.excel_parser import FunctionRow, ParsedData

# Soft yellow cho ô PIC/Status cần chú ý khi re-import
_FL_YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
_NOTE_PREFIX = "[PIC-Overload]"


def export_pic_overload_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    *,
    mode: str = "both",
    include_fl: bool = False,
    project_data: Optional[dict[str, ParsedData]] = None,
    project_dirs: Optional[dict[str, str]] = None,
) -> str:
    """
    Xuất workbook PIC Overload.

    Args:
        payload: kết quả compute_pic_overload
        mode: summary | detail | both
        include_fl: thêm sheet FL_<slug> re-import được
        project_data: slug → ParsedData (cần khi include_fl)
        project_dirs: slug → project folder (để resolve schema template nếu có)
    """
    mode = _normalize_export_mode(mode)
    grain = payload.get("grain", "day")
    thr = payload.get("thresholds") or {}
    summary = payload.get("summary") or {}
    by_pic = list(payload.get("by_pic") or [])
    by_period = list(payload.get("by_period") or [])
    detail = list(payload.get("detail") or [])
    highlight = summary.get("highlight_dates") or []

    subtitle = (
        f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}  |  "
        f"Grain: {grain}  |  Range: {payload.get('from')} → {payload.get('to')}  |  "
        f"Ngưỡng ngày > {thr.get('day_max_tasks', 5)} task  |  "
        f"Highlight: {len(highlight)} ngày"
    )

    wb = openpyxl.Workbook()
    first = True

    if _want_summary(mode):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = "Tong_hop"
        # PIC × grain period (chỉ dòng overload + top OK nếu ít)
        rows_src = [r for r in by_period if r.get("is_overload")] or by_period[:50]
        if not rows_src and by_pic:
            rows_src = [
                {
                    "pic": p["pic"],
                    "period": payload.get("from", ""),
                    "label": "Toàn khoảng",
                    "task_days": p["task_days"],
                    "overload_days": p["overload_days"],
                    "max_concurrent": p["max_concurrent"],
                    "is_overload": p["is_overload"],
                    "highlight_dates": p.get("highlight_dates") or [],
                    "projects": p.get("projects") or [],
                    "also_overdue": p.get("also_overdue"),
                }
                for p in by_pic if p.get("is_overload")
            ] or [
                {
                    "pic": p["pic"],
                    "period": "",
                    "label": "Toàn khoảng",
                    "task_days": p["task_days"],
                    "overload_days": p["overload_days"],
                    "max_concurrent": p["max_concurrent"],
                    "is_overload": p["is_overload"],
                    "highlight_dates": p.get("highlight_dates") or [],
                    "projects": p.get("projects") or [],
                    "also_overdue": p.get("also_overdue"),
                }
                for p in by_pic[:30]
            ]

        data_rows = [
            [
                idx + 1,
                r.get("pic", ""),
                r.get("label") or r.get("period", ""),
                r.get("max_concurrent", 0),
                r.get("task_days", 0),
                r.get("overload_days", 0),
                ", ".join(r.get("projects") or []),
                ", ".join((r.get("highlight_dates") or [])[:12]),
                "OVERLOAD" if r.get("is_overload") else "OK",
                "Có" if r.get("also_overdue") else "",
            ]
            for idx, r in enumerate(rows_src)
        ]

        def _sum_fill(_ri, idx):
            return RED_FILL if rows_src[idx].get("is_overload") else None

        _write_sheet(
            ws,
            title="PIC OVERLOAD — TỔNG HỢP (ĐA DỰ ÁN)",
            columns=[
                ("STT", 6),
                ("PIC", 18),
                ("Kỳ", 16),
                ("Max concurrent", 14),
                ("Task-days", 12),
                ("Ngày đỏ", 10),
                ("Projects", 28),
                ("Ngày highlight", 36),
                ("Status", 12),
                ("Vừa overdue?", 12),
            ],
            data_rows=data_rows,
            row_fill_fn=_sum_fill,
            subtitle=subtitle,
        )

    if _want_detail(mode):
        ws = wb.active if first else wb.create_sheet("Chi_tiet")
        first = False
        if ws.title != "Chi_tiet":
            ws.title = "Chi_tiet"
        data_rows = [
            [
                idx + 1,
                r.get("pic", ""),
                r.get("date", ""),
                r.get("project_name") or r.get("project_slug", ""),
                r.get("ma_cn", ""),
                r.get("ten_cn", ""),
                r.get("module", ""),
                r.get("phase", ""),
                r.get("status", ""),
                r.get("start", ""),
                r.get("end", ""),
                "Có" if r.get("is_overdue") else "",
            ]
            for idx, r in enumerate(detail)
        ]

        def _det_fill(_ri, idx):
            d = detail[idx].get("date", "")
            return RED_FILL if d in set(highlight) else None

        _write_sheet(
            ws,
            title="PIC OVERLOAD — CHI TIẾT TASK",
            columns=[
                ("STT", 6),
                ("PIC", 16),
                ("Ngày", 12),
                ("Project", 20),
                ("Mã CN", 14),
                ("Tên chức năng", 36),
                ("Module", 10),
                ("Phase", 14),
                ("Status", 12),
                ("Start", 12),
                ("End", 12),
                ("Overdue?", 10),
            ],
            data_rows=data_rows,
            row_fill_fn=_det_fill,
            subtitle=subtitle,
        )

    if include_fl and project_data:
        _append_fl_sheets(
            wb,
            detail=detail,
            project_data=project_data,
            project_dirs=project_dirs or {},
            first_is_active=first,
        )

    # Nếu mode không tạo sheet nào (edge) — sheet trống
    if first:
        wb.active.title = "Empty"
        wb.active["A1"] = "Không có dữ liệu PIC Overload để xuất."

    os.makedirs(output_dir, exist_ok=True)
    fname = f"PIC_Overload_{grain}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, fname)
    wb.save(filepath)
    wb.close()
    return filepath


def _append_fl_sheets(
    wb: openpyxl.Workbook,
    *,
    detail: list[dict],
    project_data: dict[str, ParsedData],
    project_dirs: dict[str, str],
    first_is_active: bool,
) -> None:
    """Mỗi project 1 sheet FL_<slug> — chỉ row bị overload, tô vàng PIC/Status."""
    # slug → set(ma_cn) + set((ma_cn, phase))
    by_slug: dict[str, set[str]] = defaultdict(set)
    phases_hit: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for r in detail:
        slug = r.get("project_slug") or ""
        ma = r.get("ma_cn") or ""
        if not slug or not ma:
            continue
        by_slug[slug].add(ma)
        phases_hit[slug].add((ma, r.get("phase") or ""))

    for slug, ma_set in sorted(by_slug.items()):
        data = project_data.get(slug)
        if data is None:
            continue
        sheet_name = _safe_sheet_name(f"FL_{slug}")
        if first_is_active and wb.active.title in ("Sheet", "Empty"):
            ws = wb.active
            ws.title = sheet_name
            first_is_active = False
        else:
            ws = wb.create_sheet(sheet_name)

        headers_ordered = _headers_ordered(data)
        # Try saved schema for header order
        pdir = project_dirs.get(slug)
        if pdir:
            try:
                from exporter.fl_export_schema import load_fl_export_schema
                saved = load_fl_export_schema(pdir)
                if saved and saved.get("headers"):
                    headers_ordered = list(saved["headers"])
            except Exception:
                pass

        # Header row
        for col, h in enumerate(headers_ordered, 1):
            cell = ws.cell(1, col, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = min(28, max(10, len(h) + 2))

        # Index rows by ma_cn
        rows_by_ma = {
            str(r.meta.get("ma_cn") or ""): r
            for r in data.rows
            if r.meta.get("ma_cn")
        }
        remark_col = _find_remark_col(data, headers_ordered)
        pic_status_cols = _pic_status_col_map(data, headers_ordered)

        out_row = 2
        for ma in sorted(ma_set):
            row = rows_by_ma.get(ma)
            if row is None:
                continue
            hit_phases = {ph for (m, ph) in phases_hit[slug] if m == ma}
            values = _row_values(row, data, headers_ordered)
            # Append note
            if remark_col:
                note = values.get(remark_col) or ""
                add = f"{_NOTE_PREFIX} phase overload: {', '.join(sorted(hit_phases))}"
                values[remark_col] = f"{note} | {add}".strip(" |") if note else add

            for col, h in enumerate(headers_ordered, 1):
                cell = ws.cell(out_row, col, values.get(h, ""))
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = THIN_BORDER
                # Highlight PIC/Status of hit phases
                info = pic_status_cols.get(h)
                if info and info["phase"] in hit_phases:
                    cell.fill = _FL_YELLOW
                if remark_col and h == remark_col:
                    cell.fill = _FL_YELLOW
            out_row += 1

        ws.freeze_panes = "A2"
        if out_row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers_ordered))}{out_row - 1}"


def _safe_sheet_name(name: str) -> str:
    bad = set(r'[]:*?/\\')
    cleaned = "".join("_" if c in bad else c for c in name)[:31]
    return cleaned or "FL"


def _headers_ordered(data: ParsedData) -> list[str]:
    items = sorted(data.headers.items(), key=lambda x: x[1])
    return [h for h, _ in items]


def _find_remark_col(data: ParsedData, headers: list[str]) -> Optional[str]:
    # Prefer meta remark
    if data.meta_columns.get("remark"):
        inv = {idx: h for h, idx in data.headers.items()}
        return inv.get(data.meta_columns["remark"])
    for h in headers:
        hl = h.lower()
        if " - " in h:
            continue
        if "remark" in hl or "ghi chú" in hl or "ghi chu" in hl:
            return h
    return None


def _pic_status_col_map(data: ParsedData, headers: list[str]) -> dict[str, dict]:
    """header → {phase, kind: pic|status}."""
    out: dict[str, dict] = {}
    for pg in data.phase_groups:
        for attr, col in (pg.attributes or {}).items():
            for h, idx in data.headers.items():
                if idx != col:
                    continue
                au = attr.upper()
                if "PIC" in au:
                    out[h] = {"phase": pg.name, "kind": "pic"}
                elif "STATUS" in au:
                    out[h] = {"phase": pg.name, "kind": "status"}
    return out


def _row_values(row: FunctionRow, data: ParsedData, headers: list[str]) -> dict[str, Any]:
    """Best-effort reconstruct cell values từ ParsedData (không giữ format gốc)."""
    values: dict[str, Any] = {}
    inv = {idx: h for h, idx in data.headers.items()}
    for meta_key, col in (data.meta_columns or {}).items():
        if not col:
            continue
        h = inv.get(col)
        if h:
            values[h] = row.meta.get(meta_key, "")

    for pg in data.phase_groups:
        pd = row.phases.get(pg.name)
        if not pd:
            continue
        for attr, col in (pg.attributes or {}).items():
            h = inv.get(col)
            if not h:
                continue
            au = attr.upper()
            if au in ("START", "FROM", "PLANNED"):
                values[h] = pd.start_date.isoformat() if pd.start_date else ""
            elif au in ("END", "TO", "ACTUAL"):
                values[h] = pd.end_date.isoformat() if pd.end_date else ""
            elif "STATUS" in au:
                values[h] = pd.status or ""
            elif "PIC" in au:
                values[h] = ", ".join(pd.pics or [])
            elif "ESTIMATE" in au or au.endswith("MH"):
                values[h] = pd.estimate_mh if pd.estimate_mh is not None else ""
            elif "NOTE" in au:
                values[h] = pd.note or ""
            else:
                values[h] = (pd.extra or {}).get(attr, "")
    for h in headers:
        values.setdefault(h, "")
    return values
