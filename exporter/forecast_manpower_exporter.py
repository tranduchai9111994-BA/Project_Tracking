"""Xuất Excel Forecast Manpower — Tong_hop + Chi_tiet."""
from __future__ import annotations

import os
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def export_forecast_manpower(
    result: dict[str, Any],
    upload_folder: str,
    mode: str = "both",
) -> str:
    """
    mode: summary | detail | both
    Sheets: Tong_hop (pools + stages), Chi_tiet (function × phase).
    """
    os.makedirs(upload_folder, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(upload_folder, f"Forecast_Manpower_{ts}.xlsx")
    mode = (mode or "both").strip().lower()
    if mode not in ("summary", "detail", "both"):
        mode = "both"

    wb = openpyxl.Workbook()
    # remove default later
    default_ws = wb.active
    wb.remove(default_ws)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    h2_fill = PatternFill("solid", fgColor="FFC000")
    hire_fill = PatternFill("solid", fgColor="FECACA")

    def _style_header(ws, headers: list[str]) -> None:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    unit = result.get("display_unit_label") or result.get("display_unit") or "MH"
    if mode in ("summary", "both"):
        ws = wb.create_sheet("Tong_hop", 0)
        # Meta
        ws["A1"] = "Forecast Manpower"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = result.get("basis_label") or ""
        ws["A3"] = f"Đơn vị hiển thị: {unit} · Target tháng: {result.get('target_months')}"
        ws.merge_cells("A1:H1")

        row = 5
        ws.cell(row, 1, "POOL — Dev riêng / Triển khai chung").fill = h2_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
        headers = [
            "Nhóm",
            f"Tổng ({unit})",
            f"Còn lại ({unit})",
            f"Đã Closed ({unit})",
            "Phase còn",
            "MH mặc định (số)",
            "Người hiện tại",
            "Người cần",
            "Cần tuyển thêm",
            "Ghi chú / phương pháp",
        ]
        _style_header_at(ws, row, headers, header_fill, header_font)
        row += 1
        for p in result.get("pools") or []:
            vals = [
                p.get("label"),
                p.get("display_total"),
                p.get("display_remaining"),
                p.get("display_closed"),
                p.get("count_remaining"),
                p.get("count_defaulted"),
                p.get("headcount_current"),
                p.get("people_needed"),
                p.get("hire_needed"),
                p.get("method_note"),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row, c, v)
                if c == 9 and (p.get("hire_needed") or 0) > 0:
                    cell.fill = hire_fill
            row += 1

        row += 1
        ws.cell(row, 1, "THEO CÔNG ĐOẠN (task type)").fill = h2_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
        _style_header_at(ws, row, headers, header_fill, header_font)
        row += 1
        for p in result.get("stages") or []:
            vals = [
                p.get("label"),
                p.get("display_total"),
                p.get("display_remaining"),
                p.get("display_closed"),
                p.get("count_remaining"),
                p.get("count_defaulted"),
                p.get("headcount_current"),
                p.get("people_needed"),
                p.get("hire_needed"),
                p.get("method_note"),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row, c, v)
                if c == 9 and (p.get("hire_needed") or 0) > 0:
                    cell.fill = hire_fill
            row += 1

        for col, w in enumerate([28, 14, 14, 14, 10, 12, 12, 10, 12, 55], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    if mode in ("detail", "both"):
        ws = wb.create_sheet("Chi_tiet")
        headers = [
            "Mã CN",
            "Tên chức năng",
            "Module",
            "Quy trình",
            "Phase",
            "Công đoạn",
            "Pool",
            "Status",
            "MH tính",
            "MH gốc FL",
            "Default?",
            "Cơ sở",
            "Start",
            "End",
            "PIC",
            "Ghi chú / phương pháp",
        ]
        _style_header(ws, headers)
        for r_i, d in enumerate(result.get("detail") or [], 2):
            vals = [
                d.get("ma_cn"),
                d.get("ten_cn"),
                d.get("module"),
                d.get("quy_trinh"),
                d.get("phase"),
                d.get("task_type"),
                d.get("pool"),
                d.get("status"),
                d.get("mh"),
                d.get("estimate_mh_raw"),
                "Yes" if d.get("used_default") else "",
                d.get("basis"),
                d.get("start"),
                d.get("end"),
                d.get("pic"),
                d.get("method_note"),
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(r_i, c, v)
        for col, w in enumerate([12, 36, 8, 14, 14, 14, 16, 12, 10, 10, 8, 10, 12, 12, 16, 40], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    if not wb.sheetnames:
        wb.create_sheet("Tong_hop")

    wb.save(path)
    return path


def _style_header_at(ws, row: int, headers: list[str], fill, font) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
