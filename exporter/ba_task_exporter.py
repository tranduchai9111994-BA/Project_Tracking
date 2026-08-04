"""
Xuất Excel cho BA Task Management (Gói B).

- export_ba_tasks: toàn bộ đầu việc, 1 sheet.
- export_ba_tasks_weekly: 4 sheet theo tuần — Đầu việc / Cuộc họp / Sản phẩm / Nợ KH
  (B5 sẽ nâng cấp thêm định dạng/tuần selector; ở đây đã có style đỏ/vàng cơ bản).
"""
from __future__ import annotations

import os
from datetime import date
from typing import Any, Optional

import openpyxl

from analyzer.ba_task_store import tasks_in_week, week_date_range
from exporter.excel_exporter import ORANGE_FILL, RED_FILL, _write_sheet


def _tags_str(t: dict[str, Any]) -> str:
    tags = t.get("tags") or []
    return ", ".join(str(x) for x in tags if x)


def _task_row_fill(alert_level: Optional[str]):
    if alert_level == "overdue":
        return RED_FILL
    if alert_level in ("upcoming", "blocked"):
        return ORANGE_FILL
    return None


_TASK_COLUMNS = [
    ("STT", 5), ("Tiêu đề", 34), ("Loại", 12), ("Module", 10),
    ("Trạng thái", 12), ("Ưu tiên", 10), ("PIC", 16),
    ("Hạn", 12), ("Ngày xong", 12), ("Cảnh báo", 10), ("Ghi chú", 30), ("Tags", 20),
]


def _task_rows(tasks: list[dict[str, Any]]) -> tuple[list[list[Any]], list]:
    rows, fills = [], []
    for idx, t in enumerate(tasks, 1):
        rows.append([
            idx, t.get("title", ""), t.get("type", ""), t.get("module", ""),
            t.get("status", ""), t.get("priority", ""), t.get("assignee", ""),
            t.get("due_date") or "", t.get("done_date") or "",
            t.get("alert_level") or "", t.get("notes") or "", _tags_str(t),
        ])
        fills.append(_task_row_fill(t.get("alert_level")))
    return rows, fills


def export_ba_tasks(
    tasks: list[dict[str, Any]], output_dir: str = "uploads", project_name: str = "",
) -> str:
    """Xuất toàn bộ đầu việc BA, 1 sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BA_Tasks"
    rows, fills = _task_rows(tasks)
    _write_sheet(
        ws,
        title="QUẢN LÝ ĐẦU VIỆC BA",
        subtitle=f"{project_name} · Tổng: {len(tasks)} · Ngày xuất: {date.today().strftime('%d/%m/%Y')}",
        columns=_TASK_COLUMNS,
        data_rows=rows,
        row_fill_fn=lambda row_idx, item_idx: fills[item_idx],
    )
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"BA_Tasks_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


def export_ba_tasks_weekly(
    tasks: list[dict[str, Any]], week_iso: str, output_dir: str = "uploads", project_name: str = "",
) -> str:
    """4 sheet theo tuần: Đầu việc / Cuộc họp / Sản phẩm bàn giao / Nợ KH đang chờ."""
    wb = openpyxl.Workbook()
    rng = week_date_range(week_iso)
    date_range = f"{rng[0].strftime('%d/%m/%Y')} – {rng[1].strftime('%d/%m/%Y')}" if rng else ""
    week_num = week_iso.split("-W")[-1] if "-W" in week_iso else week_iso
    header = (
        f"Dự án: {project_name} | Tuần {week_num} ({date_range}) | "
        f"Xuất lúc: {date.today().strftime('%d/%m/%Y')}"
    )

    week_tasks = [t for t in tasks if t.get("type") == "task"]
    week_tasks = tasks_in_week(week_tasks, week_iso, date_field="due_date")
    ws1 = wb.active
    ws1.title = "Dau_viec_tuan"
    rows, fills = _task_rows(week_tasks)
    _write_sheet(
        ws1, title="ĐẦU VIỆC TRONG TUẦN", subtitle=header,
        columns=_TASK_COLUMNS, data_rows=rows,
        row_fill_fn=lambda row_idx, item_idx: fills[item_idx],
    )

    meetings = [t for t in tasks if t.get("type") == "meeting"]
    meetings = tasks_in_week(meetings, week_iso, date_field="due_date")
    ws2 = wb.create_sheet("Cuoc_hop")
    meet_cols = [
        ("STT", 5), ("Tiêu đề", 30), ("Ngày họp", 12), ("Giờ", 8),
        ("Địa điểm", 20), ("Thành phần", 30), ("Agenda", 30), ("MoM", 40),
    ]
    meet_rows = []
    for idx, t in enumerate(meetings, 1):
        info = t.get("meeting_info") or {}
        attendees = info.get("attendees") or []
        meet_rows.append([
            idx, t.get("title", ""), info.get("meeting_date") or t.get("due_date") or "",
            info.get("time") or "", info.get("location") or "",
            ", ".join(str(a) for a in attendees), info.get("agenda") or "", info.get("mom_notes") or "",
        ])
    _write_sheet(ws2, title="CUỘC HỌP TRONG TUẦN", subtitle=header, columns=meet_cols, data_rows=meet_rows)

    deliverables = [t for t in tasks if t.get("type") == "deliverable"]
    deliverables = tasks_in_week(deliverables, week_iso, date_field="due_date")
    ws3 = wb.create_sheet("San_pham_ban_giao")
    dlv_cols = [
        ("STT", 5), ("Tên sản phẩm", 32), ("Format", 10), ("Hạn nộp", 12),
        ("Đã nộp", 12), ("Đã duyệt", 12), ("Reviewer", 16), ("Trạng thái", 12),
    ]
    dlv_rows = []
    for idx, t in enumerate(deliverables, 1):
        info = t.get("deliverable_info") or {}
        dlv_rows.append([
            idx, info.get("deliverable_name") or t.get("title", ""), info.get("format") or "",
            info.get("target_date") or t.get("due_date") or "", info.get("submitted_date") or "",
            info.get("approved_date") or "", info.get("reviewer") or "", t.get("status", ""),
        ])
    _write_sheet(ws3, title="SẢN PHẨM BÀN GIAO", subtitle=header, columns=dlv_cols, data_rows=dlv_rows)

    debts = [t for t in tasks if t.get("type") == "customer_debt" and t.get("status") not in ("done", "cancelled")]
    ws4 = wb.create_sheet("No_KH")
    debt_cols = [
        ("STT", 5), ("Mô tả", 34), ("Ngày yêu cầu", 14), ("Số ngày chờ", 12),
        ("Chịu trách nhiệm", 20), ("Số lần follow-up", 14), ("Follow-up gần nhất", 16),
    ]
    today = date.today()
    debt_rows, debt_fills = [], []
    for idx, t in enumerate(debts, 1):
        info = t.get("debt_info") or {}
        req = info.get("requested_date")
        wait_days = None
        if req:
            try:
                wait_days = (today - date.fromisoformat(str(req)[:10])).days
            except ValueError:
                wait_days = None
        debt_rows.append([
            idx, info.get("description") or t.get("title", ""), req or "",
            wait_days if wait_days is not None else "",
            info.get("responsible_party") or "", info.get("follow_up_count") or 0,
            info.get("last_follow_up") or "",
        ])
        debt_fills.append(RED_FILL if (wait_days or 0) > 7 else None)
    _write_sheet(
        ws4, title="NỢ KHÁCH HÀNG ĐANG CHỜ", subtitle=header, columns=debt_cols, data_rows=debt_rows,
        row_fill_fn=lambda row_idx, item_idx: debt_fills[item_idx],
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"BA_Tasks_Weekly_{week_iso}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath
