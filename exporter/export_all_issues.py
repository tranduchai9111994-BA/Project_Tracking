"""
T34 Task 1 — Xuất "Toàn bộ vấn đề" ra 1 Excel workbook duy nhất.

File: `iHRP_Van_De_Tong_Hop_<slug>_YYYYMMDD.xlsx`

Sheet layout (0-indexed theo README, thực tế openpyxl là 1-indexed):
  0. Cover        — Tổng quan project + filter + count mỗi loại (link đến sheet)
  1. Overdue      — Function trễ deadline (dedup theo Mã CN, phase merged)
  2. Chua_Co_PIC  — Function chưa assign
  3. Dinh_Tre     — Function stalled giữa 2 phase
  4. High_Risk    — Function risk score cao (>=30)
  5. Aging_WIP    — Function In-progress quá lâu
  6. Data_Quality — Row có lỗi data
  7. Bookmark     — Function đã star (nếu có)

Reuse ParsedData đã filter — không tự filter lại. Signature nhận sẵn:
  - overdue_list (deduped theo ma_cn)
  - unassigned_list
  - stalled_list
  - risk_list (>=30)
  - aging_wip_items
  - data_quality_issues
  - bookmark_functions
Giúp caller apply global filter một lần và inject vào exporter → tránh
tính lại 8 lần.

Layout theo spec:
  - Banner row 1 (merge A1:H1) tên sheet + count + màu category.
  - Header row 2 bold + fill xanh nhạt.
  - Freeze pane row 3.
  - Auto-filter, auto column width (đã tính sẵn qua bảng CFG).
"""
from __future__ import annotations

import os
from collections import defaultdict
from datetime import date
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from analyzer.i18n import normalize_lang, sheet_name as _sn, t as _t
from exporter.reason_formatters import (
    format_risk_factors_detailed,
    process_code,
    reason_aging_wip,
    reason_overdue,
    reason_stalled,
    reason_unassigned,
)


# ==========================================================================
# STYLE — Palette theo spec (banner màu category)
# ==========================================================================

BANNER_COLORS = {
    "cover":       "1F4E79",  # Xanh navy
    "overdue":     "C00000",  # Đỏ
    "unassigned":  "ED7D31",  # Cam
    "stalled":     "BF8F00",  # Vàng đậm
    "high_risk":   "E60000",  # Đỏ tươi
    "aging":       "FFC000",  # Vàng nhạt
    "data_quality":"595959",  # Xám
    "bookmark":    "7030A0",  # Tím
}

HEADER_FILL = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, size=10, color="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Fill theo mức trễ (tôn màu overdue > aging > stalled > default)
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")


def _fill_by_days(days: int) -> Optional[PatternFill]:
    if not days or days < 0:
        return None
    if days >= 30:
        return RED_FILL
    if days >= 14:
        return ORANGE_FILL
    if days >= 7:
        return YELLOW_FILL
    return None


def _fill_by_risk(score: int) -> Optional[PatternFill]:
    if score >= 80:
        return RED_FILL
    if score >= 50:
        return ORANGE_FILL
    if score >= 30:
        return YELLOW_FILL
    return None


# ==========================================================================
# Helper — banner + header + data + freeze
# ==========================================================================

def _write_banner(
    ws, title: str, count: int, category: str, n_cols: int, lang: str = "vi",
) -> None:
    """Row 1: banner merge A1:<lastcol>1, fill màu category, text trắng đậm."""
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = f"{title}  —  {_t('all_issues.total_records', lang, count=count)}"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(
        start_color=BANNER_COLORS.get(category, "1F4E79"),
        end_color=BANNER_COLORS.get(category, "1F4E79"),
        fill_type="solid",
    )
    ws.row_dimensions[1].height = 30


def _write_header(ws, columns: list[tuple[str, int]]) -> None:
    """Row 2: header bold + fill xanh nhạt."""
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28


def _write_data(
    ws,
    rows: list[list[Any]],
    row_fill_fn=None,
    n_cols: int = 0,
) -> None:
    """Ghi data từ row 3 trở đi."""
    for offset, values in enumerate(rows):
        r_idx = 3 + offset
        fill = row_fill_fn(offset) if row_fill_fn else None
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    # Empty state — thông báo "Không có record" ở A3
    if not rows and n_cols:
        last = get_column_letter(n_cols)
        ws.merge_cells(f"A3:{last}3")
        c = ws["A3"]
        c.value = "✓ Không có record nào trong nhóm này (đã áp dụng filter global)."
        c.font = Font(name="Arial", italic=True, color="666666")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 24


def _finalize_sheet(ws, n_rows: int, n_cols: int) -> None:
    """Freeze pane row 3, auto-filter row 2."""
    ws.freeze_panes = "A3"
    if n_rows > 0:
        last = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A2:{last}{2 + n_rows}"


# ==========================================================================
# Dedup helper — gom nhiều phase của 1 function thành 1 row
# ==========================================================================

def _dedup_by_ma_cn(items: list[dict], phase_field: str = "phase") -> list[dict]:
    """
    Gom nhiều phase-record của cùng 1 Mã CN thành 1 row (spec: sheet Overdue
    dedup theo Mã CN, phase merged bằng dấu phẩy).

    - Giữ ma_cn/ten_cn/module/priority của record đầu tiên.
    - Merge phase → "Analysis, UAT".
    - Chọn days_overdue = MAX (worst-case).
    - Merge PIC (unique, giữ order first-seen).
    """
    groups: dict[str, dict] = {}
    for it in items:
        key = str(it.get("ma_cn") or "").strip() or f"__norow_{id(it)}"
        if key not in groups:
            g = dict(it)  # copy
            g["_phases"] = [str(it.get(phase_field) or "")]
            g["_pics"] = list(it.get("pic") or [])
            g["_max_days"] = int(it.get("days_overdue") or 0)
            groups[key] = g
        else:
            g = groups[key]
            ph = str(it.get(phase_field) or "")
            if ph and ph not in g["_phases"]:
                g["_phases"].append(ph)
            for p in (it.get("pic") or []):
                if p and p not in g["_pics"]:
                    g["_pics"].append(p)
            days = int(it.get("days_overdue") or 0)
            if days > g["_max_days"]:
                g["_max_days"] = days

    out = []
    for g in groups.values():
        g[phase_field] = ", ".join([p for p in g["_phases"] if p])
        g["pic"] = g["_pics"]
        g["days_overdue"] = g["_max_days"]
        g.pop("_phases", None)
        g.pop("_pics", None)
        g.pop("_max_days", None)
        out.append(g)
    out.sort(key=lambda x: x.get("days_overdue", 0), reverse=True)
    return out


# ==========================================================================
# Main entry
# ==========================================================================

def export_all_issues(
    *,
    project_name: str,
    slug: str,
    overdue_list: list[dict],
    unassigned_list: list[dict],
    stalled_list: list[dict],
    risk_list: list[dict],
    aging_wip_items: list[dict],
    data_quality_issues: list[dict],
    bookmark_functions: list[dict],
    filter_info: Optional[dict] = None,
    output_dir: str = "uploads",
    lang: str = "vi",
) -> str:
    """
    Xuất 1 workbook 8 sheet (Cover + 7 loại vấn đề).

    Args:
      project_name: tên project (hiển thị ở cover).
      slug: dùng để đặt tên file.
      overdue_list / unassigned_list / …: đã filter global (caller lo).
      filter_info: {"modules": [...], "processes": [...], "pics": [...]}
        — để render dòng "Filter đang áp dụng" ở cover.
      output_dir: thư mục output (mặc định uploads/).
      lang: 'vi' | 'en' — tên sheet + banner theo ngôn ngữ UI.

    Returns:
      Path to Excel file.
    """
    lang = normalize_lang(lang)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    names = {
        "cover": _sn("cover", lang),
        "overdue": _sn("overdue", lang),
        "unassigned": _sn("unassigned", lang),
        "stalled": _sn("stalled", lang),
        "high_risk": _sn("high_risk", lang),
        "aging_wip": _sn("aging_wip", lang),
        "data_quality": _sn("data_quality", lang),
        "bookmark": _sn("bookmark", lang),
    }

    _write_cover_sheet(
        wb,
        project_name=project_name,
        slug=slug,
        counts={
            names["overdue"]: len(_dedup_by_ma_cn(overdue_list)),
            names["unassigned"]: len(unassigned_list),
            names["stalled"]: len(stalled_list),
            names["high_risk"]: len(risk_list),
            names["aging_wip"]: len(aging_wip_items),
            names["data_quality"]: len(data_quality_issues),
            names["bookmark"]: len(bookmark_functions),
        },
        filter_info=filter_info or {},
        lang=lang,
        sheet_names=names,
    )

    _write_overdue_sheet(wb, _dedup_by_ma_cn(overdue_list), lang=lang, name=names["overdue"])
    _write_unassigned_sheet(wb, unassigned_list, lang=lang, name=names["unassigned"])
    _write_stalled_sheet(wb, stalled_list, lang=lang, name=names["stalled"])
    _write_risk_sheet(wb, risk_list, lang=lang, name=names["high_risk"])
    _write_aging_sheet(wb, aging_wip_items, lang=lang, name=names["aging_wip"])
    _write_dq_sheet(wb, data_quality_issues, lang=lang, name=names["data_quality"])
    _write_bookmark_sheet(wb, bookmark_functions, lang=lang, name=names["bookmark"])

    wb.active = 0

    os.makedirs(output_dir, exist_ok=True)
    safe_slug = "".join(c for c in slug if c.isalnum() or c in ("_", "-")) or "project"
    filename = f"iHRP_Van_De_Tong_Hop_{safe_slug}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# Per-sheet writers
# ==========================================================================

def _write_cover_sheet(
    wb, *, project_name: str, slug: str, counts: dict[str, int], filter_info: dict,
    lang: str = "vi", sheet_names: Optional[dict] = None,
) -> None:
    """Cover sheet: project name, filter info, timestamp, count mỗi loại với link."""
    sn = sheet_names or {}
    cover = sn.get("cover") or _sn("cover", lang)
    ws = wb.create_sheet(cover)

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"📊 {_t('all_issues.title', lang)} — {project_name}"
    c.font = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color=BANNER_COLORS["cover"],
                         end_color=BANNER_COLORS["cover"], fill_type="solid")
    ws.row_dimensions[1].height = 32

    all_label = "(all)" if lang == "en" else "(tất cả)"
    meta_rows = [
        ("Project", project_name),
        ("Slug", slug),
        (_t("all_issues.export_date", lang), date.today().strftime("%d/%m/%Y")),
    ]
    modules = ", ".join(filter_info.get("modules") or []) or all_label
    processes = ", ".join(filter_info.get("processes") or []) or all_label
    pics = ", ".join(filter_info.get("pics") or []) or all_label
    meta_rows.extend([
        (f"{_t('all_issues.filter', lang)} Module", modules),
        (f"{_t('all_issues.filter', lang)} Process", processes),
        (f"{_t('all_issues.filter', lang)} PIC", pics),
    ])

    for idx_r, (label, value) in enumerate(meta_rows, start=3):
        lc = ws.cell(row=idx_r, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=11)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = THIN_BORDER
        vc = ws.cell(row=idx_r, column=2, value=str(value))
        vc.font = Font(name="Arial", size=11)
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc.border = THIN_BORDER
        ws.merge_cells(start_row=idx_r, start_column=2, end_row=idx_r, end_column=4)

    start = 3 + len(meta_rows) + 2
    ws.cell(row=start - 1, column=1, value=_t("all_issues.title", lang)).font = Font(
        name="Arial", bold=True, size=12, color="1F4E79"
    )
    ws.merge_cells(start_row=start - 1, start_column=1, end_row=start - 1, end_column=4)

    headers = (
        ["Issue type", "Count", "Sheet", "Link"] if lang == "en"
        else ["Loại vấn đề", "Số record", "Sheet", "Link"]
    )
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    open_lbl = "→ Open" if lang == "en" else "→ Mở sheet"
    display_map = [
        (_t("all_issues.sheet.overdue", lang), sn.get("overdue") or _sn("overdue", lang)),
        (_t("all_issues.sheet.unassigned", lang), sn.get("unassigned") or _sn("unassigned", lang)),
        (_t("all_issues.sheet.stalled", lang), sn.get("stalled") or _sn("stalled", lang)),
        (_t("all_issues.sheet.high_risk", lang), sn.get("high_risk") or _sn("high_risk", lang)),
        (_t("all_issues.sheet.aging", lang), sn.get("aging_wip") or _sn("aging_wip", lang)),
        (_t("all_issues.sheet.dq", lang), sn.get("data_quality") or _sn("data_quality", lang)),
        (_t("all_issues.sheet.bookmark", lang), sn.get("bookmark") or _sn("bookmark", lang)),
    ]
    for offset, (label, sheet) in enumerate(display_map):
        r = start + 1 + offset
        cnt = counts.get(sheet, 0)

        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BODY_FONT
        c1.alignment = BODY_ALIGN
        c1.border = THIN_BORDER

        c2 = ws.cell(row=r, column=2, value=cnt)
        c2.font = Font(name="Arial", bold=True, size=11,
                       color="C00000" if cnt > 0 else "666666")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.border = THIN_BORDER

        c3 = ws.cell(row=r, column=3, value=sheet)
        c3.font = BODY_FONT
        c3.alignment = BODY_ALIGN
        c3.border = THIN_BORDER

        c4 = ws.cell(row=r, column=4, value=open_lbl)
        c4.hyperlink = Hyperlink(
            ref=f"D{r}",
            location=f"'{sheet}'!A1",
            display=f"{open_lbl} {sheet}",
        )
        c4.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        c4.alignment = BODY_ALIGN
        c4.border = THIN_BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28


def _write_overdue_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("overdue", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Code", 14), ("Function Name", 40), ("Module", 12),
            ("Process", 22),
            ("Overdue phases", 28), ("Deadline (max)", 13), ("Days late (max)", 15),
            ("Status", 13), ("PIC", 22), ("Priority", 12), ("Note", 30),
            ("Reason", 48),
        ]
        title = "🔴 OVERDUE TASKS (dedup by Code)"
    else:
        cols = [
            ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 12),
            ("Quy trình", 22),
            ("Phase trễ (gộp)", 28), ("Deadline (max)", 13), ("Số ngày trễ (max)", 15),
            ("Status", 13), ("PIC", 22), ("Priority", 12), ("Ghi chú", 30),
            ("Lý do", 48),
        ]
        title = "🔴 TASK TRỄ DEADLINE (dedup theo Mã CN)"
    _write_banner(ws, title, len(items), "overdue", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        rows.append([
            idx + 1,
            it.get("ma_cn", ""), it.get("ten_cn", ""), it.get("module", ""),
            process_code(it),
            it.get("phase", ""), it.get("end_date", ""), it.get("days_overdue", 0),
            it.get("status", ""), ", ".join(it.get("pic", [])),
            it.get("priority", ""), it.get("note", ""),
            reason_overdue(it),
        ])

    _write_data(
        ws, rows,
        row_fill_fn=lambda i: _fill_by_days(items[i].get("days_overdue", 0)),
        n_cols=len(cols),
    )
    _finalize_sheet(ws, len(rows), len(cols))


def _write_unassigned_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("unassigned", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Code", 14), ("Rlog ID", 14), ("Function Name", 40), ("Module", 12),
            ("Process", 22), ("Prev phase", 16), ("Phase", 16), ("Status", 16),
            ("Priority", 12), ("Complexity", 12),
            ("Start", 13), ("Deadline", 13), ("Days late", 12), ("Reason", 48),
        ]
        title = "🟠 UNASSIGNED TASKS"
    else:
        cols = [
            ("STT", 6), ("Mã CN", 14), ("Rlog ID", 14), ("Tên chức năng", 40), ("Module", 12),
            ("Quy trình", 22), ("Phase trước", 16), ("Phase", 16), ("Status", 16),
            ("Priority", 12), ("Complexity", 12),
            ("Start", 13), ("Deadline", 13), ("Trễ (ngày)", 12), ("Lý do", 48),
        ]
        title = "🟠 TASK CHƯA CÓ PIC PHỤ TRÁCH"
    _write_banner(ws, title, len(items), "unassigned", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        rows.append([
            idx + 1,
            it.get("ma_cn", ""), it.get("rlog_id", ""), it.get("ten_cn", ""),
            it.get("module", ""),
            process_code(it),
            it.get("predecessor_phase", ""),
            it.get("phase", ""), it.get("status", ""),
            it.get("priority", ""), it.get("complexity", ""),
            it.get("start_date", ""),
            it.get("end_date", ""), it.get("days_overdue", 0),
            reason_unassigned(it),
        ])

    def _fill(i):
        it = items[i]
        if it.get("is_overdue"):
            return RED_FILL
        if "Must" in (it.get("priority") or ""):
            return ORANGE_FILL
        return None

    _write_data(ws, rows, row_fill_fn=_fill, n_cols=len(cols))
    _finalize_sheet(ws, len(rows), len(cols))


def _write_stalled_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("stalled", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Code", 14), ("Function Name", 40), ("Module", 12),
            ("Process", 22),
            ("Completed phase", 18), ("Waiting phase", 18),
            ("Completed on", 13), ("Wait end", 13), ("Wait days", 12),
            ("Priority", 12), ("Reason", 48),
        ]
        title = "🟡 STALLED TASKS (BETWEEN PHASES)"
    else:
        cols = [
            ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 12),
            ("Quy trình", 22),
            ("Phase đã xong", 18), ("Phase chờ", 18),
            ("Xong ngày", 13), ("End phase chờ", 13), ("Số ngày chờ", 12),
            ("Priority", 12), ("Lý do", 48),
        ]
        title = "🟡 TASK ĐÌNH TRỆ (KẸT GIỮA 2 PHASE)"
    _write_banner(ws, title, len(items), "stalled", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        rows.append([
            idx + 1,
            it.get("ma_cn", ""), it.get("ten_cn", ""), it.get("module", ""),
            process_code(it),
            it.get("completed_phase", ""), it.get("waiting_phase", ""),
            it.get("completed_date", ""), it.get("waiting_end_date", ""),
            it.get("wait_days", 0),
            it.get("priority", ""),
            reason_stalled(it),
        ])
    _write_data(
        ws, rows,
        row_fill_fn=lambda i: _fill_by_days(items[i].get("wait_days", 0)),
        n_cols=len(cols),
    )
    _finalize_sheet(ws, len(rows), len(cols))


def _write_risk_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("high_risk", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Risk Score", 12), ("Code", 14),
            ("Function Name", 40), ("Module", 12), ("Process", 22),
            ("Priority", 12), ("Complexity", 12), ("Risk Factors", 50),
            ("Detailed factors", 60),
        ]
        title = "🔺 HIGH RISK FUNCTIONS (≥30)"
    else:
        cols = [
            ("STT", 6), ("Risk Score", 12), ("Mã CN", 14),
            ("Tên chức năng", 40), ("Module", 12), ("Quy trình", 22),
            ("Priority", 12), ("Complexity", 12), ("Risk Factors", 50),
            ("Yếu tố chi tiết", 60),
        ]
        title = "🔺 FUNCTION CÓ ĐIỂM RỦI RO CAO (≥30)"
    _write_banner(ws, title, len(items), "high_risk", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        rows.append([
            idx + 1, it.get("risk_score", 0),
            it.get("ma_cn", ""), it.get("ten_cn", ""), it.get("module", ""),
            process_code(it),
            it.get("priority", ""), it.get("complexity", ""),
            " | ".join(it.get("risk_factors", [])),
            format_risk_factors_detailed(it),
        ])
    _write_data(
        ws, rows,
        row_fill_fn=lambda i: _fill_by_risk(items[i].get("risk_score", 0)),
        n_cols=len(cols),
    )
    _finalize_sheet(ws, len(rows), len(cols))


def _write_aging_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("aging_wip", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Code", 14), ("Function Name", 40), ("Module", 12),
            ("Process", 20), ("Phase", 16), ("Status", 13),
            ("Start", 13), ("End (plan)", 13), ("PIC", 22),
            ("Aging (days)", 12), ("Threshold", 10), ("Over by (days)", 12),
            ("Priority", 12), ("Complexity", 12), ("Reason", 40),
        ]
        title = "🟡 AGING WIP (In-progress too long)"
    else:
        cols = [
            ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 12),
            ("Quy trình", 20), ("Phase", 16), ("Status", 13),
            ("Start", 13), ("End (plan)", 13), ("PIC", 22),
            ("Aging (ngày)", 12), ("Ngưỡng", 10), ("Over by (ngày)", 12),
            ("Priority", 12), ("Complexity", 12), ("Lý do", 40),
        ]
        title = "🟡 TASK AGING WIP (In-progress quá lâu)"
    _write_banner(ws, title, len(items), "aging", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        thr = it.get("threshold_days")
        rows.append([
            idx + 1,
            it.get("ma_cn", ""), it.get("ten_cn", ""), it.get("module", ""),
            it.get("quy_trinh", ""), it.get("phase", ""), it.get("status", ""),
            it.get("start_date", ""), it.get("end_date", ""),
            ", ".join(it.get("pic", []) if isinstance(it.get("pic"), list) else [str(it.get("pic") or "")]),
            it.get("aging_days", 0), thr if thr is not None else "",
            it.get("over_by_days", 0),
            it.get("priority", ""), it.get("complexity", ""),
            reason_aging_wip(it, thr),
        ])
    _write_data(
        ws, rows,
        row_fill_fn=lambda i: _fill_by_days(items[i].get("aging_days", 0)),
        n_cols=len(cols),
    )
    _finalize_sheet(ws, len(rows), len(cols))


def _write_dq_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("data_quality", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Code", 14), ("Function Name", 32), ("Module", 12),
            ("Process", 22),
            ("Phase", 14), ("Issue code", 22), ("Severity", 8),
            ("Label", 40), ("Detail", 40), ("Suggestion", 30),
        ]
        title = "⚫ DATA QUALITY ISSUES"
    else:
        cols = [
            ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 32), ("Module", 12),
            ("Quy trình", 22),
            ("Phase", 14), ("Mã lỗi", 22), ("Mức", 8),
            ("Mô tả", 40), ("Chi tiết", 40), ("Gợi ý", 30),
        ]
        title = "⚫ DATA QUALITY — LỖI DỮ LIỆU"
    _write_banner(ws, title, len(items), "data_quality", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        rows.append([
            idx + 1,
            it.get("ma_cn", ""), it.get("ten_cn", ""), it.get("module", ""),
            process_code(it),
            it.get("phase", ""), it.get("code", ""), it.get("severity", ""),
            it.get("label", ""), it.get("detail", ""), it.get("suggestion", ""),
        ])

    def _fill(i):
        sev = (items[i].get("severity") or "").lower()
        if sev == "high":
            return RED_FILL
        if sev == "medium":
            return ORANGE_FILL
        if sev == "low":
            return YELLOW_FILL
        return None

    _write_data(ws, rows, row_fill_fn=_fill, n_cols=len(cols))
    _finalize_sheet(ws, len(rows), len(cols))


def _write_bookmark_sheet(wb, items: list[dict], *, lang: str = "vi", name: str = "") -> None:
    ws = wb.create_sheet(name or _sn("bookmark", lang))
    if lang == "en":
        cols = [
            ("#", 6), ("Code", 14), ("Function Name", 40), ("Module", 12),
            ("Process", 20), ("Priority", 12), ("Complexity", 12),
            ("Phase group", 16), ("FIT/GAP", 10),
        ]
        title = "🟣 BOOKMARKED FUNCTIONS"
    else:
        cols = [
            ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 12),
            ("Quy trình", 20), ("Priority", 12), ("Complexity", 12),
            ("Giai đoạn", 16), ("FIT/GAP", 10),
        ]
        title = "🟣 FUNCTION ĐÃ BOOKMARK"
    _write_banner(ws, title, len(items), "bookmark", len(cols), lang=lang)
    _write_header(ws, cols)

    rows = []
    for idx, it in enumerate(items):
        rows.append([
            idx + 1,
            it.get("ma_cn", ""), it.get("ten_cn", ""), it.get("module", ""),
            it.get("quy_trinh", ""), it.get("priority", ""),
            it.get("complexity", ""), it.get("giai_doan", ""),
            it.get("fit_gap", ""),
        ])
    _write_data(ws, rows, n_cols=len(cols))
    _finalize_sheet(ws, len(rows), len(cols))
