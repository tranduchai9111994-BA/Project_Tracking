"""
Xuất Function List re-import — chỉ các CN dính issue.

Nguồn issue: overdue / unassigned / stalled / anomalies (+ missing_deadline nếu có).
Định dạng cột theo schema mẫu (fl_export_schema) hoặc header ParsedData hiện tại.

Sheet ``Function List`` (import được luôn) + sheet phụ ``Ly_do`` (review).
Không ghi text note / Remark tracker vào FL; chỉ tô màu:
  - Vàng: PIC/Status bất thường / cần sửa
  - Xanh nhạt: From/Start auto-fill = To/End phase trước + 1 ngày làm việc (bỏ T7/CN)

Text ô (PIC/Status/tên…): 1 dòng — bỏ ``\\n`` thừa, không wrap.
"""
from __future__ import annotations

import os
import re
from datetime import date, datetime
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from exporter.excel_exporter import (
    BODY_FONT,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FONT,
    THIN_BORDER,
    _write_sheet,
)
from exporter.fl_export_schema import (
    next_business_day,
    resolve_export_schema,
)
from exporter.reason_formatters import (
    join_notes,
    process_code,
    reason_overdue,
    reason_stalled,
    reason_unassigned,
)
from parser.excel_parser import FunctionRow, ParsedData, PhaseData, VALID_STATUSES

YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
# Body: 1 dòng — không wrap (tránh PIC/Status/tên bị xuống dòng xấu khi import)
FL_BODY_ALIGN = Alignment(vertical="center", wrap_text=False)
_NL_RE = re.compile(r"[\r\n\u2028\u2029]+")
_MULTI_SPACE_RE = re.compile(r"[ \t]{2,}")


# ------------------------------------------------------------------
# Collect issue hits (union theo ma_cn)
# ------------------------------------------------------------------

def collect_issue_hits(
    *,
    overdue_list: Optional[list[dict]] = None,
    unassigned_list: Optional[list[dict]] = None,
    stalled_list: Optional[list[dict]] = None,
    anomaly_issues: Optional[list[dict]] = None,
) -> dict[str, dict[str, Any]]:
    """
    Union theo ma_cn.

    Return: ma_cn → {
      kinds: set[str],
      yellow_pic: set[phase],
      yellow_status: set[phase],
      notes: list[str],
    }
    """
    hits: dict[str, dict[str, Any]] = {}

    def _ensure(ma: str) -> dict[str, Any]:
        ma = str(ma or "").strip()
        if not ma:
            return {}
        if ma not in hits:
            hits[ma] = {
                "kinds": set(),
                "yellow_pic": set(),
                "yellow_status": set(),
                "notes": [],
            }
        return hits[ma]

    for it in overdue_list or []:
        ma = it.get("ma_cn") or ""
        h = _ensure(ma)
        if not h:
            continue
        ph = it.get("phase") or ""
        h["kinds"].add("overdue")
        if ph:
            h["yellow_status"].add(ph)
        h["notes"].append(reason_overdue(it) if it.get("end_date") or it.get("status") else (
            f"Overdue {ph}" + (
                f" ({it.get('days_overdue')}d)" if it.get("days_overdue") is not None else ""
            )
        ))

    for it in unassigned_list or []:
        ma = it.get("ma_cn") or ""
        h = _ensure(ma)
        if not h:
            continue
        ph = it.get("phase") or ""
        h["kinds"].add("unassigned")
        if ph:
            h["yellow_pic"].add(ph)
        h["notes"].append(reason_unassigned(it))

    for it in stalled_list or []:
        ma = it.get("ma_cn") or ""
        h = _ensure(ma)
        if not h:
            continue
        wait = it.get("waiting_phase") or ""
        done = it.get("completed_phase") or ""
        h["kinds"].add("stalled")
        if wait:
            h["yellow_status"].add(wait)
            h["yellow_pic"].add(wait)
        h["notes"].append(reason_stalled(it) if done or wait else (
            f"Đình trệ {done}→{wait}"
            + (f" ({it.get('wait_days')}d)" if it.get("wait_days") else "")
        ))

    for it in anomaly_issues or []:
        ma = it.get("ma_cn") or ""
        h = _ensure(ma)
        if not h:
            continue
        code = it.get("code") or "anomaly"
        ph = it.get("phase") or ""
        h["kinds"].add(f"anomaly:{code}")
        label = it.get("label") or code
        if ph:
            # Status/date anomalies → status; blank_pic → pic
            if code in ("blank_pic",):
                h["yellow_pic"].add(ph)
            elif code in (
                "invalid_status", "end_before_start", "closed_no_end",
                "missing_deadline", "phase_overlap", "estimate_vs_duration",
            ):
                h["yellow_status"].add(ph)
            h["notes"].append(f"{label} · {ph}")
        else:
            h["notes"].append(str(label))

    return hits


# ------------------------------------------------------------------
# Date chain
# ------------------------------------------------------------------

def compute_date_chain_fills(
    row: FunctionRow,
    phase_order: list[str],
) -> dict[str, date]:
    """
    phase_name → ngày Start/From cần auto-fill.

    Chỉ khi Start trống và phase trước có End; không đè ngày đã có.
    """
    fills: dict[str, date] = {}
    for i in range(1, len(phase_order)):
        prev_name = phase_order[i - 1]
        curr_name = phase_order[i]
        prev_pd = row.phases.get(prev_name)
        curr_pd = row.phases.get(curr_name)
        if not prev_pd or not prev_pd.end_date:
            continue
        if curr_pd and curr_pd.start_date:
            continue  # đã có — không đè
        nxt = next_business_day(prev_pd.end_date)
        if nxt:
            fills[curr_name] = nxt
    return fills


def format_fl_date(d: date) -> str:
    """Format dd/MM/yyyy — khớp mẫu DanhSachFunction."""
    return d.strftime("%d/%m/%Y")


# ------------------------------------------------------------------
# Row value reconstruction
# ------------------------------------------------------------------

def _headers_from_schema(schema: dict[str, Any], data: ParsedData) -> list[str]:
    headers = list(schema.get("headers") or [])
    if headers:
        return headers
    return [h for h, _ in sorted(data.headers.items(), key=lambda x: x[1])]


def _inv_headers(data: ParsedData) -> dict[int, str]:
    return {idx: h for h, idx in data.headers.items()}


def row_values_for_export(
    row: FunctionRow,
    data: ParsedData,
    headers: list[str],
    *,
    date_fills: Optional[dict[str, date]] = None,
    source_values: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Map header → value; ưu tiên source_values (copy từ file gốc) nếu có."""
    values: dict[str, Any] = {h: "" for h in headers}
    if source_values:
        for h, v in source_values.items():
            if h in values and v is not None:
                values[h] = v

    inv = _inv_headers(data)
    for meta_key, col in (data.meta_columns or {}).items():
        if not col:
            continue
        h = inv.get(col)
        if h and h in values:
            # Giữ source nếu đã có; chỉ fill khi trống
            if values[h] in ("", None) or source_values is None:
                values[h] = row.meta.get(meta_key, "")

    for pg in data.phase_groups:
        pd = row.phases.get(pg.name) or PhaseData()
        for attr, col in (pg.attributes or {}).items():
            h = inv.get(col)
            if not h or h not in values:
                # Schema header có thể khác ParsedData — match theo "Phase - Attr"
                h = f"{pg.name} - {attr}"
                if h not in values:
                    # PIC FPT / aliases
                    candidates = [hh for hh in headers if hh.startswith(f"{pg.name} - ") and attr.upper() in hh.upper()]
                    h = candidates[0] if candidates else None
            if not h or h not in values:
                continue
            au = attr.upper()
            # Không đè giá trị đã copy từ source (trừ date fill)
            existing = values.get(h)
            has_src = source_values is not None and h in (source_values or {}) and source_values.get(h) not in (None, "")

            if au in ("START", "FROM", "PLANNED"):
                if date_fills and pg.name in date_fills and not pd.start_date:
                    values[h] = format_fl_date(date_fills[pg.name])
                elif not has_src:
                    values[h] = format_fl_date(pd.start_date) if pd.start_date else (existing or "")
            elif au in ("END", "TO", "ACTUAL"):
                if not has_src:
                    values[h] = format_fl_date(pd.end_date) if pd.end_date else (existing or "")
            elif "STATUS" in au:
                if not has_src:
                    values[h] = pd.status or existing or ""
            elif "PIC" in au:
                if not has_src:
                    values[h] = ", ".join(pd.pics or []) if pd.pics else (existing or "")
            elif "ESTIMATE" in au or au.endswith("MH"):
                if not has_src:
                    values[h] = pd.estimate_mh if pd.estimate_mh is not None else (existing or "")
            elif "NOTE" in au:
                if not has_src:
                    values[h] = pd.note or existing or ""
            else:
                if not has_src:
                    values[h] = (pd.extra or {}).get(attr, existing or "")
    return values


def _pic_status_header_map(
    data: ParsedData,
    headers: list[str],
    schema: Optional[dict] = None,
) -> dict[str, dict[str, str]]:
    """header → {phase, kind: pic|status|start}."""
    out: dict[str, dict[str, str]] = {}
    # Từ ParsedData
    inv = _inv_headers(data)
    for pg in data.phase_groups:
        for attr, col in (pg.attributes or {}).items():
            h = inv.get(col)
            if not h:
                continue
            au = attr.upper()
            if "PIC" in au:
                out[h] = {"phase": pg.name, "kind": "pic"}
            elif "STATUS" in au:
                out[h] = {"phase": pg.name, "kind": "status"}
            elif au in ("START", "FROM", "PLANNED"):
                out[h] = {"phase": pg.name, "kind": "start"}
    # Bổ sung từ schema headers nếu thiếu
    for h in headers:
        if h in out or " - " not in h:
            continue
        phase, attr = h.rsplit(" - ", 1)
        au = attr.upper()
        if "PIC" in au:
            out[h] = {"phase": phase, "kind": "pic"}
        elif "STATUS" in au:
            out[h] = {"phase": phase, "kind": "status"}
        elif au in ("START", "FROM", "PLANNED"):
            out[h] = {"phase": phase, "kind": "start"}
    return out


def _flatten_cell_text(val: Any) -> Any:
    """Chuỗi 1 dòng: bỏ xuống dòng / khoảng trắng thừa (PIC, Status, tên…)."""
    if not isinstance(val, str):
        return val
    s = _NL_RE.sub(" ", val)
    s = _MULTI_SPACE_RE.sub(" ", s).strip()
    return s


def _needs_status_yellow(status: Any) -> bool:
    s = str(status or "").strip()
    if not s:
        return True
    if s not in VALID_STATUSES:
        return True
    return False


def _needs_pic_yellow(val: Any) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    return not s


# ------------------------------------------------------------------
# Source workbook row lookup (giữ cột lạ: FID, RlogID…)
# ------------------------------------------------------------------

def _load_source_rows_by_ma(
    source_xlsx: Optional[str],
    ma_cn_header_candidates: Optional[list[str]] = None,
) -> tuple[list[str], dict[str, dict[str, Any]]]:
    """
    Đọc file nguồn → (headers, ma_cn → {header: value}).
    """
    if not source_xlsx or not os.path.isfile(source_xlsx):
        return [], {}
    try:
        wb = openpyxl.load_workbook(source_xlsx, data_only=True, read_only=True)
    except Exception:
        return [], {}
    try:
        ws = wb["Function List"] if "Function List" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if not raw_headers:
            return [], {}
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]
        # Tìm cột Mã CN
        ma_col = None
        candidates = ma_cn_header_candidates or ["Mã CN", "Mã chức năng", "Function Code", "Code"]
        for i, h in enumerate(headers):
            if h in candidates or any(c.lower() == h.lower() for c in candidates):
                ma_col = i
                break
        if ma_col is None:
            for i, h in enumerate(headers):
                hl = h.lower()
                if "mã cn" in hl or "ma cn" in hl or hl == "code":
                    ma_col = i
                    break
        if ma_col is None:
            return headers, {}

        by_ma: dict[str, dict[str, Any]] = {}
        for row in rows_iter:
            if not row or ma_col >= len(row):
                continue
            ma = str(row[ma_col] or "").strip()
            if not ma:
                continue
            vals = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                vals[h] = row[i] if i < len(row) else None
            by_ma[ma] = vals
        return headers, by_ma
    finally:
        wb.close()


# ------------------------------------------------------------------
# Main export
# ------------------------------------------------------------------

def export_fl_reimport(
    data: ParsedData,
    *,
    hits: dict[str, dict[str, Any]],
    output_dir: str = "uploads",
    project_dir: Optional[str] = None,
    source_xlsx: Optional[str] = None,
    project_slug: str = "",
    schema: Optional[dict[str, Any]] = None,
) -> str:
    """
    Xuất workbook Function List (+ sheet phụ Ly_do) — chỉ row trong ``hits``.
    Chỉ tô màu (vàng/xanh) trên FL; không ghi Remark tracker.
    Sheet ``Ly_do`` chứa kinds + notes từ collect_issue_hits để BA/PM review.

    Returns: filepath
    """
    if schema is None:
        schema = resolve_export_schema(
            project_dir or "",
            fallback_data=data,
            fallback_xlsx=source_xlsx,
        )

    headers = _headers_from_schema(schema, data)
    # Nếu schema headers rỗng / lệch — ưu tiên header file nguồn
    src_headers, src_by_ma = _load_source_rows_by_ma(source_xlsx)
    if src_headers and (not headers or set(headers) != set(h for h in src_headers if h)):
        # Dùng schema nếu đã lưu có headers; ngược lại lấy từ source
        if not schema.get("headers"):
            headers = [h for h in src_headers if h]

    col_map = _pic_status_header_map(data, headers, schema)
    phase_order = list(data.all_phases) or [pg.name for pg in data.phase_groups]

    rows_by_ma = {
        str(r.meta.get("ma_cn") or "").strip(): r
        for r in data.rows
        if str(r.meta.get("ma_cn") or "").strip()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"

    # Header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = min(28, max(10, len(h) + 2))

    out_row = 2
    for ma in sorted(hits.keys()):
        row = rows_by_ma.get(ma)
        if row is None:
            continue
        hit = hits[ma]
        date_fills = compute_date_chain_fills(row, phase_order)
        src_vals = src_by_ma.get(ma)
        values = row_values_for_export(
            row, data, headers,
            date_fills=date_fills,
            source_values=src_vals,
        )

        yellow_pic = hit.get("yellow_pic") or set()
        yellow_status = hit.get("yellow_status") or set()

        for col, h in enumerate(headers, 1):
            val = values.get(h, "")
            if isinstance(val, date) and not isinstance(val, datetime):
                val = format_fl_date(val)
            elif isinstance(val, datetime):
                val = format_fl_date(val.date())
            val = _flatten_cell_text(val if val is not None else "")
            cell = ws.cell(out_row, col, val)
            cell.font = BODY_FONT
            cell.alignment = FL_BODY_ALIGN
            cell.border = THIN_BORDER

            info = col_map.get(h)
            fill = None
            if info:
                ph = info["phase"]
                kind = info["kind"]
                if kind == "start" and ph in date_fills:
                    # Auto-fill xanh — chỉ khi thực sự fill (Start trống trước đó)
                    fill = BLUE_FILL
                elif kind == "pic" and ph in yellow_pic and _needs_pic_yellow(val):
                    fill = YELLOW_FILL
                elif kind == "status" and ph in yellow_status and _needs_status_yellow(val):
                    fill = YELLOW_FILL
                elif kind == "pic" and ph in yellow_pic:
                    fill = YELLOW_FILL  # vẫn tô khi flagged dù có giá trị (stalled)
                elif kind == "status" and ph in yellow_status:
                    fill = YELLOW_FILL
            if fill:
                cell.fill = fill

        out_row += 1

    ws.freeze_panes = "A2"
    if out_row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{out_row - 1}"

    # Sheet phụ Lý do — không đụng template FL / Remark
    ly_rows = []
    for ma in sorted(hits.keys()):
        row = rows_by_ma.get(ma)
        hit = hits[ma]
        kinds = sorted(hit.get("kinds") or [])
        ly_rows.append([
            ma,
            (row.meta.get("ten_cn") if row else "") or "",
            (row.meta.get("module") if row else "") or "",
            process_code(row.meta if row else {}),
            ", ".join(kinds),
            join_notes(list(hit.get("notes") or [])),
        ])
    ws_ly = wb.create_sheet("Ly_do")
    _write_sheet(
        ws_ly,
        title="LÝ DO ISSUE — FL RE-IMPORT (không ghi vào Remark)",
        columns=[
            ("Mã CN", 16),
            ("Tên chức năng", 36),
            ("Module", 12),
            ("Quy trình", 22),
            ("Loại issue", 28),
            ("Lý do / Ghi chú", 60),
        ],
        data_rows=ly_rows,
        subtitle=(
            f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}  |  "
            f"{len(ly_rows)} CN  |  Sheet Function List giữ nguyên template tô vàng"
        ),
    )

    os.makedirs(output_dir, exist_ok=True)
    slug_part = project_slug or "project"
    fname = f"FL_Reimport_{slug_part}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, fname)
    wb.save(filepath)
    wb.close()
    return filepath
