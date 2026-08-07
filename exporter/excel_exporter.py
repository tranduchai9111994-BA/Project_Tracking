"""
Xuất báo cáo Excel:
- export_overdue_report: danh sách task trễ (single sheet)
- export_full_report:    báo cáo tổng hợp nhiều sheet (Overdue / Unassigned / Long Duration / Stalled / High Risk / Summary)
- export_by_pic:         báo cáo riêng cho 1 PIC
- export_weekly_mom:     báo cáo tuần MoM (mẫu W30) + PM Dashboard — xem exporter/weekly_mom.py
Format chuyên nghiệp, highlight theo mức trễ / risk score.
"""
import os
from datetime import date
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from exporter.reason_formatters import (
    format_risk_factors_detailed,
    process_code,
    reason_aging_wip,
    reason_capacity,
    reason_duration,
    reason_fitgap_aging,
    reason_overdue,
    reason_stalled,
    reason_unassigned,
)


# === Styles chung ===
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(vertical="center", wrap_text=True)

RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")     # nặng
ORANGE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # trung
YELLOW_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # nhẹ
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")   # ok

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_sheet(
    ws,
    title: str,
    columns: list[tuple[str, int]],
    data_rows: list[list[Any]],
    row_fill_fn=None,
    subtitle: str = "",
) -> None:
    """
    Helper: viết 1 sheet chuẩn (title + subtitle + header + data + summary + auto-filter + freeze).

    Args:
        columns: list of (col_name, col_width)
        data_rows: list of value lists tương ứng columns
        row_fill_fn: optional callable(row_idx, item_index) -> PatternFill|None
    """
    n_cols = len(columns)
    last_col_letter = get_column_letter(n_cols)

    # Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    tc = ws["A1"]
    tc.value = title
    tc.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(f"A2:{last_col_letter}2")
    sc = ws["A2"]
    sc.value = subtitle or f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    sc.font = Font(name="Arial", size=10, italic=True, color="666666")
    sc.alignment = Alignment(horizontal="center")

    # Header row (row 4)
    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 25

    # Data rows
    for row_offset, values in enumerate(data_rows):
        row_idx = header_row + 1 + row_offset
        fill = row_fill_fn(row_idx, row_offset) if row_fill_fn else None
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    # Summary
    summary_row = header_row + len(data_rows) + 2
    ws.merge_cells(f"A{summary_row}:{last_col_letter}{summary_row}")
    scell = ws.cell(row=summary_row, column=1)
    scell.value = f"Tổng: {len(data_rows)} record | Xuất ngày {date.today().strftime('%d/%m/%Y')}"
    scell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")

    # Freeze + auto-filter
    ws.freeze_panes = f"A{header_row + 1}"
    if data_rows:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row + len(data_rows)}"

COLUMNS = [
    ("STT", 6),
    ("Mã CN", 14),
    ("Tên chức năng", 40),
    ("Module", 10),
    ("Quy trình", 22),
    ("Phase bị trễ", 16),
    ("Deadline", 13),
    ("Số ngày trễ", 12),
    ("Trạng thái", 13),
    ("PIC phụ trách", 20),
    ("Priority", 12),
    ("Ghi chú", 30),
    ("Lý do", 48),
]

EXPORT_MODES = {"summary", "detail", "both"}


def _normalize_export_mode(mode: str | None) -> str:
    m = (mode or "both").strip().lower()
    return m if m in EXPORT_MODES else "both"


def _want_summary(mode: str) -> bool:
    return mode in ("summary", "both")


def _want_detail(mode: str) -> bool:
    return mode in ("detail", "both")


class _SheetBook:
    """Workbook helper: sheet đầu = rename active; sheet sau = create."""

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self._first = True

    def sheet(self, name: str):
        if self._first:
            ws = self.wb.active
            ws.title = name
            self._first = False
            return ws
        return self.wb.create_sheet(name)


def _norm_multi_filter(v) -> list[str]:
    if not v:
        return []
    if isinstance(v, (list, tuple)):
        return [str(x).strip() for x in v if str(x).strip()]
    return [x.strip() for x in str(v).split(",") if x.strip()]


def _fill_by_days(days: int) -> Optional[PatternFill]:
    """Chọn màu highlight theo số ngày."""
    if days > 14:
        return RED_FILL
    if days > 7:
        return ORANGE_FILL
    if days > 0:
        return YELLOW_FILL
    return None


def _fill_by_risk(score: int) -> Optional[PatternFill]:
    if score >= 80:
        return RED_FILL
    if score >= 50:
        return ORANGE_FILL
    if score >= 30:
        return YELLOW_FILL
    return None


def export_overdue_report(
    overdue_list: list[dict[str, Any]],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
    mode: str = "both",
) -> str:
    """
    Tạo file Excel task trễ — sheet Tong_hop + Chi_tiet theo mode.

    mode: summary | detail | both (default both)
    """
    from collections import Counter

    mode = _normalize_export_mode(mode)
    items = list(overdue_list or [])
    if filters:
        mods = _norm_multi_filter(filters.get("module"))
        if mods:
            items = [i for i in items if i.get("module") in mods]
        if filters.get("pic"):
            items = [i for i in items if filters["pic"] in i.get("pic", [])]
        phases = _norm_multi_filter(filters.get("phase"))
        if phases:
            items = [i for i in items if i.get("phase") in phases]

    filter_text = ""
    if filters:
        parts = []
        mods = _norm_multi_filter(filters.get("module"))
        if mods:
            parts.append(f"Module: {', '.join(mods)}")
        if filters.get("pic"):
            parts.append(f"PIC: {filters['pic']}")
        phases = _norm_multi_filter(filters.get("phase"))
        if phases:
            parts.append(f"Phase: {', '.join(phases)}")
        filter_text = " | ".join(parts)
    subtitle = f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    if filter_text:
        subtitle += f"  |  Bộ lọc: {filter_text}"

    book = _SheetBook()
    if _want_summary(mode):
        mod_c = Counter((i.get("module") or "(blank)") for i in items)
        ph_c = Counter((i.get("phase") or "(blank)") for i in items)
        sum_rows = (
            [[idx + 1, f"Module: {k}", v]
             for idx, (k, v) in enumerate(sorted(mod_c.items(), key=lambda x: -x[1]))]
            + [[len(mod_c) + idx + 1, f"Phase: {k}", v]
               for idx, (k, v) in enumerate(sorted(ph_c.items(), key=lambda x: -x[1]))]
        )
        _write_sheet(
            book.sheet("Tong_hop"),
            "TỔNG HỢP OVERDUE (theo Module / Phase)",
            [("STT", 6), ("Nhóm", 28), ("Số lượng", 12)],
            sum_rows,
            subtitle=subtitle,
        )
    if _want_detail(mode):
        data_rows = [
            [
                idx + 1,
                i.get("ma_cn", ""),
                i.get("ten_cn", ""),
                i.get("module", ""),
                process_code(i),
                i.get("phase", ""),
                i.get("end_date", ""),
                i.get("days_overdue", 0),
                i.get("status", ""),
                ", ".join(i.get("pic", []) if isinstance(i.get("pic"), list) else []),
                i.get("priority", ""),
                i.get("note", ""),
                reason_overdue(i),
            ]
            for idx, i in enumerate(items)
        ]
        _write_sheet(
            book.sheet("Chi_tiet"),
            "CHI TIẾT TASK TRỄ DEADLINE",
            COLUMNS,
            data_rows,
            subtitle=subtitle,
            row_fill_fn=lambda ri, idx: _fill_by_days(items[idx].get("days_overdue", 0)),
        )

    os.makedirs(output_dir, exist_ok=True)
    filename = f"Overdue_Report_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    book.wb.save(filepath)
    book.wb.close()
    return filepath


def export_stalled_report(
    stalled_items: list[dict[str, Any]],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
    mode: str = "both",
) -> str:
    """
    Tạo file Excel đình trệ — sheet Tong_hop + Chi_tiet theo mode.

    mode: summary | detail | both (default both)
    """
    from collections import Counter

    mode = _normalize_export_mode(mode)
    items = list(stalled_items or [])
    mods = _norm_multi_filter(filters.get("module")) if filters else []
    # Filter Phase chờ — khớp cột "Phase chờ" trên dashboard, để file xuất ra
    # không nhiều hơn số dòng PM đang nhìn thấy.
    waiting = _norm_multi_filter(filters.get("waiting_phase")) if filters else []
    if mods:
        items = [i for i in items if i.get("module") in mods]
    if waiting:
        items = [i for i in items if i.get("waiting_phase") in waiting]

    filter_parts = []
    if mods:
        filter_parts.append(f"Module: {', '.join(mods)}")
    if waiting:
        filter_parts.append(f"Phase chờ: {', '.join(waiting)}")
    subtitle = f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    if filter_parts:
        subtitle += "  |  Bộ lọc: " + " | ".join(filter_parts)

    book = _SheetBook()
    if _want_summary(mode):
        c = Counter(
            f"{i.get('completed_phase', '')} → {i.get('waiting_phase', '')}"
            for i in items
        )
        _write_sheet(
            book.sheet("Tong_hop"),
            "TỔNG HỢP ĐÌNH TRỆ (theo transition)",
            [("STT", 6), ("Từ → Sang", 36), ("Số lượng", 12)],
            [[idx + 1, k, v] for idx, (k, v) in enumerate(sorted(c.items(), key=lambda x: -x[1]))],
            subtitle=subtitle,
        )
    if _want_detail(mode):
        columns = [
            ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
            ("Quy trình", 22),
            ("Phase đã xong", 16), ("Phase chờ", 16), ("Xong ngày", 13),
            ("End phase chờ", 13), ("Chờ (ngày)", 12), ("Priority", 12),
            ("Lý do", 48),
        ]
        data_rows = [
            [
                idx + 1,
                i.get("ma_cn", ""),
                i.get("ten_cn", ""),
                i.get("module", ""),
                process_code(i),
                i.get("completed_phase", ""),
                i.get("waiting_phase", ""),
                i.get("completed_date", ""),
                i.get("waiting_end_date", ""),
                i.get("wait_days", 0),
                i.get("priority", ""),
                reason_stalled(i),
            ]
            for idx, i in enumerate(items)
        ]
        _write_sheet(
            book.sheet("Chi_tiet"),
            "CHI TIẾT TASK ĐÌNH TRỆ",
            columns,
            data_rows,
            subtitle=subtitle,
            row_fill_fn=lambda ri, idx: _fill_by_days(items[idx].get("wait_days", 0)),
        )

    os.makedirs(output_dir, exist_ok=True)
    filename = f"Dinh_Tre_Report_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    book.wb.save(filepath)
    book.wb.close()
    return filepath


# ======================================================================
# V2 EXPORTS
# ======================================================================


def export_full_report(
    metrics: dict,
    output_dir: str = "uploads",
) -> str:
    """
    Xuất báo cáo tổng hợp nhiều sheet:
    - Summary
    - Overdue_Report
    - Unassigned_Tasks
    - Long_Duration
    - Dinh_Tre
    - High_Risk
    """
    wb = openpyxl.Workbook()
    # Xóa sheet default
    wb.remove(wb.active)

    summary = metrics.get("summary", {})
    overdue_list = metrics.get("overdue_list", [])
    unassigned = metrics.get("unassigned_tasks", [])
    duration_items = metrics.get("duration_analysis", {}).get("items", [])
    stalled_items = metrics.get("stalled_tasks", {}).get("items", [])
    risk_scores = metrics.get("risk_scores", [])
    # Rule V4: xuất ALL high-risk có score >= 30, KHÔNG cắt top 100 như trước
    high_risk = [r for r in risk_scores if r["risk_score"] >= 30]

    # === Sheet 1: Summary ===
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    ws.merge_cells("A1:B1")
    ws["A1"] = "BÁO CÁO TỔNG HỢP DỰ ÁN"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Tổng số chức năng", summary.get("total_functions", 0)),
        ("Số Module", summary.get("modules_count", 0)),
        ("Số Phase", summary.get("phases_count", 0)),
        ("Tiến độ chung (%)", summary.get("overall_progress_pct", 0)),
        ("Task trễ deadline", summary.get("total_overdue", 0)),
        ("Task chưa có PIC", summary.get("unassigned_count", 0)),
        ("Function high-risk (>=50)", summary.get("high_risk_count", 0)),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=4):
        ws.cell(row=idx, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=idx, column=1).border = THIN_BORDER
        c = ws.cell(row=idx, column=2, value=value)
        c.font = Font(name="Arial", size=11)
        c.alignment = Alignment(horizontal="right")
        c.border = THIN_BORDER

    # === Sheet 2: Overdue_Report ===
    ws = wb.create_sheet("Overdue_Report")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Quy trình", 22),
        ("Phase", 16), ("Deadline", 13), ("Số ngày trễ", 12),
        ("Trạng thái", 13), ("PIC", 20), ("Priority", 12), ("Ghi chú", 30),
        ("Lý do", 48),
    ]
    data_rows = [
        [
            idx + 1,
            i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
            process_code(i),
            i.get("phase", ""), i.get("end_date", ""), i.get("days_overdue", 0),
            i.get("status", ""), ", ".join(i.get("pic", [])),
            i.get("priority", ""), i.get("note", ""),
            reason_overdue(i),
        ]
        for idx, i in enumerate(overdue_list)
    ]
    _write_sheet(
        ws, "DANH SÁCH TASK TRỄ DEADLINE", columns, data_rows,
        row_fill_fn=lambda ri, idx: _fill_by_days(overdue_list[idx].get("days_overdue", 0)),
    )

    # === Sheet 3: Unassigned_Tasks ===
    ws = wb.create_sheet("Unassigned_Tasks")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Rlog ID", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Quy trình", 22), ("Phase trước", 16), ("Phase", 16), ("Trạng thái", 13),
        ("Priority", 12), ("Complexity", 12),
        ("Start", 13), ("Deadline", 13), ("Trễ (ngày)", 12), ("Lý do", 48),
    ]
    data_rows = [
        [
            idx + 1,
            i.get("ma_cn", ""), i.get("rlog_id", ""), i.get("ten_cn", ""), i.get("module", ""),
            process_code(i),
            i.get("predecessor_phase", ""),
            i.get("phase", ""), i.get("status", ""),
            i.get("priority", ""), i.get("complexity", ""),
            i.get("start_date", ""),
            i.get("end_date", ""), i.get("days_overdue", 0),
            reason_unassigned(i),
        ]
        for idx, i in enumerate(unassigned)
    ]

    def _unassigned_fill(ri: int, idx: int) -> Optional[PatternFill]:
        item = unassigned[idx]
        if item.get("is_overdue"):
            return RED_FILL
        if "Must" in (item.get("priority") or ""):
            return ORANGE_FILL
        return None

    _write_sheet(
        ws, "TASK CHƯA CÓ PIC PHỤ TRÁCH", columns, data_rows,
        row_fill_fn=_unassigned_fill,
    )

    # === Sheet 4: Long_Duration ===
    ws = wb.create_sheet("Long_Duration")
    dur_thr = metrics.get("duration_analysis", {}).get("threshold_days")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Quy trình", 22),
        ("Phase", 16), ("Start", 13), ("End", 13),
        ("Duration (ngày)", 14), ("Ngưỡng", 10), ("Loại", 10), ("Status", 12),
        ("PIC", 20), ("Priority", 12), ("Estimate MH", 12), ("Lý do", 40),
    ]
    data_rows = [
        [
            idx + 1,
            i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
            process_code(i),
            i.get("phase", ""), i.get("start_date", ""), i.get("end_date", ""),
            i.get("duration_days", 0),
            i.get("threshold_days", dur_thr),
            "Đang chạy" if i.get("duration_type") == "elapsed" else "Đã lên KH",
            i.get("status", ""),
            ", ".join(i.get("pic", [])),
            i.get("priority", ""), i.get("estimate_mh", ""),
            reason_duration(i, dur_thr),
        ]
        for idx, i in enumerate(duration_items)
    ]
    _write_sheet(
        ws, "TASK CÓ DURATION BẤT THƯỜNG", columns, data_rows,
        row_fill_fn=lambda ri, idx: _fill_by_days(duration_items[idx].get("duration_days", 0)),
    )

    # === Sheet 5: Dinh_Tre ===
    ws = wb.create_sheet("Dinh_Tre")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Quy trình", 22),
        ("Phase đã xong", 16), ("Phase chờ", 16), ("Xong ngày", 13),
        ("End phase chờ", 13), ("Chờ (ngày)", 12), ("Priority", 12), ("Lý do", 48),
    ]
    data_rows = [
        [
            idx + 1,
            i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
            process_code(i),
            i.get("completed_phase", ""), i.get("waiting_phase", ""),
            i.get("completed_date", ""), i.get("waiting_end_date", ""),
            i.get("wait_days", 0),
            i.get("priority", ""),
            reason_stalled(i),
        ]
        for idx, i in enumerate(stalled_items)
    ]
    _write_sheet(
        ws, "TASK BỊ ĐÌNH TRỆ (KẸT GIỮA 2 PHASE)", columns, data_rows,
        row_fill_fn=lambda ri, idx: _fill_by_days(stalled_items[idx].get("wait_days", 0)),
    )

    # === Sheet 6: High_Risk ===
    ws = wb.create_sheet("High_Risk")
    columns = [
        ("STT", 6), ("Risk Score", 12), ("Mã CN", 14),
        ("Tên chức năng", 40), ("Module", 10), ("Quy trình", 22),
        ("Priority", 12), ("Complexity", 12),
        ("Risk Factors", 45), ("Yếu tố chi tiết", 60),
    ]
    data_rows = [
        [
            idx + 1, i.get("risk_score", 0),
            i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
            process_code(i),
            i.get("priority", ""), i.get("complexity", ""),
            " | ".join(i.get("risk_factors", [])),
            format_risk_factors_detailed(i),
        ]
        for idx, i in enumerate(high_risk)
    ]
    _write_sheet(
        ws, "TOP FUNCTION CÓ ĐIỂM RỦI RO CAO", columns, data_rows,
        row_fill_fn=lambda ri, idx: _fill_by_risk(high_risk[idx].get("risk_score", 0)),
    )

    # === Save ===
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Full_Report_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


def export_by_pic(
    metrics: dict,
    pic_name: str,
    output_dir: str = "uploads",
) -> str:
    """
    Xuất Excel chứa toàn bộ task liên quan đến 1 PIC:
    - Sheet Overdue: task overdue của PIC này
    - Sheet Active: task đang In-progress / Assigned của PIC này
    - Sheet Effort: tổng MH gánh
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    overdue_all = metrics.get("overdue_list", [])
    duration_all = metrics.get("duration_analysis", {}).get("items", [])
    effort = metrics.get("effort_analysis", {})

    my_overdue = [i for i in overdue_all if pic_name in i.get("pic", [])]
    my_active = [i for i in duration_all if pic_name in i.get("pic", [])]

    pic_effort = next(
        (p for p in effort.get("by_pic", []) if p.get("pic") == pic_name),
        {"total_mh": 0, "closed_mh": 0, "remaining_mh": 0},
    )

    # Sheet 1: Info + Effort
    ws = wb.create_sheet("Info")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    ws.merge_cells("A1:B1")
    ws["A1"] = f"BÁO CÁO CÔNG VIỆC — {pic_name}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("PIC", pic_name),
        ("Task trễ deadline", len(my_overdue)),
        ("Task đang chạy/kéo dài", len(my_active)),
        ("Tổng Estimate MH", pic_effort.get("total_mh", 0)),
        ("MH đã Closed", pic_effort.get("closed_mh", 0)),
        ("MH còn lại", pic_effort.get("remaining_mh", 0)),
    ]
    for idx, (label, value) in enumerate(info_rows, start=4):
        ws.cell(row=idx, column=1, value=label).font = Font(name="Arial", bold=True)
        ws.cell(row=idx, column=1).border = THIN_BORDER
        c = ws.cell(row=idx, column=2, value=value)
        c.alignment = Alignment(horizontal="right")
        c.border = THIN_BORDER

    # Sheet 2: Overdue
    ws = wb.create_sheet("Overdue")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Quy trình", 22),
        ("Phase", 16), ("Deadline", 13), ("Số ngày trễ", 12),
        ("Trạng thái", 13), ("Priority", 12), ("Ghi chú", 30), ("Lý do", 48),
    ]
    data_rows = [
        [
            idx + 1,
            i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
            process_code(i),
            i.get("phase", ""), i.get("end_date", ""), i.get("days_overdue", 0),
            i.get("status", ""), i.get("priority", ""), i.get("note", ""),
            reason_overdue(i),
        ]
        for idx, i in enumerate(my_overdue)
    ]
    _write_sheet(
        ws, f"TASK TRỄ CỦA {pic_name}", columns, data_rows,
        row_fill_fn=lambda ri, idx: _fill_by_days(my_overdue[idx].get("days_overdue", 0)),
    )

    # Sheet 3: Active (đang In-progress / kéo dài)
    ws = wb.create_sheet("Active")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Phase", 16), ("Start", 13), ("End", 13),
        ("Duration (ngày)", 14), ("Status", 13), ("Priority", 12), ("Estimate MH", 12),
    ]
    data_rows = [
        [
            idx + 1,
            i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
            i.get("phase", ""), i.get("start_date", ""), i.get("end_date", ""),
            i.get("duration_days", 0), i.get("status", ""),
            i.get("priority", ""), i.get("estimate_mh", ""),
        ]
        for idx, i in enumerate(my_active)
    ]
    _write_sheet(ws, f"TASK ĐANG THỰC HIỆN — {pic_name}", columns, data_rows)

    # Save
    os.makedirs(output_dir, exist_ok=True)
    safe_pic = "".join(c for c in pic_name if c.isalnum() or c in ("_", "-"))
    filename = f"PIC_{safe_pic}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


def export_compare_report(
    compare_result: dict,
    old_date: str,
    new_date: str,
    output_dir: str = "uploads",
) -> str:
    """
    Xuất báo cáo so sánh 2 snapshot:
    - Sheet Summary
    - Sheet New_Functions
    - Sheet Removed_Functions
    - Sheet Status_Changes
    - Sheet Module_Delta
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # === Sheet 1: Summary ===
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    ws.merge_cells("A1:D1")
    ws["A1"] = f"BÁO CÁO SO SÁNH: {old_date} → {new_date}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Header
    headers = ["Chỉ tiêu", "Trước", "Sau", "Chênh lệch"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER

    rows = [
        ("Tổng chức năng",       compare_result.get("old_total"),        compare_result.get("new_total"),        compare_result.get("delta_total")),
        ("Tiến độ chung (%)",    compare_result.get("old_overall_pct"),  compare_result.get("new_overall_pct"),  compare_result.get("delta_pct")),
        ("Task trễ deadline",    compare_result.get("old_overdue"),      compare_result.get("new_overdue"),      compare_result.get("delta_overdue")),
    ]
    for idx, r in enumerate(rows, start=4):
        for cidx, val in enumerate(r, 1):
            c = ws.cell(row=idx, column=cidx, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center" if cidx > 1 else "left")

    # Velocity
    velocity = compare_result.get("velocity", {})
    ws.merge_cells(f"A{len(rows) + 6}:D{len(rows) + 6}")
    ws.cell(row=len(rows) + 6, column=1, value="VELOCITY").font = Font(name="Arial", bold=True, size=12, color="1F4E79")

    vrow = len(rows) + 7
    velocity_rows = [
        ("Số ngày giữa 2 snapshot", velocity.get("days_between")),
        ("Số function đã Closed",    velocity.get("functions_closed")),
        ("Tốc độ close (func/ngày)", velocity.get("close_rate_per_day")),
        ("Ước lượng ngày còn lại",   velocity.get("est_days_remaining")),
        ("Function mới phát sinh",   velocity.get("functions_new")),
    ]
    for label, val in velocity_rows:
        ws.cell(row=vrow, column=1, value=label).font = Font(name="Arial", bold=True)
        ws.cell(row=vrow, column=1).border = THIN_BORDER
        c = ws.cell(row=vrow, column=2, value=val)
        c.alignment = Alignment(horizontal="right")
        c.border = THIN_BORDER
        vrow += 1

    # === Sheet 2: New Functions ===
    ws = wb.create_sheet("New_Functions")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40),
        ("Module", 10), ("Priority", 12),
    ]
    new_fns = compare_result.get("new_functions", [])
    data_rows = [
        [idx + 1, f.get("ma_cn", ""), f.get("ten_cn", ""), f.get("module", ""), f.get("priority", "")]
        for idx, f in enumerate(new_fns)
    ]
    _write_sheet(ws, "FUNCTIONS MỚI PHÁT SINH", columns, data_rows,
                 row_fill_fn=lambda ri, idx: ORANGE_FILL)

    # === Sheet 3: Removed ===
    ws = wb.create_sheet("Removed_Functions")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40),
        ("Module", 10), ("Priority", 12),
    ]
    removed = compare_result.get("removed_functions", [])
    data_rows = [
        [idx + 1, f.get("ma_cn", ""), f.get("ten_cn", ""), f.get("module", ""), f.get("priority", "")]
        for idx, f in enumerate(removed)
    ]
    _write_sheet(ws, "FUNCTIONS BỊ XÓA", columns, data_rows)

    # === Sheet 4: Status Changes ===
    ws = wb.create_sheet("Status_Changes")
    columns = [
        ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
        ("Phase", 16), ("Status cũ", 13), ("Status mới", 13), ("Hướng", 10),
    ]
    changes = compare_result.get("status_changes", [])
    data_rows = [
        [
            idx + 1, c.get("ma_cn", ""), c.get("ten_cn", ""), c.get("module", ""),
            c.get("phase", ""), c.get("old_status", ""), c.get("new_status", ""),
            c.get("direction", ""),
        ]
        for idx, c in enumerate(changes)
    ]

    def _fill_by_direction(ri: int, idx: int) -> Optional[PatternFill]:
        d = changes[idx].get("direction", "")
        if d == "forward":
            return GREEN_FILL
        if d == "backward":
            return RED_FILL
        return None

    _write_sheet(ws, "THAY ĐỔI TRẠNG THÁI", columns, data_rows,
                 row_fill_fn=_fill_by_direction)

    # === Sheet 5: Module Delta ===
    ws = wb.create_sheet("Module_Delta")
    columns = [
        ("Module", 12), ("% Trước", 12), ("% Sau", 12),
        ("Chênh lệch", 12), ("Closed mới", 12), ("Function mới", 12),
    ]
    module_deltas = compare_result.get("module_deltas", {})
    data_rows = [
        [
            m,
            d.get("old_pct"), d.get("new_pct"),
            d.get("delta_pct"), d.get("closed_count"),
            d.get("new_count"),
        ]
        for m, d in module_deltas.items()
    ]
    _write_sheet(ws, "CHÊNH LỆCH TIẾN ĐỘ THEO MODULE", columns, data_rows)

    # === Save ===
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Compare_{old_date}_vs_{new_date}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# Drill-Down Export — Xuất Excel từ modal chi tiết biểu đồ
# ==========================================================================

DRILL_COLUMNS = [
    ("STT", 6),
    ("Mã CN", 14),
    ("Tên chức năng", 42),
    ("Module", 10),
    ("Quy trình", 22),
    ("Giai đoạn", 10),
    ("Phase", 14),
    ("Status", 13),
    ("PIC", 22),
    ("Start", 12),
    ("End", 12),
    ("Ngày trễ", 10),
    ("Priority", 12),
    ("Complexity", 12),
    ("FIT/GAP", 10),
    ("Estimate MH", 12),
    ("Lý do", 40),
]


def _drill_row_fill(row_idx: int, row_offset: int, items: list[dict]):
    """Row fill: overdue → RED, closed → GREEN, in-progress → YELLOW."""
    if row_offset >= len(items):
        return None
    item = items[row_offset]
    if item.get("is_overdue"):
        return RED_FILL
    status = (item.get("status") or "").lower()
    if status == "closed":
        return GREEN_FILL
    if status in ("in-progress", "assigned"):
        return YELLOW_FILL
    return None


def export_drill_down(
    items: list[dict[str, Any]],
    title: str = "Chi tiết biểu đồ",
    subtitle: str = "",
    output_dir: str = "uploads",
) -> str:
    """
    Xuất Excel danh sách function chi tiết từ drill-down (click biểu đồ).

    Args:
        items: List function dict theo format của `analyzer.drill_down.drill_down`
        title: Tiêu đề chính (VD: "Priority: Must-have")
        subtitle: Dòng subtitle (VD: "Tổng 292 function")
        output_dir: Thư mục lưu file

    Returns:
        Filepath của file .xlsx đã tạo.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi_Tiet"

    data_rows = []
    for idx, it in enumerate(items):
        pics = it.get("pics", []) or []
        pic_str = ", ".join(pics) if isinstance(pics, list) else str(pics)
        # Lý do khi drill từ issue (overdue / unassigned / risk / …)
        ly_do = it.get("ly_do") or it.get("reason") or ""
        if not ly_do and (it.get("is_overdue") or (it.get("days_overdue") or 0) > 0):
            ly_do = reason_overdue(it)
        elif not ly_do and (
            it.get("completed_phase") or it.get("waiting_phase")
        ):
            ly_do = reason_stalled(it)
        elif not ly_do and (
            it.get("predecessor_phase")
            or it.get("start_gate")
            or it.get("is_first_phase") is True
        ):
            ly_do = reason_unassigned(it)
        elif not ly_do and (it.get("risk_factors") or it.get("risk_factors_detail")):
            ly_do = format_risk_factors_detailed(it)
        data_rows.append([
            idx + 1,
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("module", ""),
            it.get("quy_trinh", ""),
            it.get("giai_doan", ""),
            it.get("phase", ""),
            it.get("status", ""),
            pic_str,
            it.get("start_date", ""),
            it.get("end_date", ""),
            it.get("days_overdue", 0) or "",
            it.get("priority", ""),
            it.get("complexity", ""),
            it.get("fit_gap", ""),
            it.get("estimate_mh", "") or "",
            ly_do,
        ])

    _write_sheet(
        ws,
        title=f"CHI TIẾT — {title}",
        columns=DRILL_COLUMNS,
        data_rows=data_rows,
        subtitle=subtitle or f"Tổng: {len(items)} function | Ngày xuất: {date.today().strftime('%d/%m/%Y')}",
        row_fill_fn=lambda r, o: _drill_row_fill(r, o, items),
    )

    os.makedirs(output_dir, exist_ok=True)
    safe_title = "".join(c if c.isalnum() or c in "._- " else "_" for c in title).strip()[:60]
    filename = f"DrillDown_{safe_title}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ======================================================================
# PIC BLACKLIST — data-quality report
# ======================================================================

# Columns cho báo cáo PIC bị blacklist khỏi parser.
# Thứ tự đặt để user dễ tra Excel: Row → Header → giá trị → keyword match → context.
BLACKLIST_COLUMNS: list[tuple[str, int]] = [
    ("STT", 6),
    ("Row Excel", 10),
    ("Mã CN", 16),
    ("Module", 10),
    ("Phase", 14),
    ("Cột (header text)", 26),
    ("Giá trị bị bỏ", 20),
    ("Keyword khớp", 14),
    ("Ghi chú", 40),
]


# ======================================================================
# PORTFOLIO COMPARE — Cross-project side-by-side compare export
# ======================================================================

def export_portfolio_compare(
    compare_result: dict,
    output_dir: str = "uploads",
) -> str:
    """
    Xuất bảng so sánh N project (từ `analyzer.portfolio.compare_projects`) ra
    Excel. Layout:
    - Cột 1: tên metric
    - Cột 2..N+1: value cho từng project (đặt tên = slug + name)
    - Highlight: cell best = xanh nhạt, worst = đỏ nhạt (theo `best_worst`)

    Args:
        compare_result: dict trả về từ compare_projects — có keys:
            projects, metrics, metric_labels, best_worst, skipped
        output_dir: folder lưu

    Returns:
        Filepath .xlsx đã tạo.
    """
    projects = compare_result.get("projects", []) or []
    metrics = compare_result.get("metrics", {}) or {}
    metric_labels = compare_result.get("metric_labels", []) or []
    best_worst = compare_result.get("best_worst", {}) or {}
    skipped = compare_result.get("skipped", []) or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio_Compare"

    n_projects = len(projects)
    n_cols = 1 + n_projects  # 1 col metric + N col project

    # === Title (merge cả row) ===
    last_col_letter = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    tc = ws["A1"]
    tc.value = "BÁO CÁO SO SÁNH DỰ ÁN (PORTFOLIO)"
    tc.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # === Subtitle ===
    ws.merge_cells(f"A2:{last_col_letter}2")
    sc = ws["A2"]
    subtitle_parts = [f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"]
    subtitle_parts.append(f"So sánh {n_projects} dự án")
    if skipped:
        subtitle_parts.append(f"Bỏ qua {len(skipped)} dự án (chưa upload file)")
    sc.value = " | ".join(subtitle_parts)
    sc.font = Font(name="Arial", size=10, italic=True, color="666666")
    sc.alignment = Alignment(horizontal="center")

    # === Header row (row 4) ===
    header_row = 4
    ws.cell(row=header_row, column=1, value="Chỉ tiêu")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if col_idx == 1:
            pass  # đã set value ở trên
        else:
            proj = projects[col_idx - 2]
            cell.value = proj.get("name") or proj.get("slug", "")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Width cột
    ws.column_dimensions["A"].width = 24
    for col_idx in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.row_dimensions[header_row].height = 30

    # === Data rows: mỗi metric 1 row ===
    for row_offset, ml in enumerate(metric_labels):
        row_idx = header_row + 1 + row_offset
        key = ml["key"]
        label = ml["label"]

        # Col 1: metric label
        c = ws.cell(row=row_idx, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = BODY_ALIGN
        c.border = THIN_BORDER

        # Col 2..N: value + highlight best/worst
        bw = best_worst.get(key, {})
        best_slug = bw.get("best")
        worst_slug = bw.get("worst")
        for col_idx, proj in enumerate(projects, 2):
            slug = proj["slug"]
            val = metrics.get(key, {}).get(slug, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            if best_slug == slug:
                cell.fill = GREEN_FILL
            elif worst_slug == slug:
                cell.fill = RED_FILL

    # === Skipped section (nếu có) ===
    if skipped:
        skip_row = header_row + len(metric_labels) + 2
        ws.merge_cells(f"A{skip_row}:{last_col_letter}{skip_row}")
        sk_cell = ws.cell(row=skip_row, column=1)
        sk_cell.value = "DỰ ÁN BỊ BỎ QUA:"
        sk_cell.font = Font(name="Arial", bold=True, color="B45309")
        for i, sk in enumerate(skipped, 1):
            r = skip_row + i
            ws.cell(row=r, column=1, value=f"  • {sk.get('slug', '')}").font = BODY_FONT
            ws.cell(row=r, column=2, value=f"Lý do: {sk.get('reason', '')}").font = BODY_FONT

    # === Freeze + auto-filter ===
    ws.freeze_panes = f"B{header_row + 1}"

    os.makedirs(output_dir, exist_ok=True)
    filename = f"Portfolio_Compare_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


def export_pic_blacklist_report(
    items: list[dict[str, Any]],
    output_dir: str = "uploads",
    subtitle: str = "",
) -> str:
    """
    Xuất báo cáo Excel danh sách PIC bị blacklist khi parse Function List.

    Đây là báo cáo "chất lượng dữ liệu" — mỗi hàng = 1 token PIC bị parser bỏ
    do trùng tên với 1 status hợp lệ (dấu hiệu cột Status bị lệch qua cột PIC).

    Args:
        items: list dict {row_index, phase_name, header_text, raw_value,
                          matched_keyword, ma_cn, module}
        output_dir: thư mục lưu file
        subtitle: dòng subtitle tùy chọn

    Returns:
        Filepath .xlsx đã tạo.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PIC_Blacklist"

    data_rows = []
    for idx, it in enumerate(items):
        data_rows.append([
            idx + 1,
            it.get("row_index", ""),
            it.get("ma_cn", ""),
            it.get("module", ""),
            it.get("phase_name", ""),
            it.get("header_text", ""),
            it.get("raw_value", ""),
            it.get("matched_keyword", ""),
            # Ghi chú: hướng dẫn user tra Excel gốc
            f'Kiểm tra ô cột "{it.get("header_text", "")}" row {it.get("row_index", "?")} — '
            f'có thể user paste lệch cột Status sang PIC.',
        ])

    _write_sheet(
        ws,
        title="BÁO CÁO PIC BỊ BLACKLIST (Data Quality)",
        columns=BLACKLIST_COLUMNS,
        data_rows=data_rows,
        subtitle=subtitle or (
            f"Tổng: {len(items)} token PIC bị parser bỏ do trùng status keyword "
            f"| Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
        ),
        # Highlight vàng nhẹ toàn bảng — đây là warning (không phải error đỏ)
        row_fill_fn=lambda r, o: YELLOW_FILL,
    )

    os.makedirs(output_dir, exist_ok=True)
    filename = f"PIC_Blacklist_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# Export chart đơn lẻ + Audit Report
# ==========================================================================

SUPPORTED_EXPORT_CHARTS = {
    "effort_heatmap",
    "effort_pic",
    "module_overview",
    "phase_matrix",
    "phase_stacked",
    "overdue",
    "pic_workload",
    "risk",
    "stalled",
    "unassigned",
    "task_type",
    "priority",
    "complexity",
    "fit_gap",
    "giai_doan",
    "process",
    "duration",
    "burndown",
}


def _normalize_export_status(status: Any) -> str:
    """Chuẩn hóa status phase cho Excel (bỏ số lệch cột, map alias)."""
    from parser.excel_parser import VALID_STATUSES, STATUS_ALIASES

    if status is None:
        return ""
    s = str(status).strip()
    if not s or s.isdigit():
        return ""
    alias = STATUS_ALIASES.get(s.lower())
    if alias:
        return alias
    for valid in VALID_STATUSES:
        if s.lower() == valid.lower():
            return valid
    return s


def _status_for_task_type(row, phase_names: list[str]) -> str:
    """Gộp status các phase thuộc 1 work bucket (1 phase → 1 status; nhiều → join)."""
    statuses: list[str] = []
    seen: set[str] = set()
    for ph in phase_names:
        pd = (row.phases or {}).get(ph)
        if pd is None:
            continue
        st = _normalize_export_status(getattr(pd, "status", None))
        if st and st not in seen:
            seen.add(st)
            statuses.append(st)
    return " | ".join(statuses)


def _func_meta(row) -> dict[str, Any]:
    """Meta chuẩn cho sheet Chi_tiet."""
    from analyzer.rlog_weekly import _row_rlog_id

    m = row.meta or {}
    return {
        "ma_cn": m.get("ma_cn") or "",
        "fid": str(m.get("fid") or "").strip(),
        "rlog_id": _row_rlog_id(row) or "",
        "ten_cn": m.get("ten_cn") or "",
        "module": m.get("module") or "",
        "quy_trinh": m.get("quy_trinh") or "",
        "priority": m.get("priority") or "",
        "complexity": m.get("complexity") or "",
        "fit_gap": m.get("fit_gap") or "",
        "giai_doan": str(m.get("giai_doan") or ""),
        "ma_du_an": m.get("ma_du_an") or "",
    }


# Cột meta đứng đầu mọi sheet Chi_tiet. Thứ tự phải khớp 1:1 với
# `_meta_cell_values` — lệch là toàn bộ sheet lệch cột mà không báo lỗi.
DETAIL_META_COLUMNS: list[tuple[str, int]] = [
    ("STT", 6),
    ("Mã CN", 14),
    ("FID", 12),
    ("Rlog ID", 14),
    ("Tên chức năng", 40),
    ("Module", 10),
    ("Quy trình", 22),
    ("Priority", 12),
    ("Complexity", 12),
    ("Mã dự án", 22),
]


def _meta_cell_values(idx: int, meta: dict[str, Any]) -> list[Any]:
    return [
        idx + 1,
        meta.get("ma_cn", ""),
        meta.get("fid", ""),
        meta.get("rlog_id", ""),
        meta.get("ten_cn", ""),
        meta.get("module", ""),
        meta.get("quy_trinh", ""),
        meta.get("priority", ""),
        meta.get("complexity", ""),
        meta.get("ma_du_an", ""),
    ]


_RLOG_COL_NAME = "Rlog ID"


def _detail_meta(parsed_data) -> tuple[list[tuple[str, int]], Any]:
    """
    Trả về ``(columns, values_fn)`` cho phần meta đầu sheet Chi_tiet.

    Hai thứ này **luôn đi cùng nhau** vì chúng phải khớp số phần tử; tách rời ở
    9 call site như trước thì thêm/bớt cột là lệch cả sheet mà openpyxl không
    báo gì — dữ liệu vẫn ghi, chỉ nằm sai cột.

    Ẩn cột ``Rlog ID`` khi FL không khai cột đó (FL của dự án bỏ
    ``Analysis - RlogID`` từ 30/07: 68 → 65 cột). Bày một cột trống trơn khiến
    PM tưởng hệ thống đọc lỗi, trong khi nguồn không có dữ liệu để đọc.
    Điều kiện là *có khai cột*, không phải *có giá trị* — file khai cột nhưng
    mới điền lác đác vẫn phải hiện, nếu không PM không biết chỗ nào còn thiếu.
    """
    from analyzer.rlog_weekly import _file_has_rlog_column

    show_rlog = True
    if parsed_data is not None:
        try:
            show_rlog = _file_has_rlog_column(parsed_data)
        except Exception:
            show_rlog = True  # nghi ngờ thì hiện — thà dư cột còn hơn mất cột

    if show_rlog:
        return list(DETAIL_META_COLUMNS), _meta_cell_values

    keep = [i for i, c in enumerate(DETAIL_META_COLUMNS) if c[0] != _RLOG_COL_NAME]
    columns = [DETAIL_META_COLUMNS[i] for i in keep]

    def values(idx: int, meta: dict[str, Any]) -> list[Any]:
        full = _meta_cell_values(idx, meta)
        return [full[i] for i in keep]

    return columns, values


def _phase_status_map(row, phases: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for ph in phases:
        pd = (row.phases or {}).get(ph)
        out[ph] = _normalize_export_status(getattr(pd, "status", None) if pd else None)
    return out


def build_task_type_detail_rows(parsed_data) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Build danh sách function × status theo loại công việc.

    Returns:
        (task_types, rows) — mỗi row: meta + statuses[task_type] = status chuẩn hóa.
    """
    task_types: list[str] = []
    task_phase_map: dict[str, list[str]] = {}
    seen: set[str] = set()
    for pg in getattr(parsed_data, "phase_groups", []) or []:
        tt = pg.task_type
        task_phase_map.setdefault(tt, []).append(pg.name)
        if tt not in seen:
            task_types.append(tt)
            seen.add(tt)

    rows: list[dict[str, Any]] = []
    for r in getattr(parsed_data, "rows", []) or []:
        meta = _func_meta(r)
        statuses = {
            tt: _status_for_task_type(r, task_phase_map.get(tt) or [])
            for tt in task_types
        }
        rows.append({**meta, "statuses": statuses})
    return task_types, rows


def build_phase_status_detail_rows(parsed_data) -> tuple[list[str], list[dict[str, Any]]]:
    """Function × status từng phase (dùng cho phase_stacked / phase_matrix / …)."""
    phases = list(getattr(parsed_data, "all_phases", None) or [])
    if not phases:
        phases = [pg.name for pg in (getattr(parsed_data, "phase_groups", None) or [])]
    rows: list[dict[str, Any]] = []
    for r in getattr(parsed_data, "rows", []) or []:
        meta = _func_meta(r)
        rows.append({**meta, "statuses": _phase_status_map(r, phases)})
    return phases, rows


def _count_by(items: list[dict], key: str) -> list[list[Any]]:
    from collections import Counter

    c = Counter((i.get(key) or "(blank)") for i in items)
    return [[idx + 1, k, v] for idx, (k, v) in enumerate(sorted(c.items(), key=lambda x: -x[1]))]


def export_chart(
    chart: str,
    metrics: dict[str, Any],
    output_dir: str = "uploads",
    subtitle: str = "",
    parsed_data=None,
    group_by: str = "module",
    mode: str = "both",
) -> str:
    """
    Xuất Excel cho 1 chart.

    mode: summary | detail | both (default both)
      - Tong_hop: số liệu biểu đồ / aggregate
      - Chi_tiet: danh sách chức năng tạo nên thống kê + tình trạng liên quan
      - Theo_nhom: optional (task_type / phase_matrix khi có group) — kèm summary
    """
    if chart not in SUPPORTED_EXPORT_CHARTS:
        raise ValueError(f"Chart không hỗ trợ: {chart}. Hỗ trợ: {sorted(SUPPORTED_EXPORT_CHARTS)}")

    mode = _normalize_export_mode(mode)
    do_sum = _want_summary(mode)
    do_det = _want_detail(mode)
    book = _SheetBook()
    sub = subtitle or f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"

    # Cột meta + hàm ghi ô phải lấy cùng lúc từ 1 chỗ để không lệch nhau
    # (xem docstring `_detail_meta`). Rlog ID tự ẩn nếu FL không khai cột.
    meta_cols, meta_vals = _detail_meta(parsed_data)

    # ------------------------------------------------------------------
    # Helpers nội bộ ghi sheet
    # ------------------------------------------------------------------
    def write_summary(title: str, columns, data_rows, sheet_name: str = "Tong_hop", row_fill_fn=None):
        if not do_sum:
            return
        ws = book.sheet(sheet_name)
        _write_sheet(ws, title, columns, data_rows, subtitle=sub, row_fill_fn=row_fill_fn)

    def write_detail(title: str, columns, data_rows, sheet_name: str = "Chi_tiet", row_fill_fn=None):
        if not do_det:
            return
        ws = book.sheet(sheet_name)
        _write_sheet(ws, title, columns, data_rows, subtitle=sub, row_fill_fn=row_fill_fn)

    def write_group(title: str, columns, data_rows, sheet_name: str = "Theo_nhom"):
        """Sheet nhóm — chỉ khi summary (mode summary/both)."""
        if not do_sum:
            return
        ws = book.sheet(sheet_name)
        _write_sheet(ws, title, columns, data_rows, subtitle=sub)

    def detail_from_functions(
        title: str,
        extra_cols: list[tuple[str, int]],
        row_builder,
    ):
        """Chi_tiet từ parsed_data.rows; row_builder(row, meta) → list extra values."""
        if not do_det or parsed_data is None:
            return
        columns = list(meta_cols) + list(extra_cols)
        data_rows = []
        for idx, r in enumerate(getattr(parsed_data, "rows", []) or []):
            meta = _func_meta(r)
            data_rows.append(meta_vals(idx, meta) + list(row_builder(r, meta)))
        write_detail(title, columns, data_rows)

    def detail_from_items(
        title: str,
        columns: list[tuple[str, int]],
        data_rows: list[list[Any]],
        row_fill_fn=None,
    ):
        write_detail(title, columns, data_rows, row_fill_fn=row_fill_fn)

    # ------------------------------------------------------------------
    # Per-chart
    # ------------------------------------------------------------------
    if chart == "effort_heatmap":
        e = metrics.get("effort_analysis") or {}
        modules = e.get("modules") or []
        phases = e.get("phases") or []
        heatmap = e.get("heatmap") or {}
        columns = [("Module", 14)] + [(p, 12) for p in phases] + [("Tổng MH", 12)]
        data_rows = []
        for m in modules:
            row_vals: list[Any] = [m]
            total = 0.0
            for p in phases:
                v = float((heatmap.get(m) or {}).get(p) or 0)
                row_vals.append(v if v else "")
                total += v
            row_vals.append(round(total, 1))
            data_rows.append(row_vals)
        write_summary("EFFORT HEATMAP — Module × Phase (MH)", columns, data_rows)
        open_tasks = e.get("open_tasks_by_pic") or []
        detail_from_items(
            "CHI TIẾT TASK CÓ ESTIMATE MH (chưa Closed)",
            [
                ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
                ("Phase", 16), ("PIC", 24), ("Status", 12),
                ("End date", 13), ("Estimate MH", 12),
            ],
            [
                [
                    idx + 1, t.get("ma_cn", ""), t.get("ten_cn", ""), t.get("module", ""),
                    t.get("phase", ""),
                    ", ".join(t.get("pic") or []) if isinstance(t.get("pic"), list) else (t.get("pic") or ""),
                    t.get("status", ""), t.get("end_date", ""), t.get("estimate_mh", ""),
                ]
                for idx, t in enumerate(open_tasks)
            ],
        )

    elif chart == "effort_pic":
        e = metrics.get("effort_analysis") or {}
        by_pic = e.get("by_pic") or []
        write_summary(
            "EFFORT THEO PIC",
            [
                ("STT", 6), ("PIC", 20), ("Total MH", 12),
                ("Closed MH", 12), ("Remaining MH", 12),
            ],
            [
                [idx + 1, p.get("pic", ""), p.get("total_mh", 0),
                 p.get("closed_mh", 0), p.get("remaining_mh", 0)]
                for idx, p in enumerate(by_pic)
            ],
        )
        open_tasks = e.get("open_tasks_by_pic") or []
        detail_from_items(
            "TASK CHƯA DONE CÓ ESTIMATE MH",
            [
                ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
                ("Phase", 16), ("PIC", 24), ("Status", 12),
                ("End date", 13), ("Estimate MH", 12),
            ],
            [
                [
                    idx + 1, t.get("ma_cn", ""), t.get("ten_cn", ""), t.get("module", ""),
                    t.get("phase", ""),
                    ", ".join(t.get("pic") or []) if isinstance(t.get("pic"), list) else (t.get("pic") or ""),
                    t.get("status", ""), t.get("end_date", ""), t.get("estimate_mh", ""),
                ]
                for idx, t in enumerate(open_tasks)
            ],
        )

    elif chart == "module_overview":
        items = metrics.get("module_overview") or []
        gb = (group_by or "module").strip().lower()
        # Có delta khi caller đã gắn (xem _module_overview_rows_for_export).
        has_delta = any(isinstance(i.get("delta"), dict) for i in items)
        columns = [("STT", 6), ("Module", 14)]
        if gb == "process":
            columns.append(("Quy trình", 22))
        columns += [
            ("Số CN", 10), ("Số QT", 10), ("% Progress", 12),
            ("Phase active", 18), ("Overdue", 10),
            ("Còn lại", 10), ("Đánh giá", 16),
        ]
        # Điểm phần trăm cho Tiến độ, phần trăm tương đối cho các cột %.
        delta_cols = [
            ("± SL", 9, "total_delta"), ("±% SL", 9, "total_delta_pct"),
            ("± Tiến độ (pp)", 14, "progress_delta"), ("±% Tiến độ", 11, "progress_delta_pct"),
            ("± Trễ", 9, "overdue_delta"), ("±% Trễ", 10, "overdue_delta_pct"),
            ("± Còn lại", 11, "remaining_delta"), ("±% Còn lại", 11, "remaining_delta_pct"),
        ]
        if has_delta:
            columns += [(name, width) for name, width, _key in delta_cols]

        risk_label = {"risk": "Rủi ro", "warning": "Cần theo dõi", "safe": "An toàn"}
        rows_out = []
        for idx, i in enumerate(items):
            row = [i.get("stt", idx + 1), i.get("module", "")]
            if gb == "process":
                row.append(i.get("process", ""))
            row += [
                i.get("total", 0), i.get("quy_trinh_count", 0), i.get("progress_pct", 0),
                i.get("active_phase", ""), i.get("overdue_count", 0),
                i.get("remaining", 0),
                risk_label.get(i.get("risk_level"), i.get("risk_level") or ""),
            ]
            if has_delta:
                d = i.get("delta") or {}
                if d.get("is_new"):
                    row += ["Mới"] + [""] * (len(delta_cols) - 1)
                else:
                    row += [
                        "" if d.get(key) is None else d.get(key)
                        for _name, _w, key in delta_cols
                    ]
            rows_out.append(row)
        write_summary("TỔNG QUAN THEO MODULE", columns, rows_out)
        if parsed_data is not None:
            phs, det = build_phase_status_detail_rows(parsed_data)
            columns = list(meta_cols) + [(p, 12) for p in phs]
            data_rows = [
                meta_vals(idx, it) + [(it.get("statuses") or {}).get(p, "") for p in phs]
                for idx, it in enumerate(det)
            ]
            write_detail("CHI TIẾT CHỨC NĂNG — Status theo phase", columns, data_rows)

    elif chart == "phase_matrix":
        mx = metrics.get("phase_status_matrix") or {}
        phases = mx.get("phases") or []
        modules = mx.get("modules") or []
        data = mx.get("data") or {}
        gb = (group_by or "module").strip().lower()
        group_label = "Quy trình" if gb == "process" else "Module"
        write_summary(
            f"PHASE × {group_label.upper()} (% Closed)",
            [(group_label, 18)] + [(f"{p} %Closed", 12) for p in phases],
            [
                [m] + [
                    ((data.get(m) or {}).get(p) or {}).get("pct_closed", 0)
                    if isinstance((data.get(m) or {}).get(p), dict)
                    else ((data.get(m) or {}).get(p) or 0)
                    for p in phases
                ]
                for m in modules
            ],
        )
        if parsed_data is not None:
            phs, det = build_phase_status_detail_rows(parsed_data)
            columns = list(meta_cols) + [(p, 12) for p in phs]
            data_rows = [
                meta_vals(idx, it) + [(it.get("statuses") or {}).get(p, "") for p in phs]
                for idx, it in enumerate(det)
            ]
            write_detail("CHI TIẾT CHỨC NĂNG — Status theo phase", columns, data_rows)

    elif chart == "phase_stacked":
        d = metrics.get("phase_progress_stacked") or {}
        phases = d.get("phases") or []
        statuses = d.get("statuses") or []
        data = d.get("data") or {}
        write_summary(
            "TIẾN ĐỘ THEO PHASE (Status count)",
            [("Phase", 16)] + [(s, 12) for s in statuses],
            [[ph] + [(data.get(ph) or {}).get(s, 0) for s in statuses] for ph in phases],
        )
        if parsed_data is not None:
            phs, det = build_phase_status_detail_rows(parsed_data)
            columns = list(meta_cols) + [(p, 12) for p in phs]
            data_rows = [
                meta_vals(idx, it) + [(it.get("statuses") or {}).get(p, "") for p in phs]
                for idx, it in enumerate(det)
            ]
            write_detail("CHI TIẾT CHỨC NĂNG — Status theo phase", columns, data_rows)

    elif chart == "task_type":
        d = metrics.get("progress_by_task_type") or {}
        task_types = d.get("task_types") or []
        gb = (group_by or "module").strip().lower()
        by_source = (d.get("by_process") if gb == "process" else d.get("by_module")) or {}
        groups = list(by_source.keys())
        summary_rows = []
        for tt in task_types:
            vals = [
                by_source[g].get(tt)
                for g in groups
                if isinstance(by_source.get(g), dict) and by_source[g].get(tt) is not None
            ]
            avg = round(sum(vals) / len(vals), 2) if vals else 0
            summary_rows.append([tt, avg])
        write_summary(
            "TIẾN ĐỘ THEO CÔNG VIỆC (% Closed trung bình)",
            [("Loại công việc", 24), ("% Closed", 12)],
            summary_rows,
        )
        group_label = "Quy trình" if gb == "process" else "Module"
        write_group(
            f"TIẾN ĐỘ THEO CÔNG VIỆC — theo {group_label} (% Closed)",
            [(group_label, 18)] + [(tt, 14) for tt in task_types],
            [
                [g] + [(by_source.get(g) or {}).get(tt, 0) for tt in task_types]
                for g in groups
            ],
        )
        if do_det and parsed_data is not None:
            tt_detail, detail_items = build_task_type_detail_rows(parsed_data)
            tt_cols = list(task_types) if task_types else list(tt_detail)
            for tt in tt_detail:
                if tt not in tt_cols:
                    tt_cols.append(tt)
            columns_det = list(meta_cols) + [(tt, 14) for tt in tt_cols]
            # DETAIL_META có Complexity; task_type cũ không có — giữ đủ meta chuẩn
            detail_rows = []
            for idx, it in enumerate(detail_items):
                st_map = it.get("statuses") or {}
                detail_rows.append(
                    meta_vals(idx, it) + [st_map.get(tt, "") for tt in tt_cols]
                )
            write_detail(
                "CHI TIẾT CHỨC NĂNG — Status theo loại công việc",
                columns_det,
                detail_rows,
            )

    elif chart == "priority":
        d = metrics.get("priority_breakdown") or {}
        write_summary(
            "PHÂN BỐ PRIORITY",
            [("STT", 6), ("Priority", 20), ("Số lượng", 12)],
            [[idx + 1, k, v] for idx, (k, v) in enumerate(sorted(d.items(), key=lambda x: -x[1]))],
        )
        detail_from_functions(
            "CHI TIẾT CHỨC NĂNG — theo Priority",
            [("Priority (metric)", 14)],
            lambda r, meta: [meta.get("priority") or "N/A"],
        )

    elif chart == "complexity":
        d = metrics.get("complexity_breakdown") or {}
        write_summary(
            "PHÂN BỐ COMPLEXITY",
            [("STT", 6), ("Complexity", 20), ("Số lượng", 12)],
            [[idx + 1, k, v] for idx, (k, v) in enumerate(sorted(d.items(), key=lambda x: -x[1]))],
        )
        detail_from_functions(
            "CHI TIẾT CHỨC NĂNG — theo Complexity",
            [("Complexity (metric)", 16)],
            lambda r, meta: [meta.get("complexity") or "N/A"],
        )

    elif chart == "fit_gap":
        d = metrics.get("fit_gap_analysis") or {}
        modules = list(d.keys())
        keys: set[str] = set()
        for m in modules:
            keys.update((d.get(m) or {}).keys())
        keys_sorted = sorted(keys)
        write_summary(
            "FIT / GAP THEO MODULE",
            [("Module", 12)] + [(k, 10) for k in keys_sorted],
            [[m] + [(d.get(m) or {}).get(k, 0) for k in keys_sorted] for m in modules],
        )
        detail_from_functions(
            "CHI TIẾT CHỨC NĂNG — FIT/GAP",
            [("FIT/GAP", 12)],
            lambda r, meta: [meta.get("fit_gap") or "N/A"],
        )

    elif chart == "giai_doan":
        d = metrics.get("giai_doan_progress") or {}
        giai_doans = list(d.keys())
        phases: list[str] = []
        for gd in giai_doans:
            for p in (d.get(gd) or {}).keys():
                if p not in phases:
                    phases.append(p)
        sum_rows = []
        for gd in giai_doans:
            row = [gd]
            cell = d.get(gd) or {}
            for p in phases:
                val = cell.get(p, 0)
                if isinstance(val, dict):
                    val = val.get("pct", val.get("pct_closed", 0))
                row.append(val)
            sum_rows.append(row)
        write_summary(
            "TIẾN ĐỘ THEO GIAI ĐOẠN (% Closed)",
            [("Giai đoạn", 14)] + [(f"{p} %Closed", 12) for p in phases],
            sum_rows,
        )
        if parsed_data is not None:
            phs, det = build_phase_status_detail_rows(parsed_data)
            columns = list(meta_cols) + [("Giai đoạn", 14)] + [(p, 12) for p in phs]
            data_rows = [
                meta_vals(idx, it)
                + [it.get("giai_doan", "")]
                + [(it.get("statuses") or {}).get(p, "") for p in phs]
                for idx, it in enumerate(det)
            ]
            write_detail("CHI TIẾT CHỨC NĂNG — Giai đoạn + Status phase", columns, data_rows)

    elif chart == "process":
        items = metrics.get("process_analysis") or []
        write_summary(
            "PHÂN TÍCH THEO QUY TRÌNH",
            [
                ("STT", 6), ("Quy trình", 40), ("Số CN", 10), ("% Closed", 12),
                ("Overdue", 10), ("Modules", 24),
            ],
            [
                [
                    idx + 1, i.get("process", ""), i.get("total", 0),
                    i.get("pct_closed", 0), i.get("overdue", 0),
                    ", ".join(i.get("modules") or []),
                ]
                for idx, i in enumerate(items)
            ],
        )
        if parsed_data is not None:
            phs, det = build_phase_status_detail_rows(parsed_data)
            columns = list(meta_cols) + [(p, 12) for p in phs]
            data_rows = [
                meta_vals(idx, it) + [(it.get("statuses") or {}).get(p, "") for p in phs]
                for idx, it in enumerate(det)
            ]
            write_detail("CHI TIẾT CHỨC NĂNG — theo Quy trình", columns, data_rows)

    elif chart == "duration":
        d = metrics.get("duration_analysis") or {}
        items = d.get("items") or []
        from collections import Counter
        type_c = Counter(
            ("Elapsed" if i.get("duration_type") == "elapsed" else "Planned")
            for i in items
        )
        mod_c = Counter((i.get("module") or "(blank)") for i in items)
        write_summary(
            "TỔNG HỢP DURATION BẤT THƯỜNG",
            [("STT", 6), ("Nhóm", 24), ("Số lượng", 12)],
            (
                [[idx + 1, f"Module: {k}", v]
                 for idx, (k, v) in enumerate(sorted(mod_c.items(), key=lambda x: -x[1]))]
                + [[len(mod_c) + idx + 1, f"Loại: {k}", v]
                   for idx, (k, v) in enumerate(sorted(type_c.items(), key=lambda x: -x[1]))]
            ) if items else [],
        )
        dur_thr = d.get("threshold_days")
        detail_from_items(
            "CHI TIẾT TASK DURATION BẤT THƯỜNG",
            [
                ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
                ("Quy trình", 22),
                ("Phase", 16), ("Start", 13), ("End", 13),
                ("Duration (ngày)", 14), ("Ngưỡng", 10), ("Loại", 12),
                ("Status", 12), ("PIC", 20), ("Lý do", 40),
            ],
            [
                [
                    idx + 1, i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
                    process_code(i),
                    i.get("phase", ""), i.get("start_date", ""), i.get("end_date", ""),
                    i.get("duration_days", 0),
                    i.get("threshold_days", dur_thr),
                    "Elapsed" if i.get("duration_type") == "elapsed" else "Planned",
                    i.get("status", ""),
                    ", ".join(i.get("pic") or []) if isinstance(i.get("pic"), list) else (i.get("pic") or ""),
                    reason_duration(i, dur_thr),
                ]
                for idx, i in enumerate(items)
            ],
        )

    elif chart == "overdue":
        items = metrics.get("overdue_list") or []
        mod_rows = _count_by(items, "module")
        ph_rows = _count_by(items, "phase")
        write_summary(
            "TỔNG HỢP OVERDUE (theo Module / Phase)",
            [("STT", 6), ("Nhóm", 28), ("Số lượng", 12)],
            (
                [[r[0], f"Module: {r[1]}", r[2]] for r in mod_rows]
                + [[r[0] + 1000, f"Phase: {r[1]}", r[2]] for r in ph_rows]
            ),
        )
        detail_from_items(
            "CHI TIẾT TASK TRỄ DEADLINE",
            [
                ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
                ("Quy trình", 22),
                ("Phase", 16), ("Deadline", 13), ("Số ngày trễ", 12),
                ("Status", 12), ("PIC", 20), ("Priority", 12),
                ("Ghi chú", 30), ("Lý do", 48),
            ],
            [
                [
                    idx + 1, i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
                    process_code(i),
                    i.get("phase", ""), i.get("end_date", ""), i.get("days_overdue", 0),
                    i.get("status", ""),
                    ", ".join(i.get("pic") or []) if isinstance(i.get("pic"), list) else "",
                    i.get("priority", ""),
                    i.get("note", ""),
                    reason_overdue(i),
                ]
                for idx, i in enumerate(items)
            ],
            row_fill_fn=lambda ri, idx: _fill_by_days(items[idx].get("days_overdue", 0)),
        )

    elif chart == "pic_workload":
        items = metrics.get("pic_workload") or []
        write_summary(
            "WORKLOAD THEO PIC",
            [
                ("STT", 6), ("PIC", 20), ("Total tasks", 12), ("Closed", 10),
                ("In-progress", 12), ("Overdue", 10), ("Assigned", 10),
            ],
            [
                [
                    idx + 1,
                    i.get("pic", ""),
                    i.get("total_tasks", i.get("total", 0)),
                    i.get("closed", i.get("closed_count", "")),
                    i.get("in_progress", i.get("inprogress", "")),
                    i.get("overdue", i.get("overdue_count", "")),
                    i.get("assigned", ""),
                ]
                for idx, i in enumerate(items)
            ],
        )
        # Chi tiết: mỗi function × phase × PIC
        if do_det and parsed_data is not None:
            detail_rows = []
            for r in getattr(parsed_data, "rows", []) or []:
                meta = _func_meta(r)
                for ph, pd in (r.phases or {}).items():
                    pics = list(getattr(pd, "pics", None) or [])
                    if not pics:
                        continue
                    st = _normalize_export_status(getattr(pd, "status", None))
                    for pic in pics:
                        detail_rows.append(
                            meta_vals(len(detail_rows), meta) + [pic, ph, st]
                        )
            write_detail(
                "CHI TIẾT FUNCTION × PHASE × PIC",
                list(meta_cols) + [
                    ("PIC", 18), ("Phase", 16), ("Status", 12),
                ],
                detail_rows,
            )

    elif chart == "risk":
        all_risk = metrics.get("risk_scores") or []
        items = [r for r in all_risk if (r.get("risk_score") or 0) >= 30]
        # Tong_hop: bucket theo mức risk
        buckets = [("Cao (≥80)", 0), ("Trung bình (50–79)", 0), ("Theo dõi (30–49)", 0)]
        for r in items:
            s = r.get("risk_score") or 0
            if s >= 80:
                buckets[0] = (buckets[0][0], buckets[0][1] + 1)
            elif s >= 50:
                buckets[1] = (buckets[1][0], buckets[1][1] + 1)
            else:
                buckets[2] = (buckets[2][0], buckets[2][1] + 1)
        write_summary(
            "TỔNG HỢP RISK SCORE (≥30)",
            [("STT", 6), ("Mức rủi ro", 24), ("Số function", 12)],
            [[idx + 1, lab, n] for idx, (lab, n) in enumerate(buckets)],
        )
        detail_from_items(
            "CHI TIẾT FUNCTION RỦI RO CAO",
            [
                ("STT", 6), ("Risk Score", 12), ("Mã CN", 14), ("Tên chức năng", 40),
                ("Module", 10), ("Quy trình", 22),
                ("Priority", 12), ("Complexity", 12), ("Risk Factors", 45),
                ("Yếu tố chi tiết", 60),
            ],
            [
                [
                    idx + 1, i.get("risk_score", 0), i.get("ma_cn", ""), i.get("ten_cn", ""),
                    i.get("module", ""), process_code(i),
                    i.get("priority", ""), i.get("complexity", ""),
                    " | ".join(i.get("risk_factors") or []),
                    format_risk_factors_detailed(i),
                ]
                for idx, i in enumerate(items)
            ],
            row_fill_fn=lambda ri, idx: _fill_by_risk(items[idx].get("risk_score", 0)),
        )

    elif chart == "stalled":
        st = metrics.get("stalled_tasks") or {}
        items = st.get("items") or []
        transitions = st.get("transitions") or []
        funnel = st.get("funnel") or []
        sum_rows = (
            [[idx + 1, f"Funnel Closed: {f.get('phase', '')}", f.get("closed", 0)]
             for idx, f in enumerate(funnel)]
            + [[idx + 100, f"Kẹt: {t.get('from', '')} → {t.get('to', '')}", t.get("count", 0)]
               for idx, t in enumerate(transitions)]
        )
        write_summary(
            "TỔNG HỢP ĐÌNH TRỆ (Funnel + Transitions)",
            [("STT", 6), ("Nhóm", 40), ("Số lượng", 12)],
            sum_rows,
        )
        detail_from_items(
            "CHI TIẾT TASK ĐÌNH TRỆ",
            [
                ("STT", 6), ("Mã CN", 14), ("Tên chức năng", 40), ("Module", 10),
                ("Quy trình", 22),
                ("Phase đã xong", 16), ("Phase chờ", 16), ("Xong ngày", 13),
                ("End phase chờ", 13), ("Chờ (ngày)", 12), ("Priority", 12),
                ("Lý do", 48),
            ],
            [
                [
                    idx + 1, i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
                    process_code(i),
                    i.get("completed_phase", ""), i.get("waiting_phase", ""),
                    i.get("completed_date", ""), i.get("waiting_end_date", ""),
                    i.get("wait_days", 0), i.get("priority", ""),
                    reason_stalled(i),
                ]
                for idx, i in enumerate(items)
            ],
        )

    elif chart == "unassigned":
        items = metrics.get("unassigned_tasks") or []
        write_summary(
            "TỔNG HỢP TASK CHƯA CÓ PIC",
            [("STT", 6), ("Nhóm", 28), ("Số lượng", 12)],
            (
                [[r[0], f"Module: {r[1]}", r[2]] for r in _count_by(items, "module")]
                + [[r[0] + 1000, f"Phase: {r[1]}", r[2]] for r in _count_by(items, "phase")]
            ),
        )
        detail_from_items(
            "CHI TIẾT TASK CHƯA CÓ PIC",
            [
                ("STT", 6), ("Mã CN", 14), ("Rlog ID", 14), ("Tên chức năng", 40), ("Module", 10),
                ("Quy trình", 22), ("Phase trước", 16), ("Phase", 16), ("Status", 12),
                ("Priority", 12), ("Start", 13), ("Deadline", 13), ("Trễ (ngày)", 12),
                ("Lý do", 48),
            ],
            [
                [
                    idx + 1, i.get("ma_cn", ""), i.get("rlog_id", ""), i.get("ten_cn", ""),
                    i.get("module", ""), process_code(i),
                    i.get("predecessor_phase", ""),
                    i.get("phase", ""), i.get("status", ""), i.get("priority", ""),
                    i.get("start_date", ""),
                    i.get("end_date", ""), i.get("days_overdue", 0),
                    reason_unassigned(i),
                ]
                for idx, i in enumerate(items)
            ],
        )

    elif chart == "burndown":
        bd = metrics.get("burndown_velocity") or metrics.get("burndown") or {}
        weeks = bd.get("weeks") or []
        closed = bd.get("closed_per_week") or []
        cum = bd.get("cumulative") or []
        write_summary(
            "BURNDOWN / VELOCITY THEO TUẦN",
            [
                ("STT", 6), ("Tuần (Monday)", 16), ("Closed / tuần", 14),
                ("Lũy kế", 12), ("Velocity 4w", 12),
            ],
            [
                [
                    idx + 1, weeks[idx],
                    closed[idx] if idx < len(closed) else 0,
                    cum[idx] if idx < len(cum) else 0,
                    bd.get("velocity_4w", "") if idx == len(weeks) - 1 else "",
                ]
                for idx in range(len(weeks))
            ],
        )
        # Chi tiết: mỗi closed event
        if do_det and parsed_data is not None:
            from analyzer.advanced_metrics import _parse_iso, _week_monday

            detail_rows = []
            scope = (bd.get("scope_phase") or "").strip()
            for r in getattr(parsed_data, "rows", []) or []:
                meta = _func_meta(r)
                last_upd = _parse_iso((r.meta or {}).get("last_updated"))
                for ph, pd in (r.phases or {}).items():
                    if getattr(pd, "status", None) != "Closed":
                        continue
                    if scope and ph != scope:
                        continue
                    event = _parse_iso(getattr(pd, "end_date", None)) or last_upd
                    if event is None:
                        continue
                    detail_rows.append(
                        meta_vals(len(detail_rows), meta)
                        + [ph, event.isoformat(), _week_monday(event).isoformat()]
                    )
            write_detail(
                "CHI TIẾT SỰ KIỆN CLOSED (burndown)",
                list(meta_cols) + [
                    ("Phase", 16), ("Ngày Closed", 13), ("Tuần (Monday)", 14),
                ],
                detail_rows,
            )

    # Guard: mode không tạo sheet nào (không xảy ra) — tạo sheet trống
    if book._first:
        book.sheet("Empty")

    os.makedirs(output_dir, exist_ok=True)
    filename = f"Chart_{chart}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    book.wb.save(filepath)
    book.wb.close()
    return filepath


def export_audit_report(
    parsed,
    metrics: dict[str, Any],
    output_dir: str = "uploads",
    subtitle: str = "",
) -> str:
    """
    Xuất Report Đánh giá 11 sheet từ ParsedData + metrics.

    Returns filepath .xlsx.
    """
    from analyzer.audit_report import build_audit_issues, AUDIT_SHEET_NAMES

    issues = build_audit_issues(parsed, metrics)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sub = subtitle or f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"

    # 01 Summary
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[0])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws["A1"] = "REPORT ĐÁNH GIÁ CHẤT LƯỢNG DỮ LIỆU"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws["A2"] = sub
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")
    summary = issues["summary"]
    labels = [
        ("Tổng chức năng", "total_functions"),
        ("Số Module", "modules_count"),
        ("Số Phase", "phases_count"),
        ("Tiến độ chung (%)", "overall_progress_pct"),
        ("Meta thiếu", "missing_meta_count"),
        ("Date lỗi", "date_errors_count"),
        ("Status lệch/thiếu", "status_errors_count"),
        ("PIC blacklist", "pic_blacklist_count"),
        ("Estimate MH rejected", "estimate_rejected_count"),
        ("Unassigned", "unassigned_count"),
        ("Overdue", "overdue_count"),
        ("Đình trệ", "stalled_count"),
        ("High risk (>=50)", "high_risk_count"),
        ("Discrepancy", "discrepancy_count"),
    ]
    for idx, (label, key) in enumerate(labels, start=4):
        ws.cell(row=idx, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=idx, column=1).border = THIN_BORDER
        c = ws.cell(row=idx, column=2, value=summary.get(key, 0))
        c.font = Font(name="Arial", size=11)
        c.border = THIN_BORDER

    # 02 Meta thiếu
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[1])
    items = issues["missing_meta"]
    _write_sheet(
        ws, "META THIẾU",
        [("STT", 6), ("Row", 8), ("Mã CN", 14), ("Tên CN", 36), ("Module", 10), ("Thiếu", 40)],
        [[idx + 1, i.get("row_index", ""), i.get("ma_cn", ""), i.get("ten_cn", ""),
          i.get("module", ""), i.get("missing_fields", "")] for idx, i in enumerate(items)],
        subtitle=sub,
    )

    # 03 Date lỗi
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[2])
    items = issues["date_errors"]
    _write_sheet(
        ws, "DATE LỖI (End < Start)",
        [("STT", 6), ("Row", 8), ("Mã CN", 14), ("Tên CN", 36), ("Module", 10),
         ("Phase", 16), ("Start", 13), ("End", 13), ("Issue", 18)],
        [[idx + 1, i.get("row_index", ""), i.get("ma_cn", ""), i.get("ten_cn", ""),
          i.get("module", ""), i.get("phase", ""), i.get("start_date", ""),
          i.get("end_date", ""), i.get("issue", "")] for idx, i in enumerate(items)],
        subtitle=sub,
    )

    # 04 Status lệch
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[3])
    items = issues["status_errors"]
    _write_sheet(
        ws, "STATUS LỆCH / THIẾU",
        [("STT", 6), ("Row", 8), ("Mã CN", 14), ("Tên CN", 36), ("Module", 10),
         ("Phase", 16), ("Issue", 28), ("Chi tiết", 45)],
        [[idx + 1, i.get("row_index", ""), i.get("ma_cn", ""), i.get("ten_cn", ""),
          i.get("module", ""), i.get("phase", ""), i.get("issue", ""),
          i.get("detail", "")] for idx, i in enumerate(items)],
        subtitle=sub,
    )

    # 05 PIC blacklist
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[4])
    items = issues["pic_blacklist"]
    _write_sheet(
        ws, "PIC BLACKLIST",
        [("STT", 6), ("Row", 8), ("Mã CN", 14), ("Module", 10), ("Phase", 16),
         ("Header", 22), ("Raw value", 16), ("Keyword", 14)],
        [[idx + 1, i.get("row_index", ""), i.get("ma_cn", ""), i.get("module", ""),
          i.get("phase_name", ""), i.get("header_text", ""), i.get("raw_value", ""),
          i.get("matched_keyword", "")] for idx, i in enumerate(items)],
        subtitle=sub, row_fill_fn=lambda r, o: YELLOW_FILL,
    )

    # 06 Estimate rejected
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[5])
    items = issues["estimate_rejected"]
    _write_sheet(
        ws, "ESTIMATE MH BỊ REJECT",
        [("STT", 6), ("Row", 8), ("Mã CN", 14), ("Module", 10), ("Phase", 16),
         ("Header", 24), ("Raw value", 22), ("Reason", 22)],
        [[idx + 1, i.get("row_index", ""), i.get("ma_cn", ""), i.get("module", ""),
          i.get("phase_name", ""), i.get("header", ""), i.get("raw_value", ""),
          i.get("reason", "")] for idx, i in enumerate(items)],
        subtitle=sub, row_fill_fn=lambda r, o: ORANGE_FILL,
    )

    # 07 Unassigned
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[6])
    items = issues["unassigned"]
    _write_sheet(
        ws, "UNASSIGNED",
        [("STT", 6), ("Mã CN", 14), ("Rlog ID", 14), ("Tên CN", 40), ("Module", 10),
         ("Quy trình", 22), ("Phase trước", 16), ("Phase", 16), ("Status", 12),
         ("Priority", 12), ("Start", 13), ("Deadline", 13), ("Lý do", 48)],
        [[idx + 1, i.get("ma_cn", ""), i.get("rlog_id", ""), i.get("ten_cn", ""),
          i.get("module", ""), process_code(i),
          i.get("predecessor_phase", ""),
          i.get("phase", ""), i.get("status", ""), i.get("priority", ""),
          i.get("start_date", ""), i.get("end_date", ""),
          reason_unassigned(i)] for idx, i in enumerate(items)],
        subtitle=sub,
    )

    # 08 Overdue
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[7])
    items = issues["overdue"]
    _write_sheet(
        ws, "OVERDUE",
        [("STT", 6), ("Mã CN", 14), ("Tên CN", 40), ("Module", 10), ("Quy trình", 22),
         ("Phase", 16), ("Deadline", 13), ("Ngày trễ", 10), ("Status", 12),
         ("PIC", 20), ("Ghi chú", 30), ("Lý do", 48)],
        [[idx + 1, i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
          process_code(i),
          i.get("phase", ""), i.get("end_date", ""), i.get("days_overdue", 0),
          i.get("status", ""),
          ", ".join(i.get("pic") or []) if isinstance(i.get("pic"), list) else "",
          i.get("note", ""),
          reason_overdue(i)]
         for idx, i in enumerate(items)],
        subtitle=sub,
        row_fill_fn=lambda ri, idx: _fill_by_days(items[idx].get("days_overdue", 0)),
    )

    # 09 Đình trệ
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[8])
    items = issues["stalled"]
    _write_sheet(
        ws, "ĐÌNH TRỆ",
        [("STT", 6), ("Mã CN", 14), ("Tên CN", 40), ("Module", 10), ("Quy trình", 22),
         ("Phase xong", 16), ("Phase chờ", 16), ("End chờ", 13),
         ("Chờ (ngày)", 12), ("Priority", 12), ("Lý do", 48)],
        [[idx + 1, i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
          process_code(i),
          i.get("completed_phase", ""), i.get("waiting_phase", ""),
          i.get("waiting_end_date", ""),
          i.get("wait_days", 0), i.get("priority", ""),
          reason_stalled(i)] for idx, i in enumerate(items)],
        subtitle=sub,
    )

    # 10 High risk
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[9])
    items = issues["high_risk"]
    _write_sheet(
        ws, "HIGH RISK (>=50)",
        [("STT", 6), ("Score", 10), ("Mã CN", 14), ("Tên CN", 40), ("Module", 10),
         ("Quy trình", 22), ("Priority", 12), ("Factors", 45), ("Yếu tố chi tiết", 60)],
        [[idx + 1, i.get("risk_score", 0), i.get("ma_cn", ""), i.get("ten_cn", ""),
          i.get("module", ""), process_code(i), i.get("priority", ""),
          " | ".join(i.get("risk_factors") or []),
          format_risk_factors_detailed(i)] for idx, i in enumerate(items)],
        subtitle=sub,
        row_fill_fn=lambda ri, idx: _fill_by_risk(items[idx].get("risk_score", 0)),
    )

    # 11 Discrepancy
    ws = wb.create_sheet(AUDIT_SHEET_NAMES[10])
    items = issues["discrepancy"]
    _write_sheet(
        ws, "DISCREPANCY / MÂU THUẪN",
        [("STT", 6), ("Mã CN", 14), ("Tên CN", 36), ("Module", 10),
         ("Issue", 28), ("Chi tiết", 50)],
        [[idx + 1, i.get("ma_cn", ""), i.get("ten_cn", ""), i.get("module", ""),
          i.get("issue", ""), i.get("detail", "")] for idx, i in enumerate(items)],
        subtitle=sub,
    )

    os.makedirs(output_dir, exist_ok=True)
    filename = f"Audit_Report_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ======================================================================
# PHASE 4/5 EXPORTS: SLA / Capacity / Slow / Baseline
# Rule (V4): XEM THÌ PHÂN TRANG NHƯNG XUẤT LÀ XUẤT ALL RECORD
# ======================================================================


def _severity_fill(sev: str) -> Optional[PatternFill]:
    """Highlight theo severity: critical=đỏ, warning=cam."""
    s = (sev or "").lower()
    if s == "critical":
        return RED_FILL
    if s == "warning":
        return ORANGE_FILL
    return None


def _filter_subtitle(filters: Optional[dict]) -> str:
    """Build subtitle chuẩn: 'Ngày xuất: dd/mm/yyyy  |  Bộ lọc: Module=..., Quy trình=..., PIC=...'."""
    date_part = f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    if not filters:
        return date_part
    parts = []
    for key, label in [("modules", "Module"), ("processes", "Quy trình"), ("pics", "PIC")]:
        vals = filters.get(key) or filters.get(key.rstrip("s")) or []
        if isinstance(vals, str):
            vals = [vals] if vals else []
        if vals:
            parts.append(f"{label}=[{', '.join(vals)}]")
    return f"{date_part}  |  Bộ lọc: {' · '.join(parts)}" if parts else date_part


def export_sla_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
) -> str:
    """
    Xuất báo cáo SLA vi phạm deadline theo Priority.

    Args:
        payload: kết quả compute_sla_violations {items, total, critical_count,
                 warning_count, thresholds}
        output_dir: thư mục lưu file
        filters: dict global filter đã áp trước khi compute (chỉ để in vào subtitle)

    Rule V4: XUẤT ALL items, không cắt theo pagination FE.
    """
    items = list(payload.get("items") or [])
    thresholds = payload.get("thresholds") or {}
    critical_count = payload.get("critical_count", 0)
    warning_count = payload.get("warning_count", 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLA_Violations"

    subtitle = _filter_subtitle(filters)
    subtitle += (f"  |  Must-have quá {thresholds.get('must_have_days', '?')}d / "
                 f"Should-have quá {thresholds.get('should_have_days', '?')}d → critical")
    subtitle += f"  |  Critical: {critical_count} · Warning: {warning_count}"

    _write_sheet(
        ws,
        title="BÁO CÁO SLA — VI PHẠM DEADLINE",
        columns=[
            ("STT", 6),
            ("Severity", 12),
            ("Mã CN", 14),
            ("Tên chức năng", 40),
            ("Module", 10),
            ("Quy trình", 22),
            ("Priority", 14),
            ("Phase", 16),
            ("End (deadline)", 14),
            ("Số ngày trễ", 12),
            ("Ngưỡng (d)", 12),
            ("Status", 12),
            ("PIC", 24),
        ],
        data_rows=[
            [
                idx + 1,
                (i.get("severity") or "").upper(),
                i.get("ma_cn", ""),
                i.get("ten_cn", ""),
                i.get("module", ""),
                process_code(i),
                i.get("priority", ""),
                i.get("phase", ""),
                i.get("end_date", ""),
                i.get("days_late", 0),
                i.get("threshold_days", 0),
                i.get("status", ""),
                ", ".join(i.get("pics") or []),
            ]
            for idx, i in enumerate(items)
        ],
        row_fill_fn=lambda _ri, idx: _severity_fill(items[idx].get("severity")),
        subtitle=subtitle,
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"SLA_Violations_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


def _overload_fill(overload: bool) -> Optional[PatternFill]:
    return RED_FILL if overload else None


def export_capacity_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
) -> str:
    """
    Xuất báo cáo Capacity PIC — remaining MH vs công suất tuần.

    Args:
        payload: kết quả compute_capacity_load {by_pic, default_md_per_week, overload_count}

    Rule V4: XUẤT ALL PIC rows.
    """
    rows = list(payload.get("by_pic") or [])
    default_md = payload.get("default_md_per_week")
    overload_count = payload.get("overload_count", 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capacity_PIC"

    subtitle = _filter_subtitle(filters)
    if default_md:
        subtitle += f"  |  Default: {default_md} MD/tuần"
    subtitle += f"  |  Overload: {overload_count}"

    _write_sheet(
        ws,
        title="BÁO CÁO CAPACITY PIC",
        columns=[
            ("STT", 6),
            ("PIC", 20),
            ("Remaining (MH)", 15),
            ("Closed (MH)", 15),
            ("Capacity (MH/tuần)", 18),
            ("Số tuần cần", 14),
            ("Overload?", 12),
            ("Lý do", 40),
        ],
        data_rows=[
            [
                idx + 1,
                r.get("pic", ""),
                r.get("remaining_mh", 0),
                r.get("closed_mh", 0),
                r.get("capacity_mh_per_week", 0),
                r.get("weeks_needed") if r.get("weeks_needed") is not None else "",
                "OVERLOAD" if r.get("overload") else "",
                reason_capacity(r),
            ]
            for idx, r in enumerate(rows)
        ],
        row_fill_fn=lambda _ri, idx: _overload_fill(bool(rows[idx].get("overload"))),
        subtitle=subtitle,
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"Capacity_PIC_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


def _slow_cell_fill(count: int) -> Optional[PatternFill]:
    """Highlight ô heatmap theo số phase-record trễ."""
    if count >= 10:
        return RED_FILL
    if count >= 5:
        return ORANGE_FILL
    if count >= 1:
        return YELLOW_FILL
    return None


def export_slow_heatmap_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
) -> str:
    """
    Xuất báo cáo "Ai đang chậm" — Heatmap PIC × Phase.

    Sheet 1: Matrix PIC × Phase (highlight theo count).
    Sheet 2: Flat list (chỉ ô > 0) — dễ pivot.

    Rule V4: XUẤT ALL PIC × Phase records.
    """
    pics = list(payload.get("pics") or [])
    phases = list(payload.get("phases") or [])
    heatmap = payload.get("heatmap") or {}
    total_slow = payload.get("total_slow", 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heatmap_PIC_x_Phase"

    subtitle = _filter_subtitle(filters) + f"  |  Tổng phase-record trễ: {total_slow}"

    # === Sheet 1: matrix ===
    n_cols = 1 + len(phases) + 1  # cột PIC + N phases + Σ
    last_col_letter = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    tc = ws["A1"]
    tc.value = "HEATMAP PIC × PHASE — Task đang chậm"
    tc.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{last_col_letter}2")
    sc = ws["A2"]
    sc.value = subtitle
    sc.font = Font(name="Arial", size=10, italic=True, color="666666")
    sc.alignment = Alignment(horizontal="center")

    hdr_row = 4
    cells = ["PIC"] + phases + ["Σ"]
    widths = [22] + [max(12, len(ph) + 2) for ph in phases] + [10]
    for c_idx, (name, w) in enumerate(zip(cells, widths), 1):
        cell = ws.cell(row=hdr_row, column=c_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[hdr_row].height = 25

    for row_offset, pic in enumerate(pics):
        r_idx = hdr_row + 1 + row_offset
        row_data = heatmap.get(pic) or {}
        row_sum = 0
        pic_cell = ws.cell(row=r_idx, column=1, value=pic)
        pic_cell.font = Font(name="Arial", bold=True, size=10)
        pic_cell.border = THIN_BORDER
        pic_cell.alignment = BODY_ALIGN
        for ph_idx, ph in enumerate(phases, 2):
            v = int(row_data.get(ph) or 0)
            row_sum += v
            cell = ws.cell(row=r_idx, column=ph_idx, value=v if v > 0 else "")
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            fill = _slow_cell_fill(v)
            if fill:
                cell.fill = fill
        sum_cell = ws.cell(row=r_idx, column=n_cols, value=row_sum)
        sum_cell.font = Font(name="Arial", bold=True, size=10)
        sum_cell.alignment = Alignment(horizontal="center")
        sum_cell.border = THIN_BORDER

    if pics:
        total_row = hdr_row + len(pics) + 2
        ws.merge_cells(f"A{total_row}:{last_col_letter}{total_row}")
        s = ws.cell(row=total_row, column=1)
        s.value = (f"Tổng: {len(pics)} PIC × {len(phases)} phase | "
                   f"{total_slow} phase-record trễ | "
                   f"Xuất ngày {date.today().strftime('%d/%m/%Y')}")
        s.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
    ws.freeze_panes = f"B{hdr_row + 1}"

    # === Sheet 2: flat list ===
    ws2 = wb.create_sheet("Flat_List")
    flat_rows = []
    for pic in pics:
        row_data = heatmap.get(pic) or {}
        for ph in phases:
            v = int(row_data.get(ph) or 0)
            if v > 0:
                flat_rows.append([len(flat_rows) + 1, pic, ph, v])
    _write_sheet(
        ws2,
        title="FLAT LIST — PIC × Phase với count > 0",
        columns=[
            ("STT", 6),
            ("PIC", 22),
            ("Phase", 20),
            ("Số phase-record trễ", 18),
        ],
        data_rows=flat_rows,
        row_fill_fn=lambda _ri, idx: _slow_cell_fill(flat_rows[idx][3]),
        subtitle=subtitle,
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"Slow_Heatmap_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


def _variance_fill(days: int) -> Optional[PatternFill]:
    """Highlight theo variance ngày (trễ hoặc sớm nhiều đều đáng để ý)."""
    a = abs(int(days or 0))
    if a >= 14:
        return RED_FILL
    if a >= 7:
        return ORANGE_FILL
    if a >= 3:
        return YELLOW_FILL
    return None


def export_baseline_variance_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
    all_items: Optional[list[dict]] = None,
) -> str:
    """
    Xuất báo cáo Baseline vs Actual variance ngày.

    Args:
        payload: kết quả compute_baseline_variance {items (max 200), total_compared, ...}
        all_items: OPTIONAL — full list khi caller cần bỏ giới hạn 200 của FE.
                   Nếu None thì dùng payload["items"] (đã bị cắt 200).

    Rule V4: XUẤT ALL items — caller (endpoint) nên truyền `all_items` để bỏ giới hạn.
    """
    items = list(all_items) if all_items is not None else list(payload.get("items") or [])
    total_compared = payload.get("total_compared", len(items))
    late_count = payload.get("late_count", sum(1 for i in items if i.get("late")))
    avg_var = payload.get("avg_variance_days", 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baseline_vs_Actual"

    subtitle = _filter_subtitle(filters)
    subtitle += (f"  |  So sánh: {total_compared} · Trễ: {late_count} · "
                 f"Avg variance: {avg_var}d")

    _write_sheet(
        ws,
        title="BÁO CÁO BASELINE vs ACTUAL — VARIANCE NGÀY",
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 40),
            ("Module", 10),
            ("Quy trình", 22),
            ("Phase", 16),
            ("Plan date", 14),
            ("Actual date", 14),
            ("Variance (ngày)", 14),
            ("Trễ?", 8),
            ("Status", 12),
        ],
        data_rows=[
            [
                idx + 1,
                i.get("ma_cn", ""),
                i.get("ten_cn", ""),
                i.get("module", ""),
                process_code(i),
                i.get("phase", ""),
                i.get("plan_date", ""),
                i.get("actual_date", ""),
                i.get("variance_days", 0),
                "YES" if i.get("late") else "",
                i.get("status", ""),
            ]
            for idx, i in enumerate(items)
        ],
        row_fill_fn=lambda _ri, idx: _variance_fill(items[idx].get("variance_days", 0)),
        subtitle=subtitle,
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"Baseline_Variance_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


# --------------------------------------------------------------------------
# FIT/GAP Dashboard export (Task 2). Multi-sheet: Summary + Aging + by-Module +
# by-Process + by-Priority. Rule V4: xuất ALL — không cắt aging list theo pagination.
# --------------------------------------------------------------------------

def _aging_fill(days: Optional[int]) -> Optional[PatternFill]:
    """Fill row aging: >= 30d red, 21-29d orange, 14-20d yellow."""
    if days is None:
        return None
    if days >= 30:
        return RED_FILL
    if days >= 21:
        return ORANGE_FILL
    if days >= 14:
        return YELLOW_FILL
    return None


def export_fitgap_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    filters: Optional[dict] = None,
) -> str:
    """
    Xuất báo cáo FIT/GAP Dashboard sang Excel (multi-sheet).

    Rule V4: XUẤT ALL. Aging list dùng `payload["aging_items"]` (đã filter theo
    threshold ở backend); nếu caller muốn xuất TẤT CẢ GAP đang mở (không filter
    theo threshold), truyền vào `payload["all_open_gap_items"]` thay thế.
    """
    summary = payload.get("summary") or {}
    by_module = payload.get("by_module") or []
    by_process = payload.get("by_process") or []
    by_priority = payload.get("by_priority") or []
    aging_items = payload.get("aging_items") or []

    wb = openpyxl.Workbook()

    subtitle_base = _filter_subtitle(filters)
    thr = summary.get("aging_threshold_days", 14)

    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "Summary"
    _write_sheet(
        ws,
        title="FIT/GAP DASHBOARD — TỔNG QUAN",
        subtitle=subtitle_base + f"  |  Aging threshold: {thr} ngày",
        columns=[("Chỉ số", 30), ("Số lượng", 14)],
        data_rows=[
            ["Tổng function", summary.get("total", 0)],
            ["FIT", summary.get("fit", 0)],
            ["GAP (tổng)", summary.get("gap", 0)],
            ["GAP đã đóng (all phase Closed/Cancelled)", summary.get("gap_closed", 0)],
            ["GAP đang mở", summary.get("gap_open", 0)],
            [f"GAP aging > {thr} ngày", summary.get("gap_open_aging", 0)],
        ],
    )

    # === Sheet 2: Aging GAP list ===
    ws2 = wb.create_sheet("Aging_GAP")
    _write_sheet(
        ws2,
        title=f"GAP AGING — ĐANG MỞ > {thr} NGÀY",
        subtitle=subtitle_base + f"  |  Tổng: {len(aging_items)}",
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 40),
            ("Module", 10),
            ("Quy trình", 30),
            ("Priority", 14),
            ("Ngày mở", 14),
            ("Aging (ngày)", 12),
            ("Ngưỡng", 10),
            ("Phase đang mở", 14),
            ("Status", 12),
            ("PIC", 24),
            ("Lý do", 40),
        ],
        data_rows=[
            [
                idx + 1,
                it.get("ma_cn", ""),
                it.get("ten_cn", ""),
                it.get("module", ""),
                it.get("quy_trinh", ""),
                it.get("priority", ""),
                it.get("opened_date") or "N/A",
                it.get("aging_days") if it.get("aging_days") is not None else "N/A",
                thr,
                it.get("current_phase", ""),
                it.get("status", ""),
                ", ".join(it.get("pics") or []),
                reason_fitgap_aging(it, thr),
            ]
            for idx, it in enumerate(aging_items)
        ],
        row_fill_fn=lambda _ri, idx: _aging_fill(aging_items[idx].get("aging_days")),
    )

    def _breakdown_sheet(sheet_name: str, title: str, rows: list[dict], key_field: str, label: str):
        ws_bk = wb.create_sheet(sheet_name)
        _write_sheet(
            ws_bk,
            title=title,
            subtitle=subtitle_base,
            columns=[
                ("STT", 6),
                (label, 32),
                ("FIT", 10),
                ("GAP", 10),
                ("Tổng", 10),
                ("% GAP", 10),
            ],
            data_rows=[
                [idx + 1, r.get(key_field, ""), r.get("fit", 0), r.get("gap", 0),
                 r.get("total", 0), f"{r.get('pct_gap', 0)}%"]
                for idx, r in enumerate(rows)
            ],
        )

    _breakdown_sheet("By_Module", "FIT/GAP THEO MODULE", by_module, "module", "Module")
    _breakdown_sheet("By_Process", "FIT/GAP THEO QUY TRÌNH", by_process, "process", "Quy trình")
    _breakdown_sheet("By_Priority", "FIT/GAP THEO PRIORITY", by_priority, "priority", "Priority")

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"FITGAP_Dashboard_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


# --------------------------------------------------------------------------
# Function Diff export (Task 3) — multi-sheet, mỗi tab 1 sheet.
# Rule V4: XUẤT ALL (không pagination).
# --------------------------------------------------------------------------

def export_function_diff_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
) -> str:
    """
    Xuất báo cáo diff giữa 2 snapshot sang Excel (multi-sheet).
    Mỗi tab trong UI = 1 sheet: Summary / Added / Deleted / PIC / Priority
    (kèm Complexity) / FIT_GAP / Status.
    """
    cur_meta = payload.get("current_snapshot") or {}
    prev_meta = payload.get("previous_snapshot") or {}
    counts = payload.get("counts") or {}

    wb = openpyxl.Workbook()

    subtitle_base = (
        f"Hiện tại: {cur_meta.get('date', '—')} ({cur_meta.get('filename', '—')})  "
        f"|  Trước: {prev_meta.get('date', '—')} ({prev_meta.get('filename', '—')})"
    )

    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "Summary"
    _write_sheet(
        ws,
        title="FUNCTION DIFF — TỔNG QUAN THAY ĐỔI",
        subtitle=subtitle_base,
        columns=[("Chỉ số", 40), ("Số lượng", 14)],
        data_rows=[
            ["Tổng function hiện tại", counts.get("current_total", 0)],
            ["Tổng function snapshot trước", counts.get("previous_total", 0)],
            ["+ Mới thêm", counts.get("added", 0)],
            ["- Bị xoá", counts.get("deleted", 0)],
            ["⇄ Function đổi (distinct)", counts.get("total_changed", 0)],
            ["   ↳ Đổi PIC (bản ghi)", counts.get("pic_changed", 0)],
            ["   ↳ Đổi Priority/Complexity", counts.get("prio_complex_changed", 0)],
            ["   ↳ Đổi FIT/GAP", counts.get("fitgap_changed", 0)],
            ["   ↳ Đổi Status phase (bản ghi)", counts.get("status_changed", 0)],
        ],
    )

    # === Sheet 2: Added ===
    ws2 = wb.create_sheet("Added")
    _write_sheet(
        ws2,
        title="FUNCTION MỚI THÊM",
        subtitle=subtitle_base,
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 40),
            ("Module", 10),
            ("Quy trình", 30),
            ("Priority", 12),
            ("Complexity", 12),
            ("FIT/GAP", 10),
        ],
        data_rows=[
            [idx + 1, r.get("ma_cn", ""), r.get("ten_cn", ""), r.get("module", ""),
             r.get("quy_trinh", ""), r.get("priority", ""), r.get("complexity", ""),
             r.get("fit_gap", "")]
            for idx, r in enumerate(payload.get("added") or [])
        ],
        row_fill_fn=lambda _ri, _idx: GREEN_FILL,
    )

    # === Sheet 3: Deleted ===
    ws3 = wb.create_sheet("Deleted")
    _write_sheet(
        ws3,
        title="FUNCTION BỊ XOÁ",
        subtitle=subtitle_base,
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 40),
            ("Module", 10),
            ("Quy trình", 30),
            ("Priority", 12),
            ("Complexity", 12),
            ("FIT/GAP", 10),
        ],
        data_rows=[
            [idx + 1, r.get("ma_cn", ""), r.get("ten_cn", ""), r.get("module", ""),
             r.get("quy_trinh", ""), r.get("priority", ""), r.get("complexity", ""),
             r.get("fit_gap", "")]
            for idx, r in enumerate(payload.get("deleted") or [])
        ],
        row_fill_fn=lambda _ri, _idx: RED_FILL,
    )

    # === Sheet 4: PIC change ===
    ws4 = wb.create_sheet("PIC_Change")
    _write_sheet(
        ws4,
        title="THAY ĐỔI PIC",
        subtitle=subtitle_base,
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 30),
            ("Module", 10),
            ("Quy trình", 22),
            ("Phase", 16),
            ("PIC cũ", 30),
            ("PIC mới", 30),
        ],
        data_rows=[
            [idx + 1, r.get("ma_cn", ""), r.get("ten_cn", ""), r.get("module", ""),
             process_code(r), r.get("phase", ""), r.get("old", ""), r.get("new", "")]
            for idx, r in enumerate(payload.get("pic_changed") or [])
        ],
    )

    # === Sheet 5: Priority / Complexity change ===
    ws5 = wb.create_sheet("PrioComplex_Change")
    _write_sheet(
        ws5,
        title="THAY ĐỔI PRIORITY / COMPLEXITY",
        subtitle=subtitle_base,
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 30),
            ("Module", 10),
            ("Quy trình", 22),
            ("Field", 16),
            ("Cũ", 20),
            ("Mới", 20),
        ],
        data_rows=[
            [idx + 1, r.get("ma_cn", ""), r.get("ten_cn", ""), r.get("module", ""),
             process_code(r), r.get("field", ""), r.get("old", ""), r.get("new", "")]
            for idx, r in enumerate(payload.get("priority_complexity_changed") or [])
        ],
    )

    # === Sheet 6: FIT/GAP change ===
    ws6 = wb.create_sheet("FITGAP_Change")
    _write_sheet(
        ws6,
        title="THAY ĐỔI FIT / GAP",
        subtitle=subtitle_base,
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 30),
            ("Module", 10),
            ("Quy trình", 22),
            ("Cũ", 14),
            ("Mới", 14),
        ],
        data_rows=[
            [idx + 1, r.get("ma_cn", ""), r.get("ten_cn", ""), r.get("module", ""),
             process_code(r), r.get("old", ""), r.get("new", "")]
            for idx, r in enumerate(payload.get("fitgap_changed") or [])
        ],
    )

    # === Sheet 7: Phase status change ===
    ws7 = wb.create_sheet("Status_Change")
    _write_sheet(
        ws7,
        title="THAY ĐỔI STATUS PHASE",
        subtitle=subtitle_base,
        columns=[
            ("STT", 6),
            ("Mã CN", 14),
            ("Tên chức năng", 30),
            ("Module", 10),
            ("Quy trình", 22),
            ("Phase", 16),
            ("Status cũ", 14),
            ("Status mới", 14),
        ],
        data_rows=[
            [idx + 1, r.get("ma_cn", ""), r.get("ten_cn", ""), r.get("module", ""),
             process_code(r), r.get("phase", ""), r.get("old", ""), r.get("new", "")]
            for idx, r in enumerate(payload.get("phase_status_changed") or [])
        ],
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"Function_Diff_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# T22 — Aging WIP Report
# ==========================================================================


def export_aging_wip_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    subtitle: str = "",
) -> str:
    """
    Xuất báo cáo Excel Aging WIP.
    payload từ analyzer.advanced_metrics.compute_aging_wip().
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AgingWIP"
    items = payload.get("items") or []
    items = [it for it in items if it.get("is_aging")]
    summary = payload.get("summary") or {}
    threshold = payload.get("threshold_days", 14)

    columns = [
        ("STT", 6),
        ("Row #", 8),
        ("Mã CN", 14),
        ("Tên chức năng", 32),
        ("Module", 10),
        ("Quy trình", 24),
        ("Phase", 16),
        ("Status", 12),
        ("Start", 12),
        ("End", 12),
        ("PIC", 20),
        ("Aging (ngày)", 12),
        ("Ngưỡng (ngày)", 12),
        ("Over ngưỡng", 12),
        ("Priority", 12),
        ("Complexity", 12),
        ("Lý do", 40),
    ]
    data_rows = [
        [
            idx + 1,
            it.get("row_num", ""),
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("module", ""),
            it.get("quy_trinh", ""),
            it.get("phase", ""),
            it.get("status", ""),
            it.get("start_date", ""),
            it.get("end_date", ""),
            it.get("pic", ""),
            it.get("aging_days", 0),
            threshold,
            it.get("over_by_days", 0),
            it.get("priority", ""),
            it.get("complexity", ""),
            reason_aging_wip(it, threshold),
        ]
        for idx, it in enumerate(items)
    ]

    # Tô màu theo mức độ over: <7d = vàng, 7-30d = cam, >30d = đỏ.
    def _fill(_row_idx: int, offset: int):
        if offset >= len(items):
            return None
        over = items[offset].get("over_by_days", 0)
        if over > 30:
            return RED_FILL
        if over >= 7:
            return ORANGE_FILL
        return YELLOW_FILL

    _write_sheet(
        ws,
        title=f"AGING WIP — In-progress > {threshold} ngày",
        subtitle=subtitle or (
            f"Tổng WIP: {summary.get('total_wip', 0)} | Aging: {summary.get('total_aging', 0)} | "
            f"Avg={summary.get('avg_aging_days', 0)}d Max={summary.get('max_aging_days', 0)}d | "
            f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
        ),
        columns=columns,
        data_rows=data_rows,
        row_fill_fn=_fill,
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"Aging_WIP_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# T21 — Data Quality Report
# ==========================================================================

# Mapping severity → fill (đỏ/cam/vàng) để user scan nhanh.
_DQ_SEVERITY_FILL = {
    "high": RED_FILL,
    "medium": ORANGE_FILL,
    "low": YELLOW_FILL,
}


def export_data_quality_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    subtitle: str = "",
) -> str:
    """
    Xuất báo cáo Excel Data Quality (2 sheet).

    Args:
        payload: dict trả về từ analyzer.data_quality.compute_data_quality()
                 {issues: [...], summary: {...}}
    Returns:
        Filepath .xlsx đã tạo.
    """
    wb = openpyxl.Workbook()
    summary = payload.get("summary") or {}
    issues = payload.get("issues") or []

    # === Sheet 1: Summary ===
    ws1 = wb.active
    ws1.title = "Summary"
    by_sev = summary.get("by_severity") or {}
    by_code = summary.get("by_code") or {}
    summary_rows = [
        ["Tổng function", summary.get("total_rows", 0)],
        ["Function có issue", summary.get("affected_rows", 0)],
        ["Function clean", summary.get("clean_rows", 0)],
        ["% Clean", f"{summary.get('clean_pct', 0)}%"],
        ["", ""],
        ["Tổng issue", summary.get("total_issues", 0)],
        ["  Severity: High", by_sev.get("high", 0)],
        ["  Severity: Medium", by_sev.get("medium", 0)],
        ["  Severity: Low", by_sev.get("low", 0)],
    ]
    # Thêm break-down theo code
    if by_code:
        summary_rows.append(["", ""])
        summary_rows.append(["--- Chi tiết theo loại ---", ""])
        for code, cnt in sorted(by_code.items(), key=lambda x: -x[1]):
            summary_rows.append([f"  {code}", cnt])

    _write_sheet(
        ws1,
        title="DATA QUALITY — TỔNG QUAN",
        subtitle=subtitle or f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}",
        columns=[("Chỉ số", 42), ("Giá trị", 18)],
        data_rows=summary_rows,
    )

    # === Sheet 2: Danh sách issue chi tiết ===
    ws2 = wb.create_sheet("Issues")
    dq_columns = [
        ("STT", 6),
        ("Row #", 8),
        ("Mã CN", 14),
        ("Tên chức năng", 32),
        ("Module", 10),
        ("Quy trình", 22),
        ("Phase", 16),
        ("Mã lỗi", 22),
        ("Loại issue", 22),
        ("Severity", 10),
        ("Chi tiết", 30),
        ("Gợi ý xử lý", 40),
    ]
    data_rows = [
        [
            idx + 1,
            it.get("row_num", ""),
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("module", ""),
            process_code(it),
            it.get("phase", ""),
            it.get("code", ""),
            it.get("label", ""),
            it.get("severity", "").upper(),
            it.get("detail", ""),
            it.get("suggestion", ""),
        ]
        for idx, it in enumerate(issues)
    ]

    def _fill(_row_idx: int, offset: int):
        if offset >= len(issues):
            return None
        return _DQ_SEVERITY_FILL.get(issues[offset].get("severity"))

    _write_sheet(
        ws2,
        title="DATA QUALITY — DANH SÁCH ISSUE CHI TIẾT",
        subtitle=(
            f"Tổng: {len(issues)} issue | "
            f"High={by_sev.get('high', 0)} Medium={by_sev.get('medium', 0)} Low={by_sev.get('low', 0)} | "
            f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
        ),
        columns=dq_columns,
        data_rows=data_rows,
        row_fill_fn=_fill,
    )

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"Data_Quality_{date.today().strftime('%Y%m%d')}.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# FID issues — danh sách lỗi theo đúng lưới dashboard
# ==========================================================================

#: Nền vàng cho ô cần PM điền tay (giống quy ước tô vàng của FL re-import).
_FID_INPUT_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

_FID_TYPE_LABELS = {"missing_fid": "Thiếu FID", "duplicate_fid": "Trùng FID"}


def export_fid_issues_report(
    issues: list[dict[str, Any]],
    output_dir: str = "uploads",
    subtitle: str = "",
    project_slug: str = "",
) -> str:
    """
    Xuất danh sách lỗi FID — 7 cột đúng như lưới dashboard (bỏ cột số thứ tự),
    cộng 1 cột trống ``FID cần cập nhật`` tô vàng để PM điền tay.

    File này KHÔNG import lại được (chỉ là danh sách để lọc/soi trong Excel) —
    muốn import thì dùng /export-fl-reimport?kinds=fid. Vì vậy sheet đặt tên
    ``Loi_FID`` chứ không phải ``Function List``: upload nhầm file này sẽ bị
    parser từ chối thay vì ghi đè project bằng vài chục dòng.

    Args:
        issues: list issue từ analyzer.fid_check.compute_fid_issues()["issues"],
                đã áp filter của section.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loi_FID"

    columns = [
        ("Mã CN", 16),
        ("Tên chức năng", 44),
        ("Module", 10),
        ("FID hiện tại", 14),
        ("Loại issue", 12),
        ("Dev phase", 14),
        ("Chi tiết", 52),
        ("FID cần cập nhật", 18),
    ]
    data_rows = [
        [
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("module", ""),
            it.get("fid", "") or "",
            _FID_TYPE_LABELS.get(it.get("issue_type"), it.get("issue_type", "")),
            it.get("dev_phase", ""),
            it.get("detail", ""),
            "",  # PM điền tay
        ]
        for it in issues
    ]

    missing = sum(1 for it in issues if it.get("issue_type") == "missing_fid")
    dup = sum(1 for it in issues if it.get("issue_type") == "duplicate_fid")
    _write_sheet(
        ws,
        title="DEV CLOSED — THIẾU / TRÙNG FID",
        subtitle=(
            subtitle
            or f"Tổng {len(issues)} issue | Thiếu FID={missing} · Trùng FID={dup} "
               f"| Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
        ),
        columns=columns,
        data_rows=data_rows,
    )

    # Tô vàng cột nhập tay (header row 4 → data từ row 5)
    input_col = len(columns)
    for offset in range(len(data_rows)):
        ws.cell(row=5 + offset, column=input_col).fill = _FID_INPUT_FILL

    os.makedirs(output_dir, exist_ok=True)
    suffix = f"_{project_slug}" if project_slug else ""
    filepath = os.path.join(
        output_dir, f"Loi_FID{suffix}_{date.today().strftime('%Y%m%d')}.xlsx"
    )
    wb.save(filepath)
    wb.close()
    return filepath


# ==========================================================================
# Gantt Calendar — Excel-style export
# ==========================================================================

def _hex_to_argb(hex_color: str) -> str:
    """VD '#3b82f6' → '003b82f6' (openpyxl PatternFill dùng ARGB 8-char)."""
    h = (hex_color or "").lstrip("#")
    if len(h) == 3:
        h = "".join(ch * 2 for ch in h)
    if len(h) != 6:
        h = "000000"
    return "00" + h.upper()


# Nhạt hơn cho fill (background) — pha trắng để text vẫn đọc được.
def _lighten_hex(hex_color: str, factor: float = 0.55) -> str:
    """Trộn màu với trắng theo factor (0..1) — 1 = trắng, 0 = giữ nguyên."""
    h = (hex_color or "").lstrip("#")
    if len(h) == 3:
        h = "".join(ch * 2 for ch in h)
    if len(h) != 6:
        return "#FFFFFF"
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"#{r:02X}{g:02X}{b:02X}"


def export_gantt_calendar_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    subtitle: str = "",
) -> str:
    """
    Xuất Gantt Calendar Excel-style: header 2-3 tầng (Month/Week/Day), rows
    aggregate theo group_by, cell tô màu theo phase category, marker Today.

    payload = compute_gantt_calendar(...) shape.
    """
    from analyzer.gantt_calendar import CATEGORY_COLORS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GanttCalendar"

    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    month_spans = payload.get("month_spans") or []
    week_spans = payload.get("week_spans") or []
    granularity = payload.get("granularity") or "week"
    today_col = payload.get("today_col")
    group_by = payload.get("group_by") or "module"

    # Số cột: 1 cột label + N cột timeline
    n_time_cols = len(columns)
    total_cols = 1 + n_time_cols
    last_letter = get_column_letter(total_cols)

    # ===== Row 1: Title =====
    ws.merge_cells(f"A1:{last_letter}1")
    tc = ws["A1"]
    tc.value = f"GANTT CALENDAR — Group: {group_by} · Granularity: {granularity}"
    tc.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ===== Row 2: Subtitle =====
    ws.merge_cells(f"A2:{last_letter}2")
    sc = ws["A2"]
    sc.value = subtitle or f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    sc.font = Font(name="Arial", size=10, italic=True, color="666666")
    sc.alignment = Alignment(horizontal="center")

    # ===== Header block =====
    # Layout header:
    #   granularity=day   → 3 tầng: Month / Week / Day     (rows 4, 5, 6)
    #   granularity=week  → 2 tầng: Month / Week           (rows 4, 5)
    #   granularity=month → 1 tầng: Month                  (row 4)
    header_start_row = 4
    if granularity == "day":
        n_header_rows = 3
    elif granularity == "week":
        n_header_rows = 2
    else:
        n_header_rows = 1
    header_end_row = header_start_row + n_header_rows - 1

    # Cột A: label "Row" (merge cả n_header_rows)
    label_cell_range = f"A{header_start_row}:A{header_end_row}"
    if header_start_row != header_end_row:
        ws.merge_cells(label_cell_range)
    lc = ws.cell(row=header_start_row, column=1)
    lc.value = "Module / Quy trình / Function"
    lc.font = HEADER_FONT
    lc.fill = HEADER_FILL
    lc.alignment = HEADER_ALIGN
    lc.border = THIN_BORDER

    # === Row header Month (row 4) — merge theo month_spans ===
    col_cursor = 2  # cột B là start timeline
    for span in month_spans:
        colspan = span["colspan"]
        start_col = col_cursor
        end_col = col_cursor + colspan - 1
        cell = ws.cell(row=header_start_row, column=start_col, value=span["label"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        if colspan > 1:
            ws.merge_cells(
                start_row=header_start_row,
                start_column=start_col,
                end_row=header_start_row,
                end_column=end_col,
            )
        col_cursor += colspan

    # === Row 2 (row 5) — Week header ===
    if n_header_rows >= 2:
        col_cursor = 2
        if granularity == "day":
            # week_spans dùng cho gr=day
            for wsp in week_spans:
                colspan = wsp["colspan"]
                start_col = col_cursor
                end_col = col_cursor + colspan - 1
                cell = ws.cell(row=header_start_row + 1, column=start_col, value=wsp["label"])
                cell.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
                if colspan > 1:
                    ws.merge_cells(
                        start_row=header_start_row + 1,
                        start_column=start_col,
                        end_row=header_start_row + 1,
                        end_column=end_col,
                    )
                col_cursor += colspan
        else:  # week
            # granularity=week: hàng 2 = "Week + start day", VD "W22 · 01-Jun"
            for i, c in enumerate(columns):
                col = 2 + i
                cell = ws.cell(row=header_start_row + 1, column=col)
                lbl_extra = c.get("week_date_label") or ""
                cell.value = f"{c['label']} · {lbl_extra}" if lbl_extra else c["label"]
                cell.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER

    # === Row 3 (row 6) — Day header (chỉ granularity=day) ===
    if n_header_rows == 3:
        for i, c in enumerate(columns):
            col = 2 + i
            cell = ws.cell(row=header_start_row + 2, column=col, value=c["label"])
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

    ws.row_dimensions[header_start_row].height = 22
    if n_header_rows >= 2:
        ws.row_dimensions[header_start_row + 1].height = 18
    if n_header_rows == 3:
        ws.row_dimensions[header_start_row + 2].height = 18

    # ===== Data rows =====
    data_start_row = header_end_row + 1
    ws.column_dimensions["A"].width = 40
    # Cột timeline: rộng cố định để cell "đủ vuông" cho bar bên trong
    time_col_width = 6 if granularity == "day" else 9 if granularity == "week" else 12
    for i in range(n_time_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = time_col_width

    # Fill nhạt cho cột today (nếu có) — làm nền cả cột từ header_end_row+1 xuống
    today_fill = PatternFill(start_color="FBCFE8", end_color="FBCFE8", fill_type="solid")

    for r_idx, row in enumerate(rows):
        excel_row = data_start_row + r_idx
        # Cột A: tên row + kèm thông tin phụ (func_count, active_phase, %)
        name = row.get("name") or ""
        extra_parts = []
        if row.get("func_count"):
            extra_parts.append(f"{row['func_count']} func")
        if row.get("overdue_count"):
            extra_parts.append(f"⚠ {row['overdue_count']} trễ")
        suffix = f" ({' · '.join(extra_parts)})" if extra_parts else ""
        cell = ws.cell(row=excel_row, column=1, value=f"{name}{suffix}")
        cell.font = Font(name="Arial", bold=bool(row.get("is_aggregate")), size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

        # Timeline cells — tô theo cell_categories (phase segment), không còn
        # 1 màu summary xám full-span.
        cells_flags = row.get("cells") or []
        cell_cats = row.get("cell_categories") or []
        span_start = row.get("span_start_col")
        span_end = row.get("span_end_col")

        for i, active in enumerate(cells_flags):
            col = 2 + i
            excel_cell = ws.cell(row=excel_row, column=col)
            excel_cell.border = THIN_BORDER
            # Today column: nền hồng nhẹ cho mọi cell (kể cả row không active)
            if today_col is not None and i == today_col and not active:
                excel_cell.fill = today_fill
            if not active:
                continue
            cat = (cell_cats[i] if i < len(cell_cats) and cell_cats[i] else None) \
                or row.get("category") or "phase1"
            if cat == "summary":
                cat = "phase1"
            color_hex = CATEGORY_COLORS.get(cat, CATEGORY_COLORS["phase1"])
            bar_argb = _hex_to_argb(_lighten_hex(color_hex, 0.35))
            text_argb = _hex_to_argb(color_hex).replace("00", "FF", 1)  # ARGB alpha=FF
            excel_cell.fill = PatternFill(
                start_color=bar_argb[2:], end_color=bar_argb[2:], fill_type="solid",
            )
            mid = None
            if span_start is not None and span_end is not None:
                mid = (span_start + span_end) // 2
            if mid == i:
                excel_cell.value = f"{row.get('pct', 0)}%"
                excel_cell.font = Font(name="Arial", bold=True, size=10, color=text_argb[2:])
                excel_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Đảm bảo cột today vẫn có fill hồng nếu span không cover today
        if today_col is not None and 0 <= today_col < n_time_cols:
            tcell = ws.cell(row=excel_row, column=2 + today_col)
            # Nếu cell chưa có fill (không active) → set today_fill
            if not cells_flags[today_col]:
                tcell.fill = today_fill

    # ===== Legend cuối sheet =====
    legend_row = data_start_row + len(rows) + 2
    legend_labels = payload.get("legend") or {}
    ws.cell(row=legend_row, column=1, value="LEGEND:").font = Font(name="Arial", bold=True, size=10)
    for i, (key, meta) in enumerate(legend_labels.items()):
        c = ws.cell(row=legend_row, column=2 + i * 2)
        c.value = meta.get("label", key)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill(
            start_color=_hex_to_argb(meta.get("color", "#94a3b8"))[2:],
            end_color=_hex_to_argb(meta.get("color", "#94a3b8"))[2:],
            fill_type="solid",
        )

    # Freeze pane ở cột đầu + header
    ws.freeze_panes = f"B{data_start_row}"

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(
        output_dir, f"Gantt_Calendar_{group_by}_{granularity}_{date.today().strftime('%Y%m%d')}.xlsx",
    )
    wb.save(filepath)
    wb.close()
    return filepath
