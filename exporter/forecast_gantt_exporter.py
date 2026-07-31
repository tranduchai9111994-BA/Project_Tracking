"""Xuất Excel nhẹ — Forecast Gantt UAT/Golive theo tháng."""
from __future__ import annotations

import os
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def export_forecast_gantt(
    result: dict[str, Any],
    upload_folder: str,
) -> str:
    """
    2 sheet: Tổng hợp (project × milestone tháng) + Theo tháng (UAT/Golive).
    """
    os.makedirs(upload_folder, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(upload_folder, f"Forecast_Gantt_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tong_hop"

    milestones = result.get("milestones") or []
    projects = result.get("projects") or []

    n = len(milestones)
    headers = (
        ["Dự án", "Slug"]
        + [m["label"] for m in milestones]
        + [f"{m['label']} (span)" for m in milestones]
        + [f"{m['label']} (nguồn)" for m in milestones]
        + [f"{m['label']} — Đánh giá lý do hợp lý" for m in milestones]
    )
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    assess_fills = {
        "ok": PatternFill("solid", fgColor="D1FAE5"),
        "warn": PatternFill("solid", fgColor="FEF3C7"),
        "risk": PatternFill("solid", fgColor="FEE2E2"),
    }

    for r_i, proj in enumerate(projects, 2):
        ws.cell(r_i, 1, proj.get("name") or "")
        ws.cell(r_i, 2, proj.get("slug") or "")
        ms = proj.get("milestones") or {}
        for j, m in enumerate(milestones):
            info = ms.get(m["id"]) or {}
            ws.cell(r_i, 3 + j, info.get("month") or "")
            span_s = info.get("span_start") or ""
            span_e = info.get("span_end") or ""
            span_txt = f"{span_s} → {span_e}" if (span_s or span_e) else ""
            ws.cell(r_i, 3 + n + j, span_txt)
            src = info.get("source") or ""
            src_label = {
                "open_max": "max End còn mở",
                "closed_max": "max End Closed",
                "no_date": "không có End",
            }.get(src, src)
            ws.cell(r_i, 3 + 2 * n + j, src_label)
            assess = info.get("assessment") or {}
            cell = ws.cell(r_i, 3 + 3 * n + j, assess.get("text") or "")
            fill = assess_fills.get(assess.get("level") or "")
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # Sheet 2 — UAT / Golive theo tháng
    ws2 = wb.create_sheet("UAT_Golive")
    ws2["A1"] = "Rule"
    ws2["B1"] = result.get("rule") or ""
    ws2["A3"] = "Tháng"
    ws2["B3"] = "UAT với KH (dự án)"
    ws2["C3"] = "Golive với KH (dự án)"
    for c in range(1, 4):
        ws2.cell(3, c).fill = header_fill
        ws2.cell(3, c).font = header_font

    summary = result.get("summary") or {}
    uat = summary.get("uat_by_month") or {}
    golive = summary.get("golive_by_month") or {}
    all_m = sorted(set(uat.keys()) | set(golive.keys()))
    for i, mk in enumerate(all_m, 4):
        ws2.cell(i, 1, mk)
        ws2.cell(i, 2, ", ".join(uat.get(mk) or []))
        ws2.cell(i, 3, ", ".join(golive.get(mk) or []))

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 40

    wb.save(path)
    wb.close()
    return path
