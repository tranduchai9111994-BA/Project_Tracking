"""
Xuất Excel Rlog coded tuần này + kế hoạch tuần tới.

Workbook 3 sheet:
  1. Summary — count + week labels + definition
  2. Coded — danh sách Rlog coded tuần này
  3. Ke_hoach — danh sách kế hoạch code tuần tới

Reuse payload từ analyzer.rlog_weekly.compute_rlog_weekly / metrics.rlog_weekly.
"""
from __future__ import annotations

import os
from datetime import date
from typing import Any

import openpyxl

from exporter.excel_exporter import _write_sheet


def _pic_str(item: dict[str, Any]) -> str:
    pics = item.get("pic") or []
    if isinstance(pics, list):
        return ", ".join(str(p) for p in pics if p)
    return str(pics) if pics else ""


def export_rlog_weekly_report(
    payload: dict[str, Any],
    output_dir: str = "uploads",
    subtitle: str = "",
) -> str:
    """
    Tạo file Excel Rlog weekly từ payload compute_rlog_weekly().

    Returns:
        Filepath .xlsx đã tạo.
    """
    coded = (payload.get("rlog_coded_this_week") or {})
    plan = (payload.get("rlog_plan_next_week") or {})
    week = payload.get("week") or {}
    next_w = payload.get("next_week") or {}
    coded_items = list(coded.get("items") or [])
    plan_items = list(plan.get("items") or [])

    week_lbl = week.get("iso_week_label") or ""
    next_lbl = next_w.get("iso_week_label") or ""
    base_sub = subtitle or f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"

    wb = openpyxl.Workbook()

    # === Sheet 1: Summary ===
    ws1 = wb.active
    ws1.title = "Summary"
    summary_rows = [
        ["Coded tuần này", coded.get("count", 0)],
        ["  Nhãn tuần", week_lbl],
        ["  Từ ngày", week.get("monday_iso") or ""],
        ["  Đến ngày", week.get("sunday_iso") or ""],
        ["", ""],
        ["Kế hoạch tuần tới", plan.get("count", 0)],
        ["  Nhãn tuần", next_lbl],
        ["  Từ ngày", next_w.get("monday_iso") or ""],
        ["  Đến ngày", next_w.get("sunday_iso") or ""],
        ["", ""],
        ["Scope", payload.get("rlog_scope") or ""],
        ["Cột Rlog phát hiện", "Có" if payload.get("rlog_column_detected") else "Không"],
        ["Today", week.get("today_iso") or ""],
        ["", ""],
        ["Định nghĩa", payload.get("definition") or ""],
    ]
    _write_sheet(
        ws1,
        title="RLOG — TỔNG QUAN TUẦN",
        subtitle=base_sub,
        columns=[("Chỉ số", 28), ("Giá trị", 70)],
        data_rows=summary_rows,
    )

    # === Sheet 2: Coded tuần này ===
    ws2 = wb.create_sheet(f"Coded_{week_lbl}" if week_lbl else "Coded")
    coded_cols = [
        ("STT", 6),
        ("Mã CN", 14),
        ("Tên chức năng", 36),
        ("RlogID", 12),
        ("Module", 10),
        ("PIC", 20),
        ("Phase", 14),
        ("Status", 12),
        ("Closed", 12),
        ("Start", 12),
        ("End", 12),
    ]
    coded_rows = [
        [
            idx + 1,
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("rlog_id", ""),
            it.get("module", ""),
            _pic_str(it),
            it.get("phase", ""),
            it.get("status", ""),
            it.get("closed_date") or it.get("end_date") or "",
            it.get("start_date", ""),
            it.get("end_date", ""),
        ]
        for idx, it in enumerate(coded_items)
    ]
    _write_sheet(
        ws2,
        title=f"RLOG CODED TUẦN NÀY ({week_lbl})" if week_lbl else "RLOG CODED TUẦN NÀY",
        subtitle=f"Tổng: {len(coded_items)} | {base_sub}",
        columns=coded_cols,
        data_rows=coded_rows,
    )

    # === Sheet 3: Kế hoạch tuần tới ===
    ws3 = wb.create_sheet(f"Ke_hoach_{next_lbl}" if next_lbl else "Ke_hoach")
    plan_cols = [
        ("STT", 6),
        ("Mã CN", 14),
        ("Tên chức năng", 36),
        ("RlogID", 12),
        ("Module", 10),
        ("PIC", 20),
        ("Phase", 14),
        ("Status", 12),
        ("Deadline", 12),
        ("Start", 12),
        ("End", 12),
    ]
    plan_rows = [
        [
            idx + 1,
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("rlog_id", ""),
            it.get("module", ""),
            _pic_str(it),
            it.get("phase", ""),
            it.get("status", ""),
            it.get("end_date", ""),
            it.get("start_date", ""),
            it.get("end_date", ""),
        ]
        for idx, it in enumerate(plan_items)
    ]
    _write_sheet(
        ws3,
        title=f"KẾ HOẠCH CODE TUẦN TỚI ({next_lbl})" if next_lbl else "KẾ HOẠCH CODE TUẦN TỚI",
        subtitle=f"Tổng: {len(plan_items)} | {base_sub}",
        columns=plan_cols,
        data_rows=plan_rows,
    )

    os.makedirs(output_dir, exist_ok=True)
    filename = f"Rlog_Weekly_{week_lbl or date.today().strftime('%Y%m%d')}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    return filepath
