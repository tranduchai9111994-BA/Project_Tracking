"""
Xuất báo cáo tuần GAP hoàn thành ra Excel 2-sheet:
  Sheet 1 — "Tổng quan tuần": pivot Module × Phase
  Sheet 2 — "Chi tiết":       danh sách function/phase đầy đủ
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Styles ─────────────────────────────────────────────────────────────────
_TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1E3A5F")
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

_SUBHDR_FONT = Font(name="Arial", size=10, italic=True, color="444444")
_SUBHDR_ALIGN = Alignment(horizontal="center")

_HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HDR_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_BODY_FONT = Font(name="Arial", size=9)
_BODY_ALIGN = Alignment(vertical="center", wrap_text=True)

_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
_STRIPE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
_ORANGE_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")


def _auto_width(ws, min_w: int = 8, max_w: int = 50) -> None:
    for col_cells in ws.columns:
        best = min_w
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > best:
                best = length
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(best + 2, max_w)


def _write_title_rows(ws, title: str, subtitle: str, n_cols: int) -> None:
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = title
    c.font = _TITLE_FONT
    c.alignment = _TITLE_ALIGN
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]
    c2.value = subtitle
    c2.font = _SUBHDR_FONT
    c2.alignment = _SUBHDR_ALIGN
    ws.row_dimensions[2].height = 18


# ── Sheet 1: Tổng quan ─────────────────────────────────────────────────────
def _build_summary_sheet(ws, result: dict[str, Any]) -> None:
    week_label = result.get("week_label", "")
    items: list[dict] = result.get("items", [])
    summary = result.get("summary", {})

    # Thu thập danh sách phase và module
    phases = list(dict.fromkeys(it["phase"] for it in items))
    modules = list(dict.fromkeys(it["module"] for it in items))

    n_cols = 2 + len(phases) + 1  # STT + Module + phases + Total

    _write_title_rows(
        ws,
        f"BÁO CÁO TIẾN ĐỘ TUẦN — {week_label}",
        f"Các GAP/chức năng hoàn thành trong {week_label}",
        n_cols,
    )

    # Header row 3
    hdr_row = 3
    headers = ["STT", "Module"] + phases + ["Tổng"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _HDR_ALIGN
        c.border = _THIN
    ws.row_dimensions[hdr_row].height = 22

    # Pivot: module → phase → count
    pivot: dict[str, dict[str, int]] = {}
    for it in items:
        m = it["module"]
        p = it["phase"]
        if m not in pivot:
            pivot[m] = {}
        pivot[m][p] = pivot[m].get(p, 0) + 1

    # Data rows
    row_idx = hdr_row + 1
    for stt, mod in enumerate(modules, 1):
        phase_counts = pivot.get(mod, {})
        row_total = sum(phase_counts.get(p, 0) for p in phases)

        cells_data = [stt, mod] + [phase_counts.get(p, 0) for p in phases] + [row_total]
        for ci, val in enumerate(cells_data, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = _BODY_FONT
            c.alignment = _BODY_ALIGN
            c.border = _THIN
            # Tô xanh nếu count > 0 (chỉ cột phase)
            if ci > 2 and ci < len(cells_data) and isinstance(val, int) and val > 0:
                c.fill = _GREEN_FILL
        row_idx += 1

    # Dòng tổng cuối
    total_by_phase = summary.get("by_phase", {})
    grand_total = summary.get("total", 0)
    total_row = ["", "TỔNG"] + [total_by_phase.get(p, 0) for p in phases] + [grand_total]
    for ci, val in enumerate(total_row, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.font = Font(name="Arial", bold=True, size=9)
        c.fill = _TOTAL_FILL
        c.alignment = _HDR_ALIGN
        c.border = _THIN

    ws.freeze_panes = "A3"
    _auto_width(ws)


# ── Sheet 2: Chi tiết ──────────────────────────────────────────────────────
def _build_detail_sheet(ws, result: dict[str, Any]) -> None:
    week_label = result.get("week_label", "")
    items: list[dict] = result.get("items", [])

    columns = [
        "STT", "Rlog ID", "Mã CN", "Tên chức năng",
        "Module", "Quy trình", "FIT/GAP",
        "Phase", "Start", "End", "Trạng thái", "PIC", "Ghi chú",
    ]
    n_cols = len(columns)

    _write_title_rows(
        ws,
        f"CHI TIẾT — {week_label}",
        f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}",
        n_cols,
    )

    # Header row 3
    hdr_row = 3
    for ci, h in enumerate(columns, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _HDR_ALIGN
        c.border = _THIN
    ws.row_dimensions[hdr_row].height = 22

    today = date.today()

    for idx, it in enumerate(items):
        row_idx = hdr_row + 1 + idx
        stt = idx + 1
        status = it.get("status", "")
        end_str = it.get("end", "")

        # Xác định fill row
        end_d = None
        if end_str:
            try:
                from datetime import datetime
                end_d = datetime.strptime(end_str, "%Y-%m-%d").date()
            except ValueError:
                pass

        is_inprogress = status in {"In-progress", "In progress", "Inprogress"}
        is_overdue = end_d is not None and end_d < today

        if is_inprogress:
            row_fill = _YELLOW_FILL
        elif is_overdue:
            row_fill = _ORANGE_FILL
        elif idx % 2 == 1:
            row_fill = _STRIPE_FILL
        else:
            row_fill = None

        values = [
            stt,
            it.get("rlog_id", ""),
            it.get("ma_cn", ""),
            it.get("ten_cn", ""),
            it.get("module", ""),
            it.get("quy_trinh", ""),
            it.get("fitgap", ""),
            it.get("phase", ""),
            it.get("start", ""),
            end_str,
            status,
            it.get("pic", ""),
            "",  # Ghi chú — để trống
        ]

        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = _BODY_FONT
            c.alignment = _BODY_ALIGN
            c.border = _THIN
            if row_fill:
                c.fill = row_fill

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(n_cols)}{hdr_row}"
    _auto_width(ws)


# ── Public API ─────────────────────────────────────────────────────────────
def export_weekly_gap_excel(result: dict[str, Any]) -> bytes:
    """
    Tạo file Excel 2-sheet từ kết quả compute_weekly_gap().
    Trả về bytes của file .xlsx.
    """
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Tổng quan tuần"
    _build_summary_sheet(ws1, result)

    ws2 = wb.create_sheet("Chi tiết")
    _build_detail_sheet(ws2, result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
