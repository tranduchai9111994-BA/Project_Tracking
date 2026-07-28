"""Pytest fixtures — dùng chung cho toàn bộ test suite."""
import os
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

# Thêm root vào sys.path để import parser/analyzer/exporter
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


TODAY = date(2026, 7, 28)  # Cố định today để test reproducible


@pytest.fixture
def today():
    return TODAY


@pytest.fixture
def sample_xlsx_path(tmp_path):
    """
    Tạo 1 file Function List mẫu bao trùm nhiều case:
    - Row 1: đã Closed hết → không overdue, không unassigned
    - Row 2: overdue ở UAT (In-progress, End < today) + multi-PIC
    - Row 3: unassigned + stalled (Analysis Closed, Dev Open không PIC)
    - Row 4: long duration đang chạy (Dev In-progress, Start 30d trước, no End)
    - Row 5: Should-have Complexity High không overdue
    - Row 6: overdue nặng + PIC case bị lộn (SONHN6 vs SonHN6)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"

    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "Quy trình",
        "Priority", "Complexity", "FIT/GAP", "Giai đoạn", "Risk/Blocker",
        "Analysis - Start", "Analysis - End", "Analysis - Status",
        "Analysis - Estimate MH", "Analysis - PIC",
        "Dev - Start", "Dev - End", "Dev - Status",
        "Dev - Estimate MH", "Dev - PIC",
        "UAT - From", "UAT - To", "UAT - Status",
        "UAT - Estimate MH", "UAT - PIC FPT", "UAT - PIC MPHG",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)

    T = TODAY  # alias

    data = [
        # 1: Closed hết
        [1, "TMS.FR.01", "Chấm công app", "TMS", "TMS.BP.01 - Chấm công",
         "Must-have", "Medium", "FIT", "1", "",
         T - timedelta(days=60), T - timedelta(days=55), "Closed", 8, "SonHN6",
         T - timedelta(days=50), T - timedelta(days=40), "Closed", 16, "PhatTPT3",
         T - timedelta(days=30), T - timedelta(days=25), "Closed", 12, "BaoLQ31", "Chị Liên"],
        # 2: overdue ở UAT + multi-PIC
        [2, "TMS.FR.02", "Báo cáo tháng", "TMS", "TMS.BP.01 - Chấm công",
         "Must-have", "High", "GAP", "1", "Client thay đổi req",
         T - timedelta(days=40), T - timedelta(days=35), "Closed", 12, "SonHN6",
         T - timedelta(days=30), T - timedelta(days=15), "Closed", 24, "PhatTPT3",
         T - timedelta(days=10), T - timedelta(days=5), "In-progress", 24, "BaoLQ31, NhiVN", "Anh Cường"],
        # 3: Stalled Dev + Unassigned Dev
        [3, "HR.FR.05", "Quản lý NS", "HR", "HR.BP.02 - Onboarding",
         "Should-have", "Low", "FIT", "2", "",
         T - timedelta(days=20), T - timedelta(days=15), "Closed", 4, "CuongNM129",
         None, None, "Open", None, "",
         None, None, None, None, "", ""],
        # 4: Long duration đang chạy Dev
        [4, "PR.FR.03", "Đánh giá 360", "PR", "PR.BP.01 - Performance",
         "Could-have", "High", "Customization", "2", "Complexity cao",
         T - timedelta(days=100), T - timedelta(days=90), "Closed", 40, "TungTT83",
         T - timedelta(days=30), None, "In-progress", 60, "NhuNHT3+ HaiTD16",
         None, None, None, None, "", ""],
        # 5: Should-have + High complexity, không overdue
        [5, "SYS.FR.01", "Cấu hình phân quyền", "SYS", "SYS.BP.01 - Admin",
         "Should-have", "High", "FIT", "1", "",
         T - timedelta(days=5), T + timedelta(days=5), "Assigned", 8, "HaiTD16",
         None, None, None, None, "",
         None, None, None, None, "", ""],
        # 6: overdue nặng + PIC bị lộn case (SONHN6 vs SonHN6)
        [6, "ESS.FR.10", "Xin phép nghỉ", "ESS", "ESS.BP.03 - Leave",
         "Must-have", "Medium", "FIT", "2", "Đang chờ TL confirm",
         T - timedelta(days=45), T - timedelta(days=20), "In-progress", 8, "SONHN6",  # ALL CAPS
         None, None, None, None, "",
         None, None, None, None, "", ""],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, val)

    path = tmp_path / "sample.xlsx"
    wb.save(str(path))
    wb.close()
    return str(path)


@pytest.fixture
def parsed_data(sample_xlsx_path):
    """ParsedData từ file mẫu."""
    from parser.excel_parser import FunctionListParser
    return FunctionListParser().parse(sample_xlsx_path)


@pytest.fixture
def metrics(parsed_data, today):
    """Metrics đầy đủ."""
    from analyzer.dashboard_engine import DashboardEngine
    return DashboardEngine(today=today, long_duration_threshold=3).compute_all(parsed_data)


@pytest.fixture
def flask_client(tmp_path, sample_xlsx_path):
    """Flask test client với upload folder tạm — mỗi test có project folder riêng biệt."""
    from app import app
    import app as app_module
    from analyzer.project_manager import ProjectManager

    # Redirect toàn bộ storage về tmp_path để không đụng data thật
    app.config["UPLOAD_FOLDER"] = str(tmp_path)
    app.config["PROJECTS_FOLDER"] = str(tmp_path / "projects")
    os.makedirs(app.config["PROJECTS_FOLDER"], exist_ok=True)

    # Reset toàn bộ state và tạo ProjectManager mới trên tmp
    app_module._state.clear()
    app_module._project_mgr = ProjectManager(app.config["PROJECTS_FOLDER"])
    app_module._project_mgr.get_or_create_default()

    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client
