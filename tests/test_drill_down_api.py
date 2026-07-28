"""HTTP integration tests cho drill-down API endpoints."""
import io
import json
import openpyxl


def _upload(client, sample_xlsx_path):
    with open(sample_xlsx_path, "rb") as f:
        return client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "sample.xlsx")},
            content_type="multipart/form-data",
        )


# ==========================================================================
# GET /api/projects/<slug>/drill-down
# ==========================================================================

def test_drill_down_requires_upload(flask_client):
    r = flask_client.get("/api/projects/default/drill-down?chart=priority&priority=Must-have")
    assert r.status_code in (400, 404)  # chưa upload
    assert "error" in r.get_json()


def test_drill_down_priority(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/drill-down?chart=priority&priority=Must-have")
    assert r.status_code == 200
    data = r.get_json()
    assert data["success"] is True
    assert data["chart"] == "priority"
    assert data["total"] == 3
    assert len(data["items"]) == 3
    assert all(i["priority"] == "Must-have" for i in data["items"])


def test_drill_down_phase_matrix(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/drill-down?chart=phase_matrix&module=TMS&phase=Analysis"
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["total"] == 2
    ma_cns = {i["ma_cn"] for i in data["items"]}
    assert ma_cns == {"TMS.FR.01", "TMS.FR.02"}


def test_drill_down_pic_workload_overdue(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/drill-down?chart=pic_workload&pic=SonHN6&status=overdue"
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["total"] == 1
    assert data["items"][0]["ma_cn"] == "ESS.FR.10"
    assert data["items"][0]["is_overdue"] is True


def test_drill_down_invalid_chart_returns_400(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/drill-down?chart=foobar")
    assert r.status_code == 400
    assert "error" in r.get_json()


def test_drill_down_legacy_alias(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/drill-down?chart=priority&priority=Must-have")
    assert r.status_code == 200
    assert r.get_json()["total"] == 3


def test_drill_down_title_returned(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/drill-down?chart=fit_gap&fit_gap=GAP")
    data = r.get_json()
    assert "GAP" in data["title"]


# ==========================================================================
# POST /api/projects/<slug>/drill-down/export
# ==========================================================================

def test_drill_down_export_returns_xlsx(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/drill-down/export",
        data=json.dumps({"chart": "priority", "filters": {"priority": "Must-have"}}),
        content_type="application/json",
    )
    assert r.status_code == 200
    # Content type xlsx
    ctype = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ctype or "octet-stream" in ctype

    # Verify file is real xlsx và có đúng số row
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    # Row 4 là header, row 5+ là data
    header = [c.value for c in ws[4]]
    assert "Mã CN" in header
    # 3 function Must-have — data rows có STT là số nguyên
    data_rows = [
        row for row in ws.iter_rows(min_row=5, values_only=True)
        if isinstance(row[0], int)
    ]
    assert len(data_rows) == 3
    wb.close()


def test_drill_down_export_invalid_chart(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/drill-down/export",
        data=json.dumps({"chart": "bogus", "filters": {}}),
        content_type="application/json",
    )
    assert r.status_code == 400


def test_drill_down_export_no_upload(flask_client):
    r = flask_client.post(
        "/api/projects/default/drill-down/export",
        data=json.dumps({"chart": "priority", "filters": {"priority": "Must-have"}}),
        content_type="application/json",
    )
    assert r.status_code in (400, 404)
