"""Tests cho exporter.ba_task_exporter (B2/B5)."""
import openpyxl

from analyzer import ba_task_store as store
from exporter.ba_task_exporter import export_ba_tasks, export_ba_tasks_weekly


def _sample_tasks():
    return [
        {"id": "1", "title": "A", "type": "task", "status": "open", "priority": "high",
         "assignee": "Nhi", "due_date": "2020-01-01", "alert_level": "overdue", "tags": []},
        {"id": "2", "title": "Họp KH", "type": "meeting", "status": "open", "priority": "medium",
         "week_iso": "2026-W32", "due_date": "2026-08-05",
         "meeting_info": {"meeting_date": "2026-08-05", "attendees": ["A", "B"], "agenda": "Review", "mom_notes": "OK"},
         "tags": []},
    ]


def test_export_all_creates_file(tmp_path):
    filepath = export_ba_tasks(_sample_tasks(), str(tmp_path), project_name="Test Project")
    wb = openpyxl.load_workbook(filepath)
    assert wb.sheetnames == ["BA_Tasks"]
    ws = wb["BA_Tasks"]
    assert ws["A1"].value == "QUẢN LÝ ĐẦU VIỆC BA"


def test_export_weekly_header_has_date_range(tmp_path):
    filepath = export_ba_tasks_weekly(_sample_tasks(), "2026-W32", str(tmp_path), project_name="MPHG")
    wb = openpyxl.load_workbook(filepath)
    assert set(wb.sheetnames) == {"Dau_viec_tuan", "Cuoc_hop", "San_pham_ban_giao", "No_KH"}
    header = wb["Dau_viec_tuan"]["A2"].value
    assert "Dự án: MPHG" in header
    assert "Tuần 32" in header
    assert "–" in header  # date range separator


def test_export_weekly_meeting_sheet_has_attendees(tmp_path):
    filepath = export_ba_tasks_weekly(_sample_tasks(), "2026-W32", str(tmp_path), project_name="MPHG")
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Cuoc_hop"]
    row_values = [c.value for c in ws[5]]  # header row 4, data row 5
    assert "Họp KH" in row_values


def test_export_weekly_debt_highlight(tmp_path):
    tasks = [{
        "id": "3", "title": "Nợ", "type": "customer_debt", "status": "open",
        "debt_info": {"description": "KH chưa gửi", "requested_date": "2000-01-01"},
        "tags": [],
    }]
    filepath = export_ba_tasks_weekly(tasks, "2026-W32", str(tmp_path), project_name="MPHG")
    wb = openpyxl.load_workbook(filepath)
    ws = wb["No_KH"]
    fill = ws.cell(row=5, column=1).fill
    assert fill.fgColor.rgb in ("00FEE2E2", "FFFEE2E2")
