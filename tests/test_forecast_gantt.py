"""Tests Forecast Gantt — tháng UAT/Golive + rule open_max / closed_max + span + assessment."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from analyzer.forecast_gantt import (
    FORECAST_RULE_VI,
    assess_milestone,
    compute_forecast_gantt,
    compute_milestone_for_data,
    compute_project_forecast,
)
from analyzer.project_manager import ProjectManager
from parser.excel_parser import FunctionListParser


TODAY = date(2026, 7, 31)


def _write_fl(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


HEADERS = [
    "STT", "Mã CN", "Tên chức năng", "Module",
    "Analysis - Start", "Analysis - End", "Analysis - Status", "Analysis - PIC",
    "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
    "Config UAT - Start", "Config UAT - End", "Config UAT - Status", "Config UAT - PIC",
    "UAT - Start", "UAT - End", "UAT - Status", "UAT - PIC",
    "Golive - Start", "Golive - End", "Golive - Status", "Golive - PIC",
]


def _row(
    stt, code, name, module,
    a_end=None, a_st=None,
    d_end=None, d_st=None,
    c_end=None, c_st=None,
    u_end=None, u_st=None,
    g_end=None, g_st=None,
    a_start=None, d_start=None, c_start=None, u_start=None, g_start=None,
):
    return [
        stt, code, name, module,
        a_start, a_end, a_st, "A",
        d_start, d_end, d_st, "D",
        c_start, c_end, c_st, "C",
        u_start, u_end, u_st, "U",
        g_start, g_end, g_st, "G",
    ]


@pytest.fixture
def sample_data(tmp_path):
    path = tmp_path / "fl.xlsx"
    _write_fl(
        path,
        HEADERS,
        [
            # Analysis all Closed → closed_max May
            _row(1, "F.01", "F1", "TMS",
                 a_start=date(2026, 4, 1), a_end=date(2026, 5, 10), a_st="Closed",
                 d_start=date(2026, 5, 15), d_end=date(2026, 6, 20), d_st="In-progress",
                 c_start=date(2026, 6, 1), c_end=date(2026, 7, 5), c_st="Open",
                 u_start=date(2026, 7, 1), u_end=date(2026, 8, 15), u_st="Open",
                 g_start=date(2026, 8, 1), g_end=date(2026, 9, 30), g_st="Open"),
            _row(2, "F.02", "F2", "TMS",
                 a_start=date(2026, 4, 10), a_end=date(2026, 5, 25), a_st="Closed",
                 d_start=date(2026, 5, 20), d_end=date(2026, 6, 10), d_st="Closed",
                 c_start=date(2026, 6, 15), c_end=date(2026, 7, 20), c_st="In-progress",
                 u_start=date(2026, 7, 10), u_end=date(2026, 8, 28), u_st="Assigned",
                 g_start=date(2026, 9, 1), g_end=date(2026, 10, 5), g_st="Open"),
        ],
    )
    return FunctionListParser().parse(str(path))


class TestMilestoneRule:
    def test_open_max_preferred(self, sample_data):
        uat = compute_milestone_for_data(sample_data, ("UAT",), today=TODAY)
        assert uat["source"] == "open_max"
        assert uat["month"] == "2026-08"
        assert uat["date"] == "2026-08-28"  # max of open ends
        assert uat["open"] == 2

    def test_closed_max_when_all_closed(self, sample_data):
        analysis = compute_milestone_for_data(sample_data, ("Phân tích",), today=TODAY)
        assert analysis["source"] == "closed_max"
        assert analysis["month"] == "2026-05"
        assert analysis["date"] == "2026-05-25"
        assert analysis["open"] == 0

    def test_project_forecast_keys(self, sample_data):
        fc = compute_project_forecast(sample_data, today=TODAY)
        assert set(fc) >= {"analysis", "dev", "config", "uat", "golive"}
        assert fc["golive"]["month"] == "2026-10"
        assert fc["dev"]["source"] == "open_max"  # 1 still open
        assert fc["dev"]["month"] == "2026-06"

    def test_span_from_start_end(self, sample_data):
        uat = compute_milestone_for_data(sample_data, ("UAT",), today=TODAY)
        assert uat["span_start"] == "2026-07"  # min start Jul 1 / Jul 10
        assert uat["span_end"] == "2026-08"    # max end Aug 28
        analysis = compute_milestone_for_data(sample_data, ("Phân tích",), today=TODAY)
        assert analysis["span_start"] == "2026-04"
        assert analysis["span_end"] == "2026-05"


class TestAssessment:
    def test_ok_open_in_range(self, sample_data):
        uat = compute_milestone_for_data(sample_data, ("UAT",), today=TODAY)
        assert uat["assessment"]["level"] == "ok"
        assert "hợp lý" in uat["assessment"]["text"]

    def test_ok_all_closed(self, sample_data):
        analysis = compute_milestone_for_data(sample_data, ("Phân tích",), today=TODAY)
        assert analysis["assessment"]["level"] == "ok"
        assert "Closed" in analysis["assessment"]["text"]

    def test_missing_end_risk(self):
        info = {
            "total": 3, "open": 2, "closed": 1, "no_end": 2,
            "pct_closed": 33.3, "source": "no_date", "month": None,
        }
        a = assess_milestone(info, today=TODAY)
        assert a["level"] == "risk"
        assert "Thiếu End" in a["text"]

    def test_high_closed_far_month_warn(self):
        info = {
            "total": 10, "open": 1, "closed": 9, "no_end": 0,
            "pct_closed": 90.0, "source": "open_max", "month": "2026-12",
        }
        a = assess_milestone(info, today=TODAY)
        assert a["level"] == "warn"
        assert "Closed cao" in a["text"]

    def test_past_month_still_open_risk(self):
        info = {
            "total": 4, "open": 2, "closed": 2, "no_end": 0,
            "pct_closed": 50.0, "source": "open_max", "month": "2026-05",
        }
        a = assess_milestone(info, today=TODAY)
        assert a["level"] == "risk"
        assert "đã qua" in a["text"]


class TestMultiProject:
    def test_two_projects(self, tmp_path):
        root = tmp_path / "projects"
        root.mkdir()
        mgr = ProjectManager(str(root))
        a = mgr.create_project("Alpha")
        b = mgr.create_project("Beta")

        _write_fl(
            Path(mgr.get_project_folder(a.slug)) / "current.xlsx",
            HEADERS,
            [_row(1, "A.01", "A1", "M",
                  a_end=date(2026, 4, 1), a_st="Closed",
                  u_end=date(2026, 8, 1), u_st="Open",
                  g_end=date(2026, 9, 1), g_st="Open")],
        )
        _write_fl(
            Path(mgr.get_project_folder(b.slug)) / "current.xlsx",
            HEADERS,
            [_row(1, "B.01", "B1", "M",
                  a_end=date(2026, 5, 1), a_st="Closed",
                  u_end=date(2026, 7, 15), u_st="Closed",
                  g_end=date(2026, 11, 20), g_st="Open")],
        )

        cache = {}

        def loader(slug):
            if slug in cache:
                return cache[slug]
            p = Path(mgr.get_project_folder(slug)) / "current.xlsx"
            if not p.is_file():
                return None
            data = FunctionListParser().parse(str(p))
            cache[slug] = {"data": data}
            return cache[slug]

        result = compute_forecast_gantt(
            mgr, loader, slugs=[a.slug, b.slug], today=TODAY,
        )
        assert result["rule"] == FORECAST_RULE_VI
        assert result["summary"]["project_count"] == 2
        assert "2026-08" in result["summary"]["uat_by_month"]
        # Beta UAT all closed → closed_max July
        beta = next(p for p in result["projects"] if p["slug"] == b.slug)
        assert beta["milestones"]["uat"]["month"] == "2026-07"
        assert beta["milestones"]["uat"]["source"] == "closed_max"
        assert result["milestone_aggregate"]["golive"]["month"] == "2026-11"
        # Assessment present
        assert "assessment" in beta["milestones"]["uat"]
        assert beta["milestones"]["uat"]["assessment"]["level"] in ("ok", "warn", "risk")
