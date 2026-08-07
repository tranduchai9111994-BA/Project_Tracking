"""
Regression: section «Chưa PIC» có filter Status cục bộ.

- HTML section-unassigned có select#unassignedStatusFilter với onchange
  gọi ``onUnassignedStatusChange``.
- FE renderUnassignedSection:
    * populate options từ unique statuses trong items (preserve chọn cũ),
    * áp filter status trước khi phân trang,
    * hiển thị count "0 task chưa có PIC khi lọc không match".
- Nút «Chi tiết» đi qua ``openUnassignedDrill`` để truyền status vào drill.
- Export FL re-import qua ``_unassignedExportLocalParams`` → param
  ``l_status`` để backend lọc theo status.
- Backend /export-fl-reimport?kinds=unassigned&l_status=Open phải lọc
  Function List xuất ra chỉ giữ dòng thiếu PIC ở phase Status = Open.
"""
from datetime import date, timedelta
from pathlib import Path

import io

import openpyxl
import pytest

import app as app_module
from parser.excel_parser import FunctionRow, PhaseData, PhaseGroup, ParsedData

ROOT = Path(__file__).resolve().parents[1]
DASHBOARD_JS = (ROOT / "static" / "js" / "dashboard.js").read_text(encoding="utf-8")
INDEX_HTML = (ROOT / "templates" / "index.html").read_text(encoding="utf-8")


def test_html_has_status_filter_and_hooks():
    # Trong section-unassigned phải có select#unassignedStatusFilter
    sec = INDEX_HTML.split('id="section-unassigned"')[1].split("</section>")[0]
    assert 'id="unassignedStatusFilter"' in sec
    assert "onchange=\"onUnassignedStatusChange()\"" in sec
    # Nút Chi tiết đi qua openUnassignedDrill (không hard-code {})
    assert "openUnassignedDrill()" in sec
    assert "openDrillDown('unassigned', {})" not in sec
    # Nút export nhận extraParams = _unassignedExportLocalParams
    assert "_unassignedExportLocalParams" in sec


def test_dashboard_js_has_populate_and_filter_and_helpers():
    # Helper populate options + preserve selection
    assert "function _unassignedPopulateStatusFilter(" in DASHBOARD_JS
    pop_body = DASHBOARD_JS.split("function _unassignedPopulateStatusFilter(")[1].split(
        "function onUnassignedStatusChange("
    )[0]
    assert "Tất cả status" in pop_body
    # Reset khi status đang chọn không còn trong dataset mới
    assert 'sel.value = ""' in pop_body

    # Onchange handler reset trang 1 + re-render
    assert "function onUnassignedStatusChange(" in DASHBOARD_JS
    onc = DASHBOARD_JS.split("function onUnassignedStatusChange(")[1].split(
        "window.onUnassignedStatusChange"
    )[0]
    assert "pageState.unassigned" in onc
    assert "renderUnassignedSection()" in onc

    # Filter apply trong renderUnassignedSection
    rus = DASHBOARD_JS.split("function renderUnassignedSection(")[1].split(
        "// ========================================================================"
    )[0]
    assert "_unassignedPopulateStatusFilter(allItems)" in rus
    assert "unassignedStatusFilter" in rus
    assert "allItems.filter(" in rus
    # Count text phải phản ánh filter đang áp — hiển thị "status=<val>" dạng
    # nội suy template literal (không hardcode chuỗi thô "lọc status=").
    assert "status=${statusF}" in rus


def test_dashboard_js_export_and_drill_helpers():
    # Export local params
    assert "function _unassignedExportLocalParams(" in DASHBOARD_JS
    exp = DASHBOARD_JS.split("function _unassignedExportLocalParams(")[1].split(
        "window._unassignedExportLocalParams"
    )[0]
    assert 'params.set("l_status"' in exp

    # Drill-down truyền status
    assert "function openUnassignedDrill(" in DASHBOARD_JS
    drill = DASHBOARD_JS.split("function openUnassignedDrill(")[1].split(
        "window.openUnassignedDrill"
    )[0]
    assert "unassignedStatusFilter" in drill
    assert 'openDrillDown("unassigned"' in drill


# ── Backend: /export-fl-reimport?kinds=unassigned tôn trọng l_status ──────

def _unassigned_data() -> ParsedData:
    """2 unassigned: U1 Status=Open, U2 Status=Assigned; 1 có PIC (bỏ qua)."""
    today = date.today()
    past = today - timedelta(days=5)
    rows = [
        FunctionRow(
            row_num=2,
            meta={"ma_cn": "U1", "ten_cn": "A", "module": "HR"},
            phases={
                "Analysis": PhaseData(status="Closed", end_date=past - timedelta(days=10), pics=["P1"]),
                "Dev": PhaseData(status="Open", end_date=past, pics=[]),
            },
        ),
        FunctionRow(
            row_num=3,
            meta={"ma_cn": "U2", "ten_cn": "B", "module": "HR"},
            phases={
                "Analysis": PhaseData(status="Closed", end_date=past - timedelta(days=10), pics=["P1"]),
                "Dev": PhaseData(status="Assigned", end_date=past, pics=[]),
            },
        ),
        FunctionRow(
            row_num=4,
            meta={"ma_cn": "OK", "ten_cn": "C", "module": "HR"},
            phases={
                "Analysis": PhaseData(status="Closed", end_date=past - timedelta(days=10), pics=["P1"]),
                "Dev": PhaseData(status="Open", end_date=past, pics=["P2"]),
            },
        ),
    ]
    return ParsedData(
        headers={"Mã CN": 1, "Tên chức năng": 2, "Module": 3,
                 "Analysis - Status": 4, "Analysis - End": 5, "Analysis - PIC": 6,
                 "Dev - Status": 7, "Dev - End": 8, "Dev - PIC": 9},
        meta_columns={"ma_cn": 1, "ten_cn": 2, "module": 3},
        phase_groups=[
            PhaseGroup(name="Analysis",
                       attributes={"Status": 4, "End": 5, "PIC": 6}),
            PhaseGroup(name="Dev",
                       attributes={"Status": 7, "End": 8, "PIC": 9}),
        ],
        rows=rows,
        all_phases=["Analysis", "Dev"],
        all_modules=["HR"],
    )


@pytest.fixture
def api(flask_client):
    app_module._state["default"] = {
        "data": _unassigned_data(), "filename": "u.xlsx", "path": "",
    }
    return flask_client


def _fl_codes(resp) -> set[str]:
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb["Function List"]
    return {
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }


def test_fl_reimport_unassigned_khong_status_giu_ca_hai(api):
    r = api.get("/api/projects/default/export-fl-reimport?kinds=unassigned")
    assert r.status_code == 200
    assert _fl_codes(r) == {"U1", "U2"}


def test_fl_reimport_unassigned_l_status_open(api):
    r = api.get(
        "/api/projects/default/export-fl-reimport?kinds=unassigned&l_status=Open"
    )
    assert r.status_code == 200
    assert _fl_codes(r) == {"U1"}


def test_fl_reimport_unassigned_l_status_assigned(api):
    r = api.get(
        "/api/projects/default/export-fl-reimport?kinds=unassigned&l_status=Assigned"
    )
    assert r.status_code == 200
    assert _fl_codes(r) == {"U2"}


# ==========================================================================
# Multi-checkbox Module + Phase filter cho section «Chưa PIC»
# ==========================================================================
# User phản hồi 06/08/2026: Document task thiếu PIC gây nhiễu → default bỏ tick.
# Test HTML/JS/backend/drill_down cùng nhất quán, tránh drift.

class TestUnassignedMsHtml:
    def test_html_has_module_and_phase_multi_select_containers(self):
        sec = INDEX_HTML.split('id="section-unassigned"')[1].split("</section>")[0]
        assert 'id="unassignedModuleMS"' in sec
        assert 'id="unassignedPhaseMS"' in sec
        # Vị trí: Module trước Phase trước Status (đọc trái sang phải hợp lý)
        pos_mod = sec.index('id="unassignedModuleMS"')
        pos_ph = sec.index('id="unassignedPhaseMS"')
        pos_st = sec.index('id="unassignedStatusFilter"')
        assert pos_mod < pos_ph < pos_st


class TestUnassignedMsJs:
    def test_default_uncheck_document_phase(self):
        """Default phase selection = tất cả TRỪ Document (matcher generic
        keyword «document|tai lieu», không hardcode chính xác chữ)."""
        assert "function _unassignedDefaultPhases(" in DASHBOARD_JS
        body = DASHBOARD_JS.split("function _unassignedDefaultPhases(")[1].split(
            "function _unassignedPhaseLsKey("
        )[0]
        assert "_stalledIsDocPhase" in body  # reuse doc-phase detector
        # Fallback: nếu chỉ còn Document thì trả all — không ẩn hết section
        assert "kept.length" in body

    def test_localstorage_key_scoped_by_project(self):
        assert "function _unassignedPhaseLsKey(" in DASHBOARD_JS
        assert "unassignedPhaseSel:" in DASHBOARD_JS
        assert "currentProjectSlug" in DASHBOARD_JS.split(
            "function _unassignedPhaseLsKey("
        )[1].split("}")[0]

    def test_load_phase_sel_auto_ticks_new_phases_except_document(self):
        """Phase mới xuất hiện lần đầu (không có trong `known`) được tự tick,
        trừ khi là Document — user không bị âm thầm ẩn phase mới."""
        body = DASHBOARD_JS.split("function _loadUnassignedPhaseSel(")[1].split(
            "function _saveUnassignedPhaseSel("
        )[0]
        assert "known.has(p)" in body
        assert "!_stalledIsDocPhase(p)" in body

    def test_init_creates_both_ms_and_wires_reset_page(self):
        assert "function _initUnassignedFilters(" in DASHBOARD_JS
        body = DASHBOARD_JS.split("function _initUnassignedFilters(")[1].split(
            "function _unassignedSelectedModules("
        )[0]
        # 2 MS instance riêng, không share state với overdue/stalled
        assert 'key: "unassignedModule"' in body
        assert 'key: "unassignedPhase"' in body
        # Đổi filter → reset về trang 1 (tránh trang trống khi số dòng giảm)
        assert "pageState.unassigned" in body

    def test_render_applies_module_and_phase_filters(self):
        rus = DASHBOARD_JS.split("function renderUnassignedSection(")[1].split(
            "// ========================================================================"
        )[0]
        # Bốn filter đồng thời: status + module + phase (MS) — phase filter chỉ áp
        # khi user KHÔNG tick hết (rỗng và tick-hết đều = «không lọc»).
        assert "_unassignedSelectedModules(" in rus
        assert "_unassignedSelectedPhases(" in rus
        assert "phaseIsFiltering" in rus
        # Count hiển thị filter đang áp — user hiểu tại sao số giảm
        assert "lọc " in rus  # "lọc status=..., module (n), phase (n/N)"

    def test_export_local_params_writes_module_phase_status(self):
        exp = DASHBOARD_JS.split("function _unassignedExportLocalParams(")[1].split(
            "window._unassignedExportLocalParams"
        )[0]
        assert 'params.set("l_status"' in exp
        assert 'params.set("l_module"' in exp
        assert 'params.set("l_phase"' in exp
        # Không set l_phase khi tick hết — tránh chuỗi dài vô nghĩa
        assert "phases.length < allPhases.length" in exp

    def test_drill_down_passes_module_and_phase(self):
        drill = DASHBOARD_JS.split("function openUnassignedDrill(")[1].split(
            "window.openUnassignedDrill"
        )[0]
        assert "filters.module" in drill
        assert "filters.phase" in drill


# ── Backend: kèm l_module + l_phase khi export ────────────────────────────

def _multi_module_data() -> ParsedData:
    """3 module × 2 phase để test filter đồng thời module + phase."""
    today = date.today()
    past = today - timedelta(days=5)
    old = past - timedelta(days=10)
    rows = []
    for i, (code, module, doc_open, dev_open) in enumerate([
        ("HR1", "HR", True, False),
        ("HR2", "HR", False, True),
        ("PR1", "PR", True, False),
        ("PR2", "PR", False, True),
        ("SL1", "SL", True, True),
    ], start=2):
        rows.append(FunctionRow(
            row_num=i,
            meta={"ma_cn": code, "ten_cn": f"T{code}", "module": module},
            phases={
                "Document": PhaseData(
                    status="Open" if doc_open else "Closed",
                    end_date=past if doc_open else old,
                    pics=[] if doc_open else ["P1"],
                ),
                "Dev": PhaseData(
                    status="Open" if dev_open else "Closed",
                    end_date=past if dev_open else old,
                    pics=[] if dev_open else ["P1"],
                ),
            },
        ))
    return ParsedData(
        headers={"Mã CN": 1, "Tên chức năng": 2, "Module": 3,
                 "Document - Status": 4, "Document - End": 5, "Document - PIC": 6,
                 "Dev - Status": 7, "Dev - End": 8, "Dev - PIC": 9},
        meta_columns={"ma_cn": 1, "ten_cn": 2, "module": 3},
        phase_groups=[
            PhaseGroup(name="Document", attributes={"Status": 4, "End": 5, "PIC": 6}),
            PhaseGroup(name="Dev", attributes={"Status": 7, "End": 8, "PIC": 9}),
        ],
        rows=rows,
        all_phases=["Document", "Dev"],
        all_modules=["HR", "PR", "SL"],
    )


@pytest.fixture
def api_multi(flask_client):
    app_module._state["default"] = {
        "data": _multi_module_data(), "filename": "u.xlsx", "path": "",
    }
    return flask_client


def test_export_fl_reimport_l_phase_excludes_document(api_multi):
    """l_phase=Dev → chỉ giữ task thiếu PIC ở Dev khi Dev thực sự đến lượt
    (predecessor Closed). SL1 có Document Open → Dev chưa tới lượt, không
    coi là unassigned tại Dev. HR2 và PR2 có Document Closed → qualify."""
    r = api_multi.get(
        "/api/projects/default/export-fl-reimport?kinds=unassigned&l_phase=Dev"
    )
    assert r.status_code == 200
    assert _fl_codes(r) == {"HR2", "PR2"}


def test_export_fl_reimport_l_module_and_l_phase_together(api_multi):
    r = api_multi.get(
        "/api/projects/default/export-fl-reimport"
        "?kinds=unassigned&l_module=HR,PR&l_phase=Dev"
    )
    assert r.status_code == 200
    assert _fl_codes(r) == {"HR2", "PR2"}


def test_export_fl_reimport_no_local_filter_keeps_all(api_multi):
    r = api_multi.get(
        "/api/projects/default/export-fl-reimport?kinds=unassigned"
    )
    assert r.status_code == 200
    # Toàn bộ task có phase Open thiếu PIC — cả Document lẫn Dev
    assert _fl_codes(r) == {"HR1", "HR2", "PR1", "PR2", "SL1"}


# ── Drill-down: multi-value module + phase ────────────────────────────────

def test_drill_down_unassigned_accepts_multi_module_and_phase():
    from analyzer.drill_down import drill_down

    data = _multi_module_data()
    today = date.today()

    # Chỉ Dev, chỉ 2 module HR + PR → HR2 + PR2
    # (SL1 có Dev nhưng predecessor Document Open → chưa tới lượt Dev,
    # không được coi là unassigned tại phase Dev.)
    result = drill_down(
        data,
        chart="unassigned",
        filters={"module": "HR,PR", "phase": "Dev"},
        today=today,
    )
    codes = {r["ma_cn"] for r in result}
    assert codes == {"HR2", "PR2"}

    # Chỉ Document → HR1 + PR1 + SL1
    result_doc = drill_down(
        data,
        chart="unassigned",
        filters={"phase": "Document"},
        today=today,
    )
    assert {r["ma_cn"] for r in result_doc} == {"HR1", "PR1", "SL1"}


def test_drill_down_unassigned_multi_value_via_list_argument():
    """Filter có thể là list (giống stalled) — không bắt buộc string comma.

    SL1 Dev không đủ điều kiện (predecessor Document Open) nên chỉ HR2 hit."""
    from analyzer.drill_down import drill_down

    data = _multi_module_data()
    today = date.today()
    result = drill_down(
        data,
        chart="unassigned",
        filters={"module": ["HR", "SL"], "phase": ["Dev"]},
        today=today,
    )
    assert {r["ma_cn"] for r in result} == {"HR2"}
