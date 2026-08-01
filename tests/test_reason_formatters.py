"""Focused tests — reason formatters + Unassigned / Overload / FL Ly_do export."""
from __future__ import annotations

import os
from datetime import date

import openpyxl
import pytest

from exporter.reason_formatters import (
    format_risk_factors_detailed,
    process_code,
    reason_aging_wip,
    reason_capacity,
    reason_duration,
    reason_overdue,
    reason_pic_overload,
    reason_scope_creep,
    reason_stalled,
    reason_uat_warning,
    reason_unassigned,
)
from exporter.excel_exporter import export_full_report, export_chart
from exporter.pic_overload_exporter import export_pic_overload_report
from exporter.fl_reimport_export import collect_issue_hits, export_fl_reimport


def test_reason_unassigned_format():
    text = reason_unassigned({
        "predecessor_phase": "Analysis",
        "start_date": "2026-01-10",
        "status": "Open",
    })
    assert "Thiếu PIC" in text
    assert "«Analysis»" in text
    assert "Start 10/01/2026" in text
    assert "Status=Open" in text


def test_reason_overdue_format():
    text = reason_overdue(
        {"end_date": "2026-01-01", "status": "In-progress", "days_overdue": 12},
        today=date(2026, 1, 13),
    )
    assert "End 01/01/2026" in text
    assert "hôm nay 13/01/2026" in text
    assert "Status=In-progress" in text
    assert "trễ 12d" in text


def test_reason_stalled_and_overload():
    st = reason_stalled({
        "completed_phase": "Analysis",
        "waiting_phase": "Coding",
        "completed_date": "2026-01-01",
        "waiting_end_date": "2026-01-05",
        "wait_days": 20,
    })
    assert "«Analysis» Closed" in st
    assert "«Coding» chưa start" in st
    assert "End chờ 05/01/2026" in st

    ol = reason_pic_overload("SonHN6", "2026-01-15", 7, 5)
    assert ol == "PIC SonHN6 ngày 15/01/2026: 7 task > ngưỡng 5"


def test_reason_duration_aging_capacity_scope_uat():
    assert "ngưỡng 3" in reason_duration(
        {"duration_days": 10, "duration_type": "elapsed", "threshold_days": 3}
    )
    assert "aging 21d > ngưỡng 14" in reason_aging_wip(
        {"start_date": "2026-01-01", "aging_days": 21}, 14
    )
    assert "Remaining 80 MH" in reason_capacity(
        {"remaining_mh": 80, "capacity_mh_per_week": 40, "weeks_needed": 2}
    )
    assert "Cột «CR» = «Yes»" in reason_scope_creep(
        {"source": "column", "raw_cr": "Yes"}, column_header="CR"
    )
    assert "Reopen 2" in reason_uat_warning({"reopen_count": 2, "uat_cycle": 3})
    assert "cycle 3 ≥ 2" in reason_uat_warning({"reopen_count": 2, "uat_cycle": 3})


def test_format_risk_factors_detailed_prefers_detail_list():
    text = format_risk_factors_detailed({
        "risk_factors": ["Có phase overdue"],
        "risk_factors_detail": ["Overdue «Analysis» 12d (+20)", "Thiếu PIC «Coding» (+15)"],
        "risk_breakdown": {"overdue": 20, "unassigned": 15},
    })
    assert "Overdue «Analysis» 12d (+20)" in text
    assert "Thiếu PIC «Coding» (+15)" in text


def test_process_code_from_meta_keys():
    assert process_code({"quy_trinh": "TMS.BP.01"}) == "TMS.BP.01"
    assert process_code({"process": "HR.BP.02"}) == "HR.BP.02"
    assert process_code({}) == ""


def test_full_report_unassigned_has_ly_do_and_quy_trinh(tmp_path, metrics):
    path = export_full_report(metrics, str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert "Unassigned_Tasks" in wb.sheetnames
    ws = wb["Unassigned_Tasks"]
    header = [c.value for c in ws[4]]
    assert "Quy trình" in header
    assert "Lý do" in header
    assert "Phase trước" in header
    ly_idx = header.index("Lý do")
    # Có ít nhất 1 data row với Lý do bắt đầu Thiếu PIC (nếu sample có unassigned)
    data = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]
    if data:
        assert any(
            isinstance(r[ly_idx], str) and "Thiếu PIC" in r[ly_idx] for r in data
        )
    wb.close()


def test_chart_unassigned_has_ly_do(tmp_path, metrics):
    path = export_chart("unassigned", metrics, str(tmp_path), mode="detail")
    wb = openpyxl.load_workbook(path)
    ws = wb["Chi_tiet"]
    header = [c.value for c in ws[4]]
    assert "Lý do" in header
    assert "Quy trình" in header
    wb.close()


def test_pic_overload_chi_tiet_has_concurrent_and_ly_do(tmp_path):
    payload = {
        "grain": "day",
        "from": "2026-01-01",
        "to": "2026-01-31",
        "thresholds": {"day_max_tasks": 5},
        "summary": {"highlight_dates": ["2026-01-10"]},
        "by_pic": [],
        "by_period": [],
        "detail": [
            {
                "pic": "A",
                "date": "2026-01-10",
                "project_slug": "p1",
                "project_name": "P1",
                "ma_cn": "CN1",
                "ten_cn": "Func 1",
                "module": "TMS",
                "quy_trinh": "TMS.BP.01",
                "phase": "Coding",
                "status": "In-progress",
                "start": "2026-01-01",
                "end": "2026-01-20",
                "is_overdue": False,
                "concurrent_count": 7,
                "threshold": 5,
                "is_day_overload": True,
            }
        ],
    }
    path = export_pic_overload_report(payload, str(tmp_path), mode="detail")
    wb = openpyxl.load_workbook(path)
    ws = wb["Chi_tiet"]
    header = [c.value for c in ws[4]]
    assert "Số task cùng ngày" in header
    assert "Ngưỡng" in header
    assert "Lý do" in header
    assert "Quy trình" in header
    ly_idx = header.index("Lý do")
    row = next(ws.iter_rows(min_row=5, values_only=True))
    assert "7 task > ngưỡng 5" in str(row[ly_idx])
    wb.close()


def test_fl_reimport_has_ly_do_sheet(tmp_path, parsed_data, metrics):
    hits = collect_issue_hits(
        overdue_list=metrics.get("overdue_list") or [],
        unassigned_list=metrics.get("unassigned_tasks") or [],
        stalled_list=(metrics.get("stalled_tasks") or {}).get("items") or [],
    )
    if not hits:
        pytest.skip("sample không có issue hits")
    path = export_fl_reimport(
        parsed_data,
        hits=hits,
        output_dir=str(tmp_path),
        project_slug="default",
    )
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert "Function List" in wb.sheetnames
    assert "Ly_do" in wb.sheetnames
    ws = wb["Ly_do"]
    header = [c.value for c in ws[4]]
    assert "Mã CN" in header
    assert "Quy trình" in header
    assert "Lý do / Ghi chú" in header
    # Không ghi note tracker vào Function List header
    fl_headers = [c.value for c in wb["Function List"][1]]
    assert not any(h and "Ly_do" in str(h) for h in fl_headers)
    wb.close()
