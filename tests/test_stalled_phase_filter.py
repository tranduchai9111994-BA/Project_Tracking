"""
Filter «Phase chờ» của section Đình trệ (multi-select, mặc định bỏ Document).

Quy ước đã chốt:
  - Lọc theo cột Phase chờ (`waiting_phase`), KHÔNG lọc theo phase vừa xong.
  - Mặc định check mọi phase trừ Document → không âm thầm ẩn UAT/Golive.
  - Drill-down + Excel phải nhận cùng filter để không lệch số với bảng.
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl

from analyzer.drill_down import build_title, drill_down
from app import _parse_drill_filters
from exporter.excel_exporter import export_stalled_report
from parser.excel_parser import FunctionRow, PhaseData, PhaseGroup, ParsedData


TODAY = date(2026, 7, 30)
PAST = TODAY - timedelta(days=10)
PHASES = ["Analysis", "Dev", "Config Local", "Document"]


def _row(ma: str, module: str, closed_upto: int) -> FunctionRow:
    """Row Closed hết phase [0..closed_upto], phase kế tiếp Open → stalled ở đó."""
    phases: dict[str, PhaseData] = {}
    for idx, name in enumerate(PHASES):
        if idx <= closed_upto:
            phases[name] = PhaseData(status="Closed", end_date=PAST)
        elif idx == closed_upto + 1:
            phases[name] = PhaseData(status="Open")
        else:
            phases[name] = PhaseData(status=None)
    return FunctionRow(
        row_num=idx + 2,
        meta={"ma_cn": ma, "ten_cn": f"Func {ma}", "module": module,
              "priority": "Must-have"},
        phases=phases,
    )


def _data() -> ParsedData:
    rows = [
        _row("A.01", "HR", 0),   # chờ Dev
        _row("A.02", "HR", 1),   # chờ Config Local
        _row("A.03", "PR", 2),   # chờ Document
        _row("A.04", "PR", 2),   # chờ Document
    ]
    return ParsedData(
        headers={},
        meta_columns={},
        phase_groups=[PhaseGroup(name=n) for n in PHASES],
        rows=rows,
        all_phases=list(PHASES),
        all_modules=["HR", "PR"],
    )


# ── Drill-down backend ─────────────────────────────────────────────────────

def test_no_filter_returns_all_waiting_phases():
    items = drill_down(_data(), "stalled", {}, today=TODAY)
    assert sorted(i["waiting_phase"] for i in items) == [
        "Config Local", "Dev", "Document", "Document",
    ]


def test_waiting_phase_filter_single():
    items = drill_down(_data(), "stalled", {"waiting_phase": "Document"}, today=TODAY)
    assert {i["ma_cn"] for i in items} == {"A.03", "A.04"}


def test_waiting_phase_filter_multi_excludes_document():
    """Mặc định UI: mọi phase trừ Document."""
    keep = "Analysis,Dev,Config Local"
    items = drill_down(_data(), "stalled", {"waiting_phase": keep}, today=TODAY)
    assert {i["ma_cn"] for i in items} == {"A.01", "A.02"}
    assert all(i["waiting_phase"] != "Document" for i in items)


def test_waiting_phase_matches_only_waiting_not_completed():
    """Config Local là phase đã xong của A.03/A.04 — không được lọt vào."""
    items = drill_down(
        _data(), "stalled", {"waiting_phase": "Config Local"}, today=TODAY,
    )
    assert {i["ma_cn"] for i in items} == {"A.02"}


def test_waiting_phase_accepts_list():
    items = drill_down(
        _data(), "stalled", {"waiting_phase": ["Dev", "Document"]}, today=TODAY,
    )
    assert {i["ma_cn"] for i in items} == {"A.01", "A.03", "A.04"}


def test_module_filter_accepts_comma_separated():
    """FE gửi module="HR,PR" — trước đây so sánh == nên trả 0 item."""
    items = drill_down(_data(), "stalled", {"module": "HR,PR"}, today=TODAY)
    assert len(items) == 4
    only_hr = drill_down(_data(), "stalled", {"module": "HR"}, today=TODAY)
    assert {i["module"] for i in only_hr} == {"HR"}


def test_module_and_waiting_phase_combined():
    items = drill_down(
        _data(), "stalled",
        {"module": "PR", "waiting_phase": "Dev,Config Local"},
        today=TODAY,
    )
    assert items == []


def test_build_title_mentions_waiting_phase():
    title = build_title("stalled", {"waiting_phase": "Dev,Config Local", "module": "HR"})
    assert "Phase chờ: Dev, Config Local" in title
    assert "Module HR" in title
    assert build_title("stalled", {}) == "Task bị Đình trệ"


def test_drill_filters_whitelist_includes_waiting_phase():
    parsed = _parse_drill_filters({"chart": "stalled", "waiting_phase": "Dev", "junk": "x"})
    assert parsed == {"waiting_phase": "Dev"}


# ── Export Excel ───────────────────────────────────────────────────────────

def _stalled_items() -> list[dict]:
    return [
        {"ma_cn": "A.01", "ten_cn": "F1", "module": "HR", "completed_phase": "Analysis",
         "waiting_phase": "Dev", "completed_date": PAST.isoformat(), "wait_days": 10,
         "priority": "Must-have"},
        {"ma_cn": "A.03", "ten_cn": "F3", "module": "PR", "completed_phase": "Config Local",
         "waiting_phase": "Document", "completed_date": PAST.isoformat(), "wait_days": 10,
         "priority": "Should-have"},
    ]


def _detail_codes(path: str) -> list[str]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Chi_tiet"]
    codes = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and str(row[1] or "").startswith("A."):
            codes.append(row[1])
    return codes


def test_export_filters_by_waiting_phase(tmp_path):
    path = export_stalled_report(
        _stalled_items(),
        output_dir=str(tmp_path),
        filters={"waiting_phase": "Dev"},
    )
    assert _detail_codes(path) == ["A.01"]


def test_export_subtitle_shows_waiting_phase(tmp_path):
    path = export_stalled_report(
        _stalled_items(),
        output_dir=str(tmp_path),
        filters={"module": "HR", "waiting_phase": "Dev,Config Local"},
    )
    wb = openpyxl.load_workbook(path)
    text = " ".join(
        str(c.value or "")
        for row in wb["Chi_tiet"].iter_rows(min_row=1, max_row=4)
        for c in row
    )
    assert "Phase chờ: Dev, Config Local" in text
    assert "Module: HR" in text


def test_export_without_filter_keeps_all(tmp_path):
    path = export_stalled_report(_stalled_items(), output_dir=str(tmp_path))
    assert _detail_codes(path) == ["A.01", "A.03"]


# ── FE wiring ──────────────────────────────────────────────────────────────

def _fe_sources() -> tuple[str, str]:
    root = Path(__file__).resolve().parents[1]
    return (
        (root / "static" / "js" / "dashboard.js").read_text(encoding="utf-8"),
        (root / "templates" / "index.html").read_text(encoding="utf-8"),
    )


def test_html_has_phase_multiselect_and_reset():
    js, html = _fe_sources()
    assert 'id="stalledPhaseMS"' in html
    assert "resetStalledPhaseFilter()" in html
    assert "function resetStalledPhaseFilter(" in js


def test_js_default_skips_document_without_hardcoding_phase_list():
    js, _ = _fe_sources()
    assert "function _stalledIsDocPhase(" in js
    assert "function _stalledDefaultPhases(" in js
    # Match keyword sau khi bỏ dấu — không so tên phase cứng
    assert 'n.includes("document")' in js
    assert 'n.includes("tai lieu")' in js
    # Fallback: không có phase nào ngoài Document → trả tất cả, tránh bảng trống
    assert "return kept.length ? kept : all.slice();" in js


def test_js_persists_selection_per_project():
    js, _ = _fe_sources()
    assert "stalledPhaseSel:" in js
    assert "function _loadStalledPhaseSel(" in js
    assert "function _saveStalledPhaseSel(" in js
    # Phase mới xuất hiện phải tự check (trừ Document)
    assert "if (!known.has(p) && !_stalledIsDocPhase(p)) sel.add(p);" in js


def test_js_filters_items_by_waiting_phase():
    js, _ = _fe_sources()
    assert "function _stalledSelectedPhases(" in js
    assert 'set.has(String(i.waiting_phase || "").trim())' in js


def test_js_sends_waiting_phase_to_export_and_drill():
    js, _ = _fe_sources()
    assert 'params.set("waiting_phase", phArr.join(","))' in js
    assert 'filters.waiting_phase = phases.join(",")' in js


def test_js_funnel_not_filtered_by_phase():
    """Funnel là tiến độ Closed toàn trình — chỉ nhận Module override."""
    js, _ = _fe_sources()
    assert "const funnel = _stalledDerivedFunnel(selected);" in js


# ── API smoke ──────────────────────────────────────────────────────────────

def test_api_export_stalled_accepts_waiting_phase(flask_client, sample_xlsx_path):
    with open(sample_xlsx_path, "rb") as f:
        r = flask_client.post(
            "/api/upload",
            data={"file": (f, "sample.xlsx")},
            content_type="multipart/form-data",
        )
    assert r.status_code == 200, r.get_data(as_text=True)
    slug = r.get_json().get("project", {}).get("slug") or "default"

    r2 = flask_client.get(
        f"/api/projects/{slug}/export-stalled?mode=both&waiting_phase=Dev"
    )
    assert r2.status_code == 200, r2.get_data(as_text=True)
    assert r2.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )


def test_api_drill_down_accepts_waiting_phase(flask_client, sample_xlsx_path):
    with open(sample_xlsx_path, "rb") as f:
        r = flask_client.post(
            "/api/upload",
            data={"file": (f, "sample.xlsx")},
            content_type="multipart/form-data",
        )
    slug = r.get_json().get("project", {}).get("slug") or "default"

    r2 = flask_client.get(
        f"/api/projects/{slug}/drill-down?chart=stalled&waiting_phase=Dev"
    )
    assert r2.status_code == 200, r2.get_data(as_text=True)
    payload = r2.get_json()
    assert payload["filters"] == {"waiting_phase": "Dev"}
    assert all(i.get("waiting_phase") in ("Dev", None) for i in payload["items"])
