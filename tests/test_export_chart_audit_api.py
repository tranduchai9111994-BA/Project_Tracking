"""HTTP integration tests cho export-chart và audit-report endpoints."""
import io
import openpyxl


def _upload(client, sample_xlsx_path):
    with open(sample_xlsx_path, "rb") as f:
        return client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "sample.xlsx")},
            content_type="multipart/form-data",
        )


# ==========================================================================
# GET/POST /api/projects/<slug>/export-chart
# ==========================================================================

def test_export_chart_requires_upload(flask_client):
    r = flask_client.get("/api/projects/default/export-chart?chart=priority")
    assert r.status_code in (400, 404)
    assert "error" in r.get_json()


def test_export_chart_missing_chart_param(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/export-chart")
    assert r.status_code == 400
    assert "chart" in r.get_json()["error"].lower()


def test_export_chart_unsupported(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/export-chart?chart=not_a_real_chart")
    assert r.status_code == 400
    data = r.get_json()
    assert "supported" in data


def test_export_chart_priority_xlsx(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/export-chart?chart=priority")
    assert r.status_code == 200
    assert "spreadsheet" in (r.content_type or "") or r.data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "Tong_hop" in wb.sheetnames
    assert "Chi_tiet" in wb.sheetnames
    wb.close()


def test_export_chart_mode_summary_one_sheet(flask_client, sample_xlsx_path):
    """mode=summary → chỉ 1 sheet Tong_hop."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/export-chart?chart=priority&mode=summary")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Tong_hop"]
    wb.close()


def test_export_chart_pic_workload_both(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/export-chart?chart=pic_workload&mode=both")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "Tong_hop" in wb.sheetnames
    assert "Chi_tiet" in wb.sheetnames
    headers = [c.value for c in wb["Chi_tiet"][4]]
    assert "PIC" in headers
    assert "Status" in headers
    wb.close()


def test_export_chart_effort_heatmap_with_module_filter(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-chart?chart=effort_heatmap&module=TMS"
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert len(wb.sheetnames) >= 1
    wb.close()


def test_export_chart_post_body(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/export-chart",
        json={"chart": "task_type", "module": ["TMS"]},
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"


def test_export_task_type_detail_status_columns(flask_client, sample_xlsx_path):
    """task_type export có sheet Chi_tiet với cột status work-bucket + số row khớp filter."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-chart?chart=task_type&module=TMS"
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "Tong_hop" in wb.sheetnames
    assert "Chi_tiet" in wb.sheetnames
    ws = wb["Chi_tiet"]
    # Header row 4 (title/subtitle/_write_sheet convention)
    headers = [c.value for c in ws[4]]
    assert "Mã CN" in headers
    assert "Tên chức năng" in headers
    assert "Module" in headers
    # Work-bucket status columns từ sample: Phân tích / Lập trình / UAT
    status_cols = {"Phân tích", "Lập trình", "UAT"}
    assert status_cols.issubset(set(headers))
    # Data rows (sau header) — sample có 2 function TMS
    data_rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            break
        # Summary row bắt đầu bằng "Tổng:"
        if isinstance(row[0], str) and row[0].startswith("Tổng"):
            break
        data_rows.append(row)
    assert len(data_rows) == 2
    module_idx = headers.index("Module")
    assert all(row[module_idx] == "TMS" for row in data_rows)
    # Status cells không phải số (đã chuẩn hóa Closed / In-progress / … / blank)
    phan_tich_idx = headers.index("Phân tích")
    statuses = {row[phan_tich_idx] for row in data_rows}
    assert "Closed" in statuses
    wb.close()


def test_export_task_type_filter_row_count(flask_client, sample_xlsx_path):
    """Filter Module=PR → Chi_tiet chỉ còn function PR (sample: 1 row)."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/export-chart",
        json={"chart": "task_type", "module": ["PR"], "group_by": "module"},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb["Chi_tiet"]
    headers = [c.value for c in ws[4]]
    ma_idx = headers.index("Mã CN")
    codes = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None or (isinstance(row[0], str) and str(row[0]).startswith("Tổng")):
            break
        codes.append(row[ma_idx])
    assert codes == ["PR.FR.03"]
    wb.close()


def test_export_chart_mode_summary_one_sheet(flask_client, sample_xlsx_path):
    """mode=summary → chỉ sheet Tong_hop (priority không có Theo_nhom)."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-chart?chart=priority&mode=summary"
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Tong_hop"]
    wb.close()


def test_export_chart_mode_detail_only(flask_client, sample_xlsx_path):
    """mode=detail → chỉ sheet Chi_tiet."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-chart?chart=complexity&mode=detail"
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Chi_tiet"]
    wb.close()


def test_export_chart_mode_both_default(flask_client, sample_xlsx_path):
    """Default / mode=both → Tong_hop + Chi_tiet."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-chart?chart=fit_gap&mode=both"
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "Tong_hop" in wb.sheetnames
    assert "Chi_tiet" in wb.sheetnames
    wb.close()


def test_export_module_overview_summary_mode(flask_client, sample_xlsx_path):
    """module_overview + mode=summary → 1 sheet Tong_hop."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/export-chart",
        json={"chart": "module_overview", "mode": "summary", "module": ["TMS"]},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Tong_hop"]
    wb.close()


def test_export_task_type_summary_has_no_chi_tiet(flask_client, sample_xlsx_path):
    """task_type mode=summary → Tong_hop (+ Theo_nhom), không Chi_tiet."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-chart?chart=task_type&mode=summary"
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "Tong_hop" in wb.sheetnames
    assert "Chi_tiet" not in wb.sheetnames
    wb.close()


# ==========================================================================
# GET/POST /api/projects/<slug>/audit-report
# ==========================================================================

def test_audit_report_requires_upload(flask_client):
    r = flask_client.get("/api/projects/default/audit-report?scope=all")
    assert r.status_code in (400, 404)
    assert "error" in r.get_json()


def test_audit_report_all_xlsx(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/audit-report?scope=all")
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    # Audit report có nhiều sheet (summary + issue sheets)
    assert len(wb.sheetnames) >= 2
    wb.close()


def test_audit_report_filtered_scope(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/audit-report?scope=filtered&module=TMS"
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"


def test_audit_report_post(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/audit-report",
        json={"scope": "all"},
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
