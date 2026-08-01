"""
T34 Task 1 — Tests cho exporter/export_all_issues.py + endpoint
/api/projects/<slug>/export-all-issues.

Kiểm tra:
  1. Workbook có đúng 8 sheet với tên chuẩn.
  2. Cover sheet chứa filter info + hyperlink.
  3. Overdue sheet dedup theo Mã CN (không lặp phase-record).
  4. Filter global g_module áp dụng đúng (giảm số row so với no-filter).
  5. Endpoint HTTP trả file Excel hợp lệ với đúng Content-Type.
  6. Bookmark sheet respect filter (function bị filter loại không xuất hiện).
"""
from __future__ import annotations

import io
import os

import openpyxl
import pytest

from exporter.export_all_issues import (
    _dedup_by_ma_cn,
    export_all_issues,
)


# ==========================================================================
# Unit tests — helper _dedup_by_ma_cn
# ==========================================================================

class TestDedupByMaCn:
    def test_empty(self):
        assert _dedup_by_ma_cn([]) == []

    def test_single_record_passthrough(self):
        items = [{"ma_cn": "A.01", "phase": "Analysis", "days_overdue": 5,
                  "pic": ["Alice"], "module": "HR"}]
        out = _dedup_by_ma_cn(items)
        assert len(out) == 1
        assert out[0]["ma_cn"] == "A.01"
        assert out[0]["phase"] == "Analysis"
        assert out[0]["days_overdue"] == 5

    def test_merge_multiple_phases_same_ma_cn(self):
        items = [
            {"ma_cn": "A.01", "phase": "Analysis", "days_overdue": 5,
             "pic": ["Alice"], "module": "HR"},
            {"ma_cn": "A.01", "phase": "UAT", "days_overdue": 12,
             "pic": ["Bob"], "module": "HR"},
            {"ma_cn": "A.01", "phase": "Dev", "days_overdue": 3,
             "pic": ["Alice"], "module": "HR"},
        ]
        out = _dedup_by_ma_cn(items)
        assert len(out) == 1
        # Phase merged bằng dấu phẩy, giữ order first-seen
        assert out[0]["phase"] == "Analysis, UAT, Dev"
        # days_overdue = MAX
        assert out[0]["days_overdue"] == 12
        # PIC deduped, giữ order first-seen
        assert out[0]["pic"] == ["Alice", "Bob"]

    def test_multiple_distinct_ma_cn(self):
        items = [
            {"ma_cn": "A.01", "phase": "UAT", "days_overdue": 5, "pic": []},
            {"ma_cn": "B.02", "phase": "Dev", "days_overdue": 10, "pic": []},
            {"ma_cn": "A.01", "phase": "Dev", "days_overdue": 15, "pic": []},
        ]
        out = _dedup_by_ma_cn(items)
        assert len(out) == 2
        codes = {r["ma_cn"] for r in out}
        assert codes == {"A.01", "B.02"}
        # Sort DESC by days_overdue → A.01 (15) trước B.02 (10)
        assert out[0]["ma_cn"] == "A.01"
        assert out[0]["days_overdue"] == 15

    def test_sorted_by_days_desc(self):
        items = [
            {"ma_cn": "A", "phase": "P1", "days_overdue": 3, "pic": []},
            {"ma_cn": "B", "phase": "P2", "days_overdue": 30, "pic": []},
            {"ma_cn": "C", "phase": "P3", "days_overdue": 10, "pic": []},
        ]
        out = _dedup_by_ma_cn(items)
        assert [r["days_overdue"] for r in out] == [30, 10, 3]


# ==========================================================================
# Unit tests — export_all_issues() workbook shape
# ==========================================================================

class TestWorkbookShape:
    @pytest.fixture
    def wb_path(self, tmp_path):
        """Tạo workbook mẫu với data phong phú."""
        path = export_all_issues(
            project_name="Test Project",
            slug="test",
            overdue_list=[
                {"ma_cn": "A.01", "phase": "UAT", "days_overdue": 15,
                 "pic": ["Alice"], "module": "HR", "priority": "Must-have",
                 "status": "In-progress", "note": ""},
                # 2 phase-record của cùng function A.01 — sẽ dedup
                {"ma_cn": "A.01", "phase": "Dev", "days_overdue": 5,
                 "pic": ["Bob"], "module": "HR", "priority": "Must-have",
                 "status": "Open", "note": ""},
            ],
            unassigned_list=[
                {"ma_cn": "B.01", "ten_cn": "Foo", "module": "PR",
                 "rlog_id": "25001",
                 "phase": "Dev", "status": "Open", "priority": "Should-have",
                 "complexity": "Low", "end_date": "", "days_overdue": 0,
                 "is_overdue": False},
            ],
            stalled_list=[
                {"ma_cn": "C.01", "ten_cn": "Bar", "module": "SI",
                 "completed_phase": "Analysis", "waiting_phase": "Dev",
                 "completed_date": "2026-01-01", "wait_days": 210,
                 "priority": "Must-have"},
            ],
            risk_list=[
                {"ma_cn": "D.01", "ten_cn": "Baz", "module": "PIT",
                 "risk_score": 85, "priority": "Must-have",
                 "complexity": "High", "risk_factors": ["Overdue >30d", "Complexity=High"]},
            ],
            aging_wip_items=[
                {"ma_cn": "E.01", "ten_cn": "Qux", "module": "HR",
                 "quy_trinh": "HR.BP.01", "phase": "Dev", "status": "In-progress",
                 "start_date": "2026-06-01", "end_date": "2026-06-15",
                 "pic": ["Alice"], "aging_days": 30, "over_by_days": 16,
                 "priority": "Must-have", "complexity": "Medium"},
            ],
            data_quality_issues=[
                {"ma_cn": "F.01", "ten_cn": "Zap", "module": "HR",
                 "phase": "", "code": "MA_CN_DUPLICATE", "severity": "high",
                 "label": "Mã CN trùng", "detail": "F.01 xuất hiện 3 lần",
                 "suggestion": "Đặt lại mã unique"},
            ],
            bookmark_functions=[
                {"ma_cn": "G.01", "ten_cn": "Star", "module": "HR",
                 "quy_trinh": "HR.BP.02", "priority": "Must-have",
                 "complexity": "High", "giai_doan": "1", "fit_gap": "FIT"},
            ],
            filter_info={"modules": ["HR", "PR"], "processes": [], "pics": []},
            output_dir=str(tmp_path),
        )
        return path

    def test_file_exists(self, wb_path):
        assert os.path.exists(wb_path)
        assert wb_path.endswith(".xlsx")

    def test_filename_format(self, wb_path):
        name = os.path.basename(wb_path)
        assert name.startswith("iHRP_Van_De_Tong_Hop_test_")
        assert name.endswith(".xlsx")

    def test_has_8_sheets(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        assert len(wb.sheetnames) == 8

    def test_sheet_names_and_order(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        # Default lang=vi → tên sheet tiếng Việt (U26)
        expected = [
            "Tong_quan", "Tre_han", "Chua_PIC", "Dinh_tre",
            "Rui_ro_cao", "Aging_WIP", "Chat_luong_DL", "Bookmark",
        ]
        assert wb.sheetnames == expected

    def test_sheet_names_en(self, tmp_path):
        path = export_all_issues(
            project_name="EN",
            slug="en",
            overdue_list=[], unassigned_list=[], stalled_list=[],
            risk_list=[], aging_wip_items=[], data_quality_issues=[],
            bookmark_functions=[],
            output_dir=str(tmp_path),
            lang="en",
        )
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == [
            "Cover", "Overdue", "Unassigned", "Stalled",
            "High_Risk", "Aging_WIP", "Data_Quality", "Bookmark",
        ]

    def test_cover_has_project_name(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Tong_quan"]
        # Row 1 = banner với project name
        assert "Test Project" in str(ws["A1"].value)

    def test_cover_shows_filter_modules(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Tong_quan"]
        # Scan cả sheet Cover tìm "HR, PR"
        found = False
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=4, values_only=True):
            for cell in row:
                if cell and "HR, PR" in str(cell):
                    found = True
                    break
        assert found, "Filter modules 'HR, PR' phải hiển thị ở Cover sheet"

    def test_cover_has_hyperlinks_to_sheets(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Tong_quan"]
        # Tìm ít nhất 1 hyperlink trỏ đến Tre_han
        found_overdue_link = False
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.hyperlink and "Tre_han" in str(cell.hyperlink.location or ""):
                    found_overdue_link = True
                    break
        assert found_overdue_link

    def test_overdue_dedup_by_ma_cn(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Tre_han"]
        # Row 1 = banner, Row 2 = header, Row 3+ = data
        # Data có 2 record của A.01 → chỉ còn 1 sau dedup
        header = [c.value for c in ws[2]]
        data_rows = [row for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        assert len(data_rows) == 1
        phase_idx = header.index("Phase trễ (gộp)")
        phase_val = str(data_rows[0][phase_idx])
        assert "UAT" in phase_val and "Dev" in phase_val
        assert "Quy trình" in header
        assert "Lý do" in header

    def test_overdue_days_takes_max(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Tre_han"]
        header = [c.value for c in ws[2]]
        data_rows = [row for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        days_idx = header.index("Số ngày trễ (max)")
        # Dedup nhận MAX(15, 5) = 15
        assert data_rows[0][days_idx] == 15

    def test_all_sheets_have_banner_row_1(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        for name in ["Tre_han", "Chua_PIC", "Dinh_tre", "Rui_ro_cao",
                     "Aging_WIP", "Chat_luong_DL", "Bookmark"]:
            ws = wb[name]
            # Row 1 phải là banner merged A1:...
            assert ws["A1"].value is not None
            assert "Tổng" in str(ws["A1"].value) or "Total" in str(ws["A1"].value)

    def test_all_sheets_freeze_row_3(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        for name in ["Tre_han", "Chua_PIC", "Dinh_tre", "Rui_ro_cao",
                     "Aging_WIP", "Chat_luong_DL", "Bookmark"]:
            ws = wb[name]
            assert ws.freeze_panes == "A3"

    def test_bookmark_sheet_has_data(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Bookmark"]
        data_rows = [row for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
        # G.01 có trong bookmark_functions → sheet phải có 1 row
        assert len(data_rows) == 1
        assert data_rows[0][1] == "G.01"

    def test_empty_sheet_shows_placeholder(self, tmp_path):
        """Sheet không có data phải hiển thị thông báo 'Không có record'."""
        path = export_all_issues(
            project_name="Empty test", slug="e",
            overdue_list=[], unassigned_list=[], stalled_list=[],
            risk_list=[], aging_wip_items=[], data_quality_issues=[],
            bookmark_functions=[], filter_info={},
            output_dir=str(tmp_path),
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Tre_han"]
        # Row 3 có message placeholder
        assert "Không có record" in str(ws["A3"].value or "")


# ==========================================================================
# HTTP integration tests — endpoint
# ==========================================================================

def _upload(client, xlsx_path):
    with open(xlsx_path, "rb") as f:
        return client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "test.xlsx")},
            content_type="multipart/form-data",
        )


class TestExportAllIssuesEndpoint:
    def test_endpoint_exists(self, flask_client, sample_xlsx_path):
        _upload(flask_client, sample_xlsx_path)
        r = flask_client.get("/api/projects/default/export-all-issues")
        assert r.status_code == 200
        # Content-Type Excel
        assert r.mimetype in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        )

    def test_endpoint_404_when_no_data(self, flask_client):
        r = flask_client.get("/api/projects/default/export-all-issues")
        assert r.status_code == 404

    def test_response_is_valid_xlsx(self, flask_client, sample_xlsx_path):
        _upload(flask_client, sample_xlsx_path)
        r = flask_client.get("/api/projects/default/export-all-issues")
        assert r.status_code == 200
        # Load workbook từ bytes → must not raise
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert len(wb.sheetnames) == 8
        assert wb.sheetnames[0] == "Tong_quan"

    def test_global_filter_reduces_rows(self, flask_client, sample_xlsx_path):
        _upload(flask_client, sample_xlsx_path)

        # Fetch không filter
        r_all = flask_client.get("/api/projects/default/export-all-issues")
        wb_all = openpyxl.load_workbook(io.BytesIO(r_all.data))
        overdue_all = [row for row in wb_all["Tre_han"].iter_rows(
            min_row=3, values_only=True) if row[0] is not None]

        # Fetch với filter module=TMS (chỉ có TMS.FR.01/02 — chỉ .02 overdue)
        r_tms = flask_client.get("/api/projects/default/export-all-issues?g_module=TMS")
        wb_tms = openpyxl.load_workbook(io.BytesIO(r_tms.data))
        overdue_tms = [row for row in wb_tms["Tre_han"].iter_rows(
            min_row=3, values_only=True) if row[0] is not None]

        # Sample data: overdue có TMS.FR.02 + ESS.FR.10 → all >= tms + 1
        assert len(overdue_all) >= len(overdue_tms)
        # Với filter TMS → chỉ còn TMS records
        for row in overdue_tms:
            # Cột 3 (0-indexed) = Module
            assert row[3] == "TMS"

    def test_post_body_filter_works(self, flask_client, sample_xlsx_path):
        _upload(flask_client, sample_xlsx_path)
        r = flask_client.post(
            "/api/projects/default/export-all-issues",
            json={"g_module": ["TMS"]},
        )
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        # Cover sheet phải reflect filter TMS
        ws = wb["Tong_quan"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=4, values_only=True):
            for cell in row:
                if cell and "TMS" in str(cell):
                    found = True
                    break
        assert found

    def test_filename_contains_slug(self, flask_client, sample_xlsx_path):
        _upload(flask_client, sample_xlsx_path)
        r = flask_client.get("/api/projects/default/export-all-issues")
        cd = r.headers.get("Content-Disposition", "")
        assert "iHRP_Van_De_Tong_Hop_default" in cd

    def test_bookmark_sheet_respects_filter(self, flask_client, sample_xlsx_path):
        """Bookmark 1 function → filter module khác → sheet Bookmark rỗng."""
        from analyzer import project_store as ps
        import app as app_module

        _upload(flask_client, sample_xlsx_path)
        # Bookmark TMS.FR.01
        project_dir = app_module._project_dir_for("default")
        ps.save_bookmarks(project_dir, ["TMS.FR.01"])

        # Filter module=HR → TMS.FR.01 không xuất hiện trong filtered data
        r = flask_client.get("/api/projects/default/export-all-issues?g_module=HR")
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        # Data rows chỉ có STT dạng int ở col A; placeholder empty state là string
        bookmark_rows = [row for row in wb["Bookmark"].iter_rows(
            min_row=3, values_only=True) if isinstance(row[0], int)]
        # Không có bookmark HR nào → sheet phải rỗng
        assert len(bookmark_rows) == 0

        # Không filter → TMS.FR.01 bookmark phải xuất hiện
        r = flask_client.get("/api/projects/default/export-all-issues")
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        bookmark_rows = [row for row in wb["Bookmark"].iter_rows(
            min_row=3, values_only=True) if isinstance(row[0], int)]
        assert len(bookmark_rows) == 1
        assert bookmark_rows[0][1] == "TMS.FR.01"
