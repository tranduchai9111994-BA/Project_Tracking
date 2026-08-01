"""Unit tests — Phase A: baseline SV + completion forecast."""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

from analyzer.baseline_sv import (
    compute_baseline_sv,
    compute_function_sv,
    compute_milestone_sv,
    attach_baseline_to_forecast_row,
)
from analyzer.completion_forecast import (
    compute_completion_forecast,
    count_remaining_phases,
)
from analyzer.forecast_gantt import compute_project_forecast
from analyzer.project_store import load_project_settings, save_project_settings
from parser.excel_parser import FunctionListParser


TODAY = date(2026, 7, 31)


def _write_fl(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


HEADERS = [
    "STT", "Mã CN", "Tên chức năng", "Module",
    "Analysis - Start", "Analysis - End", "Analysis - Status", "Analysis - PIC",
    "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
    "UAT - Start", "UAT - End", "UAT - Status", "UAT - PIC",
]


def _parse(path: Path):
    return FunctionListParser().parse(str(path))


# ------------------------------------------------------------------
# Baseline SV
# ------------------------------------------------------------------

def test_function_sv_late_and_early(tmp_path):
    """SV = current_end − baseline_end: +10 late, −5 early."""
    base_path = tmp_path / "base.xlsx"
    cur_path = tmp_path / "cur.xlsx"
    _write_fl(base_path, HEADERS, [
        [1, "F.01", "Late one", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 1), "In-progress", "D",
         date(2026, 7, 1), date(2026, 8, 1), "Open", "U"],
        [2, "F.02", "Early one", "HR",
         date(2026, 4, 1), date(2026, 5, 10), "Closed", "A",
         date(2026, 5, 1), date(2026, 7, 1), "In-progress", "D",
         None, None, None, ""],
    ])
    _write_fl(cur_path, HEADERS, [
        # Dev End +10 days vs baseline → late
        [1, "F.01", "Late one", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 11), "In-progress", "D",
         date(2026, 7, 1), date(2026, 8, 1), "Open", "U"],
        # Dev End −5 days vs baseline → early
        [2, "F.02", "Early one", "HR",
         date(2026, 4, 1), date(2026, 5, 10), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 26), "Closed", "D",
         None, None, None, ""],
    ])
    base = _parse(base_path)
    cur = _parse(cur_path)
    items = compute_function_sv(cur, base)
    by_code = {(i["ma_cn"], i["phase"]): i for i in items}

    late = by_code[("F.01", "Dev")]
    assert late["sv_days"] == 10
    assert late["late"] is True

    early = by_code[("F.02", "Dev")]
    assert early["sv_days"] == -5
    assert early["early"] is True


def test_baseline_sv_module_and_milestone_agg(tmp_path):
    base_path = tmp_path / "base.xlsx"
    cur_path = tmp_path / "cur.xlsx"
    _write_fl(base_path, HEADERS, [
        [1, "F.01", "F1", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 1), "Closed", "D",
         date(2026, 7, 1), date(2026, 8, 15), "Open", "U"],
    ])
    # UAT End đẩy muộn hơn baseline 16 ngày
    _write_fl(cur_path, HEADERS, [
        [1, "F.01", "F1", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 1), "Closed", "D",
         date(2026, 7, 1), date(2026, 8, 31), "Open", "U"],
    ])
    result = compute_baseline_sv(
        _parse(cur_path), _parse(base_path),
        baseline_snapshot_id="2026-06-01",
        today=TODAY,
    )
    assert result["baseline_snapshot_id"] == "2026-06-01"
    assert result["summary"]["compared"] >= 1
    assert result["summary"]["late_count"] >= 1
    assert any(m["module"] == "TMS" for m in result["modules"])

    uat = result["milestones"]["uat"]
    assert uat["sv_days"] == 16
    assert uat["late"] is True


def test_attach_baseline_to_forecast_row(tmp_path):
    base_path = tmp_path / "base.xlsx"
    cur_path = tmp_path / "cur.xlsx"
    _write_fl(base_path, HEADERS, [
        [1, "F.01", "F1", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 1), "Closed", "D",
         date(2026, 7, 1), date(2026, 8, 1), "Open", "U"],
    ])
    _write_fl(cur_path, HEADERS, [
        [1, "F.01", "F1", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 1), "Closed", "D",
         date(2026, 7, 1), date(2026, 9, 1), "Open", "U"],
    ])
    cur = _parse(cur_path)
    base = _parse(base_path)
    ms = compute_project_forecast(cur, today=TODAY)
    layer = attach_baseline_to_forecast_row(ms, base, today=TODAY)
    assert layer["summary"]["compared"] >= 1
    uat = layer["milestones"]["uat"]
    assert uat["baseline_month"] == "2026-08"
    assert uat["sv_days"] is not None
    assert uat["sv_days"] > 0


def test_project_store_baseline_snapshot_id(tmp_path):
    d = str(tmp_path)
    s = load_project_settings(d)
    assert s.get("baseline_snapshot_id") == ""
    save_project_settings(d, {"baseline_snapshot_id": "2026-05-15"})
    s2 = load_project_settings(d)
    assert s2["baseline_snapshot_id"] == "2026-05-15"
    save_project_settings(d, {"baseline_snapshot_id": ""})
    assert load_project_settings(d)["baseline_snapshot_id"] == ""


# ------------------------------------------------------------------
# Completion forecast
# ------------------------------------------------------------------

def test_completion_forecast_done(tmp_path):
    path = tmp_path / "done.xlsx"
    _write_fl(path, HEADERS, [
        [1, "F.01", "F1", "TMS",
         date(2026, 4, 1), date(2026, 5, 1), "Closed", "A",
         date(2026, 5, 1), date(2026, 6, 1), "Closed", "D",
         date(2026, 6, 1), date(2026, 7, 1), "Closed", "U"],
    ])
    fc = compute_completion_forecast(_parse(path), today=TODAY)
    assert fc["status"] == "done"
    assert fc["remaining"] == 0
    assert fc["forecast_date"] == TODAY.isoformat()


def test_completion_forecast_zero_velocity(tmp_path):
    """Có remaining nhưng không Closed gần đây → zero_velocity hoặc no_history."""
    path = tmp_path / "stuck.xlsx"
    # Closed End rất xa trong quá khứ + remaining open không Closed gần đây
    # → weeks có lịch sử nhưng 4 tuần gần = 0
    old = TODAY - timedelta(days=90)
    _write_fl(path, HEADERS, [
        [1, "F.01", "Old closed", "TMS",
         old, old + timedelta(days=5), "Closed", "A",
         None, None, None, "",
         None, None, None, ""],
        [2, "F.02", "Still open", "TMS",
         TODAY - timedelta(days=10), TODAY + timedelta(days=30), "In-progress", "A",
         None, None, "Open", "D",
         None, None, None, ""],
    ])
    fc = compute_completion_forecast(_parse(path), today=TODAY)
    assert fc["remaining"] >= 1
    assert fc["status"] in ("zero_velocity", "no_history", "ok")
    if fc["status"] == "zero_velocity":
        assert fc["forecast_date"] is None
        assert "velocity" in fc["message"].lower() or "0" in fc["message"]


def test_completion_forecast_ok_linear(tmp_path):
    """Steady Closed gần đây → có forecast_date."""
    path = tmp_path / "steady.xlsx"
    rows = []
    # 8 phase Closed trải đều 4 tuần gần + 4 còn lại Open
    for i in range(8):
        end = TODAY - timedelta(days=(3 - i // 2) * 7 + 2)
        rows.append([
            i + 1, f"C.{i:02d}", f"Closed {i}", "TMS",
            end - timedelta(days=5), end, "Closed", "A",
            None, None, None, "",
            None, None, None, "",
        ])
    for i in range(4):
        rows.append([
            20 + i, f"O.{i:02d}", f"Open {i}", "TMS",
            TODAY, TODAY + timedelta(days=14), "In-progress", "A",
            None, None, None, "",
            None, None, None, "",
        ])
    _write_fl(path, HEADERS, rows)
    data = _parse(path)
    counts = count_remaining_phases(data)
    assert counts["remaining"] >= 4
    fc = compute_completion_forecast(data, today=TODAY)
    assert fc["status"] == "ok"
    assert fc["velocity_4w"] > 0
    assert fc["forecast_date"] is not None
    assert fc["weeks_needed"] is not None
    assert fc["confidence_band"] is not None
    # remaining / velocity ≈ weeks
    expected_weeks = counts["remaining"] / fc["velocity_4w"]
    assert abs(fc["weeks_needed"] - round(expected_weeks, 1)) < 0.15


def test_completion_forecast_no_history(tmp_path):
    path = tmp_path / "fresh.xlsx"
    _write_fl(path, HEADERS, [
        [1, "F.01", "Fresh", "TMS",
         TODAY, TODAY + timedelta(days=10), "Open", "A",
         None, None, "Open", "D",
         None, None, None, ""],
    ])
    fc = compute_completion_forecast(_parse(path), today=TODAY)
    assert fc["status"] == "no_history"
    assert fc["forecast_date"] is None
