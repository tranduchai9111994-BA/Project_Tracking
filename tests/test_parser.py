"""Tests cho parser.excel_parser."""
from datetime import date

import pytest

from parser.excel_parser import FunctionListParser, VALID_STATUSES


def test_parse_returns_all_rows(parsed_data):
    """File mẫu có 6 rows data."""
    assert len(parsed_data.rows) == 6


def test_parse_detects_modules(parsed_data):
    """Modules được phát hiện đúng và sorted."""
    assert parsed_data.all_modules == ["ESS", "HR", "PR", "SYS", "TMS"]


def test_parse_detects_phase_groups(parsed_data):
    """3 phase groups: Analysis, Dev, UAT (thứ tự xuất hiện)."""
    phase_names = [pg.name for pg in parsed_data.phase_groups]
    assert phase_names == ["Analysis", "Dev", "UAT"]


def test_phase_group_task_type_mapping(parsed_data):
    """task_type mapping đúng."""
    mapping = {pg.name: pg.task_type for pg in parsed_data.phase_groups}
    assert mapping["Analysis"] == "Phân tích"
    assert mapping["Dev"] == "Lập trình"
    assert mapping["UAT"] == "UAT"


def test_uat_uses_from_to_columns(parsed_data):
    """UAT phase có From/To và 2 loại PIC (FPT + MPHG)."""
    uat = next(pg for pg in parsed_data.phase_groups if pg.name == "UAT")
    assert uat.start_col is not None
    assert uat.end_col is not None
    assert len(uat.pic_cols) == 2  # PIC FPT + PIC MPHG


def test_meta_columns_detected(parsed_data):
    """Meta columns quan trọng đều được detect."""
    mc = parsed_data.meta_columns
    for key in ("stt", "ma_cn", "ten_cn", "module", "priority", "complexity",
                "fit_gap", "giai_doan", "quy_trinh", "risk_blocker"):
        assert mc.get(key) is not None, f"Missing meta column: {key}"


def test_pic_case_normalization(parsed_data):
    """
    Row 6 có PIC 'SONHN6' (ALL CAPS), Row 1 có 'SonHN6' (mix case).
    Sau normalize, cả 2 phải cùng key 'SonHN6'.
    """
    all_pics = parsed_data.all_pics
    # Không được có cả 2 version
    assert "SonHN6" in all_pics
    assert "SONHN6" not in all_pics


def test_pic_multi_split(parsed_data):
    """PIC 'NhuNHT3+ HaiTD16' phải tách thành 2 PIC."""
    row_pr = next(r for r in parsed_data.rows if r.meta.get("ma_cn") == "PR.FR.03")
    dev_pics = row_pr.phases["Dev"].pics
    assert "NhuNHT3" in dev_pics
    assert "HaiTD16" in dev_pics


def test_pic_multi_comma_space(parsed_data):
    """PIC 'BaoLQ31, NhiVN' phải tách thành 2 PIC."""
    row2 = next(r for r in parsed_data.rows if r.meta.get("ma_cn") == "TMS.FR.02")
    pics = row2.phases["UAT"].pics
    assert "BaoLQ31" in pics
    assert "NhiVN" in pics
    assert "Anh Cường" in pics  # PIC MPHG cũng được thu thập


def test_status_normalization(parsed_data):
    """Status hợp lệ được giữ, số sẽ bị chuyển thành None."""
    for row in parsed_data.rows:
        for pd in row.phases.values():
            if pd.status is not None:
                assert pd.status in VALID_STATUSES


def test_date_parsing(parsed_data):
    """Date được parse thành date object."""
    for row in parsed_data.rows:
        for pd in row.phases.values():
            for d in (pd.start_date, pd.end_date):
                if d is not None:
                    assert isinstance(d, date), f"Expected date, got {type(d)}"


def test_estimate_mh_is_float(parsed_data):
    """Estimate MH là float khi có giá trị."""
    row1 = next(r for r in parsed_data.rows if r.meta.get("ma_cn") == "TMS.FR.01")
    assert row1.phases["Analysis"].estimate_mh == 8.0


def test_priorities_and_complexities(parsed_data):
    """Priorities và Complexities được thu thập unique."""
    assert "Must-have" in parsed_data.all_priorities
    assert "Should-have" in parsed_data.all_priorities
    assert "High" in parsed_data.all_complexities
    assert "Medium" in parsed_data.all_complexities


def test_giai_doan_normalized_to_string(parsed_data):
    """Giai đoạn 1 hoặc 2 (numeric) phải trở thành string."""
    for row in parsed_data.rows:
        gd = row.meta.get("giai_doan")
        if gd is not None:
            assert isinstance(gd, str)


def test_row_without_data_is_skipped(tmp_path):
    """Row hoàn toàn trống bị bỏ qua."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["STT", "Mã CN", "Tên chức năng", "Module", "Analysis - Status"])
    ws.append([1, "TMS.FR.01", "Test", "TMS", "Closed"])
    ws.append([None, None, None, None, None])  # empty
    ws.append([2, "TMS.FR.02", "Test2", "TMS", "Open"])
    p = tmp_path / "sparse.xlsx"
    wb.save(str(p))
    wb.close()

    data = FunctionListParser().parse(str(p))
    assert len(data.rows) == 2


def test_pic_blacklist_status_keywords(tmp_path):
    """
    Bug 3 regression: cột PIC có giá trị "Closed" / "In-progress" (do user paste
    lệch cột Status sang cột PIC) phải bị filter ra khỏi danh sách PIC.
    Screenshot 2: "Closed" lẫn vào dropdown PIC.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Analysis - Start", "Analysis - End", "Analysis - Status", "Analysis - PIC",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    ws.append([1, "TMS.FR.01", "Test1", "TMS",
               None, None, "Closed", "SonHN6"])
    # PIC = "Closed" (lệch cột) → phải filter
    ws.append([2, "TMS.FR.02", "Test2", "TMS",
               None, None, "In-progress", "Closed"])
    # PIC = "In-progress, HaiTD16" → chỉ giữ HaiTD16
    ws.append([3, "TMS.FR.03", "Test3", "TMS",
               None, None, "Open", "In-progress, HaiTD16"])
    # PIC = "- , Cancelled" → cả 2 đều bị bỏ (rỗng list)
    ws.append([4, "TMS.FR.04", "Test4", "TMS",
               None, None, "Assigned", "-;Cancelled"])
    p = tmp_path / "pic_blacklist.xlsx"
    wb.save(str(p))
    wb.close()

    data = FunctionListParser().parse(str(p))
    pics_set = set(data.all_pics)
    assert "Closed" not in pics_set
    assert "In-progress" not in pics_set
    assert "Cancelled" not in pics_set
    assert "SonHN6" in pics_set
    assert "HaiTD16" in pics_set
    # Row 4 không có PIC hợp lệ nào
    row4 = next(r for r in data.rows if r.meta.get("ma_cn") == "TMS.FR.04")
    assert row4.phases["Analysis"].pics == []


def test_pic_blacklisted_field_exists(parsed_data):
    """
    ParsedData phải có field `pic_blacklisted` là list.
    Sample fixture không có ô lệch cột → list rỗng.
    """
    assert hasattr(parsed_data, "pic_blacklisted")
    assert isinstance(parsed_data.pic_blacklisted, list)
    # Fixture không có cell PIC nào là status keyword → rỗng
    assert parsed_data.pic_blacklisted == []


def test_pic_blacklisted_collects_status_leaks(tmp_path):
    """
    File có cell PIC = "Closed" → phải thu thập vào pic_blacklisted với đủ
    metadata: row_index (Excel row number), phase_name, header_text,
    raw_value, matched_keyword, ma_cn, module.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Analysis - Start", "Analysis - End", "Analysis - Status", "Analysis - PIC",
        "UAT - From", "UAT - To", "UAT - Status", "UAT - PIC FPT",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    # Row 2: PIC = "Closed" (lệch cột)
    ws.append([1, "TMS.FR.01", "Test1", "TMS",
               None, None, "Open", "Closed",
               None, None, None, ""])
    # Row 3: PIC = "In-progress, HaiTD16" — giữ HaiTD16, log "In-progress"
    ws.append([2, "TMS.FR.02", "Test2", "HR",
               None, None, "Open", "In-progress, HaiTD16",
               None, None, None, ""])
    # Row 4: 2 phase cùng bị lệch — PIC Analysis = "CLOSED", PIC UAT = "Cancelled"
    ws.append([3, "PR.FR.03", "Test3", "PR",
               None, None, "Assigned", "CLOSED",
               None, None, "Open", "Cancelled"])
    p = tmp_path / "blacklist_meta.xlsx"
    wb.save(str(p))
    wb.close()

    data = FunctionListParser().parse(str(p))
    bl = data.pic_blacklisted

    # 4 entry: Row2/Closed, Row3/In-progress, Row4/CLOSED, Row4/Cancelled
    assert len(bl) == 4, f"Expected 4 blacklist entries, got {len(bl)}: {bl}"

    # Row 2 — cell PIC Analysis chứa "Closed"
    e0 = next(e for e in bl if e["row_index"] == 2)
    assert e0["phase_name"] == "Analysis"
    assert e0["header_text"] == "Analysis - PIC"
    assert e0["raw_value"] == "Closed"
    assert e0["matched_keyword"] == "Closed"
    assert e0["ma_cn"] == "TMS.FR.01"
    assert e0["module"] == "TMS"

    # Row 3 — chỉ log "In-progress" (HaiTD16 vẫn được giữ trong pd.pics)
    e1 = next(e for e in bl if e["row_index"] == 3)
    assert e1["raw_value"] == "In-progress"
    assert e1["matched_keyword"] == "In-progress"
    assert e1["ma_cn"] == "TMS.FR.02"

    # Row 4 — có 2 entry (Analysis + UAT)
    row4_entries = [e for e in bl if e["row_index"] == 4]
    assert len(row4_entries) == 2
    phases = {e["phase_name"] for e in row4_entries}
    assert phases == {"Analysis", "UAT"}
    # Case gốc được giữ trong raw_value, matched_keyword là canonical
    row4_analysis = next(e for e in row4_entries if e["phase_name"] == "Analysis")
    assert row4_analysis["raw_value"] == "CLOSED"
    assert row4_analysis["matched_keyword"] == "Closed"  # canonical
    row4_uat = next(e for e in row4_entries if e["phase_name"] == "UAT")
    assert row4_uat["raw_value"] == "Cancelled"
    assert row4_uat["header_text"] == "UAT - PIC FPT"

    # Row 3: HaiTD16 vẫn được giữ trong PIC hợp lệ
    row3 = next(r for r in data.rows if r.meta.get("ma_cn") == "TMS.FR.02")
    assert "HaiTD16" in row3.phases["Analysis"].pics
    assert "In-progress" not in row3.phases["Analysis"].pics


def test_pic_blacklisted_does_not_log_dash_or_na(tmp_path):
    """
    "-" và "n/a" là placeholder trống, KHÔNG được log vào blacklist
    (chỉ log token thật sự trùng VALID_STATUSES).
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Analysis - Status", "Analysis - PIC",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    ws.append([1, "TMS.FR.01", "Test", "TMS", "Open", "-"])
    ws.append([2, "TMS.FR.02", "Test", "TMS", "Open", "n/a"])
    ws.append([3, "TMS.FR.03", "Test", "TMS", "Open", "-, N/A, Cancelled"])
    p = tmp_path / "dash_na.xlsx"
    wb.save(str(p))
    wb.close()

    data = FunctionListParser().parse(str(p))
    bl = data.pic_blacklisted
    # Chỉ log 1 entry: "Cancelled" ở row 4 (Excel row=4 vì row 1 là header)
    assert len(bl) == 1
    assert bl[0]["raw_value"] == "Cancelled"
    assert bl[0]["row_index"] == 4


def test_giai_doan_alphanumeric(tmp_path):
    """Giai đoạn '2A' không được crash (V2 fix)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["STT", "Mã CN", "Tên chức năng", "Module", "Giai đoạn", "Analysis - Status"])
    ws.append([1, "TMS.FR.01", "Test", "TMS", "2A", "Closed"])
    ws.append([2, "TMS.FR.02", "Test2", "TMS", 2.0, "Open"])
    ws.append([3, "TMS.FR.03", "Test3", "TMS", 2.5, "Open"])
    p = tmp_path / "alphanumeric_gd.xlsx"
    wb.save(str(p))
    wb.close()

    data = FunctionListParser().parse(str(p))
    gds = [r.meta.get("giai_doan") for r in data.rows]
    assert gds == ["2A", "2", "2.5"]
