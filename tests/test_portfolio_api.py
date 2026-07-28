"""
HTTP integration tests cho 4 portfolio endpoint:
- GET  /api/portfolio/search
- POST /api/portfolio/compare
- POST /api/portfolio/compare/export
- GET  /api/portfolio/rollup

Test setup: tạo 2 project fake (alpha, beta) qua /api/projects và upload 2 file
Excel khác nhau để verify cross-project logic.
"""
import io
import os
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest


TODAY = date(2026, 7, 28)


# ==========================================================================
# Fixture: file Excel thứ 2 (project beta)
# ==========================================================================

def _make_beta_xlsx(dest_path: str) -> str:
    """Beta file: 3 function, module FIN, PIC unique (HungPV) — dùng test search."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"

    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "Quy trình",
        "Priority", "Complexity",
        "Analysis - Start", "Analysis - End", "Analysis - Status",
        "Analysis - PIC",
        "Dev - Start", "Dev - End", "Dev - Status",
        "Dev - PIC",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)

    T = TODAY
    data = [
        [1, "FIN.FR.01", "Bảng lương", "FIN", "FIN.BP.01 - Payroll",
         "Must-have", "Medium",
         T - timedelta(days=30), T - timedelta(days=25), "Closed", "HungPV",
         T - timedelta(days=20), T - timedelta(days=15), "Closed", "HungPV"],
        [2, "FIN.FR.02", "Thuế TNCN", "FIN", "FIN.BP.02 - Tax",
         "Should-have", "High",
         T - timedelta(days=10), T + timedelta(days=5), "In-progress", "HungPV",
         None, None, "Open", ""],
        [3, "FIN.FR.03", "BHXH", "FIN", "FIN.BP.03 - Insurance",
         "Could-have", "Low",
         T - timedelta(days=20), T - timedelta(days=10), "In-progress", "HungPV",
         None, None, None, ""],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, val)

    wb.save(dest_path)
    wb.close()
    return dest_path


@pytest.fixture
def beta_xlsx_path(tmp_path):
    return _make_beta_xlsx(str(tmp_path / "beta_api.xlsx"))


def _upload(client, xlsx_path, project="default"):
    """Helper: upload file .xlsx vào project cụ thể."""
    with open(xlsx_path, "rb") as f:
        data = {"file": (io.BytesIO(f.read()), "test.xlsx")}
    return client.post(
        f"/api/projects/{project}/upload",
        data=data,
        content_type="multipart/form-data",
    )


@pytest.fixture
def two_project_client(flask_client, sample_xlsx_path, beta_xlsx_path):
    """
    Setup: tạo 2 project (alpha, beta), upload 2 file khác nhau vào mỗi project.
    Return: flask_client đã có sẵn 2 project.
    """
    flask_client.post("/api/projects", json={"name": "Alpha"})
    flask_client.post("/api/projects", json={"name": "Beta"})
    _upload(flask_client, sample_xlsx_path, project="alpha")
    _upload(flask_client, beta_xlsx_path, project="beta")
    return flask_client


# ==========================================================================
# SEARCH endpoint — 6 tests
# ==========================================================================

class TestPortfolioSearchAPI:
    def test_search_missing_query_returns_empty(self, two_project_client):
        r = two_project_client.get("/api/portfolio/search")
        assert r.status_code == 200
        assert r.get_json()["total"] == 0

    def test_search_by_name_cross_project(self, two_project_client):
        # "Chấm công" chỉ trong alpha
        r = two_project_client.get("/api/portfolio/search?q=Chấm công&scope=name")
        assert r.status_code == 200
        data = r.get_json()
        assert data["total"] >= 1
        slugs = {x["project_slug"] for x in data["results"]}
        assert "alpha" in slugs

    def test_search_by_code_finds_in_beta(self, two_project_client):
        r = two_project_client.get("/api/portfolio/search?q=FIN.FR&scope=code")
        assert r.status_code == 200
        data = r.get_json()
        assert data["total"] >= 3  # 3 mã CN của beta
        assert all(x["project_slug"] == "beta" for x in data["results"])

    def test_search_by_pic_across_projects(self, two_project_client):
        # HungPV chỉ trong beta
        r = two_project_client.get("/api/portfolio/search?q=HungPV&scope=pic")
        assert r.status_code == 200
        data = r.get_json()
        assert all("HungPV" in x["pic"] for x in data["results"])

    def test_search_scope_all_default(self, two_project_client):
        # Không truyền scope → default "all"
        r = two_project_client.get("/api/portfolio/search?q=TMS")
        assert r.status_code == 200
        assert r.get_json()["scope"] == "all"

    def test_search_limit_respects_query_param(self, two_project_client):
        r = two_project_client.get("/api/portfolio/search?q=F&limit=2")
        data = r.get_json()
        assert len(data["results"]) <= 2


# ==========================================================================
# COMPARE endpoint — 5 tests
# ==========================================================================

class TestPortfolioCompareAPI:
    def test_compare_two_projects_ok(self, two_project_client):
        r = two_project_client.post(
            "/api/portfolio/compare",
            json={"slugs": ["alpha", "beta"]},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert len(data["projects"]) == 2
        assert "total_functions" in data["metrics"]

    def test_compare_returns_best_worst(self, two_project_client):
        r = two_project_client.post(
            "/api/portfolio/compare",
            json={"slugs": ["alpha", "beta"]},
        )
        data = r.get_json()
        assert "best_worst" in data
        # Alpha có 6 function > beta 3 → alpha best cho total_functions
        assert data["best_worst"]["total_functions"]["best"] == "alpha"

    def test_compare_less_than_2_returns_400(self, two_project_client):
        r = two_project_client.post(
            "/api/portfolio/compare",
            json={"slugs": ["alpha"]},
        )
        assert r.status_code == 400

    def test_compare_slugs_not_a_list_returns_400(self, two_project_client):
        r = two_project_client.post(
            "/api/portfolio/compare",
            json={"slugs": "alpha"},
        )
        assert r.status_code == 400

    def test_compare_skips_project_without_file(self, two_project_client):
        # Tạo project mới nhưng KHÔNG upload
        two_project_client.post("/api/projects", json={"name": "NoUpload"})
        r = two_project_client.post(
            "/api/portfolio/compare",
            json={"slugs": ["alpha", "noupload"]},
        )
        data = r.get_json()
        assert r.status_code == 200
        assert len(data["projects"]) == 1
        assert any(s["slug"] == "noupload" for s in data["skipped"])


# ==========================================================================
# COMPARE EXPORT endpoint — 2 tests
# ==========================================================================

class TestPortfolioCompareExportAPI:
    def test_export_returns_xlsx(self, two_project_client):
        r = two_project_client.post(
            "/api/portfolio/compare/export",
            json={"slugs": ["alpha", "beta"]},
        )
        assert r.status_code == 200
        # xlsx MIME type
        assert "spreadsheet" in r.headers["Content-Type"] or "xlsx" in r.headers.get("Content-Disposition", "")
        assert len(r.data) > 500  # File non-trivial

    def test_export_less_than_2_returns_400(self, two_project_client):
        r = two_project_client.post(
            "/api/portfolio/compare/export",
            json={"slugs": ["alpha"]},
        )
        assert r.status_code == 400


# ==========================================================================
# ROLLUP endpoint — 5 tests
# ==========================================================================

class TestPortfolioRollupAPI:
    def test_rollup_all_projects_default(self, two_project_client):
        r = two_project_client.get("/api/portfolio/rollup")
        assert r.status_code == 200
        data = r.get_json()
        # 3 project: default (chưa upload), alpha, beta → chỉ 2 có file
        assert data["projects_count"] >= 2
        assert "metrics" in data
        # total = alpha 6 + beta 3
        assert data["metrics"]["summary"]["total_functions"] == 9

    def test_rollup_per_project_returned(self, two_project_client):
        r = two_project_client.get("/api/portfolio/rollup")
        data = r.get_json()
        slugs = {p["slug"] for p in data["per_project"]}
        assert "alpha" in slugs
        assert "beta" in slugs

    def test_rollup_subset_via_slugs_param(self, two_project_client):
        r = two_project_client.get("/api/portfolio/rollup?slugs=alpha")
        assert r.status_code == 200
        data = r.get_json()
        assert data["projects_count"] == 1
        assert data["metrics"]["summary"]["total_functions"] == 6

    def test_rollup_no_project_with_file_returns_404(self, flask_client):
        # Chưa upload gì hết → default cũng rỗng
        r = flask_client.get("/api/portfolio/rollup")
        assert r.status_code == 404

    def test_rollup_summary_has_projects_count(self, two_project_client):
        r = two_project_client.get("/api/portfolio/rollup")
        data = r.get_json()
        assert data["metrics"]["summary"]["projects_count"] >= 2
        assert "projects_included" in data["metrics"]["summary"]
