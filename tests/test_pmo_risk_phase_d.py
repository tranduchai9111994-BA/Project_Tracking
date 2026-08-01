"""Unit tests — Phase D: risk scoring + PIC overload feed + module cascade."""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

from analyzer.module_dependency import (
    compute_module_cascade,
    pick_gate_phase,
)
from analyzer.pic_overload import overloaded_pics_for_data
from analyzer.risk_scorer import (
    CASCADE_DELAY_POINTS,
    PIC_OVERLOAD_POINTS,
    compute_all_risk_scores,
    compute_pmo_risk,
    compute_risk_score,
)
from parser.excel_parser import FunctionListParser, FunctionRow, PhaseData


TODAY = date(2026, 7, 28)


def _make_row(**meta):
    return FunctionRow(row_num=1, meta=meta, phases={})


def _write_fl(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


def _parse(path: Path):
    return FunctionListParser().parse(str(path))


# ------------------------------------------------------------------
# Gate / cascade
# ------------------------------------------------------------------

def test_pick_gate_prefers_config():
    assert pick_gate_phase(["Analysis", "Dev", "Config+Test", "UAT"]) == "Config+Test"
    assert pick_gate_phase(["Phân tích", "Cấu hình", "UAT"]) == "Cấu hình"


def test_pick_gate_fallback_middle_when_no_config():
    # Không Config → bỏ UAT/Golive, lấy giữa
    assert pick_gate_phase(["Analysis", "Dev", "UAT", "Golive"]) == "Dev"


def test_cascade_blocks_downstream_when_pred_config_open(tmp_path):
    """Module A Config chưa Closed → module B bị cascade."""
    path = tmp_path / "fl.xlsx"
    t0 = TODAY - timedelta(days=5)
    t1 = TODAY + timedelta(days=5)
    _write_fl(
        path,
        [
            "STT", "Mã CN", "Tên chức năng", "Module",
            "Config - Start", "Config - End", "Config - Status", "Config - PIC",
            "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
        ],
        [
            # Module A: Config còn In-progress
            [1, "A.01", "A1", "ModA", t0, t1, "In-progress", "P1", None, None, None, ""],
            [2, "A.02", "A2", "ModA", t0, t1, "Open", "P1", None, None, None, ""],
            # Module B: downstream
            [3, "B.01", "B1", "ModB", t0, t1, "Open", "P2", t0, t1, "Open", "P2"],
        ],
    )
    data = _parse(path)
    result = compute_module_cascade(
        data, ["ModA", "ModB"], today=TODAY, gate_closed_threshold=0.7,
    )
    assert result["gate_phase"] and "config" in result["gate_phase"].lower()
    assert "ModB" in result["modules_blocked"]
    assert result["blocked_by_map"]["ModB"] == "ModA"
    assert result["warning_count"] >= 1


def test_cascade_ready_when_pred_mostly_closed(tmp_path):
    path = tmp_path / "fl.xlsx"
    t0 = TODAY - timedelta(days=10)
    t1 = TODAY - timedelta(days=1)
    _write_fl(
        path,
        [
            "STT", "Mã CN", "Tên chức năng", "Module",
            "Config - Start", "Config - End", "Config - Status", "Config - PIC",
        ],
        [
            [1, "A.01", "A1", "ModA", t0, t1, "Closed", "P1"],
            [2, "A.02", "A2", "ModA", t0, t1, "Closed", "P1"],
            [3, "B.01", "B1", "ModB", t0, t1, "Open", "P2"],
        ],
    )
    data = _parse(path)
    result = compute_module_cascade(data, ["ModA", "ModB"], today=TODAY)
    assert "ModB" not in result["modules_blocked"]
    assert result["warning_count"] == 0


# ------------------------------------------------------------------
# PIC overload → risk score
# ------------------------------------------------------------------

def test_pic_overload_adds_points():
    row = _make_row()
    row.phases["Dev"] = PhaseData(
        start_date=TODAY, end_date=TODAY + timedelta(days=1),
        status="In-progress", pics=["BusyGuy"],
    )
    r = compute_risk_score(
        row, TODAY, ["Dev"],
        overloaded_pics={"BusyGuy"},
    )
    assert r["breakdown"].get("pic_overload") == PIC_OVERLOAD_POINTS
    assert any("PIC overload" in f for f in r["factors"])
    assert "BusyGuy" in r["overload_pics"]


def test_cascade_adds_points():
    row = _make_row(module="ModB")
    r = compute_risk_score(
        row, TODAY, [],
        cascade_blocked_by="ModA",
    )
    assert r["breakdown"].get("cascade_delay") == CASCADE_DELAY_POINTS
    assert any("Cascade delay" in f for f in r["factors"])
    assert r["cascade_from"] == "ModA"


def test_overloaded_pics_for_data_detects_concurrency(tmp_path):
    """6 task cùng PIC chồng lịch → overload (day_max=5)."""
    path = tmp_path / "ol.xlsx"
    t0 = TODAY - timedelta(days=2)
    t1 = TODAY + timedelta(days=2)
    rows = []
    for i in range(6):
        rows.append([
            i + 1, f"T.{i:02d}", f"Task {i}", "M",
            t0, t1, "In-progress", "OverloadGuy",
        ])
    _write_fl(
        path,
        [
            "STT", "Mã CN", "Tên chức năng", "Module",
            "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
        ],
        rows,
    )
    data = _parse(path)
    ol = overloaded_pics_for_data(data, today=TODAY)
    assert "OverloadGuy" in ol


def test_compute_all_feeds_overload_and_cascade():
    row = _make_row(ma_cn="X.1", ten_cn="X", module="ModB", priority="")
    row.phases["Dev"] = PhaseData(
        start_date=TODAY, end_date=TODAY + timedelta(days=1),
        status="In-progress", pics=["Busy"],
    )
    from parser.excel_parser import ParsedData, PhaseGroup
    data = ParsedData(
        headers={},
        meta_columns={},
        phase_groups=[PhaseGroup(name="Dev")],
        rows=[row],
        all_modules=["ModB"],
        all_phases=["Dev"],
        all_pics=["Busy"],
        all_statuses=["In-progress"],
        all_priorities=[],
        all_complexities=[],
        all_giai_doan=[],
        all_processes=[],
    )
    results = compute_all_risk_scores(
        data, TODAY,
        overloaded_pics={"Busy"},
        blocked_by_map={"ModB": "ModA"},
    )
    assert len(results) == 1
    bd = results[0]["risk_breakdown"]
    assert bd.get("pic_overload") == PIC_OVERLOAD_POINTS
    assert bd.get("cascade_delay") == CASCADE_DELAY_POINTS
    assert results[0]["risk_score"] == PIC_OVERLOAD_POINTS + CASCADE_DELAY_POINTS


def test_compute_pmo_risk_dimensions(tmp_path):
    path = tmp_path / "pmo.xlsx"
    t0 = TODAY - timedelta(days=3)
    t1 = TODAY + timedelta(days=3)
    # 6 task cùng PIC → overload; ModA Config open → ModB cascade
    rows = []
    for i in range(6):
        rows.append([
            i + 1, f"A.{i:02d}", f"A{i}", "ModA",
            t0, t1, "In-progress", "BusyGuy",
            None, None, None, "",
        ])
    rows.append([
        7, "B.01", "B1", "ModB",
        t0, t1, "Open", "Other",
        t0, t1, "Open", "Other",
    ])
    _write_fl(
        path,
        [
            "STT", "Mã CN", "Tên chức năng", "Module",
            "Config - Start", "Config - End", "Config - Status", "Config - PIC",
            "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
        ],
        rows,
    )
    data = _parse(path)
    pmo = compute_pmo_risk(
        data, today=TODAY, module_order=["ModA", "ModB"],
    )
    assert "resource" in pmo["dimensions"]
    assert "dependency" in pmo["dimensions"]
    assert pmo["dimensions"]["resource"]["overload_pic_count"] >= 1
    assert "BusyGuy" in pmo["dimensions"]["resource"]["overload_pics"]
    assert "ModB" in pmo["dimensions"]["dependency"]["modules_blocked"]
    assert pmo["summary"]["overload_pic_count"] >= 1
    assert any(m["module"] == "ModB" and m["dependency_flag"] for m in pmo["modules"])
    # Function trên ModB phải có cascade factor
    b_rows = [r for r in pmo["risk_scores"] if r["module"] == "ModB"]
    assert b_rows
    assert any("Cascade" in f for f in b_rows[0]["risk_factors"])


def test_score_still_capped_with_phase_d_factors():
    row = _make_row(priority="Must-have", complexity="High", risk_blocker="X")
    row.phases["Analysis"] = PhaseData(
        start_date=None, end_date=None, status="Closed", pics=["A"],
    )
    row.phases["Dev"] = PhaseData(
        start_date=TODAY - timedelta(days=60),
        end_date=TODAY - timedelta(days=30),
        status="In-progress",
        pics=["Busy"],
    )
    r = compute_risk_score(
        row, TODAY, ["Analysis", "Dev"],
        overloaded_pics={"Busy"},
        cascade_blocked_by="ModA",
    )
    assert r["score"] == 100
