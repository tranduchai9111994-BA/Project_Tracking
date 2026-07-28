"""
Unit tests cho `analyzer/portfolio.py`.

Fixtures:
- Tạo 2 project fake (alpha, beta) với 2 file Excel khác nhau:
  - alpha: 6 row (dùng lại sample_xlsx_path chuẩn)
  - beta: 4 row với module + phase khác một phần → verify aggregate xử lý được
    schema khác nhau.
"""
import os
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

from analyzer.portfolio import (
    search_across_projects,
    compare_projects,
    aggregate_rollup,
    rollup_summary_override,
    SEARCH_SCOPES,
    COMPARE_METRICS,
    DEFAULT_SEARCH_LIMIT,
)
from analyzer.project_manager import ProjectManager
from analyzer.dashboard_engine import DashboardEngine
from parser.excel_parser import FunctionListParser


TODAY = date(2026, 7, 28)


# ==========================================================================
# Helpers — tạo file Excel khác cho project thứ 2 (beta)
# ==========================================================================

def _make_beta_xlsx(dest_path: str) -> str:
    """
    Tạo Function List thứ 2 (project beta):
    - 4 function, 2 module mới (FIN, ATT) không trùng project alpha
    - Phase khác một phần: có Analysis (trùng) + Requirement + Development (khác)
    - Có 1 overdue để test overdue_flag cross-project
    - Có PIC "SonHN6" trùng với alpha để test search PIC across project
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"

    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "Quy trình",
        "Priority", "Complexity", "Giai đoạn",
        "Analysis - Start", "Analysis - End", "Analysis - Status",
        "Analysis - Estimate MH", "Analysis - PIC",
        "Requirement - Start", "Requirement - End", "Requirement - Status",
        "Requirement - PIC",
        "Development - Start", "Development - End", "Development - Status",
        "Development - PIC",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)

    T = TODAY

    data = [
        # 1: Closed hết ở Development (phase cuối) → 100% done
        [1, "FIN.FR.01", "Bảng lương tháng", "FIN", "FIN.BP.01 - Payroll",
         "Must-have", "Medium", "1",
         T - timedelta(days=60), T - timedelta(days=55), "Closed", 8, "SonHN6",
         T - timedelta(days=50), T - timedelta(days=40), "Closed", "HungPV",
         T - timedelta(days=30), T - timedelta(days=25), "Closed", "AnhLD"],
        # 2: In-progress Development, không overdue
        [2, "FIN.FR.02", "Quyết toán thuế", "FIN", "FIN.BP.02 - Tax",
         "Should-have", "High", "1",
         T - timedelta(days=40), T - timedelta(days=35), "Closed", 12, "SonHN6",
         T - timedelta(days=30), T - timedelta(days=15), "Closed", "HungPV",
         T - timedelta(days=10), T + timedelta(days=20), "In-progress", "AnhLD"],
        # 3: Overdue nặng ở Development
        [3, "ATT.FR.01", "Chấm công OT", "ATT", "ATT.BP.01 - Overtime",
         "Must-have", "High", "2",
         T - timedelta(days=40), T - timedelta(days=35), "Closed", 8, "SonHN6",
         T - timedelta(days=30), T - timedelta(days=25), "Closed", "HungPV",
         T - timedelta(days=20), T - timedelta(days=10), "In-progress", "AnhLD"],
        # 4: Unassigned Requirement
        [4, "ATT.FR.02", "Ca gãy", "ATT", "ATT.BP.02 - Shift",
         "Could-have", "Low", "2",
         T - timedelta(days=10), T - timedelta(days=5), "Closed", 4, "SonHN6",
         None, None, "Open", "",
         None, None, None, ""],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, val)

    wb.save(dest_path)
    wb.close()
    return dest_path


@pytest.fixture
def beta_xlsx_path(tmp_path):
    """File Excel cho project beta."""
    path = tmp_path / "beta.xlsx"
    return _make_beta_xlsx(str(path))


@pytest.fixture
def portfolio_setup(tmp_path, sample_xlsx_path, beta_xlsx_path):
    """
    Tạo 2 project (alpha, beta) trong tmp_path/projects, upload 2 file khác nhau.

    Return: (project_mgr, state_loader) — state_loader dùng dict lookup nội bộ.
    """
    projects_dir = tmp_path / "projects"
    pm = ProjectManager(str(projects_dir))
    alpha = pm.create_project("Alpha")
    beta = pm.create_project("Beta")

    # Copy file → project folder + parse + build state dict
    import shutil
    shutil.copy2(sample_xlsx_path, pm.get_current_file_path(alpha.slug))
    shutil.copy2(beta_xlsx_path, pm.get_current_file_path(beta.slug))

    parser = FunctionListParser()
    engine = DashboardEngine(today=TODAY)
    states = {}
    for slug in (alpha.slug, beta.slug):
        data = parser.parse(pm.get_current_file_path(slug))
        metrics = engine.compute_all(data)
        states[slug] = {"data": data, "metrics": metrics, "filename": f"{slug}.xlsx"}

    def loader(slug):
        return states.get(slug)

    return pm, loader, states


@pytest.fixture
def empty_portfolio(tmp_path):
    """PM với 2 project nhưng CHƯA upload file — dùng test skip logic."""
    projects_dir = tmp_path / "projects_empty"
    pm = ProjectManager(str(projects_dir))
    pm.create_project("EmptyA")
    pm.create_project("EmptyB")
    return pm, (lambda slug: None)


# ==========================================================================
# SEARCH — 12 tests
# ==========================================================================

class TestSearch:
    def test_search_empty_query_returns_empty(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = search_across_projects(pm, loader, "", "all")
        assert r["total"] == 0
        assert r["results"] == []
        assert r["projects_searched"] == 0

    def test_search_by_name_finds_function(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        # "Chấm công" xuất hiện ở alpha (Chấm công app) + beta (Chấm công OT)
        r = search_across_projects(pm, loader, "chấm công", "name")
        assert r["total"] >= 2
        matched_names = {x["ten_cn"] for x in r["results"]}
        assert "Chấm công app" in matched_names
        assert "Chấm công OT" in matched_names

    def test_search_by_name_case_insensitive(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r_lower = search_across_projects(pm, loader, "báo cáo", "name")
        r_upper = search_across_projects(pm, loader, "BÁO CÁO", "name")
        assert r_lower["total"] == r_upper["total"] >= 1

    def test_search_by_code_partial_match(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        # "TMS.FR" match nhiều function của alpha
        r = search_across_projects(pm, loader, "TMS.FR", "code")
        codes = {x["ma_cn"] for x in r["results"]}
        assert "TMS.FR.01" in codes
        assert "TMS.FR.02" in codes

    def test_search_by_pic_cross_project(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        # "SonHN6" xuất hiện cả 2 project
        r = search_across_projects(pm, loader, "SonHN6", "pic")
        slugs_hit = {x["project_slug"] for x in r["results"]}
        assert "alpha" in slugs_hit
        assert "beta" in slugs_hit

    def test_search_by_process(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = search_across_projects(pm, loader, "payroll", "process")
        assert r["total"] >= 1
        assert all("Payroll" in x["quy_trinh"] for x in r["results"])

    def test_search_scope_all_finds_any_field(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        # "TMS.FR.02" là mã CN → scope=all cũng phải tìm được
        r = search_across_projects(pm, loader, "TMS.FR.02", "all")
        assert r["total"] >= 1
        assert any(x["matched_field"] == "code" for x in r["results"])

    def test_search_no_match_returns_empty(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = search_across_projects(pm, loader, "xyzzy_not_exist_string", "all")
        assert r["total"] == 0
        assert r["results"] == []

    def test_search_limits_results(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = search_across_projects(pm, loader, "F", "all", limit=3)
        # Tất cả mã CN đều có "F" → nhiều kết quả
        assert len(r["results"]) <= 3
        if r["total"] > 3:
            assert r["truncated"] is True

    def test_search_skips_projects_without_file(self, empty_portfolio):
        pm, loader = empty_portfolio
        r = search_across_projects(pm, loader, "anything", "all")
        assert r["total"] == 0
        assert r["projects_searched"] == 0
        assert len(r["projects_skipped"]) == 2
        assert all(s["reason"] == "no_file" for s in r["projects_skipped"])

    def test_search_result_has_overdue_flag(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        # ESS.FR.10 của alpha là overdue nặng
        r = search_across_projects(pm, loader, "ESS.FR.10", "code")
        assert r["total"] >= 1
        assert any(x["overdue_flag"] for x in r["results"])

    def test_search_invalid_scope_falls_back_to_all(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = search_across_projects(pm, loader, "TMS.FR.01", "unknown_scope")
        # invalid scope → fallback all → vẫn tìm được
        assert r["total"] >= 1


# ==========================================================================
# COMPARE — 8 tests
# ==========================================================================

class TestCompare:
    def test_compare_returns_metrics_for_each_project(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "beta"])
        assert len(r["projects"]) == 2
        # Mỗi metric có value cho cả 2 project
        for key, _, _ in COMPARE_METRICS:
            assert "alpha" in r["metrics"][key]
            assert "beta" in r["metrics"][key]

    def test_compare_total_functions_correct(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "beta"])
        assert r["metrics"]["total_functions"]["alpha"] == 6
        assert r["metrics"]["total_functions"]["beta"] == 4

    def test_compare_best_worst_computed(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "beta"])
        # alpha có 6 function > beta 4 → alpha là best (higher_is_better=True)
        assert r["best_worst"]["total_functions"]["best"] == "alpha"
        assert r["best_worst"]["total_functions"]["worst"] == "beta"

    def test_compare_dedupes_slugs(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "alpha", "beta"])
        # Dedupe → chỉ 2 project
        assert len(r["projects"]) == 2

    def test_compare_skips_unknown_slug(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "not_exist"])
        assert len(r["projects"]) == 1
        assert r["projects"][0]["slug"] == "alpha"
        assert any(s["slug"] == "not_exist" and s["reason"] == "not_found" for s in r["skipped"])

    def test_compare_skips_project_without_file(self, tmp_path, sample_xlsx_path):
        # 1 project có file, 1 project chưa upload
        pm = ProjectManager(str(tmp_path / "projects"))
        pm.create_project("HasFile")
        pm.create_project("NoFile")

        import shutil
        shutil.copy2(sample_xlsx_path, pm.get_current_file_path("hasfile"))
        parser = FunctionListParser()
        engine = DashboardEngine(today=TODAY)
        data = parser.parse(pm.get_current_file_path("hasfile"))
        metrics = engine.compute_all(data)
        states = {"hasfile": {"data": data, "metrics": metrics}}

        r = compare_projects(pm, lambda s: states.get(s), ["hasfile", "nofile"])
        assert len(r["projects"]) == 1
        assert any(s["slug"] == "nofile" and s["reason"] == "no_file" for s in r["skipped"])

    def test_compare_metric_labels_included(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "beta"])
        labels = {m["key"] for m in r["metric_labels"]}
        assert "total_functions" in labels
        assert "overall_progress_pct" in labels
        assert "stalled_count" in labels

    def test_compare_last_phase_not_evaluated(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = compare_projects(pm, loader, ["alpha", "beta"])
        # last_phase_name có higher_is_better=None → không xuất hiện trong best_worst
        assert "last_phase_name" not in r["best_worst"]
        # Nhưng vẫn có value
        assert isinstance(r["metrics"]["last_phase_name"]["alpha"], str)


# ==========================================================================
# ROLLUP — 8 tests
# ==========================================================================

class TestRollup:
    def test_rollup_concatenates_rows(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        # alpha 6 + beta 4 = 10
        assert len(r["aggregated"].rows) == 10
        assert r["projects_count"] == 2

    def test_rollup_adds_virtual_project_column(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        for row in r["aggregated"].rows:
            assert "_project_slug" in row.meta
            assert "_project_name" in row.meta
            assert row.meta["_project_slug"] in ("alpha", "beta")

    def test_rollup_unions_phase_groups(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        phases = {pg.name for pg in r["aggregated"].phase_groups}
        # alpha có Analysis, Dev, UAT; beta có Analysis, Requirement, Development
        assert "Analysis" in phases   # trùng nhau
        assert "Dev" in phases        # chỉ alpha
        assert "UAT" in phases        # chỉ alpha
        assert "Requirement" in phases   # chỉ beta
        assert "Development" in phases   # chỉ beta

    def test_rollup_unions_modules(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        modules = set(r["aggregated"].all_modules)
        # Alpha: TMS, HR, PR, SYS, ESS. Beta: FIN, ATT
        assert {"TMS", "HR", "PR", "SYS", "ESS", "FIN", "ATT"} <= modules

    def test_rollup_per_project_stats(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        slugs = {p["slug"] for p in r["per_project"]}
        assert slugs == {"alpha", "beta"}
        for p in r["per_project"]:
            assert p["total"] > 0
            assert "progress_pct" in p
            assert "overdue" in p
            assert "on_time" in p
            # on_time = total - overdue
            assert p["on_time"] == max(0, p["total"] - p["overdue"])

    def test_rollup_default_uses_all_non_archived(self, tmp_path, sample_xlsx_path):
        """slugs=None → dùng all non-archived."""
        pm = ProjectManager(str(tmp_path / "projects"))
        pm.create_project("Active")
        arc = pm.create_project("Arch")
        pm.archive_project(arc.slug)

        import shutil
        for slug in ("active", "arch"):
            shutil.copy2(sample_xlsx_path, pm.get_current_file_path(slug))
        parser = FunctionListParser()
        engine = DashboardEngine(today=TODAY)
        states = {}
        for slug in ("active", "arch"):
            d = parser.parse(pm.get_current_file_path(slug))
            m = engine.compute_all(d)
            states[slug] = {"data": d, "metrics": m}

        r = aggregate_rollup(pm, lambda s: states.get(s), slugs=None)
        # Chỉ active được rollup (arch bị archived, default không include)
        slugs = {p["slug"] for p in r["per_project"]}
        assert "active" in slugs
        assert "arch" not in slugs

    def test_rollup_explicit_archived_included(self, tmp_path, sample_xlsx_path):
        """Explicit slugs → include kể cả archived."""
        pm = ProjectManager(str(tmp_path / "projects"))
        arc = pm.create_project("Arch")
        pm.archive_project(arc.slug)

        import shutil
        shutil.copy2(sample_xlsx_path, pm.get_current_file_path(arc.slug))
        parser = FunctionListParser()
        engine = DashboardEngine(today=TODAY)
        d = parser.parse(pm.get_current_file_path(arc.slug))
        m = engine.compute_all(d)
        states = {arc.slug: {"data": d, "metrics": m}}

        r = aggregate_rollup(pm, lambda s: states.get(s), slugs=[arc.slug])
        assert r["projects_count"] == 1
        assert r["per_project"][0]["slug"] == arc.slug

    def test_rollup_aggregated_runs_with_engine(self, portfolio_setup):
        """Virtual ParsedData phải chạy được DashboardEngine mà không crash."""
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        engine = DashboardEngine(today=TODAY)
        m = engine.compute_all(r["aggregated"])
        assert m["summary"]["total_functions"] == 10
        # Có ít nhất 1 module trong aggregated
        assert len(m["module_overview"]) >= 1


# ==========================================================================
# ROLLUP SUMMARY OVERRIDE — 3 tests
# ==========================================================================

class TestRollupOverride:
    def test_override_uses_weighted_avg(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        engine = DashboardEngine(today=TODAY)
        m = engine.compute_all(r["aggregated"])
        overridden = rollup_summary_override(r, m)
        # Weighted = sum(pct * total) / sum(total)
        expected_num = sum(p["progress_pct"] * p["total"] for p in r["per_project"])
        expected_den = sum(p["total"] for p in r["per_project"])
        expected = round(expected_num / expected_den, 1)
        assert overridden["summary"]["overall_progress_pct"] == expected

    def test_override_adds_projects_count(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        engine = DashboardEngine(today=TODAY)
        m = engine.compute_all(r["aggregated"])
        overridden = rollup_summary_override(r, m)
        assert overridden["summary"]["projects_count"] == 2
        assert set(overridden["summary"]["projects_included"]) == {"alpha", "beta"}

    def test_override_does_not_mutate_original(self, portfolio_setup):
        pm, loader, _ = portfolio_setup
        r = aggregate_rollup(pm, loader, ["alpha", "beta"])
        engine = DashboardEngine(today=TODAY)
        m = engine.compute_all(r["aggregated"])
        original_pct = m["summary"]["overall_progress_pct"]
        rollup_summary_override(r, m)
        # Metrics gốc không bị mutate
        assert m["summary"]["overall_progress_pct"] == original_pct
