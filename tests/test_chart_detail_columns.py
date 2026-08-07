"""
Cột meta của sheet Chi_tiet (export-chart mode=detail/both).

Bối cảnh: PM xuất «Cả hai» ở Tiến độ theo module, sheet Chi_tiet thiếu FID nên
không đối chiếu được với hệ thống iHRP, còn cột Rlog ID thì trống trơn.
Nguyên nhân Rlog trống là **file FL nguồn bỏ cột `Analysis - RlogID` từ 30/07**
(68 cột → 65 cột), không phải lỗi đọc — test dưới khoá lại kết luận đó để sau này
không ai đi sửa nhầm chỗ.
"""
from __future__ import annotations

from datetime import date

import pytest

from exporter.excel_exporter import (
    DETAIL_META_COLUMNS,
    _func_meta,
    _meta_cell_values,
)
from parser.excel_parser import FunctionRow, PhaseData


def _row(ma_cn="HR.01", fid="502", rlog=None, **meta):
    phases = {"Analysis": PhaseData(status="Closed", extra={"RlogID": rlog} if rlog else {})}
    return FunctionRow(
        row_num=2,
        meta={"ma_cn": ma_cn, "ten_cn": "Chức năng A", "module": "HR",
              "quy_trinh": "QT1", "fid": fid, **meta},
        phases=phases,
    )


# ── Bất biến quan trọng nhất: header khớp số ô ────────────────────────────

def test_so_cot_header_khop_so_o_du_lieu():
    """
    Lệch 1 phần tử là cả sheet lệch cột mà openpyxl không hề báo lỗi — dữ liệu
    vẫn ghi ra, chỉ nằm sai cột. Đây là bất biến phải canh mỗi lần thêm cột.
    """
    vals = _meta_cell_values(0, _func_meta(_row()))
    assert len(vals) == len(DETAIL_META_COLUMNS)


def test_thu_tu_cot_meta():
    names = [c[0] for c in DETAIL_META_COLUMNS]
    assert names == [
        "STT", "Mã CN", "FID", "Rlog ID", "Tên chức năng", "Module",
        "Quy trình", "Priority", "Complexity", "Mã dự án",
    ]


# ── FID ───────────────────────────────────────────────────────────────────

def test_co_cot_fid():
    assert "FID" in [c[0] for c in DETAIL_META_COLUMNS]


def test_fid_lay_dung_gia_tri():
    vals = _meta_cell_values(0, _func_meta(_row(fid="6084")))
    assert vals[DETAIL_META_COLUMNS.index(("FID", 12))] == "6084"


def test_fid_dung_ngay_sau_ma_cn():
    """FID là mã định danh → xếp cạnh Mã CN cho dễ đối chiếu, không nhét cuối."""
    names = [c[0] for c in DETAIL_META_COLUMNS]
    assert names.index("FID") == names.index("Mã CN") + 1


def test_fid_trong_thi_ra_chuoi_rong_khong_phai_none():
    """None ghi ra Excel thành ô trống nhưng làm lệch kiểu khi so sánh/sort."""
    meta = _func_meta(_row(fid=None))
    assert meta["fid"] == ""


def test_fid_so_duoc_ep_ve_chuoi_va_trim():
    """FID trong FL có thể là số (502) hoặc chuỗi có khoảng trắng."""
    assert _func_meta(_row(fid=502))["fid"] == "502"
    assert _func_meta(_row(fid="  502 "))["fid"] == "502"


# ── Rlog ID ───────────────────────────────────────────────────────────────

def test_rlog_id_doc_duoc_khi_file_co_cot():
    """FL 68 cột (≤29/07) có `Analysis - RlogID` → phải lấy được."""
    assert _func_meta(_row(rlog="25259"))["rlog_id"] == "25259"


def test_rlog_id_rong_khi_file_khong_co_cot():
    """FL 65 cột (từ 30/07) bỏ cột này → trống là đúng, không phải lỗi đọc."""
    assert _func_meta(_row(rlog=None))["rlog_id"] == ""


def test_rlog_va_fid_la_hai_cot_doc_lap():
    """Từng bị hiểu lẫn: FID có dữ liệu không có nghĩa Rlog ID cũng có."""
    meta = _func_meta(_row(fid="502", rlog=None))
    assert meta["fid"] == "502"
    assert meta["rlog_id"] == ""


# ── Sheet Chi_tiet thật ───────────────────────────────────────────────────

def _parsed(rows, *, rlog_col: bool):
    """ParsedData tối thiểu; `rlog_col` mô phỏng FL 68 cột vs 65 cột."""
    from parser.excel_parser import ParsedData, PhaseGroup

    attrs = {"Status": 5}
    if rlog_col:
        attrs["RlogID"] = 6
    return ParsedData(
        headers={"Mã CN": 1, "Tên chức năng": 2, "Module": 3, "FID": 4,
                 "Analysis - Status": 5},
        meta_columns={"ma_cn": 1, "ten_cn": 2, "module": 3, "fid": 4},
        phase_groups=[PhaseGroup(name="Analysis", attributes=attrs)],
        rows=rows,
        all_phases=["Analysis"],
        all_modules=["HR", "TMS"],
    )


@pytest.fixture
def data_with_fid():
    """
    ParsedData có cột FID nhưng KHÔNG có cột Rlog — giống FL từ 30/07.
    Fixture `parsed_data` dùng chung sinh từ file mẫu **không có cột FID**,
    nên không dùng được để kiểm cột này.
    """
    return _parsed(
        [
            _row("TMS.FR.01", fid="502"),
            _row("TMS.FR.02", fid="6084"),
            _row("HR.FR.05", fid=""),  # FL thật có ~19% dòng trống FID
        ],
        rlog_col=False,
    )


# ── Ẩn cột Rlog ID khi FL không khai cột ──────────────────────────────────

def test_an_cot_rlog_khi_file_khong_khai_cot():
    """Cột trống trơn làm PM tưởng hệ thống đọc lỗi → không bày ra."""
    from exporter.excel_exporter import _detail_meta

    cols, _vals = _detail_meta(_parsed([_row(rlog=None)], rlog_col=False))
    assert "Rlog ID" not in [c[0] for c in cols]
    assert "FID" in [c[0] for c in cols], "ẩn Rlog không được kéo FID đi theo"


def test_hien_cot_rlog_khi_file_co_khai_cot():
    from exporter.excel_exporter import _detail_meta

    cols, _vals = _detail_meta(_parsed([_row(rlog="25259")], rlog_col=True))
    assert "Rlog ID" in [c[0] for c in cols]


def test_hien_cot_rlog_khi_khai_cot_nhung_chua_dien_gia_tri():
    """
    Điều kiện là *có khai cột*, không phải *có giá trị*: file mới khai cột mà
    ẩn đi thì PM không biết còn thiếu ở đâu.
    """
    from exporter.excel_exporter import _detail_meta

    cols, _vals = _detail_meta(_parsed([_row(rlog=None)], rlog_col=True))
    assert "Rlog ID" in [c[0] for c in cols]


def test_hien_cot_rlog_khi_khong_biet_parsed_data():
    """Không có parsed_data để suy đoán → thà dư cột còn hơn mất cột."""
    from exporter.excel_exporter import _detail_meta

    cols, _vals = _detail_meta(None)
    assert "Rlog ID" in [c[0] for c in cols]


@pytest.mark.parametrize("rlog_col", [True, False])
def test_an_cot_thi_so_o_giam_theo_dung_so_header(rlog_col):
    """Bất biến then chốt của cơ chế ẩn cột: header và ô phải giảm cùng nhau."""
    from exporter.excel_exporter import _detail_meta

    cols, vals = _detail_meta(_parsed([_row()], rlog_col=rlog_col))
    assert len(vals(0, _func_meta(_row()))) == len(cols)


def test_an_cot_khong_lam_lech_gia_tri_con_lai():
    """Ô sau vị trí bị xoá phải dịch lên đúng cột, không giữ chỗ trống."""
    from exporter.excel_exporter import _detail_meta

    cols, vals = _detail_meta(_parsed([_row()], rlog_col=False))
    row = vals(0, _func_meta(_row(ma_cn="HR.09", fid="777")))
    names = [c[0] for c in cols]
    assert row[names.index("Mã CN")] == "HR.09"
    assert row[names.index("FID")] == "777"
    assert row[names.index("Tên chức năng")] == "Chức năng A"


@pytest.fixture
def detail_sheet(tmp_path, data_with_fid):
    """Xuất chart module_overview mode=both rồi trả về sheet Chi_tiet."""
    import openpyxl

    from analyzer.dashboard_engine import DashboardEngine
    from exporter.excel_exporter import export_chart

    path = export_chart(
        "module_overview",
        DashboardEngine().compute_all(data_with_fid),
        output_dir=str(tmp_path),
        mode="both",
        parsed_data=data_with_fid,
    )
    wb = openpyxl.load_workbook(path)
    assert "Chi_tiet" in wb.sheetnames
    yield wb["Chi_tiet"]
    wb.close()


def test_sheet_chi_tiet_co_header_fid(detail_sheet):
    hdr = [detail_sheet.cell(row=4, column=c).value
           for c in range(1, detail_sheet.max_column + 1)]
    assert "FID" in hdr
    # FL mẫu không khai cột Rlog → cột đó phải được ẩn
    assert "Rlog ID" not in hdr


def test_sheet_chi_tiet_ghi_fid_vao_dung_cot(detail_sheet):
    """Không chỉ có header — giá trị phải nằm đúng cột đó."""
    hdr = [detail_sheet.cell(row=4, column=c).value
           for c in range(1, detail_sheet.max_column + 1)]
    fid_col = hdr.index("FID") + 1
    macn_col = hdr.index("Mã CN") + 1
    seen = [
        (detail_sheet.cell(row=r, column=macn_col).value,
         detail_sheet.cell(row=r, column=fid_col).value)
        for r in range(5, min(detail_sheet.max_row, 12) + 1)
    ]
    assert any(fid not in (None, "") for _ma, fid in seen), (
        f"không dòng nào có FID: {seen}"
    )
