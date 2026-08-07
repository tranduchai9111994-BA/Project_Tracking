"""
Hai chế độ xuất cho các tab issue: «Danh sách lỗi» vs «FL để import».

Quy ước đã chốt:
  - File xuất chỉ chứa record CÓ VẤN ĐỀ của đúng tab đang mở. Trước đây
    /export-fl-reimport luôn union mọi loại issue (224 dòng ở MPHG) nên đứng ở
    tab Thiếu FID cũng nhận về hàng trăm dòng không liên quan.
  - `kinds` thu hẹp union; không truyền = union đầy đủ (backward compat).
  - Filter cục bộ của widget (`l_module`, `l_phase`, `l_pic`, `l_waiting_phase`,
    `fid_module`, `fid_type`) áp vào cả 2 chế độ để file khớp bảng.
  - «Danh sách lỗi» FID = 7 cột như lưới + 1 cột trống «FID cần cập nhật».
  - Upload file ít dòng bất thường → cảnh báo critical (không chặn).
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

import app as app_module
from exporter.excel_exporter import export_fid_issues_report
from parser.excel_parser import FunctionRow, PhaseData, PhaseGroup, ParsedData


JS = Path(__file__).resolve().parents[1] / "static" / "js" / "dashboard.js"
HTML = Path(__file__).resolve().parents[1] / "templates" / "index.html"


def _issue(ma: str, module: str, fid: str, kind: str) -> dict:
    return {
        "ma_cn": ma, "ten_cn": f"Func {ma}", "module": module, "quy_trinh": "",
        "fid": fid, "issue_type": kind, "dev_phase": "Dev",
        "detail": "Phase 'Dev' đã Closed nhưng FID trống",
    }


# ── Exporter: danh sách lỗi FID ────────────────────────────────────────────

def _load(path: str):
    wb = openpyxl.load_workbook(path)
    return wb, wb["Loi_FID"]


def test_danh_sach_loi_dung_7_cot_nhu_luoi_cong_cot_dien_tay(tmp_path):
    path = export_fid_issues_report(
        [_issue("HR.01", "HR", "", "missing_fid")], output_dir=str(tmp_path)
    )
    _wb, ws = _load(path)
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers == [
        "Mã CN", "Tên chức năng", "Module", "FID hiện tại",
        "Loại issue", "Dev phase", "Chi tiết", "FID cần cập nhật",
    ]


def test_cot_dien_tay_duoc_to_vang_va_de_trong(tmp_path):
    path = export_fid_issues_report(
        [_issue("HR.01", "HR", "", "missing_fid")], output_dir=str(tmp_path)
    )
    _wb, ws = _load(path)
    cell = ws.cell(row=5, column=8)
    assert cell.value in (None, "")
    assert "FFF59D" in str(cell.fill.start_color.rgb)


def test_co_autofilter_de_loc_trong_excel(tmp_path):
    path = export_fid_issues_report(
        [_issue(f"HR.{i:02d}", "HR", "", "missing_fid") for i in range(5)],
        output_dir=str(tmp_path),
    )
    _wb, ws = _load(path)
    assert ws.auto_filter.ref == "A4:H9"  # header row 4 + 5 dòng data


def test_sheet_khong_ten_function_list_de_khong_import_nham(tmp_path):
    """File danh sách lỗi không phải FL — đặt tên khác để upload nhầm bị chặn."""
    path = export_fid_issues_report(
        [_issue("HR.01", "HR", "", "missing_fid")], output_dir=str(tmp_path)
    )
    wb, _ws = _load(path)
    assert wb.sheetnames == ["Loi_FID"]
    assert "Function List" not in wb.sheetnames


def test_loai_issue_ghi_bang_tieng_viet(tmp_path):
    path = export_fid_issues_report(
        [
            _issue("HR.01", "HR", "", "missing_fid"),
            _issue("HR.02", "HR", "233", "duplicate_fid"),
        ],
        output_dir=str(tmp_path),
    )
    _wb, ws = _load(path)
    assert [ws.cell(row=r, column=5).value for r in (5, 6)] == ["Thiếu FID", "Trùng FID"]


def _data_codes(ws) -> list[str]:
    """Mã CN của các dòng data — dừng ở autofilter để bỏ dòng «Tổng: N record»."""
    last = int(ws.auto_filter.ref.split(":")[1][1:])
    return [
        ws.cell(row=r, column=1).value
        for r in range(5, last + 1)
        if ws.cell(row=r, column=1).value
    ]


def test_chi_xuat_record_co_van_de(tmp_path):
    """Không có dòng nào ngoài danh sách issue được đưa vào file."""
    issues = [_issue("HR.01", "HR", "", "missing_fid")]
    path = export_fid_issues_report(issues, output_dir=str(tmp_path))
    _wb, ws = _load(path)
    assert _data_codes(ws) == ["HR.01"]


# ── Cảnh báo tụt số dòng khi upload ───────────────────────────────────────

def test_canh_bao_khi_file_upload_it_dong_bat_thuong():
    hist = [{"row_count": 30}, {"row_count": 389, "filename": "FL.xlsx"}]
    warns = app_module._row_count_drop_warning(hist, 30)
    assert len(warns) == 1
    assert warns[0]["level"] == "critical"
    assert warns[0]["code"] == "row_count_drop"
    assert "389" in warns[0]["message"] and "30" in warns[0]["message"]
    assert warns[0]["detail"]["prev_rows"] == 389


def test_khong_canh_bao_khi_giam_nhe():
    """Giảm ít là bình thường (đóng bớt scope) — chỉ cảnh báo khi tụt > 30%."""
    assert app_module._row_count_drop_warning(
        [{"row_count": 380}, {"row_count": 389}], 380
    ) == []


def test_khong_canh_bao_lan_upload_dau_tien():
    assert app_module._row_count_drop_warning([{"row_count": 30}], 30) == []


def test_bo_qua_entry_row_count_0_khi_tim_ban_truoc():
    hist = [{"row_count": 30}, {"row_count": 0}, {"row_count": 389}]
    warns = app_module._row_count_drop_warning(hist, 30)
    assert len(warns) == 1 and warns[0]["detail"]["prev_rows"] == 389


def test_file_rong_khong_bao_trung_voi_empty_rows():
    assert app_module._row_count_drop_warning([{"row_count": 0}, {"row_count": 389}], 0) == []


# ── API: kinds scoping + local filter ─────────────────────────────────────

def _mini_data() -> ParsedData:
    """2 module: HR có FID, APP không có FID nào."""
    rows = [
        FunctionRow(
            row_num=2,
            meta={"ma_cn": "HR.01", "ten_cn": "A", "module": "HR", "fid": "F1"},
            phases={"Dev": PhaseData(status="Closed")},
        ),
        FunctionRow(
            row_num=3,
            meta={"ma_cn": "HR.02", "ten_cn": "B", "module": "HR", "fid": ""},
            phases={"Dev": PhaseData(status="Closed")},
        ),
        FunctionRow(
            row_num=4,
            meta={"ma_cn": "APP.01", "ten_cn": "C", "module": "APP", "fid": ""},
            phases={"Dev": PhaseData(status="Closed")},
        ),
    ]
    return ParsedData(
        headers={"Mã CN": 1, "Tên chức năng": 2, "Module": 3, "FID": 4,
                 "Dev - Status": 5},
        meta_columns={"ma_cn": 1, "ten_cn": 2, "module": 3, "fid": 4},
        phase_groups=[PhaseGroup(name="Dev", attributes={"Status": 5})],
        rows=rows,
        all_phases=["Dev"],
        all_modules=["APP", "HR"],
    )


@pytest.fixture
def api(flask_client):
    """flask_client từ conftest (đã login) + state giả cho project default."""
    app_module._state["default"] = {
        "data": _mini_data(), "filename": "mini.xlsx", "path": "",
    }
    return flask_client


def test_kinds_khong_hop_le_tra_400(api):
    r = api.get("/api/projects/default/export-fl-reimport?kinds=xxx")
    assert r.status_code == 400
    assert "kinds không hợp lệ" in r.get_json()["error"]


def test_export_fid_issues_chi_dong_loi_fid(api):
    r = api.get("/api/projects/default/export-fid-issues")
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.data))["Loi_FID"]
    # HR.01 có FID → không phải issue, phải bị bỏ qua
    assert set(_data_codes(ws)) == {"HR.02", "APP.01"}


def test_export_fid_issues_ton_trong_fid_module(api):
    r = api.get("/api/projects/default/export-fid-issues?fid_module=HR")
    ws = openpyxl.load_workbook(io.BytesIO(r.data))["Loi_FID"]
    assert set(_data_codes(ws)) == {"HR.02"}


def test_export_fid_issues_400_khi_filter_ra_rong(api):
    r = api.get("/api/projects/default/export-fid-issues?fid_module=KHONGCO")
    assert r.status_code == 400
    assert "phạm vi đang lọc" in r.get_json()["error"]


def test_fl_reimport_kinds_fid_chi_lay_dong_loi_fid(api):
    r = api.get("/api/projects/default/export-fl-reimport?kinds=fid")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "Function List" in wb.sheetnames  # vẫn import lại được
    ws = wb["Function List"]
    assert ws.max_row - 1 == 2  # HR.02 + APP.01


def test_fl_reimport_kinds_fid_ton_trong_fid_module(api):
    r = api.get("/api/projects/default/export-fl-reimport?kinds=fid&fid_module=HR")
    ws = openpyxl.load_workbook(io.BytesIO(r.data))["Function List"]
    assert ws.max_row - 1 == 1


def test_fl_reimport_khong_kinds_van_union_nhu_cu(api):
    """Nút «Xuất FL chỉnh sửa» ở Archive không truyền kinds → không được đổi."""
    r = api.get("/api/projects/default/export-fl-reimport?fid_issues=1")
    assert r.status_code == 200


# ── Frontend wiring ───────────────────────────────────────────────────────

def test_menu_export_co_2_nhom():
    js = JS.read_text(encoding="utf-8")
    assert "Danh sách lỗi" in js
    assert "FL để import" in js
    assert "data-fl-group" in js
    assert "data-single-list" in js


def test_openExportModePicker_nhan_opts_flkinds():
    js = JS.read_text(encoding="utf-8")
    start = js.index("function openExportModePicker")
    # Cắt tới cuối hàm chứ không cắt theo số ký tự cố định — thêm nhóm menu mới
    # là window trượt ra ngoài phần cần kiểm.
    body = js[start:js.index("window.openExportModePicker", start)]
    assert "opts" in body
    assert "flKinds" in body
    assert "exportFlReimportScoped" in body
    assert "singleList" in body


def test_export_fl_scoped_gui_kinds():
    js = JS.read_text(encoding="utf-8")
    start = js.index("async function exportFlReimportScoped")
    body = js[start:start + 1500]
    assert 'params.set("kinds", kinds)' in body
    assert "extraParams" in body


@pytest.mark.parametrize("kind", ["overdue", "unassigned", "stalled", "dq", "fid"])
def test_moi_tab_issue_deu_noi_flkinds(kind):
    html = HTML.read_text(encoding="utf-8")
    js = JS.read_text(encoding="utf-8")
    assert f"flKinds: '{kind}'" in html or f'flKinds: "{kind}"' in js, (
        f"tab {kind} chưa có nhóm «FL để import»"
    )


def test_local_filter_forward_cho_overdue_va_stalled():
    js = JS.read_text(encoding="utf-8")
    assert "function _overdueExportLocalParams" in js
    assert "function _stalledExportLocalParams" in js
    assert 'params.set("l_module"' in js
    assert 'params.set("l_waiting_phase"' in js
    assert 'params.set("l_phase"' in js


def test_narrow_xu_ly_field_list():
    """`pic` trong overdue_list là list → filter phải khớp giao, không so string."""
    src = Path(app_module.__file__).read_text(encoding="utf-8")
    start = src.index("def _narrow(items: list, field: str")
    body = src[start:start + 700]
    assert "isinstance(val, (list, tuple, set))" in body
    assert "any(" in body


def test_banner_critical_thay_toast():
    """Toast tự tắt sau 3.5s → cảnh báo mất dòng phải là banner cố định."""
    js = JS.read_text(encoding="utf-8")
    html = HTML.read_text(encoding="utf-8")
    assert 'id="uploadCriticalWarn"' in html
    start = js.index("function _showUploadWarnings")
    body = js[start:start + 1800]
    assert 'w.level === "critical"' in body
    assert "uploadCriticalWarn" in body
    assert "_dismissUploadWarning" in js
