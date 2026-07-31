"""Tests cho PIC Overload đa dự án."""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

from analyzer.pic_overload import (
    DEFAULT_DAY_MAX_TASKS,
    compute_pic_overload,
    default_thresholds,
    load_overload_settings,
    merge_thresholds,
    period_key,
    save_overload_settings,
)
from analyzer.project_manager import ProjectManager
from parser.excel_parser import FunctionListParser


TODAY = date(2026, 7, 28)


def _write_fl(path: Path, rows: list[list], headers: list[str] | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    hdrs = headers or [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
        "Config - Start", "Config - End", "Config - Status", "Config - PIC",
    ]
    for i, h in enumerate(hdrs, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


@pytest.fixture
def two_proj_mgr(tmp_path):
    """2 project, cùng PIC OverloadGuy với 3+ task chồng lịch."""
    root = tmp_path / "projects"
    root.mkdir()
    mgr = ProjectManager(str(root))
    a = mgr.create_project("Alpha")
    b = mgr.create_project("Beta")

    t0 = TODAY - timedelta(days=3)
    t1 = TODAY + timedelta(days=3)

    # Alpha: 3 Dev tasks cùng PIC trong khoảng t0–t1
    _write_fl(
        Path(mgr.get_project_folder(a.slug)) / "current.xlsx",
        [
            [1, "A.01", "Task A1", "TMS", t0, t1, "In-progress", "OverloadGuy", None, None, None, ""],
            [2, "A.02", "Task A2", "TMS", t0, t1, "Assigned", "OverloadGuy", None, None, None, ""],
            [3, "A.03", "Task A3", "HR", t0, t1, "Open", "OverloadGuy; Helper", None, None, None, ""],
        ],
    )
    # Beta: thêm 3 task nữa → tổng 6 concurrent > default 5
    _write_fl(
        Path(mgr.get_project_folder(b.slug)) / "current.xlsx",
        [
            [1, "B.01", "Task B1", "FIN", t0, t1, "In-progress", "OverloadGuy", None, None, None, ""],
            [2, "B.02", "Task B2", "FIN", t0, t1, "In-progress", "OverloadGuy", None, None, None, ""],
            [3, "B.03", "Task B3", "FIN", t0, t1, "Pending", "OverloadGuy",
             t0, t1, "In-progress", "OtherPIC"],
        ],
    )
    return mgr


def _loader_for(mgr: ProjectManager):
    cache = {}

    def load(slug: str):
        if slug in cache:
            return cache[slug]
        path = Path(mgr.get_project_folder(slug)) / "current.xlsx"
        if not path.is_file():
            return None
        data = FunctionListParser().parse(str(path))
        cache[slug] = {"data": data, "filename": path.name}
        return cache[slug]

    return load


class TestThresholds:
    def test_defaults(self):
        t = default_thresholds()
        assert t["day_max_tasks"] == DEFAULT_DAY_MAX_TASKS == 5
        assert t["week_min_overload_days"] == 2

    def test_merge_and_persist(self, tmp_path):
        folder = str(tmp_path / "projects")
        Path(folder).mkdir()
        saved = save_overload_settings(folder, {"day_max_tasks": 3, "phase_keywords": "Dev,Config"})
        assert saved["day_max_tasks"] == 3
        assert saved["phase_keywords"] == ["Dev", "Config"]
        loaded = load_overload_settings(folder)
        assert loaded["day_max_tasks"] == 3


class TestCompute:
    def test_day_overload_cross_project(self, two_proj_mgr):
        result = compute_pic_overload(
            two_proj_mgr,
            _loader_for(two_proj_mgr),
            grain="day",
            date_from=(TODAY - timedelta(days=5)).isoformat(),
            date_to=(TODAY + timedelta(days=5)).isoformat(),
            thresholds={"day_max_tasks": 5},
            today=TODAY,
        )
        assert result["summary"]["projects_scanned"] == 2
        pics = {p["pic"]: p for p in result["by_pic"]}
        assert "OverloadGuy" in pics
        assert pics["OverloadGuy"]["is_overload"] is True
        assert pics["OverloadGuy"]["max_concurrent"] >= 6
        assert len(result["summary"]["highlight_dates"]) > 0
        # Detail có cả 2 project
        slugs = {d["project_slug"] for d in result["detail"] if d["pic"] == "OverloadGuy"}
        assert len(slugs) >= 2

    def test_closed_excluded(self, tmp_path):
        root = tmp_path / "projects"
        root.mkdir()
        mgr = ProjectManager(str(root))
        p = mgr.create_project("OnlyClosed")
        t0 = TODAY - timedelta(days=2)
        t1 = TODAY + timedelta(days=2)
        _write_fl(
            Path(mgr.get_project_folder(p.slug)) / "current.xlsx",
            [
                [1, "C.01", "Done", "TMS", t0, t1, "Closed", "BusyPIC", None, None, None, ""],
                [2, "C.02", "Done2", "TMS", t0, t1, "Cancelled", "BusyPIC", None, None, None, ""],
            ],
        )
        result = compute_pic_overload(
            mgr, _loader_for(mgr), grain="day",
            date_from=t0.isoformat(), date_to=t1.isoformat(),
            today=TODAY, thresholds={"day_max_tasks": 1},
        )
        assert all(p["pic"] != "BusyPIC" for p in result["by_pic"]) or \
            not any(p["is_overload"] for p in result["by_pic"] if p["pic"] == "BusyPIC")
        # BusyPIC không có active → không vào by_pic
        assert "BusyPIC" not in {p["pic"] for p in result["by_pic"]}

    def test_phase_keyword_filter(self, two_proj_mgr):
        result = compute_pic_overload(
            two_proj_mgr, _loader_for(two_proj_mgr),
            grain="day",
            date_from=(TODAY - timedelta(days=5)).isoformat(),
            date_to=(TODAY + timedelta(days=5)).isoformat(),
            thresholds={"day_max_tasks": 1, "phase_keywords": ["Config"]},
            today=TODAY,
        )
        # Chỉ Config phase của B.03 → OtherPIC có 1 task; OverloadGuy không trên Config
        pics = {p["pic"] for p in result["by_pic"]}
        assert "OtherPIC" in pics
        assert "OverloadGuy" not in pics

    def test_week_compare_present(self, two_proj_mgr):
        result = compute_pic_overload(
            two_proj_mgr, _loader_for(two_proj_mgr),
            grain="week",
            date_from=(TODAY - timedelta(days=14)).isoformat(),
            date_to=(TODAY + timedelta(days=7)).isoformat(),
            thresholds={"day_max_tasks": 5, "week_min_overload_days": 1},
            today=TODAY,
        )
        assert result["summary"]["week_compare"] is not None
        assert "current_week" in result["summary"]["week_compare"]

    def test_period_key(self):
        assert period_key(date(2026, 7, 28), "day") == "2026-07-28"
        assert period_key(date(2026, 7, 28), "month") == "2026-07"
        assert "W" in period_key(date(2026, 7, 28), "week")


class TestApi:
    def test_settings_and_api(self, flask_client, sample_xlsx_path):
        import io
        with open(sample_xlsx_path, "rb") as f:
            flask_client.post(
                "/api/projects/default/upload",
                data={"file": (io.BytesIO(f.read()), "sample.xlsx")},
                content_type="multipart/form-data",
            )
        r = flask_client.get("/api/pic-overload/settings")
        assert r.status_code == 200
        assert r.get_json()["settings"]["day_max_tasks"] == 5

        r2 = flask_client.put(
            "/api/pic-overload/settings",
            json={"day_max_tasks": 2},
        )
        assert r2.status_code == 200
        assert r2.get_json()["settings"]["day_max_tasks"] == 2

        r3 = flask_client.get(
            "/api/pic-overload?grain=day&from=2026-07-01&to=2026-08-15&day_max_tasks=2"
        )
        assert r3.status_code == 200
        body = r3.get_json()
        assert body["success"] is True
        assert body["grain"] == "day"
        assert "by_pic" in body
        assert "calendar" in body

        r4 = flask_client.post(
            "/api/pic-overload/export",
            json={"grain": "day", "mode": "both", "from": "2026-07-01", "to": "2026-08-15"},
        )
        assert r4.status_code == 200
        assert "spreadsheet" in (r4.content_type or "") or r4.data[:2] == b"PK"
