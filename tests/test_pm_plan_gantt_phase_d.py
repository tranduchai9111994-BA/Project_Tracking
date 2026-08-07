# -*- coding: utf-8 -*-
"""Tests PM Gantt Phase D — day detail grid."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from analyzer.pm_plan_gantt import (
    build_pm_day_grid,
    build_pm_gantt_view,
    compute_pm_day_axis,
)
from parser.pm_plan_parser import parse_plan

ROOT = Path(__file__).resolve().parents[1]
DASHBOARD_JS = (ROOT / "static" / "js" / "dashboard.js").read_text(encoding="utf-8")
INDEX_HTML = (ROOT / "templates" / "index.html").read_text(encoding="utf-8")


def test_parse_day_columns_from_detail_header(tmp_path: Path):
    path = tmp_path / "plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail"
    ws["A6"] = ""
    ws["M6"] = "2026-03-02"
    ws["N6"] = "2026-03-03"
    ws["O6"] = "2026-03-04"
    ws["A7"] = "STT"
    ws["B7"] = "Công việc"
    ws["F7"] = "Bắt đầu"
    ws["G7"] = "Kết thúc"
    ws["M7"] = "M"
    ws["N7"] = "T"
    ws["O7"] = "W"
    ws["A8"] = 1
    ws["B8"] = "Task span"
    ws["F8"] = "2026-03-02"
    ws["G8"] = "2026-03-03"
    wb.save(path)

    out = parse_plan(str(path), {"Detail": "schedule"})
    assert len(out.get("day_columns") or []) == 3
    axis = compute_pm_day_axis(out)
    assert len(axis) == 3
    grid = build_pm_day_grid(out, axis, today=date(2026, 8, 6))
    tasks = [r for r in grid["rows"] if r["kind"] == "task"]
    assert len(tasks[0]["day_active"]) == 2


def test_gantt_view_includes_day_bundle():
    plan = {
        "schedule": [
            {
                "name": "T1",
                "start": "2026-03-02",
                "end": "2026-03-02",
                "is_phase_header": False,
            },
        ],
        "day_columns": [
            {"label": "M", "date": "2026-03-02"},
            {"label": "T", "date": "2026-03-03"},
        ],
    }
    view = build_pm_gantt_view(plan)
    assert view and view.get("day_axis")
    assert view.get("day_grid", {}).get("rows")


class TestStaticPhaseD:
    def test_html_detail_view(self):
        assert 'id="pmGanttViewDetail"' in INDEX_HTML
        assert 'id="pmPlanDetailWrap"' in INDEX_HTML

    def test_js_detail_renderer(self):
        assert "function renderPmPlanDetailGrid(" in DASHBOARD_JS
        assert "day_axis" in DASHBOARD_JS
        assert "viewMode === \"detail\"" in DASHBOARD_JS
