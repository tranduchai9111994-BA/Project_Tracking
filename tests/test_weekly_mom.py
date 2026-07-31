"""Tests cho export_weekly_mom — date swap, week window, smoke workbook."""
import os
from datetime import date, timedelta
from types import SimpleNamespace

import openpyxl

from exporter.weekly_mom import (
    export_weekly_mom,
    _collect_week_plan,
    _fmt_date,
    _iso_week_label,
    _normalize_date_pair,
    _parse_date,
    _ranges_overlap,
)

TODAY = date(2026, 7, 28)


def test_parse_and_fmt_date_variants():
    assert _parse_date(date(2026, 7, 31)) == date(2026, 7, 31)
    assert _parse_date("2026-07-31") == date(2026, 7, 31)
    assert _parse_date("31/07/2026") == date(2026, 7, 31)
    assert _fmt_date("2026-08-31") == "31/08/2026"
    assert _fmt_date(None) == ""
    assert _fmt_date("") == ""


def test_normalize_date_pair_swaps_when_start_after_end():
    s, e, swapped = _normalize_date_pair(date(2026, 8, 31), date(2026, 7, 31))
    assert swapped is True
    assert s == date(2026, 7, 31)
    assert e == date(2026, 8, 31)

    s2, e2, swapped2 = _normalize_date_pair("31/08/2026", "31/07/2026")
    assert swapped2 is True
    assert s2 == date(2026, 7, 31)
    assert e2 == date(2026, 8, 31)

    s3, e3, swapped3 = _normalize_date_pair(date(2026, 7, 1), date(2026, 7, 31))
    assert swapped3 is False
    assert s3 == date(2026, 7, 1)
    assert e3 == date(2026, 7, 31)


def test_ranges_overlap_point_and_span():
    ws, we = date(2026, 7, 27), date(2026, 8, 2)
    assert _ranges_overlap(date(2026, 7, 31), date(2026, 7, 31), ws, we)
    assert _ranges_overlap(date(2026, 7, 1), date(2026, 8, 15), ws, we)
    assert not _ranges_overlap(date(2026, 8, 10), date(2026, 8, 20), ws, we)
    assert _ranges_overlap(None, date(2026, 7, 28), ws, we)
    assert _ranges_overlap(date(2026, 7, 28), None, ws, we)


def _fake_parsed(rows):
    return SimpleNamespace(rows=rows, all_modules=[], all_phases=[])


def _phase(**kwargs):
    return SimpleNamespace(
        start_date=kwargs.get("start"),
        end_date=kwargs.get("end"),
        status=kwargs.get("status", "In-progress"),
        pics=kwargs.get("pics", ["NhiVN"]),
        note=kwargs.get("note", ""),
    )


def test_collect_week_plan_includes_overlap_and_swaps_dates():
    """End/Start trong tuần + swap Start>End; không lấy overlap-only dài."""
    week_start, week_end = date(2026, 7, 27), date(2026, 8, 2)
    row = SimpleNamespace(
        meta={"ma_cn": "PR.FR.40", "ten_cn": "HSĐC", "module": "PR"},
        phases={
            # End đúng trong tuần
            "Config Local": _phase(start=date(2026, 7, 31), end=date(2026, 7, 31)),
            # Start > End (bug FL) — End 31/07 trong tuần sau swap
            "Config UAT": _phase(start=date(2026, 8, 31), end=date(2026, 7, 31)),
            # Overlap dài (không end/start trong tuần) — bỏ
            "Analysis": _phase(start=date(2026, 4, 1), end=date(2026, 8, 8)),
            # Ngoài tuần
            "Dev": _phase(start=date(2026, 9, 1), end=date(2026, 9, 10)),
            # Closed — bỏ
            "Document": _phase(start=date(2026, 7, 28), end=date(2026, 7, 29), status="Closed"),
        },
    )
    items = _collect_week_plan(_fake_parsed([row]), week_start, week_end, limit=20)
    names = [it["ten"] for it in items]
    assert any("Config Local" in n for n in names)
    assert any("Config UAT" in n for n in names)
    assert not any("Analysis" in n for n in names)
    assert not any("Dev" in n for n in names)
    assert not any("Document" in n for n in names)
    for it in items:
        if it["from"] and it["to"]:
            d0 = _parse_date(it["from"])
            d1 = _parse_date(it["to"])
            assert d0 <= d1, it
    uat = next(it for it in items if "Config UAT" in it["ten"])
    assert uat["from"] == "31/07/2026"
    assert uat["to"] == "31/08/2026"
    assert "swap" in (uat.get("note") or "").lower()
    assert uat["pic"] == "NhiVN"


def test_collect_week_plan_next_week_not_empty_when_deadlines_exist():
    week_start = date(2026, 8, 3)
    week_end = date(2026, 8, 9)
    row = SimpleNamespace(
        meta={"ma_cn": "SI.FR.01", "ten_cn": "BHXH", "module": "SI"},
        phases={
            "Analysis": _phase(start=date(2026, 8, 3), end=date(2026, 8, 3), status="Assigned"),
            "Dev": _phase(start=date(2026, 8, 5), end=date(2026, 8, 5), status=""),
        },
    )
    items = _collect_week_plan(_fake_parsed([row]), week_start, week_end)
    assert len(items) == 2


def test_export_weekly_mom_sheets_and_headers(tmp_path, metrics, parsed_data):
    """Workbook có Cover / Master / Gantt / MoM / PM Dashboard."""
    path = export_weekly_mom(
        metrics,
        str(tmp_path),
        project_code="KDG_iHRP_2026_PM",
        parsed_data=parsed_data,
        today=TODAY,
    )
    assert os.path.exists(path)
    week = _iso_week_label(TODAY)
    assert week in os.path.basename(path)

    wb = openpyxl.load_workbook(path)
    names = wb.sheetnames
    assert names[0] == "Cover Page"
    assert "Master plan" in names
    assert "Gantt" in names
    mom_name = f"MoM_{week}"
    assert mom_name in names
    assert "PM Dashboard" in names

    cover = wb["Cover Page"]
    assert cover["C4"].value == "KDG_iHRP_2026_PM"
    assert week in str(cover["F5"].value)
    # TOC khớp thứ tự sheet: Master → Gantt → MoM → PM Dashboard
    assert cover["B9"].value == "Master plan"
    assert cover["B10"].value == "Gantt"
    assert cover["B11"].value == mom_name
    assert cover["B12"].value == "PM Dashboard"
    # Presentation: header TOC xanh #0070C0
    assert "0070C0" in str(cover["B8"].fill.fgColor.rgb or "").upper()

    mp = wb["Master plan"]
    assert mp["B2"].value == "STT"
    assert mp["C2"].value == "Công việc"
    # Có dữ liệu (không còn N/A placeholder duy nhất)
    assert mp["C4"].value is not None
    assert "N/A — Master plan (WBS" not in str(mp["C4"].value)
    # LU có ngày; v1.0 để trống (baseline)
    assert mp["I4"].value or mp["J4"].value
    assert mp["G4"].value in (None, "")
    assert mp["H4"].value in (None, "")
    assert "0070C0" in str(mp["B2"].fill.fgColor.rgb or "").upper()
    assert mp.column_dimensions["C"].width >= 40

    gantt = wb["Gantt"]
    assert "GANTT" in str(gantt["A1"].value).upper()
    assert gantt["B2"].value  # legend
    # Cột tuần đủ rộng để đọc Wxx
    assert (gantt.column_dimensions["E"].width or 0) >= 6.5

    mom = wb[mom_name]
    assert "Ngày" in str(mom["B2"].value)
    assert mom["B8"].value == "STT"
    assert mom["C8"].value == "Công việc"
    assert mom["H8"].value == "Tình trạng"
    assert mom["B9"].value == "A"
    assert "KẾ HOẠCH TUẦN NÀY" in str(mom["C9"].value)
    assert "0070C0" in str(mom["B8"].fill.fgColor.rgb or "").upper()
    # Wrap tên CN
    assert mom.column_dimensions["C"].width >= 40
    assert mom["C10"].alignment.wrap_text is True or mom["C9"].alignment.wrap_text is True


    # Section B tuần tới tồn tại
    found_b = False
    for r in range(9, 40):
        if mom.cell(r, 2).value == "B":
            found_b = True
            assert "TUẦN TỚI" in str(mom.cell(r, 3).value)
            break
    assert found_b

    # From <= To trên mọi dòng kế hoạch có đủ 2 ngày
    for r in range(10, 60):
        f = mom.cell(r, 6).value
        t = mom.cell(r, 7).value
        if f and t and "/" in str(f) and "/" in str(t):
            d0 = _parse_date(f)
            d1 = _parse_date(t)
            if d0 and d1:
                assert d0 <= d1, (r, f, t)

    dash = wb["PM Dashboard"]
    a_vals = [str(c.value or "") for c in dash["A"] if c.value]
    joined = " | ".join(a_vals)
    assert "TÓM TẮT" in joined
    assert "TRỄ" in joined or "OVERDUE" in joined
    assert "PHASE" in joined.upper() or "TIẾN ĐỘ" in joined
    assert "MODULE" in joined.upper()
    assert "PIC" in joined
    # Có ít nhất 2 chart; pie data không gồm dòng title trống
    assert len(dash._charts) >= 2
    pie = next((c for c in dash._charts if type(c).__name__ == "PieChart"), None)
    if pie is not None and pie.series:
        val_f = pie.series[0].val.numRef.f
        # Range dạng $B$71:$B$72 — không gồm title row
        assert "$B$" in val_f
        # Không phải 3-cell spanning title
        import re
        m = re.search(r"\$B\$(\d+):\$B\$(\d+)", val_f)
        assert m, val_f
        assert int(m.group(2)) - int(m.group(1)) == 1, val_f

    wb.close()


def test_export_weekly_mom_empty_metrics_still_creates(tmp_path):
    path = export_weekly_mom({}, str(tmp_path), project_code="demo", today=TODAY)
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert "PM Dashboard" in wb.sheetnames
    assert "Gantt" in wb.sheetnames
    wb.close()


def test_export_mom_week_sections_use_overlap(tmp_path):
    """Smoke: plan tuần này chứa End/Start trong tuần; tuần tới chứa deadline week+1."""
    today = date(2026, 7, 31)
    week_start = today - timedelta(days=today.weekday())
    next_start = week_start + timedelta(days=7)

    row = SimpleNamespace(
        meta={"ma_cn": "X.1", "ten_cn": "Demo", "module": "M"},
        phases={
            "Analysis": _phase(start=week_start, end=week_start + timedelta(days=2)),
            "Dev": _phase(start=next_start, end=next_start),
            # Overlap dài không Start/End trong tuần — không vào MoM
            "UAT": _phase(start=week_start - timedelta(days=60), end=next_start + timedelta(days=60)),
        },
    )
    parsed = _fake_parsed([row])
    metrics = {
        "summary": {"total_functions": 1, "overall_progress_pct": 10},
        "overdue_list": [],
        "phase_progress_stacked": {
            "phases": ["Analysis", "Dev", "UAT"],
            "statuses": ["Closed", "In-progress", "(Blank)"],
            "data": {
                "Analysis": {"Closed": 0, "In-progress": 1, "(Blank)": 0},
                "Dev": {"Closed": 0, "In-progress": 1, "(Blank)": 0},
                "UAT": {"Closed": 0, "In-progress": 0, "(Blank)": 1},
            },
        },
        "module_overview": [
            {"stt": 1, "module": "M", "total": 1, "progress_pct": 0, "active_phase": "Analysis", "overdue_count": 0},
        ],
        "pic_workload": [{"pic": "NhiVN", "total_tasks": 2, "closed": 0, "in_progress": 2, "assigned": 0, "overdue": 0}],
        "timeline_data": {
            "modules": ["M"],
            "phases": ["Analysis", "Dev", "UAT"],
            "data": {
                "M": {
                    "Analysis": {
                        "start": week_start.isoformat(),
                        "end": (week_start + timedelta(days=2)).isoformat(),
                        "total": 1, "closed": 0, "pct_closed": 0,
                    },
                    "Dev": {
                        "start": next_start.isoformat(),
                        "end": next_start.isoformat(),
                        "total": 1, "closed": 0, "pct_closed": 0,
                    },
                    "UAT": {
                        "start": (week_start - timedelta(days=60)).isoformat(),
                        "end": (next_start + timedelta(days=60)).isoformat(),
                        "total": 1, "closed": 0, "pct_closed": 0,
                    },
                },
            },
        },
    }
    path = export_weekly_mom(metrics, str(tmp_path), project_code="demo", parsed_data=parsed, today=today)
    wb = openpyxl.load_workbook(path)
    mom = wb[f"MoM_{_iso_week_label(today)}"]
    texts = [str(mom.cell(r, 3).value or "") for r in range(1, 50)]
    assert any("Analysis" in t for t in texts)
    assert any("Dev" in t for t in texts)
    assert not any("[UAT]" in t for t in texts)
    # Master có module M; LU có ngày, v1.0 trống
    mp = wb["Master plan"]
    mp_texts = [str(mp.cell(r, 3).value or "") for r in range(1, 30)]
    assert any(t.strip() == "M" for t in mp_texts)
    # Gantt không hiện hàng ngoài cửa sổ-only empty — UAT dài vẫn giao cửa sổ nên có thể hiện
    gantt = wb["Gantt"]
    assert "GANTT" in str(gantt["A1"].value).upper()

    dash = wb["PM Dashboard"]
    # Phase % loại Blank: UAT Closed=0, total status=0 → 0%; Analysis 0/1=0
    # Tìm dòng UAT trong block phase
    uat_row = None
    for r in range(1, 80):
        if dash.cell(r, 1).value == "UAT":
            uat_row = r
            break
    assert uat_row is not None
    # total có status = 0 (chỉ Blank) → % = 0, cột D = 0
    assert dash.cell(uat_row, 4).value == 0

    mod_chart = next(
        (c for c in dash._charts if c.title and "Module" in str(
            "".join(
                (rr.t or "")
                for p in (c.title.tx.rich.p or [])
                for rr in (p.r or [])
            ) if c.title.tx and c.title.tx.rich else ""
        )),
        None,
    )
    # Fallback: series trỏ cột D (tiến độ %)
    found_progress_chart = False
    for c in dash._charts:
        if not c.series:
            continue
        val_f = getattr(getattr(c.series[0].val, "numRef", None), "f", "") or ""
        if "$D$" in val_f:
            found_progress_chart = True
            break
    assert found_progress_chart, "Module chart phải gắn cột Tiến độ % (D)"
    wb.close()


def test_pm_dashboard_phase_pct_excludes_blank(tmp_path):
    """% Closed = Closed / tổng status ≠ (Blank); pie range 2 ô số."""
    today = date(2026, 7, 31)
    metrics = {
        "summary": {
            "total_functions": 10,
            "overall_progress_pct": 40.0,
            "total_overdue": 0,
            "total_overdue_records": 0,
            "unassigned_count": 0,
            "high_risk_count": 0,
            "modules_count": 1,
            "phases_count": 1,
        },
        "overdue_list": [],
        "phase_progress_stacked": {
            "phases": ["Analysis"],
            "statuses": ["Closed", "In-progress", "(Blank)"],
            "data": {"Analysis": {"Closed": 4, "In-progress": 1, "(Blank)": 5}},
        },
        "module_overview": [
            {"module": "PR", "total": 10, "progress_pct": 40, "active_phase": "Analysis", "overdue_count": 0},
        ],
        "pic_workload": [
            {"pic": "A", "total_tasks": 5, "closed": 2, "in_progress": 1, "assigned": 0, "overdue": 0},
        ],
        "timeline_data": {"modules": [], "phases": [], "data": {}},
    }
    path = export_weekly_mom(metrics, str(tmp_path), project_code="t", today=today)
    wb = openpyxl.load_workbook(path)
    dash = wb["PM Dashboard"]
    analysis_row = None
    for r in range(1, 40):
        if dash.cell(r, 1).value == "Analysis":
            analysis_row = r
            break
    assert analysis_row is not None
    assert dash.cell(analysis_row, 2).value == 80.0  # 4/5
    assert dash.cell(analysis_row, 3).value == 4
    assert dash.cell(analysis_row, 4).value == 5
    pie = next(c for c in dash._charts if type(c).__name__ == "PieChart")
    val_f = pie.series[0].val.numRef.f
    import re
    m = re.search(r"\$B\$(\d+):\$B\$(\d+)", val_f)
    assert m and int(m.group(2)) - int(m.group(1)) == 1
    wb.close()
