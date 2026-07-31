"""Tests FL re-import: date chain (bỏ T7/CN), schema, yellow/blue highlight, union."""
from __future__ import annotations

import json
import os
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

from exporter.fl_export_schema import (
    build_schema_from_parsed,
    next_business_day,
    review_from_xlsx,
    save_fl_export_template,
    schema_from_xlsx,
)
from exporter.fl_reimport_export import (
    BLUE_FILL,
    YELLOW_FILL,
    collect_issue_hits,
    compute_date_chain_fills,
    export_fl_reimport,
    format_fl_date,
)
from parser.excel_parser import FunctionListParser, FunctionRow, PhaseData


# ------------------------------------------------------------------
# next_business_day / date chain
# ------------------------------------------------------------------

def test_next_business_day_skips_weekend():
    # Friday → Monday
    fri = date(2026, 7, 31)  # Friday
    assert fri.weekday() == 4
    assert next_business_day(fri) == date(2026, 8, 3)  # Monday

    # Saturday → Monday
    sat = date(2026, 8, 1)
    assert next_business_day(sat) == date(2026, 8, 3)

    # Thursday → Friday
    thu = date(2026, 7, 30)
    assert next_business_day(thu) == date(2026, 7, 31)


def test_compute_date_chain_fills_prev_end_plus_one(parsed_data):
    # Row HR.FR.05: Analysis Closed có End, Dev Start trống
    row = next(r for r in parsed_data.rows if r.meta.get("ma_cn") == "HR.FR.05")
    phases = list(parsed_data.all_phases)
    fills = compute_date_chain_fills(row, phases)
    assert "Dev" in fills
    prev_end = row.phases["Analysis"].end_date
    assert fills["Dev"] == next_business_day(prev_end)
    # Không đè UAT nếu Dev chưa có End (Dev có End trong sample → có thể fill UAT)
    # Dev Start được fill; Dev đã có End → UAT From có thể fill nếu trống
    if row.phases.get("UAT") and not row.phases["UAT"].start_date and row.phases["Dev"].end_date:
        assert "UAT" in fills


def test_date_chain_does_not_overwrite_existing_start():
    row = FunctionRow(row_num=1, meta={"ma_cn": "X.1"})
    row.phases["A"] = PhaseData(end_date=date(2026, 7, 28), status="Closed")
    row.phases["B"] = PhaseData(start_date=date(2026, 8, 1), status="Open")  # đã có
    fills = compute_date_chain_fills(row, ["A", "B"])
    assert "B" not in fills


# ------------------------------------------------------------------
# collect_issue_hits union
# ------------------------------------------------------------------

def test_collect_issue_hits_union_by_ma_cn():
    hits = collect_issue_hits(
        overdue_list=[{"ma_cn": "A.1", "phase": "UAT", "days_overdue": 5}],
        unassigned_list=[{"ma_cn": "A.1", "phase": "Dev"}, {"ma_cn": "B.2", "phase": "Dev"}],
        stalled_list=[{"ma_cn": "B.2", "completed_phase": "Analysis", "waiting_phase": "Dev", "wait_days": 3}],
        anomaly_issues=[{"ma_cn": "C.3", "phase": "Dev", "code": "end_before_start", "label": "End<Start"}],
    )
    assert set(hits.keys()) == {"A.1", "B.2", "C.3"}
    assert "overdue" in hits["A.1"]["kinds"]
    assert "unassigned" in hits["A.1"]["kinds"]
    assert "UAT" in hits["A.1"]["yellow_status"]
    assert "Dev" in hits["A.1"]["yellow_pic"]
    assert "Dev" in hits["B.2"]["yellow_pic"]
    assert "stalled" in hits["B.2"]["kinds"]
    assert any("end_before_start" in k for k in hits["C.3"]["kinds"])


# ------------------------------------------------------------------
# Schema from uploaded headers
# ------------------------------------------------------------------

def test_schema_from_reference_like_headers(tmp_path):
    """Schema auto-detect từ header giống DanhSachFunction (Remark + From/To)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "Priority", "Remark",
        "Analysis - Start", "Analysis - End", "Analysis - Status", "Analysis - PIC",
        "UAT - From", "UAT - To", "UAT - Status", "UAT - PIC FPT",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    ws.cell(2, 1, 1)
    ws.cell(2, 2, "X.01")
    ws.cell(2, 3, "Demo")
    ws.cell(2, 4, "HR")
    path = tmp_path / "mau.xlsx"
    wb.save(path)
    wb.close()

    schema = schema_from_xlsx(str(path), "mau.xlsx")
    assert schema["headers"][0] == "STT"
    assert "Remark" in schema["headers"]
    assert schema["note_column"]["kind"] == "meta_remark"
    assert schema["note_column"]["header"] == "Remark"
    assert schema["meta_map"].get("ma_cn") == "Mã CN"
    phase_names = [p["name"] for p in schema["phase_groups"]]
    assert "Analysis" in phase_names
    assert "UAT" in phase_names

    review = review_from_xlsx(str(path), "mau.xlsx")
    assert review["slots"]
    ma_slot = next(s for s in review["slots"] if s["id"] == "meta:ma_cn")
    assert ma_slot["header"] == "Mã CN"
    assert ma_slot["confidence"] == "high"


def test_save_fl_export_template_per_project(tmp_path, sample_xlsx_path):
    pdir = tmp_path / "proj"
    pdir.mkdir()
    schema = schema_from_xlsx(sample_xlsx_path, "sample.xlsx")
    saved = save_fl_export_template(str(pdir), sample_xlsx_path, schema)
    assert (pdir / "fl_export_template.xlsx").is_file()
    assert (pdir / "fl_export_schema.json").is_file()
    raw = json.loads((pdir / "fl_export_schema.json").read_text(encoding="utf-8"))
    assert raw["version"] == 1
    assert saved["source_filename"] == "sample.xlsx"


# ------------------------------------------------------------------
# Export workbook — yellow / blue / union rows
# ------------------------------------------------------------------

def test_export_fl_reimport_highlights_and_date_chain(tmp_path, parsed_data, sample_xlsx_path):
    hits = collect_issue_hits(
        overdue_list=[{"ma_cn": "TMS.FR.02", "phase": "UAT", "days_overdue": 5}],
        unassigned_list=[{"ma_cn": "HR.FR.05", "phase": "Dev"}],
        stalled_list=[{
            "ma_cn": "HR.FR.05",
            "completed_phase": "Analysis",
            "waiting_phase": "Dev",
            "wait_days": 10,
        }],
    )
    # Thêm Remark vào schema bằng cách build từ parsed + fake note
    schema = build_schema_from_parsed(parsed_data, source_filename="sample.xlsx")
    # Sample không có Remark — note_column kind none; vẫn OK

    out = export_fl_reimport(
        parsed_data,
        hits=hits,
        output_dir=str(tmp_path),
        project_dir=str(tmp_path / "p"),
        source_xlsx=sample_xlsx_path,
        project_slug="demo",
        schema=schema,
    )
    assert os.path.isfile(out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Function List"]  # không sheet hướng dẫn
    ws = wb["Function List"]
    headers = [c.value for c in ws[1]]
    assert "Mã CN" in headers
    ma_idx = headers.index("Mã CN") + 1

    # Chỉ 2 function
    mas = []
    for r in range(2, ws.max_row + 1):
        mas.append(str(ws.cell(r, ma_idx).value or ""))
    assert set(mas) == {"TMS.FR.02", "HR.FR.05"}

    # Không ghi [Tracker] vào Remark (nếu có cột)
    if "Remark" in headers:
        remark_col = headers.index("Remark") + 1
        for r in range(2, ws.max_row + 1):
            rv = str(ws.cell(r, remark_col).value or "")
            assert "[Tracker]" not in rv

    # HR.FR.05: Dev Start auto-fill xanh
    dev_start_h = "Dev - Start"
    assert dev_start_h in headers
    ds_col = headers.index(dev_start_h) + 1
    pic_col = headers.index("Dev - PIC") + 1

    hr_row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, ma_idx).value) == "HR.FR.05":
            hr_row = r
            break
    assert hr_row
    start_cell = ws.cell(hr_row, ds_col)
    # Dev Start blank trong sample → fill
    assert start_cell.value  # filled
    assert start_cell.fill.fgColor.rgb in ("00DDEBF7", "DDEBF7") or \
        start_cell.fill.start_color.rgb in ("00DDEBF7", "DDEBF7")
    assert start_cell.alignment.wrap_text is not True

    pic_cell = ws.cell(hr_row, pic_col)
    # PIC trống + unassigned → vàng
    assert pic_cell.fill.fgColor.rgb in ("00FFF59D", "FFF59D") or \
        pic_cell.fill.start_color.rgb in ("00FFF59D", "FFF59D")

    # Verify date = Analysis End + 1 business day
    row = next(r for r in parsed_data.rows if r.meta.get("ma_cn") == "HR.FR.05")
    expected = format_fl_date(next_business_day(row.phases["Analysis"].end_date))
    assert start_cell.value == expected
    wb.close()


def test_flatten_cell_text_collapses_newlines():
    from exporter.fl_reimport_export import _flatten_cell_text
    assert _flatten_cell_text("CuongNM129;\n\nTungTT83") == "CuongNM129; TungTT83"
    assert _flatten_cell_text("A\r\nB") == "A B"
    assert _flatten_cell_text(12) == 12


def test_export_fl_endpoint(flask_client, sample_xlsx_path):
    import io
    with open(sample_xlsx_path, "rb") as f:
        flask_client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "test.xlsx")},
            content_type="multipart/form-data",
        )
    r = flask_client.get("/api/projects/default/export-fl-reimport")
    assert r.status_code == 200
    assert "spreadsheet" in (r.content_type or "") or r.data[:2] == b"PK"


def test_fl_template_upload_and_get(flask_client, sample_xlsx_path):
    with open(sample_xlsx_path, "rb") as f:
        r = flask_client.post(
            "/api/projects/default/fl-export-template?save=1",
            data={"file": (f, "sample.xlsx")},
            content_type="multipart/form-data",
        )
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body.get("saved") is True
    assert body.get("schema", {}).get("headers")

    g = flask_client.get("/api/projects/default/fl-export-template")
    assert g.status_code == 200
    gd = g.get_json()
    assert gd["has_template"] is True
    assert gd["schema"]["headers"]

    # Save slot override (file đã nằm sẵn — không SameFileError)
    assignments = {s["id"]: s["header"] for s in gd["schema"]["slots"] if s.get("header")}
    p = flask_client.post(
        "/api/projects/default/fl-export-template?save=1",
        json={"slot_assignments": assignments},
    )
    assert p.status_code == 200, p.get_data(as_text=True)

    d = flask_client.delete("/api/projects/default/fl-export-template")
    assert d.status_code == 200
    assert d.get_json()["has_template"] is False
