"""Tests HTTP integration cho Flask app."""
import io
import json
import os

import pytest


def _upload(client, xlsx_path):
    """Upload helper."""
    with open(xlsx_path, "rb") as f:
        data = {"file": (io.BytesIO(f.read()), "test.xlsx")}
    return client.post(
        "/api/upload",
        data=data,
        content_type="multipart/form-data",
    )


def test_index_page_loads(flask_client):
    """GET / trả về HTML 200."""
    r = flask_client.get("/")
    assert r.status_code == 200
    assert b"iHRP" in r.data or b"Function" in r.data


def test_upload_no_file_returns_400(flask_client):
    """POST /api/upload không có file → 400."""
    r = flask_client.post("/api/upload")
    assert r.status_code == 400


def test_upload_wrong_extension(flask_client):
    """Upload file không phải .xlsx → 400."""
    data = {"file": (io.BytesIO(b"hello"), "test.txt")}
    r = flask_client.post(
        "/api/upload",
        data=data,
        content_type="multipart/form-data",
    )
    assert r.status_code == 400


def test_upload_success(flask_client, sample_xlsx_path):
    """Upload thành công → JSON có metrics."""
    r = _upload(flask_client, sample_xlsx_path)
    assert r.status_code == 200
    data = r.get_json()
    assert data["success"] is True
    assert data["rows_count"] == 6
    assert "metrics" in data
    assert data["metrics"]["summary"]["total_functions"] == 6


def test_dashboard_endpoint_after_upload(flask_client, sample_xlsx_path):
    """GET /api/dashboard sau upload → JSON metrics."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/dashboard")
    assert r.status_code == 200
    data = r.get_json()
    assert "metrics" in data


def test_dashboard_endpoint_without_upload_returns_error(flask_client):
    """GET /api/dashboard chưa upload → 4xx."""
    r = flask_client.get("/api/dashboard")
    assert r.status_code in (400, 404)  # code trả 404 ("Chưa upload")


def test_overdue_endpoint(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/overdue")
    assert r.status_code == 200
    data = r.get_json()
    assert "overdue" in data
    assert "total" in data


def test_overdue_filter_module(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/overdue?module=TMS")
    assert r.status_code == 200
    for item in r.get_json()["overdue"]:
        assert item["module"] == "TMS"


def test_unassigned_endpoint(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/unassigned")
    assert r.status_code == 200
    assert "items" in r.get_json()


def test_long_duration_endpoint(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/long-duration")
    assert r.status_code == 200


def test_stalled_endpoint(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/stalled")
    assert r.status_code == 200


def _count_stalled_rows_in_xlsx(resp_data: bytes) -> int:
    """Đếm số row data trong file stalled export (header ở row 4)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb.active
    n = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        first = row[0]
        if first is None:
            continue
        try:
            int(str(first))
            n += 1
        except (TypeError, ValueError):
            pass
    wb.close()
    return n


def _stalled_modules_in_xlsx(resp_data: bytes) -> set[str]:
    """Lấy set Module từ cột Module (index 3) của file stalled export."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb.active
    mods: set[str] = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        try:
            int(str(row[0]))
        except (TypeError, ValueError):
            continue
        if row[3]:
            mods.add(str(row[3]))
    wb.close()
    return mods


def test_export_stalled_download(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/projects/default/export-stalled")
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ctype or "excel" in ctype
    assert _count_stalled_rows_in_xlsx(r.data) >= 1


def test_export_stalled_local_module_filter(flask_client, sample_xlsx_path):
    """Local module=HR → chỉ HR; module=NONEXISTENT → 0 row."""
    _upload(flask_client, sample_xlsx_path)
    r_all = flask_client.get("/api/projects/default/export-stalled")
    n_all = _count_stalled_rows_in_xlsx(r_all.data)
    assert n_all >= 1

    r_hr = flask_client.get("/api/projects/default/export-stalled?module=HR")
    assert r_hr.status_code == 200
    n_hr = _count_stalled_rows_in_xlsx(r_hr.data)
    assert n_hr >= 1
    assert n_hr <= n_all
    assert _stalled_modules_in_xlsx(r_hr.data) == {"HR"}

    r_none = flask_client.get("/api/projects/default/export-stalled?module=NONEXISTENT_XYZ")
    assert r_none.status_code == 200
    assert _count_stalled_rows_in_xlsx(r_none.data) == 0


def test_export_stalled_global_and_local_intersect(flask_client, sample_xlsx_path):
    """g_module=HR,TMS + local module=HR → chỉ HR."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-stalled?g_module=HR,TMS&module=HR"
    )
    assert r.status_code == 200
    assert _stalled_modules_in_xlsx(r.data) <= {"HR"}
    assert _count_stalled_rows_in_xlsx(r.data) >= 1


def test_risk_scores_endpoint(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/risk-scores?top=5")
    assert r.status_code == 200
    data = r.get_json()
    assert "items" in data
    assert len(data["items"]) <= 5


def test_snapshots_endpoint(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/snapshots")
    assert r.status_code == 200
    data = r.get_json()
    assert "snapshots" in data
    assert len(data["snapshots"]) >= 1


def test_delete_snapshot(flask_client, sample_xlsx_path):
    """Upload rồi delete snapshot."""
    _upload(flask_client, sample_xlsx_path)
    snapshots = flask_client.get("/api/snapshots").get_json()["snapshots"]
    if snapshots:
        r = flask_client.delete(f"/api/snapshots/{snapshots[0]['date']}")
        assert r.status_code == 200


def test_export_overdue_download(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/export-overdue")
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ctype or "excel" in ctype


def _count_overdue_rows_in_xlsx(resp_data: bytes) -> int:
    """Đếm số row data trong file overdue export (header ở row 4)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb.active
    n = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        # Cell đầu tiên là STT — check số nguyên
        first = row[0]
        if first is None:
            continue
        try:
            int(str(first))
            n += 1
        except (TypeError, ValueError):
            pass
    wb.close()
    return n


def test_export_overdue_respects_global_filter(flask_client, sample_xlsx_path):
    """
    Bug-fix regression: export-overdue phải áp global filter (g_module/g_process/g_pic)
    khi user apply header dashboard filter. Trước fix, endpoint dùng st["metrics"]
    (full unfiltered) → export ra tất cả record.
    """
    _upload(flask_client, sample_xlsx_path)

    # Baseline: no filter → toàn bộ overdue
    r_all = flask_client.get("/api/projects/default/export-overdue")
    assert r_all.status_code == 200
    n_all = _count_overdue_rows_in_xlsx(r_all.data)
    assert n_all >= 2  # sample_xlsx có ≥ 2 function overdue

    # Global filter g_module=TMS → chỉ TMS
    r_tms = flask_client.get("/api/projects/default/export-overdue?g_module=TMS")
    assert r_tms.status_code == 200
    n_tms = _count_overdue_rows_in_xlsx(r_tms.data)
    assert n_tms >= 1
    assert n_tms < n_all, f"Global filter phải thu hẹp: {n_tms} vs {n_all}"

    # Global filter g_module=NONEXISTENT → 0 row
    r_none = flask_client.get("/api/projects/default/export-overdue?g_module=NONEXISTENT_XYZ")
    assert r_none.status_code == 200
    assert _count_overdue_rows_in_xlsx(r_none.data) == 0


def test_export_overdue_local_widget_backward_compat(flask_client, sample_xlsx_path):
    """
    Backward compat: query cũ chỉ có `module`/`pic`/`phase` (local widget)
    vẫn giữ semantics filter local trên overdue_list gốc.
    """
    _upload(flask_client, sample_xlsx_path)
    r_all = flask_client.get("/api/projects/default/export-overdue")
    n_all = _count_overdue_rows_in_xlsx(r_all.data)
    r_tms = flask_client.get("/api/projects/default/export-overdue?module=TMS")
    n_tms = _count_overdue_rows_in_xlsx(r_tms.data)
    assert n_tms <= n_all


def test_export_overdue_global_and_local_intersect(flask_client, sample_xlsx_path):
    """
    Kết hợp global + local: global g_module=TMS,ESS và local module=TMS
    → chỉ TMS (local là subset của global — hành vi giống dashboard hiện tại).
    """
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get(
        "/api/projects/default/export-overdue?g_module=TMS,ESS&module=TMS"
    )
    assert r.status_code == 200
    r_tms_only = flask_client.get("/api/projects/default/export-overdue?g_module=TMS")
    assert _count_overdue_rows_in_xlsx(r.data) == _count_overdue_rows_in_xlsx(r_tms_only.data)


def test_export_overdue_post_body_global_filter(flask_client, sample_xlsx_path):
    """POST body {g_module: [...]} — cùng semantics như query g_module."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.post(
        "/api/projects/default/export-overdue",
        json={"g_module": ["TMS"]},
    )
    assert r.status_code == 200
    n_post = _count_overdue_rows_in_xlsx(r.data)
    r_get = flask_client.get("/api/projects/default/export-overdue?g_module=TMS")
    assert n_post == _count_overdue_rows_in_xlsx(r_get.data)


def test_export_full_report(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/export-full-report")
    assert r.status_code == 200
    assert len(r.data) > 100  # File phải có nội dung


def test_export_by_pic(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    # Lấy 1 PIC bất kỳ từ metrics
    metrics = flask_client.get("/api/dashboard").get_json()["metrics"]
    if metrics["pic_workload"]:
        pic = metrics["pic_workload"][0]["pic"]
        r = flask_client.get(f"/api/export-by-pic?pic={pic}")
        assert r.status_code == 200


def test_export_by_pic_missing_param_returns_400(flask_client, sample_xlsx_path):
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/export-by-pic")
    assert r.status_code == 400


# ==========================================================================
# PIC Blacklist endpoint
# ==========================================================================

def _upload_with_blacklist(client, tmp_path):
    """Tạo + upload 1 file có PIC lệch cột để tests blacklist endpoint."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Analysis - Status", "Analysis - PIC",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    ws.append([1, "TMS.FR.01", "Test", "TMS", "Open", "Closed"])
    ws.append([2, "TMS.FR.02", "Test", "TMS", "Open", "In-progress, HaiTD16"])
    p = tmp_path / "with_blacklist.xlsx"
    wb.save(str(p))
    wb.close()

    with open(str(p), "rb") as f:
        return client.post(
            "/api/upload",
            data={"file": (io.BytesIO(f.read()), "with_blacklist.xlsx")},
            content_type="multipart/form-data",
        )


def test_pic_blacklist_endpoint_empty_file(flask_client, sample_xlsx_path):
    """Sample fixture không có PIC lệch cột → items rỗng, total=0."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/pic-blacklist")
    assert r.status_code == 200
    data = r.get_json()
    assert data["success"] is True
    assert data["items"] == []
    assert data["total"] == 0
    assert data["keywords"] == []


def test_pic_blacklist_endpoint_with_data(flask_client, tmp_path):
    """File có PIC lệch cột → endpoint trả items đầy đủ metadata."""
    _upload_with_blacklist(flask_client, tmp_path)
    r = flask_client.get("/api/pic-blacklist")
    assert r.status_code == 200
    data = r.get_json()
    assert data["success"] is True
    assert data["total"] == 2  # "Closed" + "In-progress"
    # Keywords sorted, unique
    assert set(data["keywords"]) == {"Closed", "In-progress"}
    # Items có đủ field
    for it in data["items"]:
        for key in ("row_index", "phase_name", "header_text",
                    "raw_value", "matched_keyword", "ma_cn", "module"):
            assert key in it, f"Missing field '{key}' in item: {it}"


def test_pic_blacklist_endpoint_no_state_returns_404(flask_client):
    """Chưa upload → endpoint trả 404."""
    r = flask_client.get("/api/pic-blacklist")
    assert r.status_code == 404


def test_pic_blacklist_export_returns_xlsx(flask_client, tmp_path):
    """POST/GET export → trả file .xlsx (Content-Type spreadsheet)."""
    _upload_with_blacklist(flask_client, tmp_path)
    r = flask_client.get("/api/pic-blacklist/export")
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ctype or "excel" in ctype
    assert len(r.data) > 100  # File có content


def test_pic_blacklist_export_empty_still_works(flask_client, sample_xlsx_path):
    """
    Không có blacklist token → vẫn xuất được file (bảng rỗng có header).
    Edge case: user click "Xuất Excel" khi PIC sạch → không được crash.
    """
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/pic-blacklist/export")
    assert r.status_code == 200
    assert len(r.data) > 100


def test_upload_response_includes_pic_blacklist_count(flask_client, tmp_path):
    """Response upload phải có field top-level pic_blacklist_count để FE badge."""
    r = _upload_with_blacklist(flask_client, tmp_path)
    assert r.status_code == 200
    data = r.get_json()
    assert "pic_blacklist_count" in data
    assert data["pic_blacklist_count"] == 2


def test_dashboard_response_includes_pic_blacklist_count(flask_client, sample_xlsx_path):
    """Response dashboard cũng có pic_blacklist_count (sample = 0)."""
    _upload(flask_client, sample_xlsx_path)
    r = flask_client.get("/api/dashboard")
    assert r.status_code == 200
    data = r.get_json()
    assert "pic_blacklist_count" in data
    assert data["pic_blacklist_count"] == 0


def test_dashboard_blacklist_count_stable_under_filter(flask_client, tmp_path):
    """
    Blacklist count là data-quality info → KHÔNG cascade theo global filter.
    Filter Module=TMS phải vẫn trả cùng count như no-filter.
    """
    _upload_with_blacklist(flask_client, tmp_path)
    r_no = flask_client.get("/api/projects/default/dashboard")
    r_yes = flask_client.get("/api/projects/default/dashboard?module=TMS")
    c_no = r_no.get_json()["pic_blacklist_count"]
    c_yes = r_yes.get_json()["pic_blacklist_count"]
    assert c_no == c_yes == 2
