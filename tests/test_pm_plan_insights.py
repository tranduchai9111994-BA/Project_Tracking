# -*- coding: utf-8 -*-
"""Tests PM schedule insights (Phase B) + parser status/actual columns."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from analyzer.pm_plan_gantt import compute_pm_schedule_insights
from parser.pm_plan_parser import parse_plan, propose_sheet_mapping


def test_propose_mapping_vietinak_style():
    names = ["Kế hoạch khung", "Kế hoạch chi tiết", "LookupValue"]
    m = propose_sheet_mapping(names)
    assert m["Kế hoạch khung"] == "gantt"
    assert m["Kế hoạch chi tiết"] == "schedule"
    assert m["LookupValue"] == "ignore"


def test_parse_schedule_status_and_actual(tmp_path: Path):
    path = tmp_path / "plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail"
    ws["A7"] = "STT"
    ws["B7"] = "Công việc"
    ws["C7"] = "FPT"
    ws["D7"] = "Client"
    ws["E7"] = "Trạng thái"
    ws["F7"] = "Bắt đầu"
    ws["G7"] = "Kết thúc"
    ws["H7"] = "Ngày hoàn thành"
    ws["I7"] = "% Hoàn thành"
    ws["A8"] = 1
    ws["B8"] = "Task done"
    ws["C8"] = "FPT"
    ws["D8"] = "KH"
    ws["E8"] = "Completed"
    ws["F8"] = "2026-03-01"
    ws["G8"] = "2026-03-05"
    ws["H8"] = "2026-03-06"
    ws["I8"] = 100
    ws["A9"] = 2
    ws["B9"] = "Task late"
    ws["C9"] = "FPT"
    ws["D9"] = "KH"
    ws["E9"] = "Open"
    ws["F9"] = "2026-03-01"
    ws["G9"] = "2026-03-03"
    ws["H9"] = "2026-03-10"
    ws["I9"] = 50
    wb.save(path)

    out = parse_plan(str(path), {"Detail": "schedule"})
    sched = out["schedule"]
    assert len(sched) == 2
    done = sched[0]
    assert done["status"] == "Closed"
    assert done["actual_end"] == "2026-03-06"
    assert done["percent_complete"] == 100

    late = sched[1]
    assert late["status"] == "Open"
    assert late["actual_end"] == "2026-03-10"


def test_insights_overdue_and_slip():
    plan = {
        "schedule": [
            {
                "name": "Done task",
                "start": "2026-01-01",
                "end": "2026-01-10",
                "status": "Closed",
                "is_phase_header": False,
            },
            {
                "name": "Overdue open",
                "start": "2026-01-01",
                "end": "2026-01-05",
                "status": "Open",
                "is_phase_header": False,
            },
            {
                "name": "Slip actual",
                "start": "2026-02-01",
                "end": "2026-02-05",
                "actual_end": "2026-02-08",
                "status": "In-progress",
                "is_phase_header": False,
            },
            {
                "name": "Section A",
                "start": "2026-01-01",
                "end": "2026-03-01",
                "is_phase_header": True,
            },
        ],
    }
    today = date(2026, 8, 6)
    ins = compute_pm_schedule_insights(plan, today=today)
    s = ins["summary"]
    assert s["total_tasks"] == 3
    assert s["done"] == 1
    assert s["overdue"] == 2  # Overdue open + Slip actual (end < today, chưa Closed)
    assert s["in_progress"] == 0
    assert s["max_slip_days"] == 3
    assert len(ins["overdue_items"]) == 2
    assert ins["overdue_items"][0]["name"] == "Overdue open"
    assert ins["slip_items"][0]["name"] == "Slip actual"
