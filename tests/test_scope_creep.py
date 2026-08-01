"""Unit tests — Phase C: Change Request / Scope Creep tracking."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from analyzer.scope_creep import (
    CR_TAG,
    compute_scope_creep,
    is_cr_cell_value,
)
from analyzer.project_store import (
    VALID_FUNCTION_TAGS,
    load_project_settings,
    save_project_settings,
)
from parser.excel_parser import FunctionListParser


def _write_fl(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


def _parse(path: Path):
    return FunctionListParser().parse(str(path))


BASE_HEADERS = [
    "STT", "Mã CN", "Tên chức năng", "Module",
    "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC", "Dev - Estimate MH",
]


# ------------------------------------------------------------------
# Cell / detection helpers
# ------------------------------------------------------------------

def test_is_cr_cell_truthy_and_falsey():
    assert is_cr_cell_value("Yes") is True
    assert is_cr_cell_value("1") is True
    assert is_cr_cell_value("X") is True
    assert is_cr_cell_value("CR-001") is True
    assert is_cr_cell_value("Phát sinh") is True
    assert is_cr_cell_value(True) is True
    assert is_cr_cell_value(2) is True

    assert is_cr_cell_value(None) is False
    assert is_cr_cell_value("") is False
    assert is_cr_cell_value("No") is False
    assert is_cr_cell_value("0") is False
    assert is_cr_cell_value("-") is False
    assert is_cr_cell_value(False) is False


def test_cr_tag_in_valid_tags():
    assert CR_TAG in VALID_FUNCTION_TAGS


# ------------------------------------------------------------------
# Column auto-detect (primary)
# ------------------------------------------------------------------

def test_detect_exact_cr_column_not_description(tmp_path):
    """Exact header «CR» được nhận; «Description» không bị nhầm."""
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "Description", "CR",
        "Dev - Status", "Dev - Estimate MH",
    ]
    path = tmp_path / "cr.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "A", "TMS", "some description text", "Yes", "Closed", 16],
        [2, "F.02", "B", "TMS", "desc", "", "Open", 8],
        [3, "F.03", "C", "HR", "x", "No", "Open", 8],
    ])
    data = _parse(path)
    assert data.meta_columns.get("is_cr") is not None
    # Description không phải is_cr
    cr_col = data.meta_columns["is_cr"]
    assert data.headers.get("CR") == cr_col

    result = compute_scope_creep(data)
    assert result["detection"]["mode"] == "column"
    assert result["detection"]["column_header"] == "CR"
    assert result["summary"]["total_functions"] == 3
    assert result["summary"]["cr_count"] == 1
    assert result["summary"]["original_count"] == 2
    assert result["summary"]["creep_rate_pct"] == pytest.approx(33.3, abs=0.1)
    assert result["summary"]["mh_cr"] == 16.0
    assert result["summary"]["mh_original"] == 16.0


def test_detect_phat_sinh_and_cr_date(tmp_path):
    headers = BASE_HEADERS + ["Phát sinh", "Ngày phát sinh"]
    path = tmp_path / "ps.xlsx"
    _write_fl(path, headers, [
        [1, "A1", "FN1", "HR",
         date(2026, 7, 1), date(2026, 7, 2), "Closed", "P", 10,
         "Yes", date(2026, 6, 15)],
        [2, "A2", "FN2", "HR",
         None, None, "Open", "Q", 20,
         "", None],
        [3, "A3", "FN3", "TMS",
         None, None, "Open", "R", 30,
         "X", "01/07/2026"],
    ])
    data = _parse(path)
    assert data.meta_columns.get("is_cr") is not None
    assert data.meta_columns.get("cr_date") is not None

    result = compute_scope_creep(data)
    assert result["detection"]["mode"] == "column"
    assert "Phát sinh" in (result["detection"]["column_header"] or "")
    assert result["summary"]["cr_count"] == 2
    assert result["summary"]["creep_rate_pct"] == pytest.approx(66.7, abs=0.1)
    assert result["summary"]["mh_cr"] == 40.0
    assert result["summary"]["mh_original"] == 20.0
    assert result["summary"]["cr_with_raised_date"] == 2
    # Module breakdown
    by_mod = {m["module"]: m for m in result["modules"]}
    assert by_mod["HR"]["cr"] == 1
    assert by_mod["TMS"]["cr"] == 1


def test_default_mh_when_estimate_blank(tmp_path):
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "CR",
        "Dev - Status",
    ]
    path = tmp_path / "blank_mh.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "CR fn", "M", "Yes", "Open"],
        [2, "F.02", "Orig", "M", "", "Open"],
    ])
    result = compute_scope_creep(_parse(path), default_mh=8.0)
    # Mỗi function 1 phase trống MH → 8 MH
    assert result["summary"]["mh_cr"] == 8.0
    assert result["summary"]["mh_original"] == 8.0
    assert result["summary"]["phases_default_mh"] == 2


def test_cancelled_phase_excluded_from_effort(tmp_path):
    headers = BASE_HEADERS + ["Change Request"]
    path = tmp_path / "cancel.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "CR", "M",
         date(2026, 7, 1), date(2026, 7, 2), "Cancelled", "A", 100, "Yes"],
        [2, "F.02", "OK", "M",
         date(2026, 7, 1), date(2026, 7, 2), "Closed", "B", 16, "Yes"],
    ])
    result = compute_scope_creep(_parse(path))
    assert result["summary"]["cr_count"] == 2
    # F.01 Cancelled → 0 MH; F.02 = 16
    assert result["summary"]["mh_cr"] == 16.0


# ------------------------------------------------------------------
# Fallback: tag + settings codes (no column)
# ------------------------------------------------------------------

def test_fallback_tag_and_settings_when_no_column(tmp_path):
    path = tmp_path / "nocol.xlsx"
    _write_fl(path, BASE_HEADERS, [
        [1, "F.01", "A", "TMS",
         None, None, "Open", "P", 10],
        [2, "F.02", "B", "TMS",
         None, None, "Open", "Q", 20],
        [3, "F.03", "C", "HR",
         None, None, "Open", "R", 30],
    ])
    data = _parse(path)
    assert data.meta_columns.get("is_cr") is None

    result = compute_scope_creep(
        data,
        function_tags={"F.01": ["CR", "đã review"]},
        cr_function_codes=["F.03", "unknown"],
    )
    assert result["detection"]["mode"] == "tag_or_settings"
    assert result["detection"]["primary"] == "tag_or_settings"
    assert result["summary"]["cr_count"] == 2
    assert result["summary"]["mh_cr"] == 40.0  # 10 + 30
    assert result["summary"]["mh_original"] == 20.0
    sources = {it["ma_cn"]: it["source"] for it in result["cr_functions"]}
    assert sources["F.01"] == "tag"
    assert sources["F.03"] == "settings"


def test_no_column_no_fallback_all_original(tmp_path):
    path = tmp_path / "clean.xlsx"
    _write_fl(path, BASE_HEADERS, [
        [1, "F.01", "A", "M", None, None, "Open", "P", 8],
    ])
    result = compute_scope_creep(_parse(path))
    assert result["detection"]["mode"] == "none"
    assert result["summary"]["cr_count"] == 0
    assert result["summary"]["creep_rate_pct"] == 0.0
    assert any("Chưa có tag" in m or "scope gốc" in m for m in result["messages"])


def test_column_present_ignores_tag_settings(tmp_path):
    """Khi có cột CR → cột là nguồn chính (tag/settings không cộng thêm)."""
    headers = BASE_HEADERS + ["CR"]
    path = tmp_path / "col_wins.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "A", "M", None, None, "Open", "P", 8, ""],  # không CR trên cột
        [2, "F.02", "B", "M", None, None, "Open", "Q", 8, "Yes"],
    ])
    result = compute_scope_creep(
        _parse(path),
        function_tags={"F.01": ["CR"]},
        cr_function_codes=["F.01"],
    )
    assert result["detection"]["mode"] == "column"
    assert result["summary"]["cr_count"] == 1
    assert result["cr_functions"][0]["ma_cn"] == "F.02"


# ------------------------------------------------------------------
# Settings persistence
# ------------------------------------------------------------------

def test_cr_function_codes_settings_roundtrip(tmp_path):
    d = str(tmp_path / "proj")
    Path(d).mkdir()
    s = load_project_settings(d)
    assert s["cr_function_codes"] == []
    save_project_settings(d, {"cr_function_codes": "A1, B2\nC3"})
    s2 = load_project_settings(d)
    assert s2["cr_function_codes"] == ["A1", "B2", "C3"]
    save_project_settings(d, {"cr_function_codes": []})
    assert load_project_settings(d)["cr_function_codes"] == []
