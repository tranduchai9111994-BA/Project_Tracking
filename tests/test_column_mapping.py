"""
T32 — Tests cho Column Mapping Wizard.

Cover:
  · Fuzzy match: các case điển hình (English → Việt / snake_case → space).
  · Parser hoạt động với column_mapping override header lạ.
  · Preset CRUD (project_store): list, save (upsert), delete, cap.
  · API endpoints: /api/upload-preview + /api/upload-confirm + preset routes.
  · Edge cases: file lỗi, tmp_id không tồn tại, path traversal attack.
"""
from __future__ import annotations

import io
import os
from pathlib import Path

import openpyxl
import pytest

from parser import column_mapping as cm_mod
from parser.excel_parser import FunctionListParser
from analyzer import project_store as ps


# ---------------------------------------------------------------------------
# Fixtures — file Excel với header không chuẩn iHRP
# ---------------------------------------------------------------------------


@pytest.fixture
def non_standard_xlsx(tmp_path) -> str:
    """File có header English + phase pattern lạ (không có ' - ')."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    ws.append([
        "Function Code",   # → Mã CN
        "Function Name",   # → Tên chức năng
        "Phan he",         # → Module
        "AnalysisStart",   # → Analysis - Start (không có ' - ')
        "AnalysisEnd",     # → Analysis - End
        "AnalysisStatus",  # → Analysis - Status
        "AnalysisPIC",     # → Analysis - PIC
    ])
    ws.append(["PR.FR.01", "Tính lương cơ bản", "PR",
               "2026-01-01", "2026-01-15", "Closed", "AnhTV"])
    ws.append(["HR.HRM.05", "Chấm công", "HR",
               "2026-02-01", "2026-02-20", "In-progress", "BaoLQ"])
    path = tmp_path / "non_standard.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def project_dir(tmp_path) -> str:
    d = tmp_path / "proj"
    d.mkdir()
    return str(d)


# ---------------------------------------------------------------------------
# 1. Fuzzy match
# ---------------------------------------------------------------------------


def test_ihrp_standard_columns_present():
    """IHRP_STANDARD_COLUMNS phải bao gồm meta chính + phase phổ biến."""
    cols = cm_mod.IHRP_STANDARD_COLUMNS
    assert "Mã CN" in cols
    assert "Tên chức năng" in cols
    assert "Module" in cols
    assert "Analysis - Start" in cols
    assert "UAT - Status" in cols
    # Không được có duplicate
    assert len(cols) == len(set(cols))


@pytest.mark.parametrize("ihrp, actual, min_score", [
    ("Function Name", "Tên chức năng", 0.0),   # chỉ cần > 0 (ngôn ngữ khác)
    ("Analysis - Start", "AnalysisStart", 0.6),
    ("Analysis - Start", "Analysis Start Date", 0.5),
    ("Module", "Phân hệ", 0.0),                # xu hướng thấp, vẫn cần > 0
    ("Mã CN", "Function Code", 0.0),
    ("Mã CN", "Ma CN", 0.5),
    ("Mã CN", "Mã chức năng", 0.4),
])
def test_fuzzy_score_min_threshold(ihrp, actual, min_score):
    """Verify fuzzy score đủ cao cho các case điển hình."""
    score = cm_mod._fuzzy_score(ihrp, actual)
    assert score > min_score, f"Score {score} for '{ihrp}' vs '{actual}'"


def test_fuzzy_score_normalizes_case_and_sep():
    """`Analysis-Start`, `analysis_start`, `AnalysisStart` cùng score."""
    base = "Analysis - Start"
    s1 = cm_mod._fuzzy_score(base, "Analysis-Start")
    s2 = cm_mod._fuzzy_score(base, "analysis_start")
    s3 = cm_mod._fuzzy_score(base, "AnalysisStart")
    # Chấp nhận sai lệch nhỏ do bonus token overlap
    assert s1 >= 0.7 and s2 >= 0.7 and s3 >= 0.7


def test_suggest_mapping_returns_top_k():
    """Với 1 iHRP col, gợi ý top-3 header thực tế theo score."""
    actual = ["Function Code", "Function Name", "Code123", "Name X"]
    result = cm_mod.suggest_mapping(actual, ["Mã CN", "Tên chức năng"], top_k=2)
    assert len(result["Mã CN"]) <= 2
    assert len(result["Tên chức năng"]) <= 2
    # "Function Code" phải là top cho "Mã CN"
    assert result["Mã CN"][0]["header"] in {"Function Code", "Code123"}


def test_suggest_mapping_empty_when_no_match():
    """Không header nào match → list rỗng."""
    actual = ["completely", "unrelated", "headers"]
    result = cm_mod.suggest_mapping(actual, ["Analysis - Start"], min_score=0.7)
    assert result["Analysis - Start"] == []


def test_suggest_mapping_sorted_by_score_desc():
    actual = ["Function Code", "Function Name"]
    result = cm_mod.suggest_mapping(actual, ["Mã CN"])
    scores = [x["score"] for x in result["Mã CN"]]
    assert scores == sorted(scores, reverse=True)


# ---------------------------------------------------------------------------
# 2. read_headers_and_preview
# ---------------------------------------------------------------------------


def test_read_headers_and_preview_basic(non_standard_xlsx):
    headers, preview, sheet = cm_mod.read_headers_and_preview(non_standard_xlsx)
    assert headers[0] == "Function Code"
    assert headers[1] == "Function Name"
    assert len(preview) == 2  # 2 data rows
    assert preview[0][0] == "PR.FR.01"
    assert sheet == "Function List"


def test_read_headers_and_preview_missing_file(tmp_path):
    headers, preview, sheet = cm_mod.read_headers_and_preview(
        str(tmp_path / "nonexistent.xlsx")
    )
    assert headers == []
    assert preview == []
    assert sheet is None


def test_read_headers_datetime_serialized(tmp_path):
    """Datetime cell trong preview phải convert sang ISO string (JSON-safe)."""
    from datetime import datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Code", "Date"])
    ws.append(["A", datetime(2026, 1, 15)])
    path = tmp_path / "dt.xlsx"
    wb.save(str(path))
    headers, preview, _ = cm_mod.read_headers_and_preview(str(path))
    assert preview[0][1].startswith("2026-01-15")


# ---------------------------------------------------------------------------
# 3. Parser với column_mapping
# ---------------------------------------------------------------------------


def test_parser_with_column_mapping_maps_meta(non_standard_xlsx):
    """
    Không có mapping → parser detect được 'Function Code' → "Code" keyword
    match cho ma_cn. Nhưng 'AnalysisStart' không match 'Phase - Attr' pattern
    → 0 phase group.
    """
    parsed_no_map = FunctionListParser().parse(non_standard_xlsx)
    assert len(parsed_no_map.phase_groups) == 0  # không detect được phase

    # Có mapping → phase group detect thành công
    mapping = {
        "Mã CN": "Function Code",
        "Tên chức năng": "Function Name",
        "Module": "Phan he",
        "Analysis - Start": "AnalysisStart",
        "Analysis - End": "AnalysisEnd",
        "Analysis - Status": "AnalysisStatus",
        "Analysis - PIC": "AnalysisPIC",
    }
    parsed = FunctionListParser().parse(non_standard_xlsx, column_mapping=mapping)
    assert len(parsed.phase_groups) == 1
    assert parsed.phase_groups[0].name == "Analysis"
    assert len(parsed.rows) == 2
    assert parsed.rows[0].meta.get("ma_cn") == "PR.FR.01"
    assert parsed.rows[0].meta.get("module") == "PR"


def test_parser_column_mapping_none_still_works(non_standard_xlsx):
    """column_mapping=None → parser hoạt động như trước (backward compat)."""
    parsed = FunctionListParser().parse(non_standard_xlsx, column_mapping=None)
    assert len(parsed.rows) == 2


def test_parser_apply_column_mapping_missing_actual_skipped(non_standard_xlsx):
    """Actual header không tồn tại trong file → skip thầm lặng."""
    mapping = {
        "Mã CN": "Function Code",
        "Tên chức năng": "Missing Header Doesnt Exist",  # skip
    }
    parsed = FunctionListParser().parse(non_standard_xlsx, column_mapping=mapping)
    assert parsed.rows[0].meta.get("ma_cn") == "PR.FR.01"
    # ten_cn có thể vẫn detect được qua auto-detect keyword "Function Name"
    # → giá trị vẫn có; test không assert None.


def test_sanitize_column_mapping():
    """Drop entry rỗng, trim, cap 200 keys."""
    inp = {
        "Mã CN": "  Function Code  ",
        "": "empty key",
        "empty value": "",
        None: "None key",
    }
    out = cm_mod.sanitize_column_mapping(inp)
    assert out == {"Mã CN": "Function Code"}


def test_sanitize_column_mapping_non_dict():
    assert cm_mod.sanitize_column_mapping("not a dict") == {}
    assert cm_mod.sanitize_column_mapping(None) == {}


# ---------------------------------------------------------------------------
# 4. Preset CRUD (project_store)
# ---------------------------------------------------------------------------


def test_list_mapping_presets_empty(project_dir):
    assert ps.list_mapping_presets(project_dir) == []


def test_save_and_list_mapping_preset(project_dir):
    ps.save_mapping_preset(project_dir, "iHRP MPHG", {
        "Mã CN": "Function Code",
        "Tên chức năng": "Function Name",
    })
    presets = ps.list_mapping_presets(project_dir)
    assert len(presets) == 1
    assert presets[0]["name"] == "iHRP MPHG"
    assert presets[0]["mapping"]["Mã CN"] == "Function Code"


def test_save_mapping_preset_upsert(project_dir):
    """Save cùng name → overwrite mapping."""
    ps.save_mapping_preset(project_dir, "T1", {"A": "1"})
    ps.save_mapping_preset(project_dir, "T1", {"B": "2"})
    presets = ps.list_mapping_presets(project_dir)
    assert len(presets) == 1
    assert presets[0]["mapping"] == {"B": "2"}


def test_save_mapping_preset_empty_name_raises(project_dir):
    with pytest.raises(ValueError):
        ps.save_mapping_preset(project_dir, "", {"A": "B"})


def test_delete_mapping_preset(project_dir):
    ps.save_mapping_preset(project_dir, "P1", {"A": "1"})
    ps.save_mapping_preset(project_dir, "P2", {"B": "2"})
    deleted, remaining = ps.delete_mapping_preset(project_dir, "P1")
    assert deleted is True
    assert len(remaining) == 1
    assert remaining[0]["name"] == "P2"


def test_delete_mapping_preset_missing(project_dir):
    deleted, _ = ps.delete_mapping_preset(project_dir, "no_such")
    assert deleted is False


def test_mapping_preset_cap(project_dir):
    """Vượt _MAX_PRESETS → drop preset cũ nhất."""
    for i in range(35):
        ps.save_mapping_preset(project_dir, f"P{i:02d}", {"A": str(i)})
    presets = ps.list_mapping_presets(project_dir)
    assert len(presets) == 30  # _MAX_PRESETS


def test_mapping_preset_persistence(project_dir):
    """Save → tạo file JSON đúng path."""
    ps.save_mapping_preset(project_dir, "Test", {"A": "1"})
    fp = os.path.join(project_dir, "excel_mapping_presets.json")
    assert os.path.isfile(fp)


# ---------------------------------------------------------------------------
# 5. API endpoints (Flask client)
# ---------------------------------------------------------------------------


def test_api_upload_preview_returns_schema(flask_client, non_standard_xlsx):
    with open(non_standard_xlsx, "rb") as f:
        r = flask_client.post(
            "/api/upload-preview",
            data={"file": (f, "test.xlsx")},
            content_type="multipart/form-data",
        )
    assert r.status_code == 200
    body = r.get_json()
    assert body["success"] is True
    assert "tmp_id" in body
    assert body["headers"][0] == "Function Code"
    assert len(body["preview_rows"]) == 2
    assert "ihrp_columns" in body
    assert "auto_suggest" in body
    assert body["auto_suggest"]["Mã CN"]  # có suggestion cho Mã CN


def test_api_upload_preview_reject_non_xlsx(flask_client):
    r = flask_client.post(
        "/api/upload-preview",
        data={"file": (io.BytesIO(b"x"), "foo.txt")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400
    assert "xlsx" in r.get_json()["error"].lower()


def test_api_upload_preview_no_file(flask_client):
    r = flask_client.post("/api/upload-preview")
    assert r.status_code == 400


def test_api_upload_confirm_applies_mapping(flask_client, non_standard_xlsx):
    """Full flow: preview → confirm → dashboard payload phải có rows_count > 0."""
    with open(non_standard_xlsx, "rb") as f:
        r = flask_client.post(
            "/api/upload-preview",
            data={"file": (f, "test.xlsx")},
            content_type="multipart/form-data",
        )
    tmp_id = r.get_json()["tmp_id"]

    r2 = flask_client.post("/api/upload-confirm", json={
        "tmp_id": tmp_id,
        "project_slug": "default",
        "column_mapping": {
            "Mã CN": "Function Code",
            "Tên chức năng": "Function Name",
            "Module": "Phan he",
            "Analysis - Start": "AnalysisStart",
            "Analysis - End": "AnalysisEnd",
            "Analysis - Status": "AnalysisStatus",
        },
    })
    body = r2.get_json()
    assert r2.status_code == 200, body
    assert body["success"] is True
    assert body["rows_count"] == 2
    assert body["column_mapping_applied"] is True
    assert body["column_mapping_count"] >= 5


def test_api_upload_confirm_missing_tmp_id(flask_client):
    r = flask_client.post("/api/upload-confirm", json={
        "project_slug": "default",
        "column_mapping": {},
    })
    assert r.status_code == 400


def test_api_upload_confirm_bad_tmp_id_path_traversal(flask_client):
    """tmp_id có ký tự lạ (attack path traversal) → 400."""
    r = flask_client.post("/api/upload-confirm", json={
        "tmp_id": "../../../etc/passwd",
        "project_slug": "default",
    })
    assert r.status_code == 400


def test_api_upload_confirm_expired_tmp(flask_client):
    """tmp_id hợp lệ format nhưng file không tồn tại → 404."""
    r = flask_client.post("/api/upload-confirm", json={
        "tmp_id": "abcdef1234567890",
        "project_slug": "default",
    })
    assert r.status_code == 404


def test_api_upload_confirm_no_mapping_uses_auto_detect(flask_client, tmp_path):
    """
    File chuẩn iHRP → upload-confirm không cần mapping (rỗng) → parser
    auto-detect vẫn hoạt động (backward compat).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    ws.append(["Mã CN", "Tên chức năng", "Module",
               "Analysis - Start", "Analysis - End", "Analysis - Status"])
    ws.append(["PR.01", "Test", "PR", "2026-01-01", "2026-01-15", "Closed"])
    path = tmp_path / "standard.xlsx"
    wb.save(str(path))

    with open(str(path), "rb") as f:
        r = flask_client.post(
            "/api/upload-preview",
            data={"file": (f, "standard.xlsx")},
            content_type="multipart/form-data",
        )
    tmp_id = r.get_json()["tmp_id"]
    r2 = flask_client.post("/api/upload-confirm", json={
        "tmp_id": tmp_id,
        "project_slug": "default",
        "column_mapping": {},
    })
    body = r2.get_json()
    assert body["success"] is True
    assert body["column_mapping_applied"] is False
    assert body["rows_count"] == 1


def test_api_mapping_preset_crud(flask_client):
    # GET rỗng
    r = flask_client.get("/api/projects/default/mapping-presets")
    assert r.status_code == 200
    assert r.get_json()["presets"] == []

    # POST save
    r = flask_client.post("/api/projects/default/mapping-presets", json={
        "name": "MPHG Template",
        "mapping": {"Mã CN": "Function Code"},
    })
    assert r.status_code == 201
    presets = r.get_json()["presets"]
    assert len(presets) == 1
    assert presets[0]["name"] == "MPHG Template"

    # GET có 1 preset
    r = flask_client.get("/api/projects/default/mapping-presets")
    assert len(r.get_json()["presets"]) == 1

    # DELETE
    r = flask_client.delete("/api/projects/default/mapping-presets/MPHG Template")
    assert r.status_code == 200
    r = flask_client.get("/api/projects/default/mapping-presets")
    assert r.get_json()["presets"] == []


def test_api_mapping_preset_missing_name_400(flask_client):
    r = flask_client.post("/api/projects/default/mapping-presets", json={"name": ""})
    assert r.status_code == 400


def test_api_mapping_preset_delete_not_found_404(flask_client):
    r = flask_client.delete("/api/projects/default/mapping-presets/no_such")
    assert r.status_code == 404


def test_api_upload_preview_with_project_slug_returns_presets(flask_client, non_standard_xlsx):
    """Nếu FE gửi project_slug trong query → response kèm list preset."""
    # Save 1 preset trước
    flask_client.post("/api/projects/default/mapping-presets", json={
        "name": "Vendor A", "mapping": {"Mã CN": "Function Code"}
    })
    with open(non_standard_xlsx, "rb") as f:
        r = flask_client.post(
            "/api/upload-preview?project_slug=default",
            data={"file": (f, "test.xlsx")},
            content_type="multipart/form-data",
        )
    body = r.get_json()
    assert body["success"] is True
    assert len(body["presets"]) == 1
    assert body["presets"][0]["name"] == "Vendor A"
