"""Unit tests — Phase E: UAT / Customer feedback quality."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from analyzer.project_store import VALID_FUNCTION_TAGS
from analyzer.uat_quality import (
    UAT_ISSUE_TAG,
    compute_uat_quality,
    find_uat_phase_name,
    parse_count,
)
from parser.excel_parser import FunctionListParser


def _write_fl(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


def _parse(path: Path):
    return FunctionListParser().parse(str(path))


BASE_HEADERS = [
    "STT", "Mã CN", "Tên chức năng", "Module",
    "UAT - Start", "UAT - End", "UAT - Status", "UAT - PIC",
]


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def test_parse_count():
    assert parse_count(3) == 3
    assert parse_count(2.0) == 2
    assert parse_count("5") == 5
    assert parse_count("1,5") == 1  # float truncated via int
    assert parse_count(None) is None
    assert parse_count("") is None
    assert parse_count("-") is None
    assert parse_count("n/a") is None
    assert parse_count(-1) is None


def test_uat_issue_tag_in_valid_tags():
    assert UAT_ISSUE_TAG in VALID_FUNCTION_TAGS


# ------------------------------------------------------------------
# Column auto-detect
# ------------------------------------------------------------------

def test_detect_so_loi_not_debug(tmp_path):
    """«Số lỗi» được nhận; «Debug Notes» không bị nhầm Bug."""
    headers = BASE_HEADERS + ["Debug Notes", "Số lỗi", "Reopen", "Số vòng UAT"]
    path = tmp_path / "def.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "A", "TMS",
         date(2026, 7, 1), date(2026, 7, 10), "Closed", "P",
         "some debug", 3, 1, 2],
        [2, "F.02", "B", "TMS",
         date(2026, 7, 1), date(2026, 7, 10), "Closed", "Q",
         "", 0, 0, 1],
        [3, "F.03", "C", "HR",
         date(2026, 7, 1), date(2026, 7, 10), "In-progress", "R",
         "", 5, 2, 3],
    ])
    data = _parse(path)
    assert data.meta_columns.get("defect_count") is not None
    assert data.headers.get("Số lỗi") == data.meta_columns["defect_count"]
    assert data.meta_columns.get("reopen_count") is not None
    assert data.meta_columns.get("uat_cycle") is not None
    # Debug Notes không phải defect
    assert data.headers.get("Debug Notes") != data.meta_columns["defect_count"]

    result = compute_uat_quality(data)
    assert result["detection"]["mode"] == "column"
    assert result["detection"]["defect_header"] == "Số lỗi"
    assert result["summary"]["total_defects"] == 8  # 3+0+5
    assert result["summary"]["fns_with_defects"] == 2
    assert result["summary"]["total_reopens"] == 3  # 1+0+2
    assert result["summary"]["fns_with_reopen"] == 2
    # denom: F.01 Closed+reopen, F.02 Closed, F.03 reopen>0 → 3
    assert result["summary"]["reopen_denom"] == 3
    assert result["summary"]["reopen_rate_pct"] == pytest.approx(66.7, abs=0.1)
    assert result["summary"]["avg_uat_cycles"] == pytest.approx(2.0, abs=0.01)
    assert result["summary"]["multi_cycle_count"] == 2  # cycle 2 and 3
    assert find_uat_phase_name(data) == "UAT"

    by_mod = {m["module"]: m for m in result["modules"]}
    assert by_mod["TMS"]["defects"] == 3
    assert by_mod["HR"]["defects"] == 5


def test_exact_bug_header_not_debug(tmp_path):
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module", "Bug", "Debug Notes",
        "UAT - Status",
    ]
    path = tmp_path / "bug.xlsx"
    _write_fl(path, headers, [
        [1, "A1", "FN", "M", 4, "debug text", "Closed"],
    ])
    data = _parse(path)
    assert data.meta_columns["defect_count"] == data.headers["Bug"]
    result = compute_uat_quality(data)
    assert result["summary"]["total_defects"] == 4


def test_feedback_and_phat_hoi(tmp_path):
    headers = BASE_HEADERS + ["Phản hồi", "Feedback Count"]
    # Feedback Count is more specific — may win for feedback_count;
    # Phản hồi also matches. First found by keyword order in META.
    path = tmp_path / "fb.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "A", "HR",
         None, None, "Closed", "P", 2, 7],
    ])
    data = _parse(path)
    assert data.meta_columns.get("feedback_count") is not None
    result = compute_uat_quality(data)
    assert result["detection"]["mode"] == "column"
    assert result["summary"]["total_feedback"] is not None
    assert result["summary"]["total_feedback"] >= 2


def test_prefer_uat_over_config_uat(tmp_path):
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Config UAT - Status", "UAT - Status", "Số lỗi",
    ]
    path = tmp_path / "phases.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "A", "M", "Closed", "Closed", 1],
    ])
    data = _parse(path)
    assert find_uat_phase_name(data) == "UAT"


# ------------------------------------------------------------------
# Empty / tag fallback (no inventing defects)
# ------------------------------------------------------------------

def test_no_columns_empty_no_fake_defects(tmp_path):
    path = tmp_path / "empty.xlsx"
    _write_fl(path, BASE_HEADERS, [
        [1, "F.01", "A", "M", None, None, "Closed", "P"],
        [2, "F.02", "B", "M", None, None, "Open", "Q"],
    ])
    data = _parse(path)
    assert data.meta_columns.get("defect_count") is None
    result = compute_uat_quality(data)
    assert result["detection"]["mode"] == "none"
    assert result["detection"]["has_quality_columns"] is False
    assert result["summary"]["total_defects"] is None
    assert result["summary"]["total_feedback"] is None
    assert result["summary"]["reopen_rate_pct"] is None
    assert result["summary"]["tagged_uat_issue"] == 0
    assert any("Không có cột" in m for m in result["messages"])


def test_tag_fallback_qualitative_only(tmp_path):
    path = tmp_path / "tag.xlsx"
    _write_fl(path, BASE_HEADERS, [
        [1, "F.01", "A", "TMS", None, None, "Closed", "P"],
        [2, "F.02", "B", "TMS", None, None, "Open", "Q"],
    ])
    result = compute_uat_quality(
        _parse(path),
        function_tags={"F.01": ["UAT issue", "đã review"]},
    )
    assert result["detection"]["mode"] == "tag"
    assert result["summary"]["tagged_uat_issue"] == 1
    assert result["summary"]["total_defects"] is None  # không bịa
    assert any("UAT issue" in m for m in result["messages"])


def test_column_present_lists_functions(tmp_path):
    headers = BASE_HEADERS + ["Defect Count", "Reopen Count", "UAT Cycle"]
    path = tmp_path / "list.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "Hot", "HR",
         None, None, "Closed", "P", 10, 3, 2],
        [2, "F.02", "Ok", "HR",
         None, None, "Closed", "Q", 0, 0, 1],
    ])
    result = compute_uat_quality(_parse(path))
    assert result["functions"][0]["ma_cn"] == "F.01"
    assert result["functions"][0]["defect_count"] == 10
    assert result["summary"]["reopen_rate_pct"] == 50.0  # 1/2
