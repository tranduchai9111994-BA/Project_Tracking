"""Tests smoke cho exporter.rlog_exporter + API export-rlog-weekly."""
from datetime import date
from pathlib import Path

from parser.excel_parser import ParsedData, FunctionRow, PhaseData, PhaseGroup
from analyzer.rlog_weekly import compute_rlog_weekly
from exporter.rlog_exporter import export_rlog_weekly_report


TODAY = date(2026, 7, 31)


def _pg() -> list[PhaseGroup]:
    return [
        PhaseGroup(name="Analysis", attributes={"Start": 1, "End": 2, "Status": 3, "RlogID": 4}),
        PhaseGroup(name="Dev", attributes={"Start": 5, "End": 6, "Status": 7}),
    ]


def _row(ma, module, rlog_id, status, start, end):
    return FunctionRow(
        row_num=2,
        meta={"ma_cn": ma, "ten_cn": ma, "module": module},
        phases={
            "Analysis": PhaseData(status="Closed", end_date=date(2026, 3, 1), extra={"RlogID": rlog_id}),
            "Dev": PhaseData(status=status, start_date=start, end_date=end, pics=["DevA"]),
        },
    )


def _data(rows) -> ParsedData:
    return ParsedData(
        headers={},
        meta_columns={},
        phase_groups=_pg(),
        rows=rows,
        all_modules=sorted({r.meta.get("module") for r in rows if r.meta.get("module")}),
        all_phases=["Analysis", "Dev"],
        all_pics=["DevA"],
        all_statuses=["Closed", "In-progress", "Open"],
        all_priorities=[],
        all_complexities=[],
        all_giai_doan=[],
        all_processes=[],
    )


def test_export_rlog_weekly_workbook(tmp_path: Path):
    rows = [
        _row("A1", "PR", "R1", "Closed", date(2026, 7, 20), date(2026, 7, 29)),
        _row("A2", "PR", "R2", "In-progress", date(2026, 8, 3), date(2026, 8, 7)),
        _row("B1", "SI", "R3", "Open", date(2026, 8, 3), date(2026, 8, 5)),
    ]
    payload = compute_rlog_weekly(_data(rows), today=TODAY)
    path = export_rlog_weekly_report(payload, output_dir=str(tmp_path), subtitle="Test")
    assert Path(path).exists()
    assert path.endswith(".xlsx")

    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert any(n.startswith("Coded") for n in wb.sheetnames)
    assert any(n.startswith("Ke_hoach") for n in wb.sheetnames)
    # Summary có count coded / plan
    ws = wb["Summary"]
    vals = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, 25) if ws.cell(r, 1).value}
    assert vals.get("Coded tuần này") == 1
    assert vals.get("Kế hoạch tuần tới") == 2
    wb.close()


def test_export_rlog_weekly_endpoint(flask_client, sample_xlsx_path):
    """GET export-rlog-weekly trả xlsx; filter module thu hẹp được."""
    # Upload sample
    with open(sample_xlsx_path, "rb") as f:
        flask_client.post(
            "/api/projects/default/upload",
            data={"file": (f, "fl.xlsx")},
            content_type="multipart/form-data",
        )
    r_all = flask_client.get("/api/projects/default/export-rlog-weekly")
    assert r_all.status_code == 200
    assert "spreadsheet" in (r_all.content_type or "") or r_all.data[:2] == b"PK"

    r_none = flask_client.get("/api/projects/default/export-rlog-weekly?module=NONEXISTENT_XYZ")
    assert r_none.status_code == 200
    assert r_none.data[:2] == b"PK"
