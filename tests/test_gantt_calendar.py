"""Tests cho analyzer.gantt_calendar — compute payload cho Gantt Calendar UI."""
from __future__ import annotations

from datetime import date, timedelta

import pytest

from analyzer.gantt_calendar import (
    CATEGORY_COLORS,
    _build_columns,
    _build_month_spans,
    _build_week_spans,
    _choose_granularity,
    _month_add,
    _week_start,
    compute_gantt_calendar,
)
from parser.excel_parser import (
    FunctionRow, ParsedData, PhaseData, PhaseGroup,
)


# ==========================================================================
# Helpers — build ParsedData nhân tạo (không cần Excel thật)
# ==========================================================================
def _make_data(today: date, funcs: list[dict]) -> ParsedData:
    """
    Build ParsedData từ list function config.

    Mỗi function config: {
      "ma_cn": "F1", "module": "M1", "quy_trinh": "P1",
      "phases": {
        "Analysis": {"start": -30, "end": -20, "status": "Closed"},
        "Dev":      {"start": -18, "end":  -5, "status": "In-progress"},
        ...
      }
    }
    """
    rows = []
    all_modules_seen = []
    all_phases_seen: list[str] = []
    for i, fc in enumerate(funcs, 1):
        phases = {}
        for ph_name, ph_cfg in fc.get("phases", {}).items():
            phases[ph_name] = PhaseData(
                start_date=(today + timedelta(days=ph_cfg["start"])) if ph_cfg.get("start") is not None else None,
                end_date=(today + timedelta(days=ph_cfg["end"])) if ph_cfg.get("end") is not None else None,
                status=ph_cfg.get("status"),
                pics=ph_cfg.get("pics", []),
            )
            if ph_name not in all_phases_seen:
                all_phases_seen.append(ph_name)
        rows.append(FunctionRow(row_num=i + 1, meta={
            "ma_cn": fc.get("ma_cn", f"F{i}"),
            "ten_cn": fc.get("ten_cn", fc.get("ma_cn", f"F{i}")),
            "module": fc.get("module", "M1"),
            "quy_trinh": fc.get("quy_trinh", ""),
        }, phases=phases))
        m = fc.get("module")
        if m and m not in all_modules_seen:
            all_modules_seen.append(m)

    return ParsedData(
        headers={}, meta_columns={},
        phase_groups=[PhaseGroup(name=p, attributes={}) for p in all_phases_seen],
        rows=rows,
        all_modules=all_modules_seen,
        all_phases=all_phases_seen,
        all_pics=[], all_statuses=[], all_priorities=[], all_complexities=[],
        all_giai_doan=[], all_processes=[],
    )


TODAY = date(2026, 7, 30)


# ==========================================================================
# 1. Column building
# ==========================================================================
def test_choose_granularity_thresholds():
    d0 = date(2026, 1, 1)
    assert _choose_granularity(d0, d0 + timedelta(days=30)) == "day"
    assert _choose_granularity(d0, d0 + timedelta(days=200)) == "week"
    assert _choose_granularity(d0, d0 + timedelta(days=800)) == "month"


def test_build_columns_day_granularity():
    cols = _build_columns(date(2026, 6, 1), date(2026, 6, 7), "day")
    assert len(cols) == 7
    assert cols[0]["label"] == "01-Jun"
    assert cols[-1]["label"] == "07-Jun"
    for c in cols:
        assert c["start"] == c["end"]  # day granularity: start = end


def test_build_columns_week_granularity_snaps_to_monday():
    # 1/6/2026 là thứ Hai → start-of-week = chính nó
    cols = _build_columns(date(2026, 6, 1), date(2026, 6, 21), "week")
    # 3 tuần: W23, W24, W25
    assert len(cols) == 3
    assert all(c["label"].startswith("W") for c in cols)
    # Column 0 start phải là Monday
    assert date.fromisoformat(cols[0]["start"]).weekday() == 0


def test_build_columns_month_granularity():
    cols = _build_columns(date(2026, 6, 1), date(2026, 8, 31), "month")
    assert len(cols) == 3
    assert cols[0]["label"] == "Jun-26"
    assert cols[-1]["label"] == "Aug-26"


def test_month_spans_group_columns():
    cols = _build_columns(date(2026, 6, 1), date(2026, 7, 31), "week")
    spans = _build_month_spans(cols)
    labels = [s["label"] for s in spans]
    assert "Jun-26" in labels and "Jul-26" in labels
    # Tổng colspan = len(cols)
    assert sum(s["colspan"] for s in spans) == len(cols)


def test_week_spans_only_for_day_granularity():
    cols = _build_columns(date(2026, 6, 1), date(2026, 6, 14), "day")
    ws = _build_week_spans(cols, "day")
    assert sum(s["colspan"] for s in ws) == len(cols)
    # Ít nhất 2 tuần khác nhau (14 ngày sẽ vắt qua 3 tuần ISO)
    assert len(ws) >= 2

    # Không có week_spans cho granularity != day
    cols_w = _build_columns(date(2026, 6, 1), date(2026, 6, 21), "week")
    assert _build_week_spans(cols_w, "week") == []


# ==========================================================================
# 2. compute_gantt_calendar — end-to-end
# ==========================================================================
def test_gantt_calendar_5_functions_range_detect():
    """5 function Start/End khác nhau → range = [minStart .. maxEnd]."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {"Analysis": {"start": -60, "end": -30, "status": "Closed"}}},
        {"ma_cn": "F2", "module": "M1", "phases": {"Analysis": {"start": -50, "end": -20, "status": "Closed"}}},
        {"ma_cn": "F3", "module": "M2", "phases": {"Dev":      {"start": -40, "end": -10, "status": "In-progress"}}},
        {"ma_cn": "F4", "module": "M2", "phases": {"UAT":      {"start": -20, "end":  10, "status": "Assigned"}}},
        {"ma_cn": "F5", "module": "M2", "phases": {"Golive":   {"start":  20, "end":  40, "status": "Open"}}},
    ])
    payload = compute_gantt_calendar(data, group_by="function", granularity="week", today=TODAY)
    # min date = TODAY-60 (F1), max date = TODAY+40 (F5)
    assert date.fromisoformat(payload["min_date"]) <= TODAY - timedelta(days=60)
    assert date.fromisoformat(payload["max_date"]) >= TODAY + timedelta(days=40)
    # today_col phải nằm trong range
    assert payload["today_col"] is not None
    assert 0 <= payload["today_col"] < len(payload["columns"])


def test_gantt_calendar_group_by_module_aggregates():
    """Group by module → mỗi module 1 row, aggregate min/max/pct."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -30, "end": -25, "status": "Closed"},
            "Dev":      {"start": -20, "end": -10, "status": "Closed"},
            "UAT":      {"start":  -5, "end":   5, "status": "In-progress"},
        }},
        {"ma_cn": "F2", "module": "M1", "phases": {
            "Analysis": {"start": -40, "end": -35, "status": "Closed"},
            "Dev":      {"start": -30, "end": -20, "status": "Closed"},
            "UAT":      {"start": -15, "end":  -5, "status": "In-progress"},
        }},
        {"ma_cn": "F3", "module": "M2", "phases": {
            "Analysis": {"start": -20, "end":  -15, "status": "Closed"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="week", today=TODAY)
    rows = payload["rows"]
    # 2 module → 2 row
    assert len(rows) == 2
    m1 = next(r for r in rows if r["module"] == "M1")
    m2 = next(r for r in rows if r["module"] == "M2")
    # M1 aggregate: start = min(-40) → TODAY-40, end = max(5) → TODAY+5
    assert date.fromisoformat(m1["start"]) == TODAY - timedelta(days=40)
    assert date.fromisoformat(m1["end"]) == TODAY + timedelta(days=5)
    # M1: 2 function × 3 phase = 6 slots; closed=4 (Analysis*2, Dev*2) → 66%
    assert m1["pct"] == pytest.approx(round(4 / 6 * 100), abs=1)
    # M1 category = summary (aggregate mode)
    assert m1["category"] == "summary"
    # M1 func_count = 2
    assert m1["func_count"] == 2
    # M2 chỉ 1 func, Analysis Closed → pct = 1/3
    assert m2["func_count"] == 1


def test_gantt_calendar_granularity_week_num_columns():
    """granularity=week → số cột ≈ ceil((max-min)/7). Cho range 10 tuần chẵn."""
    data = _make_data(TODAY, [
        # 10 tuần = 70 ngày
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -35, "end": 35, "status": "In-progress"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="week", today=TODAY)
    # min snap về Monday, max snap về Sunday → 11 tuần (bao gồm cả tuần vượt biên)
    n_cols = len(payload["columns"])
    assert 10 <= n_cols <= 12  # cho phép ±1 do snap tuần
    # Mỗi cột span 7 ngày
    for c in payload["columns"]:
        start = date.fromisoformat(c["start"])
        end = date.fromisoformat(c["end"])
        assert (end - start).days == 6


def test_gantt_calendar_cells_overlap_row_range():
    """Cell active khi cột overlap [row.start, row.end]."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -14, "end": -1, "status": "In-progress"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="function", granularity="week", today=TODAY)
    row = payload["rows"][0]
    active_indices = [i for i, a in enumerate(row["cells"]) if a]
    assert len(active_indices) >= 2  # ít nhất 2 tuần được cover
    # span index đúng với first/last active
    assert row["span_start_col"] == active_indices[0]
    assert row["span_end_col"] == active_indices[-1]


def test_gantt_calendar_today_marker_in_range():
    """Nếu today nằm trong range, today_col != None và inside columns."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -30, "end": 30, "status": "In-progress"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="function", granularity="week", today=TODAY)
    assert payload["today_col"] is not None
    col = payload["columns"][payload["today_col"]]
    assert col["start"] <= TODAY.isoformat() <= col["end"]


def test_gantt_calendar_empty_dataset_returns_skeleton():
    """Không có phase Start/End nào → empty flag + rỗng cột/row."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": None, "end": None, "status": "Open"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="week", today=TODAY)
    assert payload["empty"] is True
    assert payload["columns"] == []
    assert payload["rows"] == []


def test_gantt_calendar_group_by_process():
    """Group by process → 1 row / (module, quy_trinh)."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "quy_trinh": "P1", "phases": {"Analysis": {"start": -10, "end": -1, "status": "Closed"}}},
        {"ma_cn": "F2", "module": "M1", "quy_trinh": "P1", "phases": {"Analysis": {"start": -8, "end": -2, "status": "Closed"}}},
        {"ma_cn": "F3", "module": "M1", "quy_trinh": "P2", "phases": {"Dev":      {"start": -6, "end": -1, "status": "Open"}}},
    ])
    payload = compute_gantt_calendar(data, group_by="process", granularity="week", today=TODAY)
    # 2 quy trình → 2 row
    assert len(payload["rows"]) == 2
    names = {r["name"] for r in payload["rows"]}
    assert any("P1" in n for n in names)
    assert any("P2" in n for n in names)


def test_gantt_calendar_group_by_phan_he_alias():
    """`phan_he` là alias của `module`."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "MX", "phases": {"Analysis": {"start": -10, "end": -1, "status": "Closed"}}},
    ])
    p1 = compute_gantt_calendar(data, group_by="module", granularity="week", today=TODAY)
    p2 = compute_gantt_calendar(data, group_by="phan_he", granularity="week", today=TODAY)
    assert p1["group_by"] == "module"
    assert p2["group_by"] == "phan_he"
    # Số row giống nhau
    assert len(p1["rows"]) == len(p2["rows"]) == 1


def test_gantt_calendar_category_for_function_mode_reflects_active_phase():
    """Function mode: category theo phase đang active."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -30, "end": -20, "status": "Closed"},
            "Dev":      {"start": -15, "end":  -1, "status": "In-progress"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="function", granularity="week", today=TODAY)
    row = payload["rows"][0]
    # Dev đang In-progress → category = phase2 (Lập trình)
    assert row["category"] == "phase2"
    assert row["active_phase"] == "Dev"


def test_gantt_calendar_legend_matches_category_colors():
    """Legend trả về đủ 6 category với màu khớp CATEGORY_COLORS."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {"Analysis": {"start": -5, "end": 5, "status": "Open"}}},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="week", today=TODAY)
    lg = payload["legend"]
    for key in ("phase1", "phase2", "phase3", "milestone", "summary", "idle"):
        assert key in lg
        assert lg[key]["color"] == CATEGORY_COLORS[key]


def test_gantt_calendar_auto_granularity_short_range():
    """Range 30 ngày → auto chọn 'day'."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -10, "end": 10, "status": "In-progress"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="auto", today=TODAY)
    assert payload["granularity"] == "day"
    # Có week_spans khi granularity=day
    assert len(payload["week_spans"]) > 0


def test_gantt_calendar_auto_granularity_long_range():
    """Range > 400 ngày → auto chọn 'month'."""
    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -200, "end": 250, "status": "In-progress"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="auto", today=TODAY)
    assert payload["granularity"] == "month"


# ==========================================================================
# T35 Task 2 — Outlier date detection (year 1936 bug)
# ==========================================================================
# (covered by TestOutlierDateDetection bên dưới)


def test_gantt_calendar_export_writes_valid_xlsx(tmp_path):
    """Smoke test: export_gantt_calendar_report tạo file .xlsx hợp lệ."""
    import io
    import openpyxl
    from exporter.excel_exporter import export_gantt_calendar_report

    data = _make_data(TODAY, [
        {"ma_cn": "F1", "module": "M1", "phases": {
            "Analysis": {"start": -20, "end": -10, "status": "Closed"},
            "Dev":      {"start":  -8, "end":   5, "status": "In-progress"},
        }},
        {"ma_cn": "F2", "module": "M2", "phases": {
            "Analysis": {"start": -15, "end":  -5, "status": "Closed"},
        }},
    ])
    payload = compute_gantt_calendar(data, group_by="module", granularity="week", today=TODAY)
    filepath = export_gantt_calendar_report(payload, output_dir=str(tmp_path), subtitle="Test")
    # File tồn tại + đọc được bằng openpyxl
    with open(filepath, "rb") as f:
        raw = f.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert "GanttCalendar" in wb.sheetnames
    ws = wb["GanttCalendar"]
    # Row 4 là header — cell A4 phải có value label
    assert ws.cell(row=4, column=1).value  # non-empty
    # Có ít nhất 2 data row cho 2 module
    # Data start: row 4 + n_header_rows (week=2). data rows từ row 6.
    data_row_labels = [ws.cell(row=6 + i, column=1).value for i in range(2)]
    assert any("M1" in (v or "") for v in data_row_labels)
    assert any("M2" in (v or "") for v in data_row_labels)
    wb.close()


# ==========================================================================
# 3. API endpoint smoke test
# ==========================================================================
def test_gantt_calendar_endpoint_returns_json(flask_client, sample_xlsx_path):
    """POST /upload rồi GET /gantt-calendar phải trả JSON hợp lệ."""
    import io
    with open(sample_xlsx_path, "rb") as f:
        flask_client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "sample.xlsx")},
            content_type="multipart/form-data",
        )
    r = flask_client.get("/api/projects/default/gantt-calendar?group_by=module&granularity=week")
    assert r.status_code == 200
    payload = r.get_json()
    assert "columns" in payload
    assert "rows" in payload
    assert "legend" in payload
    assert payload["group_by"] == "module"


def test_gantt_calendar_export_endpoint_returns_xlsx(flask_client, sample_xlsx_path):
    """GET /export-gantt-calendar trả file xlsx (Content-Type spreadsheet)."""
    import io
    with open(sample_xlsx_path, "rb") as f:
        flask_client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "sample.xlsx")},
            content_type="multipart/form-data",
        )
    r = flask_client.get("/api/projects/default/export-gantt-calendar?group_by=module&granularity=week")
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ctype or "octet-stream" in ctype


# ==========================================================================
# 4. T35 Task 2 — Outlier date detection
# ==========================================================================


class TestOutlierDateDetection:
    """
    Verify range Gantt Calendar không bị kéo dài 90 năm khi data có date
    năm < 2000 hoặc > (current_year + 10).

    Bug thực tế: API iHRP Task Daily trả record có Config UAT.Start='1936-03-26'
    (do source system rounding empty date thành 1899-12-30 + delta) → khi ko
    filter, min_date=1936 max_date=2027 → 90 năm timeline không đọc được.
    """

    def test_is_outlier_helper(self):
        from analyzer.gantt_calendar import _is_outlier_date
        today = date(2026, 7, 30)
        # Year quá cũ
        assert _is_outlier_date(date(1936, 3, 26), today) is True
        assert _is_outlier_date(date(1999, 12, 31), today) is True
        # Year quá xa tương lai
        assert _is_outlier_date(date(2050, 1, 1), today) is True
        # Trong range hợp lệ [2000, 2036]
        assert _is_outlier_date(date(2000, 1, 1), today) is False
        assert _is_outlier_date(date(2026, 7, 30), today) is False
        assert _is_outlier_date(date(2035, 12, 31), today) is False
        assert _is_outlier_date(date(2036, 1, 1), today) is False  # đúng ranh giới
        # None → False (không throw)
        assert _is_outlier_date(None, today) is False

    def test_outlier_excluded_from_range(self):
        """
        Function có 2 phase: 1 phase date 2026 (bình thường), 1 phase date 1936
        (outlier) → range chỉ tính 2026, KHÔNG kéo dài về 1936.
        """
        # Custom function with 1936 date — build tay không dùng _make_data
        # (helper dùng timedelta relative to today).
        # LƯU Ý: _group_rows dùng meta.get("module") lowercase → phải set
        # meta lowercase, không viết "Module" (capitalized).
        row = FunctionRow(row_num=2, meta={
            "ma_cn": "F1", "module": "HR",
        }, phases={
            "Analysis": PhaseData(
                start_date=date(2026, 5, 1),
                end_date=date(2026, 6, 1),
                status="Closed",
            ),
            "Config UAT": PhaseData(
                start_date=date(1936, 3, 26),   # outlier
                end_date=date(1936, 3, 26),     # outlier
                status="Closed",
            ),
        })
        data = ParsedData(
            headers={}, meta_columns={},
            phase_groups=[
                PhaseGroup(name="Analysis", attributes={}),
                PhaseGroup(name="Config UAT", attributes={}),
            ],
            rows=[row],
            all_modules=["HR"],
            all_phases=["Analysis", "Config UAT"],
            all_pics=[], all_statuses=[], all_priorities=[],
            all_complexities=[], all_giai_doan=[], all_processes=[],
        )
        payload = compute_gantt_calendar(
            data, group_by="module", granularity="month",
            today=date(2026, 7, 30),
        )
        # Range max KHÔNG được lùi về 1936 → phải trong 2020-2040
        min_year = int(payload["min_date"].split("-")[0])
        max_year = int(payload["max_date"].split("-")[0])
        assert min_year >= 2020, f"min_date {payload['min_date']} bị kéo về outlier"
        assert max_year <= 2040, f"max_date {payload['max_date']} bị kéo về outlier"

    def test_outlier_reported_in_skipped_dates(self):
        row = FunctionRow(row_num=2, meta={
            "ma_cn": "F999", "ten_cn": "Test outlier",
            "module": "HR",
            # Tương thích thêm với _collect_dates_with_outliers (đọc cả Mã CN)
            "Mã CN": "F999", "Tên chức năng": "Test outlier",
            "Module": "HR",
        }, phases={
            "Config UAT": PhaseData(
                start_date=date(1936, 3, 26),
                end_date=date(1936, 3, 26),
                status="Closed",
            ),
            "Analysis": PhaseData(
                start_date=date(2026, 5, 1),
                end_date=date(2026, 6, 1),
                status="Closed",
            ),
        })
        data = ParsedData(
            headers={}, meta_columns={},
            phase_groups=[
                PhaseGroup(name="Analysis", attributes={}),
                PhaseGroup(name="Config UAT", attributes={}),
            ],
            rows=[row],
            all_modules=["HR"], all_phases=["Analysis", "Config UAT"],
            all_pics=[], all_statuses=[], all_priorities=[],
            all_complexities=[], all_giai_doan=[], all_processes=[],
        )
        payload = compute_gantt_calendar(
            data, group_by="module", today=date(2026, 7, 30),
        )
        # skipped_dates + skipped_count present
        assert "skipped_dates" in payload
        assert "skipped_count" in payload
        assert payload["skipped_count"] == 2  # 2 field (start + end) outlier
        # Mỗi entry có field cần thiết
        for entry in payload["skipped_dates"]:
            assert "ma_cn" in entry
            assert "phase" in entry
            assert "attr" in entry
            assert "value" in entry
            assert entry["ma_cn"] == "F999"
            assert entry["phase"] == "Config UAT"
            assert entry["value"].startswith("1936")

    def test_no_outlier_empty_skipped_list(self):
        """Data sạch → skipped_dates = [] và skipped_count = 0."""
        data = _make_data(date(2026, 7, 30), [
            {"ma_cn": "F1", "module": "HR", "phases": {
                "Analysis": {"start": -30, "end": -20, "status": "Closed"},
                "Dev":      {"start": -10, "end":  5,  "status": "In-progress"},
            }},
        ])
        payload = compute_gantt_calendar(
            data, group_by="module", today=date(2026, 7, 30),
        )
        assert payload["skipped_count"] == 0
        assert payload["skipped_dates"] == []

    def test_row_aggregate_ignores_outlier_dates(self):
        """
        Aggregate row start/end phải bỏ qua outlier — nếu 1 function có
        Analysis 2026 + Config UAT 1936 → agg.start = 2026 (NOT 1936).
        """
        row = FunctionRow(row_num=2, meta={
            "ma_cn": "F1", "module": "HR",
        }, phases={
            "Analysis": PhaseData(
                start_date=date(2026, 5, 1),
                end_date=date(2026, 6, 1),
                status="Closed",
            ),
            "Config UAT": PhaseData(
                start_date=date(1936, 3, 26),   # outlier
                end_date=date(1936, 3, 26),
                status="Closed",
            ),
        })
        data = ParsedData(
            headers={}, meta_columns={},
            phase_groups=[
                PhaseGroup(name="Analysis", attributes={}),
                PhaseGroup(name="Config UAT", attributes={}),
            ],
            rows=[row],
            all_modules=["HR"], all_phases=["Analysis", "Config UAT"],
            all_pics=[], all_statuses=[], all_priorities=[],
            all_complexities=[], all_giai_doan=[], all_processes=[],
        )
        payload = compute_gantt_calendar(
            data, group_by="module", today=date(2026, 7, 30),
        )
        # Có 1 row aggregate cho module HR — start/end phải là 2026, không 1936
        assert len(payload["rows"]) == 1
        agg_start = payload["rows"][0]["start"]
        agg_end = payload["rows"][0]["end"]
        assert agg_start is not None and agg_start.startswith("2026"), (
            f"Aggregate row start={agg_start} — không được là 1936"
        )
        assert agg_end is not None and agg_end.startswith("2026"), (
            f"Aggregate row end={agg_end} — không được là 1936"
        )

    def test_all_outliers_treated_as_empty(self):
        """
        Function chỉ có outlier date → coi như không có date → empty=True
        (giống case function không có phase nào).
        """
        row = FunctionRow(row_num=2, meta={"ma_cn": "F1", "module": "HR"}, phases={
            "Analysis": PhaseData(
                start_date=date(1936, 3, 26),
                end_date=date(1936, 3, 26),
                status="Closed",
            ),
        })
        data = ParsedData(
            headers={}, meta_columns={},
            phase_groups=[PhaseGroup(name="Analysis", attributes={})],
            rows=[row],
            all_modules=["HR"], all_phases=["Analysis"],
            all_pics=[], all_statuses=[], all_priorities=[],
            all_complexities=[], all_giai_doan=[], all_processes=[],
        )
        payload = compute_gantt_calendar(
            data, group_by="module", today=date(2026, 7, 30),
        )
        assert payload["empty"] is True
        assert payload["skipped_count"] == 2  # start + end đều outlier
