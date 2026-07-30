"""
T34 Task 3B — Tests for analyzer/type_infer.py.

Tests for infer_type + compatible_ihrp_cols + infer_all_headers +
validate_mapping_dry_run.
"""
from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pytest

from analyzer.type_infer import (
    _classify_single,
    _RE_DATE_ISO,
    compatible_ihrp_cols,
    infer_all_headers,
    infer_type,
    validate_mapping_dry_run,
)


class TestClassifySingle:
    def test_none(self):
        assert _classify_single(None) == "empty"

    def test_empty_string(self):
        assert _classify_single("") == "empty"
        assert _classify_single("   ") == "empty"

    def test_bool_native(self):
        assert _classify_single(True) == "boolean"
        assert _classify_single(False) == "boolean"

    def test_int_native(self):
        assert _classify_single(42) == "integer"
        assert _classify_single(-1) == "integer"
        assert _classify_single(0) == "integer"

    def test_float_native(self):
        assert _classify_single(3.14) == "decimal"
        assert _classify_single(-0.5) == "decimal"

    def test_float_excel_serial(self):
        # Serial 45000 ~ 2023-03-15
        assert _classify_single(45000.0) == "date_excel_serial"
        assert _classify_single(46500.0) == "date_excel_serial"

    def test_datetime_native(self):
        assert _classify_single(datetime(2026, 1, 15)) == "date_iso"
        assert _classify_single(date(2026, 3, 20)) == "date_iso"

    def test_string_date_iso(self):
        assert _classify_single("2026-07-30") == "date_iso"
        assert _classify_single("2026/01/15") == "date_iso"

    def test_string_date_dmy(self):
        assert _classify_single("30/07/2026") == "date_dmy"
        assert _classify_single("15-01-26") == "date_dmy"

    def test_string_int(self):
        assert _classify_single("42") == "integer"
        assert _classify_single("-100") == "integer"

    def test_string_decimal(self):
        assert _classify_single("3.14") == "decimal"
        assert _classify_single("100,5") == "decimal"

    def test_string_status_enum(self):
        assert _classify_single("Open") == "status_enum"
        assert _classify_single("In-progress") == "status_enum"
        assert _classify_single("closed") == "status_enum"
        assert _classify_single("CANCELLED") == "status_enum"

    def test_string_boolean(self):
        assert _classify_single("TRUE") == "boolean"
        assert _classify_single("no") == "boolean"
        assert _classify_single("Yes") == "boolean"

    def test_string_pic_list(self):
        assert _classify_single("Alice, Bob") == "pic_list"
        assert _classify_single("BaoLQ31; NhiVN") == "pic_list"
        assert _classify_single("Cai\nDee\nEve") == "pic_list"
        assert _classify_single("CuongNM129+ TungTT83") == "pic_list"

    def test_string_single_pic_not_list(self):
        # 1 tên đơn không phải pic_list
        assert _classify_single("Alice") == "string"

    def test_string_arbitrary(self):
        assert _classify_single("Chấm công app") == "string"


class TestInferType:
    def test_empty_list(self):
        assert infer_type([]) == "empty"

    def test_all_empty_values(self):
        assert infer_type([None, "", "   "]) == "empty"

    def test_all_date_iso(self):
        assert infer_type(["2026-01-15", "2026-03-20", "2026-06-01"]) == "date_iso"

    def test_all_date_dmy(self):
        assert infer_type(["15/01/2026", "20/03/2026", "01/06/2026"]) == "date_iso"

    def test_mixed_date_formats_canonicalized(self):
        # date_iso + date_dmy → gộp thành date_iso
        assert infer_type(["2026-01-15", "20/03/2026", "2026-06-01"]) == "date_iso"

    def test_all_integer(self):
        assert infer_type(["10", "20", "30"]) == "integer"

    def test_all_status(self):
        assert infer_type(["Open", "Closed", "In-progress"]) == "status_enum"

    def test_all_pic_list(self):
        assert infer_type(["Alice, Bob", "Cai; Dee", "Eve\nFred"]) == "pic_list"

    def test_majority_wins(self):
        # 3 date + 1 string → date (75% > 60%)
        assert infer_type(["2026-01-15", "2026-02-20", "2026-03-30", "random text"]) == "date_iso"

    def test_mix_falls_back_to_string(self):
        # Mix nhiều loại → string
        assert infer_type(["hello", "42", "2026-01-15"]) == "string"

    def test_empty_values_ignored(self):
        # Empty ignored, còn lại all int → integer
        assert infer_type([None, "10", "20", "", "30"]) == "integer"

    def test_datetime_objects(self):
        assert infer_type([datetime(2026, 1, 1), datetime(2026, 2, 1)]) == "date_iso"


class TestCompatibleIhrpCols:
    def test_date_returns_start_end_columns(self):
        cols = compatible_ihrp_cols("date_iso")
        assert "Analysis - Start" in cols
        assert "Dev - End" in cols
        assert "UAT - Start" in cols
        assert "Last Updated Date" in cols
        # Không có PIC / Status trong đây
        assert "Analysis - PIC" not in cols
        assert "Dev - Status" not in cols

    def test_pic_returns_pic_columns(self):
        cols = compatible_ihrp_cols("pic_list")
        assert "Analysis - PIC" in cols
        assert "Dev - PIC" in cols
        assert "UAT - PIC" in cols
        # Không có Start/End
        assert "Analysis - Start" not in cols

    def test_status_returns_status_columns(self):
        cols = compatible_ihrp_cols("status_enum")
        assert "Analysis - Status" in cols
        assert "UAT - Status" in cols

    def test_integer_returns_estimate_mh_and_priority(self):
        cols = compatible_ihrp_cols("integer")
        assert "Analysis - Estimate MH" in cols
        assert "Priority" in cols
        assert "Giai đoạn" in cols

    def test_string_returns_all_ihrp_columns(self):
        cols = compatible_ihrp_cols("string")
        assert "Mã CN" in cols
        assert "Analysis - Start" in cols
        assert "Dev - PIC" in cols

    def test_empty_returns_empty(self):
        assert compatible_ihrp_cols("empty") == []

    def test_boolean_returns_empty(self):
        # Boolean không có col iHRP tương ứng
        assert compatible_ihrp_cols("boolean") == []

    def test_excel_serial_treated_as_date(self):
        cols = compatible_ihrp_cols("date_excel_serial")
        assert "Analysis - Start" in cols


class TestInferAllHeaders:
    def test_basic_shape(self):
        headers = ["Ma CN", "Deadline", "PIC"]
        preview = [
            ["A.01", "2026-07-30", "Alice, Bob"],
            ["A.02", "2026-08-15", "Cai; Dee"],
            ["A.03", "2026-09-01", "Eve"],
        ]
        result = infer_all_headers(headers, preview)
        assert "Ma CN" in result
        assert "Deadline" in result
        assert "PIC" in result
        # Deadline → date
        assert result["Deadline"]["type"] == "date_iso"
        # Ma CN → string (không match pattern nào)
        assert result["Ma CN"]["type"] == "string"

    def test_sample_values_included(self):
        headers = ["X"]
        preview = [["v1"], ["v2"], ["v3"], ["v4"]]
        result = infer_all_headers(headers, preview)
        # Chỉ lấy 3 sample
        assert len(result["X"]["samples"]) == 3

    def test_badge_included(self):
        result = infer_all_headers(["date_col"], [["2026-01-15"]])
        assert "badge" in result["date_col"]
        assert result["date_col"]["badge"]["icon"] == "📅"

    def test_empty_header_skipped(self):
        result = infer_all_headers(["A", "", "B"], [["1", "x", "2"]])
        assert "A" in result
        assert "B" in result
        assert "" not in result

    def test_non_empty_samples_preferred(self):
        headers = ["X"]
        preview = [[None], [""], ["actual"], ["value"]]
        result = infer_all_headers(headers, preview)
        # Phải ưu tiên non-empty
        assert "actual" in result["X"]["samples"] or "value" in result["X"]["samples"]


class TestValidateMappingDryRun:
    @pytest.fixture
    def excel_path(self, tmp_path):
        """Tạo Excel với header lạ để test mapping."""
        path = tmp_path / "dry-run.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Function List"
        # Header lạ — không chuẩn iHRP
        ws["A1"] = "Function Code"      # → Mã CN
        ws["B1"] = "Function Name"       # → Tên chức năng
        ws["C1"] = "Sub-System"          # → Module
        ws["D1"] = "Analysis Start"      # → Analysis - Start
        ws["E1"] = "Analysis Deadline"   # → Analysis - End
        ws["F1"] = "Analysis Status"     # → Analysis - Status
        ws["G1"] = "Analysis Assignee"   # → Analysis - PIC
        # Data
        ws["A2"] = "FN.01"; ws["B2"] = "Login"; ws["C2"] = "HR"
        ws["D2"] = date(2026, 1, 1); ws["E2"] = date(2026, 1, 15)
        ws["F2"] = "Closed"; ws["G2"] = "Alice"
        ws["A3"] = "FN.02"; ws["B3"] = "Logout"; ws["C3"] = "HR"
        ws["D3"] = date(2026, 2, 1); ws["E3"] = date(2026, 2, 20)
        ws["F3"] = "In-progress"; ws["G3"] = "Bob"
        wb.save(str(path))
        wb.close()
        return str(path)

    def test_dry_run_with_valid_mapping(self, excel_path):
        mapping = {
            "Mã CN": "Function Code",
            "Tên chức năng": "Function Name",
            "Module": "Sub-System",
            "Analysis - Start": "Analysis Start",
            "Analysis - End": "Analysis Deadline",
            "Analysis - Status": "Analysis Status",
            "Analysis - PIC": "Analysis Assignee",
        }
        result = validate_mapping_dry_run(excel_path, mapping, n_rows=5)
        assert result["success"] is True
        assert result["row_count_scanned"] == 2  # chỉ 2 record trong file
        rows = result["rows"]
        assert len(rows) == 2
        assert rows[0]["ma_cn"] == "FN.01"
        assert rows[0]["ten_cn"] == "Login"
        assert rows[0]["module"] == "HR"
        # Analysis phase
        an = rows[0]["phases"].get("Analysis")
        assert an is not None
        assert an["status"] == "Closed"
        assert "Alice" in an["pic"]

    def test_warning_when_header_not_in_file(self, excel_path):
        mapping = {
            "Mã CN": "Function Code",
            "Dev - Start": "Non-Existent Column",  # header không tồn tại
        }
        result = validate_mapping_dry_run(excel_path, mapping, n_rows=5)
        # Có warning về "Non-Existent Column"
        assert len(result["warnings"]) >= 1
        assert any("Non-Existent Column" in w for w in result["warnings"])

    def test_no_mapping_uses_auto_detect(self, excel_path):
        """Empty mapping → parser auto-detect (best-effort)."""
        result = validate_mapping_dry_run(excel_path, {}, n_rows=5)
        assert result["success"] is True
        # Auto-detect có thể không nhận được hết vì headers không chuẩn
        # → chấp nhận rows có thể có ma_cn rỗng hoặc partial

    def test_n_rows_clamp(self, excel_path):
        result = validate_mapping_dry_run(excel_path, {"Mã CN": "Function Code"}, n_rows=100)
        # File chỉ có 2 record → return 2
        assert result["row_count_scanned"] == 2

    def test_result_json_serializable(self, excel_path):
        """Verify dry-run result serializable qua json.dumps."""
        import json
        mapping = {"Mã CN": "Function Code"}
        result = validate_mapping_dry_run(excel_path, mapping)
        json.dumps(result)  # Must not raise

    def test_invalid_file_returns_error(self, tmp_path):
        # File không tồn tại
        result = validate_mapping_dry_run(str(tmp_path / "no.xlsx"), {})
        assert result["success"] is False
        assert len(result["errors"]) >= 1


# ==========================================================================
# HTTP integration tests
# ==========================================================================

def _upload_preview(client, xlsx_path):
    with open(xlsx_path, "rb") as f:
        return client.post(
            "/api/upload-preview",
            data={"file": (io.BytesIO(f.read()), "test.xlsx")},
            content_type="multipart/form-data",
        )


class TestUploadPreviewIncludesColumnTypes:
    def test_preview_returns_column_types(self, flask_client, sample_xlsx_path):
        r = _upload_preview(flask_client, sample_xlsx_path)
        assert r.status_code == 200
        d = r.get_json()
        assert "column_types" in d
        # Mỗi header phải có info type
        ct = d["column_types"]
        assert len(ct) > 0
        for header, info in ct.items():
            assert "type" in info
            assert "badge" in info
            assert "samples" in info

    def test_column_types_detect_date(self, flask_client, sample_xlsx_path):
        r = _upload_preview(flask_client, sample_xlsx_path)
        d = r.get_json()
        # sample_xlsx có "Analysis - Start" với date
        ct = d["column_types"]
        # Filter các header có type là date
        date_headers = [h for h, info in ct.items()
                       if info["type"] in ("date_iso", "date_dmy", "date_excel_serial")]
        # Ít nhất Analysis - Start phải là date
        assert any("Start" in h or "End" in h for h in date_headers)


class TestValidateMappingEndpoint:
    def test_validate_missing_tmp_id(self, flask_client):
        r = flask_client.post("/api/validate-mapping", json={})
        assert r.status_code == 400

    def test_validate_bad_tmp_id(self, flask_client):
        r = flask_client.post("/api/validate-mapping",
                              json={"tmp_id": "not-hex-chars-xxx", "column_mapping": {}})
        assert r.status_code == 400

    def test_validate_nonexistent_tmp(self, flask_client):
        r = flask_client.post("/api/validate-mapping",
                              json={"tmp_id": "deadbeefcafe1234", "column_mapping": {}})
        assert r.status_code == 404

    def test_validate_full_flow(self, flask_client, sample_xlsx_path):
        # 1. Upload preview để lấy tmp_id
        r_prev = _upload_preview(flask_client, sample_xlsx_path)
        assert r_prev.status_code == 200
        tmp_id = r_prev.get_json()["tmp_id"]

        # 2. Validate với mapping có mã CN
        r = flask_client.post("/api/validate-mapping", json={
            "tmp_id": tmp_id,
            "column_mapping": {"Mã CN": "Mã CN", "Tên chức năng": "Tên chức năng"},
            "n_rows": 3,
        })
        assert r.status_code == 200
        d = r.get_json()
        assert d["success"] is True
        assert d["row_count_scanned"] >= 1

    def test_validate_n_rows_clamp(self, flask_client, sample_xlsx_path):
        r_prev = _upload_preview(flask_client, sample_xlsx_path)
        tmp_id = r_prev.get_json()["tmp_id"]
        r = flask_client.post("/api/validate-mapping", json={
            "tmp_id": tmp_id,
            "column_mapping": {},
            "n_rows": 999,  # sẽ clamp về 20
        })
        assert r.status_code == 200
        # Sample xlsx chỉ có 6 rows → scanned ≤ 6
        assert r.get_json()["row_count_scanned"] <= 6


class TestIntegrationMappingPresetCRUD:
    def test_list_empty(self, flask_client):
        r = flask_client.get(
            "/api/projects/default/integrations/integ-123/mapping-presets"
        )
        assert r.status_code == 200
        assert r.get_json()["presets"] == []

    def test_save_and_list(self, flask_client):
        r = flask_client.post(
            "/api/projects/default/integrations/integ-123/mapping-presets",
            json={
                "name": "Vendor X preset",
                "mapping": {"Mã CN": "code", "Tên chức năng": "name"},
            },
        )
        assert r.status_code == 201
        presets = r.get_json()["presets"]
        assert len(presets) == 1
        assert presets[0]["name"] == "Vendor X preset"

    def test_isolate_per_integration(self, flask_client):
        """Preset của integ-A không xuất hiện ở integ-B."""
        flask_client.post(
            "/api/projects/default/integrations/integ-A/mapping-presets",
            json={"name": "A preset", "mapping": {"Mã CN": "code_a"}},
        )
        r_b = flask_client.get(
            "/api/projects/default/integrations/integ-B/mapping-presets"
        )
        assert r_b.get_json()["presets"] == []

    def test_delete_preset(self, flask_client):
        flask_client.post(
            "/api/projects/default/integrations/integ-X/mapping-presets",
            json={"name": "todelete", "mapping": {"X": "y"}},
        )
        r = flask_client.delete(
            "/api/projects/default/integrations/integ-X/mapping-presets/todelete"
        )
        assert r.status_code == 200
        # Verify empty
        r_list = flask_client.get(
            "/api/projects/default/integrations/integ-X/mapping-presets"
        )
        assert r_list.get_json()["presets"] == []

    def test_delete_missing_returns_404(self, flask_client):
        r = flask_client.delete(
            "/api/projects/default/integrations/integ-Z/mapping-presets/nonexistent"
        )
        assert r.status_code == 404

    def test_save_missing_name_returns_400(self, flask_client):
        r = flask_client.post(
            "/api/projects/default/integrations/integ-X/mapping-presets",
            json={"mapping": {"X": "y"}},
        )
        assert r.status_code == 400

    def test_save_updates_existing(self, flask_client):
        """POST cùng name → overwrite mapping."""
        for i in range(2):
            r = flask_client.post(
                "/api/projects/default/integrations/integ-X/mapping-presets",
                json={"name": "same", "mapping": {f"K{i}": f"V{i}"}},
            )
            assert r.status_code == 201
        r_list = flask_client.get(
            "/api/projects/default/integrations/integ-X/mapping-presets"
        )
        presets = r_list.get_json()["presets"]
        assert len(presets) == 1  # không duplicate
        # Mapping mới nhất
        assert "K1" in presets[0]["mapping"]
