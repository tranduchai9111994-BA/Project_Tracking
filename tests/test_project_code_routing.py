"""
Tests: multi-project routing theo Mã dự án (project_code_field / map / filter).
"""
from __future__ import annotations

from analyzer import integrations as integ_mod


def test_group_records_by_map():
    records = [
        {"project": "MPHG_IHRP_2025_PM", "functionCode": "A"},
        {"project": "MPHG_IHRP_2025_PM", "functionCode": "B"},
        {"project": "OTHER_X", "functionCode": "C"},
        {"project": "OTHER_X", "functionCode": "D"},
        {"project": "UNKNOWN", "functionCode": "E"},
        {"functionCode": "F"},  # missing project
    ]
    groups, skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field="project",
        project_code_map={
            "MPHG_IHRP_2025_PM": "mphg",
            "OTHER_X": "other",
        },
        default_slug="mphg",
    )
    assert set(groups.keys()) == {"mphg", "other"}
    assert len(groups["mphg"]) == 2
    assert len(groups["other"]) == 2
    # UNKNOWN + empty
    reasons = {(s["code"], s["reason"]): s["count"] for s in skipped}
    assert reasons[("UNKNOWN", "unmapped")] == 1
    assert reasons[("(empty)", "empty")] == 1


def test_group_records_filter_only():
    records = [
        {"project": "MPHG_IHRP_2025_PM", "id": 1},
        {"project": "OTHER", "id": 2},
        {"project": "MPHG_IHRP_2025_PM", "id": 3},
    ]
    groups, skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field="project",
        project_code_map={},
        project_code_filter="MPHG_IHRP_2025_PM",
        default_slug="mphg",
    )
    assert list(groups.keys()) == ["mphg"]
    assert len(groups["mphg"]) == 2
    assert any(s["reason"] == "filtered" and s["count"] == 1 for s in skipped)


def test_group_records_filter_plus_map():
    """Filter thu hẹp trước; map quyết định slug."""
    records = [
        {"project": "KEEP", "id": 1},
        {"project": "DROP", "id": 2},
        {"project": "KEEP", "id": 3},
    ]
    groups, skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field="project",
        project_code_map={"KEEP": "alpha"},
        project_code_filter="KEEP",
        default_slug="default",
    )
    assert groups == {"alpha": [records[0], records[2]]}
    assert any(s["reason"] == "filtered" for s in skipped)


def test_sanitize_endpoint_keeps_project_routing():
    ep = integ_mod._sanitize_endpoint({
        "name": "FL",
        "path": "/api/x",
        "response_type": "json",
        "data_path": "data",
        "field_mapping": {"Mã CN": "functionCode", "Mã dự án": "project"},
        "project_code_field": "project",
        "project_code_map": {"MPHG_IHRP_2025_PM": "mphg", "": "bad", "X": "Bad Slug"},
        "project_code_filter": "MPHG_IHRP_2025_PM",
    })
    assert ep is not None
    assert ep["project_code_field"] == "project"
    assert ep["project_code_filter"] == "MPHG_IHRP_2025_PM"
    assert ep["project_code_map"] == {"MPHG_IHRP_2025_PM": "mphg"}
    assert ep["field_mapping"]["Mã dự án"] == "project"


def test_group_records_filter_list():
    """project_code_filter nhận list mã."""
    records = [
        {"project": "A", "id": 1},
        {"project": "B", "id": 2},
        {"project": "C", "id": 3},
        {"project": "A", "id": 4},
    ]
    groups, skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field="project",
        project_code_map={"A": "alpha", "B": "beta", "C": "gamma"},
        project_code_filter=["A", "B"],
        default_slug="default",
    )
    assert set(groups.keys()) == {"alpha", "beta"}
    assert len(groups["alpha"]) == 2
    assert len(groups["beta"]) == 1
    assert any(s["code"] == "C" and s["reason"] == "filtered" for s in skipped)


def test_extract_unique_project_codes():
    records = [
        {"project": "MPHG", "x": 1},
        {"project": "OTHER", "x": 2},
        {"project": "MPHG", "x": 3},
        {"project": "  ", "x": 4},
        {"x": 5},
    ]
    codes = integ_mod.extract_unique_project_codes(records, "project")
    assert codes == [
        {"code": "MPHG", "count": 2},
        {"code": "OTHER", "count": 1},
    ]


def test_apply_sync_routing_overrides_selected_map():
    ep = {
        "id": "ep1",
        "project_code_field": "project",
        "project_code_map": {"OLD": "old"},
        "project_code_filter": "OLD",
    }
    out = integ_mod._apply_sync_routing_overrides(
        ep,
        selected_map={"A": "alpha", "B": "beta"},
    )
    assert out["project_code_map"] == {"A": "alpha", "B": "beta"}
    assert set(out["project_code_filter"]) == {"A", "B"}
    # Nhiều mã → chưa gắn params.project (sync sẽ multi-fetch)
    assert "project" not in (out.get("params") or {})
    # Không mutate endpoint gốc
    assert ep["project_code_map"] == {"OLD": "old"}


def test_apply_sync_routing_overrides_injects_project_param():
    """1 mã filter → params.project để API server-side (?project=)."""
    ep = {
        "id": "ep1",
        "params": {"foo": "1"},
        "project_code_field": "project",
        "project_code_map": {},
        "project_code_filter": "",
    }
    out = integ_mod._apply_sync_routing_overrides(
        ep,
        selected_map={"MPHG_IHRP_2025_PM": "mphg"},
    )
    assert out["params"]["project"] == "MPHG_IHRP_2025_PM"
    assert out["params"]["foo"] == "1"
    assert ep.get("params") == {"foo": "1"}  # không mutate gốc


def test_project_query_codes_prefers_filter_over_params():
    ep = {
        "params": {"project": "FROM_PARAMS"},
        "project_code_filter": ["FROM_FILTER"],
    }
    assert integ_mod._project_query_codes(ep) == ["FROM_FILTER"]
    assert integ_mod._project_query_codes({"params": {"project": "P1"}}) == ["P1"]
    assert integ_mod._project_query_codes({"params": {}}) == []


def test_selected_map_list_filter_must_not_str_coerce():
    """
    Regression P0: modal gửi selected_map → filter = list mã.
    Nếu caller str(list) thì filter thành \"['MPHG_...']\" → lọc bỏ hết.
    """
    records = [
        {"project": "MPHG_IHRP_2025_PM", "id": i}
        for i in range(10)
    ] + [
        {"project": "OTHER_X", "id": 99},
        {"project": "  ", "id": 100},
        {"id": 101},
    ]
    ep = {
        "project_code_field": "project",
        "project_code_map": {"OLD": "old"},
        "project_code_filter": "OLD",
    }
    routed = integ_mod._apply_sync_routing_overrides(
        ep,
        selected_map={"MPHG_IHRP_2025_PM": "mphg"},
    )
    assert routed["project_code_map"] == {"MPHG_IHRP_2025_PM": "mphg"}
    assert isinstance(routed["project_code_filter"], list)
    assert routed["project_code_filter"] == ["MPHG_IHRP_2025_PM"]

    # Đúng: truyền list nguyên (giống fix sync path)
    groups, skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field=routed["project_code_field"],
        project_code_map=routed["project_code_map"],
        project_code_filter=routed["project_code_filter"] or "",
        default_slug="mphg",
    )
    assert len(groups["mphg"]) == 10
    reasons = {s["reason"]: s["count"] for s in skipped}
    assert reasons.get("filtered") == 1  # OTHER_X
    assert reasons.get("empty") == 2

    # Sai cũ: str(list) → 0 nhóm, toàn bộ bị filtered
    bad_groups, bad_skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field=routed["project_code_field"],
        project_code_map=routed["project_code_map"],
        project_code_filter=str(routed["project_code_filter"] or ""),
        default_slug="mphg",
    )
    assert bad_groups == {}
    assert sum(s["count"] for s in bad_skipped if s["reason"] == "filtered") == 11


def test_group_records_whitespace_and_empty_code():
    records = [
        {"project": "  KEEP  ", "id": 1},
        {"project": "", "id": 2},
        {"project": None, "id": 3},
        {"project": "\n", "id": 4},
    ]
    groups, skipped = integ_mod.group_records_by_project_code(
        records,
        project_code_field="project",
        project_code_map={"KEEP": "alpha"},
        project_code_filter=["KEEP"],
        default_slug="default",
    )
    assert len(groups["alpha"]) == 1
    assert sum(s["count"] for s in skipped if s["reason"] == "empty") == 3


def test_selected_map_overrides_saved_filter():
    """selected_map thay map + tự set filter = keys; bỏ filter cũ."""
    ep = {
        "project_code_field": "project",
        "project_code_map": {"OLD": "old"},
        "project_code_filter": "OLD",
    }
    out = integ_mod._apply_sync_routing_overrides(
        ep,
        selected_map={"NEW": "mphg"},
        project_code_filter=None,
    )
    assert out["project_code_map"] == {"NEW": "mphg"}
    assert out["project_code_filter"] == ["NEW"]
    # Explicit filter vẫn thắng khi truyền kèm
    out2 = integ_mod._apply_sync_routing_overrides(
        ep,
        selected_map={"NEW": "mphg", "KEEP": "keep"},
        project_code_filter=["KEEP"],
    )
    assert out2["project_code_map"] == {"NEW": "mphg", "KEEP": "keep"}
    assert out2["project_code_filter"] == ["KEEP"]


def test_merge_endpoint_project_code_map(tmp_path):
    folder = str(tmp_path)
    created = integ_mod.create_integration(folder, {
        "name": "T",
        "base_url": "https://example.com",
        "auth": {"method": "api_key", "api_key_env": "X"},
        "endpoints": [{
            "name": "FL",
            "path": "/api/x",
            "response_type": "json",
            "project_code_field": "project",
            "project_code_map": {"KEEP": "keep"},
        }],
    })
    ep_id = created["endpoints"][0]["id"]
    integ_mod.merge_endpoint_project_code_map(
        folder, created["id"], ep_id,
        {"NEW": "new-slug", "KEEP": "keep2"},
    )
    got = integ_mod.get_integration(folder, created["id"])
    assert got["endpoints"][0]["project_code_map"] == {
        "KEEP": "keep2",
        "NEW": "new-slug",
    }


def test_routing_enabled_requires_field_and_map_or_filter():
    assert integ_mod._routing_enabled({"project_code_field": ""}) is False
    assert integ_mod._routing_enabled({
        "project_code_field": "project",
        "project_code_map": {},
        "project_code_filter": "",
    }) is False
    assert integ_mod._routing_enabled({
        "project_code_field": "project",
        "project_code_map": {"A": "a"},
    }) is True
    assert integ_mod._routing_enabled({
        "project_code_field": "project",
        "project_code_filter": "A",
    }) is True


def test_ma_du_an_meta_detect(tmp_path):
    """Parser nhận cột Mã dự án vào meta.ma_du_an."""
    import openpyxl
    from parser.excel_parser import FunctionListParser

    path = tmp_path / "fl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    ws.append(["Mã CN", "Tên chức năng", "Module", "Mã dự án", "Analysis - Status"])
    ws.append(["X.01", "Test", "TMS", "MPHG_IHRP_2025_PM", "Closed"])
    ws.append(["X.02", "Test2", "HR", "OTHER_CODE", "Open"])
    wb.save(path)

    parsed = FunctionListParser().parse(str(path))
    assert parsed.meta_columns.get("ma_du_an") is not None
    assert parsed.rows[0].meta.get("ma_du_an") == "MPHG_IHRP_2025_PM"
    assert parsed.all_project_codes == ["MPHG_IHRP_2025_PM", "OTHER_CODE"]


def test_filter_by_project_codes(tmp_path):
    """_filter_parsed_data cắt rows theo meta.ma_du_an."""
    import openpyxl
    from app import _filter_parsed_data
    from parser.excel_parser import FunctionListParser

    path = tmp_path / "fl2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    ws.append(["Mã CN", "Tên chức năng", "Module", "Mã dự án", "Analysis - Status"])
    ws.append(["A.01", "A", "TMS", "MPHG_IHRP_2025_PM", "Closed"])
    ws.append(["B.01", "B", "HR", "OTHER_CODE", "Open"])
    ws.append(["C.01", "C", "PR", "MPHG_IHRP_2025_PM", "Open"])
    wb.save(path)

    parsed = FunctionListParser().parse(str(path))
    filtered = _filter_parsed_data(parsed, project_codes=["MPHG_IHRP_2025_PM"])
    assert len(filtered.rows) == 2
    assert filtered.all_project_codes == ["MPHG_IHRP_2025_PM"]
    assert all(r.meta.get("ma_du_an") == "MPHG_IHRP_2025_PM" for r in filtered.rows)


def test_structure_includes_all_project_codes(tmp_path):
    import openpyxl
    from analyzer.dashboard_engine import DashboardEngine
    from parser.excel_parser import FunctionListParser

    path = tmp_path / "fl3.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    ws.append(["Mã CN", "Tên chức năng", "Module", "Mã dự án", "Analysis - Status"])
    ws.append(["A.01", "A", "TMS", "CODE_A", "Closed"])
    wb.save(path)
    parsed = FunctionListParser().parse(str(path))
    metrics = DashboardEngine().compute_all(parsed)
    assert metrics["structure"]["all_project_codes"] == ["CODE_A"]


def test_dashboard_g_project_and_alias(flask_client, tmp_path):
    """Dashboard nhận g_project / g_ma_du_an; clear → full dataset."""
    import io
    import openpyxl

    path = tmp_path / "upload.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Function List"
    ws.append(["Mã CN", "Tên chức năng", "Module", "Mã dự án", "Analysis - Status"])
    ws.append(["A.01", "A", "TMS", "CODE_A", "Closed"])
    ws.append(["A.02", "B", "HR", "CODE_B", "Open"])
    wb.save(path)

    with open(path, "rb") as f:
        up = flask_client.post(
            "/api/projects/default/upload",
            data={"file": (io.BytesIO(f.read()), "upload.xlsx")},
            content_type="multipart/form-data",
        )
    assert up.status_code == 200

    r_a = flask_client.get("/api/projects/default/dashboard?g_project=CODE_A")
    assert r_a.status_code == 200
    body = r_a.get_json()
    assert body["applied_filter"]["project_codes"] == ["CODE_A"]
    assert body["applied_filter"]["row_count"] == 1
    assert body["metrics"]["summary"]["total_functions"] == 1

    r_clear = flask_client.get("/api/projects/default/dashboard")
    assert r_clear.get_json()["applied_filter"] is None
    assert r_clear.get_json()["metrics"]["summary"]["total_functions"] == 2

    r_alias = flask_client.get("/api/projects/default/dashboard?g_ma_du_an=CODE_B")
    assert r_alias.get_json()["applied_filter"]["row_count"] == 1


def test_saved_view_persists_project_codes(tmp_path):
    from analyzer.project_store import load_saved_views, upsert_saved_view

    d = str(tmp_path)
    upsert_saved_view(d, {
        "name": "Ma A",
        "modules": [],
        "processes": [],
        "pics": [],
        "project_codes": ["CODE_A"],
    })
    views = load_saved_views(d)
    assert views[0]["project_codes"] == ["CODE_A"]
