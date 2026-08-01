"""Unit tests — Phase B: Earned Value (EV/PV/AC → SPI/CPI)."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from analyzer.earned_value import (
    STATUS_PCT,
    compute_earned_value,
    _safe_ratio,
    _schedule_pct,
    _working_days,
)
from parser.excel_parser import FunctionListParser


TODAY = date(2026, 7, 31)


def _write_fl(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            ws.cell(r_i, c_i, v)
    wb.save(path)
    wb.close()


HEADERS = [
    "STT", "Mã CN", "Tên chức năng", "Module",
    "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC", "Dev - Estimate MH",
]


def _parse(path: Path):
    return FunctionListParser().parse(str(path))


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def test_safe_ratio_guards():
    assert _safe_ratio(10, 0) is None
    assert _safe_ratio(10, None) is None
    assert _safe_ratio(10, -1) is None
    assert _safe_ratio(10, 5) == 2.0


def test_schedule_pct_linear_and_edges():
    # End đã qua → 100%
    assert _schedule_pct(date(2026, 6, 1), date(2026, 6, 30), TODAY) == 1.0
    # Start chưa tới → 0%
    assert _schedule_pct(date(2026, 8, 1), date(2026, 8, 31), TODAY) == 0.0
    # Chỉ End quá hạn → 100%
    assert _schedule_pct(None, date(2026, 7, 1), TODAY) == 1.0
    assert _schedule_pct(None, date(2026, 8, 1), TODAY) == 0.0
    # Không ngày → None
    assert _schedule_pct(None, None, TODAY) is None
    # Giữa kỳ: Mon 2026-07-20 → Fri 2026-07-31 = 10 ngày làm; today=31 → 100%
    assert _schedule_pct(date(2026, 7, 20), date(2026, 7, 31), TODAY) == 1.0
    # Mid: Start 7/20 End 8/14, today 7/31
    mid = _schedule_pct(date(2026, 7, 20), date(2026, 8, 14), TODAY)
    assert mid is not None and 0 < mid < 1


# ------------------------------------------------------------------
# EV / CPI without baseline
# ------------------------------------------------------------------

def test_evm_without_baseline_spi_na(tmp_path):
    """Không baseline → SPI=None, vẫn có EV/AC/CPI."""
    path = tmp_path / "cur.xlsx"
    # Closed 16 MH, 2 ngày làm (Mon-Tue) → AC = 2*8 = 16 → CPI = 1
    _write_fl(path, HEADERS, [
        [1, "F.01", "Done", "TMS",
         date(2026, 7, 27), date(2026, 7, 28), "Closed", "A", 16],
        [2, "F.02", "WIP", "TMS",
         date(2026, 7, 27), date(2026, 8, 15), "In-progress", "B", 40],
    ])
    data = _parse(path)
    result = compute_earned_value(data, baseline=None, today=TODAY)

    assert result["has_baseline"] is False
    assert result["summary"]["spi"] is None
    assert result["summary"]["pv"] is None
    assert result["summary"]["spi_label"] == "N/A"
    # EV = 16*1.0 + 40*0.5 = 36
    assert result["summary"]["ev"] == 36.0
    assert result["summary"]["bac"] == 56.0
    assert result["summary"]["cpi"] is not None
    assert any("baseline" in m.lower() or "SPI" in m for m in result["messages"])


def test_evm_closed_ev_equals_mh(tmp_path):
    path = tmp_path / "cur.xlsx"
    _write_fl(path, HEADERS, [
        [1, "F.01", "Done", "HR",
         date(2026, 7, 1), date(2026, 7, 2), "Closed", "A", 8],
        [2, "F.02", "Open", "HR",
         None, None, "Open", "B", 8],
    ])
    result = compute_earned_value(_parse(path), today=TODAY)
    assert result["summary"]["ev"] == 8.0
    assert result["summary"]["bac"] == 16.0
    assert result["summary"]["ev_pct_bac"] == 50.0


def test_evm_cancelled_excluded(tmp_path):
    path = tmp_path / "cur.xlsx"
    _write_fl(path, HEADERS, [
        [1, "F.01", "X", "M",
         date(2026, 7, 1), date(2026, 7, 2), "Cancelled", "A", 100],
        [2, "F.02", "Y", "M",
         date(2026, 7, 1), date(2026, 7, 2), "Closed", "B", 10],
    ])
    result = compute_earned_value(_parse(path), today=TODAY)
    assert result["summary"]["bac"] == 10.0
    assert result["summary"]["ev"] == 10.0


def test_evm_default_mh_when_blank(tmp_path):
    """Ô Estimate MH trống → DEFAULT 8."""
    headers = [
        "STT", "Mã CN", "Tên chức năng", "Module",
        "Dev - Start", "Dev - End", "Dev - Status", "Dev - PIC",
    ]
    path = tmp_path / "cur.xlsx"
    _write_fl(path, headers, [
        [1, "F.01", "No MH", "M",
         date(2026, 7, 1), date(2026, 7, 2), "Closed", "A"],
    ])
    result = compute_earned_value(_parse(path), today=TODAY, default_mh=8.0)
    assert result["summary"]["bac"] == 8.0
    assert result["summary"]["ev"] == 8.0
    assert result["summary"]["phases_default_mh"] == 1


# ------------------------------------------------------------------
# PV / SPI with baseline
# ------------------------------------------------------------------

def test_evm_with_baseline_spi(tmp_path):
    """
    Baseline: cả 2 phase End ≤ today → PV = BAC.
    Current: 1 Closed (EV=16), 1 Open (EV=0) → SPI = 16/32 = 0.5
    """
    base_path = tmp_path / "base.xlsx"
    cur_path = tmp_path / "cur.xlsx"
    rows_base = [
        [1, "F.01", "A", "TMS",
         date(2026, 6, 1), date(2026, 6, 30), "Open", "X", 16],
        [2, "F.02", "B", "TMS",
         date(2026, 6, 1), date(2026, 6, 30), "Open", "Y", 16],
    ]
    rows_cur = [
        [1, "F.01", "A", "TMS",
         date(2026, 6, 1), date(2026, 6, 28), "Closed", "X", 16],
        [2, "F.02", "B", "TMS",
         date(2026, 7, 1), date(2026, 8, 15), "Open", "Y", 16],
    ]
    _write_fl(base_path, HEADERS, rows_base)
    _write_fl(cur_path, HEADERS, rows_cur)

    result = compute_earned_value(
        _parse(cur_path),
        baseline=_parse(base_path),
        baseline_snapshot_id="2026-06-01",
        today=TODAY,
    )
    assert result["has_baseline"] is True
    assert result["baseline_snapshot_id"] == "2026-06-01"
    assert result["summary"]["ev"] == 16.0
    assert result["summary"]["pv"] == 32.0
    assert result["summary"]["spi"] == 0.5
    assert result["summary"]["cpi"] is not None
    # Module row
    assert any(m["module"] == "TMS" and m["spi"] == 0.5 for m in result["modules"])


def test_evm_cpi_divide_by_zero(tmp_path):
    """Closed nhưng không Start → AC=0 → CPI=None."""
    path = tmp_path / "cur.xlsx"
    _write_fl(path, HEADERS, [
        [1, "F.01", "No start", "M",
         None, date(2026, 7, 2), "Closed", "A", 8],
    ])
    result = compute_earned_value(_parse(path), today=TODAY)
    assert result["summary"]["ev"] == 8.0
    assert result["summary"]["ac"] == 0.0
    assert result["summary"]["cpi"] is None
    assert result["summary"]["cpi_label"] == "N/A"


def test_status_pct_partial():
    assert STATUS_PCT["Closed"] == 1.0
    assert STATUS_PCT["In-progress"] == 0.5
    assert STATUS_PCT["Assigned"] == 0.25
