"""Tests cho exporter.excel_exporter."""
import os

import openpyxl
import pytest

from exporter.excel_exporter import (
    export_overdue_report,
    export_stalled_report,
    export_full_report,
    export_by_pic,
    export_compare_report,
)


def test_export_overdue_creates_file(tmp_path, metrics):
    """export_overdue_report tạo file .xlsx hợp lệ (Tong_hop + Chi_tiet)."""
    path = export_overdue_report(metrics["overdue_list"], str(tmp_path))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) >= {"Tong_hop", "Chi_tiet"}
    wb.close()


def test_export_overdue_with_filter(tmp_path, metrics):
    """Có filter → chỉ export row match module=TMS trên Chi_tiet."""
    path = export_overdue_report(
        metrics["overdue_list"], str(tmp_path),
        filters={"module": "TMS"},
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Chi_tiet"]
    rows = list(ws.iter_rows(values_only=True))
    # Header ở row 4 (row 1=title, row 2=subtitle, row 3=trống)
    header = rows[3]
    module_idx = None
    for i, h in enumerate(header):
        if h and "Module" in str(h):
            module_idx = i
            break
    assert module_idx is not None, f"Không tìm thấy cột Module trong header: {header}"
    for r in rows[4:]:
        val = r[module_idx]
        if val and str(val).startswith(("TMS", "ESS", "HR", "SI")):
            assert val == "TMS", f"Row bị lọt: {val}"
    wb.close()


def test_export_stalled_creates_file(tmp_path, metrics):
    """export_stalled_report tạo file .xlsx với Tong_hop + Chi_tiet."""
    items = metrics["stalled_tasks"]["items"]
    path = export_stalled_report(items, str(tmp_path))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) >= {"Tong_hop", "Chi_tiet"}
    assert "ĐÌNH TRỆ" in str(wb["Chi_tiet"]["A1"].value).upper() or "ĐÌNH TRỆ" in str(wb["Chi_tiet"]["A1"].value)
    wb.close()


def test_export_stalled_with_module_filter(tmp_path, metrics):
    """Filter module → chỉ giữ module đã chọn trên Chi_tiet."""
    items = metrics["stalled_tasks"]["items"]
    if not items:
        pytest.skip("sample không có stalled items")
    target = items[0]["module"]
    path = export_stalled_report(items, str(tmp_path), filters={"module": target})
    wb = openpyxl.load_workbook(path)
    ws = wb["Chi_tiet"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        try:
            int(str(row[0]))
        except (TypeError, ValueError):
            continue
        assert row[3] == target
    wb.close()


def test_export_full_report_has_multiple_sheets(tmp_path, metrics):
    """Full report có nhiều sheet: Summary + Overdue + Unassigned + Long Duration + Stalled + High Risk."""
    path = export_full_report(metrics, str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) >= 4
    wb.close()


def test_export_by_pic(tmp_path, metrics):
    """Export theo PIC tạo file .xlsx với sheet Info."""
    pic_list = metrics["pic_workload"]
    # Sample data trong conftest luôn có PIC (SonHN6/PhatTPT3/...); nếu không có,
    # nghĩa là parser hoặc fixture đã hỏng — assert thay vì silent-skip.
    assert pic_list, "Sample data phải có ít nhất 1 PIC (fixture bị hỏng?)"
    pic = pic_list[0]["pic"]
    path = export_by_pic(metrics, pic, str(tmp_path))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) >= 1
    wb.close()


def test_export_compare_report(tmp_path, parsed_data):
    """export_compare_report chạy được, tạo file .xlsx."""
    from analyzer.compare_engine import CompareEngine
    result = CompareEngine().compare(
        parsed_data, parsed_data,
        old_date="2026-07-01", new_date="2026-07-28",
    )
    path = export_compare_report(result, "2026-07-01", "2026-07-28", str(tmp_path))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) >= 1
    wb.close()


def test_export_empty_overdue_still_creates_file(tmp_path):
    """Empty overdue list vẫn tạo file (không crash)."""
    path = export_overdue_report([], str(tmp_path))
    assert os.path.exists(path)


def test_export_task_type_chart_detail_sheet(tmp_path, parsed_data, metrics):
    """export_chart(task_type) có Chi_tiet với cột status + đủ số function."""
    from exporter.excel_exporter import export_chart, build_task_type_detail_rows

    tt, rows = build_task_type_detail_rows(parsed_data)
    assert "Phân tích" in tt
    assert "Lập trình" in tt
    assert len(rows) == len(parsed_data.rows)
    assert rows[0]["statuses"]["Phân tích"] == "Closed"

    path = export_chart(
        "task_type",
        metrics,
        output_dir=str(tmp_path),
        parsed_data=parsed_data,
        group_by="module",
        mode="both",
    )
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) >= {"Tong_hop", "Theo_nhom", "Chi_tiet"}
    ws = wb["Chi_tiet"]
    headers = [c.value for c in ws[4]]
    assert "Phân tích" in headers
    assert "Lập trình" in headers
    assert "UAT" in headers
    # Đếm data rows
    n = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None or (isinstance(row[0], str) and str(row[0]).startswith("Tổng")):
            break
        n += 1
    assert n == len(parsed_data.rows)
    wb.close()


def test_export_chart_mode_summary_only(tmp_path, parsed_data, metrics):
    """mode=summary → chỉ Tong_hop (không Chi_tiet)."""
    from exporter.excel_exporter import export_chart

    path = export_chart(
        "priority", metrics, output_dir=str(tmp_path),
        parsed_data=parsed_data, mode="summary",
    )
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Tong_hop"]
    wb.close()


def test_export_chart_phase_stacked_both_sheets(tmp_path, parsed_data, metrics):
    """phase_stacked mode=both → Tong_hop + Chi_tiet có cột phase status."""
    from exporter.excel_exporter import export_chart

    path = export_chart(
        "phase_stacked", metrics, output_dir=str(tmp_path),
        parsed_data=parsed_data, mode="both",
    )
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"Tong_hop", "Chi_tiet"}
    headers = [c.value for c in wb["Chi_tiet"][4]]
    assert "Mã CN" in headers
    assert "Analysis" in headers or any(h for h in headers if h and "Analy" in str(h))
    wb.close()


def test_export_chart_module_overview_both_sheets(tmp_path, parsed_data, metrics):
    """module_overview mode=both → Tong_hop + Chi_tiet."""
    from exporter.excel_exporter import export_chart

    path = export_chart(
        "module_overview", metrics, output_dir=str(tmp_path),
        parsed_data=parsed_data, mode="both",
    )
    wb = openpyxl.load_workbook(path)
    assert "Tong_hop" in wb.sheetnames
    assert "Chi_tiet" in wb.sheetnames
    wb.close()


def test_export_overdue_mode_summary_only(tmp_path, metrics):
    """overdue mode=summary → chỉ Tong_hop."""
    path = export_overdue_report(metrics["overdue_list"], str(tmp_path), mode="summary")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Tong_hop"]
    wb.close()
